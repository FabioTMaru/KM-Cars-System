import tkinter as tk
from tkinter import ttk, font
import tkinter.messagebox as msgbox
import base64
from io import BytesIO
import sqlite3
import os
import logging
import shutil
from datetime import datetime

# ─── Logging ──────────────────────────────────────────────────────────────────
_log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(_log_dir, exist_ok=True)
logging.basicConfig(
    filename=os.path.join(_log_dir, "kmcars.log"),
    level=logging.ERROR,
    format="%(asctime)s [%(levelname)s] %(funcName)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    encoding="utf-8",
)
logger = logging.getLogger("kmcars")

# ─── FTM Logo (ICO embutido em base64) ────────────────────────────────────────
FTM_ICO_B64 = (
    "AAABAAEAIBIAAAEAIABwCQAAFgAAACgAAAAgAAAAJAAAAAEAIAAAAAAAAAkAAMMOAADDDgAAAAAAAA"
    "AAAAAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8A"
    "AAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/w"
    "AAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8BAQT/CQkm/wwLM/8CAgj/AAAA/wAAAP8AAAD/"
    "AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wUFE/"
    "8NDDj/Bwcd/wAAAf8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AgIG/xgYbP8tLMX/Kyq4"
    "/wYGG/8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD"
    "/AAAA/wAAAP8AAAD/Dg08/y0sy/8rKbv/ERFK/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP"
    "8KCyr/Li3H/ywrw/8WFlv/BgYN/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8CAgL/ERER/"
    "wQDBP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8GBRf/HRx2/y8t0f8mJqn/BAQP/wAAAP8AAAD"
    "/AAAA/wAAAP8AAAD/AAAA/xAQSf8xMNb/HR2A/x4eHP9SUVD/Q0ND/1hYWP87Ozv/S0tL/1VVV"
    "P9IR0f/SUhI/1RUVP9tbWz/VlZV/0FAQP82Njb/P0BA/1paWv9BQUH/ODc3/0dHRv8hISj/IiGh"
    "/y4tyP8JCSb/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/FBRZ/y8u1/8TE2L/Li0r/1tbW/9lZWb"
    "/amtr/1VWVv9PT1D/aGho/2NjY/9jY2T/Wlpa/11dXf9dXV3/Xl5f/2NjY/9UVVX/Y2Nj/1ZWV"
    "/9eXl//bW1t/ycnKP8dH4v/LSzO/woKMP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8XFmL/MC/Z"
    "/xMTW/8NDQz/Ly4w/xMTFP8DAwP/AwMD/wEBAf8DAwP/ERAR/xYVF/8EBAX/BAQE/wUFBf8EBAb"
    "/BwYH/wMCA/8CAgH/AwMD/wcGB/8QEBH/CAgJ/x0dh/8tLdL/DQ04/wAAAP8AAAD/AAAA/wAAAP"
    "8AAAD/AQEF/x0chP8xMNj/ERFK/wcHH/8cHIf/Dg5E/wAAAP8AAAD/AAAA/wAAAP8MDDn/Hx6L/"
    "wsLMP8AAAD/CAgm/x0dhv8LDDP/AQEF/wQEEv8AAAH/FBRb/xoac/8CAgn/Ghp2/zAw2f8UFFn/"
    "AAAB/wAAAP8AAAD/AAAA/wUFFf8eHoT/LS3L/yoquv8HBx//Cwsy/y8v0P8XF2T/AAAA/wAAAP8"
    "AAAD/AAAA/xQTWP8wL9f/ERFK/wAAAP8NDTv/Li7R/xESTv8REk7/JCSh/wcHHf8gII//KCi0/w"
    "MEDf8PEEH/LS3O/ysqwP8ZGWz/AgIH/wAAAP8AAAD/Bwcf/ycnr/8wMNf/Jiaq/wQEEP8MDDP/LS"
    "3N/yAgk/8TE1b/EhFQ/wQEEv8AAAD/FBNX/y8u0/8QEEn/AAAA/wwNOv8sLMz/GBlr/yYmrv8vL"
    "9T/Gxp1/yAfkP8nJ7T/AwQP/woKLv8rKsX/MC/U/yEhkf8CAwr/AAAA/wAAAP8BAQP/BwYc/yIh"
    "mf8vLtL/Dw8+/wsKMP8tLMv/Kiq//yYmq/8lJaH/CAgj/wAAAP8UE1j/Ly7T/xAQSP8AAAD/DAw"
    "5/ywry/8pKbv/Jyew/xkZcf8qKr7/KSm4/yYmsP8DAw7/GBht/zAu2P8bGnj/BAQQ/wAAAf8AAA"
    "D/AAAA/wAAAP8AAAD/FhZk/zEw2P8TE1b/Cwsv/y4uzf8aGnL/BQUU/wUFFP8DAwn/AgIH/xUVX"
    "P8vL9P/EhJO/wICCP8NDTz/LCzN/y4u0v8UFFj/AwIL/yIjmP8wL9n/JSau/wQEEf8dHIf/MC7U"
    "/w4OPf8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8UFVr/Ly/X/xUWX/8MDTL/Li/P/ykpu/8kJaD"
    "/JCOf/x8giP8jI57/KSm3/zAw1/8pKrX/IySb/xcXZv8uLdH/IyOb/wMDDf8AAAD/DA03/ywsyv"
    "8nKLT/BQUU/x8ej/8uLc//DQ01/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/xARS/8xMdb/HBx1/"
    "wYGGf8ZGW//Gxx4/xsceP8aG3b/Fhdk/xobdP8bHHb/Ght1/xscd/8bG3L/Dw8+/xkYbf8LCzD/"
    "AAAA/wAAAP8BAQT/FBRV/xUVXf8EBBD/JSWj/y0syv8JCir/AAAA/wAAAP8AAAD/AAAA/wAAAP8"
    "AAAD/DAww/zAwzf8pKLX/Dw4+/wEBBP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/w"
    "AAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/BAMN/xQTWf8sK8r/Kiq1/wUFFf8AAAD/AA"
    "AA/wAAAP8AAAD/AAAA/wAAAP8DAgn/HRx+/zAv0v8sK7r/BgYa/wAAAP8AAAD/AAAA/wAAAP8AA"
    "AD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8ODjv/Ly7N/y0r"
    "xv8VFVv/AQEC/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8CAgr/Dw89/xISTf8DAwz/AAAA"
    "/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/"
    "AAAA/wYGF/8SEUn/Cwos/wEBA/8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8A"
    "AAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAA"
    "AP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAD/AAAA/wAAAP8AAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA="
)


# ─── Color Palette — Clean Light Theme ────────────────────────────────────────
COLORS = {
    "bg_main":       "#F0F2F5",   # cinza claro principal
    "bg_content":    "#F7F8FA",   # fundo das páginas
    "bg_card":       "#FFFFFF",   # cards brancos
    "bg_sidebar":    "#555B6A",   # sidebar cinza médio claro
    "sidebar_hover": "#636A7A",
    "sidebar_active":"#5D6375",
    "accent":        "#CC0000",   # vermelho KM Cars
    "accent2":       "#FF3333",
    "accent_soft":   "#FDECEA",   # vermelho bem suave p/ hover
    "text_primary":  "#1A1A2E",
    "text_secondary":"#555770",
    "text_muted":    "#9DA3B4",
    "text_sidebar":  "#F0F2F5",
    "text_side_mut": "#C8CDD8",
    "border":        "#E2E5ED",
    "border_dark":   "#D0D4DF",
    "submenu_bg":    "#FFFFFF",
    "submenu_border":"#E2E5ED",
    "red":           "#CC0000",
    "inativo":       "#CC0000",
    "green":         "#2D9E6B",
    "blue":          "#2563EB",
    "orange":        "#D97706",
}

# ─── Font Palette — centralizado ──────────────────────────────────────────────
FONTS = {
    "nav":        ("Helvetica", 10, "bold"),
    "sub":        ("Helvetica",  9, "bold"),
    "body":       ("Helvetica", 10),
    "body_bold":  ("Helvetica", 10, "bold"),
    "small":      ("Helvetica",  9),
    "small_bold": ("Helvetica",  9, "bold"),
    "tiny":       ("Helvetica",  8),
    "tiny_bold":  ("Helvetica",  8, "bold"),
    "micro":      ("Helvetica",  7),
    "micro_bold": ("Helvetica",  7, "bold"),
    "title":      ("Georgia",   13, "bold"),
    "card_num":   ("Georgia",   28, "bold"),
    "logo":       ("Georgia",   34, "bold"),
}
SUBMENUS = {
    "Dashboard":     ["Visão Geral", "Indicadores", "Relatório Mensal", "Relatório Anual", "Estoque", "Vendas do Mês", "Shaken a Vencer", "Serviços", "Dossiê Cliente", "Dossiê Carro"],
    "Cadastros":     ["Clientes", "Carros", "Tipos de Custo", "Tipos de Serviço", "Categoria Desp. Fixas"],
    "Entrada/Compra":["Nova Compra", "Custos", "Despesas Fixas"],
    "Estoque":       ["Veículos", "Peças"],
    "Venda":         ["Nova Venda", "Lucro de Vendas"],
    "Serviço":       ["Ordem de Serviço", "Shaken", "Custo OS", "Custo SK", "Lucro OS", "Lucro SK", "Relatório"],
    "Parcelamentos": ["Financiamentos", "Previsão de Recebimento", "Dívidas Clientes"],
    "Configurações":  ["Geral", "Interface", "IS", "Banco de Dados"],
}

NAV_ICONS = {
    "Dashboard":     "⊞",
    "Cadastros":     "◈",
    "Entrada/Compra":"⊕",
    "Estoque":       "▦",
    "Venda":         "◆",
    "Serviço":       "⚙",
    "Parcelamentos": "◎",
    "Configurações":  "⚙",
}


class KMCars(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("KM Cars — Sistema de Gestão")
        self.configure(bg=COLORS["bg_main"])

        # ── Aplica escala TK antes de construir qualquer widget ───────────────
        try:
            import json, os, sys as _sys
            base = os.path.dirname(_sys.executable) if getattr(_sys,"frozen",False) else os.path.dirname(os.path.abspath(__file__))
            cfg_path = os.path.join(base, "kmcars_config.json")
            with open(cfg_path, encoding="utf-8") as _f:
                _cfg = json.load(_f)
            _escala = float(_cfg.get("escala", 1.0))
        except Exception:
            _escala = 1.0
        # tk scaling: valor padrão do sistema é ~1.33 em Windows 96dpi
        # Ajustamos relativamente ao padrão detectado
        try:
            _base_scaling = self.tk.call("tk", "scaling")
            self.tk.call("tk", "scaling", _base_scaling * _escala)
        except Exception:
            pass

        self.geometry("1340x820")
        self.minsize(1100, 700)

        self.current_page = tk.StringVar(value="Dashboard")
        self.current_sub  = {page: subs[0] for page, subs in SUBMENUS.items()}

        self._init_db()          # ← banco de dados primeiro
        self._load_ftm_logo()
        self._load_fonts()
        self._build_ui()
        self._cfg_apply_on_startup()
        self._show_page("Dashboard")


    # ── Banco de Dados SQLite ─────────────────────────────────────────────────
    def _init_db(self):
        import sys as _sys
        if getattr(_sys, "frozen", False):
            # Executando como .exe gerado pelo PyInstaller
            base = os.path.dirname(_sys.executable)
        else:
            # Executando como .py normal
            base = os.path.dirname(os.path.abspath(__file__))
        self._db_path = os.path.join(base, "kmcars.db")
        self.conn = sqlite3.connect(self._db_path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        cur = self.conn.cursor()

        # ── Tabelas base (versão inicial) ─────────────────────────────────────
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS clientes (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                nome     TEXT NOT NULL,
                telefone TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS carros (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                carro      TEXT NOT NULL,
                ano        TEXT,
                cor        TEXT,
                placa      TEXT,
                chassi     TEXT,
                status     TEXT NOT NULL DEFAULT 'Estoque',
                cliente_id INTEGER
            );
            CREATE TABLE IF NOT EXISTS compras (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                carro_id        INTEGER,
                carro           TEXT NOT NULL,
                cor             TEXT,
                placa           TEXT,
                tipo            TEXT NOT NULL,
                valor           INTEGER,
                data_entrada    TEXT,
                taxa_leilao     INTEGER,
                taxa_reciclagem INTEGER,
                a_venda         INTEGER DEFAULT 0,
                preco_venda     INTEGER,
                preco_avista    INTEGER
            );
            CREATE TABLE IF NOT EXISTS tipos_custo (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS tipos_servico (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS servicos (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                os_num         TEXT NOT NULL,
                carro_id       INTEGER,
                carro          TEXT,
                cliente_id     INTEGER,
                data_servico   TEXT,
                tipo_servico   TEXT,
                descricao      TEXT,
                valor          INTEGER,
                status         TEXT DEFAULT 'Aberto'
            );
            CREATE TABLE IF NOT EXISTS custos (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                compra_id   INTEGER NOT NULL,
                tipo_custo  TEXT NOT NULL,
                descricao   TEXT,
                valor       INTEGER,
                FOREIGN KEY (compra_id) REFERENCES compras(id)
            );
            CREATE TABLE IF NOT EXISTS vendas (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                carro_id            INTEGER,
                carro               TEXT NOT NULL,
                cor                 TEXT,
                placa               TEXT,
                cliente_id          INTEGER,
                tipo_venda          TEXT NOT NULL,
                valor_venda         INTEGER,
                entrada             INTEGER,
                parcela_mensal      INTEGER,
                num_parcelas        INTEGER,
                valor_ultima_parc   INTEGER,
                data_primeira_parc  TEXT,
                data_venda          TEXT,
                carro_troca_id      INTEGER,
                carro_troca         TEXT,
                valor_troca         INTEGER,
                volta_paga          INTEGER,
                parcelas_pagas      INTEGER DEFAULT 0,
                obs                 TEXT
            );
            CREATE TABLE IF NOT EXISTS pagamentos (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                venda_id        INTEGER NOT NULL,
                num_parcela     INTEGER NOT NULL,
                data_pagamento  TEXT NOT NULL,
                valor_pago      INTEGER NOT NULL,
                FOREIGN KEY (venda_id) REFERENCES vendas(id)
            );
            CREATE TABLE IF NOT EXISTS shaken (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                sk_num          TEXT NOT NULL UNIQUE,
                carro_id        INTEGER NOT NULL,
                cliente_id      INTEGER,
                valor           INTEGER,
                data_vencimento TEXT,
                por_conta       INTEGER DEFAULT 0,
                data_registro   TEXT,
                obs             TEXT
            );
            CREATE TABLE IF NOT EXISTS despesas_fixas (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                descricao   TEXT NOT NULL,
                valor       INTEGER NOT NULL,
                categoria   TEXT DEFAULT 'Geral',
                data_ref    TEXT,
                recorrente  INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS categorias_despesas (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS configuracoes (
                chave TEXT PRIMARY KEY,
                valor TEXT
            );
            CREATE TABLE IF NOT EXISTS custos_os (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                servico_id  INTEGER NOT NULL,
                tipo_custo  TEXT NOT NULL,
                descricao   TEXT,
                valor       INTEGER,
                data_custo  TEXT,
                FOREIGN KEY (servico_id) REFERENCES servicos(id)
            );
            CREATE TABLE IF NOT EXISTS custos_sk (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                shaken_id   INTEGER NOT NULL,
                tipo_custo  TEXT NOT NULL,
                descricao   TEXT,
                valor       INTEGER,
                data_custo  TEXT,
                FOREIGN KEY (shaken_id) REFERENCES shaken(id)
            );
        """)

        # ── Migrações versionadas via PRAGMA user_version ─────────────────────
        self._run_migrations(cur)

        # ── Dados padrão ──────────────────────────────────────────────────────
        self.conn.execute(
            "INSERT OR IGNORE INTO configuracoes (chave, valor) VALUES ('cambio_brl_jpy', '0')")
        tipos_padrao = ["Peça", "Gasolina", "Óleo", "Transferência",
                        "Jibaiseki", "Juuryouzei", "Placa", "Entrada Polícia",
                        "Taxa Reciclagem", "Taxa Leilão"]
        for t in tipos_padrao:
            cur.execute("INSERT OR IGNORE INTO tipos_custo (nome) VALUES (?)", (t,))
        for ts in ["Troca de Óleo", "Manutenção"]:
            cur.execute("INSERT OR IGNORE INTO tipos_servico (nome) VALUES (?)", (ts,))
        for cat in ["Geral", "Aluguel", "Utilities", "Salários", "Marketing", "Impostos"]:
            cur.execute("INSERT OR IGNORE INTO categorias_despesas (nome) VALUES (?)", (cat,))
        self.conn.commit()

        # ── Backup automático ao iniciar ──────────────────────────────────────
        self._auto_backup()

    def _run_migrations(self, cur):
        """Migrações versionadas usando PRAGMA user_version."""
        version = self.conn.execute("PRAGMA user_version").fetchone()[0]
        logger.info(f"Versão do banco: {version}")

        # ── v0 → v1: colunas adicionais que podem não existir em BDs antigos ──
        if version < 1:
            migrations_v1 = [
                "ALTER TABLE compras  ADD COLUMN cor              TEXT",
                "ALTER TABLE compras  ADD COLUMN placa            TEXT",
                "ALTER TABLE compras  ADD COLUMN data_entrada     TEXT",
                "ALTER TABLE compras  ADD COLUMN taxa_leilao      INTEGER",
                "ALTER TABLE compras  ADD COLUMN taxa_reciclagem  INTEGER",
                "ALTER TABLE compras  ADD COLUMN carro_id         INTEGER",
                "ALTER TABLE compras  ADD COLUMN a_venda          INTEGER DEFAULT 0",
                "ALTER TABLE compras  ADD COLUMN preco_venda      INTEGER",
                "ALTER TABLE compras  ADD COLUMN preco_avista     INTEGER",
                "ALTER TABLE compras  ADD COLUMN cliente_id       INTEGER",
                "ALTER TABLE compras  ADD COLUMN data_compra      TEXT",
                "ALTER TABLE custos   ADD COLUMN data_custo       TEXT",
                "ALTER TABLE carros   ADD COLUMN cliente_id       INTEGER",
                "ALTER TABLE carros   ADD COLUMN data_shaken      TEXT",
                "ALTER TABLE carros   ADD COLUMN historico_status TEXT",
                "ALTER TABLE vendas   ADD COLUMN parcelas_pagas   INTEGER DEFAULT 0",
                "ALTER TABLE clientes ADD COLUMN cidade           TEXT    DEFAULT ''",
                "ALTER TABLE vendas   ADD COLUMN compra_id        INTEGER",
                "ALTER TABLE vendas   ADD COLUMN lucro            INTEGER",
                "ALTER TABLE servicos ADD COLUMN os_num           TEXT",
                "ALTER TABLE servicos ADD COLUMN valor            INTEGER",
                "ALTER TABLE servicos ADD COLUMN status           TEXT DEFAULT 'Aberto'",
                "ALTER TABLE servicos ADD COLUMN obs              TEXT",
                "ALTER TABLE shaken   ADD COLUMN custo_shaken     INTEGER",
                "ALTER TABLE shaken   ADD COLUMN renovado         INTEGER DEFAULT 0",
                "ALTER TABLE shaken   ADD COLUMN renovacao_de     INTEGER",
            ]
            for sql in migrations_v1:
                try:
                    self.conn.execute(sql)
                except sqlite3.OperationalError as e:
                    if "duplicate column" not in str(e).lower():
                        logger.warning(f"Migração v1 ignorada ({sql[:50]}...): {e}")
            self.conn.execute("PRAGMA user_version = 1")
            self.conn.commit()
            logger.info("Migração v1 concluída.")

        # ── v1 → v2: normalização tipo custo Shaken ───────────────────────────
        if version < 2:
            try:
                self.conn.execute(
                    "UPDATE custos SET tipo_custo='Serviço Shaken' WHERE tipo_custo='Custo Shaken'")
                self.conn.execute(
                    "UPDATE tipos_custo SET nome='Serviço Shaken' WHERE nome='Custo Shaken'")
                self.conn.execute("PRAGMA user_version = 2")
                self.conn.commit()
                logger.info("Migração v2 concluída.")
            except Exception as e:
                logger.error(f"Erro na migração v2: {e}", exc_info=True)

        # ── v2 → v3: NFI (SNF/CNF) e config IS ─────────────────────────────────
        if version < 3:
            mig_v3 = [
                "ALTER TABLE vendas   ADD COLUMN nfi TEXT DEFAULT 'SNF'",
                "ALTER TABLE servicos ADD COLUMN nfi TEXT DEFAULT 'SNF'",
                "ALTER TABLE shaken   ADD COLUMN nfi TEXT DEFAULT 'SNF'",
                "INSERT OR IGNORE INTO configuracoes (chave,valor) VALUES ('is_percentual','10')",
            ]
            for sql in mig_v3:
                try:
                    self.conn.execute(sql)
                except Exception as e:
                    if "duplicate column" not in str(e).lower():
                        logger.warning(f"Migração v3 ignorada: {e}")
            self.conn.execute("PRAGMA user_version = 3")
            self.conn.commit()
            logger.info("Migração v3 concluída.")

        # ── v3 → v4: IS retroativo para registros CNF já existentes ──────────
        if version < 4:
            try:
                is_perc_mig = 10.0
                try:
                    r = self.conn.execute(
                        "SELECT valor FROM configuracoes WHERE chave='is_percentual'"
                    ).fetchone()
                    if r: is_perc_mig = float(r[0])
                except Exception: pass

                def _ins_is_if_missing(table_name, id_col, val_col, date_col, ref_col):
                    """Insere IS em table_name para registros CNF sem IS ainda."""
                    rows = self.conn.execute(
                        f"SELECT id, {val_col}, {date_col} FROM {ref_col} WHERE nfi='CNF'"
                    ).fetchall()
                    for rid, val, dt in rows:
                        val_int = 0
                        try: val_int = int(str(val or "0").replace(",",""))
                        except Exception: pass
                        if not val_int: continue
                        # Verifica se IS já existe
                        chk = self.conn.execute(
                            f"SELECT COUNT(*) FROM {table_name} WHERE {id_col}=? AND tipo_custo='IS'",
                            (rid,)).fetchone()
                        if chk and chk[0] > 0: continue
                        is_v = int(val_int * is_perc_mig / 100)
                        if is_v <= 0: continue
                        self.conn.execute(
                            f"INSERT INTO {table_name} ({id_col},tipo_custo,descricao,valor,data_custo) "
                            "VALUES (?,?,?,?,?)",
                            (rid, "IS", "IS automático CNF", str(is_v), dt or ""))

                # Vendas CNF → custos (via compra_id)
                vendas_cnf = self.conn.execute(
                    "SELECT v.id, v.valor_venda, v.data_venda, v.compra_id, v.carro_id "
                    "FROM vendas v WHERE v.nfi='CNF'"
                ).fetchall()
                for vid, val, dt, compra_id, carro_id in vendas_cnf:
                    val_int = 0
                    try: val_int = int(str(val or "0").replace(",",""))
                    except Exception: pass
                    if not val_int: continue
                    cid = compra_id
                    if not cid and carro_id:
                        cr = self.conn.execute(
                            "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                            (carro_id,)).fetchone()
                        cid = cr[0] if cr else None
                    if not cid: continue
                    chk = self.conn.execute(
                        "SELECT COUNT(*) FROM custos WHERE compra_id=? AND tipo_custo='IS'",
                        (cid,)).fetchone()
                    if chk and chk[0] > 0: continue
                    is_v = int(val_int * is_perc_mig / 100)
                    if is_v <= 0: continue
                    self.conn.execute(
                        "INSERT INTO custos (compra_id,tipo_custo,descricao,valor,data_custo) "
                        "VALUES (?,?,?,?,?)",
                        (cid, "IS", "IS automático CNF", str(is_v), dt or ""))

                # OS CNF → custos_os
                _ins_is_if_missing("custos_os", "servico_id", "valor", "data_servico", "servicos")

                # Shaken CNF → custos_sk
                _ins_is_if_missing("custos_sk", "shaken_id", "valor", "data_registro", "shaken")

                self.conn.execute("PRAGMA user_version = 4")
                self.conn.commit()
                logger.info("Migração v4 (IS retroativo CNF) concluída.")
            except Exception as e:
                logger.error(f"Erro na migração v4: {e}", exc_info=True)

        # ── v5: data_conclusao em servicos ───────────────────────────────────
        if version < 5:
            try:
                self.conn.execute(
                    "ALTER TABLE servicos ADD COLUMN data_conclusao TEXT")
            except Exception:
                pass  # coluna já existe
            self.conn.execute("PRAGMA user_version = 5")
            self.conn.commit()

        # ── Próximas migrações: adicionar aqui como v6, etc. ──────────────────

    def _auto_backup(self):
        """Cria backup automático do banco, mantendo os 10 mais recentes."""
        try:
            if not os.path.exists(self._db_path):
                return
            backup_dir = os.path.join(os.path.dirname(self._db_path), "backups")
            os.makedirs(backup_dir, exist_ok=True)

            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest  = os.path.join(backup_dir, f"kmcars_{stamp}.db")

            # Copia segura usando sqlite3 backup API (flush correto)
            with sqlite3.connect(dest) as bk_conn:
                self.conn.backup(bk_conn)

            # Remove backups antigos (mantém os 10 mais recentes)
            backups = sorted(
                f for f in os.listdir(backup_dir) if f.startswith("kmcars_") and f.endswith(".db")
            )
            for old in backups[:-10]:
                try:
                    os.remove(os.path.join(backup_dir, old))
                except Exception as e:
                    logger.warning(f"Não foi possível remover backup antigo {old}: {e}")

            logger.info(f"Backup criado: {dest}")
        except Exception as e:
            logger.error(f"Falha no backup automático: {e}", exc_info=True)

    # ── Logo FTM ──────────────────────────────────────────────────────────────
    def _load_ftm_logo(self):
        import os, sys as _sys2, base64, io

        self.ftm_img          = None
        self.ftm_photo        = None
        self.ftm_photo_footer = None

        # Logo embutido em base64 (não depende de arquivo externo)
        FTM_LOGO_B64 = (
            "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAEeAfADASIAAhEBAxEB/8QAHQABAAICAwEBAAAAAAAAAAAAAAEHBggCAwUECf/EAFkQAAEDAwEEBAYLCgoIBgMAAAEAAgMEBREGBxIhMQhBUWETInGBkbEUFjI2cnSTobLB0QkVIyY1QkVSVXMXJCVDRGJjksLhKDNUZHWCorMYNFNlg9JWhKP/xAAcAQEAAgMBAQEAAAAAAAAAAAAABQcBBAYDAgj/xAA+EQACAQICBQcKBQMFAQAAAAAAAQIDBAURBiExQXESM1FhgbHBExUWNVJTcpGh0RQiIzThJTLwJEJDkrKC/9oADAMBAAIRAxEAPwDTNEUIAiIgCIpQBERAFClEBCIiAIilAQpREAREQEIiIAiIgCKUQBERAFCIgCIiAKURAEQIgIRSoQBERAEUogCIiAhFKhAEREBKhFKAhSiIAiIgChEQBERAFKIgCIiAhERASoREAREQBSiZQBEyoQBSoX32W03G71Qp7dSyTvPPdHAeU9Sw2orNn3TpzqSUILNvcj4sKCCFaFm2VSkB93uIiOM+DgbvHyZPBZJTbO9LU4G9SzVJHMyynj5hhaU8Qox1LWdLb6I4jWWcko8X4LMotFf7NH6ZjIxZaU47cn1lYdthtNqttqon0Fvp6aR05a50bcEjd5LFK/hUmoJPWfd9olc2VtK4nOLUdyz4dBWOURQt85QlQiIAiIgCIpygCZRMoAoREAREQBSiIAmUKIAoUqEAREQBEClAEyiZQBEyoQBERAFKhEBKIu6ga2StgjeMsdI1rh2glYbyMxWbyOkor1l0DpPHi214Hb4d32r5pNnGmJRhsNVCe1s2fWForEaXWddLQrEFvj839ilMKCrbrtlVC9pNDdpY3dTZYwR6QsP1DoLUNoY6V1KKqAfzkB3vSOa9qd3RqPJMi7vR3ELWPKnTbXStfcYmikgg4IwQoWyQgUomUAyiZRAQiIgCIiAIiIAiIgJUIiAIpC++x22e63Wnt9OPwk7w0cOQ6z5hxWG0lmz7p05VJKMVm3qR6+htJ1OoanwshdDQROxJL1uP6re/1K6bVbaS1UTKWhp2QRDqaOJ7yesrnbLZTWq2w0NI0NihaGgdZPWT3ld6525upVpdRcmB4HRwyktWdR7X4Lq7zmHuPAlcgexdfk5qQcLWJ3I5lVztzH8mUHxh30VYoyeoqv8Abowi0UBx/SD9FbNpz8SC0kX9LrLqXeioURF0ZTAREQBERASihEAREQBERAEREBKhSoQBERAEREAREQBERASoREAREQBERAEREBK77b+Uab9631hfOvptn5Rpv3rfWFiWw+6f96Nk3ZI86cl2OYd0cDzXU/IPHK5RM/QGtkh5bxBwVDpHuPuiuskk8VI+dD6WoxHW+iKS9RvqqFkdNcQM5aMMl7ndh71TVXTz0lTJTVEbopo3Fr2OHEELZVoyVgG2PTbJaNuoKRg8LFhlVj85vIO8o5eTyKTsbtqSpz2bjhdKtH6c6cry3WUlrklvXTxW/pKlULkRhcVMlaBERAEREAREQBERAEREAREQEjmrN2HW0S11bdXt/wBQwRRk9rufzKsVeOx2D2PowPPuqiV0mfJwHqWliE+TRaW/UdNolbKviUW9kU34L6sy9/E8+K6i30rnzTCgC3Uda83UF7oLHQmrr5t1vJjG8XPPYAvSlw2N0jjhjRlx7lQOtb3Lfb7NUl58AwlkDc8AwfbzW1aW3l569iIHSHGvNdunFZzls8X2GT3HapdXPc23UkFPH+aZBvu+xY1qPVN5v8McVzqGyMjdvsa1gbg4wvBRTlO3pU9cYlXXWNX92nGrVbT3bF8lqJKhSmF7EWQikhEBCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAilQgCIiAIiIAiIgCKVCAIinCAhdkEj4ZmTM90xwcPKFwTJQynk80Z3T7UtSxnMvsWYDqdFj1LJdObRqG51Dae5RewZnnDX5zGT9Sp/KhalSyozWpZcDoLTSjEreak6nKXQ9f8AJssDnrz2KetYZsmvkl0s7qGoeXVFGQ0OPNzDy9HJZqBxwoKrTdObg9xbNjewvbeFxT2SXy6V2M5sHpXZLTRV9FUUUzQY5o3MeD2ELrC7YXbrjx6ivPYbEoqSyew1pulK+iuFRSSe7hkcw+Yr5Sss2s0wptZVD2jDahjZfORx+cLEl1FKfLgpdJROIW/4a6qUfZbX1CIi9DTCIpQEIiICURQgCIiAIilAQtgtDQ+xtKW6HGP4s13p4/Wtf2jLgMda2QtcQht1PEB7iBjfQ0KKxR/liju9BaederPoSXzf8H1NPUuQXUMBc2nqUQWQzydeVZodFXKdrt15jEbT3uOPrWvbhgq7tsU3g9FiMH/W1LB6ASqRdzKm8NjlTb6WVbprWc72MN0Y97f8HFepSaevlXTsqKa11UsTxlr2xkgheYVfmz3PtNtueOIO3vK97u4dCKkkRuj2EU8UrypVJNJLPVxRTg0pqT9i1nya7BpHUpH5FrP7ivwO7gubZO4KOeJz6EdctBrT3svoUAdI6l/YlZ8muJ0lqUfoSt+TWwfhePuQnhOvdCec6nsoeg9p72X0NejpTUg52Wt+TKe1XUf7FrfkythS/jyCeE7gs+c6nsox6DWvvZfQ14OltRDnZq35IoNMahPKzVvyRWwjpCeoIHHdccDl2p5zn7KHoNbe9l8ka6Nsd2dWPoxbqk1LGhzotw7wB5EhfR7V9Rfsat+SKs2gkP8ACxcySf8AyTBz+Cs4MxGPFHJelTEZwy1LWkzSsdELe5jNuo1yZSju3PI169q2ov2NW/JFT7VdR/sWt+SK2FEueoIXnsC8/OlT2UbvoNa+9l8ka9DSmpDystb8mVy9qWpf2JW/JrYIydwXHwx/VGU851PZRn0Gtfey+hr97UtS/sWt+TT2pal/YlZ8mtghIT1BC89gTznU9lGPQe097L6Gvw0lqX9iVvyag6S1KP0JW/JlbBCU9gXPw2fzQnnOp7KHoPa+9l9DXj2q6jB42Wt+SK+G4W6tt8whrqaWnkI3g2RuDjtWyjnEnkFTW2txOr2DspmesrZtb2VapyWiGx3Rmhhtr5enNt5pa8t5idDarlXsc+ioZ6hrDhxjYXAFd/tcv37IrPkirC2ISFtquIHXO36KsIPdnkF8V7+VOo4JbDYwnRKjfWkLiVRpy3ZLpyNexpu/n9D1vyRXwVlLUUc7oKqGSGVvumPbghbMMlc3jgKvttNjNVQQ3+njBlhHg6jA5s6j5j8x7lmhiDqTUZLLMxi2h8bS1lXozcnHW01u3/LaVEiKVJnCkIpRAERQgJHNffQWa7V8Jno7dU1EQO6XxxkjPYum00c9xuMFDTN3pZnhrftWxen6KGy2anttKBuRN4nrc7rJ8pWnd3fkEslm2dJo/gDxWUpTbjCO/r6Pua/v09fGjJtNaP8A4Sus2O8jna6z5Fy2Pklc4cVwa92HHOOHatLzpP2Tp/QS3f8AzP5I1mqIZqeUxTxPikHNrxghcFk+1F29rWtJ7GfRCxjgpanPlwUukr28oK3uJ0U8+S2vkzM9jtT4DWUUJdhtRE6MjtOMj1K6HDDlQGhJ/Y2rrXKOGKhoPkJx9a2AlwHKGxKOVVPpRZOhNZzsZU3/ALZd6X8nFGnjwXAnqRp4qPOzKs23wBtfbakDi+FzCfI7P1qulam2+Mut1tmwfFle30gH6lVi6GxedCJTWlNPkYpV68n9EQpRFtnPhERAQiIgCIiAIiIApREB3UTd+ribjOXtHzrZPc3WAY5AD5lrlZW791pG9s7B/wBQWyD+ah8U2x7SxdBY/krPrj4nUpz6FJC4qLO/ME21zYsNDH1OqSfQ3/NVITlWjtud/Eba3tkefmCq1T+HrKgu0qDS2XKxSfUl3EgcVsDs8YPaVbDj+j/WVr8FsFs7I9pNs+L/AFleOKc2uJIaEfu6nw+KPVe3ChTIuPUoUs9EtOOtdg5LD9bavfpmtpoBQMqRNGX5c/dxg47F4g2ryY/IsPyp+xbMLOtUjyorUQl1pDh1rVdGrUyktup/YsslcCfQq0dtYef0JF8qfsXH+FV37GZ8qfsX3+Br+z9Tw9KsK979JfYszgu2Ju81w7lV42qnrszPlf8AJdrNrOAR95W8R/6qfga/smfSnC/e/R/Y9OiYRtZuY4f+Sb/hWZuxgeRU7T65bFq+qv5t4d4eEReC3+WMcc+Ze3/CozHCzD5VetazrSayjuRG4ZpDh1GNTl1Ms5ya1PY3q3FkNK55HaqzO1ZoHCzD5VdZ2qf+zt+VXj+Br+ySXpRhXvfo/sWc7jnC44PUvH0PfDqK0SV7qYU+7KYw0Ozyx9q9w4J4LWnFwk4vaTVvXp3FKNWk84vWji3guYIPJdb+0Lx9VX9unrYK99MagGUR7odjnn7EjFzeS2ma9anQpurUeUVtPZLVAOFX52sU3XZX/KhdMm1WnJ4WZw/+UfYtn8FX9khfSfC/e/R/YslrgceVU7tqb+NzSP8AZ2+sr1xtUhH6Hf8AK/5LENbagZqS7tr2U5p8RBm6XZ5Z+1bdlbVadXlSWo5/SXGrG8sXSoVM5ZrVk/sZlsTH8mXDh/Pt+irFHzrANiYH3quJ/t2/RWfE4WneP9eR0mjK/pdHg+9nIkLqqoo6mmfTTtD4ZGlrmnkQea5Zz5VK1sybkk1kzX7V1mlsN9noJMlgO9E79Zh5FeSrr2qaf++9iFZTMzV0YLwAOL2dY+v0qlMLorSv5annvW0pfH8LeHXbgl+R648Ojs2BEULaIQlEC9nSNllv17gt7AQxx3pXge4YOZXzKSinJ7j1oUZ16kacFm28kZxsbsBYx98qI/GeDHTgjq63fV6VY4cRzXGKliooI6WBoZFE0MY0dQAT1rma9V1ZubLwwrD4YfaxoR3bX0ve/wDNx2Z9K5NaDvfBXU0ruhx43wV5G/sKJ2oNxrSs8jPohYyFlO1PHt1rPgs+iFixwuntuZjwRRmL/v63xS7z67LL4K70b/1Z2H5wtjHu3iT1LWukdu1ULux7T862QjOWNPaB6lG4otcXxO00El+SvHrj4krk0ZKgD0Lm0DmossAwjbVEDpmkfj3NUPnaVTpCunbK3OkIz2VTPU5Uu7mp3DuZ7SpNMVliTfSkQiKFvnKhERAERSgIREQBSoRASoREB6Omml+oLewddTH9ILYtxy7zqg9ndM6q1lbmAEhknhD3Boyr4aT1qFxN/nS6izNBqbVtUn0y7l/JzUEcMqQc9SnGQo07gq/be7xbWz94fUq0VjbcXj75W6HPuYXOPnd/kq4XQ2KyoRKa0olysUq9ncieSv3Z28e0y2t7ID6yqBV8bOs+0+3fuT6yvDE+bXEldB3/AKyp8PijIu5S0LiPIuY8ihSzyqNuGPvvbx/u7vpKulYe252bzQD/AHc/SVeLorLmIlMaS+tK3FdyCnCBMraIMKFITCAFEUIAiIgLn2L+9Cb4y71BZkTxWGbFz+KM/wAaPqCzPHFc1dc9LiXbgHq2h8KJAysN2xtxo8O/3lnqKzIFYftiOdG//ss+tLbno8TOOr+nVvhZSxOUUHmi6UpEKRzUKRzQFsbEiPvTcP37forPz2KvtieRabh8Yb9FWDnK5y85+RdOjT/pdHg+9kKQoxlSAtYmzsj7+KpHahp/7yagdJAzFJVZkjxyaetquve4heTrKxt1DZJqPA8O38JAT1PHV5+S2bSv5Gpm9j2kFpDhSxG0cY/3x1x49Hb35GviLtqIZIJnxSsLZGOLXNPMELrXRlMtNPJgc1eGymwC0WMVlQ3drKwB7sjixn5o+tVxs2sBvN7bNOzNHSkPk7HHqarvJw7xeSiMSr/8S7SwNDMJzzvai6o+L8Pmdkvuj1rqIwV2Dioc1RJYaOvK7InY3vgrrIwpbyd5EZkpHakc60rPgs+iFixWTbT/AH51nkZ9ELGF09vzUeCKLxfXf1vil3nKI4lae8LZCkO9TxO7WNPzLW5pwQexbG2R3hbPRzDiH07D/wBIWhii1RfE6/QSX5q6+HxPraFyHeoTOFEFisxXa43e0VK79SaN3zkfWqQPNX1tCpzU6MubACSyISD/AJSD9RVCKbw1/pNdZVem1Nxv4y6YrvYREUiccEREAREQBERAFKhSBlAQi5sY97w1jS5x5ADJWWaV0Jd7xK19RC+ipAcuklGHEdw5r4qVI01nJ5G1aWVe7mqdGLk+rx6D2Ni1re6pq7u9hDGM8DEe0nifmwrQC6rZb6W12+Kgo4wyGJuGjrPaT3ruxjqXOXFby1RyLnwbDlh1pGhta1t9b/zI5NXawZBPYukZUumZDTzTSODWxt3nE9g4rxJNvLaUztiqhPrGSIcoImR+fGT61hhX3X2vfcrxV1zznw0rnDyZ4fMvh6l09GHIpxj0FEYlc/irupWWyTb7NxCvnZ2D7Trbw/mT6yqHAyr92eN/Ey28P5g+srSxPm1xOo0HX+sqfD4o9pp7ly3uGFwKepQpaGR8dfZ7Vcp2y3C3wVL2DdaZByHYuo6W0yR+Q6P+6V6bOfFc+tfSnJLU2a07S3nLlSpxb4I8Y6W01n8iUmPgoNLaaP6EpPQV65zzQcOaeUn0v5ny7C191H/qvsea3S2lwONjpPQVgu2Gz2i22qimttvhpXvnLXFg5jdyrODlX228fyHQn/ej9ErZtKk3Wjm2QmkNnbww2rKNOKaXQulFSIiLoSoAiIgLl2Le9Ko+NH1BZqVhexb3pVHxk+oLMyVzd1z0uJduAerKHwgrDtr/ALzj8ZZ9azBYftg95x+Ms+tYtuejxM456urfCymDzUKTzULpSkQpHNQg5oC19in5JuHxhv0VYAVf7FONquHxhv0VYAXOXnPyLo0a9WUeD72c2hfJPXwR3WK2vOJZonSR5/OwcEfPlfTknyKutrdTPb71ZK6ncWyRNe4HtwW8F50KXlZ8k28WvvwFs7jLNJrPg2kyxg0812ReK8HqC+TT9wp7xaae4QEbsrckdjhzC+hzsleTTTyZvQnGrFSi809aKo2yWMU9ybeqaPdhqTiYD82Tt86wGmifNM2KNpc97g1rR1krYe926C7Wue31LfEmbjP6p6isA2a6SnpdQ1VZcocNoJDHECOD5P1h3AYPnCl7a8UaL5W2JXmOaN1KuJQdFflqPX1Pf911mbaSsTLDYIKMNHhiN+Y9rzz9HJeqeC7nO3xg9S4FvWomUnJuUiwrejChTjSgsklkj5LrcKe1W2e4VLsRwsLsfrHqA8q69KVk1z0/R11QR4SeIvOBwHE8FXG2K++Fq47HTvzHB48+Ot/UPMs72fOPtQtg6vYwWxOhyKCm9rf0Ia3xT8VitS2g/wAsI/OWaz+Wz5nsnig5O8iArkwZDvItVk8UXtR9+dX8Fn0QsYWUbU/frWfBZ9ELF109vzUeCKLxf9/W+KXeSFsBs6qG1mjbfJkEsjMbvK0rX8K2NiF1aaKttT3eNG7w0Y7jwK1cSg5Us1uJzQ26VG/8m3/emu1a/AsHPWuOUJyoCgy2TjPCyoppqeUZZLGWOHcRha6XaimttyqKGdpbJC8tOfmK2Qb5FiO0HRTdQRivt+4y4sGHAnAlA6s9q3rG4VGWUtjOT0qweeIUI1KKznDd0p7e3oKTRfZc7dXW6odBXUstPIDye3GfJ2r5MFTqaazRU84ShJxksmiERSsnyQiIgJUKVCALK9lMNNUaxgiqqeKeIxvJZI0OB4dhWKLLdkvvzg/dSepeNxzUuBJYMk8Qop+0u8uhtJQ07v4rRU0ORx3IgF3Mc4dah/MeRS0YGVzOee0vFRjBZRWRJ480ITknJYMnEjCwnavfhbrI62wv/jNaN1wB4tj6z5+XpWcPLxDIYmtfKGksa44BOOAJ6lrxqqouNVfaqW6hzavfLXsIxuY6h3LesKCqVM3uOW0rxOdnaeTgtc9We5Lf2vd8zylIUKVPlSEjmr92en8S7b+4+sqgVfmz0n2mW39x9ZUbifNridtoO/8AWVPh8Ue1hSBnyIBldrG8FClnZmM6n1bb9O1kVNWwVEjpY98GMDlnHWvKG1Cxf7JW+hv2rH9t4xfKH4t/iKr5S9vZUqlNSe1lcYzpPf2l9Uo02uTF6tXUXH/CfYeujrvQ37VwftOsXVR13ob9qp/J7UK9/N9HrIz0xxLpXyLeG06x/wCyVvob9qxraJq+3aitlNTUcNRG+KYvcZAMEYwsFRfcLKlCSktqNa70nv7ujKhUa5MtuoIpRbZzxClEQFybFvenUfGT6gszPPisM2Le9Oo+Mn1BZmea5u656XEu3R/1ZR+EkcOaw7a/7zj8ZZ9azBYjtf8AeafjEf1rFtz0eJ9Y56urfCyljzUKTzRdKUgEHNOKBAWtsT/JNw+MN+irAKr/AGJ/km4fGG/RVgtC5y85+RdGjPqujwfeyWD0KttuRxNaf3cnrarLHLuVY7dHfhrT8CT1tX1Y8/H/ADceWlj/AKVU/wDn/wBI+PZFqIUNyfaKmQiGqOYSTwbJ2ef14Vr8z3rWmNzmSNexxa5pyCOYK2A0Beo79p6OpcR7KjHg5x/WHX5+a2MRocl+UW/aQ2hmL+UpuyqPXHXHhvXZ3cD2GZB71MjiUzjh1riclRZ3RIK+HVl4hsWnJrg/BkHixNJ9048l9ipvapqE3W8+wKeTNJRktGDwc/rP1LZtaHlqiT2byFx/FFhtpKov73qjx6ezaYjWTyVVTJUzPL5JHFz3HrJV86BH4pWz4qFQQWwGgx+KFr+KhSWJrKnHLpON0HbleVW9vJ8UeoDxXfF+d5F0rshON7yKEZZxR+1T361fwWfRCxZZTtTP46VnwWfRCxZdNbczHgii8Y/f1vil3sjqXraSu77Hfqevbksad2Vo62Hn9vmXk8VIGV6yipJxexmlQrTo1I1IPJp5o2Rppo6injqIHtfFI0OY4ciCu0BYFsanu0ltmhnYHW2I/gZHcw/raO0epZ/lczWpeSm457C8sMvfx1rCvycs1s/zd0Ehc2vIHilcBx5KQvI3jrq44qpng6mGKdh/NkYHD51jmq9Oafj05cKiO0Ucc7Kdz2vazBBxzCyYjHFeTrIkaWuXH+iP9S9aUpKSSZoYjQo1LecpxTaT2pdBr2iIuoKJIREQEqERAFluyXHt0gyQB4KTiT3LElyY5zTlpIPaCvirDlwcek2rK5/C3EK2WfJaeXA2bazeaHAgjHUhY4dS1xorpcKN4fTVtTC4dbJCFmml9pdwppGU95PsqnPDwoGJGd/eoaphtSKzi8yyLLTS0rzUa0XDPftXg/oWwSQuOSuqlnjqoGVFPIJYpWhzHg5BC7MEKP2HZJprNHId2Vgm13TgqrcL7TR/h6cBtRge6Z1HzeryLPGhdkkUVTQ1FNM0OjljLHA9YPBelGq6U1NGhidjC/tpUJ79nU9zNYnDC4r7b1SOt90qaF+d6CVzPLg8F8S6dPNZoo2pBwk4y2oK+9nY/E63fuPrKoVX3s5B9pluP9gfWVHYnza4nY6EfvKnw+KPfAXIHAXEHhlQTlQpZxUu245v1F8W/wARVfrPttX5covi/wDiKwFdHZ8zEpfSL1nW4+CCIi2SFClEQBERAERQgLl2L8NJT/Gj6gszPNYTsX96lR8aPqCzgDiubuuelxLs0f8AVlH4QAsP2w+80/GGfWsx5LDdsPvOPxhn1rFtz0eJ9Y56urfCyl0RF0pSIRE60Ba+xIZtVw/ft+irBVfbESPvTcf37foqwCVzl5z8i6NGvVdHg+9nLPBVhtxH4S0n+rJ62qzQq224gD71fBk/wr6sefj/AJuPHSv1VU7P/SKyHBZJoDUb9P3pr3uPsSfDJ293U7zLGkU9UgqkXGWxlS2tzUta0a1N5Si8zZVjxI0Pad5pGQRyIXIZ9KwjZZqWmnsDrdcayGCak8VjpXhu+zq59nL0LLG3e0DJN1osAZ4TN+1c3UoyhJxa2F12WJ0Lu3jWjJLNbM9nSjx9o18+8WnniNwFXU5jh7R2u83rwqLJJJJOSVkOvL86/X6WoBPseP8ABwN7Gjr86x1TlnQ8jT17WVZpHivnC8bi/wAkdS8X292RyaeK2C0F70LX8VC17WwGgDjSFr+KheGJ82uJLaDfu6nw+KPZI61LeAd5EHEKcHDvgqFLNzKL2oHOtKzyM+iFjKyXad79K3/k+iFjS6a35qPBFF4r++rfFLvZIC9LTtrmvF3prdBwfM8Aux7lvWfMF5gKtPYVbWOnrrtIBljfAxeU8SfUsXNXyVJyPTBrH8dewoPY3r4LWywKGip7bb4qClZuQwtDWjt7z3rn1ldzhkldZauabz1l4QjGEVGKySAK5tyeQXAArHtYawpdNQCPHh6yQZZCDjA7XdgX1CEpy5MVrPK6uaVrSdWtLKK3mSljj1LxtbAN0tct4gfxV/X3KobxrfUdzc4SXB8EZ/m4PEA844/OvBmqqiYfhZ5ZM8955KkqWGzTTkzhr7TShOEqdKm2mms28tvVrOhQiKYK6CIiAIiIApUIgJ6kUIgLU2LXh8tNVWeZxJhHhYcn80nDh6celWK1UfstqTTa0oxnAm3oj5x9uFeOAOtQF/TUKza36y3tELuVxhyjJ64Nrs2rvyOQ7FyBw0+RcBhRlaR05Se1ul9j6znkAw2djZPKSOPqWJKwtt8OLvQT493AW+h3+ar0d66S0lyqMWUlpBS8liVaK6c/nr8QFf2zwfiRbDyzCfWVQK9+g1jqGhoYqKlrzHBEMMbuDgPQvi8t5V4JR6TZ0cxajhlxKpVTaay1cV1ovfHeFybGT1hUX7etUftN39xv2Lk3XmqQPym7+437FH+bKvSjslptYb4y+S+57W21u7e6EZ/o3+Iqvl6V8vNxvM7J7jUGaRjdxpIAwF5qlbem6dNRltRX2L3dO8vKlemmlJ79uwIiL2I0lFClAFCIgCIiAuTYqB7VZ/jR9QWcED9YLXm06ivNqpjTW+ukgiLt4taBzX2e3XU5/SsvoH2KJrWFSpUck1rLDwzS60tLSnQnCTcVlqy+5fIYTwDhxWIbYmbujeef4yz61Ww1tqcfpWb0BfJddS3q6UvsWur5Jod4O3XYxkLFLD6kJqTa1GcS0vs7q0qUYwlnJNa8vueQeahEUuV2EREBa+xFpNpuJ6vDt+irCDOPulrxar7dbVC+K31slOyR284N6yvtGsdSj9LVHpCiq9hOpUck1rO/wnSy1srOnQnCTcejLp4l/eDw3OeKq/blnetPZiX1tWJHWepj+lp/mXn3e8XO7+C++NXJUeCzub3Vnn6ktrGdKqpto88a0ptb+ynb04STeW3Lc0+k+BEUKVOEJJUIiAnmoREAWwOgmkaPtXAjNKFr+vXpdTX6kpo6emulRHDG3dY0O4Adi1Lu3lXilFnQaO4vSwuvKpUi2mstXE2BAOV3RxOO9n9Va+e27Un7Xqf7y7Y9ZalHK8VI86j/ADZU6UdetOLPfTl9Pud+1Ju7ratHwPohYuvruNbVXCrfVVk7ppn+6e7mV8vBTFKDhCMXuRXV7XjcXNSrFapNv5shXpspgFLo6jOMGYvlPfk4HqVFrYXSMYh0xbIxjApW/OMrQxOWVNLrOs0GpKV5Um90e9o9fOetOa4A8e5c2lQhaDR11czKSknq5eEcMbpHeQDK1zvNwnulznr6lxdJM8u49Q6h5leO0ioNNoe4uBwZGCP+84A/NlUI4DKmMMguTKfYVtpzdSdWnbp6ks+16vD6nFERSpwQREQBERAEREAREQBERAevo+TwOp7bJ2VLPWthXDjwWuNidu3micOqoZ9ILY93WobFF+aLLK0Fl+hVj1ruOHqUISoUYd2VrtwaP5Lf3SD1KsirP23D+K2x39eQfMFWHNdBYcwv83lOaVrLFanZ/wCUQinC2l2CdF6x7SNmFu1dV6quFBNVvla6COmY5rdx5bwJOepbhzpq1gqSMLd93Qj09jxdd3QeWiZ9qxTWnQrvlJSPm0tq2kuMrRkU9ZCYXP7g4EgedAalZUL19X6avukb9PYtR2ye3XCA+PDK3BI6nA8nNPURwXkIApQAlTjHNAcUXLC4oAiIgCIuWEAATGAvX0bDBUattEFTE2WCWuhZIx3JzS8Ag+ZbddObZxobSGy+3XDTOmLdaqqS6NjfLTx7ri0scceTgEBpcVxREARFIGetAQi5bqbpQEIVCIApCIgIRcg0oQUBxRSVCAKVCkICVC5EJulAcEUkYWZbHNnd22n6xGmLLVUlNVGmkqPCVJIZusxkcAePEIDDQUWa7ZdnF32XasZpu9VdHVVL6VlTv0ziWhriQBxA4+KsKCAkDLgAtjbSzwdpo2Yxu07B/wBIWusPu2+ULZOlb/EKfhyiZ6gonFXqj2lgaCJcus/h8SMrm13FcFIUQWOYttfk3dFuby36iMes/UqRdzVw7ZZMaWgb21TfouVPHmp7Dl+j2lSaZyzxLLoivEhERb5yYREQBERAEUogIREQBERAfVazu3GmdnGJWn5wtkd4FgOVrRTu3J43djgfnWyVOd+Frs5ywH5lEYotcXxLE0El+WvH4fEnPWpUKQoosEr7bYzNnoJOydw9Lf8AJVSFcu2SmL9JxzAZ8HUtPpBCptT2HvOiVFphDk4nJ9KT8PADmv0X6IDnN6KdG5pLXCOuII5g771+c+cFfor0QXf6KVGP7Ov+m9bxy5+f819vQd+WLgME/wBJf9qs3Yr0gdcaCv8ARtr71W3fT/hWtqqGqlMu7GTxdGXZLHAcQAcHkQqekOXnylcoIpaieOCCN0ksjgxjGjJc4nAAHWUB+gfTT0Rada7GH6zoY4nV9niZWU1S0cZaZ+N5pPWMEOHeFplsa2Yak2paqbY9PxMa1jfCVVXLkRU0ecbziPmA4lb5bVo26P6I1bbrq8Cak03BQOzx3pfBsZj0rEfueNtpqbY7dblEyMVVXdXskkxxxHGzdB7hvOPnKwZPJoeivsZ03TxUWstZ1ctylZnLq2KlHlawgnHlJVf7dOilJp3TlRq3Z9eJL3a6eMzzUk266ZsQ4l7Ht4SADJIwDgdaybWfRkOp9TXC/XrbRa6isq53yPdLCCWgk4aMy8ABw8yt7YFp+h2XaXrNO3faVab9QPlD6ZskrGeABGHNGXnLT2IYNZ+iRsj2bbV7NeKfUVTd4L3bpWu3KWqaxr4HDAcGlhOQ4EHj1hUjtS0nV6G2gXnStaHeEt9S6Njj+fHzY/ztIPnVl7MdWUeyvpU1ktvqYzYPvxUW+UxvBjdSvlLWkEcCG+Kc9ytP7ofoJz7jp/X1rp/CGsxbavwbc77+LoXecbzc9zUMlbdEnYdQbWKi83DUU1dTWaga2GN9K8MdJUO44yQeDW8xj84LF9ebO7dXbdajZ1stirrk2Kf2K19TK15dK3/WuLgAGsacg5/VK27p/AdHjoml8gZFeRS5APOSvnHAd+7nJ7mFVV9zlt0NbrTVmoKv8NWw0scbJH8XAyPJe7Pad0ZRBnv2Ton7NNLWeCq2la1lNVIPGDKmOkgDusN3gXO8vDyLzde9ErTl007Ne9kuqXXB8bSW0dROyZkxA4tbK0Ddd2AjzhZXt26P0m0faFXagu21a3UTMiOmoJYQRSxgAbvGUcScuJwMkr2ujvswg2P3ysqjtUtV0ttXTmOWh8WNu+CC14PhDgjiOXJxWTBoppymnt+ubbSVkL4KmnuUUcsTxhzHNkAII7QVu790SIOyC1j/AN4YP/5uWvXSiprPD0pX1FjqaWop62po6l7qeQPb4V27v8RwzkZ86v77ogSNklq/40P+25AapbC9kGotrOo5KC0FlJb6XBrrhK0mOAHkAPznHjgd3Uto4ejJsM0/HDbtS6kqZbjI33VRdI6dzietrAOA7M58pXvdDqnp7B0Wje6CKOSrmFdXTY4l8ke81rT5mDh3ntWhOprzc9Q36svN4q5quuqpXSSySuySSc47h3IgbI7eeilLpnT1RqrQFyqbvb6eMzVFDOA6dkYGS9jmgB4A4kYBxyyqN2I6Zt2sdqun9MXZ87aK4VQimMLg14bgngSCBy7FuV9z/wBS3bUOy662e7zSVVPaa0Q0r5SXERPYCY8nqBzjuK132b2qlsfTUpLTQNa2lpdSTxwtHJrAX4aPIOHmQF4ag6HWj49U01TTX6ttml6emMlc+onY+Z8m9wDXFoaxuOZOfIvTpejZ0f8AVFJNQ6YvlQ+sjGDLR3Zsz2ntcwggjzDzLw/ujGpbxQ2HTenKOolht9xfNNV7hIEpj3Q1p7R4xOO4LUDQmortpbVduvdlqpaaspp2OY6N2N7iMtPaDywgM26QWxm+bI9RRUldMK611m86hr2M3RKBza4fmvGRkd4IVn9EzYNojajoO53rUkt2bV09caeMUtS2Nu7ug8QWnjkq8Om5SU956NBu9XCxlTTVFHVw8OLHvIY4DzSH0LxvudIzsnvWf2sfoNQGN6Y6MeynTEEEO0nVbJbvVeMyl++DKaOME8APznfCOAexfHtj6IlvNo++2y+rqHVLd0m3VU4kZM09ccnDB68HIPaFrBtavNxv+0nUFzulVLUzvuEzQ57s7rWvIa0dgAAAC3i6FV7uFd0cfD11RJUutlTVQwF7skRsaHtbk9Q3iPJhAYXo7ot7L7BS0tNtG1iKi+VDQTTRVzKaNpPU3PjO8vDPYvE6QvRSt2ntJVmrNn1dWTRUMRnqrdVOEjjEBlz43gDkOODngDx6lq3rO93K/wCq7jeLrVzVVXUVL3ukleXH3RwPIFv90SrzW6g6LYN3qJax0DK2j3pTvOMbQcNJPPAOEB+eVDRVNfXQUNFBJUVNRI2OGKMZc95OA0DtJW3GguiNY7Xp6K+7WtVfe0vaC+kp5mRMhJHBr5XA5d3NHnKrnoQWiiunSGtrqtjZBRU09VE1w/nGtw0+bK2S6TGySt2papphWbSrXZLZRU7RDbJmZO+cl0rvHGSeAHDkB3rAMGv/AETtnWpbNLU7NNaSezIm+KJqhlVA93UHFoDmZ7ePkWoOrtO3jSWpa3T1+pH0lwopDHNG7t6iD1gjBB7Ct0timxQbLNc0uoaPbBZJ6YZZW0WAxtRERxafwhwRwIOOBAVWfdAn6frdo9ku9juFDWy1VtLKp9NM1+HMeQ3ODwOCsgsSHom6P1Dsy0/c9PV9wobzcaSkqaiapqBJDG17GulIYGgk8TgZ869awdG3o/zTCx+2qe6Xho3JGx3eNshd14Y0fNxWQ7Ur3cbL0LqavtdVJS1LtPUEIljOHNa+ONrsHq4ZHnX570FbU0NdFXUs74qmGQSRyNcQ5rgcg5QF69Kbo9zbKY6e/wBkr5rlp2pm8BmYDw1NIQSGvI4OBAOHYHLGO26ehdpHZlRW2zantN4E2tam2SisozWh5jaX4cfBgZHJvX196y3pUTuuvRDqLhV4knlpbfUOcePjmSPJ+c+la79AFg/h6dw/Q1T9KNYzBsZts0VsLv8AraOv2j3elpry2lZG2GW6Op8xAuLctB7SeK/PC5shjuVVHTEGBszxGQcjdDjjj18FfnT+aRt6zjH8k03rete1kHZBxmYP6w9a2XgG7SRDGCI2jHmWtduYZK+nYBkulaMedbJh3ihvYMKIxV5uC4lhaCR1V5fD4kO48VxXNQQoosMr/bUcafox21X+EqpVau21+LZbout0z3egD7VVSn8PX6C7SoNLpZ4pPqS7kERSt05khSihASoREBKhEQBERAFIREABwcrYvT0wqLLSTA536Zh/6QtdFfOzWcVOjaF+clkZjPmP2KMxOOcIs7jQarlc1KfTHP5P+T3m5K7AFACn1qHLLes8XaHSey9DXFgGXRtEo/5Tn1ZWv7ua2ZqGMno5aaQZZK0sd5DwK1xvdBLbLtU0EwIfDIW+UdR9ClsMnqlDtK404tWqlK4Wxrk/LWu9/I+Jfo50Mqf2V0Y7RS53PDmtjLuzeleM/OvzjXq2/UeobfTNpbffrpSQMJLYoKuRjBnngA4UqcE1mbdHoRwOPDX8nf8AxAf/AGWbbPuj1ss2O1UesNW6iiraqiPhIJ7lIyGCFw5Oaz85w6s5OeQytGW6y1f/APld9x/xGX/7L4LldbjcXiS4XCrrHgYDp53SEedxKGTYHpfbfabaPLFpTSjpRpukm8LLUPaWurZRyO6eIYOYzxJ48F6HQb2xWrRdzrtF6nrGUdqusrZqWrkOI4KjG6WvPU1wDePIFozzWsBPWoyUBuzta6IsOptT1OotCajoaKkuEhnfSzsL42Occkxvb+aSc4x51gu03ox2TZ/sjr79fdc0sd/geHwNc3cgnGP9S0cXF55gjzgDiNf7FrbWNhh8BZdU3q3wjlHT1sjGD/lBwvhv1+vd/qRU3y719zmHAPq6h0pA7BvE4QHmr9LejXf7Vta2FWM36GO4VVpmigq2S8SJ6dzXRyHvIDHechfmnhe1p/VeptPwSwWLUF0tkUzg6VlJVPia9w5EhpGSgNifuguvhetd0Oh6Gfeo7GzwtUGng6pkHX8FuB3EuWCdEfalSbMdpnsi8PcyyXSIUtbI0E+B45ZJgcwDz7iexU/cK2ruFbLW11TNVVUzi+WaZ5e97jzJJ4kroDsID9ANuPR30/thuTddaN1NSUlZXRtM0rQJ6aqAAAflpyHYABPHOBwysEoeh5Z7NYLtW601xBE5tMTT1EMfgoaZ4478hcfGHVjgtUtP6s1Np4uNh1BdLXve6FJVPiB8oaRlTqHVuqdQtDb7qO7XNg5Nqqt8jR5icICbFBHS67oKaGpiq44rnGxk8QO5KBIAHNyAcHnxW7f3RPA2SWoZ4m8j/tuWhUb3xyNkjc5j2EOa5pwQRyIXqXrUuor1EIrxfrpcYw7fDKqqfI3e7cOJ496A2n6CW2Gx2a3VGzbVVZDRxT1DprbUVDgIi54AfC4ngMkZGeByR1rJdfdDO23jVE100rqhtpt1VIZXUk1MZvBZOSI3Bw8XsBWjzTg5WT2raFrm1UQorZq+/UlM0YbFFXyNaB2AZ4IDfaruGz/ou7IjaoKxtVcnh0sMDnD2TX1BGN8ge5YMAZ5ADHErTzo7XCrvXSe03dq2TwlXW3d9TO7te8PcfnKq+43KuuVW+ruFZUVlQ85fLPKXvd5SeKsjoneN0h9H5/20/Qchhm9PSC0Lo/aXbKLRuoLpHbbvKZKi0yhwEm80AP3QfdjBGW9nHqVPbM+h7R2LVFNedW6liulHRyiZlJDTmNshacjfcXHxe4L4vukb5IKnRM8Ej4pWeySxzHEFpHg+II5Faq3DXut7hQGgrtX36ppXDddDLXyua4dhBdxQGy/Tp2yWe+0dNs30xWxVsNPUCe51MDgY99gIZC0jgcEknHDIaOoqwPudJH8E15/4s76DVoIvWs+pdRWendTWm+3S3wOdvOjpqt8TS7tIaQM96GTt10Pxzvg7LjUf9xy3j6D8f+jJcTj3VZW/9tq0FklfLI6WV7nyPJc5zjkuJ5klelb9Saht1CaG3X66UdISSYIKuRkZJ5ndBxxQHw3D8oVP713rK/QHoStx0XJieGamvPzL89ySTkkknrXrWzU2orbRewbff7rR0uSfAQVckbMnn4oOOPWgMl2O64n2dbUbXquKN0sdJOW1MTTgyQu8V7R34OR3gLdnarsr0F0jLNb9Yaa1JFBcGQCOKtgAkDmcSIpmZBBaSewjJ6l+dpeSSScr7rNfL1Zag1Fmu9fbZjzkpKh8Tj52kIDbrT/QzoLdNNW631xEbdDG5x9hwiHHD3TnvJAA5/WtTdbWy22bVlztdnu8V4t9LUOjp66JuGztHJw9WRwOOHBdl+1jq2/QeAvep7zcof8A06qtkkZ6HHC8JAfoDtub/oOUZGPyJbOXwY1+fy9ep1PqSptQtVTqG7TW8MawUslZI6ENb7lu4TjA6hjgvJQH6I9JAH/wYSHGMW63cMf2kS1R6IutLXojbfbLne52U1uqoZKGad58WHwmN1zuwbzQCerOVW9bq3VFdaxaq3Ul3qaDda32LLWSPiw3kN0nGBgYXjb3Yhg/RTb50crXtc1VQ6qptSvtkhp2Q1BZEJmTRAktcziMHxjx4g8F+f8ArOxzaa1ZdtPzyCSS3VktM545OLHEZ8+F20mqtTUlv+99LqG7wUmMeAirZGx4+CDheO9zpHFzyXOJySTklMjKPW0TSmt1XbYAMgztcfIOJ9Sv4HjlVVsWtT5LtPdntPgoGeDjPa93P0D1q1fzlBYjU5VXJbi19C7R0bB1ZL+959i1d+ZyGT5Vza3JwutuQcLvpxl/mK0MzriqtuLx7JtkA5tje88e0gfUq1Wb7ZKps+rXQtORTwtYfKeJ9awhdFZx5NCJSukVVVcTrSXTl8ll4BEymVtEIFClQgCIiAIiIApUKUARFCAkc1cOw+tZJZK63uPjwyiRo/quH2hU8FlezO8i0aogdK7dgqB4GUnkM8j6Vq3lN1KLS4k3o7eK0xCnOT1PU+3V35MvAjBwuJK5SEdS61zy2F0pHLOeCw3aPo832NtfQbra+NuC08BK0d/aFmPcpBJ4BfdKpKlLlRNS+saN9RdGss0/p1o1urqGsoJzBWU0sEgOC17SF04Wy01PTVTd2qpYJ2/2rA71qv8AbDZ7Pb7FTVFBb4KeZ9TuudG3BI3ScKXoYgqklFx1srjFNEZWVGdeFTOMdeTWsqjzKMo4qFInGhERAEREBKZREAymVCIAiIgCIiAlERAF9Furqy21sVdb6qekqoXb0c0Lyx7D2gjiF8yID1r9qS/38wm+Xq4XMwgiL2VUOl3M88bxOMryicqEQBERAFKhSgCIoQE5UIiAIiIApUKUAQIu+3Na+4U7HtDmmVoIPWMrDeSMxXKaR1NBJAHEnqCyXSujbve6hm7Tvp6bPjzytwAO7tKueK1Wqjx7EttLDg82xDK+1krsYzgKIqYm2soLIsez0GhCSlc1OUuhLL6/wfNabZSWa2xW+iYWxRjiTzcesnvXc7gVzc4k5JJKghRjbbzZ3VOEacVCKyS2HEBdsEjYnF7zhrQST3Lq68LGto94Fp0vUFj8T1I8DF28eZ8wys04OpJRW88ry5ha0J1p7IrMqDVFf98r/XVuciWZxb5M8F5eUULqYxUUkihqtSVWbnLa3n8yVCIsnmEREAREQBERAEREAREQEpkjrUIgLn2a6qju9uZbquQNuFO3HE8ZmjkR39qzALWymnmpp2TwSOilYcte04IKsvTG0qMxspr7C4OHD2TEOfwm/Yoe6sZJ8qns6CycA0rpOmqF48mtSlufHofWWUBlcwF8luuVuuUQkoK2CoBHJjxn0c19b8t5jCjcsnkzt41I1IqUHmuoE4WA7b5CbBSN/wB6/wABWdOOVgO2wH7x0Z6vZX+ArYtOeiQ+kXqutw8UVKiIuiKXCIiAIiICVCIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgC+i2/lGm/et9YXzr6Lb+Uab9631hYlsPun/euJsg88POjThJA7sPNccEdRXKn6CzO0FO7K4tO60ucQ1o5k8AvBv2tbDZ2OaaptVOOUUB3jnvPIL6jCU3lFZmvcXVG2hy60lFdZ7lVNDSU8lVUyNiijG857jgAKi9faidqG8umYCyki8SBh7Otx7z9i7NY6wuOon+DlPgKRpyyBh4eU9pWMlTVnZ+S/PPb3FY6SaR+cP0KHNra/af2IREW+cgEREAREQBERAEREAREQBERAEREAUjgiIDtimfG4OY9zCORacFepTap1BStDYLxWNA6nSbw+deMoXzKEZbVme1K4rUXnTm48G0ZPHrzVLf0mXfCiYfqXxX7U94vdNHT3GpbLHG/faBGG4OMdS8VF8KjTTzUVmbFTE7ypB051ZNPc28giIvU0QilEBCKVCAIiIAiKUBCKUQEIiIAiIgCIiAIpRAQilQgCIiAIiIAilQgCKVCAIiIAuynkdDOyVuN5jg4Z5ZHFdakIZTyeaMyn2laokyfD0zfJAF8c+vdUy8BcvBg/qRNH1LGFK8VbUl/tRJSxnEJLJ1pf8AZn3115utcT7NuNVOD1PkJHoXxE5XFF6pKOpGhOrOo+VNtvrChSoWTzCIiAIiIApREAUIiAlEUIAiIgClQpQBERAERQgCIiAIilAMIiIAiKEAREQBEUhAE4IhQBERAFCIgCIiAlERAE4IiAKERAEREAUohQBERAFCIgCIiAKURAEREAUKVCAIilAQilEAREQBFCIAiIgP/9k="
        )

        try:
            from PIL import Image, ImageTk
            import numpy as np
            pil_img = Image.open(io.BytesIO(base64.b64decode(FTM_LOGO_B64))).convert("RGBA")

            # Remove fundo preto automaticamente
            arr = np.array(pil_img)
            r, g, b = arr[:,:,0], arr[:,:,1], arr[:,:,2]
            mask = (r < 40) & (g < 40) & (b < 40)
            arr[mask, 3] = 0
            pil_img = Image.fromarray(arr)

            self.ftm_photo        = ImageTk.PhotoImage(pil_img.resize((120, 69), Image.LANCZOS))
            self.ftm_photo_footer = ImageTk.PhotoImage(pil_img.resize((80, 46),  Image.LANCZOS))
        except Exception:
            pass


    # ── Fonts ─────────────────────────────────────────────────────────────────
    def _load_fonts(self):
        self.fnt_nav    = FONTS["nav"]
        self.fnt_sub    = FONTS["sub"]
        self.fnt_body   = FONTS["body"]
        self.fnt_tiny   = FONTS["tiny"]

    # ── Build layout ──────────────────────────────────────────────────────────
    def _build_ui(self):
        self.main_frame = tk.Frame(self, bg=COLORS["bg_main"])
        self.main_frame.pack(fill="both", expand=True)
        self._build_sidebar()
        self._build_content_area()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    def _build_sidebar(self):
        self.sidebar = tk.Frame(self.main_frame, bg=COLORS["bg_sidebar"],
                                width=190, bd=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        # Red top bar
        tk.Frame(self.sidebar, bg=COLORS["accent"], height=4).pack(fill="x")

        # ── Logo area ────────────────────────────────────────────────────────
        logo_frame = tk.Frame(self.sidebar, bg=COLORS["bg_sidebar"], pady=22)
        logo_frame.pack(fill="x", padx=22)

        # "KM" em vermelho grande
        tk.Label(logo_frame, text="KM", font=("Georgia", 34, "bold"),
                 bg=COLORS["bg_sidebar"], fg=COLORS["accent"]).pack(anchor="w")

        # "C A R S" em branco
        tk.Label(logo_frame, text="C A R S",
                 font=("Helvetica", 10, "bold"),
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_sidebar"]).pack(anchor="w")

        tk.Label(logo_frame, text="FUJINOMIYA  ·  JAPAN",
                 font=("Helvetica", 7),
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_side_mut"]).pack(anchor="w")

        # Divider
        tk.Frame(self.sidebar, bg="#676E7E", height=1).pack(fill="x", padx=18)

        # Nav label
        tk.Label(self.sidebar, text="NAVEGAÇÃO",
                 font=("Helvetica", 7, "bold"),
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_side_mut"]).pack(
                     anchor="w", padx=22, pady=(16, 6))

        # Nav buttons
        self.nav_buttons = {}
        for page in SUBMENUS:
            self.nav_buttons[page] = self._make_nav_btn(page)

        # ── Tutorial button ───────────────────────────────────────────────────
        tk.Frame(self.sidebar, bg="#676E7E", height=1).pack(fill="x", padx=18, pady=(10,6))
        tut_btn = tk.Frame(self.sidebar, bg=COLORS["bg_sidebar"], cursor="hand2")
        tut_btn.pack(fill="x", padx=10, pady=(0,4))
        tk.Label(tut_btn, text="?", font=("Helvetica", 13, "bold"),
                 bg=COLORS["bg_sidebar"], fg="#F59E0B",
                 width=3, anchor="center").pack(side="left", padx=(8,0), pady=8)
        tk.Label(tut_btn, text="Tutorial", font=self.fnt_nav,
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_sidebar"],
                 anchor="w").pack(side="left", padx=4)
        def _tut_hover_on(e):
            tut_btn.configure(bg="#4A5568")
            for ch in tut_btn.winfo_children(): ch.configure(bg="#4A5568")
        def _tut_hover_off(e):
            tut_btn.configure(bg=COLORS["bg_sidebar"])
            for ch in tut_btn.winfo_children(): ch.configure(bg=COLORS["bg_sidebar"])
        tut_btn.bind("<Button-1>", lambda e: self._abrir_tutorial())
        tut_btn.bind("<Enter>", _tut_hover_on)
        tut_btn.bind("<Leave>", _tut_hover_off)
        for _ch in tut_btn.winfo_children():
            _ch.bind("<Button-1>", lambda e: self._abrir_tutorial())
            _ch.bind("<Enter>", _tut_hover_on)
            _ch.bind("<Leave>", _tut_hover_off)

        # ── Bottom: FTM Dev ───────────────────────────────────────────────────
        self.sidebar.pack_propagate(False)
        bottom = tk.Frame(self.sidebar, bg=COLORS["bg_sidebar"])
        bottom.pack(side="bottom", fill="x", padx=18, pady=14)

        tk.Frame(bottom, bg="#676E7E", height=1).pack(fill="x", pady=(0, 10))

        # Status online
        tk.Label(bottom, text="● Online", font=("Helvetica", 8),
                 bg=COLORS["bg_sidebar"], fg=COLORS["green"]).pack(anchor="w")
        tk.Label(bottom, text="v1.0.0  —  KM Cars JP",
                 font=("Helvetica", 7),
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_side_mut"]).pack(anchor="w")

        # FTM Dev watermark
        tk.Frame(bottom, bg="#676E7E", height=1).pack(fill="x", pady=(10, 6))

        ftm_row = tk.Frame(bottom, bg=COLORS["bg_sidebar"])
        ftm_row.pack(fill="x")

        # Logo FTM apenas
        if self.ftm_photo:
            tk.Label(ftm_row, image=self.ftm_photo,
                     bg=COLORS["bg_sidebar"]).pack(side="left")

    def _abrir_tutorial(self):
        """Abre janela flutuante do Tutorial com busca por substring."""
        # Se já existe, traz para frente
        if hasattr(self, "_tutorial_win") and self._tutorial_win and                 self._tutorial_win.winfo_exists():
            self._tutorial_win.lift()
            self._tutorial_win.focus_set()
            return

        win = tk.Toplevel(self)
        win.title("❓  KM Cars — Tutorial do Sistema")
        win.configure(bg=COLORS["bg_card"])
        win.geometry("760x600")
        win.minsize(500, 400)
        self._tutorial_win = win

        # Conteúdo completo do tutorial (seções separadas por ##)
        TUTORIAL_RAW = """
# KM Cars — Guia do Sistema

## ESTRUTURA GERAL
O sistema KM Cars é dividido em módulos: Dashboard, Compras, Vendas, Serviço, Parcelamentos e Configurações.

## CADASTROS BASE — PRÉ-REQUISITOS
Antes de registrar compras, vendas ou serviços:
• Clientes — cadastre nome e telefone. Sem cliente não é possível registrar compras, vendas ou OS.
• Carros — cadastre com nome, ano, cor, placa, chassi e status.

Status de um carro:
  Estoque  → carro disponível para venda
  Cliente  → pertence a um cliente, aparece em OS e Shaken
  Daisha   → veículo substituto emprestado
  Inativo  → arquivado, não aparece em nenhuma seleção

## FLUXO DE COMPRA
Passo a passo: Cadastrar Carro → Registrar Compra → (opcional) Adicionar Custos

Tipos de compra:
  Compra Direta  → compra simples de fornecedor
  Troca          → carro entrou como parte de uma troca
  Leilão         → gera custos automáticos (taxa exposição e kensa)

Marcar como \"À Venda\":
• Na lista de compras, clique em \"À Venda\" para disponibilizar no estoque.
• Carro não marcado como \"À Venda\" NÃO aparece ao criar uma Venda.

NFI (Nota Fiscal de Importação):
  SNF = Sem Nota Fiscal
  CNF = Com Nota Fiscal → gera IS automático sobre o valor

## FLUXO DE VENDA

Venda Simples:
• Carro marcado \"À Venda\" + cliente cadastrado

Venda Parcelada:
• Requer número de parcelas, entrada e valor da parcela mensal

Venda com Troca ⚠️ (mais complexa):
  PRÉ-REQUISITOS OBRIGATÓRIOS:
  1. Carro principal: cadastrado + marcado \"À Venda\"
  2. Carro de troca: cadastrado em Carros
  3. Carro de troca: compra registrada no sistema (mesmo simbólica)
  4. Carro de troca: status Estoque ou marcado \"À Venda\"

  PASSO A PASSO:
  1. Cadastrar carro principal → registrar compra → marcar À Venda
  2. Cadastrar carro de troca recebido do cliente
  3. Registrar compra do carro de troca (tipo \"Troca\")
  4. Nova Venda → tipo \"Com Troca\" → selecionar ambos os carros
  5. Informar valor da troca (deduzido do total)

  Se o carro de troca não tiver compra registrada, NÃO aparecerá na lista.

Venda Leilão:
• Gera automaticamente custos de Taxa de Exposição e Taxa de Kensa

IS Automático em Vendas (CNF):
• Se CNF, o IS é calculado sobre o valor de venda
• Registrado como custo na compra vinculada
• Se o valor for editado, o IS é atualizado sem duplicar

## FLUXO DE SERVIÇO (OS — Ordem de Serviço)

Pré-requisitos:
• Carro com status \"Cliente\"
• Carro com cliente vinculado (em Carros Cadastrados)
• Cliente cadastrado no sistema

Como criar uma OS:
• Serviço → Ordem de Serviço
• No campo Cliente: digite para filtrar. Os carros se filtram automaticamente para o cliente selecionado.
• Apenas carros com status \"Cliente\" aparecem na seleção.

Status de OS:
  Em Andamento   → em progresso
  Concluído      → finalizado (entra nos cálculos de receita e lucro)

Apenas OS com status \"Concluído\" entram nos cálculos de receita e lucro.

## FLUXO DE SHAKEN

Shaken é a inspeção veicular obrigatória no Japão.

Por status do carro:
  Cliente  → gera receita (cliente paga). Custos vão para custos_sk
  Estoque  → custo de preparação. Vai para custos da compra
  Daisha   → registrado, sem cobrança

Opção \"Por Conta\":
• Sem data de vencimento, não aparece nos alertas.

Alertas — Dashboard \"Shaken a Vencer\":
• Exibe carros com vencimento em até 90 dias ou vencidos
• Carros sem shaken registrado também aparecem para controle

## CUSTOS E IS

Tipos de custo por tabela:
  custos     → vinculado à Compra
  custos_os  → vinculado à OS
  custos_sk  → vinculado ao Shaken

IS (Imposto sobre Serviço):
• Calculado quando NFI = CNF
• Percentual configurável em Configurações → IS
• Gerado sobre: valor da venda, OS ou Shaken
• Anti-duplicação: ao editar o valor, o IS existente é atualizado, não duplicado

Nos relatórios:
  Custo s/IS  = custos sem IS
  IS (CNF)    = apenas o imposto
  Custo Total = Custo s/IS + IS

## DASHBOARD E RELATÓRIOS

Visão Geral:
• KPIs do mês: estoque, vendas, serviços, financiamentos, shaken
• Cards de receita e quantidade (Vendas + OS + Shaken) do mês e do ano
• Gráficos anuais com tooltip ao passar o mouse

Indicadores:
• Aba Receita: pizza de distribuição por tipo no mês
• Aba Lucro: pizza de distribuição de lucro por tipo

Relatório Mensal:
• Detalhamento por mês: Vendas, OS, Shaken
• Despesas Fixas + Resumo Consolidado com Lucro Líquido

Relatório Anual:
• 12 meses do ano com gráfico Receita Mensal e Lucro Mensal
• Valores negativos (prejuízo) em vermelho

Dossiê do Veículo:
• Histórico completo: compras, vendas, OS, shaken
• Dashboard → Dossiê Carro
• Se inativado: mostra data e motivo da inativação

Estoque Cliente (PDF):
• Dashboard → Estoque → Gerar PDF Cliente
• Mostra carros \"À Venda\" com Valor de Venda (preço a prazo) e data do Shaken
• Shaken ausente ou vencido → exibe \"2 anos\"

## TABELA DE AMARRAÇÕES RÁPIDAS

Registrar Compra:
  Precisa: Carro cadastrado + Cliente (vendedor) cadastrado

Registrar Venda simples:
  Precisa: Compra registrada + Carro marcado \"À Venda\" + Cliente comprador

Venda com Troca:
  Precisa: Tudo da venda simples + Carro de troca cadastrado + Compra do carro de troca registrada

Criar OS:
  Precisa: Carro com status \"Cliente\" + Cliente vinculado ao carro + Tipo de Serviço cadastrado

Registrar Shaken:
  Precisa: Carro cadastrado (qualquer status exceto Inativo)

Carro aparecer no Estoque (vendas):
  Precisa: Status \"Estoque\" + marcado \"À Venda\"

Carro aparecer nas OS:
  Precisa: Status \"Cliente\"

IS gerado automaticamente:
  Precisa: NFI = CNF + valor preenchido

Lucro OS aparecer:
  Precisa: OS com status \"Concluído\"

Despesa entrar no Lucro Mensal:
  Precisa: Despesa Fixa com mês de referência correto ou marcada como Recorrente

## ERROS COMUNS E SOLUÇÕES

Carro não aparece na venda:
  Causa: Não marcado como \"À Venda\"
  Solução: Histórico de Compras → botão \"À Venda\"

Carro de troca não aparece:
  Causa: Sem compra registrada no sistema
  Solução: Registrar compra do carro de troca

OS sem carros disponíveis:
  Causa: Nenhum carro com status \"Cliente\"
  Solução: Alterar status do carro em Carros Cadastrados

Despesas não aparecem no relatório:
  Causa: Mês de referência errado ou ausente
  Solução: Verificar campo \"Mês de Referência\" na despesa

Lucro OS zerado:
  Causa: OS não marcada como Concluído
  Solução: Alterar status da OS para \"Concluído\"

Shaken não aparece no alerta:
  Causa: Marcado como \"Por Conta\"
  Solução: Desmarcar \"Por Conta\" ou verificar data"""

        # Divide em blocos por seção para facilitar busca
        import re as _re
        secoes = []
        bloco_atual = []
        for linha in TUTORIAL_RAW.splitlines():
            if linha.startswith("## ") and bloco_atual:
                secoes.append("\n".join(bloco_atual))  # fixed
                bloco_atual = [linha]
            else:
                bloco_atual.append(linha)
        if bloco_atual:
            secoes.append("\n".join(bloco_atual))  # fixed

        # ── Barra superior ────────────────────────────────────────────────────
        top_bar = tk.Frame(win, bg=COLORS["accent"], height=4)
        top_bar.pack(fill="x")

        hdr = tk.Frame(win, bg="#1E293B")
        hdr.pack(fill="x", padx=0)
        tk.Label(hdr, text="❓  Tutorial do Sistema — KM Cars",
                 font=("Helvetica", 11, "bold"),
                 bg="#1E293B", fg="white").pack(side="left", padx=16, pady=10)

        # ── Barra de busca ────────────────────────────────────────────────────
        search_bar = tk.Frame(win, bg=COLORS["bg_main"],
                              highlightthickness=1,
                              highlightbackground=COLORS["border"])
        search_bar.pack(fill="x", padx=12, pady=(10, 4))

        tk.Label(search_bar, text="🔍", font=("Helvetica", 10),
                 bg=COLORS["bg_main"], fg=COLORS["text_muted"]
                 ).pack(side="left", padx=(8, 4), pady=6)
        _tut_search = tk.StringVar()
        _search_entry = tk.Entry(search_bar, textvariable=_tut_search,
                                  font=("Helvetica", 10),
                                  bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                  insertbackground=COLORS["text_primary"],
                                  relief="flat", bd=0)
        _search_entry.pack(side="left", fill="x", expand=True, ipady=5)
        _result_lbl = tk.Label(search_bar, text="",
                                font=("Helvetica", 8),
                                bg=COLORS["bg_main"], fg=COLORS["text_muted"])
        _result_lbl.pack(side="right", padx=8)
        tk.Button(search_bar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: _tut_search.set("")
                  ).pack(side="right", padx=(0, 4), ipady=2)

        # ── Área de texto com scroll ──────────────────────────────────────────
        txt_frame = tk.Frame(win, bg=COLORS["bg_card"])
        txt_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        vsb_t = tk.Scrollbar(txt_frame, orient="vertical")
        vsb_t.pack(side="right", fill="y")

        txt = tk.Text(txt_frame, font=("Courier", 9),
                      bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                      insertbackground=COLORS["text_primary"],
                      relief="flat", bd=0,
                      wrap="word", state="disabled",
                      yscrollcommand=vsb_t.set,
                      highlightthickness=0,
                      padx=14, pady=10,
                      spacing1=2, spacing2=1, spacing3=4)
        txt.pack(side="left", fill="both", expand=True)
        vsb_t.configure(command=txt.yview)

        # Tags de formatação
        txt.tag_configure("h1",    font=("Helvetica", 13, "bold"),
                          foreground=COLORS["accent"], spacing1=8, spacing3=4)
        txt.tag_configure("h2",    font=("Helvetica", 10, "bold"),
                          foreground="#F59E0B", spacing1=6, spacing3=2)
        txt.tag_configure("body",  font=("Helvetica", 9),
                          foreground=COLORS["text_primary"])
        txt.tag_configure("muted", font=("Helvetica", 8),
                          foreground=COLORS["text_muted"])
        txt.tag_configure("hl",    background="#854D0E", foreground="white")
        txt.tag_configure("sep",   font=("Helvetica", 5))

        def _render(content_str, highlight_term=""):
            txt.configure(state="normal")
            txt.delete("1.0", "end")
            for linha in content_str.splitlines():
                if linha.startswith("# "):
                    txt.insert("end", linha[2:] + "\n", "h1")
                elif linha.startswith("## "):
                    txt.insert("end", "\n" + linha[3:] + "\n", "h2")
                elif linha.strip() == "":
                    txt.insert("end", "\n", "sep")
                elif linha.startswith("  ") or linha.startswith("    "):
                    txt.insert("end", linha + "\n", "muted")
                else:
                    txt.insert("end", linha + "\n", "body")
            # Highlight matches
            if highlight_term:
                start = "1.0"
                count = 0
                while True:
                    pos = txt.search(highlight_term, start,
                                     stopindex="end", nocase=True)
                    if not pos:
                        break
                    end_pos = f"{pos}+{len(highlight_term)}c"
                    txt.tag_add("hl", pos, end_pos)
                    start = end_pos
                    count += 1
                _result_lbl.configure(
                    text=f"{count} resultado(s)" if count else "Nenhum resultado",
                    fg=COLORS["green"] if count else COLORS["red"])
                if count:
                    first = txt.search(highlight_term, "1.0", nocase=True)
                    if first:
                        txt.see(first)
            else:
                _result_lbl.configure(text="")
            txt.configure(state="disabled")

        # Renderização inicial
        _render(TUTORIAL_RAW)

        def _on_search(*_):
            termo = _tut_search.get().strip()
            if not termo:
                _render(TUTORIAL_RAW)
                return
            # Filtra seções que contêm o termo
            matches = [s for s in secoes if termo.lower() in s.lower()]
            if matches:
                _render("\n\n".join(matches), highlight_term=termo)
                _render(TUTORIAL_RAW, highlight_term=termo)

        _tut_search.trace_add("write", _on_search)
        _search_entry.focus_set()

        # Rodapé
        foot = tk.Frame(win, bg="#1E293B")
        foot.pack(fill="x")
        tk.Label(foot, text="KM Cars — Sistema de Gestão  ·  Use a busca para encontrar tópicos rapidamente",
                 font=("Helvetica", 7), bg="#1E293B",
                 fg=COLORS["text_muted"]).pack(side="left", padx=12, pady=6)
        tk.Button(foot, text="Fechar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat",
                  cursor="hand2", padx=10, pady=3,
                  command=win.destroy).pack(side="right", padx=10, pady=6)

    def _make_nav_btn(self, page):
        icon = NAV_ICONS.get(page, "•")
        frame = tk.Frame(self.sidebar, bg=COLORS["bg_sidebar"], cursor="hand2")
        frame.pack(fill="x", padx=10, pady=1)

        icon_lbl = tk.Label(frame, text=icon, font=("Helvetica", 13),
                            bg=COLORS["bg_sidebar"], fg=COLORS["text_side_mut"],
                            width=3, anchor="center")
        icon_lbl.pack(side="left", padx=(8, 0), pady=9)

        txt_lbl = tk.Label(frame, text=page, font=self.fnt_nav,
                           bg=COLORS["bg_sidebar"], fg=COLORS["text_sidebar"],
                           anchor="w")
        txt_lbl.pack(side="left", fill="x", expand=True, pady=9)

        indicator = tk.Frame(frame, bg=COLORS["bg_sidebar"], width=3)
        indicator.pack(side="right", fill="y")

        for w in [frame, icon_lbl, txt_lbl, indicator]:
            w.bind("<Button-1>", lambda e, p=page: self._show_page(p))
            w.bind("<Enter>",    lambda e, p=page: self._nav_hover(p, True))
            w.bind("<Leave>",    lambda e, p=page: self._nav_hover(p, False))

        return {"frame": frame, "icon": icon_lbl, "text": txt_lbl,
                "indicator": indicator}

    def _nav_hover(self, page, entering):
        if self.current_page.get() == page:
            return
        btn = self.nav_buttons[page]
        bg = COLORS["sidebar_hover"] if entering else COLORS["bg_sidebar"]
        for w in [btn["frame"], btn["icon"], btn["text"], btn["indicator"]]:
            w.configure(bg=bg)
        btn["text"].configure(fg="#FFFFFF" if entering else COLORS["text_sidebar"])
        btn["icon"].configure(fg=COLORS["accent2"] if entering else COLORS["text_side_mut"])

    def _nav_set_active(self, page):
        for p, btn in self.nav_buttons.items():
            if p == page:
                for w in [btn["frame"], btn["icon"], btn["text"], btn["indicator"]]:
                    w.configure(bg=COLORS["sidebar_active"])
                btn["text"].configure(fg=COLORS["accent"])
                btn["icon"].configure(fg=COLORS["accent"])
                btn["indicator"].configure(bg=COLORS["accent"])
            else:
                for w in [btn["frame"], btn["icon"], btn["text"], btn["indicator"]]:
                    w.configure(bg=COLORS["bg_sidebar"])
                btn["text"].configure(fg=COLORS["text_sidebar"])
                btn["icon"].configure(fg=COLORS["text_side_mut"])

    # ── Content area ──────────────────────────────────────────────────────────
    def _build_content_area(self):
        self.content_wrapper = tk.Frame(self.main_frame, bg=COLORS["bg_main"])
        self.content_wrapper.pack(side="left", fill="both", expand=True)

        # ── Top bar ───────────────────────────────────────────────────────────
        self.topbar = tk.Frame(self.content_wrapper, bg=COLORS["bg_card"],
                               height=56, bd=0)
        self.topbar.pack(fill="x")
        self.topbar.pack_propagate(False)

        self.topbar_title = tk.Label(self.topbar, text="",
                                     font=("Georgia", 13, "bold"),
                                     bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self.topbar_title.pack(side="left", padx=26, pady=14)

        right_info = tk.Frame(self.topbar, bg=COLORS["bg_card"])
        right_info.pack(side="right", padx=22)

        tk.Label(right_info, text="Administrador",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="right")
        tk.Label(right_info, text="👤  ",
                 bg=COLORS["bg_card"], fg=COLORS["accent"],
                 font=("Helvetica", 11)).pack(side="right")

        # ── Busca Global ──────────────────────────────────────────────────────
        search_frame = tk.Frame(self.topbar, bg=COLORS["bg_card"])
        search_frame.pack(side="right", padx=(0, 6))

        tk.Label(search_frame, text="🔍", font=("Helvetica", 10),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="left")
        self._global_search_var = tk.StringVar()
        self._global_search_entry = tk.Entry(
            search_frame, textvariable=self._global_search_var,
            font=("Helvetica", 9), width=22,
            bg=COLORS["bg_main"], fg=COLORS["text_primary"],
            insertbackground=COLORS["text_primary"],
            relief="flat", bd=0,
            highlightthickness=1,
            highlightbackground=COLORS["border"])
        self._global_search_entry.pack(side="left", ipady=4, padx=(4,0))
        self._global_search_entry.insert(0, "Buscar cliente, carro, OS...")
        self._global_search_entry.configure(fg=COLORS["text_muted"])

        def _gs_focus_in(e):
            if self._global_search_var.get() == "Buscar cliente, carro, OS...":
                self._global_search_entry.delete(0, "end")
                self._global_search_entry.configure(fg=COLORS["text_primary"])
        def _gs_focus_out(e):
            if not self._global_search_var.get().strip():
                self._global_search_entry.insert(0, "Buscar cliente, carro, OS...")
                self._global_search_entry.configure(fg=COLORS["text_muted"])
        def _gs_key(e):
            if e.keysym == "Return":
                self._global_search_execute()
            elif e.keysym == "Escape":
                self._global_search_entry.delete(0, "end")
                _gs_focus_out(None)

        self._global_search_entry.bind("<FocusIn>",  _gs_focus_in)
        self._global_search_entry.bind("<FocusOut>", _gs_focus_out)
        self._global_search_entry.bind("<KeyPress>", _gs_key)

        tk.Button(search_frame, text="↵", font=("Helvetica", 9, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=6, pady=3,
                  command=self._global_search_execute).pack(side="left", padx=(4,0))

        tk.Button(search_frame, text="⌨ F1", font=("Helvetica", 7),
                  bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                  relief="flat", cursor="hand2", padx=4, pady=3,
                  command=lambda: self._show_keybind_tooltip()
                  ).pack(side="left", padx=(6,0))

        tk.Frame(self.topbar, bg=COLORS["border"], width=1
                 ).pack(side="right", fill="y", pady=12, padx=8)

        # Red line under topbar
        tk.Frame(self.content_wrapper, bg=COLORS["accent"], height=2).pack(fill="x")

        # ── Sub-menu bar ──────────────────────────────────────────────────────
        self.submenu_bar = tk.Frame(self.content_wrapper,
                                    bg=COLORS["submenu_bg"], height=42)
        self.submenu_bar.pack(fill="x")
        self.submenu_bar.pack_propagate(False)

        # Border under submenu
        tk.Frame(self.content_wrapper,
                 bg=COLORS["submenu_border"], height=1).pack(fill="x")

        # ── Page container ────────────────────────────────────────────────────
        self.page_container = tk.Frame(self.content_wrapper, bg=COLORS["bg_content"])
        self.page_container.pack(fill="both", expand=True)

        self.pages = {}
        for page in SUBMENUS:
            frame = tk.Frame(self.page_container, bg=COLORS["bg_content"])
            self.pages[page] = frame
            self._build_page(frame, page)

        # ── Footer FTM ────────────────────────────────────────────────────────
        self._build_footer()

    def _build_footer(self):
        footer = tk.Frame(self.content_wrapper, bg=COLORS["bg_card"],
                          height=34, bd=0)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        tk.Frame(footer, bg=COLORS["border"], height=1).pack(fill="x")

        inner = tk.Frame(footer, bg=COLORS["bg_card"])
        inner.pack(fill="x", expand=True, padx=20)

        # Lado esquerdo — copyright
        tk.Label(inner, text="© 2025 KM Cars — Fujinomiya, Japan",
                 font=("Helvetica", 7), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(side="left", pady=8)

        # Lado direito — FTM Dev
        right = tk.Frame(inner, bg=COLORS["bg_card"])
        right.pack(side="right", pady=6)

        tk.Label(right, text="Desenvolvido por",
                 font=("Helvetica", 7), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(side="left", padx=(0, 6))

        if self.ftm_photo_footer:
            tk.Label(right, image=self.ftm_photo_footer,
                     bg=COLORS["bg_card"]).pack(side="left", padx=(0, 6))

        tk.Label(right, text="Fabio Takeda Maruyama",
                 font=("Helvetica", 7, "bold"), bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(side="left", padx=(0, 10))

        tk.Label(right, text="+81 080-6473-4174  |  fabiotm1985@gmail.com",
                 font=("Helvetica", 7), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(side="left")

    # ── Sub-menu ──────────────────────────────────────────────────────────────
    def _build_submenu(self, page):
        for w in self.submenu_bar.winfo_children():
            w.destroy()

        subs   = SUBMENUS[page]
        active = self.current_sub[page]

        tk.Frame(self.submenu_bar,
                 bg=COLORS["submenu_bg"], width=18).pack(side="left")

        for sub in subs:
            is_active = (sub == active)
            fg = COLORS["accent"] if is_active else COLORS["text_secondary"]
            bg = COLORS["submenu_bg"]

            btn_frame = tk.Frame(self.submenu_bar, bg=bg, cursor="hand2")
            btn_frame.pack(side="left")

            lbl = tk.Label(btn_frame, text=sub, font=("Helvetica", 9, "bold"),
                           fg=fg, bg=bg, padx=14, pady=12)
            lbl.pack()

            bar = tk.Frame(btn_frame, height=2,
                           bg=COLORS["accent"] if is_active else bg)
            bar.pack(fill="x")

            for w in [btn_frame, lbl, bar]:
                w.bind("<Button-1>",
                       lambda e, p=page, s=sub: self._switch_sub(p, s))
                w.bind("<Enter>",
                       lambda e, b=lbl, ba=bar, a=is_active:
                           self._sub_hover(b, ba, a, True))
                w.bind("<Leave>",
                       lambda e, b=lbl, ba=bar, a=is_active:
                           self._sub_hover(b, ba, a, False))

    def _sub_hover(self, lbl, bar, is_active, entering):
        if is_active:
            return
        lbl.configure(fg=COLORS["text_primary"] if entering else COLORS["text_secondary"])

    def _is_form_dirty(self):
        """Verifica se há dados não salvos nos formulários principais."""
        try:
            # Nova Venda
            if (getattr(self,"_venda_edit_id",None) is None and
                    (self._venda_carro_var.get() or self._venda_cliente_var.get())):
                return "Nova Venda"
        except Exception: pass
        try:
            # Nova Compra
            if (getattr(self,"_compra_edit_id",None) is None and
                    self._compra_carro_var.get().strip()):
                return "Nova Compra"
        except Exception: pass
        try:
            # Nova OS
            if (getattr(self,"_serv_edit_id",None) is None and
                    self._serv_carro_var.get().strip()):
                return "Nova OS"
        except Exception: pass
        return None

    def _switch_sub(self, page, sub):
        # Dirty check — só alerta ao sair de telas de formulário
        dirty = self._is_form_dirty()
        current_sub = self.current_sub.get(page,"")
        if dirty and current_sub != sub:
            form_pages = {"Nova Venda","Nova Compra","Ordem de Serviço"}
            if current_sub in form_pages:
                if not self._confirm_unsaved(dirty):
                    return  # usuário cancelou
        self.current_sub[page] = sub
        self._build_submenu(page)
        # Mostra sub-tela se existir
        if page in self.sub_pages and sub in self.sub_pages[page]:
            for k, f in self.sub_pages[page].items():
                f.place_forget()
            self.sub_pages[page][sub].place(in_=self.pages[page],
                                             x=0, y=0, relwidth=1, relheight=1)

    # ── Pages ─────────────────────────────────────────────────────────────────
    def _build_page(self, parent, page):
        self.sub_pages = getattr(self, "sub_pages", {})
        self.sub_pages[page] = {}

        # Sub-telas específicas
        if page == "Dashboard":
            self._build_dashboard_subs(parent)
            return
        if page == "Cadastros":
            self._build_cadastros_subs(parent)
            return
        if page == "Entrada/Compra":
            self._build_entrada_subs(parent)
            return
        if page == "Estoque":
            self._build_estoque_subs(parent)
            return
        if page == "Venda":
            self._build_venda_subs(parent)
            return
        if page == "Parcelamentos":
            self._build_parcelamentos_subs(parent)
            return
        if page == "Configurações":
            self._build_configuracoes_subs(parent)
            return
        if page == "Serviço":
            self._build_servico_subs(parent)
            return

        # Placeholder padrão
        self._build_placeholder(parent, page)

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: VENDA
    # ══════════════════════════════════════════════════════════════════════════
    def _build_venda_subs(self, parent):
        self.sub_pages["Venda"] = {}
        frame = tk.Frame(parent, bg=COLORS["bg_content"])
        self.sub_pages["Venda"]["Nova Venda"] = frame
        self._build_nova_venda(frame)

        frame2 = tk.Frame(parent, bg=COLORS["bg_content"])
        self.sub_pages["Venda"]["Lucro de Vendas"] = frame2
        self._build_lucro_vendas(frame2)

        frame.place(in_=parent, x=0, y=0, relwidth=1, relheight=1)

    def _build_nova_venda(self, parent):
        self._venda_edit_id    = None
        self._venda_tipo_var   = tk.StringVar(value="")
        self._venda_filtro_var = tk.StringVar(value="")
        self._venda_filtro_mes = tk.StringVar(value="Todos")
        self._venda_filtro_ano = tk.StringVar(value="Todos")
        self._venda_filtro_tipo = tk.StringVar(value="Todos")

        TIPO_OPTS = ["Venda a Vista", "Venda Parcelada", "Com Troca", "Com Troca e Volta", "Venda Leilão"]

        self.vendas_data = [dict(r) for r in
            self.conn.execute("SELECT * FROM vendas ORDER BY id DESC").fetchall()]

        root_frame = tk.Frame(parent, bg=COLORS["bg_content"])
        root_frame.pack(fill="both", expand=True, padx=14, pady=14)

        # ── COLUNA DIREITA (sidebar histórico) — pack antes do centro ─────────
        hist_panel = tk.Frame(root_frame, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        hist_panel.pack(side="right", fill="both", expand=True, padx=(8, 0))

        tk.Frame(hist_panel, bg=COLORS["green"], height=4).pack(fill="x")
        hh = tk.Frame(hist_panel, bg=COLORS["bg_card"])
        hh.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(hh, text="◆  Histórico de Vendas", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hh, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=6, pady=2,
                  command=self._refresh_tabela_vendas
                  ).pack(side="right", padx=(0,8))
        self._lbl_total_vendas = tk.Label(hh, text="0 registros", font=("Helvetica", 8),
                                          bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_total_vendas.pack(side="right")

        # Filtros
        fbar = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        fbar.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0, 4), pady=5)
        tk.Entry(fbar, textvariable=self._venda_filtro_var, font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"], width=16
                 ).pack(side="left", ipady=4)
        self._venda_filtro_var.trace_add("write", lambda *_: self._refresh_tabela_vendas())
        tk.Button(fbar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._venda_filtro_var.set("")
                  ).pack(side="left", padx=(3, 8), ipady=3)

        # Filtro mês/ano
        import datetime
        anos  = ["Todos"] + [str(y) for y in range(datetime.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        tk.Label(fbar, text="Mês:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._venda_filtro_mes, values=meses,
                     state="readonly", font=("Helvetica", 8), width=4
                     ).pack(side="left", padx=(2, 6), ipady=2)
        tk.Label(fbar, text="Ano:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._venda_filtro_ano, values=anos,
                     state="readonly", font=("Helvetica", 8), width=6
                     ).pack(side="left", padx=(2, 0), ipady=2)
        self._venda_filtro_mes.trace_add("write", lambda *_: self._refresh_tabela_vendas())
        self._venda_filtro_ano.trace_add("write", lambda *_: self._refresh_tabela_vendas())

        # Filtro tipo de venda
        fbar2 = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        fbar2.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar2, text="Tipo:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar2, textvariable=self._venda_filtro_tipo,
                     values=["Todos", "Venda a Vista", "Venda Parcelada", "Com Troca",
                             "Com Troca e Volta", "Venda Leilão"],
                     state="readonly", font=("Helvetica", 8), width=18
                     ).pack(side="left", padx=(4, 0), ipady=2)
        self._venda_filtro_tipo.trace_add("write", lambda *_: self._refresh_tabela_vendas())

        # Filtro NFI
        fbar3 = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        fbar3.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar3, text="NFI:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        self._venda_filtro_nfi = tk.StringVar(value="Todos")
        ttk.Combobox(fbar3, textvariable=self._venda_filtro_nfi,
                     values=["Todos", "SNF", "CNF"],
                     state="readonly", font=("Helvetica", 8), width=7
                     ).pack(side="left", padx=(4, 8), ipady=2)
        self._venda_filtro_nfi.trace_add("write", lambda *_: self._refresh_tabela_vendas())
        # Botões PDF e Excel
        tk.Button(fbar3, text="📄 PDF", font=("Helvetica", 8, "bold"),
                  bg="#C0392B", fg="white", relief="flat", cursor="hand2",
                  command=self._exportar_vendas_pdf
                  ).pack(side="right", padx=(4,0), ipady=3, ipadx=4)
        tk.Button(fbar3, text="📊 Excel", font=("Helvetica", 8, "bold"),
                  bg="#1E8449", fg="white", relief="flat", cursor="hand2",
                  command=self._exportar_vendas_excel
                  ).pack(side="right", padx=(4,0), ipady=3, ipadx=4)

        # Cabeçalho da tabela
        tk.Frame(hist_panel, bg=COLORS["border"], height=1).pack(fill="x", padx=14)
        col_f = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=14)
        for txt, w in [("#",3),("Cliente",13),("Data",9),("Carro",13),("Tipo",11),
                       ("V.Venda",9),("Entrada",8),("Parcela",8),("Saldo",9),("Status",8),("Ações",7)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=6)

        scroll_f = tk.Frame(hist_panel, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        cv = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(scroll_f, orient="vertical", command=cv.yview)
        self._vendas_rows_frame = tk.Frame(cv, bg=COLORS["bg_card"])
        self._vendas_rows_frame.bind("<Configure>",
            lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=self._vendas_rows_frame, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        hist_panel.bind("<Enter>",
            lambda e: hist_panel.bind_all("<MouseWheel>",
                lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
        hist_panel.bind("<Leave>", lambda e: hist_panel.unbind_all("<MouseWheel>"))

        # ── COLUNA ESQUERDA: formulário ───────────────────────────────────────
        form_outer = tk.Frame(root_frame, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=320)
        form_outer.pack(side="left", fill="y", padx=(0, 0))
        form_outer.pack_propagate(False)

        fcanvas = tk.Canvas(form_outer, bg=COLORS["bg_card"], highlightthickness=0, width=318)
        fsb = tk.Scrollbar(form_outer, orient="vertical", command=fcanvas.yview)
        fcanvas.configure(yscrollcommand=fsb.set)
        fsb.pack(side="right", fill="y")
        fcanvas.pack(side="left", fill="both", expand=True)
        fcard = tk.Frame(fcanvas, bg=COLORS["bg_card"])
        fcanvas.create_window((0, 0), window=fcard, anchor="nw")
        fcard.bind("<Configure>",
            lambda e: fcanvas.configure(scrollregion=fcanvas.bbox("all")))
        form_outer.bind("<Enter>",
            lambda e: form_outer.bind_all("<MouseWheel>",
                lambda ev: fcanvas.yview_scroll(int(-1*(ev.delta/120)), "units")))
        form_outer.bind("<Leave>", lambda e: form_outer.unbind_all("<MouseWheel>"))

        tk.Frame(fcard, bg=COLORS["green"], height=4).pack(fill="x")
        self._lbl_venda_title = tk.Label(fcard, text="◆  Nova Venda",
                                          font=("Helvetica", 11, "bold"),
                                          bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._lbl_venda_title.pack(anchor="w", padx=18, pady=(14, 4))
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 12))

        def lbl(txt, required=False):
            fg = COLORS["accent"] if required else COLORS["text_secondary"]
            tk.Label(fcard, text=txt + ("  ★" if required else ""),
                     font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"], fg=fg
                     ).pack(anchor="w", padx=18)

        # Carro em estoque
        lbl("Carro", required=True)
        self._venda_carro_var = tk.StringVar(value="")
        self._venda_carro_combo = ttk.Combobox(fcard, textvariable=self._venda_carro_var,
                                               state="readonly", font=("Helvetica", 10), width=28)
        self._venda_carro_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._venda_carro_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_carros_venda())

        # Cliente (obrigatório)
        lbl("Cliente", required=True)
        self._venda_cliente_var = tk.StringVar(value="")
        self._venda_cliente_combo = ttk.Combobox(fcard, textvariable=self._venda_cliente_var,
                                                  state="readonly", font=("Helvetica", 10), width=28)
        self._venda_cliente_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._venda_cliente_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_clientes_venda())

        # Tipo de venda
        lbl("Tipo de Venda", required=True)
        self._venda_tipo_combo = ttk.Combobox(fcard, textvariable=self._venda_tipo_var,
                     values=TIPO_OPTS, state="readonly", font=("Helvetica", 10), width=28)
        self._venda_tipo_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._venda_tipo_var.trace_add("write", lambda *_: self._toggle_campos_venda())

        # ── BLOCO A VISTA ─────────────────────────────────────────────────────
        self._bloco_avista = tk.Frame(fcard, bg=COLORS["bg_card"])

        tk.Label(self._bloco_avista, text="Valor de Venda (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        self._venda_valor_entry = self._make_yen_entry(self._bloco_avista, width=20)
        self._venda_valor_entry.pack(padx=18, pady=(4, 12), ipady=7)

        tk.Label(self._bloco_avista, text="Data da Venda",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        df_av = tk.Frame(self._bloco_avista, bg=COLORS["bg_card"])
        df_av.pack(anchor="w", padx=18, pady=(4, 12))
        self._venda_dia, self._venda_mes, self._venda_ano = self._make_date_row(df_av)
        # começa oculto

        # ── BLOCO TROCA (só Com Troca / Com Troca e Volta) — direto em fcard ──
        self._bloco_troca = tk.Frame(fcard, bg=COLORS["bg_card"],
                                     highlightthickness=1,
                                     highlightbackground=COLORS["orange"])
        tk.Frame(self._bloco_troca, bg=COLORS["orange"], height=3).pack(fill="x")
        tk.Label(self._bloco_troca, text="  Carro de Troca",
                 font=("Helvetica", 8, "bold"),
                 bg=COLORS["orange"], fg="white").pack(anchor="w", pady=3)
        self._venda_troca_var = tk.StringVar(value="")
        self._venda_troca_combo = ttk.Combobox(self._bloco_troca,
                                               textvariable=self._venda_troca_var,
                                               state="readonly", font=("Helvetica", 10), width=26)
        self._venda_troca_combo.pack(padx=14, pady=(4, 4), ipady=4)
        self._venda_troca_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_troca())
        self._venda_troca_var.trace_add("write", lambda *_: self._calc_parcelas())
        self._lbl_valor_troca_info = tk.Label(self._bloco_troca, text="",
                                               font=("Helvetica", 8),
                                               bg=COLORS["bg_card"], fg=COLORS["orange"])
        self._lbl_valor_troca_info.pack(padx=14, pady=(0, 8))
        # começa oculto

        # ── BLOCO ENTRADA (Parcelada e Com Troca, não em Troca e Volta) ────────
        self._bloco_entrada_wrap = tk.Frame(fcard, bg=COLORS["bg_card"])
        tk.Label(self._bloco_entrada_wrap, text="Entrada (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        self._venda_entrada_entry = self._make_yen_entry(self._bloco_entrada_wrap, width=20)
        self._venda_entrada_entry.pack(padx=18, pady=(4, 12), ipady=7)
        self._venda_entrada_entry.bind("<KeyRelease>",
            lambda e: (self._fmt_yen(self._venda_entrada_entry), self._calc_parcelas()))
        # começa oculto

        # ── BLOCO PARCELADO parte 1: Valor de Venda (direto em fcard) ───────────
        self._bloco_valor_parc = tk.Frame(fcard, bg=COLORS["bg_card"])

        tk.Label(self._bloco_valor_parc, text="Valor de Venda (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        self._venda_valor_parc_entry = self._make_yen_entry(self._bloco_valor_parc, width=20)
        self._venda_valor_parc_entry.pack(padx=18, pady=(4, 12), ipady=7)
        self._venda_valor_parc_entry.bind("<KeyRelease>",
            lambda e: (self._fmt_yen(self._venda_valor_parc_entry), self._calc_parcelas()))
        # começa oculto

        # ── BLOCO PARCELADO parte 2: Parcela, Calc, Data (direto em fcard) ─────
        self._bloco_parc = tk.Frame(fcard, bg=COLORS["bg_card"])

        tk.Label(self._bloco_parc, text="Parcela Mensal (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        self._venda_parcela_entry = self._make_yen_entry(self._bloco_parc, width=20)
        self._venda_parcela_entry.pack(padx=18, pady=(4, 8), ipady=7)
        self._venda_parcela_entry.bind("<KeyRelease>",
            lambda e: (self._fmt_yen(self._venda_parcela_entry), self._calc_parcelas()))

        self._lbl_calc = tk.Label(self._bloco_parc, text="",
                                   font=("Helvetica", 8),
                                   bg=COLORS["bg_card"], fg=COLORS["green"],
                                   wraplength=270, justify="left")
        self._lbl_calc.pack(padx=18, pady=(0, 8))

        tk.Label(self._bloco_parc, text="Data da 1ª Parcela",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        df_p = tk.Frame(self._bloco_parc, bg=COLORS["bg_card"])
        df_p.pack(anchor="w", padx=18, pady=(4, 12))
        self._venda_pdia, self._venda_pmes, self._venda_pano = self._make_date_row(df_p)
        # começa oculto

        # ── BLOCO VOLTA (só Com Troca e Volta) — direto em fcard ─────────────
        self._bloco_volta = tk.Frame(fcard, bg=COLORS["bg_card"],
                                     highlightthickness=1,
                                     highlightbackground=COLORS["blue"])
        tk.Frame(self._bloco_volta, bg=COLORS["blue"], height=3).pack(fill="x")
        tk.Label(self._bloco_volta, text="  Volta ao Cliente",
                 font=("Helvetica", 8, "bold"),
                 bg=COLORS["blue"], fg="white").pack(anchor="w", pady=3)
        self._lbl_volta_valor = tk.Label(self._bloco_volta, text="—",
                                          font=("Helvetica", 11, "bold"),
                                          bg=COLORS["bg_card"], fg=COLORS["blue"])
        self._lbl_volta_valor.pack(padx=14, pady=(4, 8))
        # começa oculto

        # ── BLOCO LEILÃO (só Venda Leilão) ────────────────────────────────────
        self._bloco_leilao = tk.Frame(fcard, bg=COLORS["bg_card"],
                                      highlightthickness=1,
                                      highlightbackground=COLORS["orange"])
        tk.Frame(self._bloco_leilao, bg=COLORS["orange"], height=3).pack(fill="x")
        tk.Label(self._bloco_leilao, text="  Taxas Leilão Venda",
                 font=("Helvetica", 8, "bold"),
                 bg=COLORS["orange"], fg="white").pack(anchor="w", pady=3)
        tk.Label(self._bloco_leilao, text="Taxa Exposição (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=14)
        self._venda_taxa_exp_entry = self._make_yen_entry(self._bloco_leilao, width=20)
        self._venda_taxa_exp_entry.pack(padx=14, pady=(4, 10), ipady=7)
        tk.Label(self._bloco_leilao, text="Taxa Kensa (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=14)
        self._venda_taxa_kensa_entry = self._make_yen_entry(self._bloco_leilao, width=20)
        self._venda_taxa_kensa_entry.pack(padx=14, pady=(4, 14), ipady=7)
        # começa oculto


        # ── NFI ───────────────────────────────────────────────────────────────
        tk.Label(fcard, text="NFI",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        _nfi_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        _nfi_f.pack(anchor="w", padx=18, pady=(4, 12))
        self._venda_nfi_var = tk.StringVar(value="SNF")
        for _lbl, _val in [("SNF", "SNF"), ("CNF", "CNF")]:
            tk.Radiobutton(_nfi_f, text=_lbl, variable=self._venda_nfi_var, value=_val,
                           font=("Helvetica", 9),
                           bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                           activebackground=COLORS["bg_card"],
                           selectcolor=COLORS["bg_main"]).pack(anchor="w")

        # Status e botões
        self._lbl_venda_status = tk.Label(fcard, text="", font=("Helvetica", 8),
                                           bg=COLORS["bg_card"], fg=COLORS["red"])
        self._lbl_venda_status.pack(padx=18, pady=(4, 4))

        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 10))
        btn_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        btn_f.pack(padx=18, pady=(0, 20), fill="x")
        self._btn_salvar_venda = tk.Button(btn_f, text="  Registrar Venda  ",
                                            font=("Helvetica", 10, "bold"),
                                            bg=COLORS["green"], fg="white",
                                            relief="flat", cursor="hand2",
                                            activebackground=COLORS["accent2"],
                                            command=self._salvar_venda)
        self._btn_salvar_venda.pack(side="left", ipady=7, ipadx=4, fill="x", expand=True)
        tk.Button(btn_f, text="Limpar", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_venda
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=4)
        tk.Button(btn_f, text="↺", font=("Helvetica", 10, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_tabela_vendas
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=6)

        self._refresh_tabela_vendas()

    # ── Helpers Venda ─────────────────────────────────────────────────────────
    def _atualizar_combo_carros_venda(self):
        """Carros em estoque que não têm venda registrada."""
        vendidos = {r[0] for r in self.conn.execute(
            "SELECT DISTINCT carro_id FROM vendas WHERE carro_id IS NOT NULL").fetchall()}
        # Se editando, exclui o carro da venda atual do filtro
        if self._venda_edit_id:
            row = self.conn.execute(
                "SELECT carro_id FROM vendas WHERE id=?", (self._venda_edit_id,)).fetchone()
            if row and row[0]:
                vendidos.discard(row[0])
        opts = [
            f"{c['id']} — {c['carro']} {c['ano']} | {c['cor'] or '—'} | {c['placa'] or '—'}"
            for c in getattr(self, "carros_data", [])
            if c["status"] == "Estoque" and c["id"] not in vendidos
        ]
        self._venda_carro_combo["values"] = opts if opts else ["Nenhum carro disponível"]

    def _atualizar_combo_clientes_venda(self):
        clientes = [dict(r) for r in
                    self.conn.execute("SELECT * FROM clientes ORDER BY nome").fetchall()]
        opts = [f"{c['id']} — {c['nome']}" for c in clientes]
        self._venda_cliente_combo["values"] = opts if opts else ["Nenhum cliente cadastrado"]

    def _atualizar_combo_troca(self):
        """Carros do tipo Troca (compra) não vinculados a outra venda."""
        vinculados = {r[0] for r in self.conn.execute(
            "SELECT DISTINCT carro_troca_id FROM vendas WHERE carro_troca_id IS NOT NULL").fetchall()}
        if self._venda_edit_id:
            row = self.conn.execute(
                "SELECT carro_troca_id FROM vendas WHERE id=?", (self._venda_edit_id,)).fetchone()
            if row and row[0]:
                vinculados.discard(row[0])
        compras_troca = [dict(r) for r in self.conn.execute(
            "SELECT co.* FROM compras co "
            "JOIN carros ca ON co.carro_id=ca.id "
            "WHERE co.tipo='Troca' AND ca.status != 'Inativo' "
            "ORDER BY co.id DESC").fetchall()]
        opts = []
        for c in compras_troca:
            if c["id"] not in vinculados:
                val = c.get("valor") or "0"
                opts.append(
                    f"{c['id']} — {c['carro']} | {c.get('cor') or '—'} | {c.get('placa') or '—'} | ¥{val}"
                )
        self._venda_troca_combo["values"] = opts if opts else ["Nenhum carro de troca disponível"]
        # Quando seleção muda, mostra valor e auto-preenche valor de venda se "troca e volta"
        self._venda_troca_var.trace_add("write", lambda *_: self._on_troca_selecionada())

    def _on_troca_selecionada(self):
        val = self._get_valor_troca_selecionada()
        sel = self._venda_troca_var.get()
        if val > 0 and "—" in sel:
            self._lbl_valor_troca_info.configure(
                text=f"Valor de troca registrado: ¥ {val:,}")
        else:
            self._lbl_valor_troca_info.configure(text="")
        self._calc_parcelas()

    def _toggle_campos_venda(self):
        tipo = self._venda_tipo_var.get()
        # Oculta tudo
        self._bloco_avista.pack_forget()
        self._bloco_troca.pack_forget()
        self._bloco_valor_parc.pack_forget()
        self._bloco_entrada_wrap.pack_forget()
        self._bloco_parc.pack_forget()
        self._bloco_volta.pack_forget()
        if hasattr(self, "_bloco_leilao"):
            self._bloco_leilao.pack_forget()
            self._venda_taxa_exp_entry.delete(0, tk.END)
            self._venda_taxa_kensa_entry.delete(0, tk.END)
        self._lbl_calc.configure(text="")
        self._lbl_volta_valor.configure(text="—")
        self._lbl_valor_troca_info.configure(text="")

        if tipo == "Venda a Vista":
            self._bloco_avista.pack(fill="x")

        elif tipo == "Venda Parcelada":
            self._bloco_valor_parc.pack(fill="x")
            self._bloco_entrada_wrap.pack(fill="x")
            self._bloco_parc.pack(fill="x")

        elif tipo == "Com Troca":
            self._bloco_troca.pack(fill="x", padx=18, pady=(4, 10))
            self._bloco_valor_parc.pack(fill="x")
            self._bloco_entrada_wrap.pack(fill="x")
            self._bloco_parc.pack(fill="x")

        elif tipo == "Com Troca e Volta":
            self._bloco_troca.pack(fill="x", padx=18, pady=(4, 10))
            self._bloco_valor_parc.pack(fill="x")
            self._bloco_parc.pack(fill="x")
            self._bloco_volta.pack(fill="x", padx=18, pady=(4, 10))

        elif tipo == "Venda Leilão":
            self._bloco_avista.pack(fill="x")
            self._bloco_leilao.pack(fill="x", padx=18, pady=(4, 10))

        self._calc_parcelas()

    def _get_yen_int(self, entry):
        try:
            return int(entry.get().replace(",", "").strip())
        except (ValueError, AttributeError):
            return 0

    def _calc_parcelas(self):
        tipo = self._venda_tipo_var.get()
        if tipo not in ("Venda Parcelada", "Com Troca", "Com Troca e Volta"):
            self._lbl_calc.configure(text="")
            return

        valor = self._get_yen_int(self._venda_valor_parc_entry)
        parcela = self._get_yen_int(self._venda_parcela_entry)

        if tipo == "Com Troca e Volta":
            # Calcula volta
            troca_val = self._get_valor_troca_selecionada()
            volta = valor - troca_val
            if volta < 0:
                self._lbl_volta_valor.configure(
                    text=f"¥ {abs(volta):,}  (pagar ao cliente)")
            else:
                self._lbl_volta_valor.configure(text=f"¥ {volta:,}")
            self._lbl_calc.configure(text="")
            return

        entrada = self._get_yen_int(self._venda_entrada_entry)
        if tipo == "Com Troca":
            troca_val = self._get_valor_troca_selecionada()
            saldo = valor - troca_val - entrada
        else:
            saldo = valor - entrada

        if parcela <= 0 or saldo <= 0:
            self._lbl_calc.configure(text="")
            return

        n = saldo // parcela
        ultima = saldo % parcela
        if ultima == 0:
            txt = f"Saldo ¥{saldo:,} ÷ ¥{parcela:,} = {n}x ¥{parcela:,}"
        else:
            txt = (f"Saldo ¥{saldo:,} ÷ ¥{parcela:,} = {n}x ¥{parcela:,}"
                   f" + última ¥{ultima:,}")
        self._lbl_calc.configure(text=txt)

    def _get_valor_troca_selecionada(self):
        sel = self._venda_troca_var.get()
        if not sel or "—" not in sel:
            return 0
        try:
            # Formato: "ID — Carro | Cor | Placa | ¥valor"
            parte_val = sel.split("|")[-1].strip()
            return int(parte_val.replace("¥", "").replace(",", "").strip())
        except Exception:
            return 0

    def _upsert_is(self, tabela, id_col, rec_id, valor_base, data):
        """Insere ou atualiza IS em custos/custos_os/custos_sk sem duplicar."""
        try:
            is_perc = float(self.conn.execute(
                "SELECT valor FROM configuracoes WHERE chave='is_percentual'"
                ).fetchone()[0] or 10)
        except Exception:
            is_perc = 10.0
        is_val = int(int(str(valor_base or "0").replace(",","")) * is_perc / 100)
        if is_val <= 0:
            return
        existing = self.conn.execute(
            f"SELECT id FROM {tabela} WHERE {id_col}=? AND tipo_custo='IS'",
            (rec_id,)).fetchone()
        if existing:
            self.conn.execute(
                f"UPDATE {tabela} SET valor=?, data_custo=? WHERE id=?",
                (str(is_val), data or "", existing[0]))
        else:
            self.conn.execute(
                f"INSERT INTO {tabela} ({id_col},tipo_custo,descricao,valor,data_custo) "
                f"VALUES (?,?,?,?,?)",
                (rec_id, "IS", "IS automático CNF", str(is_val), data or ""))
        self.conn.commit()

    def _salvar_venda(self):
        tipo = self._venda_tipo_var.get().strip()
        carro_sel = self._venda_carro_var.get().strip()

        if not carro_sel or "Nenhum" in carro_sel:
            self._lbl_venda_status.configure(text="⚠ Selecione um carro.", fg=COLORS["red"]); return
        if not tipo:
            self._lbl_venda_status.configure(text="⚠ Selecione o tipo de venda.", fg=COLORS["red"]); return

        # Cliente obrigatório (exceto Venda Leilão que não tem cliente)
        cli_sel = self._venda_cliente_var.get().strip()
        cliente_id = None
        if tipo == "Venda Leilão":
            if cli_sel and "—" in cli_sel:
                try: cliente_id = int(cli_sel.split("—")[0].strip())
                except Exception: pass
        else:
            if not cli_sel or "—" not in cli_sel:
                self._lbl_venda_status.configure(text="⚠ Selecione um cliente.", fg=COLORS["red"]); return
            try:
                cliente_id = int(cli_sel.split("—")[0].strip())
            except Exception:
                cliente_id = None

        # Extrai dados do carro selecionado
        try:
            carro_id = int(carro_sel.split("—")[0].strip())
        except Exception:
            carro_id = None
        partes = carro_sel.split("|")
        carro_nome = carro_sel.split("|")[0].strip()
        cor_sel   = partes[1].strip() if len(partes) > 1 else ""
        placa_sel = partes[2].strip() if len(partes) > 2 else ""

        # Dados por tipo
        valor_venda = entrada = parcela_mensal = ""
        num_parcelas = 0
        valor_ultima = ""
        data_primeira_parc = data_venda = ""
        carro_troca_id = None
        carro_troca = ""
        valor_troca = ""
        volta_paga = ""

        if tipo in ("Venda a Vista", "Venda Leilão"):
            valor_venda = self._yen_raw(self._venda_valor_entry)
            if not valor_venda:
                self._lbl_venda_status.configure(text="⚠ Informe o valor de venda.", fg=COLORS["red"]); return
            data_venda = self._get_date_from_entries(self._venda_dia, self._venda_mes, self._venda_ano)

        elif tipo in ("Venda Parcelada", "Com Troca", "Com Troca e Volta"):
            valor_venda  = self._yen_raw(self._venda_valor_parc_entry)
            parcela_mensal = self._yen_raw(self._venda_parcela_entry)
            data_primeira_parc = self._get_date_from_entries(
                self._venda_pdia, self._venda_pmes, self._venda_pano)
            data_venda = data_primeira_parc

            if not valor_venda:
                self._lbl_venda_status.configure(text="⚠ Informe o valor de venda.", fg=COLORS["red"]); return

            if tipo in ("Com Troca", "Com Troca e Volta"):
                troca_sel = self._venda_troca_var.get().strip()
                if not troca_sel or "Nenhum" in troca_sel:
                    self._lbl_venda_status.configure(text="⚠ Selecione o carro de troca.", fg=COLORS["red"]); return
                try:
                    carro_troca_id = int(troca_sel.split("—")[0].strip())
                except Exception:
                    carro_troca_id = None
                carro_troca = troca_sel.split("|")[0].strip()
                valor_troca = str(self._get_valor_troca_selecionada())

            if tipo == "Com Troca e Volta":
                v_int = self._get_yen_int(self._venda_valor_parc_entry)
                t_int = self._get_valor_troca_selecionada()
                volta_paga = str(v_int - t_int)
            else:
                entrada = self._yen_raw(self._venda_entrada_entry)
                v_int = self._get_yen_int(self._venda_valor_parc_entry)
                e_int = int(entrada) if entrada else 0
                t_int = self._get_valor_troca_selecionada() if tipo == "Com Troca" else 0
                saldo = v_int - e_int - t_int
                p_int = int(parcela_mensal) if parcela_mensal else 0
                if p_int > 0 and saldo > 0:
                    num_parcelas = saldo // p_int
                    valor_ultima = str(saldo % p_int)

        # Busca compra_id ligada ao carro
        compra_id_venda = None
        if carro_id:
            cr_v = self.conn.execute(
                "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                (carro_id,)).fetchone()
            if cr_v:
                compra_id_venda = cr_v[0]

        nfi_venda = getattr(self, "_venda_nfi_var", None)
        nfi_venda = nfi_venda.get() if nfi_venda else "SNF"
        dados = (carro_id, carro_nome, cor_sel, placa_sel, cliente_id, tipo,
                 valor_venda, entrada, parcela_mensal, num_parcelas, valor_ultima,
                 data_primeira_parc, data_venda, carro_troca_id, carro_troca,
                 valor_troca, volta_paga, "", compra_id_venda, nfi_venda)

        if self._venda_edit_id is not None:
            self.conn.execute("""UPDATE vendas SET
                carro_id=?,carro=?,cor=?,placa=?,cliente_id=?,tipo_venda=?,
                valor_venda=?,entrada=?,parcela_mensal=?,num_parcelas=?,valor_ultima_parc=?,
                data_primeira_parc=?,data_venda=?,carro_troca_id=?,carro_troca=?,
                valor_troca=?,volta_paga=?,obs=?,compra_id=?,nfi=? WHERE id=?""",
                dados + (self._venda_edit_id,))
            self.conn.commit()
            self._venda_edit_id = None
            self._lbl_venda_title.configure(text="◆  Nova Venda")
            self._btn_salvar_venda.configure(text="  Registrar Venda  ")
            self._lbl_venda_status.configure(text="✔ Venda atualizada!", fg=COLORS["green"])
        else:
            cur = self.conn.execute("""INSERT INTO vendas
                (carro_id,carro,cor,placa,cliente_id,tipo_venda,valor_venda,entrada,
                 parcela_mensal,num_parcelas,valor_ultima_parc,data_primeira_parc,data_venda,
                 carro_troca_id,carro_troca,valor_troca,volta_paga,obs,compra_id,nfi)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", dados)
            self.conn.commit()
            vid = cur.lastrowid
            self._lbl_venda_status.configure(text="✔ Venda registrada!", fg=COLORS["green"])

        # IS automático para CNF (upsert — sem duplicar)
        if nfi_venda == "CNF":
            cid_is = compra_id_venda
            if not cid_is and carro_id:
                cr_is = self.conn.execute(
                    "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                    (carro_id,)).fetchone()
                cid_is = cr_is[0] if cr_is else None
            if cid_is:
                self._upsert_is("custos", "compra_id", cid_is, valor_venda, data_venda)

        # Taxas de Leilão de Venda — gerar custos automáticos
        if tipo == "Venda Leilão":
            import datetime as _dt
            data_reg = data_venda or _dt.date.today().strftime("%d/%m/%Y")
            # Busca compra vinculada ao carro
            compra_lv = self.conn.execute(
                "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                (carro_id,)).fetchone() if carro_id else None
            if compra_lv:
                cid_lv = compra_lv[0]
                taxa_exp   = self._yen_raw(self._venda_taxa_exp_entry)
                taxa_kensa = self._yen_raw(self._venda_taxa_kensa_entry)
                self.conn.execute("INSERT OR IGNORE INTO tipos_custo (nome) VALUES ('Taxa Exposição')")
                self.conn.execute("INSERT OR IGNORE INTO tipos_custo (nome) VALUES ('Taxa Kensa')")
                if taxa_exp:
                    self.conn.execute(
                        "INSERT INTO custos (compra_id,tipo_custo,descricao,valor,data_custo) "
                        "VALUES (?,?,?,?,?)",
                        (cid_lv, "Taxa Exposição", f"Venda Leilão automático", taxa_exp, data_reg))
                if taxa_kensa:
                    self.conn.execute(
                        "INSERT INTO custos (compra_id,tipo_custo,descricao,valor,data_custo) "
                        "VALUES (?,?,?,?,?)",
                        (cid_lv, "Taxa Kensa", f"Venda Leilão automático", taxa_kensa, data_reg))
                self.conn.commit()

        # Calcula e salva lucro da venda ligada à compra (se houver compra vinculada)
        try:
            if compra_id_venda:
                compra_row = self.conn.execute(
                    "SELECT * FROM compras WHERE id=?", (compra_id_venda,)).fetchone()
                if compra_row:
                    try:
                        v_int = int(str(valor_venda).replace(",", "")) if valor_venda else 0
                    except Exception:
                        v_int = 0
                    try:
                        c_int = int(str(compra_row["valor"]).replace(",", "")) if compra_row["valor"] else 0
                    except Exception:
                        c_int = 0
                    custo_total = self._get_custo_total(compra_id_venda)
                    lucro = int(v_int) - int(c_int) - int(custo_total)
                    # Atualiza venda com lucro e marca compra como não à venda
                    if 'vid' in locals():
                        self.conn.execute("UPDATE vendas SET lucro=? WHERE id=?", (str(lucro), vid))
                    self.conn.execute("UPDATE compras SET a_venda=0, preco_venda=? WHERE id=?",
                                      (valor_venda, compra_id_venda))
                    self.conn.commit()
        except Exception:
            pass

        # Atualiza status do carro após venda
        if carro_id:
            import datetime as _dt_s
            carro_row = self.conn.execute(
                "SELECT status, historico_status FROM carros WHERE id=?", (carro_id,)).fetchone()
            if carro_row:
                old_st = carro_row[0]
                hist_s = carro_row[1] or ""
                if tipo == "Venda Leilão":
                    # Leilão → Inativo; registra histórico
                    novo_st = "Inativo"
                    novo_cli = None
                else:
                    # Outras vendas → Cliente (se era Estoque)
                    novo_st = "Cliente" if old_st == "Estoque" else old_st
                    novo_cli = cliente_id
                if old_st != novo_st:
                    entradas = [e.strip() for e in hist_s.split("|") if e.strip()]
                    nova_ent = f"{old_st} ({_dt_s.date.today().strftime('%d/%m/%Y')})"
                    if nova_ent not in entradas:
                        entradas.append(nova_ent)
                    self.conn.execute(
                        "UPDATE carros SET historico_status=? WHERE id=?",
                        (" | ".join(entradas), carro_id))
                self.conn.execute(
                    "UPDATE carros SET status=?, cliente_id=? WHERE id=?",
                    (novo_st, novo_cli if novo_st != "Inativo" else None, carro_id))
                self.conn.commit()
        self.vendas_data = [dict(r) for r in
            self.conn.execute("SELECT * FROM vendas ORDER BY id DESC").fetchall()]
        self._limpar_form_venda(clear_status=False)
        self._refresh_tabela_vendas()


    def _limpar_form_venda(self, clear_status=True):
        import datetime
        self._venda_carro_var.set("")
        self._venda_cliente_var.set("")
        self._venda_tipo_var.set("")
        self._venda_troca_var.set("")
        if hasattr(self, "_venda_taxa_exp_entry"):
            self._venda_taxa_exp_entry.delete(0, tk.END)
        if hasattr(self, "_venda_taxa_kensa_entry"):
            self._venda_taxa_kensa_entry.delete(0, tk.END)
        self._venda_edit_id = None
        self._lbl_venda_title.configure(text="◆  Nova Venda")
        self._btn_salvar_venda.configure(text="  Registrar Venda  ")
        for e in [self._venda_valor_entry, self._venda_valor_parc_entry,
                  self._venda_entrada_entry, self._venda_parcela_entry]:
            e.delete(0, tk.END)
        hoje = datetime.date.today()
        for sets in [(self._venda_dia, self._venda_mes, self._venda_ano),
                     (self._venda_pdia, self._venda_pmes, self._venda_pano)]:
            for e, val in zip(sets, [str(hoje.day).zfill(2),
                                     str(hoje.month).zfill(2), str(hoje.year)]):
                e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        self._bloco_avista.pack_forget()
        self._bloco_troca.pack_forget()
        self._bloco_entrada_wrap.pack_forget()
        self._bloco_valor_parc.pack_forget()
        self._bloco_parc.pack_forget()
        self._bloco_volta.pack_forget()
        if clear_status:
            self._lbl_venda_status.configure(text="")

    def _editar_venda(self, vid):
        v = next((x for x in self.vendas_data if x["id"] == vid), None)
        if not v: return
        self._venda_edit_id = vid
        self._lbl_venda_title.configure(text="✏  Editar Venda")
        self._btn_salvar_venda.configure(text="  Salvar Alterações  ")

        # Carro
        self._atualizar_combo_carros_venda()
        cid = v.get("carro_id")
        vals = list(self._venda_carro_combo["values"])
        match = next((x for x in vals if x.startswith(f"{cid} —")), None)
        self._venda_carro_var.set(match or v["carro"])

        # Cliente
        self._atualizar_combo_clientes_venda()
        clid = v.get("cliente_id")
        cli_vals = list(self._venda_cliente_combo["values"])
        cmatch = next((x for x in cli_vals if x.startswith(f"{clid} —")), "") if clid else ""
        self._venda_cliente_var.set(cmatch)

        self._venda_tipo_var.set(v["tipo_venda"])
        self._toggle_campos_venda()

        def fill_yen(entry, val):
            entry.delete(0, tk.END)
            if val:
                try: entry.insert(0, f"{int(str(val).replace(',','')):,}")
                except: entry.insert(0, val)

        tipo = v["tipo_venda"]
        if tipo == "Venda a Vista":
            fill_yen(self._venda_valor_entry, v.get("valor_venda"))
            self._restore_date(self._venda_dia, self._venda_mes, self._venda_ano, v.get("data_venda"))
        else:
            fill_yen(self._venda_valor_parc_entry, v.get("valor_venda"))
            fill_yen(self._venda_entrada_entry,    v.get("entrada"))
            fill_yen(self._venda_parcela_entry,    v.get("parcela_mensal"))
            self._restore_date(self._venda_pdia, self._venda_pmes, self._venda_pano, v.get("data_primeira_parc"))
            if tipo in ("Com Troca", "Com Troca e Volta"):
                self._atualizar_combo_troca()
                tid = v.get("carro_troca_id")
                tvals = list(self._venda_troca_combo["values"])
                tmatch = next((x for x in tvals if x.startswith(f"{tid} —")), "")
                self._venda_troca_var.set(tmatch or v.get("carro_troca", ""))
            self._calc_parcelas()
        # NFI
        if hasattr(self, "_venda_nfi_var"):
            self._venda_nfi_var.set(v.get("nfi") or "SNF")
        self._lbl_venda_status.configure(text="")

    def _restore_date(self, e_dia, e_mes, e_ano, data_str):
        import datetime
        hoje = datetime.date.today()
        if data_str:
            partes = data_str.split("/")
            dd = partes[0] if len(partes) > 0 else str(hoje.day).zfill(2)
            mm = partes[1] if len(partes) > 1 else str(hoje.month).zfill(2)
            aa = partes[2] if len(partes) > 2 else str(hoje.year)
        else:
            dd, mm, aa = str(hoje.day).zfill(2), str(hoje.month).zfill(2), str(hoje.year)
        for e, val in [(e_dia, dd), (e_mes, mm), (e_ano, aa)]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])

    def _excluir_venda(self, vid):
        # Pede confirmação digitando "excluir"
        dlg = tk.Toplevel(self)
        dlg.title("Confirmar Exclusão")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("340x170")
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 340) // 2
        y = self.winfo_y() + (self.winfo_height() - 170) // 2
        dlg.geometry(f"+{x}+{y}")

        tk.Frame(dlg, bg=COLORS["red"], height=4).pack(fill="x")
        tk.Label(dlg, text="Confirmar Exclusão",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(pady=(14, 4))
        tk.Label(dlg, text='Digite  excluir  para confirmar:',
                 font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack()

        conf_var = tk.StringVar()
        conf_entry = tk.Entry(dlg, textvariable=conf_var, font=("Helvetica", 10),
                              bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                              insertbackground=COLORS["text_primary"],
                              relief="flat", bd=0,
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=20, justify="center")
        conf_entry.pack(pady=(6, 10), ipady=6)
        conf_entry.focus()

        btn_row = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_row.pack()

        def do_delete():
            if conf_var.get().strip().lower() == "excluir":
                self.conn.execute("DELETE FROM vendas WHERE id=?", (vid,))
                self.conn.commit()
                self.vendas_data = [dict(r) for r in
                    self.conn.execute("SELECT * FROM vendas ORDER BY id DESC").fetchall()]
                self._refresh_tabela_vendas()
                dlg.destroy()
            else:
                conf_entry.configure(highlightbackground=COLORS["red"])

        tk.Button(btn_row, text="  Excluir  ", font=("Helvetica", 10, "bold"),
                  bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                  command=do_delete).pack(side="left", ipady=5, ipadx=6)
        tk.Button(btn_row, text="  Cancelar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(side="left", padx=(10, 0), ipady=5, ipadx=6)
        conf_entry.bind("<Return>", lambda e: do_delete())

    def _refresh_tabela_vendas(self):
        for w in self._vendas_rows_frame.winfo_children():
            w.destroy()

        # Busca com join de cliente
        lista = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id = c.id "
            "ORDER BY v.id DESC").fetchall()]

        # Filtro texto
        termo = self._venda_filtro_var.get().strip().lower() if hasattr(self, "_venda_filtro_var") else ""
        if termo:
            lista = [v for v in lista if
                     termo in (v.get("carro") or "").lower() or
                     termo in (v.get("cliente_nome") or "").lower() or
                     termo in (v.get("tipo_venda") or "").lower() or
                     termo in (v.get("data_venda") or "").lower()]

        # Filtro mês/ano
        fmes = getattr(self, "_venda_filtro_mes", None)
        fano = getattr(self, "_venda_filtro_ano", None)
        fmes = fmes.get() if fmes else "Todos"
        fano = fano.get() if fano else "Todos"
        if fmes != "Todos" or fano != "Todos":
            def match_data(v):
                data = v.get("data_venda") or ""
                pts = data.split("/")
                vm = pts[1] if len(pts) > 1 else ""
                va = pts[2] if len(pts) > 2 else ""
                if fmes != "Todos" and vm != fmes: return False
                if fano != "Todos" and va != fano: return False
                return True
            lista = [v for v in lista if match_data(v)]

        # Filtro tipo
        ftipo = getattr(self, "_venda_filtro_tipo", None)
        ftipo = ftipo.get() if ftipo else "Todos"
        if ftipo != "Todos":
            lista = [v for v in lista if v.get("tipo_venda") == ftipo]

        # Filtro NFI
        fnfi_v = getattr(self, "_venda_filtro_nfi", None)
        fnfi_v = fnfi_v.get() if fnfi_v else "Todos"
        if fnfi_v != "Todos":
            lista = [v for v in lista if (v.get("nfi") or "SNF") == fnfi_v]

        self._lbl_total_vendas.configure(text=f"{len(lista)} registro(s)")
        self._vendas_lista_filtrada = lista[:]

        if not lista:
            tk.Label(self._vendas_rows_frame, text="Nenhuma venda encontrada.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        TIPO_COLORS = {
            "Venda a Vista":      COLORS["green"],
            "Venda Parcelada":    COLORS["blue"],
            "Com Troca":          COLORS["orange"],
            "Com Troca e Volta":  COLORS["accent"],
        }

        for i, v in enumerate(lista):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._vendas_rows_frame, bg=rb)
            row.pack(fill="x")

            # Verifica se é parcelada e calcula saldo/status
            tipo = v.get("tipo_venda", "")
            pagas     = v.get("parcelas_pagas") or 0
            total_p   = v.get("num_parcelas") or 0
            is_parc   = tipo in ("Venda Parcelada", "Com Troca", "Com Troca e Volta")
            quitado   = is_parc and total_p > 0 and pagas >= total_p

            try:
                total_pago = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                    "FROM pagamentos WHERE venda_id=?", (v["id"],)).fetchone()[0]
                val_venda = int(str(v.get("valor_venda") or "0").replace(",",""))
                entrada   = int(str(v.get("entrada") or "0").replace(",",""))
                val_troca = int(str(v.get("valor_troca") or "0").replace(",",""))
                saldo_orig = val_venda - entrada - val_troca
                saldo_rest = max(0, saldo_orig - total_pago)
            except Exception:
                saldo_rest = 0

            tk.Label(row, text=str(v["id"]), font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_muted"], width=3, anchor="w").pack(side="left", padx=2, pady=6)
            cli_txt = (v.get("cliente_nome") or "—")[:13]
            tk.Label(row, text=cli_txt, font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_primary"], width=13, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=v.get("data_venda") or "—", font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_secondary"], width=9, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=(v.get("carro") or "—")[:13], font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_primary"], width=13, anchor="w").pack(side="left", padx=2)
            tc = TIPO_COLORS.get(tipo, COLORS["text_muted"])
            tipo_short = {"Venda a Vista":"À Vista","Venda Parcelada":"Parcelada",
                          "Com Troca":"C/Troca","Com Troca e Volta":"T+Volta"}.get(tipo, tipo)
            tk.Label(row, text=tipo_short, font=("Helvetica",7,"bold"),
                     bg=tc, fg="white", width=11, anchor="center").pack(side="left", padx=2, pady=3)
            tk.Label(row, text=self._fmt_yen_display(v.get("valor_venda")),
                     font=("Helvetica",9), bg=rb, fg=COLORS["green"],
                     width=9, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(v.get("entrada")) if v.get("entrada") else "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=8, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(v.get("parcela_mensal")) if is_parc else "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["blue"],
                     width=8, anchor="w").pack(side="left", padx=2)
            # Saldo devedor
            if is_parc and not quitado:
                saldo_disp = f"¥{saldo_rest:,}"
                saldo_fg   = COLORS["orange"]
            elif is_parc and quitado:
                saldo_disp = "Quitado"
                saldo_fg   = COLORS["green"]
            else:
                saldo_disp = "—"
                saldo_fg   = COLORS["text_muted"]
            tk.Label(row, text=saldo_disp, font=("Helvetica",9), bg=rb,
                     fg=saldo_fg, width=9, anchor="w").pack(side="left", padx=2)
            # Status badge
            if not is_parc:
                st_txt, st_bg = "À Vista", COLORS["green"]
            elif quitado:
                st_txt, st_bg = "Quitado", COLORS["green"]
            else:
                st_txt, st_bg = "Em Aberto", COLORS["blue"]
            tk.Label(row, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_bg, fg="white", width=8, anchor="center").pack(side="left", padx=2, pady=3)

            # NFI badge
            nfi_v = v.get("nfi") or "SNF"
            nfi_bg = "#8E44AD" if nfi_v == "CNF" else "#7F8C8D"
            tk.Label(row, text=nfi_v, font=("Helvetica",7,"bold"),
                     bg=nfi_bg, fg="white", width=5, anchor="center"
                     ).pack(side="left", padx=2, pady=3)

            acts = tk.Frame(row, bg=rb)
            acts.pack(side="left", padx=2)
            tk.Button(acts, text="✏", font=("Helvetica",8), bg=COLORS["blue"], fg="white",
                      relief="flat", cursor="hand2", padx=3, pady=1,
                      command=lambda vid=v["id"]: self._editar_venda(vid)
                      ).pack(side="left", padx=(0, 2))
            tk.Button(acts, text="✕", font=("Helvetica",8), bg=COLORS["red"], fg="white",
                      relief="flat", cursor="hand2", padx=3, pady=1,
                      command=lambda vid=v["id"]: self._excluir_venda(vid)
                      ).pack(side="left")

    # ══════════════════════════════════════════════════════════════════════════
    # SUBMENU: LUCRO DE VENDAS
    # ══════════════════════════════════════════════════════════════════════════
    def _build_lucro_vendas(self, parent):
        """Tela de análise de lucro/prejuízo por venda."""
        self._lucro_filtro_var  = tk.StringVar(value="")
        self._lucro_filtro_tipo = tk.StringVar(value="Todos")
        self._lucro_filtro_ano  = tk.StringVar(value="Todos")
        self._lucro_filtro_mes  = tk.StringVar(value="Todos")
        self._lucro_filtro_nfi  = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 8))
        tk.Label(hdr, text="◈  Lucro de Vendas",
                 font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["green"]).pack(side="left")
        self._lucro_lbl_total = tk.Label(hdr, text="",
                 font=("Helvetica", 9),
                 bg=COLORS["bg_content"], fg=COLORS["text_muted"])
        self._lucro_lbl_total.pack(side="right")

        # ── KPIs ──────────────────────────────────────────────────────────────
        kpi_f = tk.Frame(root, bg=COLORS["bg_content"])
        kpi_f.pack(fill="x", pady=(0, 10))
        self._lucro_kpi_frames = {}
        for key, label, color in [
            ("total_vendas",  "Total Vendas",    COLORS["green"]),
            ("total_custo",   "Total Custo",     COLORS["orange"]),
            ("lucro_total",   "Lucro Total",     COLORS["green"]),
            ("margem_media",  "Margem Média",    COLORS["blue"]),
            ("qtd_lucro",     "Com Lucro",       COLORS["green"]),
            ("qtd_prejuizo",  "Com Prejuízo",    COLORS["red"]),
        ]:
            kf = tk.Frame(kpi_f, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            kf.pack(side="left", expand=True, fill="x", padx=(0,6))
            tk.Label(kf, text=label, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(8,0))
            lbl = tk.Label(kf, text="—", font=("Helvetica",10,"bold"),
                           bg=COLORS["bg_card"], fg=color)
            lbl.pack(pady=(2,8))
            self._lucro_kpi_frames[key] = lbl

        # ── Filtros ────────────────────────────────────────────────────────────
        import datetime as _dt
        anos  = ["Todos"] + [str(y) for y in range(_dt.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        tipos = ["Todos", "Venda a Vista", "Venda Parcelada", "Com Troca",
                 "Com Troca e Volta", "Venda Leilão"]

        fbar = tk.Frame(root, bg=COLORS["bg_main"])
        fbar.pack(fill="x", pady=(0,6))

        tk.Label(fbar, text="🔍", font=("Helvetica",9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4))
        tk.Entry(fbar, textvariable=self._lucro_filtro_var, font=("Helvetica",9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["green"], width=18
                 ).pack(side="left", ipady=4)
        self._lucro_filtro_var.trace_add("write", lambda *_: self._refresh_lucro())
        tk.Button(fbar, text="✕", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._lucro_filtro_var.set("")
                  ).pack(side="left", padx=(3,12), ipady=3)

        for txt, var, opts in [
            ("Tipo:", self._lucro_filtro_tipo, tipos),
            ("Mês:",  self._lucro_filtro_mes,  meses),
            ("Ano:",  self._lucro_filtro_ano,  anos),
        ]:
            tk.Label(fbar, text=txt, font=("Helvetica",8),
                     bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,3))
            ttk.Combobox(fbar, textvariable=var, values=opts,
                         state="readonly", font=("Helvetica",8),
                         width=14 if txt=="Tipo:" else 5
                         ).pack(side="left", padx=(0,8), ipady=2)
            var.trace_add("write", lambda *_: self._refresh_lucro())

        tk.Label(fbar, text="NFI:", font=("Helvetica",8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(4,2))
        ttk.Combobox(fbar, textvariable=self._lucro_filtro_nfi,
                     values=["Todos","SNF","CNF"],
                     state="readonly", font=("Helvetica",8), width=6
                     ).pack(side="left", padx=(0,8), ipady=2)
        self._lucro_filtro_nfi.trace_add("write", lambda *_: self._refresh_lucro())
        tk.Button(fbar, text="↺ Atualizar", font=("Helvetica",8,"bold"),
                  bg=COLORS["green"], fg="white", relief="flat", cursor="hand2",
                  padx=8, command=self._refresh_lucro
                  ).pack(side="right", ipady=4)

        # ── Tabela com dual-canvas ─────────────────────────────────────────────
        LUCRO_COLS = [
            ("Data",   80), ("#",   35), ("Carro",   125), ("Ano", 42), ("Cor", 55),
            ("Placa",  65), ("Cliente", 100), ("NFI", 42),
            ("T.Compra", 85), ("V.Compra", 80), ("Custos", 75), ("IS", 65), ("T.Total", 80),
            ("T.Venda",  80), ("V.Venda",  80), ("Carro Troca", 110), ("V.Troca", 80),
            ("Saldo",  75), ("Status", 70),
            ("Lucro/Perda", 95), ("Margem", 65),
        ]
        self._lucro_cols = LUCRO_COLS
        LUCRO_W = sum(c[1] for c in LUCRO_COLS)

        tbl_outer = tk.Frame(root, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        tbl_outer.pack(fill="both", expand=True)
        tk.Frame(tbl_outer, bg=COLORS["green"], height=4).pack(fill="x")

        tbl_area = tk.Frame(tbl_outer, bg=COLORS["bg_card"])
        tbl_area.pack(fill="both", expand=True, padx=8, pady=(4,8))

        vsb = tk.Scrollbar(tbl_area, orient="vertical")
        hsb = tk.Scrollbar(tbl_area, orient="horizontal")
        self._lucro_hdr_cv = tk.Canvas(tbl_area, bg=COLORS["bg_main"],
                                        height=28, highlightthickness=0)
        self._lucro_body_cv = tk.Canvas(tbl_area, bg=COLORS["bg_card"],
                                         highlightthickness=0,
                                         yscrollcommand=vsb.set)

        def _lx(*a):
            self._lucro_hdr_cv.xview(*a)
            self._lucro_body_cv.xview(*a)
        hsb.config(command=_lx)
        vsb.config(command=self._lucro_body_cv.yview)

        self._lucro_hdr_cv.grid(row=0, column=0, sticky="ew")
        self._lucro_body_cv.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        tbl_area.grid_rowconfigure(1, weight=1)
        tbl_area.grid_columnconfigure(0, weight=1)

        # Header frame
        lhf = tk.Frame(self._lucro_hdr_cv, bg=COLORS["bg_main"])
        lhw = self._lucro_hdr_cv.create_window((0,0), window=lhf, anchor="nw")
        for col_name, col_w in LUCRO_COLS:
            f = tk.Frame(lhf, bg=COLORS["bg_main"], width=col_w, height=28)
            f.pack_propagate(False); f.pack(side="left")
            tk.Label(f, text=col_name, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w").pack(fill="both", padx=3)
        tk.Frame(lhf, bg=COLORS["border"], height=1,
                 width=LUCRO_W).pack(fill="x")

        # Body frame
        self._lucro_rows_frame = tk.Frame(self._lucro_body_cv, bg=COLORS["bg_card"])
        lbw = self._lucro_body_cv.create_window((0,0), window=self._lucro_rows_frame, anchor="nw")
        self._lucro_rows_frame.bind("<Configure>",
            lambda e: self._lucro_body_cv.configure(
                scrollregion=self._lucro_body_cv.bbox("all")))

        def _lcr(e, _bw=lbw, _hw=lhw):
            w = max(e.width, LUCRO_W)
            self._lucro_body_cv.itemconfig(_bw, width=w)
            self._lucro_hdr_cv.itemconfig(_hw, width=w)
            self._lucro_hdr_cv.configure(scrollregion=(0,0,w,28))
        self._lucro_body_cv.bind("<Configure>", _lcr)

        def _lxmove(first, last):
            hsb.set(first, last)
            self._lucro_hdr_cv.xview_moveto(first)
        self._lucro_body_cv.configure(xscrollcommand=_lxmove)

        tbl_outer.bind("<Enter>", lambda e: tbl_outer.bind_all(
            "<MouseWheel>", lambda ev: self._lucro_body_cv.yview_scroll(
                int(-1*(ev.delta/120)), "units")))
        tbl_outer.bind("<Leave>", lambda e: tbl_outer.unbind_all("<MouseWheel>"))

        self._refresh_lucro()

    def _refresh_lucro(self):
        """Popula a tabela de Lucro de Vendas."""
        import datetime as _dt

        for w in self._lucro_rows_frame.winfo_children():
            w.destroy()

        LUCRO_COLS = self._lucro_cols
        for ci, (_, pw) in enumerate(LUCRO_COLS):
            self._lucro_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        termo      = self._lucro_filtro_var.get().strip().lower()
        ftipo      = self._lucro_filtro_tipo.get()
        fmes       = self._lucro_filtro_mes.get()
        fano       = self._lucro_filtro_ano.get()

        # ── Busca todas as vendas com dados do cliente e compra ──────────────
        rows = self.conn.execute(
            "SELECT v.*, cl.nome as cli_nome, "
            "  ca.ano as c_ano, ca.cor as c_cor, ca.placa as c_placa, "
            "  ca.status as c_status "
            "FROM vendas v "
            "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
            "LEFT JOIN carros ca ON v.carro_id=ca.id "
            "ORDER BY v.id DESC").fetchall()
        vendas = [dict(r) for r in rows]

        # ── Filtros ───────────────────────────────────────────────────────────
        def _ym(dv):
            p = (dv or "").split("/")
            return (p[2] if len(p)>2 else "", p[1] if len(p)>1 else "")

        fnfi_lc = getattr(self, "_lucro_filtro_nfi", None)
        fnfi_lc = fnfi_lc.get() if fnfi_lc else "Todos"

        filtrado = []
        for v in vendas:
            ano_v, mes_v = _ym(v.get("data_venda",""))
            if ftipo != "Todos" and v.get("tipo_venda") != ftipo: continue
            if fmes  != "Todos" and mes_v != fmes:                continue
            if fano  != "Todos" and ano_v != fano:                continue
            if fnfi_lc != "Todos" and (v.get("nfi") or "SNF") != fnfi_lc: continue
            if termo:
                hay = " ".join(filter(None, [
                    v.get("carro",""), v.get("c_placa",""), v.get("cli_nome",""),
                    v.get("tipo_venda","")])).lower()
                if termo not in hay: continue
            filtrado.append(v)

        # ── KPIs ─────────────────────────────────────────────────────────────
        tot_venda = tot_custo = tot_lucro = 0
        cnt_l = cnt_p = 0
        margens = []

        def _int(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        # ── Renderizar linhas ─────────────────────────────────────────────────
        TIPO_COMP_C = {"Leilão": COLORS["orange"], "Troca": COLORS["blue"],
                       "Compra Direta": COLORS["green"]}
        TIPO_VEND_C = {"Venda a Vista": COLORS["green"],
                       "Venda Parcelada": COLORS["blue"],
                       "Com Troca": COLORS["orange"],
                       "Com Troca e Volta": COLORS["accent"],
                       "Venda Leilão": COLORS["orange"]}

        for i, v in enumerate(filtrado):
            # ── Dados do carro e compra ───────────────────────────────────────
            carro_id = v.get("carro_id")
            compra = None
            custo_total = 0
            v_compra = 0
            tipo_compra = "—"

            # Usa compra_id direto se disponível (mais preciso), fallback por carro_id
            compra_id_direto = v.get("compra_id")
            if compra_id_direto:
                cr = self.conn.execute(
                    "SELECT * FROM compras WHERE id=?", (compra_id_direto,)).fetchone()
            elif carro_id:
                cr = self.conn.execute(
                    "SELECT * FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                    (carro_id,)).fetchone()
            else:
                cr = None
            if cr:
                compra = dict(cr)
                v_compra    = _int(compra.get("valor"))
                custo_total = self._get_custo_total(compra["id"])
                tipo_compra = compra.get("tipo","—")

            # IS da compra (custo IS em custos)
            is_compra = 0
            if compra:
                try:
                    ri_is = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos WHERE compra_id=? AND tipo_custo='IS'",
                        (compra["id"],)).fetchone()
                    is_compra = int(ri_is[0]) if ri_is[0] else 0
                except Exception: pass
            custo_sem_is = custo_total - is_compra
            total_compra = v_compra + custo_total

            # ── Dados da venda ────────────────────────────────────────────────
            v_venda   = _int(v.get("valor_venda"))
            v_entrada = _int(v.get("entrada"))
            v_troca   = _int(v.get("valor_troca"))
            pagas     = v.get("parcelas_pagas") or 0
            tot_p     = v.get("num_parcelas") or 0
            parcela   = _int(v.get("parcela_mensal"))
            is_p      = v.get("tipo_venda","") in (
                "Venda Parcelada","Com Troca","Com Troca e Volta")

            # Saldo devedor
            try:
                tp = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) FROM pagamentos WHERE venda_id=?",
                    (v["id"],)).fetchone()[0]
                saldo = max(0, v_venda - v_entrada - v_troca - tp) if is_p else 0
            except:
                saldo = 0

            # Status pagamento
            quitado = is_p and tot_p > 0 and pagas >= tot_p
            if v.get("tipo_venda") == "Venda a Vista" or v.get("tipo_venda") == "Venda Leilão":
                status_txt = "À Vista"
                status_col = COLORS["green"]
            elif quitado:
                status_txt = "Quitado"
                status_col = COLORS["green"]
            else:
                status_txt = "Em Aberto"
                status_col = COLORS["orange"]

            # Lucro = valor venda - total de custo/compra
            lucro = v_venda - total_compra
            margem = (lucro / v_venda * 100) if v_venda > 0 else 0.0

            tot_venda  += v_venda
            tot_custo  += total_compra
            tot_lucro  += lucro
            margens.append(margem)
            if lucro >= 0: cnt_l += 1
            else:          cnt_p += 1

            nfi_lc = v.get("nfi") or "SNF"

            # ── Layout da linha ───────────────────────────────────────────────
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            tk.Frame(self._lucro_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=i*2, column=0, columnspan=len(LUCRO_COLS), sticky="ew")
            ri = i*2 + 1
            self._lucro_rows_frame.grid_rowconfigure(ri, minsize=34)

            def cell(col_i, txt, fg=None, bg=None, bold=False):
                f = tk.Frame(self._lucro_rows_frame, bg=bg or rb)
                f.grid(row=ri, column=col_i, sticky="nsew")
                tk.Label(f, text=str(txt), font=("Helvetica", 7, "bold" if bold else "normal"),
                         bg=bg or rb, fg=fg or COLORS["text_primary"],
                         anchor="w", padx=4, wraplength=130
                         ).pack(fill="both", expand=True, padx=2, pady=3)

            def badge(col_i, txt, bg_c):
                f = tk.Frame(self._lucro_rows_frame, bg=rb)
                f.grid(row=ri, column=col_i, sticky="nsew")
                tk.Label(f, text=txt, font=("Helvetica",7,"bold"),
                         bg=bg_c, fg="white", anchor="center"
                         ).pack(fill="both", expand=True, padx=3, pady=5)

            carro_txt = (v.get("carro","—") or "—").split("|")[0].strip()
            if " — " in carro_txt: carro_txt = carro_txt.split(" — ",1)[1]

            col = 0
            cell(col, v.get("data_venda") or "—", COLORS["text_secondary"]); col+=1
            cell(col, f"#{v['id']}",              COLORS["text_muted"]); col+=1
            cell(col, carro_txt[:18],             COLORS["text_primary"]); col+=1
            cell(col, v.get("c_ano") or "—",      COLORS["text_secondary"]); col+=1
            cell(col, v.get("c_cor") or "—",      COLORS["text_secondary"]); col+=1
            cell(col, v.get("c_placa") or "—",    COLORS["text_secondary"]); col+=1
            cell(col, (v.get("cli_nome") or "—")[:14], COLORS["accent"]); col+=1

            # NFI badge (após cliente)
            nfi_lc_bg = "#8E44AD" if nfi_lc == "CNF" else "#7F8C8D"
            f_nfi = tk.Frame(self._lucro_rows_frame, bg=rb)
            f_nfi.grid(row=ri, column=col, sticky="nsew")
            tk.Label(f_nfi, text=nfi_lc, font=("Helvetica",7,"bold"),
                     bg=nfi_lc_bg, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            col+=1

            # Tipo compra badge
            badge(col, tipo_compra, TIPO_COMP_C.get(tipo_compra, COLORS["text_muted"])); col+=1
            cell(col, self._fmt_yen_display(v_compra), COLORS["orange"]); col+=1
            cell(col, self._fmt_yen_display(custo_sem_is), COLORS["text_secondary"]); col+=1
            # IS badge
            is_bg_lv = "#8E44AD" if is_compra > 0 else COLORS["border"]
            f_is = tk.Frame(self._lucro_rows_frame, bg=rb)
            f_is.grid(row=ri, column=col, sticky="nsew")
            tk.Label(f_is, text=self._fmt_yen_display(is_compra) if is_compra else "—",
                     font=("Helvetica",7,"bold"), bg=is_bg_lv, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            col+=1
            cell(col, self._fmt_yen_display(total_compra), COLORS["orange"], bold=True); col+=1

            # Tipo venda badge
            badge(col, v.get("tipo_venda","—"), TIPO_VEND_C.get(v.get("tipo_venda",""), COLORS["text_muted"])); col+=1
            cell(col, self._fmt_yen_display(v_venda), COLORS["green"]); col+=1
            # Carro troca e valor troca
            carro_troca_txt = (v.get("carro_troca") or "—").split("|")[0].strip()[:16]
            cell(col, carro_troca_txt, COLORS["blue"]); col+=1
            cell(col, self._fmt_yen_display(v_troca) if v_troca else "—", COLORS["blue"]); col+=1
            cell(col, self._fmt_yen_display(saldo) if saldo > 0 else "—", COLORS["orange"]); col+=1

            # Status badge
            badge(col, status_txt, status_col); col+=1

            # Lucro/Perda com fundo verde ou vermelho
            lucro_txt = f"¥ {lucro:,}" if lucro >= 0 else f"¥ {lucro:,}"
            lucro_bg  = COLORS["green"] if lucro >= 0 else COLORS["red"]
            f_l = tk.Frame(self._lucro_rows_frame, bg=rb)
            f_l.grid(row=ri, column=col, sticky="nsew")
            tk.Label(f_l, text=lucro_txt, font=("Helvetica",7,"bold"),
                     bg=lucro_bg, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            col+=1

            # Margem
            margem_txt = f"{margem:+.1f}%"
            margem_col = COLORS["green"] if margem >= 0 else COLORS["red"]
            cell(col, margem_txt, margem_col, bold=True)

        # Linha final separadora
        if filtrado:
            tk.Frame(self._lucro_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=len(filtrado)*2, column=0,
                            columnspan=len(LUCRO_COLS), sticky="ew")
        else:
            tk.Label(self._lucro_rows_frame,
                     text="Nenhuma venda encontrada com os filtros selecionados.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]
                     ).grid(row=0, column=0, columnspan=len(LUCRO_COLS), pady=40)

        # ── Atualiza KPIs ─────────────────────────────────────────────────────
        mm = (sum(margens)/len(margens)) if margens else 0
        lucro_col_kpi = COLORS["green"] if tot_lucro >= 0 else COLORS["red"]
        self._lucro_kpi_frames["total_vendas"].configure(
            text=self._fmt_yen_display(tot_venda))
        self._lucro_kpi_frames["total_custo"].configure(
            text=self._fmt_yen_display(tot_custo))
        self._lucro_kpi_frames["lucro_total"].configure(
            text=f"¥ {tot_lucro:,}", fg=lucro_col_kpi)
        self._lucro_kpi_frames["margem_media"].configure(
            text=f"{mm:+.1f}%",
            fg=COLORS["green"] if mm >= 0 else COLORS["red"])
        self._lucro_kpi_frames["qtd_lucro"].configure(text=str(cnt_l))
        self._lucro_kpi_frames["qtd_prejuizo"].configure(
            text=str(cnt_p), fg=COLORS["red"] if cnt_p > 0 else COLORS["text_muted"])
        self._lucro_lbl_total.configure(
            text=f"{len(filtrado)} venda(s) exibidas")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: PARCELAMENTOS
    # ══════════════════════════════════════════════════════════════════════════
    def _build_parcelamentos_subs(self, parent):
        self.sub_pages["Parcelamentos"] = {}
        subs = SUBMENUS["Parcelamentos"]
        for sub in subs:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Parcelamentos"][sub] = frame
            if sub == "Financiamentos":
                self._build_financiamentos(frame)
            elif sub == "Previsão de Recebimento":
                self._build_previsao_recebimento(frame)
            elif sub == "Dívidas Clientes":
                self._build_dividas_clientes(frame)
        first = subs[0]
        self.sub_pages["Parcelamentos"][first].place(in_=parent, x=0, y=0, relwidth=1, relheight=1)

    def _build_financiamentos(self, parent):
        import datetime
        _hoje = datetime.date.today()
        self._fin_filtro_var    = tk.StringVar(value="")
        self._fin_filtro_mes    = tk.StringVar(value="Todos")
        self._fin_filtro_ano    = tk.StringVar(value="Todos")
        self._fin_filtro_status = tk.StringVar(value="Todos")
        self._fin_nav_mes       = tk.IntVar(value=_hoje.month)
        self._fin_nav_ano       = tk.IntVar(value=_hoje.year)

        # Definição de colunas (nome, largura px, âncora)
        FIN_COLS = [
            ("Cliente",   140, "w"),
            ("Carro",     120, "w"),
            ("Tipo",       78, "c"),
            ("V.Venda",    82, "w"),
            ("Entrada",    72, "w"),
            ("Parcela",    72, "w"),
            ("Progresso",  90, "w"),
            ("Próx.Parc",  86, "w"),
            ("Saldo",      86, "w"),
            ("Status",     72, "c"),
            ("Ações",      80, "w"),
        ]
        self._fin_cols = FIN_COLS
        TOTAL_W = sum(c[1] for c in FIN_COLS)  # largura total da tabela

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        card = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=COLORS["accent"], height=4).pack(fill="x")

        # Header
        hdr = tk.Frame(card, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12, 4))
        tk.Label(hdr, text="◎  Financiamentos Ativos",
                 font=("Helvetica", 12, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_financiamentos).pack(side="right", padx=(8,0))
        self._lbl_total_fin = tk.Label(hdr, text="0 registros", font=("Helvetica", 8),
                                        bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_total_fin.pack(side="right")

        # ── Navegação de mês ──────────────────────────────────────────────────
        MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun",
                    "Jul","Ago","Set","Out","Nov","Dez"]

        def _fin_prev_mes():
            m = self._fin_nav_mes.get(); a = self._fin_nav_ano.get()
            m -= 1
            if m < 1: m = 12; a -= 1
            self._fin_nav_mes.set(m); self._fin_nav_ano.set(a)
            self._lbl_fin_mes.configure(text=f"{MESES_PT[m-1]} {a}")
            self._build_fin_kpis(); self._refresh_financiamentos()

        def _fin_next_mes():
            m = self._fin_nav_mes.get(); a = self._fin_nav_ano.get()
            m += 1
            if m > 12: m = 1; a += 1
            self._fin_nav_mes.set(m); self._fin_nav_ano.set(a)
            self._lbl_fin_mes.configure(text=f"{MESES_PT[m-1]} {a}")
            self._build_fin_kpis(); self._refresh_financiamentos()

        nav_fin = tk.Frame(card, bg=COLORS["bg_card"])
        nav_fin.pack(fill="x", padx=16, pady=(0,2))
        tk.Button(nav_fin, text="◀", font=("Helvetica",10,"bold"),
                  bg=COLORS["bg_main"], fg=COLORS["accent"], relief="flat",
                  cursor="hand2", padx=6, command=_fin_prev_mes
                  ).pack(side="left", ipady=2)
        _m0 = self._fin_nav_mes.get(); _a0 = self._fin_nav_ano.get()
        self._lbl_fin_mes = tk.Label(nav_fin,
                  text=f"{MESES_PT[_m0-1]} {_a0}",
                  font=("Helvetica",10,"bold"), width=12,
                  bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._lbl_fin_mes.pack(side="left", padx=6)
        tk.Button(nav_fin, text="▶", font=("Helvetica",10,"bold"),
                  bg=COLORS["bg_main"], fg=COLORS["accent"], relief="flat",
                  cursor="hand2", padx=6, command=_fin_next_mes
                  ).pack(side="left", ipady=2)
        tk.Label(nav_fin, text="← filtra parcelas do mês",
                 font=("Helvetica",7), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(side="left", padx=8)

        # Filtros
        import datetime
        anos  = ["Todos"] + [str(y) for y in range(datetime.date.today().year + 1, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]

        fbar = tk.Frame(card, bg=COLORS["bg_main"])
        fbar.pack(fill="x", padx=16, pady=(0, 4))
        tk.Label(fbar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(fbar, textvariable=self._fin_filtro_var, font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"], width=14
                 ).pack(side="left", ipady=4)
        self._fin_filtro_var.trace_add("write", lambda *_: self._refresh_financiamentos())
        tk.Button(fbar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=3,
                  command=lambda: self._fin_filtro_var.set("")
                  ).pack(side="left", padx=(2,8), ipady=3)
        tk.Label(fbar, text="Próx:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._fin_filtro_mes, values=meses,
                     state="readonly", font=("Helvetica", 8), width=3
                     ).pack(side="left", padx=(2,2), ipady=2)
        ttk.Combobox(fbar, textvariable=self._fin_filtro_ano, values=anos,
                     state="readonly", font=("Helvetica", 8), width=5
                     ).pack(side="left", padx=(0,8), ipady=2)
        self._fin_filtro_mes.trace_add("write", lambda *_: self._refresh_financiamentos())
        self._fin_filtro_ano.trace_add("write", lambda *_: self._refresh_financiamentos())
        tk.Label(fbar, text="Status:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._fin_filtro_status,
                     values=["Todos", "Em Aberto", "Quitado"],
                     state="readonly", font=("Helvetica", 8), width=9
                     ).pack(side="left", padx=(2,0), ipady=2)
        self._fin_filtro_status.trace_add("write", lambda *_: self._refresh_financiamentos())
        tk.Button(fbar, text="↺ Atualizar", font=("Helvetica", 9, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_financiamentos
                  ).pack(side="right", ipady=3, padx=(0,4))

        # ── KPIs do mês ──────────────────────────────────────────────────────
        kpi_fin = tk.Frame(card, bg=COLORS["bg_card"])
        kpi_fin.pack(fill="x", padx=16, pady=(2,6))
        self._fin_kpi_frame = kpi_fin
        self._build_fin_kpis()

        # ── Área de tabela com scroll H+V sincronizado ────────────────────────
        # Estrutura: header_canvas (sync h-scroll) + body_canvas (v+h scroll)
        table_area = tk.Frame(card, bg=COLORS["bg_card"])
        table_area.pack(fill="both", expand=True, padx=16, pady=(0,12))

        # Scrollbars
        vsb = tk.Scrollbar(table_area, orient="vertical")
        hsb = tk.Scrollbar(table_area, orient="horizontal")

        # Canvas do cabeçalho (não rola verticalmente)
        self._fin_hdr_canvas = tk.Canvas(table_area, bg=COLORS["bg_main"],
                                          height=28, highlightthickness=0)
        # Canvas do corpo (rola em ambas as direções)
        self._fin_body_canvas = tk.Canvas(table_area, bg=COLORS["bg_card"],
                                           highlightthickness=0,
                                           yscrollcommand=vsb.set,
                                           xscrollcommand=hsb.set)

        # Sincroniza scroll horizontal entre header e body
        def _sync_x(*args):
            self._fin_hdr_canvas.xview(*args)
            self._fin_body_canvas.xview(*args)
        hsb.config(command=_sync_x)
        vsb.config(command=self._fin_body_canvas.yview)

        # Grid layout
        self._fin_hdr_canvas.grid(row=0, column=0, sticky="ew")
        self._fin_body_canvas.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        table_area.grid_rowconfigure(1, weight=1)
        table_area.grid_columnconfigure(0, weight=1)

        # Frame dentro do header canvas
        hdr_frame = tk.Frame(self._fin_hdr_canvas, bg=COLORS["bg_main"])
        hdr_win = self._fin_hdr_canvas.create_window((0, 0), window=hdr_frame, anchor="nw")
        for name, pw, anch in FIN_COLS:
            f = tk.Frame(hdr_frame, bg=COLORS["bg_main"], width=pw, height=28)
            f.pack_propagate(False)
            f.pack(side="left")
            tk.Label(f, text=name, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w" if anch == "w" else "center"
                     ).pack(fill="both", padx=4)
        # Separador inferior do header
        tk.Frame(hdr_frame, bg=COLORS["border"], height=1, width=TOTAL_W).pack(fill="x")

        # Frame dentro do body canvas
        self._fin_rows_frame = tk.Frame(self._fin_body_canvas, bg=COLORS["bg_card"])
        body_win = self._fin_body_canvas.create_window((0, 0), window=self._fin_rows_frame, anchor="nw")

        def _on_rows_configure(e):
            self._fin_body_canvas.configure(scrollregion=self._fin_body_canvas.bbox("all"))
        self._fin_rows_frame.bind("<Configure>", _on_rows_configure)

        # Ao redimensionar canvas, ajusta a largura mínima dos frames internos
        def _on_canvas_resize(e, _win=body_win, _hw=hdr_win):
            w = max(e.width, TOTAL_W)
            self._fin_body_canvas.itemconfig(_win, width=w)
            self._fin_hdr_canvas.itemconfig(_hw, width=w)
            self._fin_hdr_canvas.configure(scrollregion=(0, 0, max(e.width, TOTAL_W), 28))
        self._fin_body_canvas.bind("<Configure>", _on_canvas_resize)

        # Sincroniza scroll do header quando body move no eixo X
        def _body_xmove(first, last):
            hsb.set(first, last)
            self._fin_hdr_canvas.xview_moveto(first)
        self._fin_body_canvas.configure(xscrollcommand=_body_xmove)

        # Mouse wheel
        def _mwheel(ev):
            self._fin_body_canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        card.bind("<Enter>", lambda e: card.bind_all("<MouseWheel>", _mwheel))
        card.bind("<Leave>", lambda e: card.unbind_all("<MouseWheel>"))

        self._refresh_financiamentos()

    def _calc_fin_kpis(self, mes=None, ano=None):
        """Calcula KPIs de parcelas para o mês/ano indicado (padrão=hoje)."""
        import datetime
        hoje = datetime.date.today()
        if mes is None: mes = hoje.month
        if ano is None: ano = hoje.year
        ms = f"{mes:02d}"; ys = str(ano)

        todas = [dict(r) for r in self.conn.execute(
            "SELECT * FROM vendas WHERE tipo_venda IN "
            "('Venda Parcelada','Com Troca','Com Troca e Volta')").fetchall()]

        # A receber = valor das parcelas cujo vencimento cai neste mês
        a_receber = 0
        for v in todas:
            pagas      = v.get("parcelas_pagas") or 0
            total_parc = v.get("num_parcelas")   or 0
            parcela_v  = 0
            try: parcela_v = int(str(v.get("parcela_mensal") or "0").replace(",",""))
            except: pass
            # Percorre todas as parcelas (pagas e em aberto) procurando vencimento no mês
            for n in range(0, total_parc):
                prox = self._proxima_parcela(v.get("data_primeira_parc"), n)
                if prox == "—": continue
                pts = prox.split("/")
                if len(pts) < 3: continue
                if pts[1] == ms and pts[2] == ys:
                    a_receber += parcela_v

        # Pagas no mês = registros em pagamentos com data no mês
        pagas_mes = 0
        for v in todas:
            rows = self.conn.execute(
                "SELECT valor_pago FROM pagamentos WHERE venda_id=? "
                "AND (data_pagamento LIKE ? OR data_pagamento LIKE ?)",
                (v["id"], f"__/{ms}/{ys}", f"%/{ms}/{ys}")).fetchall()
            for p in rows:
                try: pagas_mes += int(str(p[0]).replace(",",""))
                except: pass

        # Em aberto = parcela mais próxima vencida até hoje ainda não paga
        aberto_mes = 0
        for v in todas:
            pagas_v    = v.get("parcelas_pagas") or 0
            total_parc = v.get("num_parcelas")   or 0
            if pagas_v >= total_parc: continue
            parcela_v  = 0
            try: parcela_v = int(str(v.get("parcela_mensal") or "0").replace(",",""))
            except: pass
            prox = self._proxima_parcela(v.get("data_primeira_parc"), pagas_v)
            if prox == "—": continue
            pts = prox.split("/")
            if len(pts) < 3: continue
            try:
                d_prox = datetime.date(int(pts[2]), int(pts[1]), int(pts[0]))
                if d_prox <= hoje:
                    aberto_mes += parcela_v
            except: pass
        return a_receber, pagas_mes, aberto_mes

    def _build_fin_kpis(self):
        for w in self._fin_kpi_frame.winfo_children():
            w.destroy()
        mes = getattr(self, "_fin_nav_mes", None)
        ano = getattr(self, "_fin_nav_ano", None)
        m = mes.get() if mes else None
        a = ano.get() if ano else None
        a_receber, pagas_mes, aberto_mes = self._calc_fin_kpis(mes=m, ano=a)
        kpis = [
            ("A Receber (Mês)", f"¥ {a_receber:,}", COLORS["blue"]),
            ("Pagas (Mês)",     f"¥ {pagas_mes:,}", COLORS["green"]),
            ("Em Aberto",       f"¥ {aberto_mes:,}", COLORS["orange"]),
        ]
        for label, val, cor in kpis:
            c = tk.Frame(self._fin_kpi_frame, bg=COLORS["bg_content"],
                         highlightthickness=1, highlightbackground=cor)
            c.pack(side="left", padx=(0,8), pady=4, ipadx=10, ipady=4)
            tk.Frame(c, bg=cor, height=3).pack(fill="x")
            tk.Label(c, text=val, font=("Helvetica",12,"bold"),
                     bg=COLORS["bg_content"], fg=cor).pack(padx=10, pady=(4,0))
            tk.Label(c, text=label, font=("Helvetica",7),
                     bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(padx=10, pady=(0,4))

    def _build_parc_kpis(self):
        for w in self._parc_kpi_frame.winfo_children():
            w.destroy()
        mes = getattr(self, "_parc_nav_mes", None)
        ano = getattr(self, "_parc_nav_ano", None)
        m = mes.get() if mes else None
        a = ano.get() if ano else None
        a_receber, pagas_mes, aberto_mes = self._calc_fin_kpis(mes=m, ano=a)
        kpis = [
            ("A Receber (Mês)", f"¥ {a_receber:,}", COLORS["blue"]),
            ("Pagas (Mês)",     f"¥ {pagas_mes:,}", COLORS["green"]),
            ("Em Aberto",       f"¥ {aberto_mes:,}", COLORS["orange"]),
        ]
        for label, val, cor in kpis:
            c = tk.Frame(self._parc_kpi_frame, bg=COLORS["bg_content"],
                         highlightthickness=1, highlightbackground=cor)
            c.pack(side="left", padx=(0,8), pady=4, ipadx=10, ipady=4)
            tk.Frame(c, bg=cor, height=3).pack(fill="x")
            tk.Label(c, text=val, font=("Helvetica",12,"bold"),
                     bg=COLORS["bg_content"], fg=cor).pack(padx=10, pady=(4,0))
            tk.Label(c, text=label, font=("Helvetica",7),
                     bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(padx=10, pady=(0,4))

    def _proxima_parcela(self, data_primeira_parc, parcelas_pagas):
        """Calcula a data da próxima parcela baseada na data da 1ª e parcelas já pagas."""
        import datetime
        if not data_primeira_parc:
            return "—"
        try:
            partes = data_primeira_parc.split("/")
            d = int(partes[0]); m = int(partes[1]); a = int(partes[2])
            mes_total = m + parcelas_pagas
            ano = a + (mes_total - 1) // 12
            mes = ((mes_total - 1) % 12) + 1
            # Dia: mantém o dia original, ajusta para último dia do mês se necessário
            import calendar
            ultimo_dia = calendar.monthrange(ano, mes)[1]
            dia = min(d, ultimo_dia)
            return f"{dia:02d}/{mes:02d}/{ano}"
        except Exception:
            return "—"

    def _marcar_parcela_paga(self, venda_id):
        """Abre diálogo para registrar pagamento com data e valor variável."""
        v = self.conn.execute("SELECT * FROM vendas WHERE id=?", (venda_id,)).fetchone()
        if not v:
            return
        v = dict(v)
        pagas = v.get("parcelas_pagas") or 0
        total = v.get("num_parcelas") or 0
        p_int = int(str(v.get("parcela_mensal") or "0").replace(",", ""))
        ult   = int(str(v.get("valor_ultima_parc") or "0").replace(",", ""))
        # Valor padrão = parcela ou última se for a última
        restam = max(0, total - pagas)
        val_padrao = ult if restam == 1 and ult > 0 else p_int

        import datetime
        dlg = tk.Toplevel(self)
        dlg.title("Registrar Pagamento")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("360x290")
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 360) // 2
        y = self.winfo_y() + (self.winfo_height() - 290) // 2
        dlg.geometry(f"+{x}+{y}")

        tk.Frame(dlg, bg=COLORS["green"], height=4).pack(fill="x")
        tk.Label(dlg, text="Registrar Pagamento",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(pady=(12, 2))
        tk.Label(dlg, text=f"Parcela {pagas+1}/{total}  |  {v.get('carro','—')}",
                 font=("Helvetica", 8), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack()

        tk.Frame(dlg, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(8, 0))

        # Data de pagamento
        tk.Label(dlg, text="Data do Pagamento", font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20, pady=(10, 0))
        df = tk.Frame(dlg, bg=COLORS["bg_card"])
        df.pack(anchor="w", padx=20, pady=(4, 0))
        d_dia, d_mes, d_ano = self._make_date_row(df)
        hoje = datetime.date.today()
        for e, val in [(d_dia, str(hoje.day).zfill(2)),
                       (d_mes, str(hoje.month).zfill(2)),
                       (d_ano, str(hoje.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])

        # Valor pago
        tk.Label(dlg, text="Valor Pago (¥)", font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20, pady=(12, 0))
        val_entry = self._make_yen_entry(dlg, width=20)
        val_entry.pack(padx=20, pady=(4, 0), ipady=7)
        if val_padrao:
            val_entry.insert(0, f"{val_padrao:,}")
        val_entry.bind("<KeyRelease>", lambda e: self._fmt_yen(val_entry))

        lbl_err = tk.Label(dlg, text="", font=("Helvetica", 8),
                           bg=COLORS["bg_card"], fg=COLORS["red"])
        lbl_err.pack(pady=(4, 0))

        btn_row = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_row.pack(pady=(8, 16))

        def confirmar():
            data_pag = self._get_date_from_entries(d_dia, d_mes, d_ano)
            raw = val_entry.get().replace(",", "").strip()
            try:
                valor_pago = int(raw)
            except ValueError:
                lbl_err.configure(text="⚠ Valor inválido."); return
            if not data_pag:
                lbl_err.configure(text="⚠ Informe a data."); return

            # Registra pagamento na tabela
            num_parc = pagas + 1
            self.conn.execute(
                "INSERT INTO pagamentos (venda_id, num_parcela, data_pagamento, valor_pago) "
                "VALUES (?,?,?,?)", (venda_id, num_parc, data_pag, str(valor_pago)))

            # Recalcula saldo devedor e ajusta parcelas
            self._recalcular_saldo_venda(venda_id, valor_pago)
            dlg.destroy()
            self._refresh_financiamentos()

        tk.Button(btn_row, text="  Confirmar  ", font=("Helvetica", 10, "bold"),
                  bg=COLORS["green"], fg="white", relief="flat", cursor="hand2",
                  command=confirmar).pack(side="left", ipady=5, ipadx=6)
        tk.Button(btn_row, text="  Cancelar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(side="left", padx=(10, 0), ipady=5, ipadx=6)

    def _recalcular_saldo_venda(self, venda_id, valor_pago_agora=0):
        """Após qualquer pagamento, recalcula num_parcelas e ultima baseado no saldo real."""
        v = dict(self.conn.execute("SELECT * FROM vendas WHERE id=?", (venda_id,)).fetchone())

        p_int = int(str(v.get("parcela_mensal") or "0").replace(",", ""))

        # Saldo original = valor_venda - entrada - valor_troca
        try:
            vv = int(str(v.get("valor_venda") or "0").replace(",",""))
            ev = int(str(v.get("entrada") or "0").replace(",",""))
            vt = int(str(v.get("valor_troca") or "0").replace(",",""))
            saldo_total_orig = max(0, vv - ev - vt)
        except Exception:
            saldo_total_orig = 0

        # Soma todos os pagamentos registrados
        novas_pagas = self.conn.execute(
            "SELECT COUNT(*) FROM pagamentos WHERE venda_id=?", (venda_id,)).fetchone()[0]
        total_pago = self.conn.execute(
            "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) FROM pagamentos WHERE venda_id=?",
            (venda_id,)).fetchone()[0]

        saldo_restante = max(0, saldo_total_orig - total_pago)

        # Recalcula num_parcelas e ultima com base no saldo restante e parcela padrão
        if saldo_restante == 0 or p_int == 0:
            novas_parc = novas_pagas
            nova_ultima = 0
        else:
            n = saldo_restante // p_int
            r = saldo_restante % p_int
            novas_parc = novas_pagas + n + (1 if r > 0 else 0)
            nova_ultima = r if r > 0 else 0

        self.conn.execute(
            "UPDATE vendas SET parcelas_pagas=?, num_parcelas=?, valor_ultima_parc=? WHERE id=?",
            (novas_pagas, novas_parc, str(nova_ultima), venda_id))
        self.conn.commit()

    def _abrir_historico_pagamentos(self, venda_id):
        """Abre janela com histórico de pagamentos — editar e excluir individualmente."""
        import datetime

        def get_venda():
            return dict(self.conn.execute("SELECT * FROM vendas WHERE id=?", (venda_id,)).fetchone())

        v = get_venda()
        p_int_orig = int(str(v.get("parcela_mensal") or "0").replace(",", ""))
        try:
            val_venda = int(str(v.get("valor_venda") or "0").replace(",",""))
            entrada   = int(str(v.get("entrada") or "0").replace(",",""))
            val_troca = int(str(v.get("valor_troca") or "0").replace(",",""))
            saldo_orig = val_venda - entrada - val_troca
        except Exception:
            saldo_orig = 0

        dlg = tk.Toplevel(self)
        dlg.title(f"Pagamentos — {v.get('carro','—')}")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(True, True)
        dlg.geometry("560x460")
        dlg.grab_set()
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 560) // 2
        y = self.winfo_y() + (self.winfo_height() - 460) // 2
        dlg.geometry(f"+{x}+{y}")

        tk.Frame(dlg, bg=COLORS["blue"], height=4).pack(fill="x")

        # Cabeçalho
        hdr_f = tk.Frame(dlg, bg=COLORS["bg_card"])
        hdr_f.pack(fill="x", padx=16, pady=(12, 0))
        tk.Label(hdr_f, text=f"📋  {v.get('carro','—')}",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        self._lbl_pag_resumo = tk.Label(hdr_f, text="",
                                         font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_pag_resumo.pack(side="right")

        # Info parcela
        tk.Label(dlg, text=f"Parcela padrão: ¥ {p_int_orig:,}  |  Saldo total: ¥ {saldo_orig:,}",
                 font=("Helvetica", 8), bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(anchor="w", padx=16, pady=(2, 0))

        tk.Frame(dlg, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(8, 0))

        # Cabeçalho tabela
        col_h = tk.Frame(dlg, bg=COLORS["bg_main"])
        col_h.pack(fill="x", padx=16, pady=(4, 0))
        for txt, w in [("Nº",5),("Data",13),("Valor Pago",14),("Ações",14)]:
            tk.Label(col_h, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=4, pady=4)

        # Área scroll
        sf = tk.Frame(dlg, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=16, pady=(0, 4))
        cv = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(sf, orient="vertical", command=cv.yview)
        rows_frame = tk.Frame(cv, bg=COLORS["bg_card"])
        rows_frame.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=rows_frame, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        def recalcular_venda():
            """Recalcula parcelas_pagas, num_parcelas e ultima a partir dos pagamentos."""
            nova_contagem = self.conn.execute(
                "SELECT COUNT(*) FROM pagamentos WHERE venda_id=?", (venda_id,)).fetchone()[0]
            total_pago = self.conn.execute(
                "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) FROM pagamentos WHERE venda_id=?",
                (venda_id,)).fetchone()[0]
            saldo_rest = max(0, saldo_orig - total_pago)
            if p_int_orig > 0 and saldo_rest > 0:
                n = saldo_rest // p_int_orig
                r = saldo_rest % p_int_orig
                novas_parc = nova_contagem + n + (1 if r > 0 else 0)
                nova_ult   = r if r > 0 else 0
            else:
                novas_parc = nova_contagem
                nova_ult   = 0
            self.conn.execute(
                "UPDATE vendas SET parcelas_pagas=?, num_parcelas=?, valor_ultima_parc=? WHERE id=?",
                (nova_contagem, novas_parc, str(nova_ult), venda_id))
            self.conn.commit()
            return nova_contagem, novas_parc, saldo_orig - total_pago

        def update_resumo_and_main():
            """Atualiza resumo E a tela de financiamentos."""
            update_resumo()
            self._refresh_financiamentos()

        def update_resumo():
            nova_contagem, novas_parc, saldo_rest = recalcular_venda()
            saldo_rest = max(0, saldo_rest)
            txt = f"{nova_contagem} pago(s) / {novas_parc} total  |  Saldo: ¥ {saldo_rest:,}"
            self._lbl_pag_resumo.configure(text=txt)

        def populate():
            for w in rows_frame.winfo_children():
                w.destroy()
            pags = [dict(r) for r in self.conn.execute(
                "SELECT * FROM pagamentos WHERE venda_id=? ORDER BY num_parcela",
                (venda_id,)).fetchall()]
            total_pago = sum(int(str(p["valor_pago"]).replace(",","")) for p in pags)

            if not pags:
                tk.Label(rows_frame, text="Nenhum pagamento registrado ainda.",
                         font=("Helvetica", 9), bg=COLORS["bg_card"],
                         fg=COLORS["text_muted"]).pack(pady=20, padx=16)
            else:
                for j, p in enumerate(pags):
                    rb = COLORS["bg_card"] if j % 2 == 0 else COLORS["bg_content"]
                    r = tk.Frame(rows_frame, bg=rb)
                    r.pack(fill="x")
                    tk.Label(r, text=str(p["num_parcela"]),
                             font=("Helvetica",9,"bold"),
                             bg=rb, fg=COLORS["accent"], width=5, anchor="w"
                             ).pack(side="left", padx=4, pady=7)
                    tk.Label(r, text=p["data_pagamento"],
                             font=("Helvetica",9),
                             bg=rb, fg=COLORS["text_secondary"], width=13, anchor="w"
                             ).pack(side="left", padx=4)
                    try:
                        vp = int(str(p["valor_pago"]).replace(",",""))
                        vp_txt = f"¥ {vp:,}"
                    except Exception:
                        vp_txt = str(p["valor_pago"])
                    tk.Label(r, text=vp_txt, font=("Helvetica",9,"bold"),
                             bg=rb, fg=COLORS["green"], width=14, anchor="w"
                             ).pack(side="left", padx=4)
                    acts = tk.Frame(r, bg=rb)
                    acts.pack(side="left", padx=4)
                    tk.Button(acts, text="✏", font=("Helvetica",8),
                              bg=COLORS["blue"], fg="white", relief="flat",
                              cursor="hand2", padx=4, pady=1,
                              command=lambda p=p: editar_pag(p)
                              ).pack(side="left", padx=(0,2))
                    tk.Button(acts, text="✕", font=("Helvetica",8),
                              bg=COLORS["red"], fg="white", relief="flat",
                              cursor="hand2", padx=4, pady=1,
                              command=lambda pid=p["id"]: excluir_pag(pid)
                              ).pack(side="left")

                # Total
                tk.Frame(rows_frame, bg=COLORS["border"], height=1).pack(fill="x", padx=4, pady=(4,0))
                tot_row = tk.Frame(rows_frame, bg=COLORS["bg_main"])
                tot_row.pack(fill="x")
                tk.Label(tot_row, text="Total pago:",
                         font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                         width=22, anchor="e").pack(side="left", padx=4, pady=6)
                saldo_rest_txt = max(0, saldo_orig - total_pago)
                tk.Label(tot_row, text=f"¥ {total_pago:,}",
                         font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_main"], fg=COLORS["green"],
                         width=14, anchor="w").pack(side="left", padx=4)
                tk.Label(tot_row, text=f"Saldo: ¥ {saldo_rest_txt:,}",
                         font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_main"],
                         fg=COLORS["orange"] if saldo_rest_txt > 0 else COLORS["green"],
                         anchor="w").pack(side="left", padx=8)

        def excluir_pag(pid):
            if not msgbox.askyesno("Confirmar",
                                    "Excluir este pagamento?\nO saldo será recalculado automaticamente."):
                return
            self.conn.execute("DELETE FROM pagamentos WHERE id=?", (pid,))
            self.conn.commit()
            update_resumo_and_main()
            populate()

        def editar_pag(pag):
            """Abre sub-diálogo para editar data e valor de um pagamento existente."""
            ed = tk.Toplevel(dlg)
            ed.title(f"Editar Parcela {pag['num_parcela']}")
            ed.configure(bg=COLORS["bg_card"])
            ed.resizable(False, False)
            ed.geometry("340x240")
            ed.grab_set()
            ed.update_idletasks()
            ex = dlg.winfo_x() + (dlg.winfo_width() - 340) // 2
            ey = dlg.winfo_y() + (dlg.winfo_height() - 240) // 2
            ed.geometry(f"+{ex}+{ey}")

            tk.Frame(ed, bg=COLORS["blue"], height=4).pack(fill="x")
            tk.Label(ed, text=f"Editar Parcela {pag['num_parcela']}",
                     font=("Helvetica", 11, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(pady=(12,2))

            tk.Label(ed, text="Data do Pagamento", font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20, pady=(8,0))
            df_ed = tk.Frame(ed, bg=COLORS["bg_card"])
            df_ed.pack(anchor="w", padx=20, pady=(4,0))
            e_dia, e_mes, e_ano = self._make_date_row(df_ed)
            # Preenche data atual
            pts = pag["data_pagamento"].split("/")
            for e, val in [(e_dia, pts[0] if pts else ""),
                           (e_mes, pts[1] if len(pts)>1 else ""),
                           (e_ano, pts[2] if len(pts)>2 else "")]:
                e.delete(0, tk.END); e.insert(0, val)
                e.configure(fg=COLORS["text_primary"])

            tk.Label(ed, text="Valor Pago (¥)", font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20, pady=(10,0))
            val_ed = self._make_yen_entry(ed, width=20)
            val_ed.pack(padx=20, pady=(4,0), ipady=7)
            try:
                vp_curr = int(str(pag["valor_pago"]).replace(",",""))
                val_ed.insert(0, f"{vp_curr:,}")
            except Exception:
                pass
            val_ed.bind("<KeyRelease>", lambda e: self._fmt_yen(val_ed))

            lbl_ed_err = tk.Label(ed, text="", font=("Helvetica",8),
                                   bg=COLORS["bg_card"], fg=COLORS["red"])
            lbl_ed_err.pack()

            def salvar_edicao():
                data_nova = self._get_date_from_entries(e_dia, e_mes, e_ano)
                raw = val_ed.get().replace(",","").strip()
                try:
                    val_novo = int(raw)
                except ValueError:
                    lbl_ed_err.configure(text="⚠ Valor inválido."); return
                if not data_nova:
                    lbl_ed_err.configure(text="⚠ Informe a data."); return
                self.conn.execute(
                    "UPDATE pagamentos SET data_pagamento=?, valor_pago=? WHERE id=?",
                    (data_nova, str(val_novo), pag["id"]))
                self.conn.commit()
                update_resumo_and_main()
                populate()
                ed.destroy()

            btn_r = tk.Frame(ed, bg=COLORS["bg_card"])
            btn_r.pack(pady=(4,12))
            tk.Button(btn_r, text="  Salvar  ", font=("Helvetica",10,"bold"),
                      bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                      command=salvar_edicao).pack(side="left", ipady=5, ipadx=6)
            tk.Button(btn_r, text="  Cancelar  ", font=("Helvetica",10),
                      bg=COLORS["border"], fg=COLORS["text_secondary"],
                      relief="flat", cursor="hand2",
                      command=ed.destroy).pack(side="left", padx=(8,0), ipady=5, ipadx=6)

        # Botão de fechar
        tk.Button(dlg, text="  Fechar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(pady=(4, 12), ipadx=6, ipady=4)

        update_resumo()
        populate()

    def _refresh_financiamentos(self):
        import datetime
        for w in self._fin_rows_frame.winfo_children():
            w.destroy()

        # Mês/ano do navegador de mês
        nav_m  = getattr(self, "_fin_nav_mes", None)
        nav_a  = getattr(self, "_fin_nav_ano", None)
        nav_ms = f"{nav_m.get():02d}" if nav_m else None
        nav_ys = str(nav_a.get())      if nav_a else None

        todas = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id = c.id "
            "WHERE v.tipo_venda IN ('Venda Parcelada','Com Troca','Com Troca e Volta') "
            "ORDER BY v.id DESC").fetchall()]

        # Filtro de texto
        termo = self._fin_filtro_var.get().strip().lower()
        if termo:
            todas = [v for v in todas if
                     termo in (v.get("carro") or "").lower() or
                     termo in (v.get("cliente_nome") or "").lower()]

        # Filtro do navegador: mantém apenas vendas com parcela no mês/ano navegado
        if nav_ms and nav_ys:
            def tem_parcela_no_mes(v):
                total_parc = v.get("num_parcelas") or 0
                for n in range(total_parc):
                    prx = self._proxima_parcela(v.get("data_primeira_parc"), n)
                    if prx == "—": continue
                    pts = prx.split("/")
                    if len(pts) >= 3 and pts[1] == nav_ms and pts[2] == nav_ys:
                        return True
                return False
            todas = [v for v in todas if tem_parcela_no_mes(v)]

        # Filtros de combobox de mês/ano (filtro de próxima parcela)
        fmes = self._fin_filtro_mes.get()
        fano = self._fin_filtro_ano.get()
        if fmes != "Todos" or fano != "Todos":
            def match_data(v):
                pagas = v.get("parcelas_pagas") or 0
                prox = self._proxima_parcela(v.get("data_primeira_parc"), pagas)
                if prox == "\u2014": return False
                pts = prox.split("/")
                vm = pts[1] if len(pts) > 1 else ""
                va = pts[2] if len(pts) > 2 else ""
                if fmes != "Todos" and vm != fmes: return False
                if fano != "Todos" and va != fano: return False
                return True
            todas = [v for v in todas if match_data(v)]

        fstatus = self._fin_filtro_status.get()
        if fstatus == "Em Aberto":
            todas = [v for v in todas if
                     (v.get("parcelas_pagas") or 0) < (v.get("num_parcelas") or 0)]
        elif fstatus == "Quitado":
            todas = [v for v in todas if
                     (v.get("num_parcelas") or 0) > 0 and
                     (v.get("parcelas_pagas") or 0) >= (v.get("num_parcelas") or 0)]

        self._lbl_total_fin.configure(text=f"{len(todas)} financiamento(s)")

        if not todas:
            tk.Label(self._fin_rows_frame,
                     text="Nenhum financiamento encontrado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).grid(row=0, column=0, pady=40)
            return

        FIN_COLS = getattr(self, "_fin_cols", [
            ("Cliente",140,"w"),("Carro",120,"w"),("Tipo",78,"c"),
            ("V.Venda",82,"w"),("Entrada",72,"w"),("Parcela",72,"w"),
            ("Progresso",90,"w"),("Prox.Parc",86,"w"),("Saldo",86,"w"),
            ("Status",72,"c"),("Acoes",80,"w"),
        ])
        for ci, (_, pw, _a) in enumerate(FIN_COLS):
            self._fin_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        TIPO_COLORS = {"Venda Parcelada": COLORS["blue"],
                       "Com Troca": COLORS["orange"],
                       "Com Troca e Volta": COLORS["accent"]}
        TIPO_SHORT  = {"Venda Parcelada": "Parcelada",
                       "Com Troca": "C/Troca",
                       "Com Troca e Volta": "T+Volta"}

        for i, v in enumerate(todas):
            pagas      = v.get("parcelas_pagas") or 0
            total_parc = v.get("num_parcelas") or 0
            quitado    = total_parc > 0 and pagas >= total_parc

            try:
                total_pago = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                    "FROM pagamentos WHERE venda_id=?", (v["id"],)).fetchone()[0]
                vv = int(str(v.get("valor_venda") or "0").replace(",",""))
                ev = int(str(v.get("entrada") or "0").replace(",",""))
                vt = int(str(v.get("valor_troca") or "0").replace(",",""))
                saldo_rest = max(0, vv - ev - vt - total_pago)
            except Exception:
                saldo_rest = 0

            # Detecta se a parcela do mês navegado já foi paga
            pago_no_mes = False
            if nav_ms and nav_ys and not quitado:
                for n in range(total_parc):
                    prx = self._proxima_parcela(v.get("data_primeira_parc"), n)
                    if prx == "—": continue
                    pts = prx.split("/")
                    if len(pts) >= 3 and pts[1] == nav_ms and pts[2] == nav_ys:
                        p_row = self.conn.execute(
                            "SELECT id FROM pagamentos WHERE venda_id=? AND num_parcela=?",
                            (v["id"], n + 1)).fetchone()
                        if p_row:
                            pago_no_mes = True
                        break

            rb  = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            bdr = COLORS["green"] if (quitado or pago_no_mes) else COLORS["border"]

            tk.Frame(self._fin_rows_frame, bg=bdr, height=1
                     ).grid(row=i*2, column=0, columnspan=len(FIN_COLS), sticky="ew")
            ri = i*2 + 1
            self._fin_rows_frame.grid_rowconfigure(ri, minsize=36)

            tk.Label(self._fin_rows_frame,
                     text=(v.get("cliente_nome") or "\u2014")[:18],
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_primary"],
                     anchor="w", padx=4).grid(row=ri, column=0, sticky="nsew")

            car = v.get("carro") or "\u2014"
            car = car[:15]+"\u2026" if len(car) > 16 else car
            tk.Label(self._fin_rows_frame, text=car,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=1, sticky="nsew")

            tc = TIPO_COLORS.get(v["tipo_venda"], COLORS["text_muted"])
            f2 = tk.Frame(self._fin_rows_frame, bg=rb)
            f2.grid(row=ri, column=2, sticky="nsew")
            tk.Label(f2, text=TIPO_SHORT.get(v["tipo_venda"], "\u2014"),
                     font=("Helvetica",7,"bold"), bg=tc, fg="white", anchor="center"
                     ).pack(padx=3, pady=6, fill="both", expand=True)

            tk.Label(self._fin_rows_frame,
                     text=self._fmt_yen_display(v.get("valor_venda")),
                     font=("Helvetica",8,"bold"), bg=rb, fg=COLORS["green"],
                     anchor="w", padx=4).grid(row=ri, column=3, sticky="nsew")

            ent_txt = self._fmt_yen_display(v.get("entrada")) if v.get("entrada") else "\u2014"
            tk.Label(self._fin_rows_frame, text=ent_txt,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=4, sticky="nsew")

            tk.Label(self._fin_rows_frame,
                     text=self._fmt_yen_display(v.get("parcela_mensal")),
                     font=("Helvetica",8), bg=rb, fg=COLORS["blue"],
                     anchor="w", padx=4).grid(row=ri, column=5, sticky="nsew")

            prog_txt = f"{pagas}/{total_parc}" if total_parc else "\u2014"
            pc = COLORS["green"] if quitado else COLORS["accent"]
            pf = tk.Frame(self._fin_rows_frame, bg=rb)
            pf.grid(row=ri, column=6, sticky="nsew")
            tk.Label(pf, text=prog_txt, font=("Helvetica",7,"bold"),
                     bg=rb, fg=pc, anchor="w").pack(anchor="w", padx=4, pady=(4,0))
            if total_parc > 0:
                pw6 = FIN_COLS[6][1] - 12
                bw = max(2, int(pw6 * min(pagas / total_parc, 1.0)))
                bc = tk.Canvas(pf, width=pw6, height=6,
                               bg=COLORS["border"], highlightthickness=0)
                bc.pack(padx=4, pady=(1,4))
                bc.create_rectangle(0, 0, bw, 6, fill=pc, outline="")

            prox_txt = "\u2714 Quitado" if quitado else self._proxima_parcela(
                v.get("data_primeira_parc"), pagas)
            prox_fg = COLORS["green"] if quitado else COLORS["text_primary"]
            tk.Label(self._fin_rows_frame, text=prox_txt,
                     font=("Helvetica",8), bg=rb, fg=prox_fg,
                     anchor="w", padx=4).grid(row=ri, column=7, sticky="nsew")

            saldo_txt = f"\xa5{saldo_rest:,}" if not quitado else "\u2014"
            saldo_fg  = COLORS["orange"] if not quitado else COLORS["text_muted"]
            tk.Label(self._fin_rows_frame, text=saldo_txt,
                     font=("Helvetica",8,"bold"), bg=rb, fg=saldo_fg,
                     anchor="w", padx=4).grid(row=ri, column=8, sticky="nsew")

            # Status: Quitado / Pago (mês navegado) / Em Aberto
            if quitado:
                st_txt, st_bg = "Quitado", COLORS["green"]
            elif pago_no_mes:
                st_txt, st_bg = "Pago", COLORS["green"]
            else:
                st_txt, st_bg = "Em Aberto", COLORS["blue"]
            f9 = tk.Frame(self._fin_rows_frame, bg=rb)
            f9.grid(row=ri, column=9, sticky="nsew")
            tk.Label(f9, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_bg, fg="white", anchor="center"
                     ).pack(padx=3, pady=6, fill="both", expand=True)

            fa = tk.Frame(self._fin_rows_frame, bg=rb)
            fa.grid(row=ri, column=10, sticky="nsew")
            if not quitado:
                tk.Button(fa, text="\u2714", font=("Helvetica",7,"bold"),
                          bg=COLORS["green"], fg="white", relief="flat",
                          cursor="hand2", padx=2,
                          command=lambda vid=v["id"]: self._marcar_parcela_paga(vid)
                          ).pack(side="left", padx=(3,2), pady=5)
            tk.Button(fa, text="\U0001f4cb", font=("Helvetica",8),
                      bg=COLORS["blue"], fg="white", relief="flat",
                      cursor="hand2", padx=2,
                      command=lambda vid=v["id"]: self._abrir_historico_pagamentos(vid)
                      ).pack(side="left", padx=(0,2), pady=5)

        tk.Frame(self._fin_rows_frame, bg=COLORS["border"], height=1
                 ).grid(row=len(todas)*2, column=0,
                        columnspan=len(FIN_COLS), sticky="ew")


    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: TIPOS DE SERVIÇO (Cadastros)
    # ══════════════════════════════════════════════════════════════════════════
    def _build_tipos_servico_cadastro(self, parent):
        container = tk.Frame(parent, bg=COLORS["bg_content"])
        container.pack(fill="both", expand=True, padx=24, pady=20)

        form_card = tk.Frame(container, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        form_card.pack(side="left", fill="y", padx=(0, 16), ipadx=10)
        tk.Frame(form_card, bg=COLORS["accent"], height=4).pack(fill="x")
        tk.Label(form_card, text="⚙  Tipos de Serviço", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=20, pady=(14, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=(0, 12))

        tk.Label(form_card, text="Nome do Tipo", font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        self._ts_nome_entry = tk.Entry(form_card, font=("Helvetica", 11),
                                       bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                       insertbackground=COLORS["text_primary"],
                                       relief="flat", bd=0,
                                       highlightthickness=1,
                                       highlightbackground=COLORS["border"],
                                       highlightcolor=COLORS["accent"], width=26)
        self._ts_nome_entry.pack(padx=20, pady=(4, 14), ipady=7)

        self._lbl_ts_status = tk.Label(form_card, text="", font=("Helvetica", 8),
                                        bg=COLORS["bg_card"], fg=COLORS["red"])
        self._lbl_ts_status.pack(padx=20)

        btn_f = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_f.pack(padx=20, pady=(6, 20), fill="x")
        self._btn_salvar_ts = tk.Button(btn_f, text="  Salvar  ",
                                         font=("Helvetica", 10, "bold"),
                                         bg=COLORS["accent"], fg="white",
                                         relief="flat", cursor="hand2",
                                         command=self._salvar_tipo_servico)
        self._btn_salvar_ts.pack(side="left", ipady=6, ipadx=6)
        tk.Button(btn_f, text="  Limpar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_ts
                  ).pack(side="left", padx=(8, 0), ipady=6, ipadx=6)
        self._ts_edit_nome = None

        # Tabela
        table_card = tk.Frame(container, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True)
        tk.Frame(table_card, bg=COLORS["accent"], height=4).pack(fill="x")

        tbl_h = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_h.pack(fill="x", padx=20, pady=(14, 4))
        tk.Label(tbl_h, text="⚙  Tipos Cadastrados", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_h, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_tabela_ts).pack(side="right", padx=(8,0))
        self._lbl_total_ts = tk.Label(tbl_h, text="", font=("Helvetica", 8),
                                       bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_total_ts.pack(side="right")

        tk.Frame(table_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20)
        col_f = tk.Frame(table_card, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=20)
        for txt, w in [("#",4),("Nome do Tipo",30),("Em Uso",8),("Ações",10)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=4, pady=6)

        scroll_f = tk.Frame(table_card, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=20, pady=(0, 14))
        canv = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(scroll_f, orient="vertical", command=canv.yview)
        self._ts_rows_frame = tk.Frame(canv, bg=COLORS["bg_card"])
        self._ts_rows_frame.bind("<Configure>",
            lambda e: canv.configure(scrollregion=canv.bbox("all")))
        canv.create_window((0, 0), window=self._ts_rows_frame, anchor="nw")
        canv.configure(yscrollcommand=sb.set)
        canv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._refresh_tabela_ts()

    def _tipo_servico_em_uso(self, nome):
        try:
            return self.conn.execute(
                "SELECT COUNT(*) FROM servicos WHERE tipo_servico=?", (nome,)).fetchone()[0] > 0
        except Exception:
            return False

    def _salvar_tipo_servico(self):
        nome = self._ts_nome_entry.get().strip()
        if not nome:
            self._lbl_ts_status.configure(text="⚠ Informe o nome.", fg=COLORS["red"]); return
        if self._ts_edit_nome is not None:
            if self._tipo_servico_em_uso(self._ts_edit_nome):
                self._lbl_ts_status.configure(
                    text="⚠ Tipo em uso — não pode ser editado.", fg=COLORS["red"]); return
            try:
                self.conn.execute("UPDATE tipos_servico SET nome=? WHERE nome=?",
                                  (nome, self._ts_edit_nome))
                self.conn.commit()
                self._lbl_ts_status.configure(text="✔ Atualizado!", fg=COLORS["green"])
            except Exception:
                self._lbl_ts_status.configure(text="⚠ Nome já existe.", fg=COLORS["red"]); return
            self._ts_edit_nome = None
            self._btn_salvar_ts.configure(text="  Salvar  ")
        else:
            try:
                self.conn.execute("INSERT INTO tipos_servico (nome) VALUES (?)", (nome,))
                self.conn.commit()
                self._lbl_ts_status.configure(text="✔ Adicionado!", fg=COLORS["green"])
            except Exception:
                self._lbl_ts_status.configure(text="⚠ Tipo já existe.", fg=COLORS["red"]); return
        self._limpar_form_ts(clear_status=False)
        self._atualizar_combo_tipos_servico()
        self._refresh_tabela_ts()

    def _limpar_form_ts(self, clear_status=True):
        self._ts_nome_entry.delete(0, tk.END)
        self._ts_edit_nome = None
        self._btn_salvar_ts.configure(text="  Salvar  ")
        if clear_status:
            self._lbl_ts_status.configure(text="")

    def _excluir_tipo_servico(self, nome):
        if self._tipo_servico_em_uso(nome):
            msgbox.showwarning("Em Uso", f"'{nome}' já foi usado em serviços e não pode ser excluído.")
            return
        if not msgbox.askyesno("Confirmar", f"Excluir tipo '{nome}'?"):
            return
        self.conn.execute("DELETE FROM tipos_servico WHERE nome=?", (nome,))
        self.conn.commit()
        self._atualizar_combo_tipos_servico()
        self._refresh_tabela_ts()

    def _atualizar_combo_tipos_servico(self):
        tipos = [r[0] for r in self.conn.execute(
            "SELECT nome FROM tipos_servico ORDER BY nome").fetchall()]
        if hasattr(self, "_serv_tipo_combo"):
            self._serv_tipo_combo["values"] = tipos

    def _refresh_tabela_ts(self):
        for w in self._ts_rows_frame.winfo_children():
            w.destroy()
        tipos = [r[0] for r in self.conn.execute(
            "SELECT nome FROM tipos_servico ORDER BY nome").fetchall()]
        self._lbl_total_ts.configure(text=f"{len(tipos)} tipo(s)")
        for i, t in enumerate(tipos):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._ts_rows_frame, bg=rb)
            row.pack(fill="x")
            tk.Label(row, text=str(i+1), font=("Helvetica", 9),
                     bg=rb, fg=COLORS["text_muted"],
                     width=4, anchor="w").pack(side="left", padx=4, pady=7)
            tk.Label(row, text=t, font=("Helvetica", 9),
                     bg=rb, fg=COLORS["text_primary"],
                     width=30, anchor="w").pack(side="left", padx=4)
            em_uso = self._tipo_servico_em_uso(t)
            tk.Label(row, text="✔" if em_uso else "—",
                     font=("Helvetica", 9),
                     bg=rb, fg=COLORS["green"] if em_uso else COLORS["text_muted"],
                     width=8, anchor="w").pack(side="left", padx=4)
            acts = tk.Frame(row, bg=rb)
            acts.pack(side="left", padx=4)
            if not em_uso:
                tk.Button(acts, text="✕", font=("Helvetica", 9),
                          bg=COLORS["red"], fg="white", relief="flat",
                          cursor="hand2", padx=4, pady=1,
                          command=lambda n=t: self._excluir_tipo_servico(n)
                          ).pack(side="left")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: SERVIÇO
    # ══════════════════════════════════════════════════════════════════════════
    def _build_servico_subs(self, parent):
        self.sub_pages["Serviço"] = {}
        for sub in SUBMENUS["Serviço"]:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Serviço"][sub] = frame
            if sub == "Ordem de Serviço":
                self._build_ordem_servico(frame)
            elif sub == "Shaken":
                self._build_shaken(frame)
            elif sub == "Custo OS":
                self._build_custo_os(frame)
            elif sub == "Custo SK":
                self._build_custo_sk(frame)
            elif sub == "Lucro OS":
                self._build_lucro_os(frame)
            elif sub == "Lucro SK":
                self._build_lucro_sk(frame)
            elif sub == "Relatório":
                self._build_relatorio_geral(frame)
        self.sub_pages["Serviço"][SUBMENUS["Serviço"][0]].place(in_=parent, x=0, y=0, relwidth=1, relheight=1)

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: CUSTO OS
    # ══════════════════════════════════════════════════════════════════════════

    def _build_custo_os(self, parent):
        """Layout idêntico à Nova OS: form com scroll à esquerda, histórico à direita."""
        import datetime
        self._cos_edit_id     = None
        self._cos_filtro_var  = tk.StringVar(value="")
        self._cos_filtro_mes  = tk.StringVar(value="Todos")
        self._cos_filtro_ano  = tk.StringVar(value="Todos")
        self._cos_filtro_tipo = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── COLUNA DIREITA: histórico (pack antes) ────────────────────────────
        hist = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        hist.pack(side="right", fill="both", expand=True, padx=(8, 0))
        tk.Frame(hist, bg=COLORS["blue"], height=4).pack(fill="x")

        hh = tk.Frame(hist, bg=COLORS["bg_card"])
        hh.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(hh, text="▦  Custos OS Registrados",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        self._cos_total_lbl = tk.Label(hh, text="0 registros",
                 font=("Helvetica", 8), bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._cos_total_lbl.pack(side="right")

        # Filtros
        fbar = tk.Frame(hist, bg=COLORS["bg_main"])
        fbar.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(fbar, textvariable=self._cos_filtro_var, font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["blue"], width=16
                 ).pack(side="left", ipady=4)
        self._cos_filtro_var.trace_add("write", lambda *_: self._refresh_cos())
        tk.Button(fbar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._cos_filtro_var.set("")
                  ).pack(side="left", padx=(3, 8), ipady=3)

        anos  = ["Todos"] + [str(y) for y in range(datetime.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        tk.Label(fbar, text="Mês:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._cos_filtro_mes, values=meses,
                     state="readonly", font=("Helvetica", 8), width=4
                     ).pack(side="left", padx=(2, 6), ipady=2)
        self._cos_filtro_mes.trace_add("write", lambda *_: self._refresh_cos())
        tk.Label(fbar, text="Ano:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._cos_filtro_ano, values=anos,
                     state="readonly", font=("Helvetica", 8), width=6
                     ).pack(side="left", padx=(2, 6), ipady=2)
        self._cos_filtro_ano.trace_add("write", lambda *_: self._refresh_cos())
        tk.Label(fbar, text="Tipo:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        self._cos_filtro_tipo_combo = ttk.Combobox(
                 fbar, textvariable=self._cos_filtro_tipo,
                 state="readonly", font=("Helvetica", 8), width=14)
        self._cos_filtro_tipo_combo.pack(side="left", padx=(2, 0), ipady=2)
        self._cos_filtro_tipo.trace_add("write", lambda *_: self._refresh_cos())

        # Cabeçalho colunas
        tk.Frame(hist, bg=COLORS["border"], height=1).pack(fill="x", padx=14)
        col_hdr = tk.Frame(hist, bg=COLORS["bg_main"])
        col_hdr.pack(fill="x", padx=14)
        for txt, w in [("#",3),("OS",9),("Carro",15),("Cliente",13),
                       ("Tipo Custo",14),("Descrição",16),("Valor",10),("Data",9),("Ações",8)]:
            tk.Label(col_hdr, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=5)

        scroll_f = tk.Frame(hist, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        cos_cv = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        cos_sb = tk.Scrollbar(scroll_f, orient="vertical", command=cos_cv.yview)
        self._cos_rows = tk.Frame(cos_cv, bg=COLORS["bg_card"])
        self._cos_rows.bind("<Configure>",
            lambda e: cos_cv.configure(scrollregion=cos_cv.bbox("all")))
        cos_cv.create_window((0, 0), window=self._cos_rows, anchor="nw")
        cos_cv.configure(yscrollcommand=cos_sb.set)
        cos_cv.pack(side="left", fill="both", expand=True)
        cos_sb.pack(side="right", fill="y")
        hist.bind("<Enter>", lambda e: hist.bind_all("<MouseWheel>",
            lambda ev: cos_cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
        hist.bind("<Leave>", lambda e: hist.unbind_all("<MouseWheel>"))

        # ── COLUNA ESQUERDA: formulário com scroll ────────────────────────────
        form_outer = tk.Frame(root, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=310)
        form_outer.pack(side="left", fill="y")
        form_outer.pack_propagate(False)

        fcanvas = tk.Canvas(form_outer, bg=COLORS["bg_card"], highlightthickness=0, width=308)
        fsb = tk.Scrollbar(form_outer, orient="vertical", command=fcanvas.yview)
        fcanvas.configure(yscrollcommand=fsb.set)
        fsb.pack(side="right", fill="y")
        fcanvas.pack(side="left", fill="both", expand=True)
        fcard = tk.Frame(fcanvas, bg=COLORS["bg_card"])
        fcanvas.create_window((0, 0), window=fcard, anchor="nw")
        fcard.bind("<Configure>",
            lambda e: fcanvas.configure(scrollregion=fcanvas.bbox("all")))
        form_outer.bind("<Enter>", lambda e: form_outer.bind_all("<MouseWheel>",
            lambda ev: fcanvas.yview_scroll(int(-1*(ev.delta/120)), "units")))
        form_outer.bind("<Leave>", lambda e: form_outer.unbind_all("<MouseWheel>"))

        tk.Frame(fcard, bg=COLORS["blue"], height=4).pack(fill="x")
        self._cos_title_lbl = tk.Label(fcard, text="⊕  Adicionar Custo OS",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._cos_title_lbl.pack(anchor="w", padx=18, pady=(14, 2))
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(4, 12))

        def lbl(txt, req=False):
            fg = COLORS["accent"] if req else COLORS["text_secondary"]
            tk.Label(fcard, text=txt + ("  ★" if req else ""),
                     font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"], fg=fg
                     ).pack(anchor="w", padx=18)

        # Seleção da OS
        lbl("Ordem de Serviço", req=True)
        self._cos_os_var = tk.StringVar(value="")
        self._cos_os_combo = ttk.Combobox(fcard, textvariable=self._cos_os_var,
                                           state="readonly", font=("Helvetica", 10), width=28)
        self._cos_os_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._cos_os_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_os())

        # Tipo de Custo
        lbl("Tipo de Custo", req=True)
        self._cos_tipo_var = tk.StringVar(value="")
        self._cos_tipo_combo = ttk.Combobox(fcard, textvariable=self._cos_tipo_var,
                                             state="readonly", font=("Helvetica", 10), width=28)
        self._cos_tipo_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._atualizar_tipos_cos()

        # Data
        lbl("Data do Custo")
        df = tk.Frame(fcard, bg=COLORS["bg_card"])
        df.pack(anchor="w", padx=18, pady=(4, 12))
        self._cos_dia, self._cos_mes, self._cos_ano = self._make_date_row(df)

        # Descrição
        lbl("Descrição (opcional)")
        self._cos_desc = tk.Entry(fcard, font=("Helvetica", 10),
                                   bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                   insertbackground=COLORS["text_primary"],
                                   relief="flat", bd=0,
                                   highlightthickness=1, highlightbackground=COLORS["border"],
                                   highlightcolor=COLORS["blue"], width=28)
        self._cos_desc.pack(padx=18, pady=(4, 12), ipady=6)

        # Valor
        lbl("Valor (¥)")
        self._cos_valor = self._make_yen_entry(fcard, width=22)
        self._cos_valor.pack(padx=18, pady=(4, 12), ipady=7)

        self._cos_status_lbl = tk.Label(fcard, text="", font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["red"])
        self._cos_status_lbl.pack(padx=18, pady=(0, 4))
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 10))

        btn_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        btn_f.pack(padx=18, pady=(0, 20), fill="x")
        self._cos_btn_salvar = tk.Button(btn_f, text="  Salvar Custo OS  ",
                 font=("Helvetica", 10, "bold"),
                 bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                 activebackground=COLORS["accent2"],
                 command=self._salvar_custo_os)
        self._cos_btn_salvar.pack(side="left", ipady=7, ipadx=4, fill="x", expand=True)
        tk.Button(btn_f, text="Limpar", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_cos
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=4)

        self._refresh_cos()
        self._atualizar_tipos_filtro_cos()

    # ── Helpers Custo OS ──────────────────────────────────────────────────────

    def _atualizar_combo_os(self):
        rows = self.conn.execute(
            "SELECT s.id, s.os_num, s.carro, cl.nome as cli_nome "
            "FROM servicos s LEFT JOIN clientes cl ON s.cliente_id=cl.id "
            "ORDER BY s.id DESC").fetchall()
        opts = []
        for r in rows:
            carro = (r["carro"] or "—").split("|")[0].strip()
            cli   = r["cli_nome"] or "—"
            opts.append(f"#{r['id']}  {r['os_num']}  —  {carro}  [{cli}]")
        self._cos_os_combo["values"] = opts if opts else ["Nenhuma OS cadastrada"]

    def _atualizar_tipos_cos(self):
        tipos = [r[0] for r in self.conn.execute(
            "SELECT nome FROM tipos_custo ORDER BY nome").fetchall()]
        if hasattr(self, "_cos_tipo_combo"):
            self._cos_tipo_combo["values"] = tipos

    def _atualizar_tipos_filtro_cos(self):
        tipos = ["Todos"] + [r[0] for r in self.conn.execute(
            "SELECT DISTINCT tipo_custo FROM custos_os ORDER BY tipo_custo").fetchall()]
        if hasattr(self, "_cos_filtro_tipo_combo"):
            self._cos_filtro_tipo_combo["values"] = tipos

    def _salvar_custo_os(self):
        os_sel = self._cos_os_var.get().strip()
        tipo   = self._cos_tipo_var.get().strip()
        desc   = self._cos_desc.get().strip()
        valor  = self._yen_raw(self._cos_valor)
        data_c = self._get_date_from_entries(self._cos_dia, self._cos_mes, self._cos_ano)

        if not os_sel or "Nenhuma" in os_sel:
            self._cos_status_lbl.configure(text="⚠ Selecione uma OS.", fg=COLORS["red"]); return
        if not tipo:
            self._cos_status_lbl.configure(text="⚠ Selecione o tipo de custo.", fg=COLORS["red"]); return
        try:
            servico_id = int(os_sel.lstrip("#").split()[0])
        except Exception:
            self._cos_status_lbl.configure(text="⚠ OS inválida.", fg=COLORS["red"]); return

        if self._cos_edit_id is not None:
            self.conn.execute(
                "UPDATE custos_os SET servico_id=?,tipo_custo=?,descricao=?,valor=?,data_custo=? WHERE id=?",
                (servico_id, tipo, desc, valor, data_c, self._cos_edit_id))
            self.conn.commit()
            self._cos_status_lbl.configure(text="✔ Custo OS atualizado!", fg=COLORS["green"])
            self._cos_edit_id = None
            self._cos_btn_salvar.configure(text="  Salvar Custo OS  ")
            self._cos_title_lbl.configure(text="⊕  Adicionar Custo OS")
        else:
            self.conn.execute(
                "INSERT INTO custos_os (servico_id,tipo_custo,descricao,valor,data_custo) VALUES (?,?,?,?,?)",
                (servico_id, tipo, desc, valor, data_c))
            self.conn.commit()
            self._cos_status_lbl.configure(text="✔ Custo OS registrado!", fg=COLORS["green"])
        self._limpar_cos(clear_status=False)
        self._atualizar_tipos_filtro_cos()
        self._refresh_cos()

    def _limpar_cos(self, clear_status=True):
        import datetime
        self._cos_os_var.set("")
        self._cos_tipo_var.set("")
        self._cos_desc.delete(0, tk.END)
        self._cos_valor.delete(0, tk.END)
        self._cos_edit_id = None
        self._cos_btn_salvar.configure(text="  Salvar Custo OS  ")
        self._cos_title_lbl.configure(text="⊕  Adicionar Custo OS")
        hoje = datetime.date.today()
        for e, val in [(self._cos_dia, str(hoje.day).zfill(2)),
                       (self._cos_mes, str(hoje.month).zfill(2)),
                       (self._cos_ano, str(hoje.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        if clear_status:
            self._cos_status_lbl.configure(text="")

    def _editar_custo_os(self, cid):
        row = self.conn.execute("SELECT * FROM custos_os WHERE id=?", (cid,)).fetchone()
        if not row: return
        c = dict(row)
        self._atualizar_combo_os()
        srv = self.conn.execute(
            "SELECT s.*, cl.nome as cli_nome FROM servicos s "
            "LEFT JOIN clientes cl ON s.cliente_id=cl.id WHERE s.id=?",
            (c["servico_id"],)).fetchone()
        if srv:
            sv = dict(srv)
            carro = (sv.get("carro") or "—").split("|")[0].strip()
            cli   = sv.get("cli_nome") or "—"
            self._cos_os_var.set(f"#{sv['id']}  {sv['os_num']}  —  {carro}  [{cli}]")
        self._cos_tipo_var.set(c["tipo_custo"])
        self._cos_desc.delete(0, tk.END)
        self._cos_desc.insert(0, c.get("descricao") or "")
        self._cos_valor.delete(0, tk.END)
        if c.get("valor"):
            try: self._cos_valor.insert(0, f"{int(c['valor']):,}")
            except: self._cos_valor.insert(0, c["valor"])
        data = c.get("data_custo") or ""
        p = data.split("/")
        import datetime; h = datetime.date.today()
        for e, val in [
            (self._cos_dia, p[0] if len(p)>0 and p[0] else str(h.day).zfill(2)),
            (self._cos_mes, p[1] if len(p)>1 and p[1] else str(h.month).zfill(2)),
            (self._cos_ano, p[2] if len(p)>2 and p[2] else str(h.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        self._cos_edit_id = cid
        self._cos_btn_salvar.configure(text="  Atualizar Custo OS  ")
        self._cos_title_lbl.configure(text="✏  Editar Custo OS")
        self._cos_status_lbl.configure(text="")

    def _excluir_custo_os(self, cid):
        if not msgbox.askyesno("Confirmar", "Deseja excluir este custo OS?"):
            return
        self.conn.execute("DELETE FROM custos_os WHERE id=?", (cid,))
        self.conn.commit()
        self._atualizar_tipos_filtro_cos()
        self._refresh_cos()

    def _refresh_cos(self):
        for w in self._cos_rows.winfo_children():
            w.destroy()

        termo = self._cos_filtro_var.get().strip().lower()
        fmes  = self._cos_filtro_mes.get()
        fano  = self._cos_filtro_ano.get()
        ftipo = self._cos_filtro_tipo.get()

        rows = [dict(r) for r in self.conn.execute(
            "SELECT co.*, s.os_num, s.carro as s_carro, s.data_servico, "
            "  cl.nome as cli_nome "
            "FROM custos_os co "
            "LEFT JOIN servicos s  ON co.servico_id=s.id "
            "LEFT JOIN clientes cl ON s.cliente_id=cl.id "
            "ORDER BY co.id DESC").fetchall()]

        def _ym(dv):
            p = (dv or "").split("/")
            return (p[2] if len(p)>2 else ""), (p[1] if len(p)>1 else "")

        filtered = []
        for r in rows:
            ano_v, mes_v = _ym(r.get("data_custo",""))
            if fmes  != "Todos" and mes_v != fmes:  continue
            if fano  != "Todos" and ano_v != fano:  continue
            if ftipo != "Todos" and r.get("tipo_custo") != ftipo: continue
            if termo:
                hay = " ".join(filter(None, [
                    f"#{r.get('id','')}",
                    r.get("os_num",""), r.get("s_carro",""),
                    r.get("cli_nome",""), r.get("tipo_custo",""),
                    r.get("descricao","")])).lower()
                if termo not in hay: continue
            filtered.append(r)

        self._cos_total_lbl.configure(text=f"{len(filtered)} registro(s)")
        if not filtered:
            tk.Label(self._cos_rows, text="Nenhum custo OS registrado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        for i, c in enumerate(filtered):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._cos_rows, bg=rb)
            row.pack(fill="x")
            carro_txt = (c.get("s_carro") or "—").split("|")[0].strip()
            cli_txt   = (c.get("cli_nome") or "—")
            for txt, w, fg in [
                (str(c["id"]),              3,  COLORS["text_muted"]),
                (c.get("os_num") or "—",   9,  COLORS["text_secondary"]),
                (carro_txt[:14],           15, COLORS["text_primary"]),
                (cli_txt[:12],             13, COLORS["accent"]),
            ]:
                tk.Label(row, text=txt, font=("Helvetica",9), bg=rb, fg=fg,
                         width=w, anchor="w").pack(side="left", padx=2, pady=7)
            tk.Label(row, text=c.get("tipo_custo","—"),
                     font=("Helvetica",8,"bold"),
                     bg=COLORS["blue"], fg="white",
                     width=14, anchor="center").pack(side="left", padx=2, pady=4)
            tk.Label(row, text=(c.get("descricao") or "—")[:18],
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=16, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(c.get("valor")),
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_primary"],
                     width=10, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=c.get("data_custo") or "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=9, anchor="w").pack(side="left", padx=2)
            acts = tk.Frame(row, bg=rb)
            acts.pack(side="left", padx=4)
            tk.Button(acts, text="✏", font=("Helvetica",9),
                      bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                      padx=5, pady=2,
                      command=lambda cid=c["id"]: self._editar_custo_os(cid)
                      ).pack(side="left", padx=(0,3))
            tk.Button(acts, text="✕", font=("Helvetica",9),
                      bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                      padx=5, pady=2,
                      command=lambda cid=c["id"]: self._excluir_custo_os(cid)
                      ).pack(side="left")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: CUSTO SK
    # ══════════════════════════════════════════════════════════════════════════

    def _build_custo_sk(self, parent):
        """Layout idêntico à Nova OS: form com scroll à esquerda, histórico à direita."""
        import datetime
        self._csk_edit_id     = None
        self._csk_filtro_var  = tk.StringVar(value="")
        self._csk_filtro_mes  = tk.StringVar(value="Todos")
        self._csk_filtro_ano  = tk.StringVar(value="Todos")
        self._csk_filtro_tipo = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── COLUNA DIREITA: histórico (pack antes) ────────────────────────────
        hist = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        hist.pack(side="right", fill="both", expand=True, padx=(8, 0))
        tk.Frame(hist, bg=COLORS["orange"], height=4).pack(fill="x")

        hh = tk.Frame(hist, bg=COLORS["bg_card"])
        hh.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(hh, text="▦  Custos SK Registrados",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        self._csk_total_lbl = tk.Label(hh, text="0 registros",
                 font=("Helvetica", 8), bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._csk_total_lbl.pack(side="right")

        # Filtros
        fbar = tk.Frame(hist, bg=COLORS["bg_main"])
        fbar.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(fbar, textvariable=self._csk_filtro_var, font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["orange"], width=16
                 ).pack(side="left", ipady=4)
        self._csk_filtro_var.trace_add("write", lambda *_: self._refresh_csk())
        tk.Button(fbar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._csk_filtro_var.set("")
                  ).pack(side="left", padx=(3, 8), ipady=3)

        anos  = ["Todos"] + [str(y) for y in range(datetime.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        tk.Label(fbar, text="Mês:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._csk_filtro_mes, values=meses,
                     state="readonly", font=("Helvetica", 8), width=4
                     ).pack(side="left", padx=(2, 6), ipady=2)
        self._csk_filtro_mes.trace_add("write", lambda *_: self._refresh_csk())
        tk.Label(fbar, text="Ano:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._csk_filtro_ano, values=anos,
                     state="readonly", font=("Helvetica", 8), width=6
                     ).pack(side="left", padx=(2, 0), ipady=2)
        self._csk_filtro_ano.trace_add("write", lambda *_: self._refresh_csk())
        tk.Label(fbar, text="Tipo:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(6,0))
        ttk.Combobox(fbar, textvariable=self._csk_filtro_tipo,
                     values=["Todos", "Serviço Shaken", "IS"],
                     state="readonly", font=("Helvetica", 8), width=14
                     ).pack(side="left", padx=(2, 0), ipady=2)
        self._csk_filtro_tipo.trace_add("write", lambda *_: self._refresh_csk())

        # Cabeçalho colunas
        tk.Frame(hist, bg=COLORS["border"], height=1).pack(fill="x", padx=14)
        col_hdr = tk.Frame(hist, bg=COLORS["bg_main"])
        col_hdr.pack(fill="x", padx=14)
        for txt, w in [("#",3),("SK",10),("Carro",15),("Cliente",13),
                       ("Tipo",14),("Descrição",16),("Valor",10),("Data",9),("Ações",8)]:
            tk.Label(col_hdr, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=5)

        scroll_f = tk.Frame(hist, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        csk_cv = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        csk_sb = tk.Scrollbar(scroll_f, orient="vertical", command=csk_cv.yview)
        self._csk_rows = tk.Frame(csk_cv, bg=COLORS["bg_card"])
        self._csk_rows.bind("<Configure>",
            lambda e: csk_cv.configure(scrollregion=csk_cv.bbox("all")))
        csk_cv.create_window((0, 0), window=self._csk_rows, anchor="nw")
        csk_cv.configure(yscrollcommand=csk_sb.set)
        csk_cv.pack(side="left", fill="both", expand=True)
        csk_sb.pack(side="right", fill="y")
        hist.bind("<Enter>", lambda e: hist.bind_all("<MouseWheel>",
            lambda ev: csk_cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
        hist.bind("<Leave>", lambda e: hist.unbind_all("<MouseWheel>"))

        # ── COLUNA ESQUERDA: formulário com scroll ────────────────────────────
        form_outer = tk.Frame(root, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=310)
        form_outer.pack(side="left", fill="y")
        form_outer.pack_propagate(False)

        fcanvas = tk.Canvas(form_outer, bg=COLORS["bg_card"], highlightthickness=0, width=308)
        fsb = tk.Scrollbar(form_outer, orient="vertical", command=fcanvas.yview)
        fcanvas.configure(yscrollcommand=fsb.set)
        fsb.pack(side="right", fill="y")
        fcanvas.pack(side="left", fill="both", expand=True)
        fcard = tk.Frame(fcanvas, bg=COLORS["bg_card"])
        fcanvas.create_window((0, 0), window=fcard, anchor="nw")
        fcard.bind("<Configure>",
            lambda e: fcanvas.configure(scrollregion=fcanvas.bbox("all")))
        form_outer.bind("<Enter>", lambda e: form_outer.bind_all("<MouseWheel>",
            lambda ev: fcanvas.yview_scroll(int(-1*(ev.delta/120)), "units")))
        form_outer.bind("<Leave>", lambda e: form_outer.unbind_all("<MouseWheel>"))

        tk.Frame(fcard, bg=COLORS["orange"], height=4).pack(fill="x")
        self._csk_title_lbl = tk.Label(fcard, text="⊕  Adicionar Custo SK",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._csk_title_lbl.pack(anchor="w", padx=18, pady=(14, 2))
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(4, 12))

        def lbl(txt, req=False):
            fg = COLORS["accent"] if req else COLORS["text_secondary"]
            tk.Label(fcard, text=txt + ("  ★" if req else ""),
                     font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"], fg=fg
                     ).pack(anchor="w", padx=18)

        # Seleção do SK
        lbl("Shaken", req=True)
        self._csk_sk_var = tk.StringVar(value="")
        self._csk_sk_combo = ttk.Combobox(fcard, textvariable=self._csk_sk_var,
                                           state="readonly", font=("Helvetica", 10), width=28)
        self._csk_sk_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._csk_sk_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_sk_custo())

        # Tipo fixo
        lbl("Tipo de Custo")
        tk.Label(fcard, text="Serviço Shaken",
             font=("Helvetica", 10, "bold"),
             bg=COLORS["bg_card"], fg=COLORS["orange"]).pack(anchor="w", padx=18, pady=(4, 12))
        self._csk_tipo_fixo = "Serviço Shaken"

        # Data
        lbl("Data do Custo")
        df = tk.Frame(fcard, bg=COLORS["bg_card"])
        df.pack(anchor="w", padx=18, pady=(4, 12))
        self._csk_dia, self._csk_mes, self._csk_ano = self._make_date_row(df)

        # Descrição
        lbl("Descrição (opcional)")
        self._csk_desc = tk.Entry(fcard, font=("Helvetica", 10),
                                   bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                   insertbackground=COLORS["text_primary"],
                                   relief="flat", bd=0,
                                   highlightthickness=1, highlightbackground=COLORS["border"],
                                   highlightcolor=COLORS["orange"], width=28)
        self._csk_desc.pack(padx=18, pady=(4, 12), ipady=6)

        # Valor
        lbl("Valor (¥)")
        self._csk_valor = self._make_yen_entry(fcard, width=22)
        self._csk_valor.pack(padx=18, pady=(4, 12), ipady=7)

        self._csk_status_lbl = tk.Label(fcard, text="", font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["red"])
        self._csk_status_lbl.pack(padx=18, pady=(0, 4))
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 10))

        btn_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        btn_f.pack(padx=18, pady=(0, 20), fill="x")
        self._csk_btn_salvar = tk.Button(btn_f, text="  Salvar Custo SK  ",
                 font=("Helvetica", 10, "bold"),
                 bg=COLORS["orange"], fg="white", relief="flat", cursor="hand2",
                 activebackground=COLORS["accent2"],
                 command=self._salvar_custo_sk)
        self._csk_btn_salvar.pack(side="left", ipady=7, ipadx=4, fill="x", expand=True)
        tk.Button(btn_f, text="Limpar", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_csk
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=4)

        self._refresh_csk()

    # ── Helpers Custo SK ──────────────────────────────────────────────────────

    def _atualizar_combo_sk_custo(self):
        rows = self.conn.execute(
            "SELECT sk.id, sk.sk_num, c.carro, cl.nome as cli_nome "
            "FROM shaken sk "
            "LEFT JOIN carros c    ON sk.carro_id=c.id "
            "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
            "ORDER BY sk.id DESC").fetchall()
        opts = []
        for r in rows:
            carro = (r["carro"] or "—").split("|")[0].strip()
            cli   = r["cli_nome"] or "—"
            opts.append(f"#{r['id']}  {r['sk_num']}  —  {carro}  [{cli}]")
        self._csk_sk_combo["values"] = opts if opts else ["Nenhum Shaken cadastrado"]

    def _salvar_custo_sk(self):
        sk_sel = self._csk_sk_var.get().strip()
        tipo   = self._csk_tipo_fixo
        desc   = self._csk_desc.get().strip()
        valor  = self._yen_raw(self._csk_valor)
        data_c = self._get_date_from_entries(self._csk_dia, self._csk_mes, self._csk_ano)

        if not sk_sel or "Nenhum" in sk_sel:
            self._csk_status_lbl.configure(text="⚠ Selecione um Shaken.", fg=COLORS["red"]); return
        if not valor:
            self._csk_status_lbl.configure(text="⚠ Informe o valor.", fg=COLORS["red"]); return
        try:
            shaken_id = int(sk_sel.lstrip("#").split()[0])
        except Exception:
            self._csk_status_lbl.configure(text="⚠ SK inválido.", fg=COLORS["red"]); return

        if self._csk_edit_id is not None:
            self.conn.execute(
                "UPDATE custos_sk SET shaken_id=?,tipo_custo=?,descricao=?,valor=?,data_custo=? WHERE id=?",
                (shaken_id, tipo, desc, valor, data_c, self._csk_edit_id))
            self.conn.commit()
            self._csk_status_lbl.configure(text="✔ Custo SK atualizado!", fg=COLORS["green"])
            self._csk_edit_id = None
            self._csk_btn_salvar.configure(text="  Salvar Custo SK  ")
            self._csk_title_lbl.configure(text="⊕  Adicionar Custo SK")
        else:
            self.conn.execute(
                "INSERT INTO custos_sk (shaken_id,tipo_custo,descricao,valor,data_custo) VALUES (?,?,?,?,?)",
                (shaken_id, tipo, desc, valor, data_c))
            self.conn.commit()
            self._csk_status_lbl.configure(text="✔ Custo SK registrado!", fg=COLORS["green"])
        self._limpar_csk(clear_status=False)
        self._refresh_csk()

    def _limpar_csk(self, clear_status=True):
        import datetime
        self._csk_sk_var.set("")
        self._csk_desc.delete(0, tk.END)
        self._csk_valor.delete(0, tk.END)
        self._csk_edit_id = None
        self._csk_btn_salvar.configure(text="  Salvar Custo SK  ")
        self._csk_title_lbl.configure(text="⊕  Adicionar Custo SK")
        hoje = datetime.date.today()
        for e, val in [(self._csk_dia, str(hoje.day).zfill(2)),
                       (self._csk_mes, str(hoje.month).zfill(2)),
                       (self._csk_ano, str(hoje.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        if clear_status:
            self._csk_status_lbl.configure(text="")

    def _editar_custo_sk(self, cid):
        row = self.conn.execute("SELECT * FROM custos_sk WHERE id=?", (cid,)).fetchone()
        if not row: return
        c = dict(row)
        self._atualizar_combo_sk_custo()
        sk = self.conn.execute(
            "SELECT sk.*, ca.carro, cl.nome as cli_nome "
            "FROM shaken sk "
            "LEFT JOIN carros ca   ON sk.carro_id=ca.id "
            "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
            "WHERE sk.id=?", (c["shaken_id"],)).fetchone()
        if sk:
            sv = dict(sk)
            carro = (sv.get("carro") or "—").split("|")[0].strip()
            cli   = sv.get("cli_nome") or "—"
            self._csk_sk_var.set(f"#{sv['id']}  {sv['sk_num']}  —  {carro}  [{cli}]")
        self._csk_desc.delete(0, tk.END)
        self._csk_desc.insert(0, c.get("descricao") or "")
        self._csk_valor.delete(0, tk.END)
        if c.get("valor"):
            try: self._csk_valor.insert(0, f"{int(c['valor']):,}")
            except: self._csk_valor.insert(0, c["valor"])
        data = c.get("data_custo") or ""
        p = data.split("/")
        import datetime; h = datetime.date.today()
        for e, val in [
            (self._csk_dia, p[0] if len(p)>0 and p[0] else str(h.day).zfill(2)),
            (self._csk_mes, p[1] if len(p)>1 and p[1] else str(h.month).zfill(2)),
            (self._csk_ano, p[2] if len(p)>2 and p[2] else str(h.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        self._csk_edit_id = cid
        self._csk_btn_salvar.configure(text="  Atualizar Custo SK  ")
        self._csk_title_lbl.configure(text="✏  Editar Custo SK")
        self._csk_status_lbl.configure(text="")

    def _excluir_custo_sk(self, cid):
        if not msgbox.askyesno("Confirmar", "Deseja excluir este custo SK?"):
            return
        self.conn.execute("DELETE FROM custos_sk WHERE id=?", (cid,))
        self.conn.commit()
        self._refresh_csk()

    def _refresh_csk(self):
        for w in self._csk_rows.winfo_children():
            w.destroy()

        termo = self._csk_filtro_var.get().strip().lower()
        fmes  = self._csk_filtro_mes.get()
        fano  = self._csk_filtro_ano.get()

        rows = [dict(r) for r in self.conn.execute(
            "SELECT co.*, sk.sk_num, c.carro as s_carro, "
            "  cl.nome as cli_nome "
            "FROM custos_sk co "
            "LEFT JOIN shaken sk   ON co.shaken_id=sk.id "
            "LEFT JOIN carros c    ON sk.carro_id=c.id "
            "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
            "ORDER BY co.id DESC").fetchall()]

        def _ym(dv):
            p = (dv or "").split("/")
            return (p[2] if len(p)>2 else ""), (p[1] if len(p)>1 else "")

        ftipo_csk = getattr(self, "_csk_filtro_tipo", None)
        ftipo_csk = ftipo_csk.get() if ftipo_csk else "Todos"

        filtered = []
        for r in rows:
            ano_v, mes_v = _ym(r.get("data_custo",""))
            if fmes != "Todos" and mes_v != fmes: continue
            if fano != "Todos" and ano_v != fano: continue
            if ftipo_csk != "Todos" and r.get("tipo_custo","") != ftipo_csk: continue
            if termo:
                hay = " ".join(filter(None, [
                    f"#{r.get('id','')}",
                    r.get("sk_num",""), r.get("s_carro",""),
                    r.get("cli_nome",""), r.get("descricao","")])).lower()
                if termo not in hay: continue
            filtered.append(r)

        self._csk_total_lbl.configure(text=f"{len(filtered)} registro(s)")
        if not filtered:
            tk.Label(self._csk_rows, text="Nenhum custo SK registrado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        for i, c in enumerate(filtered):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._csk_rows, bg=rb)
            row.pack(fill="x")
            carro_txt = (c.get("s_carro") or "—").split("|")[0].strip()
            cli_txt   = (c.get("cli_nome") or "—")
            for txt, w, fg in [
                (str(c["id"]),              3,  COLORS["text_muted"]),
                (c.get("sk_num") or "—",   10, COLORS["text_secondary"]),
                (carro_txt[:14],           15, COLORS["text_primary"]),
                (cli_txt[:12],             13, COLORS["accent"]),
            ]:
                tk.Label(row, text=txt, font=("Helvetica",9), bg=rb, fg=fg,
                         width=w, anchor="w").pack(side="left", padx=2, pady=7)
            tk.Label(row, text=c.get("tipo_custo","—"),
                     font=("Helvetica",8,"bold"),
                     bg=COLORS["orange"], fg="white",
                     width=14, anchor="center").pack(side="left", padx=2, pady=4)
            tk.Label(row, text=(c.get("descricao") or "—")[:18],
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=16, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(c.get("valor")),
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_primary"],
                     width=10, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=c.get("data_custo") or "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=9, anchor="w").pack(side="left", padx=2)
            acts = tk.Frame(row, bg=rb)
            acts.pack(side="left", padx=4)
            tk.Button(acts, text="✏", font=("Helvetica",9),
                      bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                      padx=5, pady=2,
                      command=lambda cid=c["id"]: self._editar_custo_sk(cid)
                      ).pack(side="left", padx=(0,3))
            tk.Button(acts, text="✕", font=("Helvetica",9),
                      bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                      padx=5, pady=2,
                      command=lambda cid=c["id"]: self._excluir_custo_sk(cid)
                      ).pack(side="left")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: LUCRO OS
    # ══════════════════════════════════════════════════════════════════════════

    def _build_lucro_os(self, parent):
        """Tela de análise de lucro/prejuízo por OS — espelho de Lucro de Vendas."""
        import datetime as _dt
        self._los_filtro_var = tk.StringVar(value="")
        self._los_filtro_ano = tk.StringVar(value="Todos")
        self._los_filtro_mes = tk.StringVar(value="Todos")
        self._los_filtro_nfi = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 8))
        tk.Label(hdr, text="◈  Lucro de OS",
                 font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["blue"]).pack(side="left")
        self._los_lbl_total = tk.Label(hdr, text="",
                 font=("Helvetica", 9),
                 bg=COLORS["bg_content"], fg=COLORS["text_muted"])
        self._los_lbl_total.pack(side="right")

        # ── KPIs ──────────────────────────────────────────────────────────────
        kpi_f = tk.Frame(root, bg=COLORS["bg_content"])
        kpi_f.pack(fill="x", pady=(0, 10))
        self._los_kpi = {}
        for key, label, color in [
            ("total_receita", "Total Receita",  COLORS["blue"]),
            ("total_custo",   "Custo s/IS",     COLORS["orange"]),
            ("total_is",      "IS (CNF)",        "#8E44AD"),
            ("custo_total",   "Custo Total",    COLORS["red"]),
            ("lucro_total",   "Lucro Total",    COLORS["green"]),
            ("margem_media",  "Margem Média",   COLORS["blue"]),
            ("qtd_lucro",     "Com Lucro",      COLORS["green"]),
            ("qtd_prejuizo",  "Com Prejuízo",   COLORS["red"]),
        ]:
            kf = tk.Frame(kpi_f, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            kf.pack(side="left", expand=True, fill="x", padx=(0,6))
            tk.Label(kf, text=label, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(8,0))
            lbl_k = tk.Label(kf, text="—", font=("Helvetica",10,"bold"),
                           bg=COLORS["bg_card"], fg=color)
            lbl_k.pack(pady=(2,8))
            self._los_kpi[key] = lbl_k

        # ── Filtros ────────────────────────────────────────────────────────────
        anos  = ["Todos"] + [str(y) for y in range(_dt.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        fbar = tk.Frame(root, bg=COLORS["bg_main"])
        fbar.pack(fill="x", pady=(0,6))
        tk.Label(fbar, text="🔍", font=("Helvetica",9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4))
        tk.Entry(fbar, textvariable=self._los_filtro_var, font=("Helvetica",9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["blue"], width=18
                 ).pack(side="left", ipady=4)
        self._los_filtro_var.trace_add("write", lambda *_: self._refresh_lucro_os())
        tk.Button(fbar, text="✕", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._los_filtro_var.set("")
                  ).pack(side="left", padx=(3,12), ipady=3)
        for txt, var, opts in [
            ("Mês:",  self._los_filtro_mes, meses),
            ("Ano:",  self._los_filtro_ano, anos),
        ]:
            tk.Label(fbar, text=txt, font=("Helvetica",8),
                     bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,3))
            ttk.Combobox(fbar, textvariable=var, values=opts,
                         state="readonly", font=("Helvetica",8), width=5
                         ).pack(side="left", padx=(0,8), ipady=2)
            var.trace_add("write", lambda *_: self._refresh_lucro_os())
        tk.Label(fbar, text="NFI:", font=("Helvetica",8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(4,2))
        ttk.Combobox(fbar, textvariable=self._los_filtro_nfi,
                     values=["Todos","SNF","CNF"],
                     state="readonly", font=("Helvetica",8), width=6
                     ).pack(side="left", padx=(0,8), ipady=2)
        self._los_filtro_nfi.trace_add("write", lambda *_: self._refresh_lucro_os())
        tk.Button(fbar, text="↺ Atualizar", font=("Helvetica",8,"bold"),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  padx=8, command=self._refresh_lucro_os
                  ).pack(side="right", ipady=4)

        # ── Tabela dual-canvas ─────────────────────────────────────────────────
        LOS_COLS = [
            ("Data",  80), ("OS",  60), ("Carro", 130), ("Cliente", 110),
            ("Tipo Serviço", 110), ("NFI", 50),
            ("Receita", 85), ("Custo s/IS", 85), ("IS", 65),
            ("C.Total", 85), ("Lucro/Perda", 95), ("Margem", 70),
        ]
        self._los_cols = LOS_COLS
        LOS_W = sum(c[1] for c in LOS_COLS)

        tbl_outer = tk.Frame(root, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        tbl_outer.pack(fill="both", expand=True)
        tk.Frame(tbl_outer, bg=COLORS["blue"], height=4).pack(fill="x")
        tbl_area = tk.Frame(tbl_outer, bg=COLORS["bg_card"])
        tbl_area.pack(fill="both", expand=True, padx=8, pady=(4,8))

        vsb = tk.Scrollbar(tbl_area, orient="vertical")
        hsb = tk.Scrollbar(tbl_area, orient="horizontal")
        self._los_hdr_cv  = tk.Canvas(tbl_area, bg=COLORS["bg_main"], height=28, highlightthickness=0)
        self._los_body_cv = tk.Canvas(tbl_area, bg=COLORS["bg_card"], highlightthickness=0,
                                       yscrollcommand=vsb.set)

        def _lox(*a):
            self._los_hdr_cv.xview(*a)
            self._los_body_cv.xview(*a)
        hsb.config(command=_lox)
        vsb.config(command=self._los_body_cv.yview)

        self._los_hdr_cv.grid(row=0, column=0, sticky="ew")
        self._los_body_cv.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        tbl_area.grid_rowconfigure(1, weight=1)
        tbl_area.grid_columnconfigure(0, weight=1)

        lhf = tk.Frame(self._los_hdr_cv, bg=COLORS["bg_main"])
        lhw = self._los_hdr_cv.create_window((0,0), window=lhf, anchor="nw")
        for col_name, col_w in LOS_COLS:
            f = tk.Frame(lhf, bg=COLORS["bg_main"], width=col_w, height=28)
            f.pack_propagate(False); f.pack(side="left")
            tk.Label(f, text=col_name, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w").pack(fill="both", padx=3)
        tk.Frame(lhf, bg=COLORS["border"], height=1, width=LOS_W).pack(fill="x")

        self._los_rows_frame = tk.Frame(self._los_body_cv, bg=COLORS["bg_card"])
        lbw = self._los_body_cv.create_window((0,0), window=self._los_rows_frame, anchor="nw")
        self._los_rows_frame.bind("<Configure>",
            lambda e: self._los_body_cv.configure(
                scrollregion=self._los_body_cv.bbox("all")))

        def _locr(e, _bw=lbw, _hw=lhw):
            w = max(e.width, LOS_W)
            self._los_body_cv.itemconfig(_bw, width=w)
            self._los_hdr_cv.itemconfig(_hw, width=w)
            self._los_hdr_cv.configure(scrollregion=(0,0,w,28))
        self._los_body_cv.bind("<Configure>", _locr)

        def _loxmove(first, last):
            hsb.set(first, last)
            self._los_hdr_cv.xview_moveto(first)
        self._los_body_cv.configure(xscrollcommand=_loxmove)

        tbl_outer.bind("<Enter>", lambda e: tbl_outer.bind_all(
            "<MouseWheel>", lambda ev: self._los_body_cv.yview_scroll(
                int(-1*(ev.delta/120)), "units")))
        tbl_outer.bind("<Leave>", lambda e: tbl_outer.unbind_all("<MouseWheel>"))

        self._refresh_lucro_os()

    def _refresh_lucro_os(self):
        import datetime as _dt
        for w in self._los_rows_frame.winfo_children():
            w.destroy()
        LOS_COLS = self._los_cols
        for ci, (_, pw) in enumerate(LOS_COLS):
            self._los_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        termo = self._los_filtro_var.get().strip().lower()
        fmes  = self._los_filtro_mes.get()
        fano  = self._los_filtro_ano.get()

        rows = [dict(r) for r in self.conn.execute(
            "SELECT s.*, cl.nome as cli_nome "
            "FROM servicos s LEFT JOIN clientes cl ON s.cliente_id=cl.id "
            "ORDER BY s.id DESC").fetchall()]

        def _ym(dv):
            p = (dv or "").split("/")
            return (p[2] if len(p)>2 else ""), (p[1] if len(p)>1 else "")

        fnfi_los = getattr(self, "_los_filtro_nfi", None)
        fnfi_los = fnfi_los.get() if fnfi_los else "Todos"

        filtrado = []
        for s in rows:
            ano_v, mes_v = _ym(s.get("data_servico",""))
            if fmes != "Todos" and mes_v != fmes: continue
            if fano != "Todos" and ano_v != fano: continue
            if fnfi_los != "Todos" and (s.get("nfi") or "SNF") != fnfi_los: continue
            if termo:
                hay = " ".join(filter(None,[
                    s.get("os_num",""), s.get("carro",""),
                    s.get("cli_nome",""), s.get("tipo_servico","")])).lower()
                if termo not in hay: continue
            filtrado.append(s)

        def _int(v):
            try: return int(str(v or "0").replace(",",""))
            except: return 0

        tot_rec = tot_cus = tot_luc = tot_is_kpi = 0
        cnt_l = cnt_p = 0
        margens = []

        for i, s in enumerate(filtrado):
            receita = _int(s.get("valor"))
            custo_os = 0
            is_os = 0
            try:
                row_c = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_os WHERE servico_id=?", (s["id"],)).fetchone()
                custo_os = int(row_c[0]) if row_c[0] else 0
            except Exception:
                pass
            try:
                row_is = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_os WHERE servico_id=? AND tipo_custo='IS'",
                    (s["id"],)).fetchone()
                is_os = int(row_is[0]) if row_is[0] else 0
            except Exception:
                pass
            custo_sem_is = custo_os - is_os
            lucro  = receita - custo_os
            margem = (lucro / receita * 100) if receita > 0 else 0.0
            tot_rec    += receita
            tot_cus    += custo_os - is_os  # sem IS
            tot_is_kpi += is_os
            tot_luc    += lucro
            margens.append(margem)
            if lucro >= 0: cnt_l += 1
            else:          cnt_p += 1

            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            tk.Frame(self._los_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=i*2, column=0, columnspan=len(LOS_COLS), sticky="ew")
            ri = i*2+1
            self._los_rows_frame.grid_rowconfigure(ri, minsize=34)

            def cell(ci, txt, fg=None, bg=None, bold=False, _rb=rb):
                f = tk.Frame(self._los_rows_frame, bg=bg or _rb)
                f.grid(row=ri, column=ci, sticky="nsew")
                tk.Label(f, text=str(txt), font=("Helvetica",7,"bold" if bold else "normal"),
                         bg=bg or _rb, fg=fg or COLORS["text_primary"],
                         anchor="w", padx=4).pack(fill="both", expand=True, padx=2, pady=3)

            carro_txt = (s.get("carro","—") or "—").split("|")[0].strip()
            if " — " in carro_txt: carro_txt = carro_txt.split(" — ",1)[1]

            col = 0
            cell(col, s.get("data_servico") or "—", COLORS["text_secondary"]); col+=1
            cell(col, s.get("os_num") or f"#{s['id']}", COLORS["accent"]); col+=1
            cell(col, carro_txt[:18], COLORS["text_primary"]); col+=1
            cell(col, (s.get("cli_nome") or "—")[:14], COLORS["accent"]); col+=1
            cell(col, s.get("tipo_servico") or "—", COLORS["text_secondary"]); col+=1
            # NFI badge
            nfi_os_lc = s.get("nfi") or "SNF"
            nfi_os_bg = "#8E44AD" if nfi_os_lc == "CNF" else "#7F8C8D"
            f_nfi_os = tk.Frame(self._los_rows_frame, bg=rb)
            f_nfi_os.grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(f_nfi_os, text=nfi_os_lc, font=("Helvetica",7,"bold"),
                     bg=nfi_os_bg, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            cell(col, self._fmt_yen_display(receita), COLORS["blue"]); col+=1
            cell(col, self._fmt_yen_display(custo_sem_is), COLORS["orange"]); col+=1
            is_fg = "#8E44AD" if is_os > 0 else COLORS["text_muted"]
            cell(col, self._fmt_yen_display(is_os) if is_os else "—", is_fg); col+=1
            cell(col, self._fmt_yen_display(custo_os), COLORS["red"]); col+=1

            # Lucro badge
            lucro_bg = COLORS["green"] if lucro >= 0 else COLORS["red"]
            fl = tk.Frame(self._los_rows_frame, bg=rb)
            fl.grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(fl, text=f"¥ {lucro:,}", font=("Helvetica",7,"bold"),
                     bg=lucro_bg, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)

            margem_col = COLORS["green"] if margem >= 0 else COLORS["red"]
            cell(col, f"{margem:+.1f}%", margem_col, bold=True)

        if filtrado:
            tk.Frame(self._los_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=len(filtrado)*2, column=0,
                            columnspan=len(LOS_COLS), sticky="ew")
        else:
            tk.Label(self._los_rows_frame,
                     text="Nenhuma OS encontrada com os filtros selecionados.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]
                     ).grid(row=0, column=0, columnspan=len(LOS_COLS), pady=40)

        mm = (sum(margens)/len(margens)) if margens else 0
        lc = COLORS["green"] if tot_luc >= 0 else COLORS["red"]
        self._los_kpi["total_receita"].configure(text=self._fmt_yen_display(tot_rec))
        self._los_kpi["total_custo"].configure(text=self._fmt_yen_display(tot_cus))
        self._los_kpi["total_is"].configure(text=self._fmt_yen_display(tot_is_kpi))
        self._los_kpi["custo_total"].configure(text=self._fmt_yen_display(tot_cus + tot_is_kpi))
        self._los_kpi["lucro_total"].configure(text=f"¥ {tot_luc:,}", fg=lc)
        self._los_kpi["margem_media"].configure(
            text=f"{mm:+.1f}%", fg=COLORS["green"] if mm >= 0 else COLORS["red"])
        self._los_kpi["qtd_lucro"].configure(text=str(cnt_l))
        self._los_kpi["qtd_prejuizo"].configure(
            text=str(cnt_p), fg=COLORS["red"] if cnt_p > 0 else COLORS["text_muted"])
        self._los_lbl_total.configure(text=f"{len(filtrado)} OS exibidas")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: LUCRO SK
    # ══════════════════════════════════════════════════════════════════════════

    def _build_lucro_sk(self, parent):
        """Tela de análise de lucro/prejuízo por Shaken."""
        import datetime as _dt
        self._lsk_filtro_var = tk.StringVar(value="")
        self._lsk_filtro_ano = tk.StringVar(value="Todos")
        self._lsk_filtro_mes = tk.StringVar(value="Todos")
        self._lsk_filtro_nfi = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 8))
        tk.Label(hdr, text="◈  Lucro de Shaken",
                 font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg="#D4AC0D").pack(side="left")
        self._lsk_lbl_total = tk.Label(hdr, text="",
                 font=("Helvetica", 9),
                 bg=COLORS["bg_content"], fg=COLORS["text_muted"])
        self._lsk_lbl_total.pack(side="right")

        kpi_f = tk.Frame(root, bg=COLORS["bg_content"])
        kpi_f.pack(fill="x", pady=(0, 10))
        self._lsk_kpi = {}
        for key, label, color in [
            ("total_receita", "Total Receita",  "#D4AC0D"),
            ("total_custo",   "Custo s/IS",     COLORS["orange"]),
            ("is_total",      "IS (CNF)",        "#8E44AD"),
            ("custo_total",   "Custo Total",    COLORS["red"]),
            ("lucro_total",   "Lucro Total",    COLORS["green"]),
            ("margem_media",  "Margem Média",   "#D4AC0D"),
            ("qtd_lucro",     "Com Lucro",      COLORS["green"]),
            ("qtd_prejuizo",  "Com Prejuízo",   COLORS["red"]),
        ]:
            kf = tk.Frame(kpi_f, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            kf.pack(side="left", expand=True, fill="x", padx=(0,6))
            tk.Label(kf, text=label, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(8,0))
            lbl_k = tk.Label(kf, text="—", font=("Helvetica",10,"bold"),
                           bg=COLORS["bg_card"], fg=color)
            lbl_k.pack(pady=(2,8))
            self._lsk_kpi[key] = lbl_k

        import datetime as _dt
        anos  = ["Todos"] + [str(y) for y in range(_dt.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        fbar = tk.Frame(root, bg=COLORS["bg_main"])
        fbar.pack(fill="x", pady=(0,6))
        tk.Label(fbar, text="🔍", font=("Helvetica",9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4))
        tk.Entry(fbar, textvariable=self._lsk_filtro_var, font=("Helvetica",9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor="#D4AC0D", width=18
                 ).pack(side="left", ipady=4)
        self._lsk_filtro_var.trace_add("write", lambda *_: self._refresh_lucro_sk())
        tk.Button(fbar, text="✕", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._lsk_filtro_var.set("")
                  ).pack(side="left", padx=(3,12), ipady=3)
        for txt, var, opts in [
            ("Mês:",  self._lsk_filtro_mes, meses),
            ("Ano:",  self._lsk_filtro_ano, anos),
        ]:
            tk.Label(fbar, text=txt, font=("Helvetica",8),
                     bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,3))
            ttk.Combobox(fbar, textvariable=var, values=opts,
                         state="readonly", font=("Helvetica",8), width=5
                         ).pack(side="left", padx=(0,8), ipady=2)
            var.trace_add("write", lambda *_: self._refresh_lucro_sk())
        tk.Label(fbar, text="NFI:", font=("Helvetica",8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(4,2))
        ttk.Combobox(fbar, textvariable=self._lsk_filtro_nfi,
                     values=["Todos","SNF","CNF"],
                     state="readonly", font=("Helvetica",8), width=6
                     ).pack(side="left", padx=(0,8), ipady=2)
        self._lsk_filtro_nfi.trace_add("write", lambda *_: self._refresh_lucro_sk())
        tk.Button(fbar, text="↺ Atualizar", font=("Helvetica",8,"bold"),
                  bg="#D4AC0D", fg="white", relief="flat", cursor="hand2",
                  padx=8, command=self._refresh_lucro_sk
                  ).pack(side="right", ipady=4)

        LSK_COLS = [
            ("Data",  80), ("SK",  70), ("Carro", 130), ("Cliente", 110),
            ("NFI", 50), ("Receita", 85), ("Custo s/IS", 85),
            ("IS", 65), ("C.Total", 85),
            ("Lucro/Perda", 95), ("Margem", 70),
        ]
        self._lsk_cols = LSK_COLS
        LSK_W = sum(c[1] for c in LSK_COLS)

        tbl_outer = tk.Frame(root, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        tbl_outer.pack(fill="both", expand=True)
        tk.Frame(tbl_outer, bg="#D4AC0D", height=4).pack(fill="x")
        tbl_area = tk.Frame(tbl_outer, bg=COLORS["bg_card"])
        tbl_area.pack(fill="both", expand=True, padx=8, pady=(4,8))

        vsb = tk.Scrollbar(tbl_area, orient="vertical")
        hsb = tk.Scrollbar(tbl_area, orient="horizontal")
        self._lsk_hdr_cv  = tk.Canvas(tbl_area, bg=COLORS["bg_main"], height=28, highlightthickness=0)
        self._lsk_body_cv = tk.Canvas(tbl_area, bg=COLORS["bg_card"], highlightthickness=0,
                                       yscrollcommand=vsb.set)

        def _lskx(*a):
            self._lsk_hdr_cv.xview(*a)
            self._lsk_body_cv.xview(*a)
        hsb.config(command=_lskx)
        vsb.config(command=self._lsk_body_cv.yview)

        self._lsk_hdr_cv.grid(row=0, column=0, sticky="ew")
        self._lsk_body_cv.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        tbl_area.grid_rowconfigure(1, weight=1)
        tbl_area.grid_columnconfigure(0, weight=1)

        lhf = tk.Frame(self._lsk_hdr_cv, bg=COLORS["bg_main"])
        lhw = self._lsk_hdr_cv.create_window((0,0), window=lhf, anchor="nw")
        for col_name, col_w in LSK_COLS:
            f = tk.Frame(lhf, bg=COLORS["bg_main"], width=col_w, height=28)
            f.pack_propagate(False); f.pack(side="left")
            tk.Label(f, text=col_name, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w").pack(fill="both", padx=3)
        tk.Frame(lhf, bg=COLORS["border"], height=1, width=LSK_W).pack(fill="x")

        self._lsk_rows_frame = tk.Frame(self._lsk_body_cv, bg=COLORS["bg_card"])
        lbw = self._lsk_body_cv.create_window((0,0), window=self._lsk_rows_frame, anchor="nw")
        self._lsk_rows_frame.bind("<Configure>",
            lambda e: self._lsk_body_cv.configure(
                scrollregion=self._lsk_body_cv.bbox("all")))

        def _lskcr(e, _bw=lbw, _hw=lhw):
            w = max(e.width, LSK_W)
            self._lsk_body_cv.itemconfig(_bw, width=w)
            self._lsk_hdr_cv.itemconfig(_hw, width=w)
            self._lsk_hdr_cv.configure(scrollregion=(0,0,w,28))
        self._lsk_body_cv.bind("<Configure>", _lskcr)

        def _lskmove(first, last):
            hsb.set(first, last)
            self._lsk_hdr_cv.xview_moveto(first)
        self._lsk_body_cv.configure(xscrollcommand=_lskmove)

        tbl_outer.bind("<Enter>", lambda e: tbl_outer.bind_all(
            "<MouseWheel>", lambda ev: self._lsk_body_cv.yview_scroll(
                int(-1*(ev.delta/120)), "units")))
        tbl_outer.bind("<Leave>", lambda e: tbl_outer.unbind_all("<MouseWheel>"))

        self._refresh_lucro_sk()

    def _refresh_lucro_sk(self):
        for w in self._lsk_rows_frame.winfo_children():
            w.destroy()
        LSK_COLS = self._lsk_cols
        for ci, (_, pw) in enumerate(LSK_COLS):
            self._lsk_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        termo = self._lsk_filtro_var.get().strip().lower()
        fmes  = self._lsk_filtro_mes.get()
        fano  = self._lsk_filtro_ano.get()

        rows = [dict(r) for r in self.conn.execute(
            "SELECT sk.*, c.carro as c_nome, cl.nome as cli_nome "
            "FROM shaken sk "
            "LEFT JOIN carros c  ON sk.carro_id=c.id "
            "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
            "ORDER BY sk.id DESC").fetchall()]

        def _ym(dv):
            p = (dv or "").split("/")
            return (p[2] if len(p)>2 else ""), (p[1] if len(p)>1 else "")

        def _int(v):
            try: return int(str(v or "0").replace(",","").replace("¥","").strip())
            except: return 0

        fnfi_lsk = getattr(self, "_lsk_filtro_nfi", None)
        fnfi_lsk = fnfi_lsk.get() if fnfi_lsk else "Todos"

        filtrado = []
        for sk in rows:
            if not sk.get("valor"): continue  # só SK com receita registrada
            ano_v, mes_v = _ym(sk.get("data_registro",""))
            if fmes != "Todos" and mes_v != fmes: continue
            if fano != "Todos" and ano_v != fano: continue
            if fnfi_lsk != "Todos" and (sk.get("nfi") or "SNF") != fnfi_lsk: continue
            if termo:
                hay = " ".join(filter(None,[
                    sk.get("sk_num",""), sk.get("c_nome",""),
                    sk.get("cli_nome","")])).lower()
                if termo not in hay: continue
            filtrado.append(sk)

        tot_rec = tot_cus = tot_luc = tot_is = 0
        cnt_l = cnt_p = 0
        margens = []

        for i, sk in enumerate(filtrado):
            receita = _int(sk.get("valor"))
            # Custo = custos_sk (inclui IS inserido automaticamente)
            extra_sk = 0
            is_sk_v = 0
            try:
                row_c = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_sk WHERE shaken_id=?", (sk["id"],)).fetchone()
                extra_sk = int(row_c[0]) if row_c[0] else 0
            except Exception:
                pass
            try:
                row_is = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_sk WHERE shaken_id=? AND tipo_custo='IS'",
                    (sk["id"],)).fetchone()
                is_sk_v = int(row_is[0]) if row_is[0] else 0
            except Exception:
                pass
            custo_sem_is_sk = extra_sk - is_sk_v
            custo_total = extra_sk
            lucro  = receita - custo_total
            margem = (lucro / receita * 100) if receita > 0 else 0.0
            tot_rec += receita
            tot_cus += (custo_total - is_sk_v)   # sem IS
            tot_is  += is_sk_v
            tot_luc += lucro
            margens.append(margem)
            if lucro >= 0: cnt_l += 1
            else:          cnt_p += 1

            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            tk.Frame(self._lsk_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=i*2, column=0, columnspan=len(LSK_COLS), sticky="ew")
            ri = i*2+1
            self._lsk_rows_frame.grid_rowconfigure(ri, minsize=34)

            def cell(ci, txt, fg=None, bg=None, bold=False, _rb=rb):
                f = tk.Frame(self._lsk_rows_frame, bg=bg or _rb)
                f.grid(row=ri, column=ci, sticky="nsew")
                tk.Label(f, text=str(txt), font=("Helvetica",7,"bold" if bold else "normal"),
                         bg=bg or _rb, fg=fg or COLORS["text_primary"],
                         anchor="w", padx=4).pack(fill="both", expand=True, padx=2, pady=3)

            carro_txt = (sk.get("c_nome") or "—").split("|")[0].strip()

            col = 0
            cell(col, sk.get("data_registro") or "—", COLORS["text_secondary"]); col+=1
            cell(col, sk.get("sk_num") or f"#{sk['id']}", "#D4AC0D"); col+=1
            cell(col, carro_txt[:18], COLORS["text_primary"]); col+=1
            cell(col, (sk.get("cli_nome") or "—")[:14], COLORS["accent"]); col+=1
            # NFI badge
            nfi_sk_lc = sk.get("nfi") or "SNF"
            nfi_sk_bg = "#8E44AD" if nfi_sk_lc == "CNF" else "#7F8C8D"
            f_nfi_sk = tk.Frame(self._lsk_rows_frame, bg=rb)
            f_nfi_sk.grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(f_nfi_sk, text=nfi_sk_lc, font=("Helvetica",7,"bold"),
                     bg=nfi_sk_bg, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            cell(col, self._fmt_yen_display(receita), "#D4AC0D"); col+=1
            cell(col, self._fmt_yen_display(custo_sem_is_sk), COLORS["orange"]); col+=1
            is_fg_sk = "#8E44AD" if is_sk_v > 0 else COLORS["text_muted"]
            cell(col, self._fmt_yen_display(is_sk_v) if is_sk_v else "—", is_fg_sk); col+=1
            cell(col, self._fmt_yen_display(extra_sk), COLORS["red"]); col+=1

            lucro_bg = COLORS["green"] if lucro >= 0 else COLORS["red"]
            fl = tk.Frame(self._lsk_rows_frame, bg=rb)
            fl.grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(fl, text=f"¥ {lucro:,}", font=("Helvetica",7,"bold"),
                     bg=lucro_bg, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)

            margem_col = COLORS["green"] if margem >= 0 else COLORS["red"]
            cell(col, f"{margem:+.1f}%", margem_col, bold=True)

        if filtrado:
            tk.Frame(self._lsk_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=len(filtrado)*2, column=0,
                            columnspan=len(LSK_COLS), sticky="ew")
        else:
            tk.Label(self._lsk_rows_frame,
                     text="Nenhum Shaken com receita registrada encontrado.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]
                     ).grid(row=0, column=0, columnspan=len(LSK_COLS), pady=40)

        mm = (sum(margens)/len(margens)) if margens else 0
        lc = COLORS["green"] if tot_luc >= 0 else COLORS["red"]
        self._lsk_kpi["total_receita"].configure(text=self._fmt_yen_display(tot_rec))
        self._lsk_kpi["total_custo"].configure(text=self._fmt_yen_display(tot_cus))
        self._lsk_kpi["is_total"].configure(text=self._fmt_yen_display(tot_is) if tot_is else "—")
        self._lsk_kpi["custo_total"].configure(text=self._fmt_yen_display(tot_cus + tot_is))
        self._lsk_kpi["lucro_total"].configure(text=f"¥ {tot_luc:,}", fg=lc)
        self._lsk_kpi["margem_media"].configure(
            text=f"{mm:+.1f}%", fg=COLORS["green"] if mm >= 0 else COLORS["red"])
        self._lsk_kpi["qtd_lucro"].configure(text=str(cnt_l))
        self._lsk_kpi["qtd_prejuizo"].configure(
            text=str(cnt_p), fg=COLORS["red"] if cnt_p > 0 else COLORS["text_muted"])
        self._lsk_lbl_total.configure(text=f"{len(filtrado)} SK exibidos")


    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: RELATÓRIO GERAL
    # ══════════════════════════════════════════════════════════════════════════

    def _build_relatorio_geral(self, parent):
        """Relatório consolidado: Vendas + OS + Shaken por período."""
        import datetime as _dt
        self._rel_filtro_mes = tk.StringVar(value="Todos")
        self._rel_filtro_ano = tk.StringVar(value=str(_dt.date.today().year))
        self._rel_filtro_nfi = tk.StringVar(value="Todos")
        self._rel_secao      = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=16, pady=14)

        # ── Cabeçalho ────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 10))
        tk.Label(hdr, text="📋  Relatório Geral",
                 font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="📄 PDF", font=("Helvetica", 9, "bold"),
                  bg="#C0392B", fg="white", relief="flat", cursor="hand2",
                  command=self._relatorio_exportar_pdf
                  ).pack(side="right", ipady=5, ipadx=8)
        tk.Button(hdr, text="📊 Excel", font=("Helvetica", 9, "bold"),
                  bg="#1E8449", fg="white", relief="flat", cursor="hand2",
                  command=self._relatorio_exportar_excel
                  ).pack(side="right", padx=(0, 8), ipady=5, ipadx=8)

        # ── Filtros ───────────────────────────────────────────────────────────
        fbar = tk.Frame(root, bg=COLORS["bg_main"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        fbar.pack(fill="x", pady=(0, 10))
        anos  = [str(y) for y in range(_dt.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        for lbl_txt, var, opts, w in [
            ("Mês:",   self._rel_filtro_mes, meses, 5),
            ("Ano:",   self._rel_filtro_ano, anos,  7),
            ("NFI:",   self._rel_filtro_nfi, ["Todos","SNF","CNF"], 7),
            ("Seção:", self._rel_secao,       ["Todos","Vendas","OS","Shaken"], 8),
        ]:
            tk.Label(fbar, text=lbl_txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_secondary"]
                     ).pack(side="left", padx=(10, 3), pady=8)
            ttk.Combobox(fbar, textvariable=var, values=opts,
                         state="readonly", font=("Helvetica", 8), width=w
                         ).pack(side="left", padx=(0, 6), ipady=3)
            var.trace_add("write", lambda *_: self._relatorio_refresh())
        tk.Button(fbar, text="↺ Atualizar", font=("Helvetica", 8, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._relatorio_refresh
                  ).pack(side="right", padx=10, ipady=4)

        # ── KPIs ──────────────────────────────────────────────────────────────
        kpi_f = tk.Frame(root, bg=COLORS["bg_content"])
        kpi_f.pack(fill="x", pady=(0, 10))
        self._rel_kpis = {}
        for key, label, cor in [
            ("rec_v",   "Receita Vendas",  COLORS["green"]),
            ("rec_os",  "Receita OS",      COLORS["blue"]),
            ("rec_sk",  "Receita Shaken",  "#D4AC0D"),
            ("rec_tot", "Receita Total",   COLORS["accent"]),
            ("cus_sem", "Custo s/ IS",     COLORS["orange"]),
            ("is_tot",  "IS (CNF)",        "#8E44AD"),
            ("cus_tot", "Custo Total",     COLORS["red"]),
            ("luc_tot", "Lucro Líquido",   COLORS["green"]),
        ]:
            kf = tk.Frame(kpi_f, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            kf.pack(side="left", expand=True, fill="x", padx=(0, 5))
            tk.Label(kf, text=label, font=("Helvetica", 7, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(6, 0))
            v_lbl = tk.Label(kf, text="—", font=("Helvetica", 10, "bold"),
                             bg=COLORS["bg_card"], fg=cor)
            v_lbl.pack(pady=(2, 6))
            self._rel_kpis[key] = v_lbl

        # ── Tabela ────────────────────────────────────────────────────────────
        tbl = tk.Frame(root, bg=COLORS["bg_card"],
                       highlightthickness=1, highlightbackground=COLORS["border"])
        tbl.pack(fill="both", expand=True)
        tk.Frame(tbl, bg=COLORS["accent"], height=4).pack(fill="x")

        hdr_f = tk.Frame(tbl, bg=COLORS["bg_main"])
        hdr_f.pack(fill="x", padx=12, pady=(4, 0))
        for txt, w in [("Tipo",5),("Data",9),("Ref.",12),("Cliente",14),
                        ("Carro",14),("NFI",5),("Receita",10),
                        ("Custo s/IS",10),("IS",9),("C.Total",10),("Lucro",11)]:
            tk.Label(hdr_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=5)
        tk.Frame(tbl, bg=COLORS["border"], height=1).pack(fill="x", padx=12)

        sf = tk.Frame(tbl, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        cv = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(sf, orient="vertical", command=cv.yview)
        self._rel_rows_frame = tk.Frame(cv, bg=COLORS["bg_card"])
        self._rel_rows_frame.bind("<Configure>",
            lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=self._rel_rows_frame, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        tbl.bind("<Enter>", lambda e: tbl.bind_all("<MouseWheel>",
            lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
        tbl.bind("<Leave>", lambda e: tbl.unbind_all("<MouseWheel>"))

        self._relatorio_refresh()

    def _relatorio_refresh(self):
        """Recarrega dados do relatório conforme filtros."""
        for w in self._rel_rows_frame.winfo_children():
            w.destroy()

        fmes = getattr(self, "_rel_filtro_mes", None)
        fano = getattr(self, "_rel_filtro_ano", None)
        fnfi = getattr(self, "_rel_filtro_nfi", None)
        fsec = getattr(self, "_rel_secao",      None)
        fmes = fmes.get() if fmes else "01"
        fano = fano.get() if fano else "2024"
        fnfi = fnfi.get() if fnfi else "Todos"
        fsec = fsec.get() if fsec else "Todos"

        def _int(v):
            try: return int(str(v or "0").replace(",", ""))
            except: return 0

        def _match(data_str):
            pts = (data_str or "").split("/")
            m = pts[1] if len(pts)>1 else ""
            a = pts[2] if len(pts)>2 else ""
            if fmes != "Todos" and m != fmes: return False
            if a != fano: return False
            return True

        items = []  # (tipo, data, ref, cliente, carro, nfi, receita, custo, lucro)

        # ── Vendas ────────────────────────────────────────────────────────────
        if fsec in ("Todos", "Vendas"):
            for v in [dict(r) for r in self.conn.execute(
                    "SELECT v.*, cl.nome as cli_nome FROM vendas v "
                    "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
                    "ORDER BY v.id DESC").fetchall()]:
                if not _match(v.get("data_venda","")): continue
                nfi_v = v.get("nfi") or "SNF"
                if fnfi != "Todos" and nfi_v != fnfi: continue
                rec = _int(v.get("valor_venda"))
                cus = 0
                cid = v.get("compra_id")
                if not cid and v.get("carro_id"):
                    cr = self.conn.execute(
                        "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                        (v["carro_id"],)).fetchone()
                    cid = cr[0] if cr else None
                is_v_r = 0
                if cid:
                    cpr = self.conn.execute(
                        "SELECT valor FROM compras WHERE id=?", (cid,)).fetchone()
                    cus += _int(cpr[0]) if cpr else 0
                    cus += self._get_custo_total(cid)
                    try:
                        ri_v = self.conn.execute(
                            "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                            "FROM custos WHERE compra_id=? AND tipo_custo='IS'",
                            (cid,)).fetchone()
                        is_v_r = int(ri_v[0]) if ri_v[0] else 0
                    except Exception: pass
                cus_sem_v = cus - is_v_r
                items.append(("V", v.get("data_venda",""), f"#{v['id']}",
                               (v.get("cli_nome") or "—"), (v.get("carro") or "—"),
                               nfi_v, rec, cus_sem_v, is_v_r, rec - cus))

        # ── OS ────────────────────────────────────────────────────────────────
        if fsec in ("Todos", "OS"):
            for s in [dict(r) for r in self.conn.execute(
                    "SELECT s.*, cl.nome as cli_nome FROM servicos s "
                    "LEFT JOIN clientes cl ON s.cliente_id=cl.id "
                    "ORDER BY s.id DESC").fetchall()]:
                if not _match(s.get("data_servico","")): continue
                nfi_s = s.get("nfi") or "SNF"
                if fnfi != "Todos" and nfi_s != fnfi: continue
                rec = _int(s.get("valor"))
                cus = 0
                is_os_r = 0
                try:
                    r2 = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_os WHERE servico_id=?", (s["id"],)).fetchone()
                    cus = int(r2[0]) if r2[0] else 0
                    ri_os = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_os WHERE servico_id=? AND tipo_custo='IS'",
                        (s["id"],)).fetchone()
                    is_os_r = int(ri_os[0]) if ri_os[0] else 0
                except Exception: pass
                cus_sem_os = cus - is_os_r
                items.append(("OS", s.get("data_servico",""), s.get("os_num","") or f"#{s['id']}",
                               (s.get("cli_nome") or "—"), (s.get("carro") or "—"),
                               nfi_s, rec, cus_sem_os, is_os_r, rec - cus))

        # ── Shaken ────────────────────────────────────────────────────────────
        if fsec in ("Todos", "Shaken"):
            for sk in [dict(r) for r in self.conn.execute(
                    "SELECT sk.*, c.carro as c_nome, cl.nome as cli_nome "
                    "FROM shaken sk "
                    "LEFT JOIN carros c ON sk.carro_id=c.id "
                    "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
                    "ORDER BY sk.id DESC").fetchall()]:
                if not sk.get("valor"): continue
                if not _match(sk.get("data_registro","")): continue
                nfi_sk = sk.get("nfi") or "SNF"
                if fnfi != "Todos" and nfi_sk != fnfi: continue
                rec = _int(sk.get("valor"))
                cus = 0
                try:
                    r2 = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_sk WHERE shaken_id=?", (sk["id"],)).fetchone()
                    cus = int(r2[0]) if r2[0] else 0
                except Exception: pass
                # IS from custos_sk
                is_sk_r = 0
                try:
                    ri_is2 = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_sk WHERE shaken_id=? AND tipo_custo='IS'",
                        (sk["id"],)).fetchone()
                    is_sk_r = int(ri_is2[0]) if ri_is2[0] else 0
                except Exception: pass
                cus_sem_r = cus - is_sk_r
                items.append(("SK", sk.get("data_registro",""), sk.get("sk_num","") or f"#{sk['id']}",
                               (sk.get("cli_nome") or "—"), (sk.get("c_nome") or "—"),
                               nfi_sk, rec, cus_sem_r, is_sk_r, rec - cus))

        # ── KPIs ──────────────────────────────────────────────────────────────
        self._rel_items_cache = items[:]
        tot_v   = sum(x[6] for x in items if x[0]=="V")
        tot_os  = sum(x[6] for x in items if x[0]=="OS")
        tot_sk  = sum(x[6] for x in items if x[0]=="SK")
        tot_rec = sum(x[6] for x in items)
        # Todos os items são tuple de 10: (tipo, data, ref, cli, carro, nfi, rec, cus_sem, is_val, lucro)
        tot_sem = sum(x[7] for x in items)
        tot_is  = sum(x[8] for x in items)
        tot_cus = tot_sem + tot_is
        tot_luc = sum(x[9] for x in items)

        kpis = self._rel_kpis
        kpis["rec_v"].configure(text=self._fmt_yen_display(tot_v))
        kpis["rec_os"].configure(text=self._fmt_yen_display(tot_os))
        kpis["rec_sk"].configure(text=self._fmt_yen_display(tot_sk))
        kpis["rec_tot"].configure(text=self._fmt_yen_display(tot_rec))
        kpis["cus_sem"].configure(text=self._fmt_yen_display(tot_sem))
        kpis["is_tot"].configure(text=self._fmt_yen_display(tot_is))
        kpis["cus_tot"].configure(text=self._fmt_yen_display(tot_cus))
        kpis["luc_tot"].configure(
            text=self._fmt_yen_display(tot_luc),
            fg=COLORS["green"] if tot_luc >= 0 else COLORS["red"])

        # ── Linhas ────────────────────────────────────────────────────────────
        if not items:
            tk.Label(self._rel_rows_frame,
                     text="Nenhum registro para o período selecionado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        TC = {"V": COLORS["green"], "OS": COLORS["blue"], "SK": "#D4AC0D"}
        for i, row_data in enumerate(items):
            tipo, data, ref, cli, carro, nfi_r, rec, cus_sem, is_val, luc = row_data
            rb = COLORS["bg_card"] if i%2==0 else COLORS["bg_content"]
            row = tk.Frame(self._rel_rows_frame, bg=rb)
            row.pack(fill="x")
            tc = TC.get(tipo, COLORS["text_muted"])
            tk.Label(row, text=tipo, font=("Helvetica",7,"bold"),
                     bg=tc, fg="white", width=5, anchor="center"
                     ).pack(side="left", padx=3, pady=4)
            for txt, w, fg in [
                (data,        9,  COLORS["text_secondary"]),
                (ref[:12],   12,  COLORS["accent"]),
                (cli[:14],   14,  COLORS["text_primary"]),
                (carro[:14], 14,  COLORS["text_primary"]),
            ]:
                tk.Label(row, text=txt, font=("Helvetica",9),
                         bg=rb, fg=fg, width=w, anchor="w"
                         ).pack(side="left", padx=2)
            nb = "#8E44AD" if nfi_r=="CNF" else "#7F8C8D"
            tk.Label(row, text=nfi_r, font=("Helvetica",7,"bold"),
                     bg=nb, fg="white", width=5, anchor="center"
                     ).pack(side="left", padx=2, pady=3)
            tk.Label(row, text=self._fmt_yen_display(rec),
                     font=("Helvetica",9), bg=rb, fg=COLORS["green"],
                     width=10, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(cus_sem),
                     font=("Helvetica",9), bg=rb, fg=COLORS["orange"],
                     width=10, anchor="w").pack(side="left", padx=2)
            is_fg = "#8E44AD" if is_val > 0 else COLORS["text_muted"]
            tk.Label(row, text=self._fmt_yen_display(is_val) if is_val else "—",
                     font=("Helvetica",9), bg=rb, fg=is_fg,
                     width=9, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(cus_sem + is_val),
                     font=("Helvetica",9,"bold"), bg=rb, fg=COLORS["red"],
                     width=10, anchor="w").pack(side="left", padx=2)
            lf = COLORS["green"] if luc>=0 else COLORS["red"]
            tk.Label(row, text=f"¥{luc:,}",
                     font=("Helvetica",9,"bold"), bg=rb, fg=lf,
                     width=11, anchor="w").pack(side="left", padx=2)

    def _relatorio_exportar_pdf(self):
        items = getattr(self, "_rel_items_cache", [])
        fmes = getattr(self, "_rel_filtro_mes", None); fmes = fmes.get() if fmes else ""
        fano = getattr(self, "_rel_filtro_ano", None); fano = fano.get() if fano else ""
        path = self._get_export_path(f"relatorio_{fano}_{fmes}", "pdf")
        if not path: return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Paragraph, Spacer)
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.units import cm
        except ImportError:
            msgbox.showerror("Erro","reportlab não instalado.\nExecute: pip install reportlab")
            return

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"<b>Relatório Geral — KM Cars — {fmes}/{fano}</b>", styles["Title"]),
            Spacer(1, 0.4*cm),
        ]
        # KPI summary table
        tot_v   = sum(x[6] for x in items if x[0]=="V")
        tot_os  = sum(x[6] for x in items if x[0]=="OS")
        tot_sk  = sum(x[6] for x in items if x[0]=="SK")
        tot_rec = sum(x[6] for x in items)
        tot_sem = sum(x[7] for x in items)
        tot_is  = sum(x[8] for x in items)
        tot_cus = tot_sem + tot_is
        tot_luc = sum(x[9] for x in items)
        cnt_cnf = sum(1 for x in items if x[5]=="CNF")
        kpi_data = [
            ["Rec.Vendas","Rec.OS","Rec.Shaken","Rec.Total",
             "Custo s/IS","IS (CNF)","Custo Total","Lucro","Ops CNF"],
            [f"¥{tot_v:,}", f"¥{tot_os:,}", f"¥{tot_sk:,}", f"¥{tot_rec:,}",
             f"¥{tot_sem:,}", f"¥{tot_is:,}", f"¥{tot_cus:,}",
             f"¥{tot_luc:,}", str(cnt_cnf)],
        ]
        t_kpi = Table(kpi_data, colWidths=[2.9*cm]*9)
        t_kpi.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),"#2C3E50"),
            ("TEXTCOLOR",(0,0),(-1,0),"white"),
            ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("GRID",(0,0),(-1,-1),0.4,rl_colors.grey),
            ("TOPPADDING",(0,0),(-1,-1),4),
            ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        story += [t_kpi, Spacer(1,0.4*cm)]

        headers = ["Tipo","Data","Referência","Cliente","Carro","NFI",
                   "Receita","Custo s/IS","IS","C.Total","Lucro"]
        rows = [headers]
        for tipo, data, ref, cli, carro, nfi_r, rec, cus_sem, is_v, luc in items:
            rows.append([tipo, data, ref[:14], cli[:16], carro[:16],
                         nfi_r, f"¥{rec:,}", f"¥{cus_sem:,}", f"¥{is_v:,}",
                         f"¥{cus_sem+is_v:,}", f"¥{luc:,}"])
        col_ws = [1.0*cm,2.0*cm,2.2*cm,3.0*cm,3.0*cm,1.0*cm,2.2*cm,2.2*cm,2.0*cm,2.2*cm,2.2*cm]
        t = Table(rows, colWidths=col_ws, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),"#2C3E50"),
            ("TEXTCOLOR",(0,0),(-1,0),"white"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),7),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),["#FFFFFF","#F2F3F4"]),
            ("GRID",(0,0),(-1,-1),0.3,rl_colors.grey),
            ("ALIGN",(0,0),(-1,-1),"LEFT"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),3),
            ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ]))
        story += [t, Spacer(1,0.3*cm),
                  Paragraph(f"Total: {len(items)} operações | Lucro: ¥{tot_luc:,}",
                             styles["Normal"])]
        doc.build(story)
        msgbox.showinfo("PDF Exportado", f"Arquivo salvo:\n{path}")

    def _relatorio_exportar_excel(self):
        items = getattr(self, "_rel_items_cache", [])
        fmes = getattr(self, "_rel_filtro_mes", None); fmes = fmes.get() if fmes else ""
        fano = getattr(self, "_rel_filtro_ano", None); fano = fano.get() if fano else ""
        path = self._get_export_path(f"relatorio_{fano}_{fmes}", "xlsx")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            msgbox.showerror("Erro","openpyxl não instalado.\nExecute: pip install openpyxl")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{fmes}-{fano}"
        hf = PatternFill("solid", fgColor="2C3E50")
        hfont = Font(color="FFFFFF", bold=True, size=9)
        for ci, h in enumerate(["Tipo","Data","Referência","Cliente","Carro","NFI",
                                  "Receita","Custo","Lucro"], 1):
            c = ws.cell(1, ci, h)
            c.font = hfont; c.fill = hf
            c.alignment = Alignment(horizontal="center")
        for ri_i, (tipo, data, ref, cli, carro, nfi_r, rec, cus, luc) in enumerate(items, 2):
            for ci, v in enumerate([tipo,data,ref,cli,carro,nfi_r,rec,cus,luc], 1):
                ws.cell(ri_i, ci, v)
        ri_s = len(items)+3
        ws.cell(ri_s,5,"TOTAIS").font = Font(bold=True)
        tot_luc_sum = sum(x[8] for x in items)
        ws.cell(ri_s,7,sum(x[6] for x in items)).font = Font(bold=True)
        ws.cell(ri_s,8,sum(x[7] for x in items)).font = Font(bold=True)
        ws.cell(ri_s,9,tot_luc_sum).font = Font(
            bold=True, color="1E8449" if tot_luc_sum>=0 else "C0392B")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
        wb.save(path)
        msgbox.showinfo("Excel Exportado", f"Arquivo salvo:\n{path}")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: SHAKEN
    # ══════════════════════════════════════════════════════════════════════════

    def _next_sk_num(self):
        row = self.conn.execute("SELECT MAX(id) FROM shaken").fetchone()
        nxt = (row[0] or 0) + 1
        return f"SK-{nxt:04d}"

    def _build_shaken(self, parent):
        """Tela de Shaken: sidebar de cadastro + listagem principal."""
        import datetime
        self._sk_edit_id    = None
        self._sk_carro_tipo = ""   # 'Cliente', 'Estoque', 'Daisha'

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ────────────────────────────────────────────────────────────────
        # PAINEL DIREITO — listagem
        # ────────────────────────────────────────────────────────────────
        list_panel = tk.Frame(root, bg=COLORS["bg_card"],
                              highlightthickness=1,
                              highlightbackground=COLORS["border"])
        list_panel.pack(side="right", fill="both", expand=True, padx=(10, 0))
        tk.Frame(list_panel, bg=COLORS["blue"], height=4).pack(fill="x")

        lh = tk.Frame(list_panel, bg=COLORS["bg_card"])
        lh.pack(fill="x", padx=14, pady=(10, 4))
        tk.Label(lh, text="🚗  Shaken Registrados",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(lh, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_shaken).pack(side="right", padx=(8,0))
        self._lbl_sk_total = tk.Label(lh, text="", font=("Helvetica", 8),
                                      bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_sk_total.pack(side="right")

        # Filtros
        fb = tk.Frame(list_panel, bg=COLORS["bg_main"])
        fb.pack(fill="x", padx=14, pady=(0, 4))
        self._sk_filtro      = tk.StringVar(value="")
        self._sk_filtro_tipo = tk.StringVar(value="Todos")

        tk.Label(fb, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]
                 ).pack(side="left", padx=(0, 4), pady=5)
        tk.Entry(fb, textvariable=self._sk_filtro,
                 font=("Helvetica", 9), width=16,
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["blue"]
                 ).pack(side="left", ipady=4)
        self._sk_filtro.trace_add("write", lambda *_: self._refresh_shaken())
        tk.Button(fb, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=3,
                  command=lambda: self._sk_filtro.set("")
                  ).pack(side="left", padx=(2, 8), ipady=3)

        tk.Label(fb, text="Status:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fb, textvariable=self._sk_filtro_tipo,
                     values=["Todos", "Com Shaken", "Sem Shaken",
                             "< 1 mês", "< 3 meses", "< 6 meses",
                             "6 a 12 meses", "≥ 1 ano", "Por Conta", "Renovados"],
                     state="readonly", font=("Helvetica", 8), width=12
                     ).pack(side="left", padx=(2, 6), ipady=2)
        self._sk_filtro_tipo.trace_add("write", lambda *_: self._refresh_shaken())
        # Filtro tipo de carro
        self._sk_filtro_carro_tipo = tk.StringVar(value="Todos")
        tk.Label(fb, text="Tipo:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fb, textvariable=self._sk_filtro_carro_tipo,
                     values=["Todos", "Cliente", "Estoque", "Daisha", "Inativos c/ SK"],
                     state="readonly", font=("Helvetica", 8), width=12
                     ).pack(side="left", padx=(2, 0), ipady=2)
        self._sk_filtro_carro_tipo.trace_add("write", lambda *_: self._refresh_shaken())

        # Filtro NFI
        tk.Label(fb, text="NFI:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(8,0))
        self._sk_filtro_nfi = tk.StringVar(value="Todos")
        ttk.Combobox(fb, textvariable=self._sk_filtro_nfi,
                     values=["Todos", "SNF", "CNF"],
                     state="readonly", font=("Helvetica", 8), width=7
                     ).pack(side="left", padx=(4, 0), ipady=2)
        self._sk_filtro_nfi.trace_add("write", lambda *_: self._refresh_shaken())

        # Botões exportar
        fb2 = tk.Frame(list_panel, bg=COLORS["bg_main"])
        fb2.pack(fill="x", padx=14, pady=(0, 4))
        tk.Button(fb2, text="📄 PDF", font=("Helvetica", 8, "bold"),
                  bg="#C0392B", fg="white", relief="flat", cursor="hand2",
                  command=self._exportar_shaken_pdf
                  ).pack(side="left", ipady=3, ipadx=4)
        tk.Button(fb2, text="📊 Excel", font=("Helvetica", 8, "bold"),
                  bg="#1E8449", fg="white", relief="flat", cursor="hand2",
                  command=self._exportar_shaken_excel
                  ).pack(side="left", padx=(6,0), ipady=3, ipadx=4)
        tk.Button(fb2, text="↺ Atualizar", font=("Helvetica", 9, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat",
                  cursor="hand2", command=self._refresh_shaken
                  ).pack(side="right", ipady=3, padx=(0, 4))

        # ── Tabela com cabeçalho em canvas sincronizado (pixel-perfect align) ─
        SK_COLS = [
            ("SK", 60, "w"), ("Carro", 130, "w"), ("Ano", 40, "w"),
            ("Cor", 60, "w"), ("Placa", 75, "w"), ("Chassi", 95, "w"),
            ("Tipo", 65, "c"), ("Cliente", 105, "w"),
            ("Custo SK", 80, "w"), ("Valor", 80, "w"),
            ("Data Shaken", 95, "w"), ("Status", 80, "c"),
            ("NFI", 50, "c"), ("Ações", 65, "w"),
        ]
        self._sk_cols = SK_COLS
        SK_TOTAL_W = sum(c[1] for c in SK_COLS)

        table_area = tk.Frame(list_panel, bg=COLORS["bg_card"])
        table_area.pack(fill="both", expand=True, padx=14, pady=(0, 12))

        vsb_sk = tk.Scrollbar(table_area, orient="vertical")
        hsb_sk = tk.Scrollbar(table_area, orient="horizontal")
        self._sk_hdr_canvas  = tk.Canvas(table_area, bg=COLORS["bg_main"],
                                          height=28, highlightthickness=0)
        self._sk_body_canvas = tk.Canvas(table_area, bg=COLORS["bg_card"],
                                          highlightthickness=0,
                                          yscrollcommand=vsb_sk.set)

        def _sk_sync_x(*args):
            self._sk_hdr_canvas.xview(*args)
            self._sk_body_canvas.xview(*args)
        hsb_sk.config(command=_sk_sync_x)
        vsb_sk.config(command=self._sk_body_canvas.yview)

        self._sk_hdr_canvas.grid(row=0, column=0, sticky="ew")
        self._sk_body_canvas.grid(row=1, column=0, sticky="nsew")
        vsb_sk.grid(row=1, column=1, sticky="ns")
        hsb_sk.grid(row=2, column=0, sticky="ew")
        table_area.grid_rowconfigure(1, weight=1)
        table_area.grid_columnconfigure(0, weight=1)

        sk_hdr_frame = tk.Frame(self._sk_hdr_canvas, bg=COLORS["bg_main"])
        sk_hdr_win = self._sk_hdr_canvas.create_window(
            (0, 0), window=sk_hdr_frame, anchor="nw")
        for col_name, col_w, col_a in SK_COLS:
            f = tk.Frame(sk_hdr_frame, bg=COLORS["bg_main"], width=col_w, height=28)
            f.pack_propagate(False)
            f.pack(side="left")
            tk.Label(f, text=col_name, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w" if col_a == "w" else "center"
                     ).pack(fill="both", padx=4)
        tk.Frame(sk_hdr_frame, bg=COLORS["border"], height=1,
                 width=SK_TOTAL_W).pack(fill="x")

        self._sk_rows_frame = tk.Frame(self._sk_body_canvas, bg=COLORS["bg_card"])
        sk_body_win = self._sk_body_canvas.create_window(
            (0, 0), window=self._sk_rows_frame, anchor="nw")
        self._sk_rows_frame.bind("<Configure>",
            lambda e: self._sk_body_canvas.configure(
                scrollregion=self._sk_body_canvas.bbox("all")))

        def _sk_canvas_resize(e, _bw=sk_body_win, _hw=sk_hdr_win):
            w = max(e.width, SK_TOTAL_W)
            self._sk_body_canvas.itemconfig(_bw, width=w)
            self._sk_hdr_canvas.itemconfig(_hw, width=w)
            self._sk_hdr_canvas.configure(scrollregion=(0, 0, w, 28))
        self._sk_body_canvas.bind("<Configure>", _sk_canvas_resize)

        def _sk_xmove(first, last):
            hsb_sk.set(first, last)
            self._sk_hdr_canvas.xview_moveto(first)
        self._sk_body_canvas.configure(xscrollcommand=_sk_xmove)

        def _sk_mwheel(ev):
            self._sk_body_canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        list_panel.bind("<Enter>", lambda e: list_panel.bind_all("<MouseWheel>", _sk_mwheel))
        list_panel.bind("<Leave>", lambda e: list_panel.unbind_all("<MouseWheel>"))

        # ────────────────────────────────────────────────────────────────
        # PAINEL ESQUERDO — formulário (sidebar)
        # ────────────────────────────────────────────────────────────────
        form_outer = tk.Frame(root, bg=COLORS["bg_card"],
                              highlightthickness=1,
                              highlightbackground=COLORS["border"],
                              width=295)
        form_outer.pack(side="left", fill="y")
        form_outer.pack_propagate(False)

        # Canvas scrollável para o formulário
        fscroll = tk.Canvas(form_outer, bg=COLORS["bg_card"], highlightthickness=0)
        fsb = tk.Scrollbar(form_outer, orient="vertical", command=fscroll.yview)
        fcard = tk.Frame(fscroll, bg=COLORS["bg_card"])
        fcard.bind("<Configure>",
                   lambda e: fscroll.configure(scrollregion=fscroll.bbox("all")))
        fscroll.create_window((0, 0), window=fcard, anchor="nw")
        fscroll.configure(yscrollcommand=fsb.set)
        fscroll.pack(side="left", fill="both", expand=True)
        fsb.pack(side="right", fill="y")

        # Cabeçalho form
        tk.Frame(fcard, bg=COLORS["blue"], height=4).pack(fill="x")
        self._lbl_sk_title = tk.Label(
            fcard, text="🔧  Novo Shaken",
            font=("Helvetica", 11, "bold"),
            bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._lbl_sk_title.pack(anchor="w", padx=18, pady=(14, 2))
        self._lbl_sk_num = tk.Label(
            fcard, text="",
            font=("Helvetica", 9, "bold"),
            bg=COLORS["bg_card"], fg=COLORS["blue"])
        self._lbl_sk_num.pack(anchor="w", padx=18, pady=(0, 4))
        tk.Frame(fcard, bg=COLORS["border"], height=1
                 ).pack(fill="x", padx=18, pady=(0, 10))

        def lbl(txt, req=False):
            fg = COLORS["blue"] if req else COLORS["text_secondary"]
            tk.Label(fcard, text=txt + ("  ★" if req else ""),
                     font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_card"], fg=fg).pack(anchor="w", padx=18)

        # ── Carro ──────────────────────────────────────────────────────
        lbl("Carro", req=True)
        self._sk_carro_var   = tk.StringVar(value="")
        self._sk_carro_combo = ttk.Combobox(
            fcard, textvariable=self._sk_carro_var,
            state="normal", font=("Helvetica", 9), width=30)
        self._sk_carro_combo.pack(padx=18, pady=(4, 10), ipady=4)
        self._sk_carro_combo.bind("<ButtonPress>",
                                  lambda e: self._sk_atualizar_carros())
        self._sk_carro_combo.bind("<<ComboboxSelected>>",
                                  lambda e: self._sk_on_carro_select())
        self._sk_carro_combo.bind("<KeyRelease>",
                                  lambda e: self._sk_filtrar_carros_substring())

        # ── Campos condicionais num container dinâmico ─────────────────
        # (cliente, valor, data shaken — show/hide via _sk_update_form_visibility)
        self._sk_dyn_frame = tk.Frame(fcard, bg=COLORS["bg_card"])
        self._sk_dyn_frame.pack(fill="x")

        # Widgets condicionais (criados mas não empacotados ainda)
        self._sk_cliente_var = tk.StringVar(value="")
        self._sk_cliente_lbl = tk.Label(
            fcard, text="—",
            font=("Helvetica", 9),
            bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._sk_custo_entry = self._make_yen_entry(fcard, width=26)   # Serviço Shaken (sempre visível)
        self._sk_valor_entry = self._make_yen_entry(fcard, width=26)   # Valor para cliente
        self._sk_data_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        self._sk_dia, self._sk_mes, self._sk_ano = self._make_date_row(self._sk_data_f)

        # ── Por Conta ──────────────────────────────────────────────────
        self._sk_por_conta_var = tk.IntVar(value=0)
        pc_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        pc_f.pack(anchor="w", padx=18, pady=(0, 10))
        tk.Checkbutton(
            pc_f, text="Por Conta  (sem data de vencimento)",
            variable=self._sk_por_conta_var,
            font=("Helvetica", 9),
            bg=COLORS["bg_card"], fg=COLORS["blue"],
            selectcolor=COLORS["bg_main"],
            activebackground=COLORS["bg_card"],
            command=self._sk_toggle_por_conta
        ).pack(side="left")

        # ── NFI ────────────────────────────────────────────────────────
        lbl("NFI")
        _nfi_f_sk = tk.Frame(fcard, bg=COLORS["bg_card"])
        _nfi_f_sk.pack(anchor="w", padx=18, pady=(4, 10))
        self._sk_nfi_var = tk.StringVar(value="SNF")
        for _lbl2, _val2 in [("SNF", "SNF"), ("CNF", "CNF")]:
            tk.Radiobutton(_nfi_f_sk, text=_lbl2, variable=self._sk_nfi_var, value=_val2,
                           font=("Helvetica", 9),
                           bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                           activebackground=COLORS["bg_card"],
                           selectcolor=COLORS["bg_main"]).pack(anchor="w")

        # ── Obs ────────────────────────────────────────────────────────
        lbl("Obs.")
        self._sk_obs_entry = tk.Entry(
            fcard, font=("Helvetica", 9),
            bg=COLORS["bg_main"], fg=COLORS["text_primary"],
            insertbackground=COLORS["text_primary"],
            relief="flat", bd=0,
            highlightthickness=1,
            highlightbackground=COLORS["border"],
            highlightcolor=COLORS["blue"], width=28)
        self._sk_obs_entry.pack(padx=18, pady=(4, 12), ipady=5)

        # ── Status / botões ────────────────────────────────────────────
        self._lbl_sk_status = tk.Label(
            fcard, text="", font=("Helvetica", 8),
            bg=COLORS["bg_card"], fg=COLORS["red"])
        self._lbl_sk_status.pack(padx=18, pady=(0, 4))

        tk.Frame(fcard, bg=COLORS["border"], height=1
                 ).pack(fill="x", padx=18, pady=(0, 10))
        btn_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        btn_f.pack(padx=18, pady=(0, 20), fill="x")
        self._btn_sk_salvar = tk.Button(
            btn_f, text="  Registrar Shaken  ",
            font=("Helvetica", 10, "bold"),
            bg=COLORS["blue"], fg="white",
            relief="flat", cursor="hand2",
            command=self._salvar_shaken)
        self._btn_sk_salvar.pack(
            side="left", ipady=7, ipadx=4, fill="x", expand=True)
        tk.Button(
            btn_f, text="Limpar",
            font=("Helvetica", 9),
            bg=COLORS["border"], fg=COLORS["text_secondary"],
            relief="flat", cursor="hand2",
            command=self._limpar_form_shaken
        ).pack(side="left", padx=(8, 0), ipady=7, ipadx=4)
        tk.Button(btn_f, text="↺", font=("Helvetica", 10, "bold"),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_shaken
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=6)

        # Inicializa
        self._sk_atualizar_num()
        self._sk_atualizar_carros()
        self._sk_update_form_visibility()
        self._refresh_shaken()

    # ── Helpers do formulário ─────────────────────────────────────────────────

    def _sk_atualizar_num(self):
        if hasattr(self, "_lbl_sk_num") and self._sk_edit_id is None:
            self._lbl_sk_num.configure(text=f"Próximo: {self._next_sk_num()}")

    def _sk_atualizar_carros(self):
        """Popula combo — carros sem shaken OU com shaken < 3 meses (exceto Inativo).
        Quando em modo edição (_sk_edit_id definido), inclui o carro atual."""
        import datetime
        hoje = datetime.date.today()
        carros_all = [dict(r) for r in self.conn.execute(
            "SELECT * FROM carros WHERE status != 'Inativo' "
            "ORDER BY carro").fetchall()]

        # Para cada carro, determina se deve aparecer no combo
        opts = []
        for c in carros_all:
            cid = c["id"]
            # Busca shaken mais recente deste carro (não Por Conta)
            sk = self.conn.execute(
                "SELECT data_vencimento, por_conta FROM shaken "
                "WHERE carro_id=? ORDER BY id DESC LIMIT 1", (cid,)).fetchone()
            incluir = False
            if not sk:
                incluir = True   # Sem shaken
            else:
                por_conta = sk[1]
                dv = sk[0] or ""
                if por_conta:
                    incluir = False   # Por conta → não aparece
                elif not dv:
                    incluir = True    # Sem data = sem shaken
                else:
                    try:
                        d, m, a = dv.split("/")
                        d_venc = datetime.date(int(a), int(m), int(d))
                        dias = (d_venc - hoje).days
                        if dias < 90:          # <3 meses (pode ser vencido também)
                            incluir = True
                        else:
                            incluir = False    # >3 meses → fora do combo
                    except Exception:
                        incluir = True

            # Em modo edição, sempre inclui o carro atual
            if self._sk_edit_id is not None:
                sk_atual = self.conn.execute(
                    "SELECT carro_id FROM shaken WHERE id=?",
                    (self._sk_edit_id,)).fetchone()
                if sk_atual and sk_atual[0] == cid:
                    incluir = True

            if incluir:
                st = c.get("status","")
                opts.append(
                    f"{cid} — {c['carro']} {c.get('ano') or ''}"
                    f" [{st}] | {c.get('placa') or '—'}")

        self._sk_carro_opts_all = opts[:]
        self._sk_carro_combo["values"] = (
            opts if opts else ["Nenhum carro disponível"])

    def _sk_filtrar_carros_substring(self):
        termo = self._sk_carro_var.get().lower()
        todos = getattr(self, "_sk_carro_opts_all", [])
        if not todos:
            return
        filtrado = [o for o in todos if termo in o.lower()] if termo else todos
        self._sk_carro_combo["values"] = filtrado if filtrado else todos

    def _sk_on_carro_select(self):
        """Ao selecionar carro: determina tipo, reseta cliente, atualiza form."""
        sel = self._sk_carro_var.get()
        if not sel or "—" not in sel:
            return
        try:
            cid   = int(sel.split("—")[0].strip())
            carro = dict(self.conn.execute(
                "SELECT * FROM carros WHERE id=?", (cid,)).fetchone())
        except Exception:
            return
        self._sk_carro_tipo = carro.get("status", "")

        # Sempre reseta cliente e valor ao trocar de carro
        self._sk_cliente_var.set("")
        self._sk_cliente_lbl.configure(text="—", fg=COLORS["text_muted"])
        self._sk_valor_entry.delete(0, tk.END)

        # Vincula cliente apenas se tipo=Cliente
        if self._sk_carro_tipo == "Cliente":
            cli_id = carro.get("cliente_id")
            if cli_id:
                cli = self.conn.execute(
                    "SELECT nome FROM clientes WHERE id=?",
                    (cli_id,)).fetchone()
                self._sk_cliente_lbl.configure(
                    text=f"#{cli_id} — {cli[0]}" if cli else f"#{cli_id}",
                    fg=COLORS["text_primary"])
                self._sk_cliente_var.set(str(cli_id))
            else:
                self._sk_cliente_lbl.configure(
                    text="— sem cliente vinculado",
                    fg=COLORS["text_muted"])

        self._sk_update_form_visibility()

    def _sk_toggle_por_conta(self):
        self._sk_update_form_visibility()

    def _sk_update_form_visibility(self):
        """Mostra/oculta campos conforme tipo do carro e Por Conta."""
        is_cliente = self._sk_carro_tipo == "Cliente"
        por_conta  = bool(self._sk_por_conta_var.get())

        # Recria seção dinâmica no container reservado
        for w in self._sk_dyn_frame.winfo_children():
            w.destroy()
        bg = COLORS["bg_card"]

        if is_cliente:
            # Carro Cliente → Custo SK (custo do serviço shaken prestado)
            tk.Label(self._sk_dyn_frame,
                     text="Custo SK — Serviço Shaken (¥)",
                     font=("Helvetica",9,"bold"), bg=bg,
                     fg=COLORS["orange"]).pack(anchor="w", padx=18)
            tk.Label(self._sk_dyn_frame,
                     text="Custo do serviço → Lucro Shaken",
                     font=("Helvetica",7), bg=bg,
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=18)
            self._sk_custo_entry.pack(in_=self._sk_dyn_frame,
                                       padx=18, pady=(4,10), ipady=6)
            # Cliente + Valor
            tk.Label(self._sk_dyn_frame, text="Cliente",
                     font=("Helvetica",9,"bold"), bg=bg,
                     fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
            self._sk_cliente_lbl.pack(in_=self._sk_dyn_frame,
                                       anchor="w", padx=24, pady=(2,6))
            tk.Label(self._sk_dyn_frame, text="Valor Cobrado Cliente (¥)",
                     font=("Helvetica",9,"bold"), bg=bg,
                     fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
            self._sk_valor_entry.pack(in_=self._sk_dyn_frame,
                                       padx=18, pady=(4,12), ipady=6)
        else:
            # Carro Estoque / Daisha → Serviço Shaken (custo de entrada/compra)
            tk.Label(self._sk_dyn_frame,
                     text="Serviço Shaken — Entrada/Compra (¥)",
                     font=("Helvetica",9,"bold"), bg=bg,
                     fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
            tk.Label(self._sk_dyn_frame,
                     text="Custo de compra → Lucro da Venda",
                     font=("Helvetica",7), bg=bg,
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=18)
            self._sk_custo_entry.pack(in_=self._sk_dyn_frame,
                                       padx=18, pady=(4,10), ipady=6)

        # Data Shaken — oculta quando Por Conta
        if not por_conta:
            tk.Label(self._sk_dyn_frame,
                     text="Data Shaken (Vencimento)  ★",
                     font=("Helvetica",9,"bold"), bg=bg,
                     fg=COLORS["blue"]).pack(anchor="w", padx=18, pady=(4,0))
            self._sk_data_f.pack(in_=self._sk_dyn_frame,
                                  anchor="w", padx=18, pady=(4,10))

    def _salvar_shaken(self):
        import datetime
        carro_sel = self._sk_carro_var.get().strip()
        if not carro_sel or "—" not in carro_sel:
            self._lbl_sk_status.configure(
                text="⚠ Selecione um carro.", fg=COLORS["red"])
            return
        try:
            carro_id = int(carro_sel.split("—")[0].strip())
            carro = dict(self.conn.execute(
                "SELECT * FROM carros WHERE id=?", (carro_id,)).fetchone())
        except Exception:
            self._lbl_sk_status.configure(
                text="⚠ Carro inválido.", fg=COLORS["red"])
            return

        por_conta  = int(self._sk_por_conta_var.get())
        is_cliente = self._sk_carro_tipo == "Cliente"

        # Serviço Shaken (sempre presente)
        custo_shaken = self._yen_raw(self._sk_custo_entry)

        # Cliente (apenas tipo Cliente)
        cliente_id = None
        if is_cliente:
            cv = self._sk_cliente_var.get().strip()
            try:
                cliente_id = int(cv) if cv else None
            except Exception:
                pass

        # Valor cobrado ao cliente (apenas tipo Cliente)
        valor = self._yen_raw(self._sk_valor_entry) if is_cliente else ""

        # Data vencimento (obrigatória, exceto Por Conta)
        if por_conta:
            data_venc = ""
        else:
            data_venc = self._get_date_from_entries(
                self._sk_dia, self._sk_mes, self._sk_ano)
            if not data_venc:
                self._lbl_sk_status.configure(
                    text="⚠ Informe a data de vencimento.", fg=COLORS["red"])
                return

        obs      = self._sk_obs_entry.get().strip()
        data_reg = datetime.date.today().strftime("%d/%m/%Y")

        if self._sk_edit_id is not None:
            nfi_sk = getattr(self, "_sk_nfi_var", None)
            nfi_sk = nfi_sk.get() if nfi_sk else "SNF"
            self.conn.execute(
                "UPDATE shaken SET carro_id=?,cliente_id=?,valor=?,"
                "custo_shaken=?,data_vencimento=?,por_conta=?,obs=?,nfi=? WHERE id=?",
                (carro_id, cliente_id, valor, custo_shaken,
                 data_venc, por_conta, obs, nfi_sk, self._sk_edit_id))
            self.conn.commit()
            # Atualiza IS se CNF
            if nfi_sk == "CNF" and valor:
                self._upsert_is("custos_sk", "shaken_id",
                                self._sk_edit_id, valor, data_reg)
        else:
            sk_num = self._next_sk_num()
            renovacao_de = getattr(self, "_sk_renovacao_de", None)
            nfi_sk = getattr(self, "_sk_nfi_var", None)
            nfi_sk = nfi_sk.get() if nfi_sk else "SNF"
            self.conn.execute(
                "INSERT INTO shaken "
                "(sk_num,carro_id,cliente_id,valor,custo_shaken,"
                "data_vencimento,por_conta,data_registro,obs,renovacao_de,nfi) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (sk_num, carro_id, cliente_id, valor, custo_shaken,
                 data_venc, por_conta, data_reg, obs, renovacao_de, nfi_sk))
            self._sk_renovacao_de = None
            # IS automático para CNF (Shaken) → upsert em custos_sk
            if nfi_sk == "CNF" and valor:
                sk_id_is = self.conn.execute(
                    "SELECT id FROM shaken ORDER BY id DESC LIMIT 1").fetchone()
                if sk_id_is:
                    self._upsert_is("custos_sk", "shaken_id",
                                    sk_id_is[0], valor, data_reg)

        # Salva data_shaken no cadastro do carro
        self.conn.execute(
            "UPDATE carros SET data_shaken=? WHERE id=?",
            (data_venc, carro_id))

        # Rota do custo conforme tipo do carro:
        #   Cliente      → custos_sk (custo do serviço shaken)
        #   Estoque/Daisha → custos da compra (custo de entrada)
        if custo_shaken and custo_shaken.strip():
            if is_cliente:
                # Custo SK do serviço shaken prestado
                if self._sk_edit_id is not None:
                    sk_id_for_csk = self._sk_edit_id
                else:
                    sk_id_for_csk = self.conn.execute(
                        "SELECT id FROM shaken WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                        (carro_id,)).fetchone()
                    sk_id_for_csk = sk_id_for_csk[0] if sk_id_for_csk else None
                if sk_id_for_csk:
                    self.conn.execute(
                        "INSERT OR IGNORE INTO tipos_custo (nome) VALUES ('Serviço Shaken')")
                    self.conn.execute(
                        "INSERT INTO custos_sk (shaken_id,tipo_custo,descricao,valor,data_custo) "
                        "VALUES (?,?,?,?,?)",
                        (sk_id_for_csk, "Serviço Shaken",
                         f"Custo SK serviço — {data_reg}",
                         custo_shaken, data_reg))
            else:
                # Custo de compra/entrada (Estoque ou Daisha)
                compra_row = self.conn.execute(
                    "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                    (carro_id,)).fetchone()
                if compra_row:
                    self.conn.execute(
                        "INSERT OR IGNORE INTO tipos_custo (nome) VALUES ('Serviço Shaken')")
                    self.conn.execute(
                        "INSERT INTO custos (compra_id,tipo_custo,descricao,valor,data_custo) "
                        "VALUES (?,?,?,?,?)",
                        (compra_row[0], "Serviço Shaken",
                         f"Shaken entrada — {data_reg}",
                         custo_shaken, data_reg))

        self.conn.commit()

        self._lbl_sk_status.configure(
            text="✔ Shaken registrado!", fg=COLORS["green"])
        self._limpar_form_shaken(clear_status=False)
        self._sk_atualizar_num()
        self._refresh_shaken()

    def _limpar_form_shaken(self, clear_status=True):
        import datetime
        self._sk_edit_id    = None
        self._sk_carro_tipo = ""
        self._sk_renovacao_de = None
        self._sk_carro_var.set("")
        self._sk_cliente_var.set("")
        self._sk_cliente_lbl.configure(text="—", fg=COLORS["text_muted"])
        self._sk_custo_entry.delete(0, tk.END)
        self._sk_valor_entry.delete(0, tk.END)
        self._sk_por_conta_var.set(0)
        self._sk_obs_entry.delete(0, tk.END)
        hoje = datetime.date.today()
        for e, v in [(self._sk_dia, f"{hoje.day:02d}"),
                     (self._sk_mes, f"{hoje.month:02d}"),
                     (self._sk_ano, str(hoje.year))]:
            e.delete(0, tk.END)
            e.insert(0, v)
        self._lbl_sk_title.configure(text="🔧  Novo Shaken", fg=COLORS["text_primary"])
        self._btn_sk_salvar.configure(text="  Registrar Shaken  ")
        self._sk_atualizar_num()
        self._lbl_sk_num.configure(fg=COLORS["blue"])
        self._sk_update_form_visibility()
        if clear_status:
            self._lbl_sk_status.configure(text="")

    # ── Status ────────────────────────────────────────────────────────────────

    def _sk_status_info(self, data_venc, por_conta):
        """Retorna (texto, cor) para o status do shaken."""
        import datetime
        if por_conta:
            return ("Por Conta", COLORS["blue"])
        if not data_venc:
            return ("Sem Shaken", COLORS["red"])
        try:
            d, m, a = data_venc.split("/")
            d_venc  = datetime.date(int(a), int(m), int(d))
        except Exception:
            return ("Sem Shaken", COLORS["red"])
        hoje  = datetime.date.today()
        delta = (d_venc - hoje).days
        if delta < 0:
            return ("Vencido",     COLORS["red"])
        if delta < 30:
            return ("< 1 mês",    "#E74C3C")   # vermelho vivo
        if delta < 90:
            return ("< 3 meses",  COLORS["orange"])
        if delta < 180:
            return ("< 6 meses",  "#C9A800")   # amarelo escuro
        if delta < 365:
            return ("6 a 12 meses","#2E86AB")  # azul aço
        return ("≥ 1 ano",        COLORS["green"])

    # ── Listagem ──────────────────────────────────────────────────────────────

    def _refresh_shaken(self):
        if not hasattr(self, "_sk_rows_frame"):
            return
        for w in self._sk_rows_frame.winfo_children():
            w.destroy()

        import datetime
        hoje = datetime.date.today()

        # ── Definição de colunas (sincronizada com _build_shaken) ──────────────
        SK_COLS = getattr(self, "_sk_cols", [
            ("SK",60,"w"),("Carro",130,"w"),("Ano",40,"w"),("Cor",60,"w"),
            ("Placa",75,"w"),("Chassi",95,"w"),("Tipo",65,"c"),
            ("Cliente",105,"w"),("Custo SK",80,"w"),("Valor",80,"w"),
            ("Data Shaken",95,"w"),("Status",80,"c"),("Ações",65,"w"),
        ])
        for ci, (_, pw, _a) in enumerate(SK_COLS):
            self._sk_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        # ── Filtros ─────────────────────────────────────────────────────────────
        termo      = self._sk_filtro.get().strip().lower()
        ftipo      = self._sk_filtro_tipo.get()
        fcar_tipo  = getattr(self, "_sk_filtro_carro_tipo",
                             tk.StringVar(value="Todos")).get()

        def status_match(st):
            # "Renovados" filter handled separately via is_renovado flag
            if ftipo == "Renovados":    return False  # handled by is_renovado check
            if ftipo == "Todos":        return st not in ("Por Conta",)
            if ftipo == "Com Shaken":   return st in ("< 1 mês","< 3 meses","< 6 meses","6 a 12 meses","≥ 1 ano")
            if ftipo == "Sem Shaken":   return st in ("Sem Shaken","Vencido")
            if ftipo == "< 1 mês":      return st == "< 1 mês"
            if ftipo == "< 3 meses":    return st == "< 3 meses"
            if ftipo == "< 6 meses":    return st == "< 6 meses"
            if ftipo == "6 a 12 meses": return st == "6 a 12 meses"
            if ftipo == "≥ 1 ano":      return st == "≥ 1 ano"
            if ftipo == "Por Conta":    return st == "Por Conta"
            return True

        def carro_tipo_match(c_status):
            if fcar_tipo == "Inativos c/ SK": return c_status == "Inativo"
            if c_status == "Inativo": return False  # nunca no Todos/outros
            if fcar_tipo == "Todos": return True
            return c_status == fcar_tipo

        def text_match_reg(r):
            if not termo: return True
            hay = " ".join(filter(None, [
                r.get("sk_num",""), r.get("c_nome",""),
                r.get("cli_nome",""), r.get("c_placa",""),
                r.get("c_status","")])).lower()
            return termo in hay

        def text_match_no(r):
            if not termo: return True
            hay = " ".join(filter(None, [
                r.get("carro",""), r.get("cli_nome",""),
                r.get("placa",""), r.get("status","")])).lower()
            return termo in hay

        fnfi_sk = getattr(self, "_sk_filtro_nfi", None)
        fnfi_sk = fnfi_sk.get() if fnfi_sk else "Todos"

        # ── Carros com shaken registrado (inclui Inativos, filtrado por carro_tipo_match) ──
        registrados = [dict(r) for r in self.conn.execute(
            "SELECT s.*, "
            "  c.carro  AS c_nome,  c.ano    AS c_ano, "
            "  c.cor    AS c_cor,   c.placa  AS c_placa, "
            "  c.chassi AS c_chassi,c.status AS c_status, "
            "  cl.nome  AS cli_nome "
            "FROM shaken s "
            "JOIN  carros c  ON s.carro_id  = c.id "
            "LEFT JOIN clientes cl ON s.cliente_id = cl.id "
            "ORDER BY s.id DESC").fetchall()]

        # Apenas SKs não-renovadas são "ativos" para o ids_com_sk (p/ Sem Shaken)
        ids_com_sk = {r["carro_id"] for r in registrados if not r.get("renovado")}

        # ── Carros sem shaken (exceto Inativos) — ordenados ID maior primeiro ──
        sem_sk_raw = [dict(r) for r in self.conn.execute(
            "SELECT c.*, cl.nome AS cli_nome "
            "FROM carros c "
            "LEFT JOIN clientes cl ON c.cliente_id = cl.id "
            "WHERE c.status != 'Inativo' "
            "ORDER BY c.id DESC").fetchall()
            if r["id"] not in ids_com_sk]

        # ══════════════════════════════════════════════════════════
        # Montagem dos itens com ordenação correta:
        # 1) Sem Shaken (ID DESC — mais recente em cima)
        # 2) Registrados ordenados por urgência:
        #    Vencido (mais recente = mais próximo de hoje = ASC) →
        #    <1 mês → <3 meses → <6 meses → 6a12 → ≥1ano → Por Conta
        #    Dentro de cada grupo: data de venc ASC (quem vence antes fica em cima)
        # ══════════════════════════════════════════════════════════
        items = []  # (kind, row, st_txt, st_cor)

        STATUS_PRIORITY = {
            "Vencido":      0,
            "< 1 mês":      1,
            "< 3 meses":    2,
            "< 6 meses":    3,
            "6 a 12 meses": 4,
            "≥ 1 ano":      5,
            "Por Conta":    6,
        }

        def _dv_to_date(dv):
            try:
                d, m, a = dv.split("/")
                return datetime.date(int(a), int(m), int(d))
            except Exception:
                return datetime.date(9999, 12, 31)

        def sort_key_reg(r):
            st_t, _ = self._sk_status_info(
                r.get("data_vencimento"), r.get("por_conta"))
            prio  = STATUS_PRIORITY.get(st_t, 9)
            d_obj = _dv_to_date(r.get("data_vencimento", ""))
            return (prio, d_obj)  # ASC: urgente primeiro, dentro do grupo vence mais cedo fica em cima

        # 1) Sem shaken (somente quando filtro permite) — já vem ORDER BY c.id DESC
        if ftipo in ("Todos", "Sem Shaken"):
            for r in sem_sk_raw:
                if not carro_tipo_match(r.get("status","")): continue
                if not text_match_no(r): continue
                items.append(("no_sk", r, "Sem Shaken", COLORS["red"]))

        # 2) Registrados filtrados e ordenados por urgência (data ASC dentro de cada grupo)
        reg_filtrados = []
        for r in registrados:
            is_renovado = bool(r.get("renovado"))
            # Filtro NFI
            if fnfi_sk != "Todos" and (r.get("nfi") or "SNF") != fnfi_sk:
                continue
            if ftipo == "Renovados":
                # Filtro Renovados: mostra apenas SKs marcadas como renovadas
                if not is_renovado: continue
                if not carro_tipo_match(r.get("c_status","")): continue
                if not text_match_reg(r): continue
                reg_filtrados.append((r, "Renovado", COLORS["text_muted"]))
                continue
            # Para qualquer outro filtro: exclui renovados
            if is_renovado: continue
            st, cor = self._sk_status_info(
                r.get("data_vencimento"), r.get("por_conta"))
            if not status_match(st): continue
            if not carro_tipo_match(r.get("c_status","")): continue
            if not text_match_reg(r): continue
            reg_filtrados.append((r, st, cor))

        if ftipo == "Renovados":
            reg_filtrados.sort(key=lambda x: x[0].get("id", 0), reverse=True)
        else:
            reg_filtrados.sort(key=lambda x: sort_key_reg(x[0]))

        for r, st, cor in reg_filtrados:
            items.append(("sk", r, st, cor))

        self._lbl_sk_total.configure(text=f"{len(items)} registro(s)")
        self._shaken_lista_filtrada = items

        if not items:
            tk.Label(self._sk_rows_frame,
                     text="Nenhum registro encontrado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).grid(row=0, column=0, pady=40,
                                                    columnspan=len(SK_COLS))
            return

        for i, (kind, r, st_txt, st_cor) in enumerate(items):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]

            # Separador horizontal
            tk.Frame(self._sk_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=i*2, column=0, columnspan=len(SK_COLS), sticky="ew")
            ri = i*2 + 1
            self._sk_rows_frame.grid_rowconfigure(ri, minsize=34)

            if kind == "sk":
                sk_num    = r.get("sk_num","")
                c_nome    = r.get("c_nome","")    or "—"
                c_ano     = r.get("c_ano","")     or "—"
                c_cor     = r.get("c_cor","")     or "—"
                c_placa   = r.get("c_placa","")   or "—"
                c_chas    = r.get("c_chassi","")  or "—"
                c_status  = r.get("c_status","")  or "—"
                cli_nom   = r.get("cli_nome","")  or "—"
                # Clientes: custo SK (serviço); Estoque/Daisha: serviço shaken (compra)
                if c_status == "Cliente":
                    try:
                        csk_r = self.conn.execute(
                            "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                            "FROM custos_sk WHERE shaken_id=?", (r["id"],)).fetchone()
                        csk_v = int(csk_r[0]) if csk_r[0] else 0
                        custo_txt = self._fmt_yen_display(str(csk_v)) if csk_v else "—"
                    except Exception:
                        custo_txt = "—"
                else:
                    custo_txt = (self._fmt_yen_display(r.get("custo_shaken"))
                                 if r.get("custo_shaken") else "—")
                vl_txt    = (self._fmt_yen_display(r.get("valor"))
                             if r.get("valor") else "—")
                # Data column: for renovados, show "Renov. → SK-XXXX" if successor found
                if r.get("renovado"):
                    try:
                        succ = self.conn.execute(
                            "SELECT sk_num FROM shaken WHERE renovacao_de=?",
                            (r["id"],)).fetchone()
                        d_txt = f"→ {succ[0]}" if succ else "Renovado"
                    except Exception:
                        d_txt = "Renovado"
                else:
                    d_txt = (r.get("data_vencimento","") or
                             ("Por Conta" if r.get("por_conta") else "—"))
                sk_id  = r["id"]
                car_id = r["carro_id"]
                has_cli = bool(r.get("cliente_id"))
                is_pc   = bool(r.get("por_conta"))
            else:
                sk_num    = "—"
                c_nome    = r.get("carro","")   or "—"
                c_ano     = r.get("ano","")     or "—"
                c_cor     = r.get("cor","")     or "—"
                c_placa   = r.get("placa","")   or "—"
                c_chas    = r.get("chassi","")  or "—"
                c_status  = r.get("status","")  or "—"
                cli_nom   = r.get("cli_nome","") or "—"
                custo_txt = "—"
                vl_txt    = "—"
                d_txt     = "—"
                sk_id  = None
                car_id = r["id"]
                has_cli = False
                is_pc   = False

            # Flag para inativos
            is_inativo_carro = (c_status == "Inativo")

            # Cor do badge de tipo de carro
            TIPO_C = {"Cliente": COLORS["accent"], "Estoque": COLORS["blue"],
                      "Daisha": COLORS["orange"], "Inativo": COLORS["red"]}

            col = 0
            # SK
            tk.Label(self._sk_rows_frame, text=sk_num,
                     font=("Helvetica",8,"bold"), bg=rb,
                     fg=COLORS["blue"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Carro
            disp = c_nome[:17]+"…" if len(c_nome)>18 else c_nome
            tk.Label(self._sk_rows_frame, text=disp,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_primary"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Ano
            tk.Label(self._sk_rows_frame, text=c_ano,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Cor
            tk.Label(self._sk_rows_frame, text=c_cor[:7],
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Placa
            tk.Label(self._sk_rows_frame, text=c_placa[:9],
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Chassi
            tk.Label(self._sk_rows_frame, text=c_chas[:11],
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_muted"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Tipo (badge)
            tf = tk.Frame(self._sk_rows_frame, bg=rb)
            tf.grid(row=ri, column=col, sticky="nsew"); col+=1
            tc = TIPO_C.get(c_status, COLORS["text_muted"])
            tk.Label(tf, text=c_status, font=("Helvetica",7,"bold"),
                     bg=tc, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            # Cliente
            cli_disp = cli_nom[:13]+"…" if len(cli_nom)>14 else cli_nom
            tk.Label(self._sk_rows_frame, text=cli_disp,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Custo SK
            tk.Label(self._sk_rows_frame, text=custo_txt,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["orange"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Valor
            tk.Label(self._sk_rows_frame, text=vl_txt,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["green"], anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Data Shaken — oculto para Inativos
            d_fg = st_cor if kind=="sk" else COLORS["text_muted"]
            d_txt_show = "—" if is_inativo_carro else d_txt
            tk.Label(self._sk_rows_frame, text=d_txt_show,
                     font=("Helvetica",8,"bold"), bg=rb,
                     fg=d_fg, anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Status badge — "INATIVO", "RENOVADO" ou status normal
            is_renovado_row = bool(r.get("renovado")) if kind == "sk" else False
            bf = tk.Frame(self._sk_rows_frame, bg=rb)
            bf.grid(row=ri, column=col, sticky="nsew"); col+=1
            if is_renovado_row:
                tk.Label(bf, text="RENOVADO",
                         font=("Helvetica",7,"bold"),
                         bg="#7F8C8D", fg="white", anchor="center"
                         ).pack(fill="both", expand=True, padx=3, pady=5)
            elif is_inativo_carro:
                tk.Label(bf, text="INATIVO",
                         font=("Helvetica",7,"bold"),
                         bg=COLORS["red"], fg="white", anchor="center"
                         ).pack(fill="both", expand=True, padx=3, pady=5)
            else:
                bfg = "white" if st_cor not in ("#C9A800","#D4AC0D") else "#333"
                tk.Label(bf, text=st_txt,
                         font=("Helvetica",7,"bold"),
                         bg=st_cor, fg=bfg, anchor="center"
                         ).pack(fill="both", expand=True, padx=3, pady=5)
            # Ações
            af = tk.Frame(self._sk_rows_frame, bg=rb)
            af.grid(row=ri, column=col, sticky="nsew")

            if sk_id:
                if not is_renovado_row:
                    tk.Button(af, text="✎", font=("Helvetica",8),
                              bg=COLORS["blue"], fg="white",
                              relief="flat", cursor="hand2", padx=4, pady=1,
                              command=lambda sid=sk_id: self._sk_editar(sid)
                              ).pack(side="left", padx=(3,2), pady=5)
                tk.Button(af, text="✕", font=("Helvetica",8),
                          bg=COLORS["red"], fg="white",
                          relief="flat", cursor="hand2", padx=4, pady=1,
                          command=lambda sid=sk_id: self._sk_excluir(sid)
                          ).pack(side="left", padx=(0,2), pady=5)
                if not is_renovado_row:
                    # P.Cta só para carros não-inativos
                    if not is_inativo_carro and has_cli and not is_pc and st_txt in (
                            "Vencido","< 1 mês","< 3 meses","< 6 meses","Sem Shaken"):
                        tk.Button(af, text="P.Cta", font=("Helvetica",7,"bold"),
                                  bg=COLORS["blue"], fg="white",
                                  relief="flat", cursor="hand2", padx=2, pady=1,
                                  command=lambda sid=sk_id:
                                      self._sk_marcar_por_conta(sid)
                                  ).pack(side="left", padx=(0,2), pady=5)
                    # Botão Renovar — só aparece para SK vencida ou próxima do vencimento (< 3 meses)
                    if not is_inativo_carro and st_txt in (
                            "Vencido","< 1 mês","< 3 meses"):
                        tk.Button(af, text="↺ Renovar", font=("Helvetica",7,"bold"),
                                  bg="#27AE60", fg="white",
                                  relief="flat", cursor="hand2", padx=3, pady=1,
                                  command=lambda sid=sk_id, cid=car_id:
                                      self._sk_renovar(sid, cid)
                                  ).pack(side="left", padx=(0,2), pady=5)
            elif not is_inativo_carro:
                tk.Button(af, text="+ SK", font=("Helvetica",7,"bold"),
                          bg=COLORS["blue"], fg="white",
                          relief="flat", cursor="hand2", padx=3, pady=1,
                          command=lambda cid=car_id:
                              self._sk_novo_para_carro(cid)
                          ).pack(side="left", padx=(3,0), pady=5)

        # Linha final
        tk.Frame(self._sk_rows_frame, bg=COLORS["border"], height=1
                 ).grid(row=len(items)*2, column=0,
                        columnspan=len(SK_COLS), sticky="ew")

    # ── Ações ─────────────────────────────────────────────────────────────────

    def _sk_renovar(self, sk_id_antigo, carro_id):
        """Marca SK atual como Renovada e abre formulário pré-preenchido para nova SK."""
        sk_ant = self.conn.execute(
            "SELECT * FROM shaken WHERE id=?", (sk_id_antigo,)
        ).fetchone()
        if not sk_ant:
            return
        sk_ant = dict(sk_ant)

        if not msgbox.askyesno(
            "Confirmar Renovação",
            f"Renovar {sk_ant.get('sk_num','SK')}?\n\n"
            "• O registro atual será marcado como Renovado\n"
            "• Um novo formulário será aberto pré-preenchido\n"
            "• O novo shaken ficará vinculado a este"
        ):
            return

        # Marca SK antiga como renovada
        self.conn.execute(
            "UPDATE shaken SET renovado=1 WHERE id=?", (sk_id_antigo,))
        self.conn.commit()

        # 1. Navega PRIMEIRO para Serviço → Shaken (garante widgets renderizados)
        self._show_page("Serviço")
        self._switch_sub("Serviço", "Shaken")

        # 2. Preenche o formulário depois que tk renderizou a tela
        def _preencher():
            self._sk_renovando_carro_id = carro_id
            self._limpar_form_shaken(clear_status=False)
            self._sk_edit_id      = None
            self._sk_renovacao_de = sk_id_antigo
            self._sk_renovando_carro_id = carro_id  # restaura após limpar

            # Seleciona o carro no combo (inclui mesmo se shaken recente)
            self._sk_atualizar_carros()
            for opt in self._sk_carro_combo["values"]:
                if str(carro_id) in str(opt).split("—")[0]:
                    self._sk_carro_var.set(opt)
                    break
            self._sk_on_carro_select()

            # Pré-preenche cliente
            if sk_ant.get("cliente_id"):
                self._sk_cliente_var.set(str(sk_ant["cliente_id"]))

            # Pré-preenche custo anterior como sugestão
            if sk_ant.get("custo_shaken"):
                self._sk_custo_entry.delete(0, "end")
                try:
                    v = int(str(sk_ant["custo_shaken"]).replace(",", ""))
                    self._sk_custo_entry.insert(0, f"{v:,}")
                except Exception:
                    self._sk_custo_entry.insert(0, str(sk_ant["custo_shaken"]))

            # Banner de renovação
            self._lbl_sk_title.configure(
                text=f"↺  Renovação  —  {sk_ant.get('sk_num', 'SK')}",
                fg="#27AE60")
            self._lbl_sk_num.configure(
                text="⚠  Preencha a nova data e salve",
                fg=COLORS["orange"])
            self._lbl_sk_status.configure(
                text=f"↺ Renovando {sk_ant.get('sk_num','SK')} — preencha a nova data e salve",
                fg="#27AE60")

            self._refresh_shaken()

        self.after(80, _preencher)

    def _sk_editar(self, sk_id):
        row = self.conn.execute(
            "SELECT s.*, c.status AS c_status "
            "FROM shaken s JOIN carros c ON s.carro_id=c.id "
            "WHERE s.id=?", (sk_id,)).fetchone()
        if not row:
            return
        r = dict(row)
        self._sk_edit_id    = sk_id
        self._sk_carro_tipo = r.get("c_status","")
        self._lbl_sk_title.configure(text="🔧  Editar Shaken")
        self._btn_sk_salvar.configure(text="  Salvar Alterações  ")
        self._lbl_sk_num.configure(text=f"ID: {r['sk_num']}")

        # Seleciona carro
        self._sk_atualizar_carros()
        for opt in self._sk_carro_combo["values"]:
            if opt.startswith(f"{r['carro_id']} —"):
                self._sk_carro_var.set(opt)
                break
        self._sk_on_carro_select()

        # Serviço Shaken
        self._sk_custo_entry.delete(0, tk.END)
        if r.get("custo_shaken"):
            self._sk_custo_entry.insert(0, r["custo_shaken"])

        # Valor (cobrado ao cliente)
        self._sk_valor_entry.delete(0, tk.END)
        if r.get("valor"):
            self._sk_valor_entry.insert(0, r["valor"])

        # Por Conta
        self._sk_por_conta_var.set(r.get("por_conta", 0))

        # Data
        dv = r.get("data_vencimento","")
        if dv:
            pts = dv.split("/")
            if len(pts) == 3:
                for e, v in [(self._sk_dia, pts[0]),
                             (self._sk_mes, pts[1]),
                             (self._sk_ano, pts[2])]:
                    e.delete(0, tk.END)
                    e.insert(0, v)

        # Obs
        self._sk_obs_entry.delete(0, tk.END)
        if r.get("obs"):
            self._sk_obs_entry.insert(0, r["obs"])

        self._sk_update_form_visibility()
        # NFI
        if hasattr(self, "_sk_nfi_var"):
            self._sk_nfi_var.set(r.get("nfi") or "SNF")
        self._lbl_sk_status.configure(text="")

    def _sk_marcar_por_conta(self, sk_id):
        """Confirma e marca shaken como Por Conta do cliente — sai da listagem de renovações."""
        sk_row = self.conn.execute(
            "SELECT sk.*, c.carro AS c_nome, cl.nome AS cli_nome "
            "FROM shaken sk "
            "LEFT JOIN carros c ON sk.carro_id=c.id "
            "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
            "WHERE sk.id=?", (sk_id,)).fetchone()
        if not sk_row:
            return
        sk = dict(sk_row)
        carro_txt  = sk.get("c_nome") or "—"
        cliente_txt = sk.get("cli_nome") or "—"

        dlg = tk.Toplevel(self)
        dlg.title("Por Conta do Cliente")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.lift()
        dlg.attributes("-topmost", True)
        dlg.after(100, lambda: dlg.attributes("-topmost", False))
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 420) // 2
        y = self.winfo_y() + (self.winfo_height() - 240) // 2
        dlg.geometry(f"420x240+{x}+{y}")

        tk.Frame(dlg, bg=COLORS["blue"], height=4).pack(fill="x")
        tk.Label(dlg, text="✋  Marcar como Por Conta do Cliente",
                 font=("Helvetica",11,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(pady=(14,4))
        tk.Frame(dlg, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=(0,6))

        tk.Label(dlg,
                 text=f"Carro: {carro_txt}  |  Cliente: {cliente_txt}",
                 font=("Helvetica",9),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                 justify="left").pack(anchor="w", padx=24)

        tk.Label(dlg,
                 text="Confirmar que este carro será renovado shaken por conta do cliente?",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["orange"],
                 wraplength=370,
                 justify="left").pack(anchor="w", padx=24, pady=(8, 0))

        tk.Label(dlg,
                 text="Este registro não aparecerá mais na listagem de renovações (ficará como 'Por Conta').",
                 font=("Helvetica", 8),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                 wraplength=370,
                 justify="left").pack(anchor="w", padx=24, pady=(4, 0))

        btn_f = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_f.pack(pady=(12,16))

        def confirmar():
            self.conn.execute(
                "UPDATE shaken SET por_conta=1, data_vencimento='Por Conta' WHERE id=?",
                (sk_id,))
            if sk.get("carro_id"):
                self.conn.execute(
                    "UPDATE carros SET data_shaken='Por Conta' WHERE id=?", (sk["carro_id"],))
            self.conn.commit()
            dlg.destroy()
            self._refresh_shaken()

        tk.Button(btn_f, text="  Confirmar  ", font=("Helvetica",10,"bold"),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  command=confirmar).pack(side="left", ipady=5, ipadx=6)
        tk.Button(btn_f, text="  Cancelar  ", font=("Helvetica",10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(side="left", padx=(8,0), ipady=5, ipadx=6)

    def _sk_excluir(self, sk_id):
        """Confirmação de exclusão de Shaken com digitação de 'excluir'."""
        sk_row = self.conn.execute(
            "SELECT sk.*, c.carro as c_nome FROM shaken sk "
            "LEFT JOIN carros c ON sk.carro_id=c.id "
            "WHERE sk.id=?", (sk_id,)).fetchone()
        if not sk_row:
            return
        sk = dict(sk_row)

        dlg = tk.Toplevel(self)
        dlg.title("Confirmar Exclusão de Shaken")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(False, False)
        dlg.grab_set()

        # Cabeçalho vermelho
        tk.Frame(dlg, bg=COLORS["red"], height=5).pack(fill="x")
        tk.Label(dlg, text="⚠  Excluir Shaken Registrado",
                 font=("Helvetica", 12, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["red"]).pack(pady=(14,4), padx=24)
        tk.Frame(dlg, bg=COLORS["border"], height=1).pack(fill="x", padx=24)

        # Informações do registro
        info_f = tk.Frame(dlg, bg=COLORS["bg_content"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
        info_f.pack(fill="x", padx=24, pady=10)
        tk.Frame(info_f, bg=COLORS["red"], width=4).pack(side="left", fill="y")
        det = tk.Frame(info_f, bg=COLORS["bg_content"])
        det.pack(side="left", fill="both", expand=True, padx=10, pady=8)
        tk.Label(det, text=f"SK Nº:  {sk.get('sk_num','—')}",
                 font=("Helvetica",10,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(anchor="w")
        tk.Label(det, text=f"Veículo:  {sk.get('c_nome','—')}",
                 font=("Helvetica",9),
                 bg=COLORS["bg_content"], fg=COLORS["text_secondary"]).pack(anchor="w")
        d_venc = sk.get("data_vencimento") or ("Por Conta" if sk.get("por_conta") else "—")
        tk.Label(det, text=f"Vencimento:  {d_venc}",
                 font=("Helvetica",9),
                 bg=COLORS["bg_content"], fg=COLORS["text_secondary"]).pack(anchor="w")
        vl = self._fmt_yen_display(sk.get("valor")) if sk.get("valor") else "—"
        tk.Label(det, text=f"Valor:  {vl}",
                 font=("Helvetica",9),
                 bg=COLORS["bg_content"], fg=COLORS["text_secondary"]).pack(anchor="w")

        tk.Label(dlg,
                 text="Esta ação é irreversível.\nTodos os custos SK vinculados também serão excluídos.",
                 font=("Helvetica",9), justify="center",
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(pady=(4,8), padx=24)

        tk.Label(dlg, text='Digite  excluir  para confirmar:',
                 font=("Helvetica",9,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["red"]).pack(padx=24)
        conf_var = tk.StringVar()
        conf_entry = tk.Entry(dlg, textvariable=conf_var,
                              font=("Helvetica",11),
                              bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                              insertbackground=COLORS["text_primary"],
                              relief="flat", bd=0,
                              highlightthickness=1,
                              highlightbackground=COLORS["border"],
                              highlightcolor=COLORS["red"], width=20)
        conf_entry.pack(padx=24, pady=(4,12), ipady=6)
        conf_entry.focus_set()

        lbl_err = tk.Label(dlg, text="", font=("Helvetica",8),
                           bg=COLORS["bg_card"], fg=COLORS["red"])
        lbl_err.pack()

        btn_row = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_row.pack(pady=(4,16), padx=24, fill="x")

        def confirmar():
            if conf_var.get().strip().lower() == "excluir":
                # Excluir custos_sk vinculados
                self.conn.execute("DELETE FROM custos_sk WHERE shaken_id=?", (sk_id,))
                self.conn.execute("DELETE FROM shaken WHERE id=?", (sk_id,))
                # Limpar data_shaken do carro se era o único shaken
                if sk.get("carro_id"):
                    remaining = self.conn.execute(
                        "SELECT COUNT(*) FROM shaken WHERE carro_id=?",
                        (sk["carro_id"],)).fetchone()[0]
                    if remaining == 0:
                        self.conn.execute(
                            "UPDATE carros SET data_shaken='' WHERE id=?",
                            (sk["carro_id"],))
                self.conn.commit()
                dlg.destroy()
                self._refresh_shaken()
            else:
                lbl_err.configure(text="⚠ Digite exatamente  excluir  para confirmar.")

        tk.Button(btn_row, text="  Excluir  ",
                  font=("Helvetica",10,"bold"),
                  bg=COLORS["red"], fg="white",
                  relief="flat", cursor="hand2",
                  command=confirmar
                  ).pack(side="left", ipady=6, ipadx=8)
        tk.Button(btn_row, text="  Cancelar  ",
                  font=("Helvetica",10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy
                  ).pack(side="left", padx=(12,0), ipady=6, ipadx=8)
        conf_entry.bind("<Return>", lambda e: confirmar())

        dlg.update_idletasks()
        w, h = dlg.winfo_reqwidth(), dlg.winfo_reqheight()
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"+{x}+{y}")

    def _sk_novo_para_carro(self, carro_id):
        """Abre formulário pré-selecionando o carro."""
        self._limpar_form_shaken()
        self._sk_atualizar_carros()
        for opt in self._sk_carro_combo["values"]:
            if opt.startswith(f"{carro_id} —"):
                self._sk_carro_var.set(opt)
                break
        self._sk_on_carro_select()


    def _next_os_num(self):
        row = self.conn.execute("SELECT MAX(id) FROM servicos").fetchone()
        nxt = (row[0] or 0) + 1
        return f"OS-{nxt:04d}"

    def _build_ordem_servico(self, parent):
        self._serv_edit_id   = None
        self._serv_filtro    = tk.StringVar(value="")
        self._serv_filtro_mes = tk.StringVar(value="Todos")
        self._serv_filtro_ano = tk.StringVar(value="Todos")
        self._serv_filtro_status = tk.StringVar(value="Todos")
        self.servicos_data   = [dict(r) for r in
            self.conn.execute("SELECT * FROM servicos ORDER BY id DESC").fetchall()]

        root_frame = tk.Frame(parent, bg=COLORS["bg_content"])
        root_frame.pack(fill="both", expand=True, padx=14, pady=14)

        # ── SIDEBAR: detalhes + obs da OS ────────────────────────────────────
        self._os_sidebar = tk.Frame(root_frame, bg=COLORS["bg_card"],
                                     highlightthickness=1, highlightbackground=COLORS["border"],
                                     width=240)
        self._os_sidebar.pack(side="right", fill="y", padx=(6, 0))
        self._os_sidebar.pack_propagate(False)
        tk.Frame(self._os_sidebar, bg="#F1C40F", height=4).pack(fill="x")
        self._os_sidebar_title = tk.Label(self._os_sidebar, text="📝  Obs. de OS",
                                           font=("Helvetica",10,"bold"),
                                           bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._os_sidebar_title.pack(anchor="w", padx=12, pady=(12,4))
        tk.Frame(self._os_sidebar, bg=COLORS["border"], height=1).pack(fill="x", padx=12)

        # Área detalhes
        self._os_sidebar_info = tk.Frame(self._os_sidebar, bg=COLORS["bg_card"])
        self._os_sidebar_info.pack(fill="x", padx=12, pady=8)
        tk.Label(self._os_sidebar_info, text="Clique em 📝 para\nver detalhes e\nobs. de uma OS.",
                 font=("Helvetica",9), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"], justify="left").pack(anchor="w")

        # Área obs editável
        tk.Label(self._os_sidebar, text="Observações",
                 font=("Helvetica",8,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=12)
        self._os_sidebar_txt = tk.Text(self._os_sidebar,
                                        font=("Helvetica",9),
                                        bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                        insertbackground=COLORS["text_primary"],
                                        relief="flat", bd=0,
                                        highlightthickness=1,
                                        highlightbackground=COLORS["border"],
                                        highlightcolor="#F1C40F",
                                        wrap="word", width=24, height=10,
                                        state="disabled")
        self._os_sidebar_txt.pack(padx=12, pady=(4,6), fill="both", expand=True)
        self._os_sidebar_sid = None

        btn_sav_f = tk.Frame(self._os_sidebar, bg=COLORS["bg_card"])
        btn_sav_f.pack(padx=12, pady=(0,12), fill="x")
        self._os_sidebar_save_btn = tk.Button(btn_sav_f, text="Salvar Obs",
                                               font=("Helvetica",9,"bold"),
                                               bg="#F1C40F", fg="#333333",
                                               relief="flat", cursor="hand2",
                                               command=self._os_sidebar_salvar)
        self._os_sidebar_save_btn.pack(side="left", ipady=4, ipadx=6)
        self._os_sidebar_status = tk.Label(btn_sav_f, text="",
                                            font=("Helvetica",7),
                                            bg=COLORS["bg_card"], fg=COLORS["green"])
        self._os_sidebar_status.pack(side="left", padx=(6,0))

        # ── COLUNA DIREITA: histórico ─────────────────────────────────────────
        hist_panel = tk.Frame(root_frame, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        hist_panel.pack(side="right", fill="both", expand=True, padx=(8, 0))
        tk.Frame(hist_panel, bg=COLORS["accent"], height=4).pack(fill="x")

        hh = tk.Frame(hist_panel, bg=COLORS["bg_card"])
        hh.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(hh, text="⚙  Histórico de Serviços", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hh, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_tabela_servicos).pack(side="right", padx=(8,0))
        self._lbl_total_serv = tk.Label(hh, text="0 registros", font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_total_serv.pack(side="right")

        # Filtros
        fbar = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        fbar.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(fbar, textvariable=self._serv_filtro, font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"], width=16
                 ).pack(side="left", ipady=4)
        self._serv_filtro.trace_add("write", lambda *_: self._refresh_tabela_servicos())
        tk.Button(fbar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._serv_filtro.set("")
                  ).pack(side="left", padx=(3, 8), ipady=3)

        import datetime
        anos  = ["Todos"] + [str(y) for y in range(datetime.date.today().year, 2019, -1)]
        meses = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        tk.Label(fbar, text="Mês:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._serv_filtro_mes, values=meses,
                     state="readonly", font=("Helvetica", 8), width=4
                     ).pack(side="left", padx=(2, 6), ipady=2)
        tk.Label(fbar, text="Ano:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fbar, textvariable=self._serv_filtro_ano, values=anos,
                     state="readonly", font=("Helvetica", 8), width=6
                     ).pack(side="left", padx=(2, 0), ipady=2)
        self._serv_filtro_mes.trace_add("write", lambda *_: self._refresh_tabela_servicos())
        self._serv_filtro_ano.trace_add("write", lambda *_: self._refresh_tabela_servicos())

        tk.Label(fbar, text="Status:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(8,0))
        ttk.Combobox(fbar, textvariable=self._serv_filtro_status,
                     values=["Todos", "Em Andamento", "Concluído", "Cancelado", "Aberto"],
                     state="readonly", font=("Helvetica", 8), width=11
                     ).pack(side="left", padx=(2,0), ipady=2)
        self._serv_filtro_status.trace_add("write", lambda *_: self._refresh_tabela_servicos())

        # Filtro NFI + botões exportar
        fbar_nfi_os = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        fbar_nfi_os.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(fbar_nfi_os, text="NFI:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        self._serv_filtro_nfi = tk.StringVar(value="Todos")
        ttk.Combobox(fbar_nfi_os, textvariable=self._serv_filtro_nfi,
                     values=["Todos", "SNF", "CNF"],
                     state="readonly", font=("Helvetica", 8), width=7
                     ).pack(side="left", padx=(4, 8), ipady=2)
        self._serv_filtro_nfi.trace_add("write", lambda *_: self._refresh_tabela_servicos())
        tk.Button(fbar_nfi_os, text="📄 PDF", font=("Helvetica", 8, "bold"),
                  bg="#C0392B", fg="white", relief="flat", cursor="hand2",
                  command=self._exportar_os_pdf
                  ).pack(side="right", padx=(4,0), ipady=3, ipadx=4)
        tk.Button(fbar_nfi_os, text="📊 Excel", font=("Helvetica", 8, "bold"),
                  bg="#1E8449", fg="white", relief="flat", cursor="hand2",
                  command=self._exportar_os_excel
                  ).pack(side="right", padx=(4,0), ipady=3, ipadx=4)

        # Cabeçalho tabela
        tk.Frame(hist_panel, bg=COLORS["border"], height=1).pack(fill="x", padx=14)
        col_f = tk.Frame(hist_panel, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=14)
        for txt, w in [("OS",8),("Data",9),("Carro",16),("Cliente",20),
                       ("Serviço",14),("Valor",10),("Status",13),("Ações",10)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=6)

        scroll_f = tk.Frame(hist_panel, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        cv = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(scroll_f, orient="vertical", command=cv.yview)
        sb_h = tk.Scrollbar(scroll_f, orient="horizontal", command=cv.xview)
        self._serv_rows_frame = tk.Frame(cv, bg=COLORS["bg_card"])
        self._serv_rows_frame.bind("<Configure>",
            lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=self._serv_rows_frame, anchor="nw")
        cv.configure(xscrollcommand=sb_h.set)
        sb_h.pack(side="bottom", fill="x")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        hist_panel.bind("<Enter>", lambda e: hist_panel.bind_all("<MouseWheel>",
            lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
        hist_panel.bind("<Leave>", lambda e: hist_panel.unbind_all("<MouseWheel>"))

        # ── COLUNA ESQUERDA: formulário ───────────────────────────────────────
        form_outer = tk.Frame(root_frame, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=320)
        form_outer.pack(side="left", fill="y")
        form_outer.pack_propagate(False)

        fcanvas = tk.Canvas(form_outer, bg=COLORS["bg_card"], highlightthickness=0, width=318)
        fsb = tk.Scrollbar(form_outer, orient="vertical", command=fcanvas.yview)
        fcanvas.configure(yscrollcommand=fsb.set)
        fsb.pack(side="right", fill="y")
        fcanvas.pack(side="left", fill="both", expand=True)
        fcard = tk.Frame(fcanvas, bg=COLORS["bg_card"])
        fcanvas.create_window((0, 0), window=fcard, anchor="nw")
        fcard.bind("<Configure>",
            lambda e: fcanvas.configure(scrollregion=fcanvas.bbox("all")))
        form_outer.bind("<Enter>", lambda e: form_outer.bind_all("<MouseWheel>",
            lambda ev: fcanvas.yview_scroll(int(-1*(ev.delta/120)), "units")))
        form_outer.bind("<Leave>", lambda e: form_outer.unbind_all("<MouseWheel>"))

        tk.Frame(fcard, bg=COLORS["accent"], height=4).pack(fill="x")
        self._lbl_serv_title = tk.Label(fcard, text="⚙  Nova Ordem de Serviço",
                                         font=("Helvetica", 11, "bold"),
                                         bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._lbl_serv_title.pack(anchor="w", padx=18, pady=(14, 2))

        # Número OS (gerado auto, exibido)
        self._lbl_os_num = tk.Label(fcard, text="",
                                     font=("Helvetica", 9, "bold"),
                                     bg=COLORS["bg_card"], fg=COLORS["accent"])
        self._lbl_os_num.pack(anchor="w", padx=18, pady=(0, 4))
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 12))

        def lbl(txt, required=False):
            fg = COLORS["accent"] if required else COLORS["text_secondary"]
            tk.Label(fcard, text=txt + ("  ★" if required else ""),
                     font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"], fg=fg
                     ).pack(anchor="w", padx=18)

        # Cliente (primeiro para filtrar carros)
        # ── Cliente: busca + listbox ──────────────────────────────────────
        lbl("Cliente", required=True)
        self._serv_cliente_var = tk.StringVar(value="")
        self._serv_cliente_id  = None
        _sv_cli_filt = tk.StringVar()
        tk.Entry(fcard, textvariable=_sv_cli_filt,
                 font=("Helvetica",9), bg=COLORS["bg_main"],
                 fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"]
                 ).pack(fill="x", padx=18, pady=(4,0), ipady=5)
        self._serv_cli_sel_lbl = tk.Label(fcard,
            text="— nenhum —", font=("Helvetica",8,"bold"),
            bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._serv_cli_sel_lbl.pack(anchor="w", padx=20, pady=(2,0))
        _sv_cli_lb = tk.Listbox(fcard, font=("Helvetica",9), height=4,
                                bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                selectbackground=COLORS["accent"],
                                selectforeground="white",
                                relief="flat", bd=0, highlightthickness=1,
                                highlightbackground=COLORS["border"])
        _sv_cli_lb.pack(fill="x", padx=18, pady=(0,8))

        def _sv_cli_fill(*_):
            t = _sv_cli_filt.get().lower()
            _sv_cli_lb.delete(0, "end")
            for o in getattr(self, "_serv_cli_opts_all", []):
                if not t or t in o.lower():
                    _sv_cli_lb.insert("end", o)

        def _sv_cli_pick(e):
            sel = _sv_cli_lb.curselection()
            if not sel: return
            val = _sv_cli_lb.get(sel[0])
            self._serv_cliente_var.set(val)
            try: self._serv_cliente_id = int(val.split("—")[0].strip())
            except: self._serv_cliente_id = None
            name = val.split("—", 1)[1].strip() if "—" in val else val
            self._serv_cli_sel_lbl.configure(text=f"✔ {name}", fg=COLORS["accent"])
            self._atualizar_combo_carros_servico(filtro_cli_id=self._serv_cliente_id)

        _sv_cli_filt.trace_add("write", _sv_cli_fill)
        _sv_cli_lb.bind("<<ListboxSelect>>", _sv_cli_pick)
        self._serv_cliente_combo = _sv_cli_lb   # alias para compatibilidade

        # ── Carro: busca + listbox ────────────────────────────────────────
        lbl("Carro (Cliente)", required=True)
        self._serv_carro_var = tk.StringVar(value="")
        _sv_car_filt = tk.StringVar()
        tk.Entry(fcard, textvariable=_sv_car_filt,
                 font=("Helvetica",9), bg=COLORS["bg_main"],
                 fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"]
                 ).pack(fill="x", padx=18, pady=(4,0), ipady=5)
        self._serv_car_sel_lbl = tk.Label(fcard,
            text="— nenhum —", font=("Helvetica",8,"bold"),
            bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._serv_car_sel_lbl.pack(anchor="w", padx=20, pady=(2,0))
        _sv_car_lb = tk.Listbox(fcard, font=("Helvetica",9), height=4,
                                bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                selectbackground=COLORS["blue"],
                                selectforeground="white",
                                relief="flat", bd=0, highlightthickness=1,
                                highlightbackground=COLORS["border"])
        _sv_car_lb.pack(fill="x", padx=18, pady=(0,8))

        def _sv_car_fill(*_):
            t = _sv_car_filt.get().lower()
            _sv_car_lb.delete(0, "end")
            for o in getattr(self, "_serv_carro_opts_all", []):
                if not t or t in o.lower():
                    _sv_car_lb.insert("end", o)

        def _sv_car_pick(e):
            sel = _sv_car_lb.curselection()
            if not sel: return
            val = _sv_car_lb.get(sel[0])
            self._serv_carro_var.set(val)
            name = val.split("|")[0].split("—",1)[-1].strip()
            self._serv_car_sel_lbl.configure(text=f"✔ {name}", fg=COLORS["blue"])
            self._serv_auto_cliente()

        _sv_car_filt.trace_add("write", _sv_car_fill)
        _sv_car_lb.bind("<<ListboxSelect>>", _sv_car_pick)
        self._serv_carro_combo = _sv_car_lb   # alias para compatibilidade

        # Tipo de Serviço
        lbl("Tipo de Serviço", required=True)
        self._serv_tipo_combo = ttk.Combobox(fcard, state="readonly",
                                              font=("Helvetica", 10), width=28)
        self._serv_tipo_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self._atualizar_combo_tipos_servico()

        # Data do Serviço
        lbl("Data do Serviço")
        df = tk.Frame(fcard, bg=COLORS["bg_card"])
        df.pack(anchor="w", padx=18, pady=(4, 12))
        self._serv_dia, self._serv_mes, self._serv_ano = self._make_date_row(df)

        # Descrição
        lbl("Descrição")
        self._serv_desc_entry = tk.Entry(fcard, font=("Helvetica", 10),
                                          bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                          insertbackground=COLORS["text_primary"],
                                          relief="flat", bd=0,
                                          highlightthickness=1,
                                          highlightbackground=COLORS["border"],
                                          highlightcolor=COLORS["accent"], width=28)
        self._serv_desc_entry.pack(padx=18, pady=(4, 12), ipady=6)

        # Valor
        lbl("Valor (¥)")
        self._serv_valor_entry = self._make_yen_entry(fcard, width=20)
        self._serv_valor_entry.pack(padx=18, pady=(4, 12), ipady=7)

        # Status
        lbl("Status")
        self._serv_status_var = tk.StringVar(value="Em Andamento")
        ttk.Combobox(fcard, textvariable=self._serv_status_var,
                     values=["Em Andamento", "Concluído"],
                     state="readonly", font=("Helvetica", 10), width=28
                     ).pack(padx=18, pady=(4, 12), ipady=4)


        # ── NFI ───────────────────────────────────────────────────────────────
        tk.Label(fcard, text="NFI",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=18)
        _nfi_f_os = tk.Frame(fcard, bg=COLORS["bg_card"])
        _nfi_f_os.pack(anchor="w", padx=18, pady=(4, 12))
        self._serv_nfi_var = tk.StringVar(value="SNF")
        for _lbl, _val in [("SNF", "SNF"), ("CNF", "CNF")]:
            tk.Radiobutton(_nfi_f_os, text=_lbl, variable=self._serv_nfi_var, value=_val,
                           font=("Helvetica", 9),
                           bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                           activebackground=COLORS["bg_card"],
                           selectcolor=COLORS["bg_main"]).pack(anchor="w")

        # Status label + botões
        self._lbl_serv_status = tk.Label(fcard, text="", font=("Helvetica", 8),
                                          bg=COLORS["bg_card"], fg=COLORS["red"])
        self._lbl_serv_status.pack(padx=18, pady=(4, 4))

        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 10))
        btn_f = tk.Frame(fcard, bg=COLORS["bg_card"])
        btn_f.pack(padx=18, pady=(0, 20), fill="x")
        self._btn_salvar_serv = tk.Button(btn_f, text="  Registrar OS  ",
                                           font=("Helvetica", 10, "bold"),
                                           bg=COLORS["accent"], fg="white",
                                           relief="flat", cursor="hand2",
                                           command=self._salvar_servico)
        self._btn_salvar_serv.pack(side="left", ipady=7, ipadx=4, fill="x", expand=True)
        tk.Button(btn_f, text="Limpar", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_servico
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=4)
        tk.Button(btn_f, text="↺", font=("Helvetica", 10, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_tabela_servicos
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=6)

        # Observações (caixa de texto, abaixo dos botões)
        tk.Frame(fcard, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(8, 8))
        lbl("Observações")
        self._serv_obs_text = tk.Text(fcard, font=("Helvetica", 9),
                                      bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                      insertbackground=COLORS["text_primary"],
                                      relief="flat", bd=0,
                                      highlightthickness=1,
                                      highlightbackground=COLORS["border"],
                                      highlightcolor=COLORS["accent"],
                                      width=28, height=5, wrap="word")
        self._serv_obs_text.pack(padx=18, pady=(4, 18), fill="x")

        # Pré-carrega listas para filtro funcionar imediatamente
        self._atualizar_combo_clientes_servico()
        self._atualizar_combo_carros_servico()
        # Popula listboxes iniciais
        for o in getattr(self, "_serv_cli_opts_all", []):
            if hasattr(self, "_serv_cliente_combo") and isinstance(self._serv_cliente_combo, tk.Listbox):
                self._serv_cliente_combo.insert("end", o)
        for o in getattr(self, "_serv_carro_opts_all", []):
            if hasattr(self, "_serv_carro_combo") and isinstance(self._serv_carro_combo, tk.Listbox):
                self._serv_carro_combo.insert("end", o)
        self._atualizar_os_num()
        self._refresh_tabela_servicos()

    def _atualizar_os_num(self):
        if hasattr(self, "_lbl_os_num"):
            num = self._next_os_num()
            self._lbl_os_num.configure(text=f"Próximo número: {num}")

    def _atualizar_combo_carros_servico(self, filtro_cli_id=None):
        """Carros com status 'Cliente', opcionalmente filtrados por cliente."""
        if filtro_cli_id:
            carros = [dict(r) for r in self.conn.execute(
                "SELECT * FROM carros WHERE status='Cliente' AND cliente_id=? ORDER BY carro",
                (filtro_cli_id,)).fetchall()]
        else:
            carros = [dict(r) for r in self.conn.execute(
                "SELECT * FROM carros WHERE status='Cliente' ORDER BY carro").fetchall()]
        self._serv_carro_opts_all = [
            f"{c['id']} — {c['carro']} {c.get('ano') or ''} | {c.get('cor') or '—'} | {c.get('placa') or '—'}"
            for c in carros]
        if hasattr(self, "_serv_carro_combo") and isinstance(self._serv_carro_combo, tk.Listbox):
            self._serv_carro_combo.delete(0, "end")
            for o in self._serv_carro_opts_all:
                self._serv_carro_combo.insert("end", o)
        elif hasattr(self, "_serv_carro_combo"):
            self._serv_carro_combo["values"] = self._serv_carro_opts_all or ["Nenhum carro do tipo Cliente"]

    def _atualizar_combo_clientes_servico(self):
        clientes = [dict(r) for r in
                    self.conn.execute("SELECT * FROM clientes ORDER BY nome").fetchall()]
        self._serv_cli_opts_all = [f"{c['id']} — {c['nome']}" for c in clientes]
        if hasattr(self, "_serv_cliente_combo") and isinstance(self._serv_cliente_combo, tk.Listbox):
            self._serv_cliente_combo.delete(0, "end")
            for o in self._serv_cli_opts_all:
                self._serv_cliente_combo.insert("end", o)
        elif hasattr(self, "_serv_cliente_combo"):
            self._serv_cliente_combo["values"] = self._serv_cli_opts_all or ["Nenhum cliente cadastrado"]

    def _serv_filtrar_clientes_substring(self):
        pass  # filtering handled by Entry trace in _build_ordem_servico

    def _serv_filtrar_carros_substring(self):
        pass  # filtering handled by Entry trace in _build_ordem_servico

    def _serv_filtrar_carros_por_cliente(self):
        """Ao selecionar cliente, filtra carros vinculados a ele."""
        sel = self._serv_cliente_var.get()
        if not sel or "—" not in sel:
            return
        try:
            cli_id = int(sel.split("—")[0].strip())
        except Exception:
            return
        self._atualizar_combo_carros_servico(filtro_cli_id=cli_id)
        self._serv_carro_var.set("")

    def _serv_auto_cliente(self):
        """Ao selecionar carro, preenche automaticamente o cliente vinculado."""
        sel = self._serv_carro_var.get().strip()
        if not sel or "Nenhum" in sel:
            return
        try:
            carro_id = int(sel.split("—")[0].strip())
        except Exception:
            return
        row = self.conn.execute(
            "SELECT cliente_id FROM carros WHERE id=?", (carro_id,)).fetchone()
        if not row or not row[0]:
            return
        cliente_id = row[0]
        # Atualiza o combo de clientes e seleciona o vinculado
        self._atualizar_combo_clientes_servico()
        for opt in self._serv_cliente_combo["values"]:
            if opt.startswith(f"{cliente_id} —"):
                self._serv_cliente_var.set(opt)
                return

    def _salvar_servico(self):
        carro_sel   = self._serv_carro_var.get().strip()
        cli_sel     = self._serv_cliente_var.get().strip()
        tipo_serv   = self._serv_tipo_combo.get().strip()
        data_serv   = self._get_date_from_entries(self._serv_dia, self._serv_mes, self._serv_ano)
        desc        = self._serv_desc_entry.get().strip()
        valor       = self._yen_raw(self._serv_valor_entry)
        status      = self._serv_status_var.get()
        obs         = self._serv_obs_text.get("1.0", tk.END).strip()

        if not carro_sel or "Nenhum" in carro_sel:
            self._lbl_serv_status.configure(text="⚠ Selecione um carro.", fg=COLORS["red"]); return
        if not cli_sel or "—" not in cli_sel:
            self._lbl_serv_status.configure(text="⚠ Selecione um cliente.", fg=COLORS["red"]); return
        if not tipo_serv:
            self._lbl_serv_status.configure(text="⚠ Selecione o tipo de serviço.", fg=COLORS["red"]); return

        try:
            carro_id = int(carro_sel.split("—")[0].strip())
        except Exception:
            carro_id = None
        carro_txt = carro_sel.split("|")[0].strip()
        try:
            cliente_id = int(cli_sel.split("—")[0].strip())
        except Exception:
            cliente_id = None

        if self._serv_edit_id is not None:
            nfi_serv = getattr(self, "_serv_nfi_var", None)
            nfi_serv = nfi_serv.get() if nfi_serv else "SNF"
            self.conn.execute("""UPDATE servicos SET
                carro_id=?,carro=?,cliente_id=?,data_servico=?,tipo_servico=?,
                descricao=?,valor=?,status=?,obs=?,nfi=? WHERE id=?""",
                (carro_id, carro_txt, cliente_id, data_serv, tipo_serv,
                 desc, valor, status, obs, nfi_serv, self._serv_edit_id))
            self.conn.commit()
            # Atualiza IS se CNF
            if nfi_serv == "CNF" and valor:
                self._upsert_is("custos_os", "servico_id",
                                self._serv_edit_id, valor, data_serv)
            self._serv_edit_id = None
            self._lbl_serv_title.configure(text="⚙  Nova Ordem de Serviço")
            self._btn_salvar_serv.configure(text="  Registrar OS  ")
            self._lbl_serv_status.configure(text="✔ OS atualizada!", fg=COLORS["green"])
        else:
            os_num = self._next_os_num()
            nfi_serv = getattr(self, "_serv_nfi_var", None)
            nfi_serv = nfi_serv.get() if nfi_serv else "SNF"
            self.conn.execute("""INSERT INTO servicos
                (os_num,carro_id,carro,cliente_id,data_servico,tipo_servico,descricao,valor,status,obs,nfi)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (os_num, carro_id, carro_txt, cliente_id, data_serv,
                 tipo_serv, desc, valor, status, obs, nfi_serv))
            self.conn.commit()
            # IS automático para CNF (OS) → upsert em custos_os
            if nfi_serv == "CNF" and valor:
                sid_is = self.conn.execute(
                    "SELECT id FROM servicos ORDER BY id DESC LIMIT 1").fetchone()
                if sid_is:
                    self._upsert_is("custos_os", "servico_id",
                                    sid_is[0], valor, data_serv)
            self._lbl_serv_status.configure(text=f"✔ {os_num} registrada!", fg=COLORS["green"])

        self.servicos_data = [dict(r) for r in
            self.conn.execute("SELECT * FROM servicos ORDER BY id DESC").fetchall()]
        self._limpar_form_servico(clear_status=False)
        self._atualizar_os_num()
        self._refresh_tabela_servicos()

    def _limpar_form_servico(self, clear_status=True):
        import datetime
        self._serv_carro_var.set("")
        self._serv_cliente_var.set("")
        self._serv_cliente_id = None
        if hasattr(self, "_serv_cli_sel_lbl"):
            self._serv_cli_sel_lbl.configure(text="— nenhum selecionado —", fg=COLORS["text_muted"])
        if hasattr(self, "_serv_car_sel_lbl"):
            self._serv_car_sel_lbl.configure(text="— nenhum selecionado —", fg=COLORS["text_muted"])
        self._atualizar_combo_carros_servico()
        self._serv_tipo_combo.set("")
        self._serv_desc_entry.delete(0, tk.END)
        self._serv_valor_entry.delete(0, tk.END)
        self._serv_status_var.set("Em Andamento")
        if hasattr(self, "_serv_obs_text"):
            self._serv_obs_text.delete("1.0", tk.END)
        self._serv_edit_id = None
        self._lbl_serv_title.configure(text="⚙  Nova Ordem de Serviço")
        self._btn_salvar_serv.configure(text="  Registrar OS  ")
        hoje = datetime.date.today()
        for e, val in [(self._serv_dia, str(hoje.day).zfill(2)),
                       (self._serv_mes, str(hoje.month).zfill(2)),
                       (self._serv_ano, str(hoje.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        if clear_status:
            self._lbl_serv_status.configure(text="")

    def _editar_servico(self, sid):
        s = next((x for x in self.servicos_data if x["id"] == sid), None)
        if not s: return
        self._serv_edit_id = sid
        self._lbl_serv_title.configure(text="✏  Editar OS")
        self._btn_salvar_serv.configure(text="  Salvar Alterações  ")
        self._lbl_os_num.configure(text=f"Editando: {s.get('os_num','—')}")

        self._atualizar_combo_carros_servico()
        cid = s.get("carro_id")
        cvals = list(self._serv_carro_combo["values"])
        cm = next((x for x in cvals if x.startswith(f"{cid} —")), "")
        self._serv_carro_var.set(cm or s.get("carro", ""))

        self._atualizar_combo_clientes_servico()
        clid = s.get("cliente_id")
        clvals = list(self._serv_cliente_combo["values"])
        clm = next((x for x in clvals if x.startswith(f"{clid} —")), "") if clid else ""
        self._serv_cliente_var.set(clm)

        self._atualizar_combo_tipos_servico()
        self._serv_tipo_combo.set(s.get("tipo_servico", ""))
        self._restore_date(self._serv_dia, self._serv_mes, self._serv_ano, s.get("data_servico"))
        self._serv_desc_entry.delete(0, tk.END)
        self._serv_desc_entry.insert(0, s.get("descricao") or "")
        try:
            v = int(str(s.get("valor") or "0").replace(",", ""))
            self._serv_valor_entry.delete(0, tk.END)
            self._serv_valor_entry.insert(0, f"{v:,}")
        except Exception:
            pass
        st_load = s.get("status", "Em Andamento")
        # Map legacy statuses
        if st_load == "Aberto":    st_load = "Em Andamento"
        if st_load == "Cancelado": st_load = "Concluído"
        self._serv_status_var.set(st_load)
        if hasattr(self, "_serv_obs_text"):
            self._serv_obs_text.delete("1.0", tk.END)
            self._serv_obs_text.insert("1.0", s.get("obs") or "")
        # NFI
        if hasattr(self, "_serv_nfi_var"):
            self._serv_nfi_var.set(s.get("nfi") or "SNF")
        self._lbl_serv_status.configure(text="")

    def _excluir_servico(self, sid):
        if not msgbox.askyesno("Confirmar", "Deseja excluir esta ordem de serviço?"):
            return
        self.conn.execute("DELETE FROM servicos WHERE id=?", (sid,))
        self.conn.commit()
        self.servicos_data = [dict(r) for r in
            self.conn.execute("SELECT * FROM servicos ORDER BY id DESC").fetchall()]
        self._refresh_tabela_servicos()
        self._atualizar_os_num()

    def _refresh_tabela_servicos(self):
        for w in self._serv_rows_frame.winfo_children():
            w.destroy()

        lista = self.servicos_data[:]
        termo = self._serv_filtro.get().strip().lower()
        if termo:
            lista = [s for s in lista if
                     termo in (s.get("carro") or "").lower() or
                     termo in (s.get("tipo_servico") or "").lower() or
                     termo in (s.get("os_num") or "").lower()]

        fmes = self._serv_filtro_mes.get()
        fano = self._serv_filtro_ano.get()
        if fmes != "Todos" or fano != "Todos":
            def match(s):
                data = s.get("data_servico") or ""
                pts = data.split("/")
                vm = pts[1] if len(pts) > 1 else ""
                va = pts[2] if len(pts) > 2 else ""
                if fmes != "Todos" and vm != fmes: return False
                if fano != "Todos" and va != fano: return False
                return True
            lista = [s for s in lista if match(s)]

        fstatus_serv = getattr(self, "_serv_filtro_status", None)
        if fstatus_serv:
            fs = fstatus_serv.get()
            if fs != "Todos":
                lista = [s for s in lista if s.get("status","") == fs]

        # Filtro NFI
        fnfi_os = getattr(self, "_serv_filtro_nfi", None)
        fnfi_os = fnfi_os.get() if fnfi_os else "Todos"
        if fnfi_os != "Todos":
            lista = [s for s in lista if (s.get("nfi") or "SNF") == fnfi_os]

        self._lbl_total_serv.configure(text=f"{len(lista)} registro(s)")
        self._servicos_lista_filtrada = lista[:]

        if not lista:
            tk.Label(self._serv_rows_frame, text="Nenhuma OS encontrada.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        STATUS_COLORS = {
            "Em Andamento": COLORS["orange"],
            "Em Andamento": COLORS["orange"],
            "Concluído":    COLORS["green"],
            "Cancelado":    COLORS["red"],
        }

        for i, s in enumerate(lista):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._serv_rows_frame, bg=rb)
            row.pack(fill="x")

            tk.Label(row, text=s.get("os_num") or "—",
                     font=("Helvetica", 8, "bold"), bg=rb,
                     fg=COLORS["accent"], width=8, anchor="w").pack(side="left", padx=2, pady=6)
            tk.Label(row, text=s.get("data_servico") or "—",
                     font=("Helvetica", 9), bg=rb, fg=COLORS["text_secondary"],
                     width=9, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=s.get("carro") or "—",
                     font=("Helvetica", 9), bg=rb, fg=COLORS["text_primary"],
                     width=16, anchor="w").pack(side="left", padx=2)

            # Busca nome do cliente
            try:
                cli_row = self.conn.execute(
                    "SELECT nome FROM clientes WHERE id=?",
                    (s.get("cliente_id"),)).fetchone()
                cli_nome = cli_row[0] if cli_row else "—"
            except Exception:
                cli_nome = "—"
            tk.Label(row, text=cli_nome, font=("Helvetica", 9), bg=rb,
                     fg=COLORS["text_secondary"], width=20, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=s.get("tipo_servico") or "—",
                     font=("Helvetica", 9), bg=rb, fg=COLORS["text_primary"],
                     width=14, anchor="w").pack(side="left", padx=2)
            tk.Label(row, text=self._fmt_yen_display(s.get("valor")),
                     font=("Helvetica", 9), bg=rb, fg=COLORS["green"],
                     width=10, anchor="w").pack(side="left", padx=2)

            sc = STATUS_COLORS.get(s.get("status", "Aberto"), COLORS["text_muted"])
            tk.Label(row, text=s.get("status", "—"),
                     font=("Helvetica", 7, "bold"), bg=sc, fg="white",
                     width=13, anchor="center").pack(side="left", padx=2, pady=3)

            # NFI badge
            nfi_os = s.get("nfi") or "SNF"
            nfi_os_bg = "#8E44AD" if nfi_os == "CNF" else "#7F8C8D"
            tk.Label(row, text=nfi_os, font=("Helvetica",7,"bold"),
                     bg=nfi_os_bg, fg="white", width=5, anchor="center"
                     ).pack(side="left", padx=2, pady=3)

            acts = tk.Frame(row, bg=rb)
            acts.pack(side="left", padx=2)
            tk.Button(acts, text="✏", font=("Helvetica", 8),
                      bg=COLORS["blue"], fg="white", relief="flat",
                      cursor="hand2", padx=4, pady=1,
                      command=lambda sid=s["id"]: self._editar_servico(sid)
                      ).pack(side="left", padx=(0, 2))
            tk.Button(acts, text="✕", font=("Helvetica", 8),
                      bg=COLORS["red"], fg="white", relief="flat",
                      cursor="hand2", padx=4, pady=1,
                      command=lambda sid=s["id"]: self._excluir_servico(sid)
                      ).pack(side="left", padx=(0, 2))
            # Botão Concluir — ativo só se não estiver Concluído
            is_done = s.get("status") == "Concluído"
            tk.Button(acts, text="✔ Concluir",
                      font=("Helvetica", 7, "bold"),
                      bg=COLORS["green"] if not is_done else COLORS["border"],
                      fg="white", relief="flat",
                      cursor="hand2" if not is_done else "arrow",
                      padx=4, pady=1,
                      state="normal" if not is_done else "disabled",
                      command=(lambda sid=s["id"]: self._concluir_servico(sid))
                               if not is_done else (lambda: None)
                      ).pack(side="left", padx=(0, 2))
            # Botão de observações — amarelo, sempre visível
            obs_color = "#F1C40F"
            obs_fg    = "#333333"
            tk.Button(acts, text="📝", font=("Helvetica", 8),
                      bg=obs_color, fg=obs_fg, relief="flat",
                      cursor="hand2", padx=4, pady=1,
                      command=lambda sid=s["id"]: self._os_sidebar_abrir(sid)
                      ).pack(side="left")

    def _concluir_servico(self, sid):
        """Marca um serviço como Concluído e atualiza as telas."""
        import datetime as _dtc
        import tkinter.messagebox as mb
        try:
            row = self.conn.execute(
                "SELECT os_num, status FROM servicos WHERE id=?", (sid,)).fetchone()
            if not row:
                return
            os_num, status = row
            if status == "Concluído":
                return
            nome_os = os_num or f"#{sid}"
            if not mb.askyesno("Concluir Serviço",
                               f"Confirmar conclusão da OS {nome_os}?"):
                return
            hoje = _dtc.date.today()
            data_str = f"{hoje.day:02d}/{hoje.month:02d}/{hoje.year}"
            # Tenta gravar data_conclusao; se coluna não existir, grava só o status
            try:
                self.conn.execute(
                    "UPDATE servicos SET status='Concluído', data_conclusao=? WHERE id=?",
                    (data_str, sid))
            except Exception:
                self.conn.execute(
                    "UPDATE servicos SET status='Concluído' WHERE id=?", (sid,))
            self.conn.commit()
            # Recarrega dados e atualiza tabela
            self.servicos_data = [dict(r) for r in
                self.conn.execute("SELECT * FROM servicos ORDER BY id DESC").fetchall()]
            self._refresh_tabela_servicos()
            # Atualiza dashboard Serviços se estiver montado
            if hasattr(self, "_dash_serv_rows"):
                try: self._refresh_dash_servicos()
                except Exception: pass
        except Exception as e:
            import tkinter.messagebox as mb2
            mb2.showerror("Erro", f"Nao foi possivel concluir o servico: {e}")

    def _os_sidebar_abrir(self, sid):
        """Preenche o sidebar com detalhes e obs da OS selecionada."""
        row = self.conn.execute(
            "SELECT s.*, cl.nome as cli_nome FROM servicos s "
            "LEFT JOIN clientes cl ON s.cliente_id=cl.id "
            "WHERE s.id=?", (sid,)).fetchone()
        if not row:
            return
        s = dict(row)
        self._os_sidebar_sid = sid

        # Atualiza título
        os_num = s.get("os_num") or f"#{sid}"
        self._os_sidebar_title.configure(text=f"📝  {os_num}")

        # Reconstrói info
        for w in self._os_sidebar_info.winfo_children():
            w.destroy()

        STATUS_COLORS = {"Aberto": COLORS["blue"], "Em Andamento": COLORS["orange"],
                         "Concluído": COLORS["green"], "Cancelado": COLORS["red"]}

        def info_row(label, val, fg=None):
            f = tk.Frame(self._os_sidebar_info, bg=COLORS["bg_card"])
            f.pack(fill="x", pady=1)
            tk.Label(f, text=label, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"], width=10, anchor="w"
                     ).pack(side="left")
            tk.Label(f, text=val, font=("Helvetica",8),
                     bg=COLORS["bg_card"], fg=fg or COLORS["text_primary"], anchor="w"
                     ).pack(side="left", fill="x", expand=True)

        carro_txt = (s.get("carro") or "—").split("|")[0].strip()
        if " — " in carro_txt: carro_txt = carro_txt.split(" — ", 1)[1]
        info_row("OS",      os_num,                             COLORS["accent"])
        info_row("Data",    s.get("data_servico") or "—",       COLORS["text_secondary"])
        info_row("Carro",   carro_txt[:22],                     COLORS["text_primary"])
        info_row("Cliente", s.get("cli_nome") or "—",           COLORS["blue"])
        info_row("Serviço", s.get("tipo_servico") or "—",       COLORS["text_secondary"])
        info_row("Valor",   self._fmt_yen_display(s.get("valor")), COLORS["green"])
        st = s.get("status") or "—"
        sc = STATUS_COLORS.get(st, COLORS["text_muted"])
        info_row("Status",  st,                                  sc)
        tk.Frame(self._os_sidebar_info, bg=COLORS["border"], height=1).pack(fill="x", pady=(4,0))

        # Preenche obs
        obs_db = s.get("obs") or ""
        self._os_sidebar_txt.configure(state="normal")
        self._os_sidebar_txt.delete("1.0", tk.END)
        self._os_sidebar_txt.insert("1.0", obs_db)
        self._os_sidebar_status.configure(text="")

    def _os_sidebar_salvar(self):
        """Salva obs da OS do sidebar."""
        sid = self._os_sidebar_sid
        if not sid:
            return
        nova = self._os_sidebar_txt.get("1.0", tk.END).strip()
        self.conn.execute("UPDATE servicos SET obs=? WHERE id=?", (nova, sid))
        self.conn.commit()
        self.servicos_data = [dict(r) for r in
            self.conn.execute("SELECT * FROM servicos ORDER BY id DESC").fetchall()]
        self._refresh_tabela_servicos()
        self._os_sidebar_status.configure(text="✔ Salvo!")

    def _ver_obs_servico(self, sid, obs_txt):
        """Abre mini janela para ler (e editar) a observação da OS."""
        # Busca dados frescos do banco
        row = self.conn.execute(
            "SELECT os_num, carro, obs FROM servicos WHERE id=?", (sid,)).fetchone()
        if not row:
            return
        os_num = row[0] or f"#{sid}"
        carro  = (row[1] or "—").split("|")[0].strip()
        obs_db = row[2] or ""

        dlg = tk.Toplevel(self)
        dlg.title(f"Observações — {os_num}")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(True, True)
        dlg.lift()
        dlg.attributes("-topmost", True)
        dlg.after(100, lambda: dlg.attributes("-topmost", False))

        tk.Frame(dlg, bg="#F1C40F", height=5).pack(fill="x")
        hdr = tk.Frame(dlg, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12, 4))
        tk.Label(hdr, text=f"📝  Observações  —  {os_num}",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Label(hdr, text=carro,
                 font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="right")

        tk.Frame(dlg, bg=COLORS["border"], height=1).pack(fill="x", padx=16)

        # Caixa de texto editável
        txt = tk.Text(dlg, font=("Helvetica", 10),
                      bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                      insertbackground=COLORS["text_primary"],
                      relief="flat", bd=0,
                      highlightthickness=1,
                      highlightbackground=COLORS["border"],
                      highlightcolor="#F1C40F",
                      wrap="word", width=52, height=12)
        txt.pack(padx=16, pady=12, fill="both", expand=True)
        txt.insert("1.0", obs_db)

        lbl_status = tk.Label(dlg, text="", font=("Helvetica", 8),
                              bg=COLORS["bg_card"], fg=COLORS["green"])
        lbl_status.pack(padx=16, pady=(0, 4))

        btn_f = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_f.pack(padx=16, pady=(0, 14), fill="x")

        def salvar_obs():
            nova = txt.get("1.0", tk.END).strip()
            self.conn.execute(
                "UPDATE servicos SET obs=? WHERE id=?", (nova, sid))
            self.conn.commit()
            # Atualiza servicos_data em memória
            self.servicos_data = [dict(r) for r in
                self.conn.execute("SELECT * FROM servicos ORDER BY id DESC").fetchall()]
            self._refresh_tabela_servicos()
            lbl_status.configure(text="✔ Observação salva!")

        tk.Button(btn_f, text="  Salvar  ",
                  font=("Helvetica", 10, "bold"),
                  bg="#F1C40F", fg="#333333", relief="flat", cursor="hand2",
                  command=salvar_obs
                  ).pack(side="left", ipady=5, ipadx=8)
        tk.Button(btn_f, text="  Fechar  ",
                  font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy
                  ).pack(side="left", padx=(10, 0), ipady=5, ipadx=8)

        dlg.update_idletasks()
        w = max(dlg.winfo_reqwidth(), 480)
        h = max(dlg.winfo_reqheight(), 320)
        x = self.winfo_x() + (self.winfo_width()  - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    def _build_dashboard_subs(self, parent):
        self.sub_pages["Dashboard"] = {}
        subs = SUBMENUS["Dashboard"]
        for sub in subs:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Dashboard"][sub] = frame
            if sub == "Visão Geral":
                self._build_dash_visao_geral(frame)
            elif sub == "Estoque":
                self._build_dash_estoque(frame)
            elif sub == "Vendas do Mês":
                self._build_dash_vendas(frame)
            elif sub == "Serviços":
                self._build_dash_servicos(frame)
            elif sub == "Parcelas Vencidas":
                self._build_dash_parcelas(frame)
            elif sub == "Shaken a Vencer":
                self._build_dash_shaken_vencer(frame)
            elif sub == "Indicadores":
                self._build_dash_indicadores(frame)
            elif sub == "Relatório Mensal":
                self._build_relatorio_mensal(frame)
            elif sub == "Relatório Anual":
                self._build_relatorio_anual(frame)
            elif sub == "Dossiê Cliente":
                self._build_dash_dossie_cliente(frame)
            elif sub == "Dossiê Carro":
                self._build_dash_dossie_carro(frame)
        first = subs[0]
        self.sub_pages["Dashboard"][first].place(in_=parent, x=0, y=0, relwidth=1, relheight=1)

    # ── Dashboard: Visão Geral ────────────────────────────────────────────────
    def _build_dash_visao_geral(self, parent):
        import datetime, calendar
        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=18, pady=18)

        tk.Label(root, text="⊞  Visão Geral", font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(anchor="w", pady=(0,14))

        # ── KPI cards ─────────────────────────────────────────────────────────
        cards_row = tk.Frame(root, bg=COLORS["bg_content"])
        cards_row.pack(fill="x", pady=(0, 16))

        hoje = datetime.date.today()
        mes_str  = f"{hoje.month:02d}"
        ano_str  = str(hoje.year)

        # Estoque
        estoque_cnt = self.conn.execute(
            "SELECT COUNT(*) FROM compras WHERE a_venda=1").fetchone()[0]
        # Vendas do mês
        vendas_mes = [dict(r) for r in self.conn.execute(
            "SELECT valor_venda FROM vendas WHERE data_venda LIKE ?",
            (f"%/{mes_str}/{ano_str}",)).fetchall()]
        vendas_cnt = len(vendas_mes)
        try:
            vendas_val = sum(int(str(v["valor_venda"] or "0").replace(",","")) for v in vendas_mes)
        except Exception:
            vendas_val = 0
        # Serviços abertos
        serv_abertos = self.conn.execute(
            "SELECT COUNT(*) FROM servicos WHERE status='Em Andamento' OR status='Aberto'"
        ).fetchone()[0]
        # Financiamentos em aberto (saldo > 0)
        fin_abertos = self.conn.execute(
            "SELECT COUNT(*) FROM vendas WHERE tipo_venda IN "
            "('Venda Parcelada','Com Troca','Com Troca e Volta') "
            "AND parcelas_pagas < num_parcelas").fetchone()[0]

        # Shaken a vencer (vencidos + < 3 meses), excluindo renovados e por conta
        import datetime as _dt
        _hoje = datetime.date.today()
        _sk_urgentes = 0
        _sk_vencidos = 0
        try:
            _sk_rows = self.conn.execute(
                "SELECT data_vencimento FROM shaken "
                "WHERE (renovado IS NULL OR renovado=0) "
                "AND (por_conta IS NULL OR por_conta=0) "
                "AND data_vencimento IS NOT NULL AND data_vencimento != ''"
            ).fetchall()
            for (_dv,) in _sk_rows:
                try:
                    _d, _m, _a = _dv.split("/")
                    _dobj = datetime.date(int(_a), int(_m), int(_d))
                    _delta = (_dobj - _hoje).days
                    if _delta < 0:
                        _sk_vencidos += 1
                        _sk_urgentes += 1
                    elif _delta <= 90:
                        _sk_urgentes += 1
                except Exception:
                    pass
        except Exception:
            pass
        _sk_sub = f"{_sk_vencidos} vencido(s)" if _sk_vencidos else "todos no prazo" if _sk_urgentes == 0 else ""
        _sk_color = COLORS["red"] if _sk_vencidos > 0 else ("#D4AC0D" if _sk_urgentes > 0 else COLORS["green"])

        # Estoque parado > 60 dias
        _parados = 0
        try:
            _comp_rows = self.conn.execute(
                "SELECT data_entrada, data_compra FROM compras WHERE a_venda=1").fetchall()
            for _row in _comp_rows:
                _de = _row[0] or _row[1] or ""
                if _de:
                    try:
                        _d,_m,_a = _de.split("/")
                        _dobj = datetime.date(int(_a),int(_m),int(_d))
                        if (hoje - _dobj).days > 60:
                            _parados += 1
                    except Exception:
                        pass
        except Exception:
            pass
        _par_color = COLORS["red"] if _parados > 0 else COLORS["green"]

        kpis = [
            ("▦  Estoque",        str(estoque_cnt),           COLORS["blue"],   "veículos"),
            ("◆  Vendas do Mês",  str(vendas_cnt),            COLORS["green"],  f"¥ {vendas_val:,}"),
            ("⚙  Serviços",       str(serv_abertos),          COLORS["orange"], "em aberto"),
            ("◎  Financiamentos", str(fin_abertos),           COLORS["accent"], "em aberto"),
            ("🔔  Shaken",        str(_sk_urgentes),          _sk_color,        _sk_sub or "a vencer / vencidos"),
            ("⏳  Parados +60d",  str(_parados),              _par_color,       "carros em estoque"),
        ]
        for label, val, color, sub in kpis:
            c = tk.Frame(cards_row, bg=COLORS["bg_card"],
                         highlightthickness=1, highlightbackground=COLORS["border"])
            c.pack(side="left", expand=True, fill="x", padx=(0,10), ipady=10, ipadx=16)
            tk.Frame(c, bg=color, height=3).pack(fill="x")
            tk.Label(c, text=val, font=("Helvetica", 28, "bold"),
                     bg=COLORS["bg_card"], fg=color).pack(pady=(10,2))
            tk.Label(c, text=label, font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack()
            tk.Label(c, text=sub, font=("Helvetica", 8),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(0,6))

        # ── Cards de Receita Total e Quantidade ──────────────────────────────
        import datetime as _dt2
        _hoje2 = _dt2.date.today()
        _mes2  = f"{_hoje2.month:02d}"
        _ano2  = str(_hoje2.year)
        _pat2  = f"%/{_mes2}/{_ano2}"

        def _int2(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        # Receitas do mês
        _rec_v  = sum(_int2(r[0]) for r in self.conn.execute(
            "SELECT valor_venda FROM vendas WHERE data_venda LIKE ?", (_pat2,)).fetchall())
        _rec_os = sum(_int2(r[0]) for r in self.conn.execute(
            "SELECT valor FROM servicos WHERE data_servico LIKE ?", (_pat2,)).fetchall())
        _rec_sk = sum(_int2(r[0]) for r in self.conn.execute(
            "SELECT valor FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL", (_pat2,)).fetchall())
        _rec_tot = _rec_v + _rec_os + _rec_sk

        # Quantidades do mês
        _cnt_v  = self.conn.execute(
            "SELECT COUNT(*) FROM vendas WHERE data_venda LIKE ?", (_pat2,)).fetchone()[0]
        _cnt_os = self.conn.execute(
            "SELECT COUNT(*) FROM servicos WHERE data_servico LIKE ?", (_pat2,)).fetchone()[0]
        _cnt_sk = self.conn.execute(
            "SELECT COUNT(*) FROM shaken WHERE data_registro LIKE ?", (_pat2,)).fetchone()[0]
        _cnt_tot = _cnt_v + _cnt_os + _cnt_sk

        charts_row = tk.Frame(root, bg=COLORS["bg_content"])
        charts_row.pack(fill="both", expand=True, pady=(4, 0))

        COLOR_V  = COLORS["green"]
        COLOR_OS = COLORS["blue"]
        COLOR_SK = "#D4AC0D"

        for card_title, rows_data, total_lbl, total_val in [
            (
                "💰  Receita do Mês",
                [
                    ("◆  Vendas",            f"¥ {_rec_v:,}",  COLOR_V),
                    ("⚙  Ordens de Serviço", f"¥ {_rec_os:,}", COLOR_OS),
                    ("🔔  Shaken",           f"¥ {_rec_sk:,}", COLOR_SK),
                ],
                "Total",
                f"¥ {_rec_tot:,}",
            ),
            (
                "📦  Quantidade do Mês",
                [
                    ("◆  Vendas",            str(_cnt_v),  COLOR_V),
                    ("⚙  Ordens de Serviço", str(_cnt_os), COLOR_OS),
                    ("🔔  Shaken",           str(_cnt_sk), COLOR_SK),
                ],
                "Total",
                str(_cnt_tot),
            ),
        ]:
            card = tk.Frame(charts_row, bg=COLORS["bg_card"],
                            highlightthickness=1, highlightbackground=COLORS["border"])
            card.pack(side="left", fill="both", expand=True, padx=(0, 10))
            tk.Frame(card, bg=COLORS["accent"], height=3).pack(fill="x")
            tk.Label(card, text=card_title, font=("Helvetica", 10, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                     ).pack(anchor="w", padx=16, pady=(10, 6))

            # Linhas por tipo
            for row_lbl, row_val, row_color in rows_data:
                row_f = tk.Frame(card, bg=COLORS["bg_card"])
                row_f.pack(fill="x", padx=16, pady=3)
                # Barra colorida lateral
                tk.Frame(row_f, bg=row_color, width=4, height=40
                         ).pack(side="left", padx=(0, 10))
                inner = tk.Frame(row_f, bg=COLORS["bg_card"])
                inner.pack(side="left", fill="both", expand=True)
                tk.Label(inner, text=row_lbl, font=("Helvetica", 8),
                         bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                         anchor="w").pack(anchor="w")
                tk.Label(inner, text=row_val, font=("Helvetica", 18, "bold"),
                         bg=COLORS["bg_card"], fg=row_color,
                         anchor="w").pack(anchor="w")

            # Separador + Total
            tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(8, 0))
            tot_f = tk.Frame(card, bg=COLORS["bg_card"])
            tot_f.pack(fill="x", padx=16, pady=(6, 12))
            tk.Label(tot_f, text=total_lbl, font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="left")
            tk.Label(tot_f, text=total_val, font=("Helvetica", 16, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="right")

        # ── Cards anuais com gráfico de colunas ──────────────────────────────
        _ano_atual = datetime.date.today().year
        annual_row = tk.Frame(root, bg=COLORS["bg_content"])
        annual_row.pack(fill="both", expand=True, pady=(12, 0))

        annual_left  = tk.Frame(annual_row, bg=COLORS["bg_content"])
        annual_right = tk.Frame(annual_row, bg=COLORS["bg_content"])
        annual_left.pack(side="left",  fill="both", expand=True, padx=(0, 6))
        annual_right.pack(side="right", fill="both", expand=True, padx=(6, 0))

        self._build_annual_charts(annual_left,  "receita",    _ano_atual)
        self._build_annual_charts(annual_right, "quantidade", _ano_atual)

    def _build_annual_charts(self, parent, mode, ano):
        """Cards anuais: receita/quantidade com colunas empilhadas + tooltip ao hover."""
        import datetime as _dta

        def _int(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        MESES_ABR = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
        COLOR_V   = COLORS["green"]
        COLOR_OS  = COLORS["blue"]
        COLOR_SK  = "#D4AC0D"
        COLOR_TOT = COLORS["accent"]

        # ── Coleta dados por mês ─────────────────────────────────────────────
        dados_v  = []; dados_os = []; dados_sk = []
        for m in range(1, 13):
            ms  = f"{m:02d}"; ys = str(ano); pat = f"%/{ms}/{ys}"
            if mode == "receita":
                v   = sum(_int(r[0]) for r in self.conn.execute(
                    "SELECT valor_venda FROM vendas WHERE data_venda LIKE ?", (pat,)).fetchall())
                os_ = sum(_int(r[0]) for r in self.conn.execute(
                    "SELECT valor FROM servicos WHERE data_servico LIKE ?", (pat,)).fetchall())
                sk  = sum(_int(r[0]) for r in self.conn.execute(
                    "SELECT valor FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL", (pat,)).fetchall())
            else:
                v   = self.conn.execute("SELECT COUNT(*) FROM vendas WHERE data_venda LIKE ?",    (pat,)).fetchone()[0]
                os_ = self.conn.execute("SELECT COUNT(*) FROM servicos WHERE data_servico LIKE ?", (pat,)).fetchone()[0]
                sk  = self.conn.execute("SELECT COUNT(*) FROM shaken WHERE data_registro LIKE ?",  (pat,)).fetchone()[0]
            dados_v.append(v); dados_os.append(os_); dados_sk.append(sk)

        tot_v = sum(dados_v); tot_os = sum(dados_os); tot_sk = sum(dados_sk)
        tot_all = tot_v + tot_os + tot_sk
        is_rec  = (mode == "receita")
        title   = f"{'💰' if is_rec else '📦'}  {'Receita' if is_rec else 'Quantidade'} Anual — {ano}"

        card = tk.Frame(parent, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True, pady=(8, 0))
        tk.Frame(card, bg=COLOR_TOT, height=3).pack(fill="x")

        # título + legenda
        hdr_f = tk.Frame(card, bg=COLORS["bg_card"])
        hdr_f.pack(fill="x", padx=14, pady=(8, 4))
        tk.Label(hdr_f, text=title, font=("Helvetica", 10, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        leg_f = tk.Frame(hdr_f, bg=COLORS["bg_card"])
        leg_f.pack(side="right")
        for lbl, clr in [("Vendas", COLOR_V), ("OS", COLOR_OS), ("Shaken", COLOR_SK)]:
            tk.Frame(leg_f, bg=clr, width=10, height=10).pack(side="left", padx=(6,2))
            tk.Label(leg_f, text=lbl, font=("Helvetica", 7),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="left", padx=(0,4))

        # ── Canvas ──────────────────────────────────────────────────────────
        cv_frame = tk.Frame(card, bg=COLORS["bg_card"])
        cv_frame.pack(fill="x", padx=14, pady=(0, 4))
        CHART_H = 140
        cv = tk.Canvas(cv_frame, bg=COLORS["bg_card"], highlightthickness=0, height=CHART_H + 30)
        cv.pack(fill="x")

        # Tooltip window
        _tip_win  = [None]
        _tip_lbl  = [None]
        # bar_zones: list of (x0, x1, mes_idx) para hit-test
        _bar_zones = []

        def _fmt_val(v):
            if is_rec: return f"¥ {v:,}"
            return str(v)

        def _draw(event=None):
            cv.delete("all")
            _bar_zones.clear()
            W = cv.winfo_width()
            if W < 40: return
            combined = [a + b + c for a, b, c in zip(dados_v, dados_os, dados_sk)]
            max_v = max(combined) if any(combined) else 1
            pad_l, pad_r = 48, 8
            avail = W - pad_l - pad_r
            slot  = avail / 12
            bar_w = max(4, int(slot * 0.65))

            for pct in [0.25, 0.5, 0.75, 1.0]:
                gy = CHART_H - int(CHART_H * pct)
                cv.create_line(pad_l, gy, W - pad_r, gy, fill=COLORS["border"], dash=(3, 3))
                lbl = (f"¥{int(max_v*pct)//1000}k" if is_rec and max_v >= 1000
                       else f"¥{int(max_v*pct):,}" if is_rec
                       else str(int(max_v * pct)))
                cv.create_text(pad_l - 4, gy, text=lbl, anchor="e",
                               font=("Helvetica", 6), fill=COLORS["text_muted"])

            for i in range(12):
                xc   = pad_l + slot * i + slot / 2
                x0   = xc - bar_w / 2; x1 = xc + bar_w / 2
                base = CHART_H
                for val, clr in [(dados_sk[i], COLOR_SK),
                                 (dados_os[i], COLOR_OS),
                                 (dados_v[i],  COLOR_V)]:
                    if val <= 0: continue
                    h = max(2, int(CHART_H * val / max_v))
                    cv.create_rectangle(x0, base - h, x1, base, fill=clr, outline="")
                    base -= h
                total_m = dados_v[i] + dados_os[i] + dados_sk[i]
                if total_m > 0:
                    top_y = CHART_H - int(CHART_H * total_m / max_v)
                    lbl_t = (f"¥{total_m//1000}k" if is_rec and total_m >= 10000
                             else _fmt_val(total_m))
                    cv.create_text(xc, top_y - 4, text=lbl_t,
                                   font=("Helvetica", 6, "bold"),
                                   fill=COLORS["text_primary"], anchor="s")
                cv.create_text(xc, CHART_H + 14, text=MESES_ABR[i],
                               font=("Helvetica", 6), fill=COLORS["text_muted"], anchor="center")
                _bar_zones.append((x0 - 4, x1 + 4, i))

        def _hide_tip():
            if _tip_win[0]:
                try: _tip_win[0].destroy()
                except Exception: pass
                _tip_win[0] = None; _tip_lbl[0] = None

        def _show_tip(x_root, y_root, mes_idx):
            _hide_tip()
            v_  = dados_v[mes_idx];  os_ = dados_os[mes_idx]; sk_ = dados_sk[mes_idx]
            tot = v_ + os_ + sk_
            if tot == 0: return
            lines = [
                f"── {MESES_ABR[mes_idx]} {ano} ──",
                f"◆ Vendas :  {_fmt_val(v_)}",
                f"⚙ OS     :  {_fmt_val(os_)}",
                f"🔔 Shaken:  {_fmt_val(sk_)}",
                f"▦ Total  :  {_fmt_val(tot)}",
            ]
            tw = tk.Toplevel(cv)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x_root+14}+{y_root-10}")
            tw.configure(bg="#1E293B")
            tk.Label(tw, text="\n".join(lines),
                     font=("Helvetica", 8), justify="left",
                     bg="#1E293B", fg="white", padx=8, pady=6
                     ).pack()
            _tip_win[0] = tw; _tip_lbl[0] = tw

        def _on_motion(event):
            mx = cv.canvasx(event.x)
            for (bx0, bx1, idx) in _bar_zones:
                if bx0 <= mx <= bx1:
                    _show_tip(event.x_root, event.y_root, idx)
                    return
            _hide_tip()

        def _on_leave(event):
            _hide_tip()

        cv.bind("<Configure>", _draw)
        cv.bind("<Motion>",    _on_motion)
        cv.bind("<Leave>",     _on_leave)
        cv.after(60, _draw)

        # ── Totais anuais ────────────────────────────────────────────────────
        tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=14, pady=(2, 0))
        tot_row = tk.Frame(card, bg=COLORS["bg_card"])
        tot_row.pack(fill="x", padx=14, pady=(6, 10))

        def _tot_cell(parent_f, lbl, val, clr):
            f = tk.Frame(parent_f, bg=COLORS["bg_card"])
            f.pack(side="left", expand=True, fill="x")
            tk.Label(f, text=lbl, font=("Helvetica", 7),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(anchor="w")
            tk.Label(f, text=val, font=("Helvetica", 11, "bold"),
                     bg=COLORS["bg_card"], fg=clr).pack(anchor="w")

        fmt = lambda v: f"¥ {v:,}" if is_rec else str(v)
        _tot_cell(tot_row, "◆ Vendas",  fmt(tot_v),   COLOR_V)
        tk.Frame(tot_row, bg=COLORS["border"], width=1).pack(side="left", fill="y", pady=2)
        _tot_cell(tot_row, "⚙ OS",      fmt(tot_os),  COLOR_OS)
        tk.Frame(tot_row, bg=COLORS["border"], width=1).pack(side="left", fill="y", pady=2)
        _tot_cell(tot_row, "🔔 Shaken", fmt(tot_sk),  COLOR_SK)
        tk.Frame(tot_row, bg=COLORS["border"], width=1).pack(side="left", fill="y", pady=2)
        _tot_cell(tot_row, "▦ Total",   fmt(tot_all), COLOR_TOT)

    def _draw_ranking_carros(self, parent):
        """Top 10 carros mais lucrativos + 5 com menor margem — canvas bar chart."""
        def _int(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        # Busca vendas com compra vinculada
        rows = self.conn.execute(
            "SELECT v.carro, v.valor_venda, v.compra_id, v.carro_id "
            "FROM vendas v ORDER BY v.id DESC").fetchall()

        lucros = []
        for r in rows:
            vv = _int(r[1])
            if not vv: continue
            compra_id = r[2]
            carro_id  = r[3]
            cr = None
            if compra_id:
                cr = self.conn.execute(
                    "SELECT * FROM compras WHERE id=?", (compra_id,)).fetchone()
            elif carro_id:
                cr = self.conn.execute(
                    "SELECT * FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                    (carro_id,)).fetchone()
            if cr:
                vc = _int(dict(cr).get("valor"))
                ct = self._get_custo_total(cr[0])
                lucro = vv - vc - ct
                lucros.append((r[0] or "—", lucro, vv))

        if not lucros:
            return

        lucros.sort(key=lambda x: x[1], reverse=True)
        top10  = lucros[:10]
        bot5   = sorted(lucros, key=lambda x: x[1])[:5]
        combined = [("TOP", n, l, v) for n,l,v in top10] + [("BOT", n, l, v) for n,l,v in bot5]

        for side_tag, title, items in [
            ("left",  "🏆  Top 10 — Maior Lucro por Carro", top10),
            ("right", "⚠  5 Piores Margens por Carro",      bot5),
        ]:
            card = tk.Frame(parent, bg=COLORS["bg_card"],
                            highlightthickness=1, highlightbackground=COLORS["border"])
            card.pack(side=side_tag, fill="both", expand=True,
                      padx=(0, 8) if side_tag=="left" else 0)
            bar_c = COLORS["green"] if side_tag=="left" else COLORS["red"]
            tk.Frame(card, bg=bar_c, height=3).pack(fill="x")
            tk.Label(card, text=title, font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                     ).pack(anchor="w", padx=14, pady=(8,4))

            cv_frame = tk.Frame(card, bg=COLORS["bg_card"])
            cv_frame.pack(fill="both", expand=True, padx=14, pady=(0,10))
            BAR_H = 160
            cv = tk.Canvas(cv_frame, bg=COLORS["bg_card"],
                           highlightthickness=0, height=BAR_H+30)
            cv.pack(fill="x")

            def _draw_rank(event=None, _cv=cv, _items=items, _bar_c=bar_c, _BH=BAR_H):
                _cv.delete("all")
                W = _cv.winfo_width()
                if W < 20 or not _items: return
                max_v = max(abs(x[1]) for x in _items) or 1
                n = len(_items)
                pad_l, pad_r = 52, 8
                avail = W - pad_l - pad_r
                gap = avail / n
                bar_w = max(3, int(gap * 0.6))
                for j, (nome, lucro, venda) in enumerate(_items):
                    x_c = pad_l + gap * j + gap / 2
                    h = int(_BH * abs(lucro) / max_v) if max_v else 0
                    x0 = x_c - bar_w/2; x1 = x_c + bar_w/2
                    col = _bar_c if lucro >= 0 else COLORS["red"]
                    if h > 0:
                        _cv.create_rectangle(x0, _BH-h, x1, _BH, fill=col, outline="")
                        val_txt = f"¥{lucro//1000}k" if abs(lucro)>=1000 else f"¥{lucro}"
                        _cv.create_text(x_c, _BH-h-3, text=val_txt,
                                        font=("Helvetica",6,"bold"), fill=col, anchor="s")
                    # Grid lines
                    for pct in [0.5, 1.0]:
                        gy = _BH - int(_BH*pct)
                        _cv.create_line(pad_l, gy, W-pad_r, gy,
                                        fill=COLORS["border"], dash=(2,3))
                        _cv.create_text(pad_l-4, gy,
                                        text=f"¥{int(max_v*pct)//1000}k",
                                        anchor="e", font=("Helvetica",6),
                                        fill=COLORS["text_muted"])
                    # Label carro (truncado)
                    lbl = nome.split("|")[0].strip()[:12]
                    _cv.create_text(x_c, _BH+14, text=lbl,
                                    font=("Helvetica",6), fill=COLORS["text_muted"],
                                    anchor="center")

            cv.bind("<Configure>", _draw_rank)
            cv.after(150, _draw_rank)

    def _get_ultimos_12_meses(self):
        """Retorna lista de (ano, mes) dos últimos 12 meses, mais antigo primeiro."""
        import datetime
        hoje = datetime.date.today()
        meses = []
        for i in range(11, -1, -1):
            m = hoje.month - i
            a = hoje.year
            while m <= 0:
                m += 12; a -= 1
            meses.append((a, m))
        return meses

    def _draw_bar_chart(self, parent, title, mode, bar_color, side="left"):
        """Desenha gráfico de barras de vendas dos últimos 12 meses (Canvas)."""
        import datetime
        meses = self._get_ultimos_12_meses()
        nomes_mes = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

        # Busca dados
        dados = []
        for (a, m) in meses:
            ms = f"{m:02d}"; ys = str(a)
            rows = self.conn.execute(
                "SELECT valor_venda FROM vendas WHERE data_venda LIKE ?",
                (f"%/{ms}/{ys}",)).fetchall()
            if mode == "valor":
                try: val = sum(int(str(r[0] or "0").replace(",","")) for r in rows)
                except: val = 0
            else:
                val = len(rows)
            dados.append((nomes_mes[m-1], val))

        card = tk.Frame(parent, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(side=side, fill="both", expand=True, padx=(0 if side=="right" else 0, 8 if side=="left" else 0))
        tk.Frame(card, bg=bar_color, height=3).pack(fill="x")
        tk.Label(card, text=title, font=("Helvetica", 10, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                 ).pack(anchor="w", padx=14, pady=(10, 4))

        cv_frame = tk.Frame(card, bg=COLORS["bg_card"])
        cv_frame.pack(fill="both", expand=True, padx=14, pady=(0,12))

        CHART_H = 200
        cv = tk.Canvas(cv_frame, bg=COLORS["bg_card"], highlightthickness=0, height=CHART_H+40)
        cv.pack(fill="x")

        def _draw(event=None):
            cv.delete("all")
            W = cv.winfo_width()
            if W < 20:
                return
            n = len(dados)
            max_v = max((d[1] for d in dados), default=1) or 1
            pad_l, pad_r, pad_b = 46, 10, 36
            avail_w = W - pad_l - pad_r
            bar_w   = max(4, int(avail_w / n * 0.6))
            gap     = avail_w / n

            # Grid lines
            for pct in [0.25, 0.5, 0.75, 1.0]:
                y = CHART_H - int(CHART_H * pct)
                cv.create_line(pad_l, y, W-pad_r, y, fill=COLORS["border"], dash=(3,3))
                if mode == "valor":
                    lbl = f"¥{int(max_v*pct):,}"
                else:
                    lbl = str(int(max_v*pct))
                cv.create_text(pad_l-4, y, text=lbl, anchor="e",
                               font=("Helvetica",7), fill=COLORS["text_muted"])

            for j, (mes_lbl, val) in enumerate(dados):
                x_center = pad_l + gap * j + gap / 2
                bar_h    = int(CHART_H * val / max_v) if max_v else 0
                x0 = x_center - bar_w/2
                x1 = x_center + bar_w/2
                y0 = CHART_H - bar_h
                y1 = CHART_H
                if bar_h > 0:
                    cv.create_rectangle(x0, y0, x1, y1, fill=bar_color, outline="")
                    # Valor acima da barra
                    if mode == "valor":
                        vtext = f"¥{val//1000}k" if val >= 1000 else f"¥{val}"
                    else:
                        vtext = str(val)
                    cv.create_text(x_center, y0-3, text=vtext,
                                   font=("Helvetica",7,"bold"), fill=bar_color, anchor="s")
                # Label mês
                cv.create_text(x_center, CHART_H+14, text=mes_lbl,
                               font=("Helvetica",7), fill=COLORS["text_muted"], anchor="center")

        cv.bind("<Configure>", _draw)
        cv.after(100, _draw)

    # ── Dashboard: Estoque (visão cliente) ────────────────────────────────────
    def _build_dash_estoque(self, parent):
        self._dash_est_filtro = tk.StringVar(value="")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        card = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=COLORS["blue"], height=4).pack(fill="x")

        hdr = tk.Frame(card, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(hdr, text="▦  Veículos em Estoque",
                 font=("Helvetica", 12, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        self._lbl_dash_est_total = tk.Label(hdr, text="0 veículos",
                                             font=("Helvetica", 8),
                                             bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_dash_est_total.pack(side="right")

        # Filtro
        fb = tk.Frame(card, bg=COLORS["bg_main"])
        fb.pack(fill="x", padx=16, pady=(0,6))
        tk.Label(fb, text="🔍", font=("Helvetica",9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(fb, textvariable=self._dash_est_filtro, font=("Helvetica",9), width=24,
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"]).pack(side="left", ipady=4)
        self._dash_est_filtro.trace_add("write", lambda *_: self._refresh_dash_estoque())
        tk.Button(fb, text="✕", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=4,
                  command=lambda: self._dash_est_filtro.set("")
                  ).pack(side="left", padx=(3,0), ipady=3)
        tk.Button(fb, text="↺ Atualizar", font=("Helvetica",9,"bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_dash_estoque
                  ).pack(side="right", ipady=3, padx=(0,4))
        tk.Button(fb, text="📄 PDF Cliente", font=("Helvetica",8,"bold"),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  command=self._pdf_estoque_cliente
                  ).pack(side="right", ipady=3, padx=(0,6))

        # Cabeçalho de colunas (sem preço de compra/custo)
        tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16)
        col_f = tk.Frame(card, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=16)
        for txt, w in [("#",3),("Veículo",20),("Ano",5),("Cor",8),("Placa",9),
                       ("Chassi",14),("Preço de Venda",14)]:
            tk.Label(col_f, text=txt, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=6)

        # Scroll
        sf = tk.Frame(card, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=16, pady=(0,12))
        cv = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(sf, orient="vertical", command=cv.yview)
        self._dash_est_rows = tk.Frame(cv, bg=COLORS["bg_card"])
        self._dash_est_rows.bind("<Configure>",
            lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0,0), window=self._dash_est_rows, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        card.bind("<Enter>", lambda e: card.bind_all("<MouseWheel>",
            lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        card.bind("<Leave>", lambda e: card.unbind_all("<MouseWheel>"))

        self._refresh_dash_estoque()

    def _refresh_dash_estoque(self):
        for w in self._dash_est_rows.winfo_children():
            w.destroy()
        termo = self._dash_est_filtro.get().strip().lower()
        veiculos = [dict(r) for r in self.conn.execute(
            "SELECT * FROM compras WHERE a_venda=1 ORDER BY id DESC").fetchall()]
        if termo:
            veiculos = [v for v in veiculos if
                        termo in (v.get("carro") or "").lower() or
                        termo in (v.get("cor")   or "").lower() or
                        termo in (v.get("placa") or "").lower()]
        self._lbl_dash_est_total.configure(text=f"{len(veiculos)} veículo(s)")
        if not veiculos:
            tk.Label(self._dash_est_rows,
                     text="Nenhum veículo em estoque.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=40)
            return
        for i, v in enumerate(veiculos):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._dash_est_rows, bg=rb)
            row.pack(fill="x")
            # Busca dados do carro
            carro_info = {}
            if v.get("carro_id"):
                rc = self.conn.execute("SELECT * FROM carros WHERE id=?", (v["carro_id"],)).fetchone()
                if rc: carro_info = dict(rc)
            carro_txt = v["carro"]
            if " — " in carro_txt: carro_txt = carro_txt.split(" — ",1)[1]
            carro_txt = carro_txt.split("|")[0].strip()
            ano    = carro_info.get("ano")    or v.get("ano")    or "—"
            cor    = carro_info.get("cor")    or v.get("cor")    or "—"
            placa  = carro_info.get("placa")  or v.get("placa")  or "—"
            chassi = carro_info.get("chassi") or "—"
            preco  = self._fmt_yen_display(v.get("preco_venda")) if v.get("preco_venda") else "—"
            preco_fg = COLORS["green"] if v.get("preco_venda") else COLORS["text_muted"]

            tk.Label(row, text=str(v["id"]), font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_muted"], width=3, anchor="w").pack(side="left",padx=2,pady=7)
            tk.Label(row, text=carro_txt, font=("Helvetica",9,"bold"), bg=rb,
                     fg=COLORS["text_primary"], width=20, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=ano,    font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_secondary"], width=5, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=cor,    font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_secondary"], width=8, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=placa,  font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_secondary"], width=9, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=chassi, font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_secondary"], width=14, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=preco,  font=("Helvetica",10,"bold"), bg=rb,
                     fg=preco_fg, width=14, anchor="w").pack(side="left",padx=2)

    def _pdf_estoque_cliente(self):
        """Gera PDF de estoque para cliente: apenas carro, ano, cor e Valor de Venda (preco_avista)."""
        from tkinter import filedialog
        import datetime
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Spacer, Paragraph, HRFlowable)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.units import cm
        except ImportError:
            import tkinter.messagebox as msgbox
            msgbox.showerror("Erro", "reportlab nao instalado. Execute: pip install reportlab")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="Estoque_Cliente.pdf",
            title="Salvar PDF Estoque Cliente")
        if not path:
            return

        # Busca veículos à venda com preco_avista
        veiculos = [dict(r) for r in self.conn.execute(
            "SELECT * FROM compras WHERE a_venda=1 ORDER BY id DESC").fetchall()]

        hoje_str = datetime.date.today().strftime("%d/%m/%Y")
        BLUE = rl_colors.HexColor("#2563EB")
        GRAY = rl_colors.HexColor("#E5E7EB")
        DARK = rl_colors.HexColor("#1F2937")

        styles = getSampleStyleSheet()
        def title_p(txt):
            return Paragraph(f"<b>{txt}</b>",
                             ParagraphStyle("t", parent=styles["Normal"],
                                            fontSize=18, textColor=BLUE, spaceAfter=4))
        def sub_p(txt):
            return Paragraph(txt,
                             ParagraphStyle("s", parent=styles["Normal"],
                                            fontSize=9, textColor=rl_colors.HexColor("#6B7280")))
        def body(txt):
            return Paragraph(txt,
                             ParagraphStyle("b", parent=styles["Normal"],
                                            fontSize=9, textColor=DARK))

        doc = SimpleDocTemplate(path, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        story = []

        story.append(title_p("KM Cars — Estoque de Veículos"))
        story.append(sub_p(f"Gerado em {hoje_str}  ·  {len(veiculos)} veículo(s) disponível(is)"))
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=BLUE))
        story.append(Spacer(1, 0.5*cm))

        # Tabela de veículos
        headers = ["#", "Veículo", "Ano", "Cor", "Shaken Válido até", "Valor de Venda"]
        col_w   = [1*cm, 6*cm, 1.8*cm, 2.5*cm, 3.2*cm, 3.5*cm]

        data = [[Paragraph(f"<b>{h}</b>",
                            ParagraphStyle("h", parent=styles["Normal"],
                                           fontSize=9, textColor=rl_colors.white)) for h in headers]]

        for i, v in enumerate(veiculos):
            # Nome limpo
            carro_txt = v.get("carro","")
            if " — " in carro_txt: carro_txt = carro_txt.split(" — ",1)[1]
            carro_txt = carro_txt.split("|")[0].strip()
            # Busca dados do carro cadastrado (ano, cor)
            ano = "—"; cor = "—"
            if v.get("carro_id"):
                rc = self.conn.execute("SELECT ano, cor FROM carros WHERE id=?",
                                       (v["carro_id"],)).fetchone()
                if rc:
                    ano = rc[0] or "—"
                    cor = rc[1] or "—"
            if ano == "—": ano = v.get("ano") or "—"
            if cor == "—": cor = v.get("cor") or "—"

            # Valor de venda = preco_aprazo (sugerido a prazo)
            preco_ref = v.get("preco_aprazo") or v.get("preco_avista")
            if preco_ref:
                try:
                    val_str = f"¥ {int(str(preco_ref).replace(',','')):,}"
                except:
                    val_str = str(preco_ref)
            else:
                val_str = "A consultar"

            rb_hex = "#F0F7FF" if i % 2 == 0 else "#FFFFFF"
            row_style = ParagraphStyle("r", parent=styles["Normal"],
                                       fontSize=9, textColor=DARK)
            # Shaken do carro — "2 anos" se ausente ou vencido
            import datetime as _pdf_dt
            sk_data_pdf = "2 anos"
            sk_color_pdf = rl_colors.HexColor("#9CA3AF")  # cinza padrão
            if v.get("carro_id"):
                sk_r = self.conn.execute(
                    "SELECT data_vencimento, por_conta FROM shaken "
                    "WHERE carro_id=? AND renovado=0 ORDER BY id DESC LIMIT 1",
                    (v["carro_id"],)).fetchone()
                if sk_r and sk_r[0] and not sk_r[1]:
                    dv_str = sk_r[0]
                    try:
                        _d, _m, _a = dv_str.split("/")
                        _dobj = _pdf_dt.date(int(_a), int(_m), int(_d))
                        if _dobj >= _pdf_dt.date.today():
                            sk_data_pdf  = dv_str
                            sk_color_pdf = rl_colors.HexColor("#0369A1")
                        else:
                            sk_data_pdf  = "2 anos"
                            sk_color_pdf = rl_colors.HexColor("#9CA3AF")
                    except Exception:
                        sk_data_pdf = dv_str
                        sk_color_pdf = rl_colors.HexColor("#0369A1")

            data.append([
                Paragraph(str(i+1), row_style),
                Paragraph(f"<b>{carro_txt}</b>", row_style),
                Paragraph(str(ano), row_style),
                Paragraph(str(cor), row_style),
                Paragraph(sk_data_pdf,
                          ParagraphStyle("sk", parent=styles["Normal"],
                                         fontSize=9, textColor=sk_color_pdf)),
                Paragraph(f"<b>{val_str}</b>",
                          ParagraphStyle("v", parent=styles["Normal"],
                                         fontSize=10, textColor=BLUE)),
            ])

        tbl = Table(data, colWidths=col_w)
        row_bgs = []
        for i in range(1, len(data)):
            c = rl_colors.HexColor("#F0F7FF") if i % 2 == 1 else rl_colors.white
            row_bgs.append(("BACKGROUND", (0,i), (-1,i), c))

        tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), BLUE),
            ("TEXTCOLOR",    (0,0), (-1,0), rl_colors.white),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [rl_colors.HexColor("#F0F7FF"), rl_colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.3, GRAY),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",        (0,0), (0,-1), "CENTER"),
            ("ALIGN",        (5,0), (5,-1), "RIGHT"),
        ]))
        story.append(tbl)

        story.append(Spacer(1, 0.8*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
        story.append(sub_p(f"KM Cars — Estoque Cliente gerado em {hoje_str}"))

        doc.build(story)
        try:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass
        import tkinter.messagebox as msgbox
        msgbox.showinfo("PDF Gerado", f"Estoque Cliente salvo em:\n{path}")

    # ── Dashboard: Vendas do Mês ──────────────────────────────────────────────
    def _build_dash_vendas(self, parent):
        import datetime
        hoje = datetime.date.today()
        self._dash_vd_mes = tk.IntVar(value=hoje.month)
        self._dash_vd_ano = tk.IntVar(value=hoje.year)
        self._dash_vd_filtro = tk.StringVar(value="")
        self._dash_vd_tipo   = tk.StringVar(value="Todos")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Card principal ─────────────────────────────────────────────────────
        card = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=COLORS["green"], height=4).pack(fill="x")

        # Header com KPIs
        hdr = tk.Frame(card, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(hdr, text="◆  Vendas do Mês",
                 font=("Helvetica",12,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_dash_vendas).pack(side="right")
        kpi_row = tk.Frame(hdr, bg=COLORS["bg_card"])
        kpi_row.pack(side="right")
        self._lbl_dash_vd_qtd = tk.Label(kpi_row, text="0 vendas",
                                          font=("Helvetica",9,"bold"),
                                          bg=COLORS["bg_card"], fg=COLORS["green"])
        self._lbl_dash_vd_qtd.pack(side="left", padx=(0,16))
        self._lbl_dash_vd_total = tk.Label(kpi_row, text="¥ 0",
                                            font=("Helvetica",9,"bold"),
                                            bg=COLORS["bg_card"], fg=COLORS["green"])
        self._lbl_dash_vd_total.pack(side="left")

        # Filtros
        fb = tk.Frame(card, bg=COLORS["bg_main"])
        fb.pack(fill="x", padx=16, pady=(0,4))

        tk.Label(fb, text="🔍", font=("Helvetica",9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(fb, textvariable=self._dash_vd_filtro, font=("Helvetica",9), width=16,
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"]).pack(side="left", ipady=4)
        self._dash_vd_filtro.trace_add("write", lambda *_: self._refresh_dash_vendas())
        tk.Button(fb, text="✕", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=3,
                  command=lambda: self._dash_vd_filtro.set("")
                  ).pack(side="left", padx=(2,10), ipady=3)

        # Tipo
        tk.Label(fb, text="Tipo:", font=("Helvetica",8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fb, textvariable=self._dash_vd_tipo,
                     values=["Todos","Venda a Vista","Venda Parcelada","Com Troca","Com Troca e Volta"],
                     state="readonly", font=("Helvetica",8), width=14
                     ).pack(side="left", padx=(3,12), ipady=2)
        self._dash_vd_tipo.trace_add("write", lambda *_: self._refresh_dash_vendas())

        # Navegação de mês ← Mês/Ano →
        nav = tk.Frame(fb, bg=COLORS["bg_main"])
        nav.pack(side="left")
        tk.Button(nav, text="◀", font=("Helvetica",10,"bold"),
                  bg=COLORS["border"], fg=COLORS["text_primary"],
                  relief="flat", cursor="hand2", padx=6,
                  command=self._dash_vd_prev_mes
                  ).pack(side="left", ipady=2)
        self._lbl_dash_vd_mes = tk.Label(nav, text="",
                                          font=("Helvetica",10,"bold"),
                                          bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                          width=14, anchor="center")
        self._lbl_dash_vd_mes.pack(side="left", padx=6)
        tk.Button(nav, text="▶", font=("Helvetica",10,"bold"),
                  bg=COLORS["border"], fg=COLORS["text_primary"],
                  relief="flat", cursor="hand2", padx=6,
                  command=self._dash_vd_next_mes
                  ).pack(side="left", ipady=2)

        # Cabeçalho da tabela — idêntico ao histórico de vendas
        tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16)
        col_f = tk.Frame(card, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=16)
        for txt, w in [("#",3),("Cliente",13),("Data",9),("Carro",13),("Tipo",11),
                       ("V.Venda",9),("Entrada",8),("Parcela",8),("Saldo",9),("Status",8)]:
            tk.Label(col_f, text=txt, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=6)

        # Scroll
        sf = tk.Frame(card, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=16, pady=(0,12))
        cv2 = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb2 = tk.Scrollbar(sf, orient="vertical", command=cv2.yview)
        self._dash_vd_rows = tk.Frame(cv2, bg=COLORS["bg_card"])
        self._dash_vd_rows.bind("<Configure>",
            lambda e: cv2.configure(scrollregion=cv2.bbox("all")))
        cv2.create_window((0,0), window=self._dash_vd_rows, anchor="nw")
        cv2.configure(yscrollcommand=sb2.set)
        cv2.pack(side="left", fill="both", expand=True)
        sb2.pack(side="right", fill="y")
        card.bind("<Enter>", lambda e: card.bind_all("<MouseWheel>",
            lambda ev: cv2.yview_scroll(int(-1*(ev.delta/120)),"units")))
        card.bind("<Leave>", lambda e: card.unbind_all("<MouseWheel>"))

        self._refresh_dash_vendas()

    def _dash_vd_prev_mes(self):
        m = self._dash_vd_mes.get() - 1
        a = self._dash_vd_ano.get()
        if m < 1: m = 12; a -= 1
        self._dash_vd_mes.set(m); self._dash_vd_ano.set(a)
        self._refresh_dash_vendas()

    def _dash_vd_next_mes(self):
        m = self._dash_vd_mes.get() + 1
        a = self._dash_vd_ano.get()
        if m > 12: m = 1; a += 1
        self._dash_vd_mes.set(m); self._dash_vd_ano.set(a)
        self._refresh_dash_vendas()

    def _refresh_dash_vendas(self):
        import calendar
        MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
        for w in self._dash_vd_rows.winfo_children():
            w.destroy()

        m = self._dash_vd_mes.get(); a = self._dash_vd_ano.get()
        ms = f"{m:02d}"; ys = str(a)
        self._lbl_dash_vd_mes.configure(text=f"{MESES_PT[m-1]} {a}")

        lista = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id=c.id "
            "WHERE v.data_venda LIKE ? ORDER BY v.id DESC",
            (f"%/{ms}/{ys}",)).fetchall()]

        # Filtro texto
        termo = self._dash_vd_filtro.get().strip().lower()
        if termo:
            lista = [v for v in lista if
                     termo in (v.get("carro") or "").lower() or
                     termo in (v.get("cliente_nome") or "").lower()]
        # Filtro tipo
        ftipo = self._dash_vd_tipo.get()
        if ftipo != "Todos":
            lista = [v for v in lista if v.get("tipo_venda") == ftipo]

        # KPIs
        try:
            total_val = sum(int(str(v["valor_venda"] or "0").replace(",","")) for v in lista)
        except Exception:
            total_val = 0
        self._lbl_dash_vd_qtd.configure(text=f"{len(lista)} venda(s)")
        self._lbl_dash_vd_total.configure(text=f"¥ {total_val:,}")

        if not lista:
            tk.Label(self._dash_vd_rows,
                     text="Nenhuma venda neste mês.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        TIPO_COLORS = {"Venda a Vista":COLORS["green"],"Venda Parcelada":COLORS["blue"],
                       "Com Troca":COLORS["orange"],"Com Troca e Volta":COLORS["accent"]}
        TIPO_SHORT  = {"Venda a Vista":"À Vista","Venda Parcelada":"Parcelada",
                       "Com Troca":"C/Troca","Com Troca e Volta":"T+Volta"}

        for i, v in enumerate(lista):
            tipo   = v.get("tipo_venda","")
            pagas  = v.get("parcelas_pagas") or 0
            tot_p  = v.get("num_parcelas") or 0
            is_p   = tipo in ("Venda Parcelada","Com Troca","Com Troca e Volta")
            quitado = is_p and tot_p > 0 and pagas >= tot_p
            try:
                tp = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                    "FROM pagamentos WHERE venda_id=?", (v["id"],)).fetchone()[0]
                vv = int(str(v.get("valor_venda") or "0").replace(",",""))
                et = int(str(v.get("entrada") or "0").replace(",",""))
                vt = int(str(v.get("valor_troca") or "0").replace(",",""))
                saldo_rest = max(0, vv-et-vt-tp)
            except Exception:
                saldo_rest = 0

            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._dash_vd_rows, bg=rb)
            row.pack(fill="x")

            tk.Label(row, text=str(v["id"]), font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_muted"], width=3, anchor="w").pack(side="left",padx=2,pady=6)
            tk.Label(row, text=(v.get("cliente_nome") or "—")[:13],
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_primary"],
                     width=13, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=v.get("data_venda") or "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=9, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=(v.get("carro") or "—")[:13],
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_primary"],
                     width=13, anchor="w").pack(side="left",padx=2)
            tc = TIPO_COLORS.get(tipo, COLORS["text_muted"])
            tk.Label(row, text=TIPO_SHORT.get(tipo, tipo),
                     font=("Helvetica",7,"bold"), bg=tc, fg="white",
                     width=11, anchor="center").pack(side="left",padx=2,pady=3)
            tk.Label(row, text=self._fmt_yen_display(v.get("valor_venda")),
                     font=("Helvetica",9), bg=rb, fg=COLORS["green"],
                     width=9, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=self._fmt_yen_display(v.get("entrada")) if v.get("entrada") else "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"],
                     width=8, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=self._fmt_yen_display(v.get("parcela_mensal")) if is_p else "—",
                     font=("Helvetica",9), bg=rb, fg=COLORS["blue"],
                     width=8, anchor="w").pack(side="left",padx=2)
            if is_p and not quitado:
                sd, sf2 = f"¥{saldo_rest:,}", COLORS["orange"]
            elif is_p:
                sd, sf2 = "Quitado", COLORS["green"]
            else:
                sd, sf2 = "—", COLORS["text_muted"]
            tk.Label(row, text=sd, font=("Helvetica",9), bg=rb, fg=sf2,
                     width=9, anchor="w").pack(side="left",padx=2)
            st_txt = "À Vista" if not is_p else ("Quitado" if quitado else "Em Aberto")
            st_bg  = COLORS["green"] if (not is_p or quitado) else COLORS["blue"]
            tk.Label(row, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_bg, fg="white", width=8, anchor="center").pack(side="left",padx=2,pady=3)

    # ── Dashboard: Serviços ───────────────────────────────────────────────────
    def _build_dash_servicos(self, parent):
        import datetime
        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # Top: tabela + gráfico lado a lado
        top = tk.Frame(root, bg=COLORS["bg_content"])
        top.pack(fill="both", expand=True)

        # ── Coluna esquerda: lista serviços em aberto ──────────────────────────
        left = tk.Frame(top, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        left.pack(side="left", fill="both", expand=True, padx=(0,8))
        tk.Frame(left, bg=COLORS["orange"], height=4).pack(fill="x")

        hdr_l = tk.Frame(left, bg=COLORS["bg_card"])
        hdr_l.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(hdr_l, text="⚙  Serviços em Aberto",
                 font=("Helvetica",11,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr_l, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_dash_servicos).pack(side="right", padx=(8,0))
        self._lbl_dash_serv_total = tk.Label(hdr_l, text="",
                                              font=("Helvetica",9,"bold"),
                                              bg=COLORS["bg_card"], fg=COLORS["orange"])
        self._lbl_dash_serv_total.pack(side="right")

        tk.Frame(left, bg=COLORS["border"], height=1).pack(fill="x", padx=16)
        col_f = tk.Frame(left, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=16)
        for txt, w in [("OS",7),("Data",10),("Carro",16),("Cliente",14),
                       ("Tipo",12),("Valor",10),("Status",10)]:
            tk.Label(col_f, text=txt, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left",padx=2,pady=5)

        sf = tk.Frame(left, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=16, pady=(0,12))
        cv_s = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb_s = tk.Scrollbar(sf, orient="vertical", command=cv_s.yview)
        self._dash_serv_rows = tk.Frame(cv_s, bg=COLORS["bg_card"])
        self._dash_serv_rows.bind("<Configure>",
            lambda e: cv_s.configure(scrollregion=cv_s.bbox("all")))
        cv_s.create_window((0,0), window=self._dash_serv_rows, anchor="nw")
        cv_s.configure(yscrollcommand=sb_s.set)
        cv_s.pack(side="left", fill="both", expand=True)
        sb_s.pack(side="right", fill="y")

        # ── Coluna direita: KPI + gráfico ──────────────────────────────────────
        right = tk.Frame(top, bg=COLORS["bg_content"])
        right.pack(side="right", fill="both", expand=True)

        # KPI mês atual
        hoje = datetime.date.today()
        ms = f"{hoje.month:02d}"; ys = str(hoje.year)
        MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
        try:
            total_mes = self.conn.execute(
                "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                "FROM servicos WHERE data_servico LIKE ? AND status='Concluído'",
                (f"%/{ms}/{ys}",)).fetchone()[0]
        except Exception:
            total_mes = 0
        kpi_card = tk.Frame(right, bg=COLORS["bg_card"],
                            highlightthickness=1, highlightbackground=COLORS["border"])
        kpi_card.pack(fill="x", pady=(0,10))
        tk.Frame(kpi_card, bg=COLORS["orange"], height=3).pack(fill="x")
        tk.Label(kpi_card, text=f"¥ {total_mes:,}",
                 font=("Helvetica",24,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["orange"]).pack(pady=(10,2))
        tk.Label(kpi_card, text=f"Total de serviços concluídos — {MESES_PT[hoje.month-1]} {hoje.year}",
                 font=("Helvetica",8),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(0,10))

        self._dash_serv_kpi_card = kpi_card
        self._dash_serv_kpi_lbl  = kpi_card.winfo_children()[1]  # label ¥ valor

        # ── helper genérico: cria card gráfico com tooltip ────────────────────
        def _make_chart_card(parent_frame, title, bar_color, query_fn, height=150):
            """Cria card com KPI + gráfico de barras + tooltip hover."""
            cc = tk.Frame(parent_frame, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            cc.pack(fill="x", pady=(0, 8))
            tk.Frame(cc, bg=bar_color, height=3).pack(fill="x")
            tk.Label(cc, text=title, font=("Helvetica",10,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                     ).pack(anchor="w", padx=14, pady=(10,4))

            cv = tk.Canvas(cc, bg=COLORS["bg_card"], highlightthickness=0, height=height+28)
            cv.pack(fill="x", padx=14, pady=(0,10))

            _tip  = [None]
            _zones = []

            def _hide_tip():
                if _tip[0]:
                    try: _tip[0].destroy()
                    except Exception: pass
                    _tip[0] = None

            def _show_tip(xr, yr, lbl, val):
                _hide_tip()
                if val == 0: return
                tw = tk.Toplevel(cv)
                tw.wm_overrideredirect(True)
                tw.wm_geometry(f"+{xr+12}+{yr-10}")
                tw.configure(bg="#1E293B")
                text = f"{lbl}\n¥ {val:,}"
                tk.Label(tw, text=text, font=("Helvetica",8),
                         justify="left", bg="#1E293B", fg="white",
                         padx=8, pady=5).pack()
                _tip[0] = tw

            def _load_and_draw():
                meses12 = self._get_ultimos_12_meses()
                nomes_m = ["Jan","Fev","Mar","Abr","Mai","Jun",
                           "Jul","Ago","Set","Out","Nov","Dez"]
                dados = []
                for (a, m2) in meses12:
                    ms2 = f"{m2:02d}"; ys2 = str(a)
                    val = query_fn(ms2, ys2)
                    dados.append((nomes_m[m2-1], int(val or 0)))

                def _draw(event=None):
                    cv.delete("all")
                    _zones.clear()
                    W = cv.winfo_width()
                    if W < 20: return
                    CH = height
                    max_v = max((d[1] for d in dados), default=1) or 1
                    n = len(dados)
                    pad_l, pad_r = 52, 10
                    avail_w = W - pad_l - pad_r
                    bar_w = max(4, int(avail_w / n * 0.6))
                    gap   = avail_w / n
                    for pct in [0.25, 0.5, 0.75, 1.0]:
                        y = CH - int(CH * pct)
                        cv.create_line(pad_l, y, W-pad_r, y,
                                       fill=COLORS["border"], dash=(3,3))
                        cv.create_text(pad_l-4, y,
                                       text=f"¥{int(max_v*pct)//1000}k" if max_v>=1000 else str(int(max_v*pct)),
                                       anchor="e", font=("Helvetica",7), fill=COLORS["text_muted"])
                    for j, (lbl, val) in enumerate(dados):
                        xc = pad_l + gap*j + gap/2
                        bh = int(CH * val / max_v) if max_v else 0
                        x0b = xc - bar_w/2; x1b = xc + bar_w/2
                        if bh > 0:
                            cv.create_rectangle(x0b, CH-bh, x1b, CH,
                                                fill=bar_color, outline="")
                            vt = f"¥{val//1000}k" if val>=1000 else f"¥{val}"
                            cv.create_text(xc, CH-bh-3, text=vt,
                                           font=("Helvetica",7,"bold"),
                                           fill=bar_color, anchor="s")
                        cv.create_text(xc, CH+14, text=lbl,
                                       font=("Helvetica",7), fill=COLORS["text_muted"])
                        _zones.append((x0b-4, x1b+4, lbl, val))

                def _on_motion(event):
                    mx = cv.canvasx(event.x)
                    for (x0z, x1z, lbl, val) in _zones:
                        if x0z <= mx <= x1z:
                            _show_tip(event.x_root, event.y_root, lbl, val)
                            return
                    _hide_tip()

                cv.bind("<Configure>", _draw)
                cv.bind("<Motion>",    _on_motion)
                cv.bind("<Leave>",     lambda e: _hide_tip())
                cv.after(10, _draw)

            _load_and_draw()
            return cc, _load_and_draw

        # ── Gráfico OS ────────────────────────────────────────────────────────
        def _q_serv(ms2, ys2):
            try:
                return self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM servicos WHERE data_servico LIKE ? AND status='Concluido'",
                    (f"%/{ms2}/{ys2}",)).fetchone()[0]
            except Exception:
                return 0

        def _q_serv_real(ms2, ys2):
            try:
                return self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM servicos WHERE data_servico LIKE ?",
                    (f"%/{ms2}/{ys2}",)).fetchone()[0]
            except Exception:
                return 0

        _cc_os, _load_os = _make_chart_card(
            right,
            "⚙  Receita de Serviços — últimos 12 meses",
            COLORS["orange"],
            _q_serv_real,
            height=150,
        )

        # ── KPI Shaken mês atual ──────────────────────────────────────────────
        try:
            total_sk_mes = self.conn.execute(
                "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                "FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL",
                (f"%/{ms}/{ys}",)).fetchone()[0]
        except Exception:
            total_sk_mes = 0

        kpi_sk = tk.Frame(right, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
        kpi_sk.pack(fill="x", pady=(0, 8))
        tk.Frame(kpi_sk, bg="#D4AC0D", height=3).pack(fill="x")
        self._dash_sk_kpi_lbl = tk.Label(kpi_sk, text=f"¥ {int(total_sk_mes or 0):,}",
                 font=("Helvetica",22,"bold"),
                 bg=COLORS["bg_card"], fg="#D4AC0D")
        self._dash_sk_kpi_lbl.pack(pady=(10,2))
        tk.Label(kpi_sk, text=f"Total Shaken — {MESES_PT[hoje.month-1]} {hoje.year}",
                 font=("Helvetica",8),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(0,10))

        # ── Gráfico Shaken ────────────────────────────────────────────────────
        def _q_shaken(ms2, ys2):
            try:
                return self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL",
                    (f"%/{ms2}/{ys2}",)).fetchone()[0]
            except Exception:
                return 0

        _cc_sk, _load_sk = _make_chart_card(
            right,
            "🔔  Receita de Shaken — últimos 12 meses",
            "#D4AC0D",
            _q_shaken,
            height=150,
        )

        # ── Guarda referências para atualização pelo botão ────────────────────
        self._dash_serv_load_chart = lambda: (_load_os(), _load_sk())

        self._refresh_dash_servicos()

    def _refresh_dash_servicos(self):
        # ── Actualiza gráfico de 12 meses ────────────────────────────────────
        if hasattr(self, "_dash_serv_load_chart"):
            try: self._dash_serv_load_chart()
            except Exception: pass
        # ── Actualiza KPI cards do mês atual ─────────────────────────────────
        if hasattr(self, "_dash_serv_kpi_lbl"):
            try:
                import datetime as _dts
                _h = _dts.date.today()
                _ms = f"{_h.month:02d}"; _ys = str(_h.year)
                _tm = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM servicos WHERE data_servico LIKE ?",
                    (f"%/{_ms}/{_ys}",)).fetchone()[0]
                self._dash_serv_kpi_lbl.configure(text=f"¥ {int(_tm or 0):,}")
            except Exception: pass
        if hasattr(self, "_dash_sk_kpi_lbl"):
            try:
                import datetime as _dts2
                _h2 = _dts2.date.today()
                _ms2 = f"{_h2.month:02d}"; _ys2 = str(_h2.year)
                _tsk = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL",
                    (f"%/{_ms2}/{_ys2}",)).fetchone()[0]
                self._dash_sk_kpi_lbl.configure(text=f"¥ {int(_tsk or 0):,}")
            except Exception: pass

        for w in self._dash_serv_rows.winfo_children():
            w.destroy()
        servicos = [dict(r) for r in self.conn.execute(
            "SELECT s.*, c.nome as cliente_nome FROM servicos s "
            "LEFT JOIN clientes c ON s.cliente_id=c.id "
            "WHERE s.status IN ('Aberto','Em Andamento') "
            "ORDER BY s.id DESC").fetchall()]

        STATUS_COLORS = {
            "Aberto":       COLORS["blue"],
            "Em Andamento": COLORS["orange"],
        }
        try:
            total_aberto = sum(
                int(str(s.get("valor") or "0").replace(",","")) for s in servicos)
        except Exception:
            total_aberto = 0
        self._lbl_dash_serv_total.configure(
            text=f"{len(servicos)} serviço(s)  •  ¥ {total_aberto:,}")

        if not servicos:
            tk.Label(self._dash_serv_rows,
                     text="Nenhum serviço em aberto.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        for i, s in enumerate(servicos):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._dash_serv_rows, bg=rb)
            row.pack(fill="x")
            sc = STATUS_COLORS.get(s.get("status",""), COLORS["text_muted"])
            tk.Label(row, text=s.get("os_num") or f"#{s['id']}",
                     font=("Helvetica",8,"bold"), bg=rb,
                     fg=COLORS["accent"], width=7, anchor="w").pack(side="left",padx=2,pady=6)
            tk.Label(row, text=s.get("data_servico") or "—",
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], width=10, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=(s.get("carro") or "—")[:16],
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_primary"], width=16, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=(s.get("cliente_nome") or "—")[:14],
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], width=14, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=(s.get("tipo_servico") or "—")[:12],
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_muted"], width=12, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=self._fmt_yen_display(s.get("valor")),
                     font=("Helvetica",8,"bold"), bg=rb,
                     fg=COLORS["green"], width=10, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=s.get("status",""),
                     font=("Helvetica",7,"bold"), bg=sc,
                     fg="white", width=10, anchor="center").pack(side="left",padx=2,pady=3)

    # ── Dashboard: Parcelas Vencidas ──────────────────────────────────────────
    def _build_dash_parcelas(self, parent):
        import datetime
        _hoje2 = datetime.date.today()
        self._dash_parc_filtro  = tk.StringVar(value="")
        self._dash_parc_status  = tk.StringVar(value="Em Aberto")
        self._parc_nav_mes      = tk.IntVar(value=_hoje2.month)
        self._parc_nav_ano      = tk.IntVar(value=_hoje2.year)

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        card = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=COLORS["red"], height=4).pack(fill="x")

        hdr = tk.Frame(card, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(hdr, text="◎  Parcelas — Visão Geral",
                 font=("Helvetica",12,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        self._lbl_dash_parc_total = tk.Label(hdr, text="",
                                              font=("Helvetica",8),
                                              bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_dash_parc_total.pack(side="right")

        # Filtros
        fb = tk.Frame(card, bg=COLORS["bg_main"])
        fb.pack(fill="x", padx=16, pady=(0,4))
        tk.Label(fb, text="🔍", font=("Helvetica",9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left",padx=(0,4),pady=5)
        tk.Entry(fb, textvariable=self._dash_parc_filtro, font=("Helvetica",9), width=16,
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"]).pack(side="left", ipady=4)
        self._dash_parc_filtro.trace_add("write", lambda *_: self._refresh_dash_parcelas())
        tk.Button(fb, text="✕", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=3,
                  command=lambda: self._dash_parc_filtro.set("")
                  ).pack(side="left", padx=(2,10), ipady=3)
        tk.Label(fb, text="Status:", font=("Helvetica",8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left")
        ttk.Combobox(fb, textvariable=self._dash_parc_status,
                     values=["Todos","Em Aberto","Quitado"],
                     state="readonly", font=("Helvetica",8), width=9
                     ).pack(side="left", padx=(2,0), ipady=2)
        self._dash_parc_status.trace_add("write", lambda *_: self._refresh_dash_parcelas())
        tk.Button(fb, text="↺ Atualizar", font=("Helvetica",9,"bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_dash_parcelas
                  ).pack(side="right", ipady=3, padx=(0,4))

        # ── Navegação de mês ──────────────────────────────────────────────────
        _MESES_PT2 = ["Jan","Fev","Mar","Abr","Mai","Jun",
                      "Jul","Ago","Set","Out","Nov","Dez"]

        def _parc_prev():
            m = self._parc_nav_mes.get(); a = self._parc_nav_ano.get()
            m -= 1
            if m < 1: m = 12; a -= 1
            self._parc_nav_mes.set(m); self._parc_nav_ano.set(a)
            self._lbl_parc_mes.configure(text=f"{_MESES_PT2[m-1]} {a}")
            self._build_parc_kpis(); self._refresh_dash_parcelas()

        def _parc_next():
            m = self._parc_nav_mes.get(); a = self._parc_nav_ano.get()
            m += 1
            if m > 12: m = 1; a += 1
            self._parc_nav_mes.set(m); self._parc_nav_ano.set(a)
            self._lbl_parc_mes.configure(text=f"{_MESES_PT2[m-1]} {a}")
            self._build_parc_kpis(); self._refresh_dash_parcelas()

        nav_parc = tk.Frame(card, bg=COLORS["bg_card"])
        nav_parc.pack(fill="x", padx=16, pady=(0,2))
        tk.Button(nav_parc, text="◀", font=("Helvetica",10,"bold"),
                  bg=COLORS["bg_main"], fg=COLORS["red"], relief="flat",
                  cursor="hand2", padx=6, command=_parc_prev
                  ).pack(side="left", ipady=2)
        _pm0 = self._parc_nav_mes.get(); _pa0 = self._parc_nav_ano.get()
        self._lbl_parc_mes = tk.Label(nav_parc,
                  text=f"{_MESES_PT2[_pm0-1]} {_pa0}",
                  font=("Helvetica",10,"bold"), width=12,
                  bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._lbl_parc_mes.pack(side="left", padx=6)
        tk.Button(nav_parc, text="▶", font=("Helvetica",10,"bold"),
                  bg=COLORS["bg_main"], fg=COLORS["red"], relief="flat",
                  cursor="hand2", padx=6, command=_parc_next
                  ).pack(side="left", ipady=2)
        tk.Label(nav_parc, text="← filtra parcelas do mês",
                 font=("Helvetica",7), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(side="left", padx=8)

        # ── KPIs do mês ──────────────────────────────────────────────────────
        kpi_parc = tk.Frame(card, bg=COLORS["bg_card"])
        kpi_parc.pack(fill="x", padx=16, pady=(2,6))
        self._parc_kpi_frame = kpi_parc
        self._build_parc_kpis()

        # Cabeçalho de colunas — igual ao Financiamentos Ativos sem ações
        FIN_COLS = [
            ("Cliente",140,"w"),("Carro",120,"w"),("Tipo",78,"c"),
            ("V.Venda",82,"w"),("Entrada",72,"w"),("Parcela",72,"w"),
            ("Progresso",90,"w"),("Próx.Parc",86,"w"),("Saldo",86,"w"),("Status",72,"c"),
        ]
        self._dash_parc_cols = FIN_COLS
        PARC_TOTAL_W = sum(c[1] for c in FIN_COLS)

        # ── Área de tabela com scroll H+V sincronizado ────────────────────────
        table_area = tk.Frame(card, bg=COLORS["bg_card"])
        table_area.pack(fill="both", expand=True, padx=16, pady=(0,12))

        vsb = tk.Scrollbar(table_area, orient="vertical")
        hsb = tk.Scrollbar(table_area, orient="horizontal")

        self._parc_hdr_canvas  = tk.Canvas(table_area, bg=COLORS["bg_main"],
                                            height=28, highlightthickness=0)
        self._parc_body_canvas = tk.Canvas(table_area, bg=COLORS["bg_card"],
                                            highlightthickness=0,
                                            yscrollcommand=vsb.set)

        def _parc_sync_x(*args):
            self._parc_hdr_canvas.xview(*args)
            self._parc_body_canvas.xview(*args)
        hsb.config(command=_parc_sync_x)
        vsb.config(command=self._parc_body_canvas.yview)

        self._parc_hdr_canvas.grid(row=0, column=0, sticky="ew")
        self._parc_body_canvas.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        table_area.grid_rowconfigure(1, weight=1)
        table_area.grid_columnconfigure(0, weight=1)

        # Header frame
        parc_hdr_frame = tk.Frame(self._parc_hdr_canvas, bg=COLORS["bg_main"])
        parc_hdr_win = self._parc_hdr_canvas.create_window((0,0), window=parc_hdr_frame, anchor="nw")
        for name, pw, anch in FIN_COLS:
            f = tk.Frame(parc_hdr_frame, bg=COLORS["bg_main"], width=pw, height=28)
            f.pack_propagate(False); f.pack(side="left")
            tk.Label(f, text=name, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w" if anch=="w" else "center").pack(fill="both", padx=4)
        tk.Frame(parc_hdr_frame, bg=COLORS["border"], height=1, width=PARC_TOTAL_W).pack(fill="x")

        # Body frame
        self._dash_parc_rows = tk.Frame(self._parc_body_canvas, bg=COLORS["bg_card"])
        parc_body_win = self._parc_body_canvas.create_window((0,0), window=self._dash_parc_rows, anchor="nw")
        self._dash_parc_rows.bind("<Configure>",
            lambda e: self._parc_body_canvas.configure(scrollregion=self._parc_body_canvas.bbox("all")))

        def _parc_canvas_resize(e, _bw=parc_body_win, _hw=parc_hdr_win):
            w = max(e.width, PARC_TOTAL_W)
            self._parc_body_canvas.itemconfig(_bw, width=w)
            self._parc_hdr_canvas.itemconfig(_hw, width=w)
            self._parc_hdr_canvas.configure(scrollregion=(0,0,w,28))
        self._parc_body_canvas.bind("<Configure>", _parc_canvas_resize)

        def _parc_xmove(first, last):
            hsb.set(first, last)
            self._parc_hdr_canvas.xview_moveto(first)
        self._parc_body_canvas.configure(xscrollcommand=_parc_xmove)

        def _parc_mwheel(ev):
            self._parc_body_canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        card.bind("<Enter>", lambda e: card.bind_all("<MouseWheel>", _parc_mwheel))
        card.bind("<Leave>", lambda e: card.unbind_all("<MouseWheel>"))

        self._refresh_dash_parcelas()

    def _refresh_dash_parcelas(self):
        import datetime
        for w in self._dash_parc_rows.winfo_children():
            w.destroy()

        # Mês/ano do navegador
        nav_m  = getattr(self, "_parc_nav_mes", None)
        nav_a  = getattr(self, "_parc_nav_ano", None)
        nav_ms = f"{nav_m.get():02d}" if nav_m else None
        nav_ys = str(nav_a.get())      if nav_a else None

        todas = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id=c.id "
            "WHERE v.tipo_venda IN ('Venda Parcelada','Com Troca','Com Troca e Volta') "
            "ORDER BY v.id DESC").fetchall()]

        termo = self._dash_parc_filtro.get().strip().lower()
        if termo:
            todas = [v for v in todas if
                     termo in (v.get("carro") or "").lower() or
                     termo in (v.get("cliente_nome") or "").lower()]

        # Filtro do navegador: mantém vendas com parcela no mês/ano navegado
        if nav_ms and nav_ys:
            def tem_parcela_no_mes_p(v):
                total_parc = v.get("num_parcelas") or 0
                for n in range(total_parc):
                    prx = self._proxima_parcela(v.get("data_primeira_parc"), n)
                    if prx == "—": continue
                    pts = prx.split("/")
                    if len(pts) >= 3 and pts[1] == nav_ms and pts[2] == nav_ys:
                        return True
                return False
            todas = [v for v in todas if tem_parcela_no_mes_p(v)]

        fst = self._dash_parc_status.get()
        if fst == "Em Aberto":
            todas = [v for v in todas if
                     (v.get("parcelas_pagas") or 0) < (v.get("num_parcelas") or 0)]
        elif fst == "Quitado":
            todas = [v for v in todas if
                     (v.get("num_parcelas") or 0) > 0 and
                     (v.get("parcelas_pagas") or 0) >= (v.get("num_parcelas") or 0)]

        self._lbl_dash_parc_total.configure(text=f"{len(todas)} financiamento(s)")

        if not todas:
            tk.Label(self._dash_parc_rows,
                     text="Nenhum financiamento encontrado.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).grid(row=0, column=0, pady=40)
            return

        FIN_COLS = getattr(self, "_dash_parc_cols", [
            ("Cliente",140,"w"),("Carro",120,"w"),("Tipo",78,"c"),
            ("V.Venda",82,"w"),("Entrada",72,"w"),("Parcela",72,"w"),
            ("Progresso",90,"w"),("Prox.Parc",86,"w"),("Saldo",86,"w"),("Status",72,"c"),
        ])
        for ci, (_, pw, _a) in enumerate(FIN_COLS):
            self._dash_parc_rows.grid_columnconfigure(ci, minsize=pw, weight=0)

        TIPO_COLORS = {"Venda Parcelada": COLORS["blue"],
                       "Com Troca": COLORS["orange"],
                       "Com Troca e Volta": COLORS["accent"]}
        TIPO_SHORT  = {"Venda Parcelada": "Parcelada",
                       "Com Troca": "C/Troca",
                       "Com Troca e Volta": "T+Volta"}

        for i, v in enumerate(todas):
            pagas      = v.get("parcelas_pagas") or 0
            total_parc = v.get("num_parcelas") or 0
            quitado    = total_parc > 0 and pagas >= total_parc

            try:
                total_pago = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                    "FROM pagamentos WHERE venda_id=?", (v["id"],)).fetchone()[0]
                vv = int(str(v.get("valor_venda") or "0").replace(",",""))
                ev = int(str(v.get("entrada") or "0").replace(",",""))
                vt = int(str(v.get("valor_troca") or "0").replace(",",""))
                saldo_rest = max(0, vv - ev - vt - total_pago)
            except Exception:
                saldo_rest = 0

            # Detecta se a parcela do mês navegado foi paga
            pago_no_mes = False
            if nav_ms and nav_ys and not quitado:
                for n in range(total_parc):
                    prx = self._proxima_parcela(v.get("data_primeira_parc"), n)
                    if prx == "—": continue
                    pts = prx.split("/")
                    if len(pts) >= 3 and pts[1] == nav_ms and pts[2] == nav_ys:
                        p_row = self.conn.execute(
                            "SELECT id FROM pagamentos WHERE venda_id=? AND num_parcela=?",
                            (v["id"], n + 1)).fetchone()
                        if p_row:
                            pago_no_mes = True
                        break

            rb  = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            bdr = COLORS["green"] if (quitado or pago_no_mes) else COLORS["border"]

            tk.Frame(self._dash_parc_rows, bg=bdr, height=1
                     ).grid(row=i*2, column=0, columnspan=len(FIN_COLS), sticky="ew")
            ri = i*2 + 1
            self._dash_parc_rows.grid_rowconfigure(ri, minsize=36)

            tk.Label(self._dash_parc_rows,
                     text=(v.get("cliente_nome") or "\u2014")[:18],
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_primary"],
                     anchor="w", padx=4).grid(row=ri, column=0, sticky="nsew")

            car = v.get("carro") or "\u2014"
            car = car[:15]+"\u2026" if len(car) > 16 else car
            tk.Label(self._dash_parc_rows, text=car,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=1, sticky="nsew")

            tc = TIPO_COLORS.get(v["tipo_venda"], COLORS["text_muted"])
            f2 = tk.Frame(self._dash_parc_rows, bg=rb)
            f2.grid(row=ri, column=2, sticky="nsew")
            tk.Label(f2, text=TIPO_SHORT.get(v["tipo_venda"], "\u2014"),
                     font=("Helvetica",7,"bold"), bg=tc, fg="white", anchor="center"
                     ).pack(padx=3, pady=6, fill="both", expand=True)

            tk.Label(self._dash_parc_rows,
                     text=self._fmt_yen_display(v.get("valor_venda")),
                     font=("Helvetica",8,"bold"), bg=rb, fg=COLORS["green"],
                     anchor="w", padx=4).grid(row=ri, column=3, sticky="nsew")

            ent_txt = self._fmt_yen_display(v.get("entrada")) if v.get("entrada") else "\u2014"
            tk.Label(self._dash_parc_rows, text=ent_txt,
                     font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_secondary"], anchor="w", padx=4
                     ).grid(row=ri, column=4, sticky="nsew")

            tk.Label(self._dash_parc_rows,
                     text=self._fmt_yen_display(v.get("parcela_mensal")),
                     font=("Helvetica",8), bg=rb, fg=COLORS["blue"],
                     anchor="w", padx=4).grid(row=ri, column=5, sticky="nsew")

            prog_txt = f"{pagas}/{total_parc}" if total_parc else "\u2014"
            pc = COLORS["green"] if quitado else COLORS["accent"]
            pf = tk.Frame(self._dash_parc_rows, bg=rb)
            pf.grid(row=ri, column=6, sticky="nsew")
            tk.Label(pf, text=prog_txt, font=("Helvetica",7,"bold"),
                     bg=rb, fg=pc, anchor="w").pack(anchor="w", padx=4, pady=(4,0))
            if total_parc > 0:
                pw6 = FIN_COLS[6][1] - 12
                bw = max(2, int(pw6 * min(pagas / total_parc, 1.0)))
                bc = tk.Canvas(pf, width=pw6, height=6,
                               bg=COLORS["border"], highlightthickness=0)
                bc.pack(padx=4, pady=(1,4))
                bc.create_rectangle(0, 0, bw, 6, fill=pc, outline="")

            prox_txt = "\u2714 Quitado" if quitado else self._proxima_parcela(
                v.get("data_primeira_parc"), pagas)
            prox_fg = COLORS["green"] if quitado else COLORS["text_primary"]
            tk.Label(self._dash_parc_rows, text=prox_txt,
                     font=("Helvetica",8), bg=rb, fg=prox_fg,
                     anchor="w", padx=4).grid(row=ri, column=7, sticky="nsew")

            saldo_txt = f"\xa5{saldo_rest:,}" if not quitado else "\u2014"
            saldo_fg  = COLORS["orange"] if not quitado else COLORS["text_muted"]
            tk.Label(self._dash_parc_rows, text=saldo_txt,
                     font=("Helvetica",8,"bold"), bg=rb, fg=saldo_fg,
                     anchor="w", padx=4).grid(row=ri, column=8, sticky="nsew")

            # Status: Quitado / Pago / Em Aberto
            if quitado:
                st_txt, st_bg = "Quitado", COLORS["green"]
            elif pago_no_mes:
                st_txt, st_bg = "Pago", COLORS["green"]
            else:
                st_txt, st_bg = "Em Aberto", COLORS["red"]
            f9 = tk.Frame(self._dash_parc_rows, bg=rb)
            f9.grid(row=ri, column=9, sticky="nsew")
            tk.Label(f9, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_bg, fg="white", anchor="center"
                     ).pack(padx=3, pady=6, fill="both", expand=True)

        tk.Frame(self._dash_parc_rows, bg=COLORS["border"], height=1
                 ).grid(row=len(todas)*2, column=0,
                        columnspan=len(FIN_COLS), sticky="ew")


    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: INDICADORES — Gráfico Pizza com navegação de mês
    # ══════════════════════════════════════════════════════════════════════════
    def _build_dash_shaken_vencer(self, parent):
        """Listagem de shaken vencidos ou com menos de 3 meses para renovação."""
        import datetime

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=18, pady=18)

        # ── Cabeçalho ──────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 12))
        tk.Label(hdr, text="🔔  Shaken a Vencer", font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
        self._sv_lbl_total = tk.Label(hdr, text="", font=("Helvetica", 9),
                                      bg=COLORS["bg_content"], fg=COLORS["text_muted"])
        self._sv_lbl_total.pack(side="left", padx=14)
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._sv_refresh).pack(side="right")

        # ── KPI mini row ────────────────────────────────────────────────────────
        self._sv_kpi_frame = tk.Frame(root, bg=COLORS["bg_content"])
        self._sv_kpi_frame.pack(fill="x", pady=(0, 10))

        # ── Filtros ─────────────────────────────────────────────────────────────
        fbar = tk.Frame(root, bg=COLORS["bg_main"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        fbar.pack(fill="x", pady=(0, 8), ipady=4)
        tk.Label(fbar, text="Tipo:", font=("Helvetica", 8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(10,4))
        self._sv_filtro_tipo = tk.StringVar(value="Todos")
        ttk.Combobox(fbar, textvariable=self._sv_filtro_tipo,
                     values=["Todos", "Cliente", "Estoque", "Daisha"],
                     state="readonly", font=("Helvetica", 8), width=10
                     ).pack(side="left", padx=(0,10), ipady=2)
        self._sv_filtro_tipo.trace_add("write", lambda *_: self._sv_refresh())
        tk.Label(fbar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(4,2))
        self._sv_filtro_txt = tk.StringVar()
        tk.Entry(fbar, textvariable=self._sv_filtro_txt, font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 width=20).pack(side="left", ipady=4)
        self._sv_filtro_txt.trace_add("write", lambda *_: self._sv_refresh())

        # ── Tabela dual-canvas (header fixo + body scroll) ──────────────────────
        SV_COLS = [
            ("Urgência",   90),
            ("SK",         70),
            ("Carro",     140),
            ("Tipo",       70),
            ("Cliente",   120),
            ("Vencimento", 95),
            ("Dias",       85),
        ]
        self._sv_cols = SV_COLS
        SV_TOTAL_W = sum(c[1] for c in SV_COLS)

        tbl_card = tk.Frame(root, bg=COLORS["bg_card"],
                            highlightthickness=1, highlightbackground=COLORS["border"])
        tbl_card.pack(fill="both", expand=True)
        tk.Frame(tbl_card, bg="#D4AC0D", height=3).pack(fill="x")

        # Scrollbars
        vsb = tk.Scrollbar(tbl_card, orient="vertical")
        hsb = tk.Scrollbar(tbl_card, orient="horizontal")

        # Header canvas
        sv_hdr_cv = tk.Canvas(tbl_card, bg=COLORS["bg_main"],
                              height=28, highlightthickness=0)
        # Body canvas
        sv_body_cv = tk.Canvas(tbl_card, bg=COLORS["bg_card"],
                               highlightthickness=0, yscrollcommand=vsb.set)
        self._sv_body_cv  = sv_body_cv
        self._sv_hdr_cv   = sv_hdr_cv

        def _sv_sync_x(*args):
            sv_hdr_cv.xview(*args)
            sv_body_cv.xview(*args)

        hsb.configure(command=_sv_sync_x)
        vsb.configure(command=sv_body_cv.yview)
        sv_body_cv.configure(xscrollcommand=hsb.set)
        sv_hdr_cv.configure(xscrollcommand=hsb.set)

        sv_hdr_cv.pack(fill="x", side="top")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        sv_body_cv.pack(side="left", fill="both", expand=True)

        # Header frame inside header canvas
        sv_hdr_frame = tk.Frame(sv_hdr_cv, bg=COLORS["bg_main"])
        sv_hdr_cv.create_window((0,0), window=sv_hdr_frame, anchor="nw")
        sv_hdr_frame.bind("<Configure>",
                          lambda e: sv_hdr_cv.configure(
                              scrollregion=sv_hdr_cv.bbox("all")))

        for ci, (txt, pw) in enumerate(SV_COLS):
            sv_hdr_frame.grid_columnconfigure(ci, minsize=pw, weight=0)
        for ci, (txt, pw) in enumerate(SV_COLS):
            tk.Label(sv_hdr_frame, text=txt,
                     font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w", padx=6
                     ).grid(row=0, column=ci, sticky="nsew", ipady=6)

        # Body rows frame
        self._sv_rows_frame = tk.Frame(sv_body_cv, bg=COLORS["bg_card"])
        sv_body_cv.create_window((0,0), window=self._sv_rows_frame, anchor="nw")
        self._sv_rows_frame.bind(
            "<Configure>",
            lambda e: sv_body_cv.configure(
                scrollregion=sv_body_cv.bbox("all")))

        for ci, (_, pw) in enumerate(SV_COLS):
            self._sv_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        sv_body_cv.bind("<Enter>",
            lambda e: sv_body_cv.bind_all(
                "<MouseWheel>",
                lambda ev: sv_body_cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        sv_body_cv.bind("<Leave>",
            lambda e: sv_body_cv.unbind_all("<MouseWheel>"))

        self._sv_refresh()

    def _sv_refresh(self):
        """Atualiza lista de shaken urgentes (vencidos ou < 3 meses)."""
        import datetime
        hoje = datetime.date.today()

        for w in self._sv_rows_frame.winfo_children():
            w.destroy()
        for w in self._sv_kpi_frame.winfo_children():
            w.destroy()

        SV_COLS = getattr(self, "_sv_cols", [])
        for ci, (_, pw) in enumerate(SV_COLS):
            self._sv_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        ftipo = getattr(self, "_sv_filtro_tipo", None)
        ftipo = ftipo.get() if ftipo else "Todos"
        termo = getattr(self, "_sv_filtro_txt", None)
        termo = (termo.get() or "").strip().lower() if termo else ""

        # Busca todos os shaken ativos (não renovados, não por conta)
        rows = [dict(r) for r in self.conn.execute(
            "SELECT s.*, "
            "  c.carro AS c_nome, c.status AS c_tipo, "
            "  cl.nome AS cli_nome "
            "FROM shaken s "
            "JOIN carros c ON s.carro_id=c.id "
            "LEFT JOIN clientes cl ON s.cliente_id=cl.id "
            "WHERE (s.renovado IS NULL OR s.renovado=0) "
            "  AND (s.por_conta IS NULL OR s.por_conta=0) "
            "  AND s.data_vencimento IS NOT NULL AND s.data_vencimento != '' "
            "ORDER BY s.data_vencimento ASC"
        ).fetchall()]

        # Filtra apenas urgentes (≤ 90 dias ou vencidos)
        urgentes = []
        for r in rows:
            try:
                d, m, a = r["data_vencimento"].split("/")
                dobj = datetime.date(int(a), int(m), int(d))
                delta = (dobj - hoje).days
                if delta > 90:
                    continue
            except Exception:
                continue
            r["_delta"] = delta
            urgentes.append(r)

        # Aplica filtros
        if ftipo != "Todos":
            urgentes = [r for r in urgentes if r.get("c_tipo","") == ftipo]
        if termo:
            urgentes = [r for r in urgentes
                        if termo in (r.get("c_nome","") or "").lower()
                        or termo in (r.get("cli_nome","") or "").lower()
                        or termo in (r.get("sk_num","") or "").lower()]

        urgentes.sort(key=lambda r: r["_delta"])

        # ── Carros SEM shaken ativo (exceto Inativos) ──────────────────────
        sem_shaken = [dict(r) for r in self.conn.execute(
            "SELECT c.id, c.carro, c.status, c.cor, c.placa "
            "FROM carros c "
            "WHERE c.status != 'Inativo' "
            "  AND c.id NOT IN ("
            "      SELECT DISTINCT carro_id FROM shaken "
            "      WHERE (renovado IS NULL OR renovado=0) "
            "        AND (por_conta IS NULL OR por_conta=0) "
            "        AND data_vencimento IS NOT NULL AND data_vencimento != ''"
            "  ) "
            "ORDER BY c.carro ASC"
        ).fetchall()]

        # Filtra sem_shaken pelos mesmos filtros de tipo/termo
        if ftipo != "Todos":
            sem_shaken = [r for r in sem_shaken if r.get("status","") == ftipo]
        if termo:
            sem_shaken = [r for r in sem_shaken
                          if termo in (r.get("carro","") or "").lower()
                          or termo in (r.get("placa","") or "").lower()]

        # ── KPI mini cards ──────────────────────────────────────────────────────
        cnt_venc = sum(1 for r in urgentes if r["_delta"] < 0)
        cnt_1m   = sum(1 for r in urgentes if 0 <= r["_delta"] <= 30)
        cnt_3m   = sum(1 for r in urgentes if 30 < r["_delta"] <= 90)
        cnt_sem  = len(sem_shaken)
        for lbl, cnt, cor in [
            ("Vencidos",    cnt_venc, COLORS["red"]),
            ("< 1 mês",     cnt_1m,   COLORS["orange"]),
            ("< 3 meses",   cnt_3m,   "#D4AC0D"),
            ("Sem Shaken",  cnt_sem,  COLORS["text_muted"]),
        ]:
            kf = tk.Frame(self._sv_kpi_frame, bg=COLORS["bg_card"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            kf.pack(side="left", padx=(0,8), ipadx=16, ipady=4)
            tk.Frame(kf, bg=cor, height=3).pack(fill="x")
            tk.Label(kf, text=str(cnt), font=("Helvetica",20,"bold"),
                     bg=COLORS["bg_card"], fg=cor).pack(pady=(6,2))
            tk.Label(kf, text=lbl, font=("Helvetica",8),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(pady=(0,4))

        total_registros = len(urgentes) + cnt_sem
        self._sv_lbl_total.configure(text=f"{len(urgentes)} a vencer  |  {cnt_sem} sem shaken")

        if not urgentes and not sem_shaken:
            tk.Label(self._sv_rows_frame,
                     text="✓  Nenhum shaken vencido, próximo do vencimento ou sem shaken.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["green"]).grid(row=0, column=0,
                                              columnspan=len(SV_COLS), pady=40)
            return

        def _urgencia(delta):
            if delta < 0:   return ("VENCIDO",   COLORS["red"])
            elif delta <= 30: return ("< 1 mês",  COLORS["orange"])
            else:             return ("< 3 meses","#D4AC0D")

        def cell(ri, ci, txt, fg, bold=False, bg=None):
            _bg = bg or (COLORS["bg_card"] if ri % 2 == 0 else COLORS["bg_content"])
            f = tk.Frame(self._sv_rows_frame, bg=_bg)
            f.grid(row=ri*2+1, column=ci, sticky="nsew")
            tk.Label(f, text=str(txt),
                     font=("Helvetica", 8, "bold" if bold else "normal"),
                     bg=_bg, fg=fg, anchor="w", padx=6
                     ).pack(fill="both", expand=True, pady=4)

        for i, r in enumerate(urgentes):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            # separator
            tk.Frame(self._sv_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=i*2, column=0, columnspan=len(SV_COLS), sticky="ew")
            self._sv_rows_frame.grid_rowconfigure(i*2+1, minsize=32)

            delta = r["_delta"]
            urg_txt, urg_cor = _urgencia(delta)
            dias_txt = f"{abs(delta)}d {'atrás' if delta < 0 else 'restantes'}"

            cell(i, 0, urg_txt,                              urg_cor,               bold=True,  bg=rb)
            cell(i, 1, r.get("sk_num") or f"#{r['id']}",    "#D4AC0D",             bold=False, bg=rb)
            cell(i, 2, (r.get("c_nome") or "—")[:18],       COLORS["text_primary"],bold=False, bg=rb)
            cell(i, 3, r.get("c_tipo") or "—",              COLORS["text_muted"],  bold=False, bg=rb)
            cell(i, 4, (r.get("cli_nome") or "—")[:16],     COLORS["accent"],      bold=False, bg=rb)
            cell(i, 5, r.get("data_vencimento") or "—",     urg_cor,               bold=True,  bg=rb)
            cell(i, 6, dias_txt,                             urg_cor,               bold=True,  bg=rb)

        # linha final urgentes
        tk.Frame(self._sv_rows_frame, bg=COLORS["border"], height=1
                 ).grid(row=len(urgentes)*2, column=0,
                        columnspan=len(SV_COLS), sticky="ew")

        # ── Sem Shaken ──────────────────────────────────────────────────────
        if sem_shaken:
            base_row = len(urgentes) * 2 + 1
            # cabeçalho de seção
            sep_f = tk.Frame(self._sv_rows_frame, bg=COLORS["text_muted"])
            sep_f.grid(row=base_row, column=0, columnspan=len(SV_COLS), sticky="ew")
            tk.Label(sep_f, text="  SEM SHAKEN REGISTRADO",
                     font=("Helvetica",8,"bold"),
                     bg=COLORS["text_muted"], fg="white"
                     ).pack(side="left", pady=3, padx=6)

            for j, r in enumerate(sem_shaken):
                ri_abs = base_row + 1 + j * 2
                rb = COLORS["bg_card"] if j % 2 == 0 else COLORS["bg_content"]
                tk.Frame(self._sv_rows_frame, bg=COLORS["border"], height=1
                         ).grid(row=ri_abs, column=0, columnspan=len(SV_COLS), sticky="ew")
                self._sv_rows_frame.grid_rowconfigure(ri_abs + 1, minsize=32)

                def _sem_cell(ri, ci, txt, fg, bg):
                    f = tk.Frame(self._sv_rows_frame, bg=bg)
                    f.grid(row=ri+1, column=ci, sticky="nsew")
                    tk.Label(f, text=str(txt),
                             font=("Helvetica",8),
                             bg=bg, fg=fg, anchor="w", padx=6
                             ).pack(fill="both", expand=True, pady=4)

                _sem_cell(ri_abs, 0, "SEM SHAKEN",          COLORS["text_muted"], rb)
                _sem_cell(ri_abs, 1, "—",                   COLORS["text_muted"], rb)
                _sem_cell(ri_abs, 2, (r.get("carro") or "—")[:18], COLORS["text_primary"], rb)
                _sem_cell(ri_abs, 3, r.get("status") or "—", COLORS["text_muted"],  rb)
                _sem_cell(ri_abs, 4, "—",                   COLORS["text_muted"], rb)
                _sem_cell(ri_abs, 5, "—",                   COLORS["text_muted"], rb)
                _sem_cell(ri_abs, 6, "—",                   COLORS["text_muted"], rb)
                _sem_cell(ri_abs, 7, "—",                   COLORS["text_muted"], rb)

            # linha final sem shaken
            last = base_row + 1 + len(sem_shaken) * 2
            tk.Frame(self._sv_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=last, column=0, columnspan=len(SV_COLS), sticky="ew")

    def _build_dash_indicadores(self, parent):
        import datetime, math
        hoje = datetime.date.today()
        self._ind_mes = tk.IntVar(value=hoje.month)
        self._ind_ano = tk.IntVar(value=hoje.year)

        MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=18, pady=18)

        # ── Header: título + navegação de mês ────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 10))
        self._ind_title_lbl = tk.Label(hdr, text="◈  Indicadores — Receita do Mês",
                 font=("Helvetica",13,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"])
        self._ind_title_lbl.pack(side="left")

        nav = tk.Frame(hdr, bg=COLORS["bg_content"])
        nav.pack(side="right")

        def prev_mes():
            m = self._ind_mes.get(); a = self._ind_ano.get()
            m -= 1
            if m < 1: m = 12; a -= 1
            self._ind_mes.set(m); self._ind_ano.set(a)
            lbl_mes.configure(text=f"{MESES_PT[m-1]}  {a}")
            _dispatch_refresh()

        def next_mes():
            m = self._ind_mes.get(); a = self._ind_ano.get()
            m += 1
            if m > 12: m = 1; a += 1
            self._ind_mes.set(m); self._ind_ano.set(a)
            lbl_mes.configure(text=f"{MESES_PT[m-1]}  {a}")
            _dispatch_refresh()

        tk.Button(nav, text="◀", font=("Helvetica",11,"bold"),
                  bg=COLORS["bg_card"], fg=COLORS["accent"], relief="flat",
                  cursor="hand2", padx=8, command=prev_mes).pack(side="left", ipady=3)
        lbl_mes = tk.Label(nav, text=f"{MESES_PT[hoje.month-1]}  {hoje.year}",
                           font=("Helvetica",11,"bold"), width=16,
                           bg=COLORS["bg_content"], fg=COLORS["text_primary"])
        lbl_mes.pack(side="left", padx=4)
        tk.Button(nav, text="▶", font=("Helvetica",11,"bold"),
                  bg=COLORS["bg_card"], fg=COLORS["accent"], relief="flat",
                  cursor="hand2", padx=8, command=next_mes).pack(side="left", ipady=3)

        # ── Abas ─────────────────────────────────────────────────────────────
        tab_bar = tk.Frame(root, bg=COLORS["bg_content"])
        tab_bar.pack(fill="x", pady=(0, 0))

        content_area = tk.Frame(root, bg=COLORS["bg_content"])
        content_area.pack(fill="both", expand=True)

        # Frames das abas
        frame_receita = tk.Frame(content_area, bg=COLORS["bg_content"])
        frame_lucro   = tk.Frame(content_area, bg=COLORS["bg_content"])

        self._ind_tab_frames    = {"receita": frame_receita, "lucro": frame_lucro}
        self._ind_tab_active    = tk.StringVar(value="receita")
        self._ind_tab_btns      = {}

        def switch_tab(name):
            self._ind_tab_active.set(name)
            for k, f in self._ind_tab_frames.items():
                if k == name:
                    f.pack(fill="both", expand=True)
                else:
                    f.pack_forget()
            titles = {"receita": "◈  Indicadores — Receita do Mês",
                      "lucro":   "◈  Indicadores — Lucro do Mês"}
            self._ind_title_lbl.configure(text=titles[name])
            for k, btn in self._ind_tab_btns.items():
                if k == name:
                    btn.configure(bg=COLORS["accent"], fg="white")
                else:
                    btn.configure(bg=COLORS["bg_card"], fg=COLORS["text_secondary"])
            root.after_idle(_dispatch_refresh)

        for key, label in [("receita","📊  Receita do Mês"), ("lucro","💹  Lucro do Mês")]:
            btn = tk.Button(tab_bar, text=label,
                            font=("Helvetica", 9, "bold"),
                            relief="flat", cursor="hand2",
                            padx=14, pady=6,
                            command=lambda k=key: switch_tab(k))
            btn.pack(side="left", padx=(0,4))
            self._ind_tab_btns[key] = btn

        # Separador visual
        tk.Frame(tab_bar, bg=COLORS["border"], height=2).pack(fill="x", side="bottom")

        # ─────────────────────────────────────────────────────────────────────
        # ABA: RECEITA DO MÊS
        # ─────────────────────────────────────────────────────────────────────
        pizza_card = tk.Frame(frame_receita, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        pizza_card.pack(side="left", fill="both", expand=True, padx=(0,12), pady=8)
        tk.Frame(pizza_card, bg=COLORS["accent"], height=3).pack(fill="x")
        self._ind_pizza_title = tk.Label(pizza_card, text="Distribuição da Receita",
                                          font=("Helvetica",10,"bold"),
                                          bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._ind_pizza_title.pack(pady=(10,4))
        self._ind_cv_pizza = tk.Canvas(pizza_card, bg=COLORS["bg_card"],
                                        highlightthickness=0, width=280, height=280)
        self._ind_cv_pizza.pack(pady=6)

        legend_outer = tk.Frame(frame_receita, bg=COLORS["bg_content"])
        legend_outer.pack(side="right", fill="both", expand=True, pady=8)
        self._ind_legend_frame = legend_outer

        TIPOS_REC = [
            ("Venda a Vista",      COLORS["green"]),
            ("Venda Parcelada",    COLORS["blue"]),
            ("Com Troca",          COLORS["orange"]),
            ("Com Troca e Volta",  COLORS["accent"]),
            ("Serviços",           "#9B59B6"),
            ("Shaken",             "#D4AC0D"),
        ]
        self._ind_tipos = TIPOS_REC

        # ─────────────────────────────────────────────────────────────────────
        # ABA: LUCRO DO MÊS
        # ─────────────────────────────────────────────────────────────────────
        lp_card = tk.Frame(frame_lucro, bg=COLORS["bg_card"],
                           highlightthickness=1, highlightbackground=COLORS["border"])
        lp_card.pack(side="left", fill="both", expand=True, padx=(0,12), pady=8)
        tk.Frame(lp_card, bg=COLORS["green"], height=3).pack(fill="x")
        tk.Label(lp_card, text="Distribuição do Lucro",
                 font=("Helvetica",10,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(pady=(10,4))
        self._ind_cv_lucro = tk.Canvas(lp_card, bg=COLORS["bg_card"],
                                        highlightthickness=0, width=280, height=280)
        self._ind_cv_lucro.pack(pady=6)

        lucro_legend_outer = tk.Frame(frame_lucro, bg=COLORS["bg_content"])
        lucro_legend_outer.pack(side="right", fill="both", expand=True, pady=8)
        self._ind_lucro_legend_frame = lucro_legend_outer

        TIPOS_LUC = [
            ("Lucro Vendas",   COLORS["green"]),
            ("Lucro OS",       COLORS["blue"]),
            ("Lucro Shaken",   "#D4AC0D"),
        ]
        self._ind_tipos_luc = TIPOS_LUC

        # ═════════════════════════════════════════════════════════════════════
        # REFRESH RECEITA
        # ═════════════════════════════════════════════════════════════════════
        def _refresh_indicadores():
            ms = f"{self._ind_mes.get():02d}"
            ys = str(self._ind_ano.get())
            valores = {}
            # Busca listas para contagem de registros
            try:
                vendas_mes = [dict(r) for r in self.conn.execute(
                    "SELECT tipo_venda, valor_venda FROM vendas WHERE data_venda LIKE ?",
                    (f"%/{ms}/{ys}",)).fetchall()]
            except Exception:
                vendas_mes = []
            try:
                servs_mes = self.conn.execute(
                    "SELECT id FROM servicos WHERE data_servico LIKE ? AND status='Concluído'",
                    (f"%/{ms}/{ys}",)).fetchall()
            except Exception:
                servs_mes = []
            try:
                sks_mes = self.conn.execute(
                    "SELECT id FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL AND valor != ''",
                    (f"%/{ms}/{ys}",)).fetchall()
            except Exception:
                sks_mes = []

            for tipo, _ in TIPOS_REC:
                if tipo == "Serviços":
                    try:
                        v = self.conn.execute(
                            "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                            "FROM servicos WHERE data_servico LIKE ? AND status='Concluído'",
                            (f"%/{ms}/{ys}",)).fetchone()[0]
                    except Exception:
                        v = 0
                elif tipo == "Shaken":
                    try:
                        v = self.conn.execute(
                            "SELECT COALESCE(SUM(CAST(REPLACE(REPLACE(valor,',',''),'¥','') AS INTEGER)),0) "
                            "FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL AND valor != ''",
                            (f"%/{ms}/{ys}",)).fetchone()[0]
                    except Exception:
                        v = 0
                else:
                    rows = self.conn.execute(
                        "SELECT valor_venda FROM vendas WHERE tipo_venda=? AND data_venda LIKE ?",
                        (tipo, f"%/{ms}/{ys}")).fetchall()
                    try:
                        v = sum(int(str(r[0] or "0").replace(",","")) for r in rows)
                    except Exception:
                        v = 0
                valores[tipo] = v

            total_geral = sum(valores.values())
            total_div   = total_geral or 1

            cv = self._ind_cv_pizza
            cv.delete("all")
            W = cv.winfo_width() or 280
            H = cv.winfo_height() or 280
            cx, cy = W//2, H//2
            R = min(W, H)//2 - 18
            start = 90.0   # começa no topo (tkinter: 90° = topo, cresce anti-horário)
            for tipo, cor in TIPOS_REC:
                val = valores[tipo]
                if val <= 0: continue
                pct    = val / total_div
                extent = pct * 360.0
                # tkinter arc: start em graus, extent positivo = anti-horário
                cv.create_arc(cx-R, cy-R, cx+R, cy+R,
                              start=start, extent=extent,
                              fill=cor, outline=COLORS["bg_card"], width=2)
                # ângulo do meio da fatia em radianos (tkinter: 0=leste, positivo=anti-horário)
                # converter para radianos com eixo Y invertido do canvas
                mid_deg = start + extent / 2
                mid_rad = math.radians(mid_deg)
                tx = cx + (R * 0.68) * math.cos(mid_rad)
                ty = cy - (R * 0.68) * math.sin(mid_rad)   # Y invertido
                if pct > 0.05:
                    cv.create_text(tx, ty, text=f"{pct*100:.1f}%",
                                   font=("Helvetica",8,"bold"), fill="white")
                start += extent
            ri = R * 0.36
            cv.create_oval(cx-ri, cy-ri, cx+ri, cy+ri,
                           fill=COLORS["bg_card"], outline="")
            cv.create_text(cx, cy-7, text=f"¥{total_geral:,}",
                           font=("Helvetica",9,"bold"), fill=COLORS["text_primary"])
            cv.create_text(cx, cy+9, text="Total do Mês",
                           font=("Helvetica",7), fill=COLORS["text_muted"])

            for w in self._ind_legend_frame.winfo_children():
                w.destroy()
            total_card = tk.Frame(self._ind_legend_frame, bg=COLORS["bg_card"],
                                   highlightthickness=2, highlightbackground=COLORS["accent"])
            total_card.pack(fill="x", pady=(0,10))
            tk.Frame(total_card, bg=COLORS["accent"], height=4).pack(fill="x")
            tc_inner = tk.Frame(total_card, bg=COLORS["bg_card"])
            tc_inner.pack(fill="x", padx=14, pady=10)
            tk.Label(tc_inner, text="TOTAL RECEITA DO MÊS",
                     font=("Helvetica",8,"bold"), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w")
            tk.Label(tc_inner, text=f"¥ {total_geral:,}",
                     font=("Helvetica",18,"bold"), bg=COLORS["bg_card"],
                     fg=COLORS["accent"]).pack(anchor="w")

            # Conta registros por tipo
            cnt_tipo = {}
            for tipo, _ in TIPOS_REC:
                if tipo == "Serviços":
                    cnt_tipo[tipo] = len(servs_mes)
                elif tipo == "Shaken":
                    cnt_tipo[tipo] = len(sks_mes)
                else:
                    cnt_tipo[tipo] = sum(1 for v in vendas_mes if v.get("tipo_venda") == tipo)

            for tipo, cor in TIPOS_REC:
                val = valores[tipo]
                pct = val / total_div * 100
                cnt = cnt_tipo.get(tipo, 0)
                ind = tk.Frame(self._ind_legend_frame, bg=COLORS["bg_card"],
                               highlightthickness=1, highlightbackground=COLORS["border"])
                ind.pack(fill="x", pady=(0,6))
                tk.Frame(ind, bg=cor, width=6).pack(side="left", fill="y")
                inner = tk.Frame(ind, bg=COLORS["bg_card"])
                inner.pack(side="left", fill="both", expand=True, padx=12, pady=7)
                rt = tk.Frame(inner, bg=COLORS["bg_card"])
                rt.pack(fill="x")
                tk.Label(rt, text=tipo, font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
                # Badge share% (igual ao lucro)
                tk.Label(rt, text=f"share {pct:.1f}%",
                         font=("Helvetica",8,"bold"),
                         bg=cor, fg="white", padx=5).pack(side="right")
                tk.Label(inner, text=f"¥ {val:,}", font=("Helvetica",11,"bold"),
                         bg=COLORS["bg_card"], fg=cor).pack(anchor="w")
                # Linha registros
                r3 = tk.Frame(inner, bg=COLORS["bg_card"])
                r3.pack(fill="x", pady=(2,0))
                tk.Label(r3, text=f"{cnt} registro(s)",
                         font=("Helvetica",8), bg=COLORS["bg_card"],
                         fg=COLORS["text_muted"]).pack(side="left")
                bar_bg = tk.Frame(inner, bg=COLORS["border"], height=4)
                bar_bg.pack(fill="x", pady=(3,0))
                def _dm(e, bg=bar_bg, p=pct/100, c=cor):
                    bw = max(2, int(bg.winfo_width() * p))
                    tk.Frame(bg, bg=c, width=bw, height=4).place(x=0, y=0)
                bar_bg.bind("<Configure>", _dm)

        # ═════════════════════════════════════════════════════════════════════
        # REFRESH LUCRO
        # ═════════════════════════════════════════════════════════════════════
        def _refresh_lucro_ind():
            ms = f"{self._ind_mes.get():02d}"
            ys = str(self._ind_ano.get())

            # ── Lucro de Vendas ──────────────────────────────────────────────
            lucro_vendas   = 0
            receita_vendas = 0
            cnt_vendas     = 0
            try:
                vendas_mes = [dict(r) for r in self.conn.execute(
                    "SELECT v.* FROM vendas v "
                    "WHERE v.data_venda LIKE ?", (f"%/{ms}/{ys}",)).fetchall()]
                for v in vendas_mes:
                    try:
                        vv = int(str(v.get("valor_venda") or "0").replace(",",""))
                    except Exception:
                        vv = 0
                    # Achar compra vinculado: prefere compra_id, fallback por carro_id
                    custo_comp = 0
                    custo_cust = 0
                    compra_id_link = v.get("compra_id")
                    if compra_id_link:
                        cr = self.conn.execute(
                            "SELECT id, valor FROM compras WHERE id=?", (compra_id_link,)
                        ).fetchone()
                    elif v.get("carro_id"):
                        cr = self.conn.execute(
                            "SELECT id, valor FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                            (v.get("carro_id"),)).fetchone()
                    else:
                        cr = None
                    if cr:
                        try:
                            custo_comp = int(str(cr[1] or "0").replace(",",""))
                        except Exception:
                            custo_comp = 0
                        custo_cust = self._get_custo_total(cr[0])
                    lucro_vendas   += vv - custo_comp - custo_cust
                    receita_vendas += vv
                    cnt_vendas     += 1
            except Exception:
                pass
            margem_v = (lucro_vendas / receita_vendas * 100) if receita_vendas > 0 else 0

            # ── Lucro de OS (receita - custos_os do mês) ─────────────────────
            lucro_os   = 0
            receita_os = 0
            cnt_os     = 0
            try:
                os_rows = self.conn.execute(
                    "SELECT id, valor FROM servicos "
                    "WHERE data_servico LIKE ? AND status='Concluído'",
                    (f"%/{ms}/{ys}",)).fetchall()
                for o in os_rows:
                    try:
                        rv = int(str(o[1] or "0").replace(",",""))
                    except Exception:
                        rv = 0
                    custo_os_row = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_os WHERE servico_id=?", (o[0],)).fetchone()
                    cv_os = int(custo_os_row[0]) if custo_os_row[0] else 0
                    lucro_os   += rv - cv_os
                    receita_os += rv
                    cnt_os     += 1
            except Exception:
                pass
            margem_os = (lucro_os / receita_os * 100) if receita_os > 0 else 0

            # ── Lucro de Shaken (receita - custos_sk de carros Cliente) ────────
            lucro_sk   = 0
            receita_sk = 0
            cnt_sk     = 0
            try:
                sk_rows = self.conn.execute(
                    "SELECT id, valor FROM shaken "
                    "WHERE data_registro LIKE ? AND valor IS NOT NULL AND valor != ''",
                    (f"%/{ms}/{ys}",)).fetchall()
                for sk in sk_rows:
                    try:
                        rv_sk = int(str(sk[1] or "0").replace(",","").replace("¥","").strip())
                    except Exception:
                        rv_sk = 0
                    # Custo = soma dos custos_sk (apenas carros Cliente geram custos_sk)
                    ex_sk = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_sk WHERE shaken_id=?", (sk[0],)).fetchone()
                    ex_sk_v = int(ex_sk[0]) if ex_sk[0] else 0
                    lucro_sk   += rv_sk - ex_sk_v
                    receita_sk += rv_sk
                    cnt_sk     += 1
            except Exception:
                pass
            margem_sk = (lucro_sk / receita_sk * 100) if receita_sk > 0 else 0

            total_lucro = lucro_vendas + lucro_os + lucro_sk

            # Despesas fixas do mês
            ms = f"{self._ind_mes.get():02d}"
            ys = str(self._ind_ano.get())
            mes_ref = f"{ms}/{ys}"
            desp_rows = self.conn.execute(
                "SELECT valor FROM despesas_fixas "
                "WHERE data_ref=? OR recorrente=1", (mes_ref,)).fetchall()
            def _int_d(s):
                try: return int(str(s or "0").replace(",",""))
                except: return 0
            total_despesas = sum(_int_d(r[0]) for r in desp_rows)
            lucro_liquido  = total_lucro - total_despesas

            total_div   = total_lucro if total_lucro > 0 else 1
            total_abs   = abs(lucro_vendas) + abs(lucro_os) + abs(lucro_sk) or 1

            valores_luc = {
                "Lucro Vendas": lucro_vendas,
                "Lucro OS":     lucro_os,
                "Lucro Shaken": lucro_sk,
            }

            # Pizza lucro
            cv2 = self._ind_cv_lucro
            cv2.delete("all")
            W2 = cv2.winfo_width() or 280
            H2 = cv2.winfo_height() or 280
            cx2, cy2 = W2//2, H2//2
            R2 = min(W2, H2)//2 - 18

            # Só fatias positivas para a pizza
            pos_total = sum(v for v in valores_luc.values() if v > 0) or 1
            has_positive = any(v > 0 for v in valores_luc.values())
            if not has_positive:
                cv2.create_text(cx2, cy2, text="Sem lucro positivo",
                                font=("Helvetica",9), fill=COLORS["text_muted"])
            else:
                start2 = 90.0
                for tipo, cor in TIPOS_LUC:
                    val = valores_luc[tipo]
                    if val <= 0: continue
                    pct2    = val / pos_total
                    extent2 = pct2 * 360.0
                    cv2.create_arc(cx2-R2, cy2-R2, cx2+R2, cy2+R2,
                                   start=start2, extent=extent2,
                                   fill=cor, outline=COLORS["bg_card"], width=2)
                    mid2_deg = start2 + extent2 / 2
                    mid2_rad = math.radians(mid2_deg)
                    tx2 = cx2 + (R2 * 0.68) * math.cos(mid2_rad)
                    ty2 = cy2 - (R2 * 0.68) * math.sin(mid2_rad)
                    if pct2 > 0.05:
                        cv2.create_text(tx2, ty2, text=f"{pct2*100:.1f}%",
                                        font=("Helvetica",8,"bold"), fill="white")
                    start2 += extent2

            ri2 = R2 * 0.36
            cv2.create_oval(cx2-ri2, cy2-ri2, cx2+ri2, cy2+ri2,
                            fill=COLORS["bg_card"], outline="")
            total_color = COLORS["green"] if total_lucro >= 0 else COLORS["red"]
            cv2.create_text(cx2, cy2-7,
                            text=f"¥{total_lucro:,}",
                            font=("Helvetica",9,"bold"), fill=total_color)
            cv2.create_text(cx2, cy2+9, text="Lucro Total",
                            font=("Helvetica",7), fill=COLORS["text_muted"])

            # Legendas lucro
            for w in self._ind_lucro_legend_frame.winfo_children():
                w.destroy()

            # Card total
            total_color = COLORS["green"] if total_lucro >= 0 else COLORS["red"]
            liq_color   = COLORS["green"] if lucro_liquido >= 0 else COLORS["red"]
            lc_total = tk.Frame(self._ind_lucro_legend_frame, bg=COLORS["bg_card"],
                                 highlightthickness=2,
                                 highlightbackground=liq_color)
            lc_total.pack(fill="x", pady=(0,10))
            tk.Frame(lc_total, bg=liq_color, height=4).pack(fill="x")
            lct_inner = tk.Frame(lc_total, bg=COLORS["bg_card"])
            lct_inner.pack(fill="x", padx=14, pady=10)
            # Linha bruto
            r_bruto = tk.Frame(lct_inner, bg=COLORS["bg_card"])
            r_bruto.pack(fill="x")
            tk.Label(r_bruto, text="LUCRO BRUTO",
                     font=("Helvetica",7), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(side="left")
            tk.Label(r_bruto, text=f"¥ {total_lucro:,}",
                     font=("Helvetica",9,"bold"), bg=COLORS["bg_card"],
                     fg=total_color).pack(side="right")
            # Linha despesas
            r_desp = tk.Frame(lct_inner, bg=COLORS["bg_card"])
            r_desp.pack(fill="x", pady=(2,0))
            tk.Label(r_desp, text="(-) Despesas Fixas",
                     font=("Helvetica",7), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(side="left")
            tk.Label(r_desp, text=f"¥ {total_despesas:,}",
                     font=("Helvetica",9,"bold"), bg=COLORS["bg_card"],
                     fg=COLORS["red"]).pack(side="right")
            # Separador
            tk.Frame(lct_inner, bg=COLORS["border"], height=1).pack(fill="x", pady=4)
            # Lucro líquido grande
            tk.Label(lct_inner, text="LUCRO LÍQUIDO DO MÊS",
                     font=("Helvetica",8,"bold"), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w")
            tk.Label(lct_inner, text=f"¥ {lucro_liquido:,}",
                     font=("Helvetica",18,"bold"), bg=COLORS["bg_card"],
                     fg=liq_color).pack(anchor="w")

            # Cards por tipo
            infos = [
                ("Lucro Vendas",  lucro_vendas, receita_vendas, margem_v,  cnt_vendas, COLORS["green"]),
                ("Lucro OS",      lucro_os,     receita_os,     margem_os, cnt_os,     COLORS["blue"]),
                ("Lucro Shaken",  lucro_sk,     receita_sk,     margem_sk, cnt_sk,     "#D4AC0D"),
            ]
            for tipo, lucro, receita, margem, cnt, cor in infos:
                share = abs(lucro) / total_abs * 100
                lucro_color = COLORS["green"] if lucro >= 0 else COLORS["red"]
                card = tk.Frame(self._ind_lucro_legend_frame, bg=COLORS["bg_card"],
                                highlightthickness=1, highlightbackground=COLORS["border"])
                card.pack(fill="x", pady=(0,7))
                tk.Frame(card, bg=cor, width=6).pack(side="left", fill="y")
                inner = tk.Frame(card, bg=COLORS["bg_card"])
                inner.pack(side="left", fill="both", expand=True, padx=12, pady=8)

                # Linha 1: Nome + share
                r1 = tk.Frame(inner, bg=COLORS["bg_card"])
                r1.pack(fill="x")
                tk.Label(r1, text=tipo, font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
                tk.Label(r1, text=f"share {share:.1f}%",
                         font=("Helvetica",8,"bold"),
                         bg=cor, fg="white", padx=5).pack(side="right")

                # Linha 2: Lucro grande
                tk.Label(inner, text=f"¥ {lucro:,}",
                         font=("Helvetica",13,"bold"),
                         bg=COLORS["bg_card"], fg=lucro_color).pack(anchor="w")

                # Linha 3: Margem e n° de registros
                r3 = tk.Frame(inner, bg=COLORS["bg_card"])
                r3.pack(fill="x", pady=(2,0))
                margem_color = COLORS["green"] if margem >= 0 else COLORS["red"]
                tk.Label(r3, text=f"Margem: {margem:+.1f}%",
                         font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_card"], fg=margem_color).pack(side="left")
                tk.Label(r3, text=f"{cnt} registro(s)",
                         font=("Helvetica",7),
                         bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="right")

                # Barra de receita
                tk.Label(inner, text=f"Receita: ¥ {receita:,}",
                         font=("Helvetica",7),
                         bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(anchor="w")
                bar_bg2 = tk.Frame(inner, bg=COLORS["border"], height=4)
                bar_bg2.pack(fill="x", pady=(3,0))
                def _dm2(e, bg=bar_bg2, p=min(share/100, 1.0), c=cor):
                    bw = max(2, int(bg.winfo_width() * p))
                    tk.Frame(bg, bg=c, width=bw, height=4).place(x=0, y=0)
                bar_bg2.bind("<Configure>", _dm2)

        # ── Dispatcher ───────────────────────────────────────────────────────
        def _dispatch_refresh():
            if self._ind_tab_active.get() == "receita":
                _refresh_indicadores()
            else:
                _refresh_lucro_ind()

        # Bind canvas resizes (guarded)
        def _on_pizza_cfg(e):
            if self._ind_tab_active.get() == "receita" and e.width > 10:
                root.after_idle(_refresh_indicadores)
        def _on_lucro_cfg(e):
            if self._ind_tab_active.get() == "lucro" and e.width > 10:
                root.after_idle(_refresh_lucro_ind)
        self._ind_cv_pizza.bind("<Configure>", _on_pizza_cfg)
        self._ind_cv_lucro.bind("<Configure>", _on_lucro_cfg)

        # Inicializa
        switch_tab("receita")
        root.after(400, _refresh_indicadores)



    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: DOSSIÊ CLIENTE
    # ══════════════════════════════════════════════════════════════════════════
    def _build_relatorio_mensal(self, parent):
        """Relatório mensal consolidado com exportação PDF."""
        import datetime, math
        hoje = datetime.date.today()
        MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=18, pady=14)

        # ── Cabeçalho + navegação ──────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 12))

        tk.Label(hdr, text="📋  Relatório Mensal", font=("Helvetica",14,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=lambda: _refresh()).pack(side="left", padx=(10,0))

        self._rm_mes = tk.IntVar(value=hoje.month)
        self._rm_ano = tk.IntVar(value=hoje.year)

        nav = tk.Frame(hdr, bg=COLORS["bg_content"])
        nav.pack(side="right")

        self._rm_lbl_mes = tk.Label(nav,
            text=f"{MESES_PT[hoje.month-1]}  {hoje.year}",
            font=("Helvetica",11,"bold"), width=16,
            bg=COLORS["bg_content"], fg=COLORS["text_primary"])

        def _prev():
            m = self._rm_mes.get(); a = self._rm_ano.get()
            m -= 1
            if m < 1: m = 12; a -= 1
            self._rm_mes.set(m); self._rm_ano.set(a)
            self._rm_lbl_mes.configure(text=f"{MESES_PT[m-1]}  {a}")
            _refresh()

        def _next():
            m = self._rm_mes.get(); a = self._rm_ano.get()
            m += 1
            if m > 12: m = 1; a += 1
            self._rm_mes.set(m); self._rm_ano.set(a)
            self._rm_lbl_mes.configure(text=f"{MESES_PT[m-1]}  {a}")
            _refresh()

        tk.Button(nav, text="◀", font=("Helvetica",11,"bold"),
                  bg=COLORS["bg_card"], fg=COLORS["accent"],
                  relief="flat", cursor="hand2", padx=8,
                  command=_prev).pack(side="left", ipady=3)
        self._rm_lbl_mes.pack(side="left", padx=4)
        tk.Button(nav, text="▶", font=("Helvetica",11,"bold"),
                  bg=COLORS["bg_card"], fg=COLORS["accent"],
                  relief="flat", cursor="hand2", padx=8,
                  command=_next).pack(side="left")

        tk.Button(nav, text="📄  Exportar PDF", font=("Helvetica",9,"bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=10, pady=4, command=lambda: _exportar_pdf()
                  ).pack(side="left", padx=(16,0), ipady=2)

        # ── Container scrollável ────────────────────────────────────────────────
        outer = tk.Frame(root, bg=COLORS["bg_content"])
        outer.pack(fill="both", expand=True)
        vsb = tk.Scrollbar(outer, orient="vertical")
        cv  = tk.Canvas(outer, bg=COLORS["bg_content"], highlightthickness=0,
                        yscrollcommand=vsb.set)
        vsb.configure(command=cv.yview)
        vsb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True)
        self._rm_inner = tk.Frame(cv, bg=COLORS["bg_content"])
        cw = cv.create_window((0,0), window=self._rm_inner, anchor="nw")
        self._rm_inner.bind("<Configure>", lambda e: cv.configure(
            scrollregion=cv.bbox("all")))
        cv.bind("<Configure>", lambda e: cv.itemconfig(cw, width=e.width))
        cv.bind("<Enter>", lambda e: cv.bind_all(
            "<MouseWheel>", lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        cv.bind("<Leave>", lambda e: cv.unbind_all("<MouseWheel>"))

        # ── Helpers de layout compacto ──────────────────────────────────────
        def sec_hdr(txt, color=COLORS["blue"]):
            """Linha de seção: barra colorida + título inline."""
            row = tk.Frame(self._rm_inner, bg=COLORS["bg_content"])
            row.pack(fill="x", pady=(10,1))
            tk.Frame(row, bg=color, width=4, height=18).pack(side="left", padx=(0,6))
            tk.Label(row, text=txt, font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=color).pack(side="left")

        def kpi_row(items):
            """Linha de KPIs: múltiplos campos lado a lado, compactos."""
            row = tk.Frame(self._rm_inner, bg=COLORS["bg_card"],
                           highlightthickness=1, highlightbackground=COLORS["border"])
            row.pack(fill="x", pady=(2,0))
            for i, (lbl, val, cor) in enumerate(items):
                if i > 0:
                    tk.Frame(row, bg=COLORS["border"], width=1
                             ).pack(side="left", fill="y", pady=4)
                cell = tk.Frame(row, bg=COLORS["bg_card"])
                cell.pack(side="left", expand=True, fill="x", padx=8, pady=5)
                tk.Label(cell, text=lbl, font=("Helvetica",7),
                         bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                         anchor="w").pack(anchor="w")
                tk.Label(cell, text=val, font=("Helvetica",10,"bold"),
                         bg=COLORS["bg_card"], fg=cor,
                         anchor="w").pack(anchor="w")

        def fc_row(items):
            """Linha de fluxo de caixa: label à esq, valor à dir."""
            for lbl, val, cor in items:
                row = tk.Frame(self._rm_inner, bg=COLORS["bg_card"])
                row.pack(fill="x")
                tk.Label(row, text=lbl, font=("Helvetica",8),
                         bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                         anchor="w").pack(side="left", padx=(12,4), pady=2)
                tk.Label(row, text=val, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_card"], fg=cor,
                         anchor="e").pack(side="right", padx=12, pady=2)

        def fc_total(lbl, val, cor):
            row = tk.Frame(self._rm_inner, bg=COLORS["bg_content"],
                           highlightthickness=1, highlightbackground=cor)
            row.pack(fill="x", pady=(1,4))
            tk.Label(row, text=lbl, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_content"], fg=cor).pack(side="left", padx=12, pady=4)
            tk.Label(row, text=val, font=("Helvetica",10,"bold"),
                     bg=COLORS["bg_content"], fg=cor).pack(side="right", padx=12, pady=4)

        def fc_saldo_widget(val, cor):
            row = tk.Frame(self._rm_inner, bg=cor)
            row.pack(fill="x", pady=(4,0))
            tk.Label(row, text="SALDO DO MÊS", font=("Helvetica",9,"bold"),
                     bg=cor, fg="white").pack(side="left", padx=14, pady=6)
            tk.Label(row, text=val, font=("Helvetica",12,"bold"),
                     bg=cor, fg="white").pack(side="right", padx=14, pady=6)

        def _int(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        self._rm_data_cache = {}  # para reusar no PDF

        def _refresh():
            for w in self._rm_inner.winfo_children():
                w.destroy()

            m = self._rm_mes.get(); a = self._rm_ano.get()
            ms = f"{m:02d}"; ys = str(a)
            mes_ref = f"{ms}/{ys}"
            pattern = f"%/{ms}/{ys}"

            def _int(s):
                try: return int(str(s or "0").replace(",",""))
                except: return 0

            # ── VENDAS ──────────────────────────────────────────────────────
            vendas = [dict(r) for r in self.conn.execute(
                "SELECT v.*, cl.nome as cli_nome FROM vendas v "
                "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
                "WHERE v.data_venda LIKE ?", (pattern,)).fetchall()]
            receita_v = sum(_int(v.get("valor_venda")) for v in vendas)
            custo_v = 0
            for v in vendas:
                cid2 = v.get("compra_id") or None
                cr2  = None
                if cid2:
                    cr2 = self.conn.execute("SELECT * FROM compras WHERE id=?", (cid2,)).fetchone()
                elif v.get("carro_id"):
                    cr2 = self.conn.execute(
                        "SELECT * FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                        (v["carro_id"],)).fetchone()
                if cr2:
                    cr2d = dict(cr2)
                    custo_v += _int(cr2d.get("valor")) + self._get_custo_total(cr2d["id"])
            # IS de Vendas CNF (em custos da compra)
            is_v = 0
            for v in vendas:
                if (v.get("nfi") or "SNF") != "CNF": continue
                cid2 = v.get("compra_id") or None
                if not cid2 and v.get("carro_id"):
                    cr2b = self.conn.execute(
                        "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                        (v["carro_id"],)).fetchone()
                    cid2 = cr2b[0] if cr2b else None
                if not cid2: continue
                ri_is = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos WHERE compra_id=? AND tipo_custo='IS'", (cid2,)).fetchone()
                is_v += int(ri_is[0]) if ri_is[0] else 0
            custo_v_sem = custo_v - is_v
            lucro_v = receita_v - custo_v
            marg_v  = (lucro_v / receita_v * 100) if receita_v else 0

            # ── OS ──────────────────────────────────────────────────────────
            os_rows = [dict(r) for r in self.conn.execute(
                "SELECT s.*, cl.nome as cli_nome FROM servicos s "
                "LEFT JOIN clientes cl ON s.cliente_id=cl.id "
                "WHERE s.data_servico LIKE ?", (pattern,)).fetchall()]
            receita_os = sum(_int(r.get("valor")) for r in os_rows)
            custo_os   = sum(
                self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_os WHERE servico_id=?", (r["id"],)).fetchone()[0]
                for r in os_rows)
            is_os = sum(
                self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_os WHERE servico_id=? AND tipo_custo='IS'", (r["id"],)
                ).fetchone()[0] or 0
                for r in os_rows)
            custo_os_sem = custo_os - is_os
            lucro_os = receita_os - custo_os
            marg_os  = (lucro_os / receita_os * 100) if receita_os else 0

            # ── SHAKEN ──────────────────────────────────────────────────────
            sk_rows_all = [dict(r) for r in self.conn.execute(
                "SELECT sk.*, c.carro as c_nome, cl.nome as cli_nome "
                "FROM shaken sk "
                "LEFT JOIN carros c ON sk.carro_id=c.id "
                "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
                "WHERE sk.data_registro LIKE ?", (pattern,)).fetchall()]
            sk_rows = [r for r in sk_rows_all if r.get("valor")]
            receita_sk = sum(_int(r.get("valor")) for r in sk_rows)
            custo_sk   = sum(
                self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_sk WHERE shaken_id=?", (r["id"],)).fetchone()[0]
                for r in sk_rows)
            is_sk = sum(
                self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_sk WHERE shaken_id=? AND tipo_custo='IS'", (r["id"],)
                ).fetchone()[0] or 0
                for r in sk_rows)
            custo_sk_sem = custo_sk - is_sk
            lucro_sk = receita_sk - custo_sk
            marg_sk  = (lucro_sk / receita_sk * 100) if receita_sk else 0

            # ── DESPESAS FIXAS ───────────────────────────────────────────────
            desp_rows = [dict(r) for r in self.conn.execute(
                "SELECT * FROM despesas_fixas WHERE data_ref=? OR recorrente=1",
                (mes_ref,)).fetchall()]
            total_desp = sum(_int(r.get("valor")) for r in desp_rows)

            # ── PARCELAS RECEBIDAS ───────────────────────────────────────────
            parc_rows = [dict(r) for r in self.conn.execute(
                "SELECT p.*, v.carro, cl.nome as cli_nome "
                "FROM pagamentos p "
                "JOIN vendas v ON p.venda_id=v.id "
                "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
                "WHERE p.data_pagamento LIKE ?", (pattern,)).fetchall()]
            total_parc = sum(_int(r.get("valor_pago")) for r in parc_rows)

            # ── RESUMO ───────────────────────────────────────────────────────
            is_total    = is_v + is_os + is_sk
            lucro_bruto = lucro_v + lucro_os + lucro_sk
            lucro_liq   = lucro_bruto - total_desp
            receita_tot = receita_v + receita_os + receita_sk
            custo_sem_tot = custo_v_sem + custo_os_sem + custo_sk_sem
            marg_geral  = (lucro_liq / receita_tot * 100) if receita_tot else 0

            # ── FLUXO DE CAIXA ───────────────────────────────────────────────
            vendas_avista  = [v for v in vendas if v.get("tipo_venda") in ("À Vista","A Vista")]
            fc_vendas_avista = sum(_int(v.get("valor_venda")) for v in vendas_avista)
            vendas_entrada = [v for v in vendas
                              if v.get("tipo_venda") not in ("À Vista","A Vista")
                              and _int(v.get("entrada"))]
            fc_entradas      = sum(_int(v.get("entrada")) for v in vendas_entrada)
            fc_volta         = sum(_int(v.get("volta_paga")) for v in vendas
                                   if _int(v.get("volta_paga")) > 0)
            fc_valor_troca   = sum(_int(v.get("valor_troca")) for v in vendas
                                   if _int(v.get("valor_troca")) > 0)
            fc_parcelas      = total_parc
            fc_os            = receita_os
            fc_sk            = receita_sk
            fc_total_entrada = fc_vendas_avista + fc_entradas + fc_volta + fc_parcelas + fc_os + fc_sk

            compras_mes = [dict(r) for r in self.conn.execute(
                "SELECT * FROM compras WHERE data_entrada LIKE ?", (pattern,)).fetchall()]
            fc_compras       = sum(_int(c.get("valor")) for c in compras_mes)
            fc_custos_entrada= sum(self._get_custo_total(c["id"]) for c in compras_mes)
            fc_custos_os     = sum(
                self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_os WHERE servico_id=? AND (data_custo LIKE ? OR data_custo IS NULL)",
                    (r["id"], pattern)).fetchone()[0]
                for r in os_rows)
            fc_custos_sk     = sum(
                self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM custos_sk WHERE shaken_id=? AND (data_custo LIKE ? OR data_custo IS NULL)",
                    (r["id"], pattern)).fetchone()[0]
                for r in sk_rows)
            fc_despesas      = total_desp
            fc_total_saida   = fc_despesas + fc_compras + fc_custos_entrada + fc_custos_os + fc_custos_sk + fc_valor_troca
            fc_saldo         = fc_total_entrada - fc_total_saida
            fc_cor           = COLORS["green"] if fc_saldo >= 0 else COLORS["red"]

            # ════════════════════════════════════════════════════════════════
            # LAYOUT COMPACTO
            # ════════════════════════════════════════════════════════════════
            G = COLORS["green"]; R = COLORS["red"]
            O = COLORS["orange"]; AC = COLORS["accent"]

            # ── Vendas ──
            sec_hdr("◆  VENDAS", G)
            kpi_row([
                ("Qtd",        str(len(vendas)),       COLORS["text_primary"]),
                ("Receita",    f"¥{receita_v:,}",      G),
                ("Custo s/IS", f"¥{custo_v_sem:,}",    O),
                ("IS (CNF)",   f"¥{is_v:,}",           "#8E44AD"),
                ("C. Total",   f"¥{custo_v:,}",        COLORS["red"]),
                ("Lucro",      f"¥{lucro_v:,}",        G if lucro_v>=0 else R),
                ("Margem",     f"{marg_v:.1f}%",       AC),
            ])

            # ── OS ──
            sec_hdr("⚙  ORDENS DE SERVIÇO", COLORS["blue"])
            kpi_row([
                ("Qtd",        str(len(os_rows)),      COLORS["text_primary"]),
                ("Receita",    f"¥{receita_os:,}",     COLORS["blue"]),
                ("Custo s/IS", f"¥{custo_os_sem:,}",   O),
                ("IS (CNF)",   f"¥{is_os:,}",          "#8E44AD"),
                ("C. Total",   f"¥{custo_os:,}",       COLORS["red"]),
                ("Lucro",      f"¥{lucro_os:,}",       G if lucro_os>=0 else R),
                ("Margem",     f"{marg_os:.1f}%",      AC),
            ])

            # ── Shaken ──
            sec_hdr("🔔  SHAKEN", "#D4AC0D")
            kpi_row([
                ("Qtd",        str(len(sk_rows)),      COLORS["text_primary"]),
                ("Receita",    f"¥{receita_sk:,}",     "#D4AC0D"),
                ("Custo s/IS", f"¥{custo_sk_sem:,}",   O),
                ("IS (CNF)",   f"¥{is_sk:,}",          "#8E44AD"),
                ("C. Total",   f"¥{custo_sk:,}",       COLORS["red"]),
                ("Lucro",      f"¥{lucro_sk:,}",       G if lucro_sk>=0 else R),
                ("Margem",     f"{marg_sk:.1f}%",      AC),
            ])

            # ── Despesas + Parcelas ──
            # Carteira global (não filtrada por mês)
            import datetime as _dtimp
            hoje_dt = _dtimp.date.today()

            fin_todas = [dict(r) for r in self.conn.execute(
                "SELECT v.*, c.nome as cliente_nome FROM vendas v "
                "LEFT JOIN clientes c ON v.cliente_id=c.id "
                "WHERE v.tipo_venda IN ('Venda Parcelada','Com Troca','Com Troca e Volta') "
                "ORDER BY v.id DESC").fetchall()]

            cart_total_receber = 0
            cart_total_pago    = 0  # somente parcelas pagas
            cart_inadimp       = 0
            cart_saldo_devedor = 0

            for fv in fin_todas:
                pagas_fv = fv.get("parcelas_pagas") or 0
                tot_p_fv = fv.get("num_parcelas") or 0
                quitado  = tot_p_fv > 0 and pagas_fv >= tot_p_fv

                # Total pago = somente parcelas (tabela pagamentos)
                total_pago_fv = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                    "FROM pagamentos WHERE venda_id=?", (fv["id"],)).fetchone()[0]
                cart_total_pago += total_pago_fv

                vv_fv    = _int(fv.get("valor_venda"))
                et_fv    = _int(fv.get("entrada"))
                troca_fv = _int(fv.get("valor_troca"))
                saldo_fv = max(0, vv_fv - et_fv - troca_fv - total_pago_fv)

                if not quitado:
                    cart_total_receber += saldo_fv
                    cart_saldo_devedor += saldo_fv

                    parc_val_fv = _int(fv.get("parcela_mensal"))
                    data_1a_fv  = fv.get("data_primeira_parc")
                    if parc_val_fv > 0 and data_1a_fv:
                        try:
                            pts2 = data_1a_fv.split("/")
                            m0i  = int(pts2[1]); a0i = int(pts2[2])
                            for idx2 in range(pagas_fv, tot_p_fv):
                                mt2 = m0i + idx2
                                ay2 = a0i + (mt2 - 1) // 12
                                mo2 = ((mt2 - 1) % 12) + 1
                                try:
                                    d_parc = _dtimp.date(ay2, mo2, int(pts2[0]))
                                except Exception:
                                    d_parc = _dtimp.date(ay2, mo2, 1)
                                if d_parc < hoje_dt:
                                    is_ult = (idx2 == tot_p_fv - 1)
                                    v_ult  = _int(fv.get("valor_ultima_parc")) or parc_val_fv
                                    cart_inadimp += v_ult if is_ult else parc_val_fv
                        except Exception:
                            pass

            # ── Despesas Fixas (separado) ──
            sec_hdr("💸  DESPESAS FIXAS", R)
            kpi_row([
                ("Registros",  str(len(desp_rows)),  COLORS["text_primary"]),
                ("Total",      f"¥{total_desp:,}",   R),
            ])

            # ── Parcelas ──
            # "Total a receber no mês" = parcelas com vencimento no mês selecionado
            parc_a_receber_mes = 0
            for fv in fin_todas:
                pagas_fv = fv.get("parcelas_pagas") or 0
                tot_p_fv = fv.get("num_parcelas") or 0
                if tot_p_fv <= 0 or pagas_fv >= tot_p_fv:
                    continue
                parc_val_fv = _int(fv.get("parcela_mensal"))
                data_1a_fv  = fv.get("data_primeira_parc")
                if parc_val_fv <= 0 or not data_1a_fv:
                    continue
                try:
                    pts3 = data_1a_fv.split("/")
                    m0r  = int(pts3[1]); a0r = int(pts3[2])
                    for idx3 in range(pagas_fv, tot_p_fv):
                        mt3 = m0r + idx3
                        ay3 = a0r + (mt3 - 1) // 12
                        mo3 = ((mt3 - 1) % 12) + 1
                        if ay3 == a and mo3 == m:
                            is_ult3 = (idx3 == tot_p_fv - 1)
                            v_ult3  = _int(fv.get("valor_ultima_parc")) or parc_val_fv
                            parc_a_receber_mes += v_ult3 if is_ult3 else parc_val_fv
                except Exception:
                    pass

            sec_hdr("◎  PARCELAS", AC)
            kpi_row([
                ("A Receber no Mês",    f"¥{parc_a_receber_mes:,}", AC),
                ("Parcelas Pagas",      f"¥{cart_total_pago:,}",    G),
                ("Inadimplência",       f"¥{cart_inadimp:,}",       R if cart_inadimp else COLORS["text_muted"]),
                ("Saldo Devedor Total", f"¥{cart_saldo_devedor:,}", R if cart_saldo_devedor else G),
            ])

            # ── Resumo Consolidado ──
            sec_hdr("📊  RESUMO CONSOLIDADO", COLORS["text_primary"])
            kpi_row([
                ("Receita Total",  f"¥{receita_tot:,}",    G),
                ("Custo s/IS",     f"¥{custo_sem_tot:,}",  O),
                ("IS Total",       f"¥{is_total:,}",       "#8E44AD"),
                ("Lucro Bruto",    f"¥{lucro_bruto:,}",    G if lucro_bruto>=0 else R),
                ("(-) Despesas",   f"¥{total_desp:,}",     R),
                ("Lucro Líquido",  f"¥{lucro_liq:,}",      G if lucro_liq>=0 else R),
                ("Margem Geral",   f"{marg_geral:.1f}%",   AC),
            ])

            # ── Fluxo de Caixa — 2 colunas ──
            sec_hdr("💰  FLUXO DE CAIXA", COLORS["blue"])

            fc_outer = tk.Frame(self._rm_inner, bg=COLORS["bg_card"],
                                highlightthickness=1, highlightbackground=COLORS["border"])
            fc_outer.pack(fill="x", pady=(2,0))

            col_ent = tk.Frame(fc_outer, bg=COLORS["bg_card"])
            col_ent.pack(side="left", fill="both", expand=True)
            col_sai = tk.Frame(fc_outer, bg=COLORS["bg_card"])
            col_sai.pack(side="right", fill="both", expand=True)
            # Divisor vertical
            tk.Frame(fc_outer, bg=COLORS["border"], width=1).place(relx=0.5, rely=0, relheight=1)

            # Cabeçalhos de coluna
            hdr_e = tk.Frame(col_ent, bg=COLORS["green"])
            hdr_e.pack(fill="x")
            tk.Label(hdr_e, text="  ENTRADAS", font=("Helvetica",8,"bold"),
                     bg=COLORS["green"], fg="white").pack(side="left", pady=3, padx=4)

            hdr_s = tk.Frame(col_sai, bg=COLORS["red"])
            hdr_s.pack(fill="x")
            tk.Label(hdr_s, text="  SAÍDAS", font=("Helvetica",8,"bold"),
                     bg=COLORS["red"], fg="white").pack(side="left", pady=3, padx=4)

            def fc_col_line(col, lbl, val, val_cor, i):
                bg = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
                r = tk.Frame(col, bg=bg)
                r.pack(fill="x")
                tk.Label(r, text=lbl, font=("Helvetica",8), bg=bg,
                         fg=COLORS["text_secondary"], anchor="w"
                         ).pack(side="left", padx=8, pady=2)
                tk.Label(r, text=val, font=("Helvetica",8,"bold"), bg=bg,
                         fg=val_cor, anchor="e"
                         ).pack(side="right", padx=8, pady=2)

            ent_items = [
                ("Vendas à Vista",          f"¥{fc_vendas_avista:,}"),
                ("Entradas financ./troca",  f"¥{fc_entradas:,}"),
                ("Volta de Troca recebida", f"¥{fc_volta:,}"),
                ("Parcelas pagas",          f"¥{fc_parcelas:,}"),
                ("Receita OS",              f"¥{fc_os:,}"),
                ("Receita Shaken",          f"¥{fc_sk:,}"),
            ]
            sai_items = [
                ("Compras de Carros",   f"¥{fc_compras:,}"),
                ("Custos de Entrada",   f"¥{fc_custos_entrada:,}"),
                ("Valor de Troca",      f"¥{fc_valor_troca:,}"),
                ("Custos de OS",        f"¥{fc_custos_os:,}"),
                ("Custos de Shaken",    f"¥{fc_custos_sk:,}"),
                ("Despesas Fixas",      f"¥{fc_despesas:,}"),
            ]
            for i, (lbl, val) in enumerate(ent_items):
                fc_col_line(col_ent, lbl, val, G, i)
            for i, (lbl, val) in enumerate(sai_items):
                fc_col_line(col_sai, lbl, val, R, i)

            # Totais lado a lado
            tot_row = tk.Frame(self._rm_inner, bg=COLORS["bg_content"])
            tot_row.pack(fill="x")
            tot_e = tk.Frame(tot_row, bg=COLORS["bg_content"])
            tot_e.pack(side="left", fill="x", expand=True)
            tot_s = tk.Frame(tot_row, bg=COLORS["bg_content"])
            tot_s.pack(side="right", fill="x", expand=True)
            tk.Label(tot_e, text="TOTAL ENTRADAS", font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_content"], fg=G).pack(side="left", padx=8, pady=4)
            tk.Label(tot_e, text=f"¥{fc_total_entrada:,}", font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=G).pack(side="right", padx=8, pady=4)
            tk.Label(tot_s, text="TOTAL SAÍDAS", font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_content"], fg=R).pack(side="left", padx=8, pady=4)
            tk.Label(tot_s, text=f"¥{fc_total_saida:,}", font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=R).pack(side="right", padx=8, pady=4)

            # Saldo
            saldo_bg = COLORS["green"] if fc_saldo >= 0 else COLORS["red"]
            saldo_row = tk.Frame(self._rm_inner, bg=saldo_bg)
            saldo_row.pack(fill="x", pady=(0,12))
            tk.Label(saldo_row, text="  SALDO DO MÊS",
                     font=("Helvetica",9,"bold"), bg=saldo_bg, fg="white"
                     ).pack(side="left", padx=8, pady=5)
            tk.Label(saldo_row, text=f"¥{fc_saldo:,}",
                     font=("Helvetica",11,"bold"), bg=saldo_bg, fg="white"
                     ).pack(side="right", padx=12, pady=5)

            # cache para PDF
            self._rm_data_cache = {
                "mes_ref": f"{MESES_PT[m-1]} {a}",
                "vendas": vendas, "receita_v": receita_v,
                "custo_v": custo_v, "lucro_v": lucro_v, "marg_v": marg_v,
                "os_rows": os_rows, "receita_os": receita_os,
                "lucro_os": lucro_os, "marg_os": marg_os,
                "sk_rows": sk_rows, "receita_sk": receita_sk, "custo_sk": custo_sk,
                "lucro_sk": lucro_sk, "marg_sk": marg_sk,
                "desp_rows": desp_rows, "total_desp": total_desp,
                "parc_rows": parc_rows, "total_parc": total_parc,
                "receita_tot": receita_tot,
                "custo_v_sem": custo_v_sem, "is_v": is_v,
                "custo_os_sem": custo_os_sem, "is_os": is_os,
                "custo_sk_sem": custo_sk_sem, "is_sk": is_sk,
                "is_total": is_total, "custo_sem_tot": custo_sem_tot,
                "lucro_bruto": lucro_bruto, "lucro_liq": lucro_liq, "marg_geral": marg_geral,
                "fc_vendas_avista": fc_vendas_avista, "fc_entradas": fc_entradas,
                "fc_volta": fc_volta, "fc_valor_troca": fc_valor_troca, "fc_parcelas": fc_parcelas,
                "fc_os": fc_os, "fc_sk": fc_sk, "fc_total_entrada": fc_total_entrada,
                "fc_compras": fc_compras, "fc_custos_entrada": fc_custos_entrada,
                "fc_custos_os": fc_custos_os, "fc_custos_sk": fc_custos_sk,
                "fc_despesas": fc_despesas, "fc_total_saida": fc_total_saida,
                "fc_saldo": fc_saldo,
                # carteira
                "cart_total_receber": cart_total_receber,
                "cart_total_pago":    cart_total_pago,
                "cart_inadimp":       cart_inadimp,
                "cart_saldo_devedor": cart_saldo_devedor,
                "parc_a_receber_mes": parc_a_receber_mes,
            }
        def _exportar_pdf():
            from tkinter import messagebox as _mb, filedialog as _fd
            import datetime as _dt
            d = self._rm_data_cache
            if not d:
                _refresh()
                d = self._rm_data_cache
            m = self._rm_mes.get(); a = self._rm_ano.get()
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                Spacer, Paragraph, HRFlowable)
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors as rl_colors
                from reportlab.lib.units import cm
            except ImportError:
                _mb.showerror("PDF", "reportlab nao instalado.\nExecute: pip install reportlab")
                return

            path = _fd.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=f"relatorio_{a}_{m:02d}.pdf",
                title="Salvar Relatorio Mensal")
            if not path:
                return

            hoje_str = _dt.date.today().strftime("%d/%m/%Y")
            RED    = rl_colors.HexColor("#C00000")
            DARK   = rl_colors.HexColor("#111827")
            GRAY   = rl_colors.HexColor("#D1D5DB")
            LGRAY  = rl_colors.HexColor("#F9FAFB")
            GREEN  = rl_colors.HexColor("#16A34A")
            ORANGE = rl_colors.HexColor("#EA580C")
            BLUE   = rl_colors.HexColor("#1D4ED8")
            GOLD   = rl_colors.HexColor("#B45309")
            MUTED  = rl_colors.HexColor("#6B7280")

            styles = getSampleStyleSheet()
            def P(txt, size=8, bold=False, color=DARK, align="LEFT"):
                return Paragraph(
                    f"<b>{txt}</b>" if bold else str(txt),
                    ParagraphStyle("_p", parent=styles["Normal"],
                                   fontSize=size, textColor=color, leading=size+3,
                                   alignment={"LEFT":0,"CENTER":1,"RIGHT":2}.get(align,0)))

            doc = SimpleDocTemplate(path, pagesize=A4,
                                    rightMargin=1.5*cm, leftMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            story = []

            # Cabecalho
            story.append(P("KM CARS - Relatorio Mensal", size=16, bold=True, color=RED))
            story.append(P(f"{d.get('mes_ref','')}   |   Gerado em: {hoje_str}", size=8, color=MUTED))
            story.append(Spacer(1, 0.2*cm))
            story.append(HRFlowable(width="100%", thickness=2, color=RED))
            story.append(Spacer(1, 0.3*cm))

            def kpi_table(title, title_color, items):
                story.append(P(title, size=9, bold=True, color=title_color))
                story.append(Spacer(1, 0.1*cm))
                lbl_row = [P(lbl, size=7, color=MUTED)         for lbl,val,vc in items]
                val_row = [P(val, size=9, bold=True, color=vc) for lbl,val,vc in items]
                n = len(items)
                cw = [17.0*cm / n] * n
                tbl = Table([lbl_row, val_row], colWidths=cw)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,-1), LGRAY),
                    ("LINEBELOW",     (0,0), (-1, 0), 0.3, GRAY),
                    ("TOPPADDING",    (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ("LEFTPADDING",   (0,0), (-1,-1), 6),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 0.25*cm))

            PURPLE = rl_colors.HexColor("#8E44AD")
            kpi_table("VENDAS", GREEN, [
                ("Qtd",       str(len(d.get("vendas",[]))),             DARK),
                ("Receita",   f"Y {d.get('receita_v',0):,}",            GREEN),
                ("Custo s/IS",f"Y {d.get('custo_v_sem',0):,}",          ORANGE),
                ("IS (CNF)",  f"Y {d.get('is_v',0):,}",                 PURPLE),
                ("C. Total",  f"Y {d.get('custo_v',0):,}",              RED),
                ("Lucro",     f"Y {d.get('lucro_v',0):,}",              GREEN if d.get('lucro_v',0)>=0 else RED),
                ("Margem",    f"{d.get('marg_v',0):.1f}%",              BLUE),
            ])
            kpi_table("ORDENS DE SERVICO", BLUE, [
                ("Qtd",       str(len(d.get("os_rows",[]))),             DARK),
                ("Receita",   f"Y {d.get('receita_os',0):,}",            BLUE),
                ("Custo s/IS",f"Y {d.get('custo_os_sem',0):,}",          ORANGE),
                ("IS (CNF)",  f"Y {d.get('is_os',0):,}",                 PURPLE),
                ("C. Total",  f"Y {d.get('custo_os',0):,}",              RED),
                ("Lucro",     f"Y {d.get('lucro_os',0):,}",              GREEN if d.get('lucro_os',0)>=0 else RED),
                ("Margem",    f"{d.get('marg_os',0):.1f}%",              BLUE),
            ])
            kpi_table("SHAKEN", GOLD, [
                ("Qtd",       str(len(d.get("sk_rows",[]))),             DARK),
                ("Receita",   f"Y {d.get('receita_sk',0):,}",            GOLD),
                ("Custo s/IS",f"Y {d.get('custo_sk_sem',0):,}",          ORANGE),
                ("IS (CNF)",  f"Y {d.get('is_sk',0):,}",                 PURPLE),
                ("C. Total",  f"Y {d.get('custo_sk',0):,}",              RED),
                ("Lucro",     f"Y {d.get('lucro_sk',0):,}",              GREEN if d.get('lucro_sk',0)>=0 else RED),
                ("Margem",    f"{d.get('marg_sk',0):.1f}%",              BLUE),
            ])
            cart_inadimp_v = d.get('cart_inadimp', 0)
            cart_saldo_v   = d.get('cart_saldo_devedor', 0)

            kpi_table("DESPESAS FIXAS", RED, [
                ("Registros", str(len(d.get("desp_rows",[]))),  DARK),
                ("Total",     f"Y {d.get('total_desp',0):,}",   RED),
            ])
            kpi_table("PARCELAS", rl_colors.HexColor("#7C3AED"), [
                ("A Receber no Mes",    f"Y {d.get('parc_a_receber_mes',0):,}",
                 rl_colors.HexColor("#7C3AED")),
                ("Parcelas Pagas",      f"Y {d.get('cart_total_pago',0):,}",     GREEN),
                ("Inadimplencia",       f"Y {cart_inadimp_v:,}",
                 RED if cart_inadimp_v > 0 else MUTED),
                ("Saldo Devedor Total", f"Y {cart_saldo_v:,}",
                 RED if cart_saldo_v > 0 else GREEN),
            ])
            story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
            story.append(Spacer(1, 0.1*cm))
            kpi_table("RESUMO CONSOLIDADO", DARK, [
                ("Receita Total", f"Y {d.get('receita_tot',0):,}",      GREEN),
                ("Custo s/IS",    f"Y {d.get('custo_sem_tot',0):,}",    ORANGE),
                ("IS Total",      f"Y {d.get('is_total',0):,}",         PURPLE),
                ("Lucro Bruto",   f"Y {d.get('lucro_bruto',0):,}",      GREEN if d.get('lucro_bruto',0)>=0 else RED),
                ("(-) Despesas",  f"Y {d.get('total_desp',0):,}",       RED),
                ("Lucro Liquido", f"Y {d.get('lucro_liq',0):,}",        GREEN if d.get('lucro_liq',0)>=0 else RED),
                ("Margem Geral",  f"{d.get('marg_geral',0):.1f}%",      BLUE),
            ])

            # ── Fluxo de Caixa — 2 colunas reais ──────────────────────────
            story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
            story.append(Spacer(1, 0.2*cm))
            story.append(P("FLUXO DE CAIXA", size=9, bold=True, color=BLUE))
            story.append(Spacer(1, 0.1*cm))

            fc_saldo_v  = d.get("fc_saldo", 0)
            saldo_c_pdf = GREEN if fc_saldo_v >= 0 else RED
            saldo_bg    = rl_colors.HexColor("#DCFCE7") if fc_saldo_v >= 0 else rl_colors.HexColor("#FEE2E2")

            ent_items_pdf = [
                ("Vendas a Vista",         f"Y {d.get('fc_vendas_avista',0):,}"),
                ("Entradas financ./troca", f"Y {d.get('fc_entradas',0):,}"),
                ("Volta de Troca recebida",f"Y {d.get('fc_volta',0):,}"),
                ("Parcelas pagas",         f"Y {d.get('fc_parcelas',0):,}"),
                ("Receita OS",             f"Y {d.get('fc_os',0):,}"),
                ("Receita Shaken",         f"Y {d.get('fc_sk',0):,}"),
            ]
            sai_items_pdf = [
                ("Compras de Carros",  f"Y {d.get('fc_compras',0):,}"),
                ("Custos de Entrada",  f"Y {d.get('fc_custos_entrada',0):,}"),
                ("Valor de Troca",     f"Y {d.get('fc_valor_troca',0):,}"),
                ("Custos de OS",       f"Y {d.get('fc_custos_os',0):,}"),
                ("Custos de Shaken",   f"Y {d.get('fc_custos_sk',0):,}"),
                ("Despesas Fixas",     f"Y {d.get('fc_despesas',0):,}"),
            ]

            # Cabeçalhos ENTRADAS | SAÍDAS
            hdr_row = [[
                P("  ENTRADAS", size=8, bold=True, color=GREEN),
                P("SAIDAS  ", size=8, bold=True, color=RED, align="RIGHT"),
            ]]
            hdr_tbl = Table(hdr_row, colWidths=[8.5*cm, 8.5*cm])
            hdr_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (0,0), rl_colors.HexColor("#DCFCE7")),
                ("BACKGROUND",    (1,0), (1,0), rl_colors.HexColor("#FEE2E2")),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                ("LINEBELOW",     (0,0), (-1,0), 0.5, GRAY),
            ]))
            story.append(hdr_tbl)

            # Linhas de itens lado a lado (preenchendo com vazios se necessário)
            max_rows = max(len(ent_items_pdf), len(sai_items_pdf))
            fc_body_rows = []
            for i in range(max_rows):
                e_lbl = ent_items_pdf[i][0] if i < len(ent_items_pdf) else ""
                e_val = ent_items_pdf[i][1] if i < len(ent_items_pdf) else ""
                s_lbl = sai_items_pdf[i][0] if i < len(sai_items_pdf) else ""
                s_val = sai_items_pdf[i][1] if i < len(sai_items_pdf) else ""
                fc_body_rows.append([
                    P(e_lbl, size=8, color=MUTED),
                    P(e_val, size=8, bold=True, color=GREEN, align="RIGHT"),
                    P(s_lbl, size=8, color=MUTED),
                    P(s_val, size=8, bold=True, color=RED, align="RIGHT"),
                ])

            body_tbl = Table(fc_body_rows, colWidths=[5.5*cm, 3*cm, 5.5*cm, 3*cm])
            body_tbl.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [rl_colors.white, LGRAY]),
                ("LINEAFTER",      (1,0), (1,-1), 0.5, GRAY),
                ("TOPPADDING",     (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",  (0,0), (-1,-1), 3),
                ("LEFTPADDING",    (0,0), (-1,-1), 6),
                ("RIGHTPADDING",   (0,0), (-1,-1), 6),
                ("ALIGN",          (1,0), (1,-1), "RIGHT"),
                ("ALIGN",          (3,0), (3,-1), "RIGHT"),
            ]))
            story.append(body_tbl)

            # Totais na mesma linha (TOTAL ENTRADAS | TOTAL SAÍDAS)
            tot_row_pdf = [[
                P("TOTAL ENTRADAS", size=8, bold=True, color=GREEN),
                P(f"Y {d.get('fc_total_entrada',0):,}", size=9, bold=True, color=GREEN, align="RIGHT"),
                P("TOTAL SAIDAS", size=8, bold=True, color=RED),
                P(f"Y {d.get('fc_total_saida',0):,}", size=9, bold=True, color=RED, align="RIGHT"),
            ]]
            tot_tbl = Table(tot_row_pdf, colWidths=[5.5*cm, 3*cm, 5.5*cm, 3*cm])
            tot_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (1,0), rl_colors.HexColor("#DCFCE7")),
                ("BACKGROUND",    (2,0), (3,0), rl_colors.HexColor("#FEE2E2")),
                ("LINEABOVE",     (0,0), (-1,0), 1, GRAY),
                ("LINEBELOW",     (0,0), (-1,0), 1, GRAY),
                ("LINEAFTER",     (1,0), (1,0), 0.5, GRAY),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                ("ALIGN",         (1,0), (1,0), "RIGHT"),
                ("ALIGN",         (3,0), (3,0), "RIGHT"),
            ]))
            story.append(tot_tbl)
            story.append(Spacer(1, 0.15*cm))

            # Saldo do mês
            saldo_tbl = Table([[
                P("SALDO DO MES", size=10, bold=True, color=saldo_c_pdf),
                P(f"Y {fc_saldo_v:,}", size=11, bold=True, color=saldo_c_pdf, align="RIGHT"),
            ]], colWidths=[9*cm, 8*cm])
            saldo_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), saldo_bg),
                ("LINEABOVE",     (0,0), (-1, 0), 2, saldo_c_pdf),
                ("LINEBELOW",     (0,0), (-1, 0), 2, saldo_c_pdf),
                ("TOPPADDING",    (0,0), (-1,-1), 7),
                ("BOTTOMPADDING", (0,0), (-1,-1), 7),
                ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ]))
            story.append(saldo_tbl)
            story.append(Spacer(1, 0.3*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
            story.append(Spacer(1, 0.1*cm))
            story.append(P(f"KM Cars - Fujinomiya, Japan  |  Gerado em {hoje_str}",
                           size=7, color=MUTED, align="CENTER"))

            doc.build(story)
            _mb.showinfo("PDF Gerado", f"Relatorio salvo:\n{path}")
            try:
                import subprocess, sys
                if sys.platform == "win32":   os.startfile(path)
                elif sys.platform == "darwin": subprocess.call(["open", path])
                else:                          subprocess.call(["xdg-open", path])
            except Exception:
                pass
        _refresh()

    def _build_relatorio_anual(self, parent):
        """Relatório anual consolidado com exportação PDF."""
        import datetime
        hoje = datetime.date.today()

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=18, pady=14)

        # ── Cabeçalho + navegação ────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 12))

        tk.Label(hdr, text="📅  Relatório Anual", font=("Helvetica",14,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=lambda: _refresh()).pack(side="left", padx=(10,0))

        self._ra_ano = tk.IntVar(value=hoje.year)

        nav = tk.Frame(hdr, bg=COLORS["bg_content"])
        nav.pack(side="right")

        self._ra_lbl_ano = tk.Label(nav, text=str(hoje.year),
            font=("Helvetica",13,"bold"), width=8,
            bg=COLORS["bg_content"], fg=COLORS["text_primary"])

        def _prev():
            self._ra_ano.set(self._ra_ano.get() - 1)
            self._ra_lbl_ano.configure(text=str(self._ra_ano.get()))
            _refresh()

        def _next():
            self._ra_ano.set(self._ra_ano.get() + 1)
            self._ra_lbl_ano.configure(text=str(self._ra_ano.get()))
            _refresh()

        tk.Button(nav, text="◀", font=("Helvetica",11,"bold"),
                  bg=COLORS["bg_card"], fg=COLORS["accent"],
                  relief="flat", cursor="hand2", padx=8,
                  command=_prev).pack(side="left", ipady=3)
        self._ra_lbl_ano.pack(side="left", padx=4)
        tk.Button(nav, text="▶", font=("Helvetica",11,"bold"),
                  bg=COLORS["bg_card"], fg=COLORS["accent"],
                  relief="flat", cursor="hand2", padx=8,
                  command=_next).pack(side="left")
        tk.Button(nav, text="📄  Exportar PDF", font=("Helvetica",9,"bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=10, pady=4, command=lambda: _exportar_pdf()
                  ).pack(side="left", padx=(16,0), ipady=2)

        # ── Container scrollável ─────────────────────────────────────────────
        outer = tk.Frame(root, bg=COLORS["bg_content"])
        outer.pack(fill="both", expand=True)
        vsb = tk.Scrollbar(outer, orient="vertical")
        cv  = tk.Canvas(outer, bg=COLORS["bg_content"], highlightthickness=0,
                        yscrollcommand=vsb.set)
        vsb.configure(command=cv.yview)
        vsb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True)
        self._ra_inner = tk.Frame(cv, bg=COLORS["bg_content"])
        cw = cv.create_window((0,0), window=self._ra_inner, anchor="nw")
        self._ra_inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>", lambda e: cv.itemconfig(cw, width=e.width))
        cv.bind("<Enter>", lambda e: cv.bind_all("<MouseWheel>",
            lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        cv.bind("<Leave>", lambda e: cv.unbind_all("<MouseWheel>"))

        # ── Layout helpers ───────────────────────────────────────────────────
        def sec_hdr(txt, color=COLORS["blue"]):
            row = tk.Frame(self._ra_inner, bg=COLORS["bg_content"])
            row.pack(fill="x", pady=(10,1))
            tk.Frame(row, bg=color, width=4, height=18).pack(side="left", padx=(0,6))
            tk.Label(row, text=txt, font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=color).pack(side="left")

        def kpi_row(items):
            row = tk.Frame(self._ra_inner, bg=COLORS["bg_card"],
                           highlightthickness=1, highlightbackground=COLORS["border"])
            row.pack(fill="x", pady=(2,0))
            for i, (lbl, val, cor) in enumerate(items):
                if i > 0:
                    tk.Frame(row, bg=COLORS["border"], width=1).pack(side="left", fill="y", pady=4)
                cell = tk.Frame(row, bg=COLORS["bg_card"])
                cell.pack(side="left", expand=True, fill="x", padx=8, pady=5)
                tk.Label(cell, text=lbl, font=("Helvetica",7),
                         bg=COLORS["bg_card"], fg=COLORS["text_muted"], anchor="w").pack(anchor="w")
                tk.Label(cell, text=val, font=("Helvetica",10,"bold"),
                         bg=COLORS["bg_card"], fg=cor, anchor="w").pack(anchor="w")

        def _int(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        self._ra_data_cache = {}

        MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun",
                    "Jul","Ago","Set","Out","Nov","Dez"]

        def _refresh():
            for w in self._ra_inner.winfo_children():
                w.destroy()

            a = self._ra_ano.get()
            ys = str(a)

            # ── Busca todos os meses do ano ──────────────────────────────────
            receita_v_tot = custo_v_tot = is_v_tot = 0
            receita_os_tot = custo_os_tot = is_os_tot = 0
            receita_sk_tot = custo_sk_tot = is_sk_tot = 0
            total_desp_tot = 0
            cnt_v = cnt_os = cnt_sk = 0
            mensal = []  # (mes, rec_v, rec_os, rec_sk, lucro_tot)

            for m in range(1, 13):
                ms = f"{m:02d}"
                pattern = f"%/{ms}/{ys}"

                # Vendas
                vendas_m = [dict(r) for r in self.conn.execute(
                    "SELECT v.*, ca.carro as carro, ca.ano as c_ano, ca.placa as c_placa, "
                    "ca.cor as c_cor, cl.nome as cli_nome "
                    "FROM vendas v "
                    "LEFT JOIN carros ca ON v.carro_id=ca.id "
                    "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
                    "WHERE v.data_venda LIKE ?", (pattern,)).fetchall()]
                cnt_v += len(vendas_m)
                for v in vendas_m:
                    receita_v_tot += _int(v.get("valor_venda"))
                    cid2 = v.get("compra_id")
                    if not cid2 and v.get("carro_id"):
                        cr2 = self.conn.execute(
                            "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                            (v["carro_id"],)).fetchone()
                        cid2 = cr2[0] if cr2 else None
                    if cid2:
                        custo_v_tot += _int(self.conn.execute(
                            "SELECT valor FROM compras WHERE id=?", (cid2,)).fetchone()[0] if
                            self.conn.execute("SELECT valor FROM compras WHERE id=?", (cid2,)).fetchone() else 0)
                        custo_v_tot += self._get_custo_total(cid2)
                        ri_is2 = self.conn.execute(
                            "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                            "FROM custos WHERE compra_id=? AND tipo_custo='IS'", (cid2,)).fetchone()
                        is_v_tot += int(ri_is2[0]) if ri_is2[0] else 0

                # OS
                os_m = [dict(r) for r in self.conn.execute(
                    "SELECT * FROM servicos WHERE data_servico LIKE ?", (pattern,)).fetchall()]
                cnt_os += len(os_m)
                for r in os_m:
                    receita_os_tot += _int(r.get("valor"))
                    custo_os_tot += self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_os WHERE servico_id=?", (r["id"],)).fetchone()[0] or 0
                    is_os_tot += self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_os WHERE servico_id=? AND tipo_custo='IS'", (r["id"],)).fetchone()[0] or 0

                # Shaken
                sk_m = [dict(r) for r in self.conn.execute(
                    "SELECT * FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL AND valor != ''",
                    (pattern,)).fetchall()]
                cnt_sk += len(sk_m)
                for r in sk_m:
                    receita_sk_tot += _int(r.get("valor"))
                    custo_sk_tot += self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_sk WHERE shaken_id=?", (r["id"],)).fetchone()[0] or 0
                    is_sk_tot += self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_sk WHERE shaken_id=? AND tipo_custo='IS'", (r["id"],)).fetchone()[0] or 0

                # Despesas fixas do mês
                mes_ref_m = f"{ms}/{ys}"
                desp_m = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM despesas_fixas WHERE data_ref=? OR recorrente=1", (mes_ref_m,)).fetchone()[0] or 0
                total_desp_tot += _int(desp_m)

                rec_m = receita_v_tot  # running totals não servem por mês, recompute
                luc_m = (_int(self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor_venda,',','') AS INTEGER)),0) "
                    "FROM vendas WHERE data_venda LIKE ?", (pattern,)).fetchone()[0])
                    + _int(self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM servicos WHERE data_servico LIKE ?", (pattern,)).fetchone()[0])
                    + _int(self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL", (pattern,)).fetchone()[0]))
                mensal.append((MESES_PT[m-1], luc_m))

            # Totais derivados
            is_total        = is_v_tot + is_os_tot + is_sk_tot
            custo_v_sem     = custo_v_tot - is_v_tot
            custo_os_sem    = custo_os_tot - is_os_tot
            custo_sk_sem    = custo_sk_tot - is_sk_tot
            custo_sem_tot   = custo_v_sem + custo_os_sem + custo_sk_sem
            lucro_v         = receita_v_tot - custo_v_tot
            lucro_os        = receita_os_tot - custo_os_tot
            lucro_sk        = receita_sk_tot - custo_sk_tot
            receita_tot     = receita_v_tot + receita_os_tot + receita_sk_tot
            lucro_bruto     = lucro_v + lucro_os + lucro_sk
            lucro_liq       = lucro_bruto - total_desp_tot
            marg_v          = (lucro_v / receita_v_tot * 100) if receita_v_tot else 0
            marg_os         = (lucro_os / receita_os_tot * 100) if receita_os_tot else 0
            marg_sk         = (lucro_sk / receita_sk_tot * 100) if receita_sk_tot else 0
            marg_geral      = (lucro_liq / receita_tot * 100) if receita_tot else 0

            G = COLORS["green"]; R = COLORS["red"]
            O = COLORS["orange"]; AC = COLORS["accent"]
            P_ = "#8E44AD"

            # ── Vendas ──
            sec_hdr(f"◆  VENDAS  ({cnt_v} vendas)", G)
            kpi_row([
                ("Receita",    f"¥{receita_v_tot:,}",  G),
                ("Custo s/IS", f"¥{custo_v_sem:,}",    O),
                ("IS (CNF)",   f"¥{is_v_tot:,}",       P_),
                ("C. Total",   f"¥{custo_v_tot:,}",    COLORS["red"]),
                ("Lucro",      f"¥{lucro_v:,}",        G if lucro_v>=0 else R),
                ("Margem",     f"{marg_v:.1f}%",       AC),
            ])

            # ── OS ──
            sec_hdr(f"⚙  ORDENS DE SERVIÇO  ({cnt_os} OS)", COLORS["blue"])
            kpi_row([
                ("Receita",    f"¥{receita_os_tot:,}", COLORS["blue"]),
                ("Custo s/IS", f"¥{custo_os_sem:,}",   O),
                ("IS (CNF)",   f"¥{is_os_tot:,}",      P_),
                ("C. Total",   f"¥{custo_os_tot:,}",   COLORS["red"]),
                ("Lucro",      f"¥{lucro_os:,}",       G if lucro_os>=0 else R),
                ("Margem",     f"{marg_os:.1f}%",      AC),
            ])

            # ── Shaken ──
            sec_hdr(f"🔔  SHAKEN  ({cnt_sk} registros)", "#D4AC0D")
            kpi_row([
                ("Receita",    f"¥{receita_sk_tot:,}", "#D4AC0D"),
                ("Custo s/IS", f"¥{custo_sk_sem:,}",   O),
                ("IS (CNF)",   f"¥{is_sk_tot:,}",      P_),
                ("C. Total",   f"¥{custo_sk_tot:,}",   COLORS["red"]),
                ("Lucro",      f"¥{lucro_sk:,}",       G if lucro_sk>=0 else R),
                ("Margem",     f"{marg_sk:.1f}%",      AC),
            ])

            # ── Despesas ──
            sec_hdr("💸  DESPESAS FIXAS (anual)", R)
            kpi_row([("Total Ano", f"¥{total_desp_tot:,}", R)])

            # ── Resumo ──
            sec_hdr("📊  RESUMO CONSOLIDADO DO ANO", COLORS["text_primary"])
            kpi_row([
                ("Receita Total",  f"¥{receita_tot:,}",    G),
                ("Custo s/IS",     f"¥{custo_sem_tot:,}",  O),
                ("IS Total",       f"¥{is_total:,}",       P_),
                ("Lucro Bruto",    f"¥{lucro_bruto:,}",    G if lucro_bruto>=0 else R),
                ("(-) Despesas",   f"¥{total_desp_tot:,}", R),
                ("Lucro Líquido",  f"¥{lucro_liq:,}",      G if lucro_liq>=0 else R),
                ("Margem Geral",   f"{marg_geral:.1f}%",   AC),
            ])

            # ── Gráfico de barras por mês — Receita + Lucro lado a lado ───────
            # Calcula receita e lucro por mês (12 entradas cada)
            mensal_rec  = []  # (label, receita_total)
            mensal_luc  = []  # (label, lucro_bruto)
            for m_idx in range(1, 13):
                ms2 = f"{m_idx:02d}"; pat2 = f"%/{ms2}/{ys}"
                rv  = _int(self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor_venda,',','') AS INTEGER)),0) "
                    "FROM vendas WHERE data_venda LIKE ?", (pat2,)).fetchone()[0])
                ros = _int(self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM servicos WHERE data_servico LIKE ?", (pat2,)).fetchone()[0])
                rsk = _int(self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                    "FROM shaken WHERE data_registro LIKE ? AND valor IS NOT NULL", (pat2,)).fetchone()[0])
                rec_m = rv + ros + rsk

                # lucro bruto do mês: receita − custos de compra − custos_os − custos_sk
                lv = 0
                for vrow in self.conn.execute(
                        "SELECT valor_venda, compra_id, carro_id FROM vendas WHERE data_venda LIKE ?",
                        (pat2,)).fetchall():
                    vv2 = _int(vrow[0])
                    cid3 = vrow[1]
                    if not cid3 and vrow[2]:
                        r3 = self.conn.execute(
                            "SELECT id FROM compras WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                            (vrow[2],)).fetchone()
                        cid3 = r3[0] if r3 else None
                    if cid3:
                        vc3 = _int(self.conn.execute(
                            "SELECT valor FROM compras WHERE id=?", (cid3,)).fetchone()[0]
                            if self.conn.execute("SELECT valor FROM compras WHERE id=?", (cid3,)).fetchone() else 0)
                        vv2 -= vc3 + self._get_custo_total(cid3)
                    lv += vv2
                los2 = 0
                for orow in self.conn.execute(
                        "SELECT id, valor FROM servicos WHERE data_servico LIKE ?", (pat2,)).fetchall():
                    rv2 = _int(orow[1])
                    cv2 = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_os WHERE servico_id=?", (orow[0],)).fetchone()[0] or 0
                    los2 += rv2 - _int(cv2)
                lsk2 = 0
                for skrow in self.conn.execute(
                        "SELECT id, valor FROM shaken WHERE data_registro LIKE ? "
                        "AND valor IS NOT NULL AND valor != ''", (pat2,)).fetchall():
                    rv3 = _int(skrow[1])
                    ck3 = self.conn.execute(
                        "SELECT COALESCE(SUM(CAST(REPLACE(valor,',','') AS INTEGER)),0) "
                        "FROM custos_sk WHERE shaken_id=?", (skrow[0],)).fetchone()[0] or 0
                    lsk2 += rv3 - _int(ck3)
                luc_m = lv + los2 + lsk2

                mensal_rec.append((MESES_PT[m_idx-1], rec_m))
                mensal_luc.append((MESES_PT[m_idx-1], luc_m))

            def _make_bar_chart(parent_frame, title, bar_color, dataset, baseline=False):
                """baseline=True desenha linha zero e barras negativas em vermelho."""
                hdr_f2 = tk.Frame(parent_frame, bg=COLORS["bg_card"])
                hdr_f2.pack(fill="x", padx=10, pady=(8,2))
                tk.Frame(hdr_f2, bg=bar_color, width=4, height=14).pack(side="left", padx=(0,6))
                tk.Label(hdr_f2, text=title, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
                cv2 = tk.Canvas(parent_frame, bg=COLORS["bg_card"], height=130,
                                highlightthickness=0)
                cv2.pack(fill="x", padx=8, pady=(0,8))

                def _draw(event=None):
                    cv2.delete("all")
                    W2 = cv2.winfo_width() or 400
                    vals2 = [d[1] for d in dataset]
                    max_v2 = max((abs(v) for v in vals2), default=1) or 1
                    min_v2 = min(vals2) if baseline else 0
                    span   = max_v2 - min(min_v2, 0)
                    if span == 0: span = 1
                    PAD_L, PAD_R, PAD_T, PAD_B = 38, 8, 14, 24
                    H2 = 130
                    avail_h = H2 - PAD_T - PAD_B
                    zero_y  = H2 - PAD_B - int(avail_h * max(0, -min_v2) / span)

                    # grid lines
                    for pct in [0.5, 1.0]:
                        gy2 = H2 - PAD_B - int(avail_h * pct * max_v2 / span)
                        cv2.create_line(PAD_L, gy2, W2 - PAD_R, gy2,
                                        fill=COLORS["border"], dash=(2,3))
                        lbl2 = f"¥{int(max_v2*pct)//1000}k" if max_v2 >= 1000 else f"¥{int(max_v2*pct):,}"
                        cv2.create_text(PAD_L-3, gy2, text=lbl2, anchor="e",
                                        font=("Helvetica",6), fill=COLORS["text_muted"])

                    # zero line
                    if baseline and min_v2 < 0:
                        cv2.create_line(PAD_L, zero_y, W2 - PAD_R, zero_y,
                                        fill=COLORS["text_muted"], dash=(4,2))

                    n2 = len(dataset)
                    slot2 = (W2 - PAD_L - PAD_R) / n2
                    bw2   = max(4, int(slot2 * 0.65))
                    for i2, (lbl_m, val2) in enumerate(dataset):
                        xc2 = PAD_L + slot2 * i2 + slot2 / 2
                        x0b = xc2 - bw2/2; x1b = xc2 + bw2/2
                        if val2 >= 0:
                            bar_h2 = int(avail_h * val2 / span) if span else 0
                            y0b = zero_y - bar_h2; y1b = zero_y
                            col2 = bar_color
                        else:
                            bar_h2 = int(avail_h * abs(val2) / span) if span else 0
                            y0b = zero_y; y1b = zero_y + bar_h2
                            col2 = COLORS["red"]
                        if bar_h2 > 0:
                            cv2.create_rectangle(x0b, y0b, x1b, y1b, fill=col2, outline="")
                        # valor acima/abaixo
                        if val2 != 0:
                            lbl_v = f"¥{abs(val2)//1000}k" if abs(val2) >= 10000 else f"¥{val2:,}"
                            anchor_v = "s" if val2 >= 0 else "n"
                            yv = y0b - 3 if val2 >= 0 else y1b + 3
                            cv2.create_text(xc2, yv, text=lbl_v,
                                            font=("Helvetica",6,"bold"),
                                            fill=col2, anchor=anchor_v)
                        cv2.create_text(xc2, H2 - PAD_B + 10, text=lbl_m,
                                        font=("Helvetica",6), fill=COLORS["text_muted"],
                                        anchor="center")
                cv2.bind("<Configure>", _draw)
                cv2.after(60, _draw)

            charts_outer = tk.Frame(self._ra_inner, bg=COLORS["bg_content"])
            charts_outer.pack(fill="x", pady=(2,8))

            left_card = tk.Frame(charts_outer, bg=COLORS["bg_card"],
                                 highlightthickness=1, highlightbackground=COLORS["border"])
            left_card.pack(side="left", fill="both", expand=True, padx=(0,6))
            tk.Frame(left_card, bg=COLORS["blue"], height=3).pack(fill="x")
            _make_bar_chart(left_card, "📈  Receita Mensal", COLORS["blue"],  mensal_rec, baseline=False)

            right_card = tk.Frame(charts_outer, bg=COLORS["bg_card"],
                                  highlightthickness=1, highlightbackground=COLORS["border"])
            right_card.pack(side="right", fill="both", expand=True, padx=(6,0))
            tk.Frame(right_card, bg=COLORS["green"], height=3).pack(fill="x")
            _make_bar_chart(right_card, "💹  Lucro Mensal",   COLORS["green"], mensal_luc, baseline=True)

            # cache para PDF
            self._ra_data_cache = {
                "ano": a,
                "cnt_v": cnt_v, "cnt_os": cnt_os, "cnt_sk": cnt_sk,
                "receita_v": receita_v_tot, "custo_v_sem": custo_v_sem,
                "is_v": is_v_tot, "custo_v": custo_v_tot, "lucro_v": lucro_v, "marg_v": marg_v,
                "receita_os": receita_os_tot, "custo_os_sem": custo_os_sem,
                "is_os": is_os_tot, "custo_os": custo_os_tot, "lucro_os": lucro_os, "marg_os": marg_os,
                "receita_sk": receita_sk_tot, "custo_sk_sem": custo_sk_sem,
                "is_sk": is_sk_tot, "custo_sk": custo_sk_tot, "lucro_sk": lucro_sk, "marg_sk": marg_sk,
                "total_desp": total_desp_tot, "is_total": is_total,
                "custo_sem_tot": custo_sem_tot, "receita_tot": receita_tot,
                "lucro_bruto": lucro_bruto, "lucro_liq": lucro_liq, "marg_geral": marg_geral,
                "mensal": mensal,
            }

        def _exportar_pdf():
            from tkinter import messagebox as _mb, filedialog as _fd
            import datetime as _dt
            d = self._ra_data_cache
            if not d:
                _refresh()
                d = self._ra_data_cache
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                Spacer, Paragraph, HRFlowable)
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors as rl_colors
                from reportlab.lib.units import cm
            except ImportError:
                _mb.showerror("PDF", "reportlab nao instalado.")
                return
            path = _fd.asksaveasfilename(
                defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
                initialfile=f"relatorio_anual_{d.get('ano','')}.pdf",
                title="Salvar Relatorio Anual")
            if not path: return

            hoje_str = _dt.date.today().strftime("%d/%m/%Y")
            RED    = rl_colors.HexColor("#C00000")
            DARK   = rl_colors.HexColor("#111827")
            GRAY   = rl_colors.HexColor("#D1D5DB")
            LGRAY  = rl_colors.HexColor("#F9FAFB")
            GREEN  = rl_colors.HexColor("#16A34A")
            ORANGE = rl_colors.HexColor("#EA580C")
            BLUE   = rl_colors.HexColor("#1D4ED8")
            GOLD   = rl_colors.HexColor("#B45309")
            PURPLE = rl_colors.HexColor("#8E44AD")
            MUTED  = rl_colors.HexColor("#6B7280")

            styles = getSampleStyleSheet()
            def P(txt, size=8, bold=False, color=DARK, align="LEFT"):
                return Paragraph(
                    f"<b>{txt}</b>" if bold else str(txt),
                    ParagraphStyle("_p", parent=styles["Normal"],
                                   fontSize=size, textColor=color, leading=size+3,
                                   alignment={"LEFT":0,"CENTER":1,"RIGHT":2}.get(align,0)))

            doc = SimpleDocTemplate(path, pagesize=A4,
                                    rightMargin=1.5*cm, leftMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            story = []

            story.append(P("KM CARS - Relatorio Anual", size=16, bold=True, color=RED))
            story.append(P(f"Ano: {d.get('ano','')}   |   Gerado em: {hoje_str}", size=8, color=MUTED))
            story.append(Spacer(1, 0.2*cm))
            story.append(HRFlowable(width="100%", thickness=2, color=RED))
            story.append(Spacer(1, 0.3*cm))

            def kpi_table(title, title_color, items):
                story.append(P(title, size=9, bold=True, color=title_color))
                story.append(Spacer(1, 0.1*cm))
                lbl_row = [P(lbl, size=7, color=MUTED)          for lbl,val,vc in items]
                val_row = [P(val, size=9, bold=True, color=vc)  for lbl,val,vc in items]
                n = len(items)
                cw = [17.0*cm / n] * n
                tbl = Table([lbl_row, val_row], colWidths=cw)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,-1), LGRAY),
                    ("LINEBELOW",     (0,0), (-1, 0), 0.3, GRAY),
                    ("TOPPADDING",    (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ("LEFTPADDING",   (0,0), (-1,-1), 6),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 0.25*cm))

            kpi_table(f"VENDAS  ({d.get('cnt_v',0)} vendas)", GREEN, [
                ("Receita",    f"Y {d.get('receita_v',0):,}",      GREEN),
                ("Custo s/IS", f"Y {d.get('custo_v_sem',0):,}",    ORANGE),
                ("IS (CNF)",   f"Y {d.get('is_v',0):,}",           PURPLE),
                ("C. Total",   f"Y {d.get('custo_v',0):,}",        RED),
                ("Lucro",      f"Y {d.get('lucro_v',0):,}",        GREEN if d.get('lucro_v',0)>=0 else RED),
                ("Margem",     f"{d.get('marg_v',0):.1f}%",        BLUE),
            ])
            kpi_table(f"ORDENS DE SERVICO  ({d.get('cnt_os',0)} OS)", BLUE, [
                ("Receita",    f"Y {d.get('receita_os',0):,}",     BLUE),
                ("Custo s/IS", f"Y {d.get('custo_os_sem',0):,}",   ORANGE),
                ("IS (CNF)",   f"Y {d.get('is_os',0):,}",          PURPLE),
                ("C. Total",   f"Y {d.get('custo_os',0):,}",       RED),
                ("Lucro",      f"Y {d.get('lucro_os',0):,}",       GREEN if d.get('lucro_os',0)>=0 else RED),
                ("Margem",     f"{d.get('marg_os',0):.1f}%",       BLUE),
            ])
            kpi_table(f"SHAKEN  ({d.get('cnt_sk',0)} registros)", GOLD, [
                ("Receita",    f"Y {d.get('receita_sk',0):,}",     GOLD),
                ("Custo s/IS", f"Y {d.get('custo_sk_sem',0):,}",   ORANGE),
                ("IS (CNF)",   f"Y {d.get('is_sk',0):,}",          PURPLE),
                ("C. Total",   f"Y {d.get('custo_sk',0):,}",       RED),
                ("Lucro",      f"Y {d.get('lucro_sk',0):,}",       GREEN if d.get('lucro_sk',0)>=0 else RED),
                ("Margem",     f"{d.get('marg_sk',0):.1f}%",       BLUE),
            ])
            kpi_table("DESPESAS FIXAS (anual)", RED, [
                ("Total", f"Y {d.get('total_desp',0):,}", RED),
            ])
            story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
            story.append(Spacer(1, 0.1*cm))
            kpi_table("RESUMO CONSOLIDADO DO ANO", DARK, [
                ("Receita Total", f"Y {d.get('receita_tot',0):,}",    GREEN),
                ("Custo s/IS",    f"Y {d.get('custo_sem_tot',0):,}",  ORANGE),
                ("IS Total",      f"Y {d.get('is_total',0):,}",       PURPLE),
                ("Lucro Bruto",   f"Y {d.get('lucro_bruto',0):,}",    GREEN if d.get('lucro_bruto',0)>=0 else RED),
                ("(-) Despesas",  f"Y {d.get('total_desp',0):,}",     RED),
                ("Lucro Liquido", f"Y {d.get('lucro_liq',0):,}",      GREEN if d.get('lucro_liq',0)>=0 else RED),
                ("Margem Geral",  f"{d.get('marg_geral',0):.1f}%",    BLUE),
            ])

            # ── Tabela mensal ────────────────────────────────────────────────
            story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
            story.append(Spacer(1, 0.2*cm))
            story.append(P("RECEITA MENSAL", size=9, bold=True, color=BLUE))
            story.append(Spacer(1, 0.1*cm))

            mensal_hdr = [P(m[0], size=7, bold=True, color=MUTED) for m in d.get("mensal",[])]
            mensal_val = [P(f"Y {m[1]:,}" if m[1] else "—", size=7,
                            bold=True, color=GREEN if m[1]>=0 else RED)
                          for m in d.get("mensal",[])]
            if mensal_hdr:
                cw12 = [17.0*cm / 12] * 12
                m_tbl = Table([mensal_hdr, mensal_val], colWidths=cw12)
                m_tbl.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,-1), LGRAY),
                    ("LINEBELOW",     (0,0), (-1, 0), 0.3, GRAY),
                    ("TOPPADDING",    (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ("LEFTPADDING",   (0,0), (-1,-1), 4),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 4),
                    ("ALIGN",         (0,1), (-1,1), "CENTER"),
                ]))
                story.append(m_tbl)

            story.append(Spacer(1, 0.4*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
            story.append(Spacer(1, 0.1*cm))
            story.append(P(f"KM Cars - Fujinomiya, Japan  |  Gerado em {hoje_str}",
                           size=7, color=MUTED, align="CENTER"))
            doc.build(story)
            _mb.showinfo("PDF Gerado", f"Relatorio anual salvo:\n{path}")
            try:
                import subprocess, sys
                if sys.platform == "win32":    os.startfile(path)
                elif sys.platform == "darwin": subprocess.call(["open", path])
                else:                          subprocess.call(["xdg-open", path])
            except Exception: pass

        _refresh()

    def _build_dash_dossie_cliente(self, parent):
        self._dc_cliente_id   = None
        self._dc_search_var   = tk.StringVar(value="")
        self._dc_results_list = []

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Coluna esquerda: busca ─────────────────────────────────────────────
        left = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"],
                        width=260)
        left.pack(side="left", fill="y", padx=(0,12))
        left.pack_propagate(False)
        tk.Frame(left, bg=COLORS["blue"], height=4).pack(fill="x")
        _dc_hdr = tk.Frame(left, bg=COLORS["bg_card"])
        _dc_hdr.pack(fill="x", padx=14, pady=(12,4))
        tk.Label(_dc_hdr, text="◈  Buscar Cliente",
                 font=("Helvetica",10,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(_dc_hdr, text="↺ Atualizar", font=("Helvetica", 7),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=6, pady=2, command=self._dc_buscar).pack(side="right")

        tk.Label(left, text="ID ou Nome:", font=("Helvetica",8,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w",padx=14)
        tk.Entry(left, textvariable=self._dc_search_var, font=("Helvetica",10),
                 bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["blue"]
                 ).pack(fill="x", padx=14, pady=(4,6), ipady=6)
        self._dc_search_var.trace_add("write", lambda *_: self._dc_buscar())

        # Lista de resultados
        sf = tk.Frame(left, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=14, pady=(0,4))
        self._dc_listbox = tk.Listbox(sf,
                                       font=("Helvetica",9),
                                       bg=COLORS["bg_main"],
                                       fg=COLORS["text_primary"],
                                       selectbackground=COLORS["blue"],
                                       selectforeground="white",
                                       relief="flat", bd=0,
                                       highlightthickness=1,
                                       highlightbackground=COLORS["border"])
        sb_l = tk.Scrollbar(sf, orient="vertical", command=self._dc_listbox.yview)
        self._dc_listbox.configure(yscrollcommand=sb_l.set)
        self._dc_listbox.pack(side="left", fill="both", expand=True)
        sb_l.pack(side="right", fill="y")
        self._dc_listbox.bind("<<ListboxSelect>>", self._dc_on_select)

        tk.Button(left, text="📋  Gerar Dossiê",
                  font=("Helvetica",10,"bold"),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  command=self._dc_gerar
                  ).pack(fill="x", padx=14, pady=8, ipady=8)
        tk.Button(left, text="📄  Exportar PDF",
                  font=("Helvetica",9),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._dc_exportar_pdf
                  ).pack(fill="x", padx=14, pady=(0,12), ipady=6)

        # ── Coluna direita: dossiê ─────────────────────────────────────────────
        right_outer = tk.Frame(root, bg=COLORS["bg_content"])
        right_outer.pack(side="right", fill="both", expand=True)

        right_card = tk.Frame(right_outer, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        right_card.pack(fill="both", expand=True)
        tk.Frame(right_card, bg=COLORS["blue"], height=4).pack(fill="x")

        # Placeholder inicial
        self._dc_content_frame = tk.Frame(right_card, bg=COLORS["bg_card"])
        self._dc_content_frame.pack(fill="both", expand=True)
        self._dc_show_placeholder()

        # Scroll interno para o dossiê
        self._dc_right_card = right_card

        # Carrega lista inicial
        self._dc_buscar()

    def _dc_show_placeholder(self):
        for w in self._dc_content_frame.winfo_children():
            w.destroy()
        tk.Label(self._dc_content_frame,
                 text="◈\n\nBusque e selecione um cliente\npara visualizar o dossiê completo.",
                 font=("Helvetica",11), justify="center",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).place(relx=.5,rely=.5,anchor="center")

    def _dc_buscar(self):
        termo = self._dc_search_var.get().strip()
        if termo.isdigit():
            rows = self.conn.execute(
                "SELECT id, nome, telefone FROM clientes WHERE id=? ORDER BY nome COLLATE NOCASE",
                (int(termo),)).fetchall()
        elif termo:
            rows = self.conn.execute(
                "SELECT id, nome, telefone FROM clientes WHERE nome LIKE ? ORDER BY nome COLLATE NOCASE",
                (f"%{termo}%",)).fetchall()
        else:
            rows = self.conn.execute(
                "SELECT id, nome, telefone FROM clientes ORDER BY nome COLLATE NOCASE LIMIT 80").fetchall()
        self._dc_results_list = [dict(r) for r in rows]
        self._dc_listbox.delete(0, tk.END)
        for r in self._dc_results_list:
            self._dc_listbox.insert(tk.END, f"#{r['id']}  {r['nome']}")

    def _dc_on_select(self, event):
        sel = self._dc_listbox.curselection()
        if not sel:
            return
        self._dc_cliente_id = self._dc_results_list[sel[0]]["id"]

    def _dc_gerar(self):
        if not self._dc_cliente_id:
            from tkinter import messagebox as msgbox
            msgbox.showwarning("Atenção", "Selecione um cliente na lista.")
            return
        self._dc_render_dossie(self._dc_cliente_id)

    def _dc_render_dossie(self, cid):
        for w in self._dc_content_frame.winfo_children():
            w.destroy()

        # Busca dados
        cli = dict(self.conn.execute("SELECT * FROM clientes WHERE id=?", (cid,)).fetchone())
        vendas = [dict(r) for r in self.conn.execute(
            "SELECT * FROM vendas WHERE cliente_id=? ORDER BY id DESC", (cid,)).fetchall()]
        servicos = [dict(r) for r in self.conn.execute(
            "SELECT * FROM servicos WHERE cliente_id=? ORDER BY id DESC", (cid,)).fetchall()]

        # Scroll interno
        cv = tk.Canvas(self._dc_content_frame, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(self._dc_content_frame, orient="vertical", command=cv.yview)
        inner = tk.Frame(cv, bg=COLORS["bg_card"])
        inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        wid = cv.create_window((0,0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.bind("<Configure>", lambda e: cv.itemconfig(wid, width=e.width))
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._dc_content_frame.bind("<Enter>", lambda e: self._dc_content_frame.bind_all(
            "<MouseWheel>", lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        self._dc_content_frame.bind("<Leave>", lambda e: self._dc_content_frame.unbind_all("<MouseWheel>"))

        pad = dict(padx=20, pady=4)

        def section(txt, color=COLORS["blue"]):
            f = tk.Frame(inner, bg=color, height=2)
            f.pack(fill="x", padx=20, pady=(14,2))
            tk.Label(inner, text=txt, font=("Helvetica",11,"bold"),
                     bg=COLORS["bg_card"], fg=color).pack(anchor="w", **pad)

        def row2(lbl, val, fg=None):
            r = tk.Frame(inner, bg=COLORS["bg_card"])
            r.pack(fill="x", padx=20, pady=1)
            tk.Label(r, text=lbl+":", font=("Helvetica",8,"bold"), width=18, anchor="w",
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
            tk.Label(r, text=str(val or "—"), font=("Helvetica",9),
                     bg=COLORS["bg_card"], fg=fg or COLORS["text_primary"]).pack(side="left")

        # ── Dados pessoais ────────────────────────────────────────────────────
        tk.Label(inner, text=f"DOSSIÊ DO CLIENTE — #{cli['id']}",
                 font=("Helvetica",13,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=20, pady=(16,4))
        section("▸ Dados Pessoais", COLORS["blue"])
        row2("Nome",     cli.get("nome"))
        row2("Telefone", cli.get("telefone"))
        row2("Cidade",   cli.get("cidade") or "—")

        # ── Resumo financeiro ─────────────────────────────────────────────────
        section("▸ Resumo Financeiro", COLORS["green"])
        try:
            total_vendas = sum(int(str(v.get("valor_venda") or "0").replace(",","")) for v in vendas)
            total_serv   = sum(int(str(s.get("valor") or "0").replace(",","")) for s in servicos)
        except Exception:
            total_vendas = total_serv = 0
        row2("Qtd. Compras",   len(vendas))
        row2("Total em Vendas", f"¥ {total_vendas:,}", COLORS["green"])
        row2("Qtd. Serviços",  len(servicos))
        row2("Total Serviços", f"¥ {total_serv:,}", COLORS["orange"])

        # ── Compras (veículos vinculados a este cliente) ──────────────────────
        compras_cli = [dict(r) for r in self.conn.execute(
            "SELECT co.* FROM compras co "
            "INNER JOIN carros ca ON co.carro_id=ca.id "
            "WHERE ca.cliente_id=? ORDER BY co.id DESC", (cid,)).fetchall()]
        if compras_cli:
            section("▸ Compras (Carros do Cliente)", COLORS["blue"])
            for cc in compras_cli:
                custo_cc = self._get_custo_total(cc["id"])
                try: vcc = int(str(cc.get("valor") or "0").replace(",",""))
                except: vcc = 0
                # Busca nome de quem vendeu o carro ao admin (comprado de)
                comprado_de = "—"
                if cc.get("cliente_id"):
                    cd_row = self.conn.execute(
                        "SELECT nome FROM clientes WHERE id=?", (cc["cliente_id"],)).fetchone()
                    comprado_de = cd_row[0] if cd_row else "—"
                cc_card = tk.Frame(inner, bg=COLORS["bg_content"],
                                   highlightthickness=1, highlightbackground=COLORS["border"])
                cc_card.pack(fill="x", padx=20, pady=(4,0))
                tk.Frame(cc_card, bg=COLORS["blue"], width=4).pack(side="left", fill="y")
                cc_inner = tk.Frame(cc_card, bg=COLORS["bg_content"])
                cc_inner.pack(side="left", fill="both", expand=True, padx=10, pady=6)
                # Busca nome do carro
                car_row_c = self.conn.execute(
                    "SELECT carro, placa FROM carros WHERE id=?", (cc["carro_id"],)).fetchone()
                car_disp = (car_row_c[0] if car_row_c else cc.get("carro","—"))
                placa_disp = (car_row_c[1] if car_row_c else "") or "—"
                hdr_f = tk.Frame(cc_inner, bg=COLORS["bg_content"])
                hdr_f.pack(fill="x")
                tk.Label(hdr_f,
                         text=f"#{cc['id']}  —  {car_disp}  |  Placa: {placa_disp}  |  {cc.get('data_entrada','—')}",
                         font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
                tk.Label(hdr_f,
                         text=f"Compra de: {comprado_de}",
                         font=("Helvetica",8,"bold"),
                         bg=COLORS["accent"], fg="white", padx=6
                         ).pack(side="right", padx=(4,0))
                row_cc = tk.Frame(cc_inner, bg=COLORS["bg_content"])
                row_cc.pack(fill="x", pady=(4,0))
                TIPO_CC = {"Leilão":COLORS["orange"],"Troca":COLORS["blue"],"Compra Direta":COLORS["green"]}
                tc_cc = TIPO_CC.get(cc.get("tipo",""), COLORS["text_muted"])
                for lbl_cc, val_cc, fg_cc in [
                    ("Tipo",       cc.get("tipo","—"),                         tc_cc),
                    ("V.Compra",   self._fmt_yen_display(cc.get("valor")),     COLORS["orange"]),
                    ("Custos",     self._fmt_yen_display(custo_cc),            COLORS["text_secondary"]),
                    ("Total",      self._fmt_yen_display(vcc+custo_cc),        COLORS["green"]),
                ]:
                    cf = tk.Frame(row_cc, bg=COLORS["bg_content"])
                    cf.pack(side="left", padx=(0,16))
                    tk.Label(cf, text=lbl_cc, font=("Helvetica",7),
                             bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                    tk.Label(cf, text=val_cc, font=("Helvetica",8,"bold"),
                             bg=COLORS["bg_content"], fg=fg_cc).pack(anchor="w")

        # ── Histórico de vendas (carros vendidos PARA este cliente) ─────────
        section("▸ Histórico de Vendas — Carros Vendidos para este Cliente", COLORS["accent"])
        if not vendas:
            tk.Label(inner, text="Nenhuma venda registrada.", font=("Helvetica",8),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        for v in vendas:
            vcard = tk.Frame(inner, bg=COLORS["bg_content"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
            vcard.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(vcard, bg=COLORS["accent"], width=4).pack(side="left", fill="y")
            vc_inner = tk.Frame(vcard, bg=COLORS["bg_content"])
            vc_inner.pack(side="left", fill="both", expand=True, padx=10, pady=6)

            pagas = v.get("parcelas_pagas") or 0
            tot_p = v.get("num_parcelas") or 0
            is_p  = v.get("tipo_venda","") in ("Venda Parcelada","Com Troca","Com Troca e Volta","Venda Leilão")
            quit  = is_p and tot_p > 0 and pagas >= tot_p
            try:
                tp = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) FROM pagamentos WHERE venda_id=?",
                    (v["id"],)).fetchone()[0]
                vv = int(str(v.get("valor_venda") or "0").replace(",",""))
                et = int(str(v.get("entrada") or "0").replace(",",""))
                vt_int = int(str(v.get("valor_troca") or "0").replace(",",""))
                saldo = max(0, vv-et-vt_int-tp)
            except Exception:
                saldo = 0

            # Header da venda
            hv = tk.Frame(vc_inner, bg=COLORS["bg_content"])
            hv.pack(fill="x")
            tk.Label(hv, text=f"#{v['id']} — {v.get('data_venda','—')}  |  {v.get('carro','—')}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
            st_txt = "Quitado" if quit else ("Em Aberto" if is_p else "À Vista")
            st_col = COLORS["green"] if (quit or not is_p) else COLORS["orange"]
            tk.Label(hv, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_col, fg="white").pack(side="right", padx=4)
            tk.Label(hv, text=f"Vendido para: {cli.get('nome','—')}",
                     font=("Helvetica",8,"bold"),
                     bg=COLORS["green"], fg="white", padx=6
                     ).pack(side="right", padx=(0,4))

            det = tk.Frame(vc_inner, bg=COLORS["bg_content"])
            det.pack(fill="x", pady=(2,0))
            has_troca = bool(v.get("carro_troca"))
            det_items = [
                ("Tipo",      v.get("tipo_venda","—"),                              COLORS["text_secondary"]),
                ("Valor",     self._fmt_yen_display(v.get("valor_venda")),          COLORS["green"]),
                ("Entrada",   self._fmt_yen_display(v.get("entrada")) if v.get("entrada") else "—",
                              COLORS["text_muted"]),
                ("Parcela",   self._fmt_yen_display(v.get("parcela_mensal")) if is_p else "—",
                              COLORS["blue"]),
                ("Progresso", f"{pagas}/{tot_p}" if is_p else "—",                 COLORS["accent"]),
                ("Saldo",     f"¥{saldo:,}" if (is_p and not quit) else "—",       COLORS["orange"]),
            ]
            if has_troca:
                ct_nm = (v.get("carro_troca") or "—").split("|")[0].strip()
                vt_d  = self._fmt_yen_display(v.get("valor_troca")) if v.get("valor_troca") else "—"
                det_items += [
                    ("Carro Troca", ct_nm[:20], COLORS["orange"]),
                    ("V. Troca",    vt_d,        COLORS["orange"]),
                ]
            for lbl_t, val_t, fg_t in det_items:
                col = tk.Frame(det, bg=COLORS["bg_content"])
                col.pack(side="left", padx=(0,14))
                tk.Label(col, text=lbl_t, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col, text=val_t, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg_t).pack(anchor="w")

        # ── Histórico de serviços ─────────────────────────────────────────────
        section("▸ Histórico de Serviços", COLORS["orange"])
        if not servicos:
            tk.Label(inner, text="Nenhum serviço registrado.", font=("Helvetica",8),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        for s in servicos:
            scard = tk.Frame(inner, bg=COLORS["bg_content"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
            scard.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(scard, bg=COLORS["orange"], width=4).pack(side="left", fill="y")
            sc_inner = tk.Frame(scard, bg=COLORS["bg_content"])
            sc_inner.pack(side="left", fill="both", expand=True, padx=10, pady=6)

            hs = tk.Frame(sc_inner, bg=COLORS["bg_content"])
            hs.pack(fill="x")
            tk.Label(hs, text=f"OS {s.get('os_num') or s['id']} — {s.get('data_servico','—')}  |  {s.get('carro','—')}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
            st_col = {"Concluído":COLORS["green"],"Em Andamento":COLORS["orange"],
                      "Aberto":COLORS["blue"]}.get(s.get("status",""), COLORS["text_muted"])
            tk.Label(hs, text=s.get("status","—"),
                     font=("Helvetica",7,"bold"), bg=st_col, fg="white").pack(side="right", padx=4)

            ds = tk.Frame(sc_inner, bg=COLORS["bg_content"])
            ds.pack(fill="x", pady=(2,0))
            for lbl, val, fg in [
                ("Tipo Serviço", s.get("tipo_servico","—"), COLORS["text_secondary"]),
                ("Valor", self._fmt_yen_display(s.get("valor")), COLORS["green"]),
                ("Descrição", (s.get("descricao") or "—")[:40], COLORS["text_muted"]),
            ]:
                col = tk.Frame(ds, bg=COLORS["bg_content"])
                col.pack(side="left", padx=(0,14))
                tk.Label(col, text=lbl, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col, text=val, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg).pack(anchor="w")

        # ── Shaken do cliente ─────────────────────────────────────────────────
        section("▸ Shaken", COLORS["blue"])
        # Busca carros do cliente e seus shakens
        carros_cli = [dict(r) for r in self.conn.execute(
            "SELECT * FROM carros WHERE cliente_id=?", (cid,)).fetchall()]
        shakens_cli = []
        for car in carros_cli:
            sks = [dict(r) for r in self.conn.execute(
                "SELECT * FROM shaken WHERE carro_id=? ORDER BY id DESC LIMIT 1",
                (car["id"],)).fetchall()]
            for sk in sks:
                sk["_carro_nome"] = car.get("carro","—")
                shakens_cli.append(sk)
        if not shakens_cli and not carros_cli:
            tk.Label(inner, text="Nenhum carro vinculado.", font=("Helvetica",8),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        elif not shakens_cli:
            tk.Label(inner, text="Nenhum shaken registrado para os carros deste cliente.",
                     font=("Helvetica",8), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        for sk in shakens_cli:
            st_txt, st_cor = self._sk_status_info(
                sk.get("data_vencimento"), sk.get("por_conta"))
            skcard = tk.Frame(inner, bg=COLORS["bg_content"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
            skcard.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(skcard, bg=st_cor, width=4).pack(side="left", fill="y")
            sk_in = tk.Frame(skcard, bg=COLORS["bg_content"])
            sk_in.pack(side="left", fill="both", expand=True, padx=10, pady=6)
            hsk = tk.Frame(sk_in, bg=COLORS["bg_content"])
            hsk.pack(fill="x")
            tk.Label(hsk, text=f"{sk['sk_num']}  —  {sk.get('_carro_nome','—')}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
            tk.Label(hsk, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_cor, fg="white").pack(side="right", padx=4)
            dsk = tk.Frame(sk_in, bg=COLORS["bg_content"])
            dsk.pack(fill="x", pady=(2,0))
            d_venc = sk.get("data_vencimento","") or ("Por Conta" if sk.get("por_conta") else "—")
            vl_sk = self._fmt_yen_display(sk.get("valor")) if sk.get("valor") else "—"
            for lbl_t, val_t, fg_t in [
                ("Vencimento", d_venc, st_cor),
                ("Valor", vl_sk, COLORS["green"]),
                ("Registrado", sk.get("data_registro","—"), COLORS["text_muted"]),
            ]:
                col = tk.Frame(dsk, bg=COLORS["bg_content"])
                col.pack(side="left", padx=(0,14))
                tk.Label(col, text=lbl_t, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col, text=val_t, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg_t).pack(anchor="w")

        tk.Frame(inner, bg=COLORS["bg_card"], height=20).pack()

    def _dc_exportar_pdf(self):
        if not self._dc_cliente_id:
            from tkinter import messagebox as msgbox
            msgbox.showwarning("Atenção", "Gere o dossiê primeiro.")
            return
        self._gerar_pdf_cliente(self._dc_cliente_id)

    def _gerar_pdf_cliente(self, cid):
        """Gera PDF do dossiê do cliente usando reportlab."""
        import os, datetime
        from tkinter import filedialog, messagebox as msgbox
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle, HRFlowable)
        except ImportError:
            msgbox.showerror("Erro", "Instale reportlab: pip install reportlab")
            return

        cli = dict(self.conn.execute("SELECT * FROM clientes WHERE id=?", (cid,)).fetchone())
        vendas   = [dict(r) for r in self.conn.execute(
            "SELECT * FROM vendas WHERE cliente_id=? ORDER BY id DESC", (cid,)).fetchall()]
        servicos = [dict(r) for r in self.conn.execute(
            "SELECT * FROM servicos WHERE cliente_id=? ORDER BY id DESC", (cid,)).fetchall()]

        # Escolhe onde salvar
        nome_arq = f"Dossie_Cliente_{cli.get('nome','').replace(' ','_')}_{cid}.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=nome_arq,
            filetypes=[("PDF", "*.pdf")],
            title="Salvar Dossiê PDF")
        if not path:
            return

        doc = SimpleDocTemplate(path, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()

        BLUE  = rl_colors.HexColor("#2B7BE5")
        GREEN = rl_colors.HexColor("#27AE60")
        ORANGE= rl_colors.HexColor("#E67E22")
        GRAY  = rl_colors.HexColor("#7F8C8D")
        DARK  = rl_colors.HexColor("#1E1E2E")

        def h1(txt, color=BLUE):
            return Paragraph(f'<font color="{color.hexval()}" size="14"><b>{txt}</b></font>', styles["Normal"])
        def h2(txt, color=BLUE):
            return Paragraph(f'<font color="{color.hexval()}" size="11"><b>{txt}</b></font>', styles["Normal"])
        def body(txt):
            return Paragraph(f'<font size="9">{txt}</font>', styles["Normal"])

        story = []
        hoje_str = datetime.date.today().strftime("%d/%m/%Y")

        # Título
        story.append(h1(f"DOSSIÊ DO CLIENTE — #{cli['id']}"))
        story.append(body(f"Gerado em {hoje_str}"))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=BLUE))
        story.append(Spacer(1, 0.3*cm))

        # Dados pessoais
        story.append(h2("▸ DADOS PESSOAIS"))
        story.append(Spacer(1, 0.15*cm))
        dados_p = [
            ["Nome",     cli.get("nome") or "—"],
            ["Telefone", cli.get("telefone") or "—"],
            ["Cidade",   cli.get("cidade") or "—"],
        ]
        t = Table([[body(f"<b>{k}</b>"), body(v)] for k,v in dados_p],
                  colWidths=[4*cm, 12*cm])
        t.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [rl_colors.HexColor("#F8F9FA"), rl_colors.white]),
            ("GRID", (0,0), (-1,-1), 0.3, rl_colors.HexColor("#E0E0E0")),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

        # Resumo financeiro
        try:
            tv = sum(int(str(v.get("valor_venda") or "0").replace(",","")) for v in vendas)
            ts = sum(int(str(s.get("valor") or "0").replace(",","")) for s in servicos)
        except Exception:
            tv = ts = 0
        story.append(h2("▸ RESUMO FINANCEIRO", GREEN))
        story.append(Spacer(1, 0.15*cm))
        rf = Table([
            [body(f"<b>Qtd. Compras</b>"), body(str(len(vendas))),
             body(f"<b>Total Vendas</b>"), body(f"¥ {tv:,}")],
            [body(f"<b>Qtd. Serviços</b>"), body(str(len(servicos))),
             body(f"<b>Total Serviços</b>"), body(f"¥ {ts:,}")],
        ], colWidths=[4*cm, 3*cm, 4*cm, 5*cm])
        rf.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.3, rl_colors.HexColor("#E0E0E0")),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [rl_colors.HexColor("#EAF6FF"), rl_colors.white]),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(rf)
        story.append(Spacer(1, 0.4*cm))

        # Vendas
        story.append(h2("▸ HISTÓRICO DE VENDAS", rl_colors.HexColor("#8E44AD")))
        story.append(Spacer(1, 0.15*cm))
        if not vendas:
            story.append(body("Nenhuma venda registrada."))
        else:
            vh = [["#", "Data", "Carro", "Tipo Venda", "Valor", "Entrada", "Parcela", "Status"]]
            for v in vendas:
                pagas  = v.get("parcelas_pagas") or 0
                tot_p  = v.get("num_parcelas") or 0
                is_p   = v.get("tipo_venda","") in ("Venda Parcelada","Com Troca","Com Troca e Volta")
                quit   = is_p and tot_p > 0 and pagas >= tot_p
                st     = "Quitado" if quit else ("Em Aberto" if is_p else "À Vista")
                tipo_v = v.get("tipo_venda") or "—"
                vh.append([
                    str(v["id"]), v.get("data_venda","—"),
                    (v.get("carro") or "—")[:20],
                    tipo_v,
                    f"¥ {v.get('valor_venda') or '—'}",
                    f"¥ {v.get('entrada') or '—'}",
                    f"¥ {v.get('parcela_mensal') or '—'}" if is_p else "—",
                    st
                ])
            vt = Table([[body(c) for c in row] for row in vh],
                       colWidths=[0.8*cm, 2*cm, 3.5*cm, 3.2*cm, 2*cm, 1.8*cm, 1.8*cm, 1.9*cm])
            vt.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), BLUE),
                ("TEXTCOLOR", (0,0), (-1,0), rl_colors.white),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_colors.HexColor("#F8F9FA"), rl_colors.white]),
                ("GRID", (0,0), (-1,-1), 0.3, rl_colors.HexColor("#E0E0E0")),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("LEFTPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING", (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ]))
            story.append(vt)
        story.append(Spacer(1, 0.4*cm))

        # Serviços
        story.append(h2("▸ HISTÓRICO DE SERVIÇOS", ORANGE))
        story.append(Spacer(1, 0.15*cm))
        if not servicos:
            story.append(body("Nenhum serviço registrado."))
        else:
            sh = [["OS", "Data", "Carro", "Tipo Serviço", "Valor", "Status"]]
            for s in servicos:
                sh.append([
                    str(s.get("os_num") or s["id"]),
                    s.get("data_servico","—"),
                    (s.get("carro") or "—")[:18],
                    (s.get("tipo_servico") or "—")[:16],
                    f"¥ {s.get('valor') or '—'}",
                    s.get("status","—"),
                ])
            st = Table([[body(c) for c in row] for row in sh],
                       colWidths=[1.2*cm, 2.2*cm, 4.5*cm, 4*cm, 2.5*cm, 2.5*cm])
            st.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), ORANGE),
                ("TEXTCOLOR", (0,0), (-1,0), rl_colors.white),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_colors.HexColor("#FFF8F0"), rl_colors.white]),
                ("GRID", (0,0), (-1,-1), 0.3, rl_colors.HexColor("#E0E0E0")),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("LEFTPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING", (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ]))
            story.append(st)

        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
        story.append(body(f"<font color='grey'>KM Cars — Dossiê gerado em {hoje_str}</font>"))

        doc.build(story)
        try:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass
        msgbox.showinfo("PDF Gerado", f"Dossiê salvo em:\n{path}")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: DOSSIÊ CARRO
    # ══════════════════════════════════════════════════════════════════════════
    def _build_dash_dossie_carro(self, parent):
        self._dcar_carro_id  = None
        self._dcar_search_var = tk.StringVar(value="")
        self._dcar_results   = []

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Coluna esquerda: busca ─────────────────────────────────────────────
        left = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"],
                        width=260)
        left.pack(side="left", fill="y", padx=(0,12))
        left.pack_propagate(False)
        tk.Frame(left, bg=COLORS["orange"], height=4).pack(fill="x")
        _dcar_hdr = tk.Frame(left, bg=COLORS["bg_card"])
        _dcar_hdr.pack(fill="x", padx=14, pady=(12,4))
        tk.Label(_dcar_hdr, text="▦  Buscar Veículo",
                 font=("Helvetica",10,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(_dcar_hdr, text="↺ Atualizar", font=("Helvetica", 7),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=6, pady=2, command=self._dcar_buscar).pack(side="right")

        tk.Label(left, text="Nome / Placa / Chassi / ID:", font=("Helvetica",8,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w",padx=14)
        tk.Entry(left, textvariable=self._dcar_search_var, font=("Helvetica",10),
                 bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["orange"]
                 ).pack(fill="x", padx=14, pady=(4,6), ipady=6)
        self._dcar_search_var.trace_add("write", lambda *_: self._dcar_buscar())

        sf = tk.Frame(left, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=14, pady=(0,4))
        self._dcar_listbox = tk.Listbox(sf,
                                         font=("Helvetica",9),
                                         bg=COLORS["bg_main"],
                                         fg=COLORS["text_primary"],
                                         selectbackground=COLORS["orange"],
                                         selectforeground="white",
                                         relief="flat", bd=0,
                                         highlightthickness=1,
                                         highlightbackground=COLORS["border"])
        sb_l = tk.Scrollbar(sf, orient="vertical", command=self._dcar_listbox.yview)
        self._dcar_listbox.configure(yscrollcommand=sb_l.set)
        self._dcar_listbox.pack(side="left", fill="both", expand=True)
        sb_l.pack(side="right", fill="y")
        # Selecionar na lista já mostra o dossiê; duplo-clique também funciona
        self._dcar_listbox.bind("<<ListboxSelect>>", self._dcar_on_select)
        self._dcar_listbox.bind("<Double-Button-1>", lambda e: self._dcar_gerar())

        tk.Button(left, text="📋  Gerar Dossiê",
                  font=("Helvetica",10,"bold"),
                  bg=COLORS["orange"], fg="white", relief="flat", cursor="hand2",
                  command=self._dcar_gerar
                  ).pack(fill="x", padx=14, pady=8, ipady=8)
        tk.Button(left, text="📄  Exportar PDF",
                  font=("Helvetica",9),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._dcar_exportar_pdf
                  ).pack(fill="x", padx=14, pady=(0,12), ipady=6)

        # ── Coluna direita ─────────────────────────────────────────────────────
        right_outer = tk.Frame(root, bg=COLORS["bg_content"])
        right_outer.pack(side="right", fill="both", expand=True)
        right_card = tk.Frame(right_outer, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        right_card.pack(fill="both", expand=True)
        tk.Frame(right_card, bg=COLORS["orange"], height=4).pack(fill="x")

        self._dcar_content_frame = tk.Frame(right_card, bg=COLORS["bg_card"])
        self._dcar_content_frame.pack(fill="both", expand=True)
        self._dcar_show_placeholder()
        self._dcar_buscar()

    def _dcar_show_placeholder(self):
        for w in self._dcar_content_frame.winfo_children():
            w.destroy()
        tk.Label(self._dcar_content_frame,
                 text="▦\n\nBusque e selecione um veículo\npara visualizar o dossiê completo.",
                 font=("Helvetica",11), justify="center",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).place(relx=.5,rely=.5,anchor="center")

    def _dcar_buscar(self):
        termo = self._dcar_search_var.get().strip().lower()
        if termo:
            # Verifica se é ID numérico
            if termo.isdigit():
                rows = self.conn.execute(
                    "SELECT id, carro, placa, chassi, ano, cor FROM carros "
                    "WHERE id=? ORDER BY carro COLLATE NOCASE LIMIT 80",
                    (int(termo),)).fetchall()
            else:
                rows = self.conn.execute(
                    "SELECT id, carro, placa, chassi, ano, cor FROM carros "
                    "WHERE LOWER(carro) LIKE ? OR LOWER(placa) LIKE ? OR LOWER(chassi) LIKE ? "
                    "ORDER BY carro COLLATE NOCASE LIMIT 80",
                    (f"%{termo}%", f"%{termo}%", f"%{termo}%")).fetchall()
        else:
            rows = self.conn.execute(
                "SELECT id, carro, placa, chassi, ano, cor FROM carros "
                "ORDER BY carro COLLATE NOCASE LIMIT 80").fetchall()
        self._dcar_results = [dict(r) for r in rows]
        self._dcar_listbox.delete(0, tk.END)
        for r in self._dcar_results:
            label = f"#{r['id']}  {r.get('carro','')}"
            if r.get("ano"):
                label += f"  {r['ano']}"
            if r.get("placa"):
                label += f"  [{r['placa']}]"
            self._dcar_listbox.insert(tk.END, label)

    def _dcar_on_select(self, event):
        sel = self._dcar_listbox.curselection()
        if not sel: return
        self._dcar_carro_id = self._dcar_results[sel[0]]["id"]
        # Renderiza automaticamente ao selecionar
        self._dcar_render_dossie(self._dcar_carro_id)

    def _dcar_gerar(self):
        if not self._dcar_carro_id:
            from tkinter import messagebox as msgbox
            msgbox.showwarning("Atenção", "Selecione um veículo na lista.")
            return
        self._dcar_render_dossie(self._dcar_carro_id)

    def _dcar_render_dossie(self, car_id):
        for w in self._dcar_content_frame.winfo_children():
            w.destroy()

        row_c = self.conn.execute("SELECT * FROM carros WHERE id=?", (car_id,)).fetchone()
        if not row_c:
            tk.Label(self._dcar_content_frame,
                     text="Veículo não encontrado no banco de dados.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["red"]).place(relx=.5, rely=.5, anchor="center")
            return
        carro = dict(row_c)
        # Compras vinculadas ao carro
        compras = [dict(r) for r in self.conn.execute(
            "SELECT co.*, cl.nome as cliente_nome FROM compras co "
            "LEFT JOIN clientes cl ON co.cliente_id=cl.id "
            "WHERE co.carro_id=? ORDER BY co.id DESC", (car_id,)).fetchall()]
        carro_nome = carro.get("carro","").strip()
        # Vendas onde o carro aparece pelo nome
        vendas = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id=c.id "
            "WHERE v.carro LIKE ? ORDER BY v.id DESC",
            (f"%{carro_nome}%",)).fetchall()]
        # Serviços vinculados ao carro
        servicos = [dict(r) for r in self.conn.execute(
            "SELECT s.*, c.nome as cliente_nome FROM servicos s "
            "LEFT JOIN clientes c ON s.cliente_id=c.id "
            "WHERE s.carro LIKE ? ORDER BY s.id DESC",
            (f"%{carro_nome}%",)).fetchall()]

        # Scroll interno
        cv = tk.Canvas(self._dcar_content_frame, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(self._dcar_content_frame, orient="vertical", command=cv.yview)
        inner = tk.Frame(cv, bg=COLORS["bg_card"])
        inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        wid = cv.create_window((0,0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.bind("<Configure>", lambda e: cv.itemconfig(wid, width=e.width))
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._dcar_content_frame.bind("<Enter>", lambda e: self._dcar_content_frame.bind_all(
            "<MouseWheel>", lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        self._dcar_content_frame.bind("<Leave>", lambda e: self._dcar_content_frame.unbind_all("<MouseWheel>"))

        def section(txt, color=COLORS["orange"]):
            tk.Frame(inner, bg=color, height=2).pack(fill="x", padx=20, pady=(14,2))
            tk.Label(inner, text=txt, font=("Helvetica",11,"bold"),
                     bg=COLORS["bg_card"], fg=color).pack(anchor="w", padx=20, pady=2)

        def row2(lbl, val, fg=None):
            r = tk.Frame(inner, bg=COLORS["bg_card"])
            r.pack(fill="x", padx=20, pady=1)
            tk.Label(r, text=lbl+":", font=("Helvetica",8,"bold"), width=18, anchor="w",
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
            tk.Label(r, text=str(val or "—"), font=("Helvetica",9),
                     bg=COLORS["bg_card"], fg=fg or COLORS["text_primary"]).pack(side="left")

        # ── Cabeçalho com badge de tipo destacado ─────────────────────────
        STATUS_BADGE_COLORS = {
            "Estoque": COLORS["green"],
            "Cliente": COLORS["blue"],
            "Daisha":  COLORS["orange"],
            "Inativo": COLORS["red"],
        }
        status_atual = carro.get("status","")
        badge_color  = STATUS_BADGE_COLORS.get(status_atual, COLORS["text_muted"])

        hdr_frame = tk.Frame(inner, bg=COLORS["bg_card"])
        hdr_frame.pack(fill="x", padx=20, pady=(16, 2))

        # Lado esquerdo: título
        hdr_left = tk.Frame(hdr_frame, bg=COLORS["bg_card"])
        hdr_left.pack(side="left", fill="both", expand=True)

        ano_cor = ""
        if carro.get("ano"):  ano_cor += f"  {carro['ano']}"
        if carro.get("cor"):  ano_cor += f"  · {carro['cor']}"
        tk.Label(hdr_left,
                 text=f"DOSSIÊ DO VEÍCULO — #{carro['id']}  {carro_nome}{ano_cor}",
                 font=("Helvetica",12,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w")

        sub_parts = []
        if carro.get("placa"):  sub_parts.append(f"Placa: {carro['placa']}")
        if carro.get("chassi"): sub_parts.append(f"Chassi: {carro['chassi']}")
        if sub_parts:
            tk.Label(hdr_left, text="  ".join(sub_parts),
                     font=("Helvetica",9),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(anchor="w", pady=(2,0))

        # Lado direito: badge grande do status atual
        badge_frame = tk.Frame(hdr_frame, bg=badge_color, padx=16, pady=8)
        badge_frame.pack(side="right", anchor="n", padx=(12,0))
        tk.Label(badge_frame, text=status_atual,
                 font=("Helvetica",14,"bold"),
                 bg=badge_color, fg="white").pack()

        # Linha separadora colorida
        tk.Frame(inner, bg=badge_color, height=3).pack(fill="x", padx=20, pady=(6,0))

        # Histórico de tipos anteriores (discreto, logo abaixo do separador)
        hist_status = carro.get("historico_status","") or ""
        if hist_status.strip():
            hist_frame = tk.Frame(inner, bg=COLORS["bg_content"])
            hist_frame.pack(fill="x", padx=20, pady=(2,0))
            tk.Label(hist_frame, text="Tipos anteriores:",
                     font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(side="left", padx=(4,6))
            for entrada in hist_status.split("|"):
                entrada = entrada.strip()
                if not entrada: continue
                tipo_h = entrada.split("(")[0].strip()
                cor_h  = STATUS_BADGE_COLORS.get(tipo_h, COLORS["text_muted"])
                tk.Label(hist_frame, text=entrada,
                         font=("Helvetica",7),
                         bg=cor_h, fg="white",
                         padx=6, pady=1).pack(side="left", padx=(0,4))

        section("▸ Dados do Veículo", COLORS["orange"])
        row2("Veículo", carro.get("carro"))
        row2("Ano",     carro.get("ano"))
        row2("Cor",     carro.get("cor"))
        row2("Placa",   carro.get("placa"))
        row2("Chassi",  carro.get("chassi"))
        row2("Status",  status_atual, badge_color)

        # ── Dados do Cliente (apenas quando carro.status == 'Cliente') ────────
        if carro.get("status") == "Cliente" and carro.get("cliente_id"):
            cli_row = self.conn.execute(
                "SELECT * FROM clientes WHERE id=?",
                (carro["cliente_id"],)).fetchone()
            if cli_row:
                cli_d = dict(cli_row)
                section("▸ Dados do Cliente Vinculado", COLORS["accent"])
                row2("Nome",       cli_d.get("nome"),       COLORS["accent"])
                row2("Telefone",   cli_d.get("telefone"))
                row2("ID Cliente", f"#{cli_d['id']}")

        # Resumo
        section("▸ Resumo", COLORS["green"])
        try:
            total_compra = sum(int(str(c.get("valor") or "0").replace(",","")) for c in compras)
            total_venda  = sum(int(str(v.get("valor_venda") or "0").replace(",","")) for v in vendas)
            total_serv   = sum(int(str(s.get("valor") or "0").replace(",","")) for s in servicos)
        except Exception:
            total_compra = total_venda = total_serv = 0
        row2("Compras Registradas", len(compras))
        row2("Total Comprado",      f"¥ {total_compra:,}", COLORS["orange"])
        row2("Vendas Registradas",  len(vendas))
        row2("Total Vendido",       f"¥ {total_venda:,}", COLORS["green"])
        row2("Serviços",            len(servicos))
        row2("Total Serviços",      f"¥ {total_serv:,}", COLORS["blue"])

        # Histórico de compras (estilo grid igual ao histórico de compras principal)
        section("▸ Histórico de Compras", COLORS["blue"])
        if not compras:
            tk.Label(inner, text="Nenhuma compra registrada.",
                     font=("Helvetica",8), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        COMP_DOSS = [
            ("#",35),("Data",80),("Tipo",80),("V.Compra",90),("Custo",90),("Total",90),("Cliente",120),
        ]
        for c in compras:
            custo_t = self._get_custo_total(c["id"])
            try: vcp = int(str(c["valor"]).replace(",","")) if c["valor"] else 0
            except: vcp = 0
            cc = tk.Frame(inner, bg=COLORS["bg_content"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            cc.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(cc, bg=COLORS["blue"], width=4).pack(side="left", fill="y")
            ci = tk.Frame(cc, bg=COLORS["bg_content"])
            ci.pack(side="left", fill="both", expand=True, padx=10, pady=6)
            # Header com badge "Compra de" à direita (padrão dossiê)
            comprado_de_nome = c.get("cliente_nome") or "—"
            hdr_ci = tk.Frame(ci, bg=COLORS["bg_content"])
            hdr_ci.pack(fill="x")
            tk.Label(hdr_ci,
                     text=f"#{c['id']}  —  {c.get('data_entrada') or c.get('data_compra') or '—'}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
            tk.Label(hdr_ci,
                     text=f"Compra de: {comprado_de_nome}",
                     font=("Helvetica",8,"bold"),
                     bg=COLORS["accent"], fg="white", padx=6
                     ).pack(side="right", padx=(4,0))
            row_d = tk.Frame(ci, bg=COLORS["bg_content"])
            row_d.pack(fill="x", pady=(4,0))
            TIPO_C = {"Leilão":COLORS["orange"],"Troca":COLORS["blue"],"Compra Direta":COLORS["green"]}
            tc = TIPO_C.get(c.get("tipo",""), COLORS["text_muted"])
            for lbl_t, val, fg in [
                ("Tipo",     c.get("tipo","—"),                             tc),
                ("V.Compra", self._fmt_yen_display(c.get("valor")),         COLORS["orange"]),
                ("Custos",   self._fmt_yen_display(custo_t),                COLORS["text_secondary"]),
                ("Total",    self._fmt_yen_display(vcp+custo_t),            COLORS["green"]),
            ]:
                col_f = tk.Frame(row_d, bg=COLORS["bg_content"])
                col_f.pack(side="left", padx=(0,16))
                tk.Label(col_f, text=lbl_t, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col_f, text=val, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg).pack(anchor="w")

        # Histórico de vendas
        section("▸ Histórico de Vendas", COLORS["accent"])
        if not vendas:
            tk.Label(inner, text="Nenhuma venda registrada.",
                     font=("Helvetica",8), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        for v in vendas:
            vc = tk.Frame(inner, bg=COLORS["bg_content"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            vc.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(vc, bg=COLORS["accent"], width=4).pack(side="left", fill="y")
            vi = tk.Frame(vc, bg=COLORS["bg_content"])
            vi.pack(side="left", fill="both", expand=True, padx=10, pady=6)
            vendido_para_nome = v.get("cliente_nome") or "—"
            hdr_vi = tk.Frame(vi, bg=COLORS["bg_content"])
            hdr_vi.pack(fill="x")
            tk.Label(hdr_vi,
                     text=f"#{v['id']} — {v.get('data_venda','—')}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
            tk.Label(hdr_vi,
                     text=f"Vendido para: {vendido_para_nome}",
                     font=("Helvetica",8,"bold"),
                     bg=COLORS["green"], fg="white", padx=6
                     ).pack(side="right", padx=(4,0))
            row_d = tk.Frame(vi, bg=COLORS["bg_content"])
            row_d.pack(fill="x", pady=(2,0))
            pagas = v.get("parcelas_pagas") or 0
            tot_p = v.get("num_parcelas") or 0
            is_p  = v.get("tipo_venda","") in ("Venda Parcelada","Com Troca","Com Troca e Volta")
            quit  = is_p and tot_p > 0 and pagas >= tot_p
            st    = "Quitado" if quit else ("Em Aberto" if is_p else "À Vista")
            dcar_det = [
                ("Tipo",      v.get("tipo_venda","—"),                              COLORS["text_secondary"]),
                ("Valor",     self._fmt_yen_display(v.get("valor_venda")),          COLORS["green"]),
                ("Parcela",   self._fmt_yen_display(v.get("parcela_mensal")) if is_p else "—", COLORS["blue"]),
                ("Progresso", f"{pagas}/{tot_p}" if is_p else "—",                 COLORS["accent"]),
                ("Status",    st, COLORS["green"] if (quit or not is_p) else COLORS["orange"]),
            ]
            if v.get("carro_troca"):
                ct2 = (v.get("carro_troca") or "—").split("|")[0].strip()
                vt2 = self._fmt_yen_display(v.get("valor_troca")) if v.get("valor_troca") else "—"
                dcar_det += [("Troca", ct2[:18], COLORS["orange"]), ("V.Troca", vt2, COLORS["orange"])]
            for lbl, val, fg in dcar_det:
                col = tk.Frame(row_d, bg=COLORS["bg_content"])
                col.pack(side="left", padx=(0,14))
                tk.Label(col, text=lbl, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col, text=val, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg).pack(anchor="w")

        # Serviços
        section("▸ Serviços Realizados", COLORS["orange"])
        if not servicos:
            tk.Label(inner, text="Nenhum serviço registrado.",
                     font=("Helvetica",8), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        for s in servicos:
            sc = tk.Frame(inner, bg=COLORS["bg_content"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
            sc.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(sc, bg=COLORS["orange"], width=4).pack(side="left", fill="y")
            si = tk.Frame(sc, bg=COLORS["bg_content"])
            si.pack(side="left", fill="both", expand=True, padx=10, pady=6)
            tk.Label(si, text=f"OS {s.get('os_num') or s['id']} — {s.get('data_servico','—')}  |  Cliente: {s.get('cliente_nome') or '—'}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(anchor="w")
            row_d = tk.Frame(si, bg=COLORS["bg_content"])
            row_d.pack(fill="x", pady=(2,0))
            st_col = {"Concluído":COLORS["green"],"Em Andamento":COLORS["orange"],
                      "Aberto":COLORS["blue"]}.get(s.get("status",""), COLORS["text_muted"])
            for lbl, val, fg in [
                ("Tipo Serviço", s.get("tipo_servico","—"), COLORS["text_secondary"]),
                ("Valor", self._fmt_yen_display(s.get("valor")), COLORS["green"]),
                ("Status", s.get("status","—"), st_col),
                ("Descrição", (s.get("descricao") or "—")[:35], COLORS["text_muted"]),
            ]:
                col = tk.Frame(row_d, bg=COLORS["bg_content"])
                col.pack(side="left", padx=(0,14))
                tk.Label(col, text=lbl, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col, text=val, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg).pack(anchor="w")

        # ── Shaken do veículo ─────────────────────────────────────────────────
        section("▸ Histórico de Shaken", COLORS["blue"])
        shakens = [dict(r) for r in self.conn.execute(
            "SELECT s.*, cl.nome as cli_nome FROM shaken s "
            "LEFT JOIN clientes cl ON s.cliente_id=cl.id "
            "WHERE s.carro_id=? ORDER BY s.id DESC", (car_id,)).fetchall()]
        if not shakens:
            tk.Label(inner, text="Nenhum shaken registrado para este veículo.",
                     font=("Helvetica",8), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(anchor="w", padx=20, pady=4)
        for sk in shakens:
            st_txt, st_cor = self._sk_status_info(
                sk.get("data_vencimento"), sk.get("por_conta"))
            skc = tk.Frame(inner, bg=COLORS["bg_content"],
                           highlightthickness=1, highlightbackground=COLORS["border"])
            skc.pack(fill="x", padx=20, pady=(4,0))
            tk.Frame(skc, bg=st_cor, width=4).pack(side="left", fill="y")
            sk_i = tk.Frame(skc, bg=COLORS["bg_content"])
            sk_i.pack(side="left", fill="both", expand=True, padx=10, pady=6)
            hs2 = tk.Frame(sk_i, bg=COLORS["bg_content"])
            hs2.pack(fill="x")
            tk.Label(hs2, text=f"{sk['sk_num']}  —  Registrado: {sk.get('data_registro','—')}",
                     font=("Helvetica",9,"bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
            tk.Label(hs2, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_cor, fg="white").pack(side="right", padx=4)
            ds2 = tk.Frame(sk_i, bg=COLORS["bg_content"])
            ds2.pack(fill="x", pady=(2,0))
            d_venc = sk.get("data_vencimento","") or ("Por Conta" if sk.get("por_conta") else "—")
            vl_sk = self._fmt_yen_display(sk.get("valor")) if sk.get("valor") else "—"
            cli_sk = sk.get("cli_nome","") or "—"
            for lbl_t, val_t, fg_t in [
                ("Vencimento", d_venc, st_cor),
                ("Valor", vl_sk, COLORS["green"]),
                ("Cliente", cli_sk, COLORS["text_secondary"]),
                ("Obs", sk.get("obs","") or "—", COLORS["text_muted"]),
            ]:
                col = tk.Frame(ds2, bg=COLORS["bg_content"])
                col.pack(side="left", padx=(0,14))
                tk.Label(col, text=lbl_t, font=("Helvetica",7),
                         bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(anchor="w")
                tk.Label(col, text=val_t, font=("Helvetica",8,"bold"),
                         bg=COLORS["bg_content"], fg=fg_t).pack(anchor="w")

        # ── Histórico de Inativação ──────────────────────────────────────
        data_inat = carro.get("data_inativacao")
        obs_inat  = carro.get("obs_inativacao")
        if carro.get("status") == "Inativo" or data_inat:
            section("⛔  Inativação", COLORS["red"])
            row2("Status",        "INATIVO", COLORS["red"])
            row2("Data Inativação", data_inat or "—", COLORS["red"])
            if obs_inat:
                tk.Frame(inner, bg=COLORS["bg_card"]).pack(fill="x", padx=20, pady=(2,0))
                obs_f = tk.Frame(inner, bg=COLORS["bg_content"],
                                 highlightthickness=1, highlightbackground=COLORS["border"])
                obs_f.pack(fill="x", padx=20, pady=(4, 6))
                tk.Label(obs_f, text="Motivo / Observação:",
                         font=("Helvetica", 8, "bold"),
                         bg=COLORS["bg_content"], fg=COLORS["text_secondary"]
                         ).pack(anchor="w", padx=10, pady=(6,2))
                tk.Label(obs_f, text=obs_inat,
                         font=("Helvetica", 9),
                         bg=COLORS["bg_content"], fg=COLORS["text_primary"],
                         wraplength=480, justify="left"
                         ).pack(anchor="w", padx=10, pady=(0,8))

        tk.Frame(inner, bg=COLORS["bg_card"], height=20).pack()

    def _dcar_exportar_pdf(self):
        if not self._dcar_carro_id:
            from tkinter import messagebox as msgbox
            msgbox.showwarning("Atenção", "Gere o dossiê primeiro.")
            return
        self._gerar_pdf_carro(self._dcar_carro_id)

    def _gerar_pdf_carro(self, car_id):
        import os, datetime
        from tkinter import filedialog, messagebox as msgbox
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle, HRFlowable)
        except ImportError:
            msgbox.showerror("Erro", "Instale reportlab: pip install reportlab")
            return

        carro = dict(self.conn.execute("SELECT * FROM carros WHERE id=?", (car_id,)).fetchone())
        carro_nome = carro.get("carro","").strip()

        compras = [dict(r) for r in self.conn.execute(
            "SELECT co.*, cl.nome as cliente_nome FROM compras co "
            "LEFT JOIN clientes cl ON co.cliente_id=cl.id "
            "WHERE co.carro_id=? ORDER BY co.id DESC", (car_id,)).fetchall()]
        vendas = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id=c.id "
            "WHERE v.carro LIKE ? ORDER BY v.id DESC", (f"%{carro_nome}%",)).fetchall()]
        servicos = [dict(r) for r in self.conn.execute(
            "SELECT s.*, c.nome as cliente_nome FROM servicos s "
            "LEFT JOIN clientes c ON s.cliente_id=c.id "
            "WHERE s.carro LIKE ? ORDER BY s.id DESC", (f"%{carro_nome}%",)).fetchall()]

        nome_arq = f"Dossie_Carro_{carro_nome.replace(' ','_')}_{car_id}.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", initialfile=nome_arq,
            filetypes=[("PDF","*.pdf")], title="Salvar Dossiê Carro PDF")
        if not path:
            return

        doc = SimpleDocTemplate(path, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        ORANGE = rl_colors.HexColor("#E67E22")
        BLUE   = rl_colors.HexColor("#2B7BE5")
        GREEN  = rl_colors.HexColor("#27AE60")
        GRAY   = rl_colors.HexColor("#7F8C8D")

        def h1(txt, color=ORANGE):
            return Paragraph(f'<font color="{color.hexval()}" size="14"><b>{txt}</b></font>', styles["Normal"])
        def h2(txt, color=ORANGE):
            return Paragraph(f'<font color="{color.hexval()}" size="11"><b>{txt}</b></font>', styles["Normal"])
        def body(txt):
            return Paragraph(f'<font size="9">{txt}</font>', styles["Normal"])

        story = []
        hoje_str = datetime.date.today().strftime("%d/%m/%Y")

        story.append(h1(f"DOSSIÊ DO VEÍCULO — #{carro['id']}  {carro_nome}"))
        story.append(body(f"Gerado em {hoje_str}"))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=ORANGE))
        story.append(Spacer(1, 0.3*cm))

        # Dados do veículo — com status e histórico
        story.append(h2("▸ DADOS DO VEÍCULO"))
        story.append(Spacer(1, 0.15*cm))
        status_atual_pdf = carro.get("status") or "—"
        STATUS_COLORS_PDF = {
            "Estoque": rl_colors.HexColor("#27AE60"),
            "Cliente": rl_colors.HexColor("#2B7BE5"),
            "Daisha":  rl_colors.HexColor("#E67E22"),
            "Inativo": rl_colors.HexColor("#E74C3C"),
        }
        status_color_pdf = STATUS_COLORS_PDF.get(status_atual_pdf, GRAY)
        campos = [["Veículo", carro.get("carro") or "—"],
                  ["Ano",     str(carro.get("ano") or "—")],
                  ["Cor",     carro.get("cor") or "—"],
                  ["Placa",   carro.get("placa") or "—"],
                  ["Chassi",  carro.get("chassi") or "—"],
                  ["Status",  status_atual_pdf]]
        rows_pdf = []
        for k, v in campos:
            if k == "Status":
                rows_pdf.append([body(f"<b>{k}</b>"),
                                  Paragraph(f'<font color="{status_color_pdf.hexval()}" size="10"><b>{v}</b></font>',
                                            styles["Normal"])])
            else:
                rows_pdf.append([body(f"<b>{k}</b>"), body(v)])
        t = Table(rows_pdf, colWidths=[4*cm,12*cm])
        t.setStyle(TableStyle([
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[rl_colors.HexColor("#FFF8F0"),rl_colors.white]),
            ("GRID",(0,0),(-1,-1),0.3,rl_colors.HexColor("#E0E0E0")),
            ("LEFTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),4),
            ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        story.append(t)
        hist_pdf = (carro.get("historico_status") or "").strip()
        if hist_pdf:
            story.append(Spacer(1, 0.2*cm))
            story.append(body(f"<font color='grey' size='8'><b>Tipos anteriores:</b>  {hist_pdf}</font>"))
        story.append(Spacer(1, 0.4*cm))

        # Compras
        story.append(h2("▸ HISTÓRICO DE COMPRAS", BLUE))
        story.append(Spacer(1, 0.15*cm))
        if not compras:
            story.append(body("Nenhuma compra registrada."))
        else:
            ch = [["#","Data","Tipo Compra","Compra de","Valor Compra","Custos","Total"]]
            for c in compras:
                val_c  = int(str(c.get("valor") or "0").replace(",",""))
                cust_c = self._get_custo_total(c["id"])
                total_c = val_c + cust_c
                ch.append([
                    str(c["id"]),
                    c.get("data_entrada","—"),
                    c.get("tipo") or "—",
                    c.get("cliente_nome") or "—",
                    f"¥ {val_c:,}" if val_c else "—",
                    f"¥ {cust_c:,}" if cust_c else "—",
                    f"¥ {total_c:,}" if total_c else "—",
                ])
            ct = Table([[body(x) for x in row] for row in ch],
                       colWidths=[0.8*cm, 2*cm, 2.5*cm, 3.5*cm, 2.7*cm, 2.5*cm, 3*cm])
            ct.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),BLUE),("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#EAF6FF"),rl_colors.white]),
                ("GRID",(0,0),(-1,-1),0.3,rl_colors.HexColor("#E0E0E0")),
                ("FONTSIZE",(0,0),(-1,-1),8),("LEFTPADDING",(0,0),(-1,-1),4),
                ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ]))
            story.append(ct)
        story.append(Spacer(1, 0.4*cm))

        # Vendas
        story.append(h2("▸ HISTÓRICO DE VENDAS", GREEN))
        story.append(Spacer(1, 0.15*cm))
        if not vendas:
            story.append(body("Nenhuma venda registrada."))
        else:
            vh = [["#","Data","Vendido para","Tipo Venda","Valor","Parcela","Status"]]
            for v in vendas:
                pagas=v.get("parcelas_pagas") or 0; tot_p=v.get("num_parcelas") or 0
                is_p=v.get("tipo_venda","") in ("Venda Parcelada","Com Troca","Com Troca e Volta")
                quit=is_p and tot_p>0 and pagas>=tot_p
                st="Quitado" if quit else ("Em Aberto" if is_p else "À Vista")
                vh.append([str(v["id"]),v.get("data_venda","—"),
                            v.get("cliente_nome") or "—",
                            v.get("tipo_venda","—"),
                            f"¥ {v.get('valor_venda') or '—'}",
                            f"¥ {v.get('parcela_mensal') or '—'}" if is_p else "—", st])
            vt = Table([[body(x) for x in row] for row in vh],
                       colWidths=[0.8*cm,2*cm,3.5*cm,3.2*cm,2.3*cm,2.3*cm,2.9*cm])
            vt.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),GREEN),("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#EAFAF1"),rl_colors.white]),
                ("GRID",(0,0),(-1,-1),0.3,rl_colors.HexColor("#E0E0E0")),
                ("FONTSIZE",(0,0),(-1,-1),8),("LEFTPADDING",(0,0),(-1,-1),4),
                ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ]))
            story.append(vt)
        story.append(Spacer(1, 0.4*cm))

        # Serviços
        story.append(h2("▸ SERVIÇOS REALIZADOS", ORANGE))
        story.append(Spacer(1, 0.15*cm))
        if not servicos:
            story.append(body("Nenhum serviço registrado."))
        else:
            sh = [["OS","Data","Cliente","Tipo Serviço","Valor","Status"]]
            for s in servicos:
                sh.append([str(s.get("os_num") or s["id"]),s.get("data_servico","—"),
                            s.get("cliente_nome") or "—",s.get("tipo_servico") or "—",
                            f"¥ {s.get('valor') or '—'}",s.get("status","—")])
            st2 = Table([[body(x) for x in row] for row in sh],
                        colWidths=[1.2*cm,2.2*cm,4*cm,4*cm,2.5*cm,3*cm])
            st2.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),ORANGE),("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#FFF8F0"),rl_colors.white]),
                ("GRID",(0,0),(-1,-1),0.3,rl_colors.HexColor("#E0E0E0")),
                ("FONTSIZE",(0,0),(-1,-1),8),("LEFTPADDING",(0,0),(-1,-1),4),
                ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ]))
            story.append(st2)

        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
        story.append(body(f"<font color='grey'>KM Cars — Dossiê gerado em {hoje_str}</font>"))
        doc.build(story)
        try:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass
        msgbox.showinfo("PDF Gerado", f"Dossiê salvo em:\n{path}")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: PREVISÃO DE RECEBIMENTO
    # ══════════════════════════════════════════════════════════════════════════
    def _build_previsao_recebimento(self, parent):
        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        card = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=COLORS["accent"], height=4).pack(fill="x")

        # Header
        hdr = tk.Frame(card, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12,8))
        tk.Label(hdr, text="◎  Previsão de Recebimento",
                 font=("Helvetica",13,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica",8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=lambda: self._refresh_previsao(body_frame, kpi_frame),
                  ).pack(side="right", ipady=3, padx=(0,4))
        tk.Button(hdr, text="📄  Exportar PDF", font=("Helvetica",8),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  command=self._pdf_previsao_recebimento,
                  ).pack(side="right", ipady=3, padx=(0,6))

        # KPI destaque
        kpi_frame = tk.Frame(card, bg=COLORS["bg_card"])
        kpi_frame.pack(fill="x", padx=16, pady=(0,10))

        # Corpo (scroll)
        sf = tk.Frame(card, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=16, pady=(0,12))
        cv = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(sf, orient="vertical", command=cv.yview)
        body_frame = tk.Frame(cv, bg=COLORS["bg_card"])
        body_frame.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        win_id = cv.create_window((0,0), window=body_frame, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.bind("<Configure>", lambda e: cv.itemconfig(win_id, width=e.width))
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        card.bind("<Enter>", lambda e: card.bind_all("<MouseWheel>",
            lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        card.bind("<Leave>", lambda e: card.unbind_all("<MouseWheel>"))

        self._refresh_previsao(body_frame, kpi_frame)

    def _calcular_previsao(self):
        """Retorna lista de (ano, mes, total_parcelas) e total geral.
        Usa saldo devedor real (igual a Dívidas Clientes) para consistência.
        A última parcela usa valor_ultima_parc se definido.
        """
        import datetime
        todas = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id=c.id "
            "WHERE v.tipo_venda IN ('Venda Parcelada','Com Troca','Com Troca e Volta') "
            "AND v.parcelas_pagas < v.num_parcelas "
            "ORDER BY v.id DESC").fetchall()]

        from collections import defaultdict
        por_mes = defaultdict(int)

        def _i(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0

        for v in todas:
            pagas      = v.get("parcelas_pagas") or 0
            total_parc = v.get("num_parcelas") or 0
            restantes  = total_parc - pagas
            if restantes <= 0:
                continue

            parc_val       = _i(v.get("parcela_mensal"))
            ult_parc_val   = _i(v.get("valor_ultima_parc")) or parc_val

            # Saldo real devedor = valor_venda - entrada - troca - total já pago
            total_pago_real = self.conn.execute(
                "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                "FROM pagamentos WHERE venda_id=?", (v["id"],)).fetchone()[0]
            saldo_real = max(0, _i(v.get("valor_venda"))
                            - _i(v.get("entrada"))
                            - _i(v.get("valor_troca"))
                            - total_pago_real)

            if saldo_real <= 0 or parc_val <= 0:
                continue

            # Distribui saldo real pelos meses futuros
            # Parcelas intermediárias = parc_val; última = ult_parc_val (ou ajustada ao saldo)
            data_1a = v.get("data_primeira_parc")
            if not data_1a:
                continue
            try:
                pts = data_1a.split("/")
                m0  = int(pts[1]); a0 = int(pts[2])
            except Exception:
                continue

            # Monta lista de valores para as parcelas restantes
            valores_restantes = []
            for idx in range(pagas, total_parc):
                is_ultima = (idx == total_parc - 1)
                valores_restantes.append(ult_parc_val if is_ultima else parc_val)

            # Ajusta proporcionalmente ao saldo real (evita divergência)
            soma_proj = sum(valores_restantes)
            if soma_proj > 0 and soma_proj != saldo_real:
                fator = saldo_real / soma_proj
                valores_restantes = [round(vv * fator) for vv in valores_restantes]
                # Ajuste de arredondamento na última
                diff = saldo_real - sum(valores_restantes)
                if diff != 0 and valores_restantes:
                    valores_restantes[-1] += diff

            for k, (idx, val) in enumerate(zip(range(pagas, total_parc), valores_restantes)):
                mt = m0 + idx
                ay = a0 + (mt - 1) // 12
                mo = ((mt - 1) % 12) + 1
                por_mes[(ay, mo)] += val

        total_geral = sum(por_mes.values())
        meses_ord   = sorted(por_mes.keys())
        return meses_ord, por_mes, total_geral, todas

    def _refresh_previsao(self, body_frame, kpi_frame):
        import datetime
        MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        for w in body_frame.winfo_children():
            w.destroy()
        for w in kpi_frame.winfo_children():
            w.destroy()

        meses_ord, por_mes, total_geral, todas = self._calcular_previsao()

        # KPI destacado
        kpi = tk.Frame(kpi_frame, bg=COLORS["bg_content"],
                       highlightthickness=2, highlightbackground=COLORS["accent"])
        kpi.pack(fill="x")
        tk.Frame(kpi, bg=COLORS["accent"], height=4).pack(fill="x")
        kpi_inner = tk.Frame(kpi, bg=COLORS["bg_content"])
        kpi_inner.pack(fill="x", padx=20, pady=12)
        tk.Label(kpi_inner, text="TOTAL A RECEBER — Todas as Parcelas em Aberto",
                 font=("Helvetica",9,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_secondary"]).pack(side="left")
        tk.Label(kpi_inner, text=f"¥ {total_geral:,}",
                 font=("Helvetica",20,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["accent"]).pack(side="right")
        tk.Label(kpi_inner, text=f"{len(todas)} financiamento(s) em aberto",
                 font=("Helvetica",8),
                 bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(side="right", padx=16)

        if not meses_ord:
            tk.Label(body_frame, text="Nenhuma parcela em aberto.",
                     font=("Helvetica",11), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=40)
            return

        # Cabeçalho tabela
        col_hdr = tk.Frame(body_frame, bg=COLORS["bg_main"])
        col_hdr.pack(fill="x")
        tk.Frame(body_frame, bg=COLORS["border"], height=1).pack(fill="x")
        for txt, w, anch in [("Mês / Ano",18,"w"),("Qtd. Parcelas",14,"c"),
                               ("Total a Receber",16,"w"),("Barra",24,"w")]:
            tk.Label(col_hdr, text=txt, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor=anch).pack(side="left", padx=4, pady=6)

        hoje = datetime.date.today()
        max_val = max(por_mes.values()) or 1

        # Contar parcelas por mês
        from collections import defaultdict
        qtd_por_mes = defaultdict(int)
        todas_ativas = [dict(r) for r in self.conn.execute(
            "SELECT * FROM vendas WHERE tipo_venda IN "
            "('Venda Parcelada','Com Troca','Com Troca e Volta') "
            "AND parcelas_pagas < num_parcelas").fetchall()]
        for v in todas_ativas:
            pagas  = v.get("parcelas_pagas") or 0
            tot_p  = v.get("num_parcelas") or 0
            data_1a = v.get("data_primeira_parc")
            if not data_1a: continue
            try:
                pts = data_1a.split("/")
                d0 = int(pts[0]); m0 = int(pts[1]); a0 = int(pts[2])
            except Exception:
                continue
            for idx in range(pagas, tot_p):
                mt = m0 + idx
                a  = a0 + (mt-1)//12
                m  = ((mt-1)%12)+1
                qtd_por_mes[(a,m)] += 1

        for i, (a, m) in enumerate(meses_ord):
            is_atual = (a == hoje.year and m == hoje.month)
            rb = COLORS["bg_content"] if is_atual else (COLORS["bg_card"] if i%2==0 else COLORS["bg_content"])
            row = tk.Frame(body_frame, bg=rb,
                           highlightthickness=1 if is_atual else 0,
                           highlightbackground=COLORS["accent"])
            row.pack(fill="x", pady=(0,1))

            val = por_mes[(a,m)]
            qtd = qtd_por_mes[(a,m)]
            pct = val / max_val

            # Mês/Ano
            mes_txt = f"{MESES_PT[m-1]}  {a}"
            fg_mes  = COLORS["accent"] if is_atual else COLORS["text_primary"]
            lbl_m = tk.Label(row, text=mes_txt + (" ◀ Atual" if is_atual else ""),
                             font=("Helvetica",9,"bold" if is_atual else "normal"),
                             bg=rb, fg=fg_mes, width=18, anchor="w")
            lbl_m.pack(side="left", padx=4, pady=8)

            # Qtd parcelas
            tk.Label(row, text=str(qtd), font=("Helvetica",9,"bold"),
                     bg=rb, fg=COLORS["blue"], width=14, anchor="center").pack(side="left", padx=4)

            # Total
            tk.Label(row, text=f"¥ {val:,}", font=("Helvetica",10,"bold"),
                     bg=rb, fg=COLORS["green"], width=16, anchor="w").pack(side="left", padx=4)

            # Mini barra
            bar_host = tk.Frame(row, bg=rb)
            bar_host.pack(side="left", fill="x", expand=True, padx=(4,8))
            bar_bg = tk.Frame(bar_host, bg=COLORS["border"], height=10)
            bar_bg.pack(fill="x", pady=6)
            def _draw_bar(e, _bg=bar_bg, _p=pct, _atual=is_atual):
                bw = max(2, int(_bg.winfo_width() * _p))
                c = COLORS["accent"] if _atual else COLORS["green"]
                tk.Frame(_bg, bg=c, width=bw, height=10).place(x=0, y=0)
            bar_bg.bind("<Configure>", _draw_bar)

    def _pdf_previsao_recebimento(self):
        import datetime
        from tkinter import filedialog, messagebox as msgbox
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle, HRFlowable)
        except ImportError:
            msgbox.showerror("Erro", "Instale reportlab: pip install reportlab"); return

        MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        meses_ord, por_mes, total_geral, todas = self._calcular_previsao()

        hoje_str = datetime.date.today().strftime("%d/%m/%Y")
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", initialfile="Previsao_Recebimento.pdf",
            filetypes=[("PDF","*.pdf")], title="Salvar Previsão PDF")
        if not path: return

        ACCENT = rl_colors.HexColor("#5E43F3")
        GREEN  = rl_colors.HexColor("#27AE60")
        BLUE   = rl_colors.HexColor("#2B7BE5")
        GRAY   = rl_colors.HexColor("#7F8C8D")

        doc = SimpleDocTemplate(path, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        def h(txt, c=ACCENT, sz=13):
            return Paragraph(f'<font color="{c.hexval()}" size="{sz}"><b>{txt}</b></font>', styles["Normal"])
        def b(txt, sz=9):
            return Paragraph(f'<font size="{sz}">{txt}</font>', styles["Normal"])

        story = []
        story.append(h("PREVISÃO DE RECEBIMENTO — Parcelas em Aberto"))
        story.append(b(f"Gerado em {hoje_str}"))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
        story.append(Spacer(1, 0.3*cm))

        # KPI total
        kpi_data = [[b("<b>TOTAL GERAL A RECEBER</b>", 11),
                     h(f"¥ {total_geral:,}", GREEN, 14)]]
        kt = Table(kpi_data, colWidths=[9*cm, 8*cm])
        kt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), rl_colors.HexColor("#F0F8FF")),
            ("BOX", (0,0), (-1,-1), 2, ACCENT),
            ("LEFTPADDING", (0,0), (-1,-1), 12),
            ("TOPPADDING", (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(kt)
        story.append(Spacer(1, 0.4*cm))

        # Tabela mensal
        story.append(h("Detalhamento por Mês", BLUE, 11))
        story.append(Spacer(1, 0.2*cm))

        rows = [[b("<b>Mês / Ano</b>"), b("<b>Total a Receber</b>"), b("<b>% do Total</b>")]]
        hoje = datetime.date.today()
        for (a, m) in meses_ord:
            val = por_mes[(a,m)]
            pct = val/total_geral*100 if total_geral else 0
            is_atual = (a==hoje.year and m==hoje.month)
            lbl = f"{MESES_PT[m-1]} {a}" + (" ← Mês Atual" if is_atual else "")
            rows.append([b(f"<b>{lbl}</b>" if is_atual else lbl),
                         b(f"¥ {val:,}"),
                         b(f"{pct:.1f}%")])

        t = Table(rows, colWidths=[7*cm, 6*cm, 4*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), BLUE),
            ("TEXTCOLOR", (0,0), (-1,0), rl_colors.white),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [rl_colors.HexColor("#F0F8FF"), rl_colors.white]),
            ("GRID", (0,0), (-1,-1), 0.3, rl_colors.HexColor("#E0E0E0")),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
        story.append(b(f"<font color='grey'>KM Cars — Previsão gerada em {hoje_str}</font>"))
        doc.build(story)
        try:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass
        msgbox.showinfo("PDF Gerado", f"Salvo em:\n{path}")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: DÍVIDAS CLIENTES
    # ══════════════════════════════════════════════════════════════════════════
    def _build_dividas_clientes(self, parent):
        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        card = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=COLORS["red"], height=4).pack(fill="x")

        # Header
        hdr = tk.Frame(card, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=16, pady=(12,8))
        tk.Label(hdr, text="◎  Dívidas dos Clientes",
                 font=("Helvetica",13,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica",8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=lambda: self._refresh_dividas(body_div, kpi_div)
                  ).pack(side="right", ipady=3, padx=(0,4))
        tk.Button(hdr, text="📄  Exportar PDF", font=("Helvetica",8),
                  bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                  command=self._pdf_dividas_clientes
                  ).pack(side="right", ipady=3, padx=(0,6))

        # KPI
        kpi_div = tk.Frame(card, bg=COLORS["bg_card"])
        kpi_div.pack(fill="x", padx=16, pady=(0,8))

        # Colunas
        tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16)
        col_hdr = tk.Frame(card, bg=COLORS["bg_main"])
        col_hdr.pack(fill="x", padx=16)
        for txt, w, anch in [("#",3,"w"),("Cliente",16,"w"),("Carro",16,"w"),
                               ("Venda",10,"w"),("V.Total",10,"w"),
                               ("Já Pago",10,"w"),("Saldo Dev.",11,"w"),("Status",9,"c")]:
            tk.Label(col_hdr, text=txt, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor=anch
                     ).pack(side="left", padx=2, pady=6)

        # Scroll
        sf = tk.Frame(card, bg=COLORS["bg_card"])
        sf.pack(fill="both", expand=True, padx=16, pady=(0,12))
        cv = tk.Canvas(sf, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(sf, orient="vertical", command=cv.yview)
        body_div = tk.Frame(cv, bg=COLORS["bg_card"])
        body_div.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        win_id = cv.create_window((0,0), window=body_div, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.bind("<Configure>", lambda e: cv.itemconfig(win_id, width=e.width))
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        card.bind("<Enter>", lambda e: card.bind_all("<MouseWheel>",
            lambda ev: cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        card.bind("<Leave>", lambda e: card.unbind_all("<MouseWheel>"))

        self._refresh_dividas(body_div, kpi_div)

    def _calcular_dividas(self):
        """Retorna lista de dívidas e total geral."""
        todas = [dict(r) for r in self.conn.execute(
            "SELECT v.*, c.nome as cliente_nome FROM vendas v "
            "LEFT JOIN clientes c ON v.cliente_id=c.id "
            "WHERE v.tipo_venda IN ('Venda Parcelada','Com Troca','Com Troca e Volta') "
            "ORDER BY v.id DESC").fetchall()]

        dividas = []
        for v in todas:
            pagas  = v.get("parcelas_pagas") or 0
            tot_p  = v.get("num_parcelas") or 0
            quitado = tot_p > 0 and pagas >= tot_p
            try:
                total_pago = self.conn.execute(
                    "SELECT COALESCE(SUM(CAST(valor_pago AS INTEGER)),0) "
                    "FROM pagamentos WHERE venda_id=?", (v["id"],)).fetchone()[0]
                vv = int(str(v.get("valor_venda") or "0").replace(",",""))
                et = int(str(v.get("entrada") or "0").replace(",",""))
                vt = int(str(v.get("valor_troca") or "0").replace(",",""))
                saldo = max(0, vv - et - vt - total_pago)
            except Exception:
                total_pago = 0; vv = 0; saldo = 0
            dividas.append({
                "id":       v["id"],
                "cliente":  v.get("cliente_nome") or "—",
                "carro":    v.get("carro") or "—",
                "data":     v.get("data_venda") or "—",
                "vv":       vv,
                "total_pago": total_pago,
                "saldo":    saldo,
                "quitado":  quitado,
                "pagas":    pagas,
                "tot_p":    tot_p,
            })
        total_divida = sum(d["saldo"] for d in dividas if not d["quitado"])
        return dividas, total_divida

    def _refresh_dividas(self, body_div, kpi_div):
        for w in body_div.winfo_children():
            w.destroy()
        for w in kpi_div.winfo_children():
            w.destroy()

        dividas, total_divida = self._calcular_dividas()

        # KPI
        kpi = tk.Frame(kpi_div, bg=COLORS["bg_content"],
                       highlightthickness=2, highlightbackground=COLORS["red"])
        kpi.pack(fill="x")
        tk.Frame(kpi, bg=COLORS["red"], height=4).pack(fill="x")
        ki = tk.Frame(kpi, bg=COLORS["bg_content"])
        ki.pack(fill="x", padx=20, pady=10)
        tk.Label(ki, text="TOTAL GERAL DE DÍVIDAS EM ABERTO",
                 font=("Helvetica",9,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_secondary"]).pack(side="left")
        tk.Label(ki, text=f"¥ {total_divida:,}",
                 font=("Helvetica",20,"bold"),
                 bg=COLORS["bg_content"], fg=COLORS["red"]).pack(side="right")
        em_aberto = sum(1 for d in dividas if not d["quitado"])
        tk.Label(ki, text=f"{em_aberto} em aberto  /  {len(dividas)} total",
                 font=("Helvetica",8),
                 bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(side="right", padx=16)

        if not dividas:
            tk.Label(body_div, text="Nenhum financiamento registrado.",
                     font=("Helvetica",11), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=40)
            return

        for i, d in enumerate(dividas):
            rb = COLORS["bg_card"] if i%2==0 else COLORS["bg_content"]
            row = tk.Frame(body_div, bg=rb)
            row.pack(fill="x")

            saldo_fg = COLORS["red"] if (d["saldo"] > 0 and not d["quitado"]) else COLORS["text_muted"]
            st_txt = "Quitado" if d["quitado"] else "Em Aberto"
            st_bg  = COLORS["green"] if d["quitado"] else COLORS["red"]

            tk.Label(row, text=str(d["id"]), font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_muted"], width=3, anchor="w").pack(side="left",padx=2,pady=7)
            tk.Label(row, text=d["cliente"][:16], font=("Helvetica",9,"bold"), bg=rb,
                     fg=COLORS["text_primary"], width=16, anchor="w").pack(side="left",padx=2)
            car = d["carro"][:16]+"…" if len(d["carro"])>16 else d["carro"]
            tk.Label(row, text=car, font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_secondary"], width=16, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=d["data"], font=("Helvetica",8), bg=rb,
                     fg=COLORS["text_muted"], width=10, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=f"¥{d['vv']:,}", font=("Helvetica",9,"bold"), bg=rb,
                     fg=COLORS["green"], width=10, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=f"¥{d['total_pago']:,}", font=("Helvetica",9), bg=rb,
                     fg=COLORS["blue"], width=10, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=f"¥{d['saldo']:,}" if d["saldo"] > 0 else "—",
                     font=("Helvetica",9,"bold"), bg=rb,
                     fg=saldo_fg, width=11, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=st_txt, font=("Helvetica",7,"bold"),
                     bg=st_bg, fg="white", width=9, anchor="center").pack(side="left",padx=2,pady=3)

    def _pdf_dividas_clientes(self):
        import datetime
        from tkinter import filedialog, messagebox as msgbox
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle, HRFlowable)
        except ImportError:
            msgbox.showerror("Erro", "Instale reportlab: pip install reportlab"); return

        dividas, total_divida = self._calcular_dividas()
        hoje_str = datetime.date.today().strftime("%d/%m/%Y")

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", initialfile="Dividas_Clientes.pdf",
            filetypes=[("PDF","*.pdf")], title="Salvar Dívidas PDF")
        if not path: return

        RED    = rl_colors.HexColor("#E74C3C")
        GREEN  = rl_colors.HexColor("#27AE60")
        BLUE   = rl_colors.HexColor("#2B7BE5")
        GRAY   = rl_colors.HexColor("#7F8C8D")

        doc = SimpleDocTemplate(path, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        def h(txt, c=RED, sz=13):
            return Paragraph(f'<font color="{c.hexval()}" size="{sz}"><b>{txt}</b></font>', styles["Normal"])
        def b(txt, sz=9):
            return Paragraph(f'<font size="{sz}">{txt}</font>', styles["Normal"])

        story = []
        story.append(h("DÍVIDAS DOS CLIENTES — Financiamentos"))
        story.append(b(f"Gerado em {hoje_str}"))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=2, color=RED))
        story.append(Spacer(1, 0.3*cm))

        # KPI
        em_aberto = sum(1 for d in dividas if not d["quitado"])
        kd = Table([[b("<b>TOTAL GERAL DE DÍVIDAS EM ABERTO</b>", 11),
                     h(f"¥ {total_divida:,}", RED, 14),
                     b(f"{em_aberto} em aberto / {len(dividas)} total")]],
                   colWidths=[7.5*cm, 5.5*cm, 4*cm])
        kd.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#FFF0F0")),
            ("BOX",(0,0),(-1,-1), 2, RED),
            ("LEFTPADDING",(0,0),(-1,-1), 10),
            ("TOPPADDING",(0,0),(-1,-1), 10),
            ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(kd)
        story.append(Spacer(1, 0.4*cm))

        # Tabela
        rows = [[b("<b>#</b>"), b("<b>Cliente</b>"), b("<b>Carro</b>"),
                 b("<b>V.Total</b>"), b("<b>Já Pago</b>"), b("<b>Saldo Dev.</b>"), b("<b>Status</b>")]]
        for d in dividas:
            st = "Quitado" if d["quitado"] else "Em Aberto"
            rows.append([
                b(str(d["id"])),
                b(d["cliente"][:22]),
                b(d["carro"][:22]),
                b(f"¥ {d['vv']:,}"),
                b(f"¥ {d['total_pago']:,}"),
                b(f"¥ {d['saldo']:,}" if d["saldo"] > 0 else "—"),
                b(st),
            ])
        t = Table(rows, colWidths=[1*cm, 4.5*cm, 4.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 2.2*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), RED),
            ("TEXTCOLOR",(0,0),(-1,0), rl_colors.white),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [rl_colors.HexColor("#FFF8F8"), rl_colors.white]),
            ("GRID",(0,0),(-1,-1), 0.3, rl_colors.HexColor("#E0E0E0")),
            ("FONTSIZE",(0,0),(-1,-1), 8),
            ("LEFTPADDING",(0,0),(-1,-1), 4),
            ("TOPPADDING",(0,0),(-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
        story.append(b(f"<font color='grey'>KM Cars — Gerado em {hoje_str}</font>"))
        doc.build(story)
        try:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass
        msgbox.showinfo("PDF Gerado", f"Salvo em:\n{path}")

    # ══════════════════════════════════════════════════════════════════════════
    # MÓDULO: CONFIGURAÇÕES
    # ══════════════════════════════════════════════════════════════════════════
    def _build_configuracoes_subs(self, parent):
        self.sub_pages["Configurações"] = {}
        for sub, fn in [
            ("Geral",          self._build_config_geral),
            ("Interface",      self._build_config_interface),
            ("IS",  self._build_config_is),
            ("Banco de Dados", self._build_config_banco),
        ]:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Configurações"][sub] = frame
            fn(frame)
        list(self.sub_pages["Configurações"].values())[0].place(
            in_=parent, x=0, y=0, relwidth=1, relheight=1)

    def _config_header(self, parent, titulo, subtitulo, icon="⚙"):
        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=24, pady=20)
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 20))
        tk.Label(hdr, text=f"{icon}  {titulo}", font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Label(hdr, text=subtitulo, font=("Helvetica", 9),
                 bg=COLORS["bg_content"], fg=COLORS["text_muted"]).pack(side="left", padx=12)
        return root

    def _config_card(self, parent, titulo, cor=None):
        cor = cor or COLORS["accent"]
        card = tk.Frame(parent, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="x", pady=(0, 14))
        tk.Frame(card, bg=cor, height=3).pack(fill="x")
        if titulo:
            tk.Label(card, text=titulo, font=("Helvetica", 10, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(
                         anchor="w", padx=16, pady=(12, 4))
            tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(0, 10))
        return card

    def _build_config_geral(self, parent):
        root = self._config_header(parent, "Configurações Gerais",
                                   "Nome da empresa, versão e preferências gerais", "⚙")
        card = self._config_card(root, "Informações da Empresa")
        cfg = self._cfg_load()

        def row(card, label, var, width=30):
            f = tk.Frame(card, bg=COLORS["bg_card"])
            f.pack(fill="x", padx=16, pady=(0, 10))
            tk.Label(f, text=label, font=("Helvetica", 8, "bold"), width=18, anchor="w",
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
            e = tk.Entry(f, textvariable=var, font=("Helvetica", 9),
                         bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                         insertbackground=COLORS["text_primary"],
                         relief="flat", highlightthickness=1,
                         highlightbackground=COLORS["border"], width=width)
            e.pack(side="left", ipady=5)
            return e

        self._cfg_empresa    = tk.StringVar(value=cfg.get("empresa", "KM Cars"))
        self._cfg_cidade_emp = tk.StringVar(value=cfg.get("cidade_emp", "Fujinomiya · Japan"))
        self._cfg_versao     = tk.StringVar(value=cfg.get("versao", "1.0.0"))
        row(card, "Nome da Empresa",   self._cfg_empresa)
        row(card, "Cidade / País",     self._cfg_cidade_emp)
        row(card, "Versão do Sistema", self._cfg_versao, width=10)

        card2 = self._config_card(root, "Backup Automático", COLORS["green"])
        self._cfg_backup_auto = tk.BooleanVar(value=cfg.get("backup_auto", True))
        self._cfg_backup_dias = tk.IntVar(value=cfg.get("backup_dias", 7))
        f2 = tk.Frame(card2, bg=COLORS["bg_card"])
        f2.pack(fill="x", padx=16, pady=(0, 10))
        tk.Label(f2, text="Backup automático ativo", font=("Helvetica", 9),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Checkbutton(f2, variable=self._cfg_backup_auto,
                       bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                       activebackground=COLORS["bg_card"],
                       selectcolor=COLORS["bg_main"]).pack(side="left", padx=8)
        f3 = tk.Frame(card2, bg=COLORS["bg_card"])
        f3.pack(fill="x", padx=16, pady=(0, 10))
        tk.Label(f3, text="Frequência (dias)", font=("Helvetica", 8, "bold"), width=18, anchor="w",
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
        tk.Spinbox(f3, from_=1, to=30, textvariable=self._cfg_backup_dias,
                   width=5, font=("Helvetica", 9),
                   bg=COLORS["bg_main"], fg=COLORS["text_primary"]).pack(side="left", ipady=4)

        btn_f = tk.Frame(root, bg=COLORS["bg_content"])
        btn_f.pack(fill="x")
        self._cfg_geral_status = tk.Label(btn_f, text="", font=("Helvetica", 8),
                                           bg=COLORS["bg_content"], fg=COLORS["green"])
        self._cfg_geral_status.pack(side="left", padx=(0, 12))
        def _salvar_geral():
            c = self._cfg_load()
            c.update({"empresa": self._cfg_empresa.get().strip(),
                       "cidade_emp": self._cfg_cidade_emp.get().strip(),
                       "versao": self._cfg_versao.get().strip(),
                       "backup_auto": self._cfg_backup_auto.get(),
                       "backup_dias": self._cfg_backup_dias.get()})
            self._cfg_save(c)
            self._cfg_geral_status.configure(text="✔ Salvo!")
            self.after(2000, lambda: self._cfg_geral_status.configure(text=""))
        tk.Button(btn_f, text="  Salvar Configurações  ", font=("Helvetica", 9, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=_salvar_geral).pack(side="right", ipady=6, ipadx=4)

    def _build_config_interface(self, parent):
        cfg = self._cfg_load()
        self._cfg_escala = tk.DoubleVar(value=cfg.get("escala", 1.0))
        self._cfg_janela = tk.StringVar(value=cfg.get("janela", "maximized"))
        self._cfg_fonte  = tk.IntVar(value=cfg.get("fonte_base", 9))

        # ── Container externo (não scrola) ────────────────────────────────────
        outer = tk.Frame(parent, bg=COLORS["bg_content"])
        outer.pack(fill="both", expand=True, padx=24, pady=(16, 0))

        # ── Cabeçalho fixo ────────────────────────────────────────────────────
        hdr = tk.Frame(outer, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 6))
        tk.Label(hdr, text="🖥  Interface & Resolução", font=("Helvetica", 13, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")

        # ── Botão Salvar FIXO no topo (sempre visível) ────────────────────────
        self._cfg_ui_status = tk.Label(hdr, text="", font=("Helvetica", 8),
                                        bg=COLORS["bg_content"], fg=COLORS["green"])
        self._cfg_ui_status.pack(side="right", padx=(0, 10))

        def _salvar_ui():
            try:
                escala = float(self._cfg_escala.get())
                escala = max(0.7, min(1.5, escala))
                self._cfg_escala.set(round(escala, 2))
            except Exception:
                escala = 1.0
            c = self._cfg_load()
            c.update({"escala": escala, "janela": self._cfg_janela.get(),
                       "fonte_base": self._cfg_fonte.get()})
            self._cfg_save(c)
            j = self._cfg_janela.get()
            if j == "maximized":
                self.state("zoomed")
            elif "x" in j:
                w2, h2 = j.split("x")
                self.geometry(f"{w2}x{h2}")
                self.state("normal")
            self._cfg_ui_status.configure(text="✔ Salvo! Reinicie para aplicar escala.")
            self.after(4000, lambda: self._cfg_ui_status.configure(text=""))

        tk.Button(hdr, text="  💾 Salvar & Aplicar  ", font=("Helvetica", 9, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=_salvar_ui).pack(side="right", ipady=5, ipadx=4)

        tk.Label(outer,
                 text="⚠  Salve e REINICIE para aplicar escala. Janela é aplicada imediatamente.",
                 font=("Helvetica", 8), bg=COLORS["bg_content"],
                 fg=COLORS["orange"]).pack(anchor="w", pady=(0, 8))

        tk.Frame(outer, bg=COLORS["border"], height=1).pack(fill="x", pady=(0, 10))

        # ── Área scrollável ───────────────────────────────────────────────────
        scroll_outer = tk.Frame(outer, bg=COLORS["bg_content"])
        scroll_outer.pack(fill="both", expand=True)

        vsb = tk.Scrollbar(scroll_outer, orient="vertical")
        vsb.pack(side="right", fill="y")
        canvas = tk.Canvas(scroll_outer, bg=COLORS["bg_content"],
                           highlightthickness=0, yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.configure(command=canvas.yview)

        inner = tk.Frame(canvas, bg=COLORS["bg_content"])
        cw = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_resize(e):
            canvas.itemconfig(cw, width=e.width)
        inner.bind("<Configure>", _on_cfg)
        canvas.bind("<Configure>", _on_resize)
        canvas.bind("<Enter>",
            lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(int(-1*(ev.delta/120)), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # helper para cards dentro do inner scrollável
        def cfg_card(titulo, cor):
            card = tk.Frame(inner, bg=COLORS["bg_card"],
                            highlightthickness=1, highlightbackground=COLORS["border"])
            card.pack(fill="x", pady=(0, 12), padx=2)
            tk.Frame(card, bg=cor, height=3).pack(fill="x")
            tk.Label(card, text=titulo, font=("Helvetica", 10, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(
                         anchor="w", padx=16, pady=(10, 4))
            tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(0, 8))
            return card

        # ── Card 1: Escala ────────────────────────────────────────────────────
        card = cfg_card("Escala & Zoom", COLORS["blue"])
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        tk.Label(card, text=f"Resolução detectada: {sw} × {sh} px",
                 font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(anchor="w", padx=16, pady=(0, 8))

        presets = [
            ("Muito pequeno  — 1366×768 / Windows 100%",     0.75),
            ("Pequeno        — 1536×864 / Windows 125%",     0.85),
            ("Médio          — 1920×1080 notebook Win 125%", 0.92),
            ("Padrão         — 1920×1080 desktop 24\"",       1.00),
            ("Grande         — monitor 27\"+ / 2K",           1.10),
            ("Extra grande   — 4K / TV",                      1.20),
        ]
        tk.Label(card, text="Selecione o preset:",
                 font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(
                     anchor="w", padx=16, pady=(0, 4))
        for label, val in presets:
            fb = tk.Frame(card, bg=COLORS["bg_card"])
            fb.pack(fill="x", padx=16, pady=1)
            tk.Radiobutton(fb, text=label, variable=self._cfg_escala, value=val,
                           font=("Helvetica", 9),
                           bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                           activebackground=COLORS["bg_card"],
                           selectcolor=COLORS["bg_main"]).pack(side="left")
            if val == 0.85:
                tk.Label(fb, text="← recomendado para 1536×864",
                         font=("Helvetica", 8), bg=COLORS["bg_card"],
                         fg=COLORS["orange"]).pack(side="left", padx=6)
            elif val == 1.0:
                tk.Label(fb, text="← padrão desktop 24\"",
                         font=("Helvetica", 8), bg=COLORS["bg_card"],
                         fg=COLORS["text_muted"]).pack(side="left", padx=6)

        fc = tk.Frame(card, bg=COLORS["bg_card"])
        fc.pack(fill="x", padx=16, pady=(6, 10))
        tk.Label(fc, text="Valor personalizado (0.7–1.5):",
                 font=("Helvetica", 8), bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(side="left")
        tk.Entry(fc, textvariable=self._cfg_escala, width=6,
                 font=("Helvetica", 9), bg=COLORS["bg_main"],
                 fg=COLORS["text_primary"], relief="flat",
                 highlightthickness=1, highlightbackground=COLORS["border"]
                 ).pack(side="left", padx=8, ipady=4)

        # ── Card 2: Janela ────────────────────────────────────────────────────
        card2 = cfg_card("Tamanho da Janela ao Iniciar", COLORS["orange"])
        for label, val in [
            ("Maximizado (recomendado)", "maximized"),
            ("1280 × 800",               "1280x800"),
            ("1440 × 900",               "1440x900"),
            ("1600 × 900",               "1600x900"),
            ("1920 × 1080",              "1920x1080"),
        ]:
            tk.Radiobutton(card2, text=label, variable=self._cfg_janela, value=val,
                           font=("Helvetica", 9),
                           bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                           activebackground=COLORS["bg_card"],
                           selectcolor=COLORS["bg_main"]).pack(anchor="w", padx=16, pady=2)
        tk.Frame(card2, bg=COLORS["bg_card"], height=6).pack()

        # ── Card 3: Fonte ─────────────────────────────────────────────────────
        card3 = cfg_card("Tamanho de Fonte Base", COLORS["green"])
        ff3 = tk.Frame(card3, bg=COLORS["bg_card"])
        ff3.pack(fill="x", padx=16, pady=(0, 10))
        tk.Label(ff3, text="Tamanho base (px):", font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
        tk.Spinbox(ff3, from_=7, to=14, textvariable=self._cfg_fonte,
                   width=4, font=("Helvetica", 9),
                   bg=COLORS["bg_main"], fg=COLORS["text_primary"]).pack(
                       side="left", padx=8, ipady=4)
        tk.Label(ff3, text="(reinicie para aplicar)", font=("Helvetica", 8),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="left")

    def _build_config_is(self, parent):
        """Sub-tela para configurar percentual IS."""
        outer = tk.Frame(parent, bg=COLORS["bg_content"])
        outer.pack(fill="both", expand=True, padx=24, pady=(16, 0))
        hdr = tk.Frame(outer, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 8))
        tk.Label(hdr, text="💴  IS — Imposto (CNF)",
                 font=("Helvetica", 13, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")

        card = tk.Frame(outer, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"])
        card.pack(fill="x", pady=(0, 14))
        tk.Frame(card, bg=COLORS["blue"], height=3).pack(fill="x")
        tk.Label(card, text="Percentual IS",
                 font=("Helvetica", 10, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(
                     anchor="w", padx=16, pady=(10, 2))
        tk.Frame(card, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(0, 10))

        desc = tk.Label(card,
            text="Aplicado automaticamente como custo 'IS' ao registrar Vendas, OS e Shaken com NFI = CNF.",
            font=("Helvetica", 8), bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
            wraplength=500, justify="left")
        desc.pack(anchor="w", padx=16, pady=(0, 10))

        try:
            cur_perc = float(self.conn.execute(
                "SELECT valor FROM configuracoes WHERE chave='is_percentual'"
                ).fetchone()[0] or 10)
        except Exception:
            cur_perc = 10.0

        ff = tk.Frame(card, bg=COLORS["bg_card"])
        ff.pack(fill="x", padx=16, pady=(0, 14))
        tk.Label(ff, text="Percentual atual (%):",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
        self._is_perc_var = tk.DoubleVar(value=cur_perc)
        sp = tk.Spinbox(ff, from_=0.0, to=99.9, increment=0.5,
                        textvariable=self._is_perc_var,
                        width=7, font=("Helvetica", 10),
                        bg=COLORS["bg_main"], fg=COLORS["text_primary"])
        sp.pack(side="left", padx=(8, 10), ipady=4)
        tk.Label(ff, text=f"(padrão: 10%)",
                 font=("Helvetica", 8), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(side="left")

        st_lbl = tk.Label(outer, text="", font=("Helvetica", 9),
                          bg=COLORS["bg_content"], fg=COLORS["green"])
        st_lbl.pack(anchor="w", pady=(0, 8))

        def _salvar_is():
            try:
                p = float(self._is_perc_var.get())
                p = max(0.0, min(99.9, round(p, 1)))
                self._is_perc_var.set(p)
            except Exception:
                p = 10.0
            self.conn.execute(
                "INSERT OR REPLACE INTO configuracoes (chave,valor) VALUES ('is_percentual',?)",
                (str(p),))
            self.conn.commit()
            st_lbl.configure(text=f"✔ IS salvo: {p}%")
            outer.after(3000, lambda: st_lbl.configure(text=""))

        tk.Button(outer, text="  💾 Salvar Percentual IS  ",
                  font=("Helvetica", 10, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=_salvar_is).pack(anchor="w", ipady=6, ipadx=6)

    def _build_config_banco(self, parent):
        import os, datetime
        root = self._config_header(parent, "Banco de Dados",
                                   "Backup manual, localizacao e manutencao", "🗄")
        card = self._config_card(root, "Localizacao do Banco", COLORS["blue"])
        db_path = getattr(self, "_db_path", "—")
        try:
            db_size_str = f"{os.path.getsize(db_path) / 1024:.1f} KB"
        except Exception:
            db_size_str = "—"
        for label, val in [("Arquivo:", db_path), ("Tamanho:", db_size_str)]:
            fr = tk.Frame(card, bg=COLORS["bg_card"])
            fr.pack(fill="x", padx=16, pady=(0, 6))
            tk.Label(fr, text=label, font=("Helvetica", 8, "bold"), width=10, anchor="w",
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(side="left")
            tk.Label(fr, text=val, font=("Helvetica", 8),
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                     wraplength=450, justify="left").pack(side="left")
        tk.Frame(card, bg=COLORS["bg_card"], height=6).pack()

        card2 = self._config_card(root, "Backup Manual", COLORS["green"])
        self._cfg_bk_status = tk.Label(card2, text="", font=("Helvetica", 8),
                                        bg=COLORS["bg_card"], fg=COLORS["green"])
        self._cfg_bk_status.pack(anchor="w", padx=16, pady=(0, 6))
        def _fazer_backup():
            try:
                import shutil
                db = getattr(self, "_db_path", None)
                if not db or not os.path.exists(db):
                    self._cfg_bk_status.configure(
                        text="Banco nao encontrado.", fg=COLORS["red"]); return
                bk_dir = os.path.join(os.path.dirname(db), "backups")
                os.makedirs(bk_dir, exist_ok=True)
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = os.path.join(bk_dir, f"kmcars_backup_{ts}.db")
                shutil.copy2(db, dest)
                self._cfg_bk_status.configure(
                    text=f"Backup salvo: {dest}", fg=COLORS["green"])
            except Exception as e:
                self._cfg_bk_status.configure(text=f"Erro: {e}", fg=COLORS["red"])
        tk.Button(card2, text="  Fazer Backup Agora  ", font=("Helvetica", 9, "bold"),
                  bg=COLORS["green"], fg="white", relief="flat", cursor="hand2",
                  command=_fazer_backup).pack(anchor="w", padx=16, pady=(0, 10),
                                             ipadx=4, ipady=6)

        card3 = self._config_card(root, "Manutencao do Banco", COLORS["orange"])
        tk.Label(card3,
                 text="VACUUM compacta o banco removendo espaco nao utilizado. Recomendado mensalmente.",
                 font=("Helvetica", 8), bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"]).pack(anchor="w", padx=16, pady=(0, 8))
        self._cfg_vac_status = tk.Label(card3, text="", font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["green"])
        self._cfg_vac_status.pack(anchor="w", padx=16)
        def _vacuum():
            try:
                self.conn.execute("VACUUM"); self.conn.commit()
                try:
                    sz = os.path.getsize(getattr(self, "_db_path", ""))
                    self._cfg_vac_status.configure(
                        text=f"VACUUM concluido. Tamanho: {sz/1024:.1f} KB", fg=COLORS["green"])
                except Exception:
                    self._cfg_vac_status.configure(text="VACUUM concluido.", fg=COLORS["green"])
            except Exception as e:
                self._cfg_vac_status.configure(text=f"Erro: {e}", fg=COLORS["red"])
        tk.Button(card3, text="  Executar VACUUM  ", font=("Helvetica", 9),
                  bg=COLORS["orange"], fg="white", relief="flat", cursor="hand2",
                  command=_vacuum).pack(anchor="w", padx=16, pady=(4, 10), ipadx=4, ipady=5)

    # ══════════════════════════════════════════════════════════════════════════
    #  EXPORTAR PDF / EXCEL — Vendas, OS, Shaken
    # ══════════════════════════════════════════════════════════════════════════

    def _get_is_perc(self):
        """Retorna percentual IS configurado (padrão 10)."""
        try:
            return float(self.conn.execute(
                "SELECT valor FROM configuracoes WHERE chave='is_percentual'"
            ).fetchone()[0])
        except Exception:
            return 10.0

    # ── helpers comuns ────────────────────────────────────────────────────────
    def _rl_header(self, story, titulo, filtros_txt, styles, rl_colors, Paragraph, Spacer, HRFlowable):
        """Gera cabeçalho padrão para PDFs de relatório."""
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.styles import ParagraphStyle
        h1 = ParagraphStyle("h1", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold",
                             alignment=TA_CENTER, textColor=rl_colors.HexColor("#1a1a2e"), spaceAfter=4)
        h2 = ParagraphStyle("h2", parent=styles["Normal"], fontSize=8,
                             alignment=TA_CENTER, textColor=rl_colors.grey, spaceAfter=8)
        story.append(Paragraph("KM Cars", h1))
        story.append(Paragraph(titulo, h2))
        if filtros_txt:
            story.append(Paragraph(filtros_txt, h2))
        story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#4361ee")))
        story.append(Spacer(1, 10))

    def _rl_table(self, data, col_widths, rl_colors, Table, TableStyle):
        """Monta Table estilizada com cabeçalho colorido."""
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), rl_colors.HexColor("#4361ee")),
            ("TEXTCOLOR",   (0,0), (-1,0), rl_colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,0), 8),
            ("FONTSIZE",    (0,1), (-1,-1), 7),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [rl_colors.HexColor("#f8f9ff"), rl_colors.white]),
            ("GRID",        (0,0), (-1,-1), 0.3, rl_colors.HexColor("#cccccc")),
            ("ALIGN",       (0,0), (-1,-1), "CENTER"),
            ("ALIGN",       (0,1), (1,-1), "LEFT"),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ]))
        return t

    def _save_pdf_dialog(self, nome_base):
        from tkinter import filedialog
        import datetime
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"{nome_base}_{ts}.pdf",
            title="Salvar PDF")
        return path or None

    def _save_excel_dialog(self, nome_base):
        from tkinter import filedialog
        import datetime
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"{nome_base}_{ts}.xlsx",
            title="Salvar Excel")
        return path or None

    def _open_file(self, path):
        import subprocess, sys as _sys, os
        try:
            if _sys.platform.startswith("win"):
                os.startfile(path)
            elif _sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass

    # ── VENDAS ────────────────────────────────────────────────────────────────
    def _exportar_vendas_pdf(self):
        from tkinter import messagebox as _mb
        lista = getattr(self, "_vendas_lista_filtrada", None)
        if not lista:
            _mb.showwarning("PDF", "Nenhuma venda para exportar (aplique os filtros primeiro).")
            return
        path = self._save_pdf_dialog("historico_vendas")
        if not path: return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Spacer, HRFlowable, Paragraph)
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.units import cm
        except ImportError:
            _mb.showerror("PDF", "Instale reportlab: pip install reportlab"); return
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        # filtros usados
        nfi_f  = getattr(self, "_venda_filtro_nfi",  None)
        mes_f  = getattr(self, "_venda_filtro_mes",  None)
        ano_f  = getattr(self, "_venda_filtro_ano",  None)
        tipo_f = getattr(self, "_venda_filtro_tipo", None)
        filtros_txt = f"NFI: {nfi_f.get() if nfi_f else 'Todos'} | " \
                      f"Mês: {mes_f.get() if mes_f else 'Todos'} | " \
                      f"Ano: {ano_f.get() if ano_f else 'Todos'} | " \
                      f"Tipo: {tipo_f.get() if tipo_f else 'Todos'} | " \
                      f"{len(lista)} registro(s)"
        self._rl_header(story, "Histórico de Vendas", filtros_txt, styles, rl_colors,
                        Paragraph, Spacer, HRFlowable)
        is_perc = self._get_is_perc()
        headers = ["ID","Data","Carro","Cliente","Tipo","V.Venda","V.Custo","V.Troca","IS","Lucro","NFI"]
        cw = [1.0*cm,2.0*cm,4.5*cm,3.5*cm,2.2*cm,2.2*cm,2.2*cm,2.2*cm,1.8*cm,2.2*cm,1.5*cm]
        rows = [headers]
        for v in lista:
            nfi  = v.get("nfi") or "SNF"
            vv   = v.get("valor_venda") or 0
            vc   = v.get("valor_custo") or v.get("custo_total") or 0
            vt   = v.get("valor_troca") or 0
            is_v = int(vv * is_perc/100) if nfi == "CNF" else 0
            lucro= int(vv) - int(vc) - int(vt) - is_v
            rows.append([
                str(v.get("id","")),
                v.get("data_venda") or "",
                (v.get("carro") or "")[:30],
                (v.get("cliente_nome") or "")[:22],
                v.get("tipo_venda") or "",
                f"¥{int(vv):,}",
                f"¥{int(vc):,}",
                f"¥{int(vt):,}",
                f"¥{is_v:,}" if is_v else "—",
                f"¥{lucro:,}",
                nfi,
            ])
        story.append(self._rl_table(rows, cw, rl_colors, Table, TableStyle))
        doc.build(story)
        _mb.showinfo("PDF", f"PDF gerado com sucesso!"); self._open_file(path)

    def _exportar_vendas_excel(self):
        from tkinter import messagebox as _mb
        lista = getattr(self, "_vendas_lista_filtrada", None)
        if not lista:
            _mb.showwarning("Excel", "Nenhuma venda para exportar."); return
        path = self._save_excel_dialog("historico_vendas")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            _mb.showerror("Excel", "Instale openpyxl: pip install openpyxl"); return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendas"
        is_perc = self._get_is_perc()
        hdr_fill  = PatternFill("solid", fgColor="4361EE")
        hdr_font  = Font(bold=True, color="FFFFFF", size=9)
        bd        = Border(left=Side(style="thin"), right=Side(style="thin"),
                           top=Side(style="thin"), bottom=Side(style="thin"))
        headers = ["ID","Data","Carro","Cliente","Tipo","V.Venda","V.Custo",
                   "V.Troca","IS (10%)","Lucro","NFI"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bd
        col_widths = [6,12,30,22,14,14,14,14,14,14,7]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        alt_fill = PatternFill("solid", fgColor="F0F4FF")
        for ri, v in enumerate(lista, 2):
            nfi  = v.get("nfi") or "SNF"
            vv   = v.get("valor_venda") or 0
            vc   = v.get("valor_custo") or v.get("custo_total") or 0
            vt   = v.get("valor_troca") or 0
            is_v = int(vv * is_perc/100) if nfi == "CNF" else 0
            lucro= int(vv) - int(vc) - int(vt) - is_v
            row_data = [v.get("id"), v.get("data_venda",""), v.get("carro",""),
                        v.get("cliente_nome",""), v.get("tipo_venda",""),
                        int(vv), int(vc), int(vt),
                        is_v, lucro, nfi]
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = bd
                c.alignment = Alignment(horizontal="center" if ci not in (3,4) else "left",
                                        vertical="center")
                if fill: c.fill = fill
                if ci == 11 and nfi == "CNF":
                    c.fill = PatternFill("solid", fgColor="FFE5B4")
        ws.freeze_panes = "A2"
        wb.save(path)
        _mb.showinfo("Excel", "Excel gerado com sucesso!"); self._open_file(path)

    # ── SHAKEN ────────────────────────────────────────────────────────────────
    def _exportar_shaken_pdf(self):
        from tkinter import messagebox as _mb
        lista = getattr(self, "_shaken_lista_filtrada", None)
        if not lista:
            _mb.showwarning("PDF", "Nenhum shaken para exportar."); return
        path = self._save_pdf_dialog("historico_shaken")
        if not path: return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Spacer, HRFlowable, Paragraph)
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.units import cm
        except ImportError:
            _mb.showerror("PDF", "Instale reportlab: pip install reportlab"); return
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        nfi_f  = getattr(self, "_sk_filtro_nfi",   None)
        tipo_f = getattr(self, "_sk_filtro_tipo",   None)
        filtros_txt = f"NFI: {nfi_f.get() if nfi_f else 'Todos'} | " \
                      f"Tipo: {tipo_f.get() if tipo_f else 'Todos'} | " \
                      f"{len(lista)} registro(s)"
        self._rl_header(story, "Histórico Shaken", filtros_txt, styles, rl_colors,
                        Paragraph, Spacer, HRFlowable)
        is_perc = self._get_is_perc()
        headers = ["SK","Carro","Placa","Cliente","Custo SK","Valor","Data SK","Status","NFI","IS"]
        cw = [1.2*cm,4.5*cm,2.0*cm,3.5*cm,2.2*cm,2.2*cm,2.5*cm,2.5*cm,1.5*cm,2.0*cm]
        rows = [headers]
        for r in lista:
            nfi  = r.get("nfi") or "SNF"
            vv   = r.get("valor") or 0
            is_v = int(float(vv) * is_perc/100) if nfi == "CNF" else 0
            rows.append([
                r.get("sk_num") or str(r.get("id","")),
                (r.get("c_nome") or r.get("carro",""))[:30],
                r.get("c_placa") or r.get("placa",""),
                (r.get("cli_nome") or "")[:22],
                f"¥{int(r.get('custo_sk') or 0):,}",
                f"¥{int(float(vv)):,}",
                r.get("data_shaken") or "",
                r.get("status_label") or r.get("status",""),
                nfi,
                f"¥{is_v:,}" if is_v else "—",
            ])
        story.append(self._rl_table(rows, cw, rl_colors, Table, TableStyle))
        doc.build(story)
        _mb.showinfo("PDF", "PDF gerado com sucesso!"); self._open_file(path)

    def _exportar_shaken_excel(self):
        from tkinter import messagebox as _mb
        lista = getattr(self, "_shaken_lista_filtrada", None)
        if not lista:
            _mb.showwarning("Excel", "Nenhum shaken para exportar."); return
        path = self._save_excel_dialog("historico_shaken")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            _mb.showerror("Excel", "Instale openpyxl: pip install openpyxl"); return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Shaken"
        is_perc = self._get_is_perc()
        hdr_fill = PatternFill("solid", fgColor="4361EE")
        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        bd = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        headers = ["SK","Carro","Placa","Cliente","Custo SK","Valor","Data SK","Status","NFI","IS"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center"); c.border = bd
        for ci, w in enumerate([8,28,12,22,12,12,14,14,7,12], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        alt = PatternFill("solid", fgColor="F0F4FF")
        for ri, r in enumerate(lista, 2):
            nfi  = r.get("nfi") or "SNF"
            vv   = float(r.get("valor") or 0)
            is_v = int(vv * is_perc/100) if nfi == "CNF" else 0
            row_data = [
                r.get("sk_num") or str(r.get("id","")),
                r.get("c_nome") or r.get("carro",""),
                r.get("c_placa") or r.get("placa",""),
                r.get("cli_nome",""),
                int(r.get("custo_sk") or 0),
                int(vv), r.get("data_shaken",""),
                r.get("status_label") or r.get("status",""),
                nfi, is_v,
            ]
            fill = alt if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = bd
                c.alignment = Alignment(horizontal="left" if ci in (2,3,4) else "center",
                                        vertical="center")
                if fill: c.fill = fill
                if ci == 9 and nfi == "CNF":
                    c.fill = PatternFill("solid", fgColor="FFE5B4")
        ws.freeze_panes = "A2"
        wb.save(path)
        _mb.showinfo("Excel", "Excel gerado com sucesso!"); self._open_file(path)

    # ── OS ────────────────────────────────────────────────────────────────────
    def _exportar_os_pdf(self):
        from tkinter import messagebox as _mb
        lista = getattr(self, "_servicos_lista_filtrada", None)
        if not lista:
            _mb.showwarning("PDF", "Nenhuma OS para exportar."); return
        path = self._save_pdf_dialog("historico_os")
        if not path: return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Spacer, HRFlowable, Paragraph)
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.units import cm
        except ImportError:
            _mb.showerror("PDF", "Instale reportlab: pip install reportlab"); return
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        nfi_f    = getattr(self, "_serv_filtro_nfi",    None)
        mes_f    = getattr(self, "_serv_filtro_mes",    None)
        ano_f    = getattr(self, "_serv_filtro_ano",    None)
        status_f = getattr(self, "_serv_filtro_status", None)
        filtros_txt = f"NFI: {nfi_f.get() if nfi_f else 'Todos'} | " \
                      f"Mês: {mes_f.get() if mes_f else 'Todos'} | " \
                      f"Ano: {ano_f.get() if ano_f else 'Todos'} | " \
                      f"Status: {status_f.get() if status_f else 'Todos'} | " \
                      f"{len(lista)} registro(s)"
        self._rl_header(story, "Histórico de Ordens de Serviço", filtros_txt, styles, rl_colors,
                        Paragraph, Spacer, HRFlowable)
        is_perc = self._get_is_perc()
        headers = ["OS","Data","Carro","Cliente","Tipo","Valor","IS","Status","NFI"]
        cw = [1.8*cm,2.2*cm,4.5*cm,3.5*cm,3.5*cm,2.5*cm,2.0*cm,2.5*cm,1.5*cm]
        rows = [headers]
        for s in lista:
            nfi  = s.get("nfi") or "SNF"
            vv   = float(s.get("valor") or 0)
            is_v = int(vv * is_perc/100) if nfi == "CNF" else 0
            # busca nome cliente
            try:
                cli_r = self.conn.execute(
                    "SELECT nome FROM clientes WHERE id=?", (s.get("cliente_id"),)).fetchone()
                cli_nome = cli_r[0] if cli_r else ""
            except Exception:
                cli_nome = ""
            rows.append([
                s.get("os_num") or str(s.get("id","")),
                s.get("data_servico",""),
                (s.get("carro",""))[:30],
                cli_nome[:22],
                (s.get("tipo_servico",""))[:22],
                f"¥{int(vv):,}",
                f"¥{is_v:,}" if is_v else "—",
                s.get("status",""),
                nfi,
            ])
        story.append(self._rl_table(rows, cw, rl_colors, Table, TableStyle))
        doc.build(story)
        _mb.showinfo("PDF", "PDF gerado com sucesso!"); self._open_file(path)

    def _exportar_os_excel(self):
        from tkinter import messagebox as _mb
        lista = getattr(self, "_servicos_lista_filtrada", None)
        if not lista:
            _mb.showwarning("Excel", "Nenhuma OS para exportar."); return
        path = self._save_excel_dialog("historico_os")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            _mb.showerror("Excel", "Instale openpyxl: pip install openpyxl"); return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "OS"
        is_perc = self._get_is_perc()
        hdr_fill = PatternFill("solid", fgColor="4361EE")
        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        bd = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        headers = ["OS","Data","Carro","Cliente","Tipo Serviço","Valor","IS","Status","NFI"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center"); c.border = bd
        for ci, w in enumerate([10,12,28,22,24,14,12,14,7], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        alt = PatternFill("solid", fgColor="F0F4FF")
        for ri, s in enumerate(lista, 2):
            nfi  = s.get("nfi") or "SNF"
            vv   = float(s.get("valor") or 0)
            is_v = int(vv * is_perc/100) if nfi == "CNF" else 0
            try:
                cli_r = self.conn.execute(
                    "SELECT nome FROM clientes WHERE id=?", (s.get("cliente_id"),)).fetchone()
                cli_nome = cli_r[0] if cli_r else ""
            except Exception:
                cli_nome = ""
            row_data = [
                s.get("os_num") or str(s.get("id","")),
                s.get("data_servico",""), s.get("carro",""),
                cli_nome, s.get("tipo_servico",""),
                int(vv), is_v, s.get("status",""), nfi,
            ]
            fill = alt if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = bd
                c.alignment = Alignment(horizontal="left" if ci in (3,4,5) else "center",
                                        vertical="center")
                if fill: c.fill = fill
                if ci == 9 and nfi == "CNF":
                    c.fill = PatternFill("solid", fgColor="FFE5B4")
        ws.freeze_panes = "A2"
        wb.save(path)
        _mb.showinfo("Excel", "Excel gerado com sucesso!"); self._open_file(path)

    def _cfg_load(self):
        import json, os, sys as _sys
        base = os.path.dirname(_sys.executable) if getattr(_sys,"frozen",False) else os.path.dirname(os.path.abspath(__file__))
        try:
            with open(os.path.join(base, "kmcars_config.json"), encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def _cfg_save(self, cfg):
        import json, os, sys as _sys
        base = os.path.dirname(_sys.executable) if getattr(_sys,"frozen",False) else os.path.dirname(os.path.abspath(__file__))
        with open(os.path.join(base, "kmcars_config.json"), "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)

    def _cfg_apply_on_startup(self):
        cfg = self._cfg_load()
        j = cfg.get("janela", "maximized")
        if j == "maximized":
            self.after(100, lambda: self.state("zoomed"))
        elif "x" in j:
            w2, h2 = j.split("x")
            self.geometry(f"{w2}x{h2}")

    def _build_placeholder(self, parent, page):
        center = tk.Frame(parent, bg=COLORS["bg_content"])
        center.place(relx=0.5, rely=0.5, anchor="center")

        icon = NAV_ICONS.get(page, "◈")
        circle = tk.Frame(center, bg=COLORS["border"], width=80, height=80)
        circle.pack()
        circle.pack_propagate(False)
        tk.Label(circle, text=icon, font=("Helvetica", 32),
                 bg=COLORS["border"], fg=COLORS["text_muted"]).place(
                     relx=0.5, rely=0.5, anchor="center")

        tk.Label(center, text=page.upper(), font=("Helvetica", 18, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_secondary"]).pack(pady=(14, 4))
        tk.Label(center,
                 text="Este módulo está em desenvolvimento.\nSelecione uma opção no submenu acima.",
                 font=("Helvetica", 10), bg=COLORS["bg_content"],
                 fg=COLORS["text_muted"], justify="center").pack()
        tk.Frame(center, bg=COLORS["accent"], height=2, width=50).pack(pady=14)

        tk.Frame(center, bg=COLORS["accent"], height=2, width=50).pack(pady=14)

    # ── Busca Global ──────────────────────────────────────────────────────────
    def _global_search_execute(self):
        """Executa busca global e abre popup com resultados."""
        termo = self._global_search_var.get().strip()
        if not termo or termo == "Buscar cliente, carro, OS...":
            return
        tl = termo.lower()

        resultados = {"Clientes": [], "Carros": [], "OS": [], "Shaken": [], "Vendas": []}

        try:
            for r in self.conn.execute(
                "SELECT id, nome, telefone FROM clientes "
                "WHERE nome LIKE ? OR telefone LIKE ? LIMIT 10",
                (f"%{termo}%", f"%{termo}%")):
                resultados["Clientes"].append(f"#{r[0]}  {r[1]}  {r[2] or ''}")
        except Exception:
            pass

        try:
            for r in self.conn.execute(
                "SELECT id, carro, placa, status FROM carros "
                "WHERE carro LIKE ? OR placa LIKE ? OR chassi LIKE ? LIMIT 10",
                (f"%{termo}%", f"%{termo}%", f"%{termo}%")):
                resultados["Carros"].append(f"#{r[0]}  {r[1]}  {r[2] or ''}  [{r[3]}]")
        except Exception:
            pass

        try:
            for r in self.conn.execute(
                "SELECT s.id, s.os_num, s.carro, c.nome FROM servicos s "
                "LEFT JOIN clientes c ON s.cliente_id=c.id "
                "WHERE s.carro LIKE ? OR s.os_num LIKE ? OR c.nome LIKE ? LIMIT 10",
                (f"%{termo}%", f"%{termo}%", f"%{termo}%")):
                resultados["OS"].append(f"{r[1]}  {r[2]}  {r[3] or ''}")
        except Exception:
            pass

        try:
            for r in self.conn.execute(
                "SELECT sk.sk_num, c.carro, cl.nome FROM shaken sk "
                "LEFT JOIN carros c ON sk.carro_id=c.id "
                "LEFT JOIN clientes cl ON sk.cliente_id=cl.id "
                "WHERE c.carro LIKE ? OR cl.nome LIKE ? OR sk.sk_num LIKE ? LIMIT 10",
                (f"%{termo}%", f"%{termo}%", f"%{termo}%")):
                resultados["Shaken"].append(f"{r[0]}  {r[1] or ''}  {r[2] or ''}")
        except Exception:
            pass

        try:
            for r in self.conn.execute(
                "SELECT v.id, v.carro, cl.nome, v.data_venda FROM vendas v "
                "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
                "WHERE v.carro LIKE ? OR cl.nome LIKE ? OR v.placa LIKE ? LIMIT 10",
                (f"%{termo}%", f"%{termo}%", f"%{termo}%")):
                resultados["Vendas"].append(f"#{r[0]}  {r[1]}  {r[2] or ''}  {r[3] or ''}")
        except Exception:
            pass

        total = sum(len(v) for v in resultados.values())

        # ── Popup de resultados ──────────────────────────────────────────────
        dlg = tk.Toplevel(self)
        dlg.title(f"Busca: {termo}")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.lift()
        dlg.attributes("-topmost", True)
        dlg.after(100, lambda: dlg.attributes("-topmost", False))
        w, h = 580, 460
        x = self.winfo_x() + (self.winfo_width()  - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

        tk.Frame(dlg, bg=COLORS["accent"], height=4).pack(fill="x")
        hdr_d = tk.Frame(dlg, bg=COLORS["bg_card"])
        hdr_d.pack(fill="x", padx=18, pady=(12,6))
        tk.Label(hdr_d, text=f'🔍  Resultados para "{termo}"',
                 font=("Helvetica",11,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Label(hdr_d, text=f"{total} encontrado(s)",
                 font=("Helvetica",8),
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(side="left", padx=10)
        tk.Button(hdr_d, text="✕ Fechar", font=("Helvetica",8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(side="right")

        tk.Frame(dlg, bg=COLORS["border"], height=1).pack(fill="x", padx=18)

        # Scroll area
        outer = tk.Frame(dlg, bg=COLORS["bg_card"])
        outer.pack(fill="both", expand=True, padx=18, pady=10)
        vsb = tk.Scrollbar(outer, orient="vertical")
        cv  = tk.Canvas(outer, bg=COLORS["bg_card"], highlightthickness=0,
                        yscrollcommand=vsb.set)
        vsb.configure(command=cv.yview)
        vsb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(cv, bg=COLORS["bg_card"])
        cw = cv.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: cv.configure(
            scrollregion=cv.bbox("all")))
        cv.bind("<Configure>", lambda e: cv.itemconfig(cw, width=e.width))
        cv.bind_all("<MouseWheel>",
                    lambda e: cv.yview_scroll(int(-1*(e.delta/120)),"units"))

        ICON = {"Clientes":"◈", "Carros":"▦", "OS":"⚙", "Shaken":"🔔", "Vendas":"◆"}
        COL  = {"Clientes":COLORS["accent"],"Carros":COLORS["blue"],
                "OS":COLORS["orange"],"Shaken":"#D4AC0D","Vendas":COLORS["green"]}

        if total == 0:
            tk.Label(inner, text="Nenhum resultado encontrado.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
        else:
            for cat, items in resultados.items():
                if not items: continue
                sec = tk.Frame(inner, bg=COLORS["bg_card"])
                sec.pack(fill="x", pady=(8,2))
                tk.Frame(sec, bg=COL[cat], width=3).pack(side="left", fill="y")
                tk.Label(sec, text=f" {ICON[cat]}  {cat}  ({len(items)})",
                         font=("Helvetica",9,"bold"),
                         bg=COLORS["bg_card"], fg=COL[cat]).pack(side="left", padx=6)
                for item in items:
                    row_f = tk.Frame(inner, bg=COLORS["bg_content"])
                    row_f.pack(fill="x", pady=1)
                    tk.Frame(row_f, bg=COL[cat], width=2).pack(side="left", fill="y")
                    tk.Label(row_f, text=item,
                             font=("Helvetica",8),
                             bg=COLORS["bg_content"], fg=COLORS["text_primary"],
                             anchor="w", padx=10).pack(fill="x", pady=3)

    # ── Despesas Fixas ────────────────────────────────────────────────────────
    def _build_despesas_fixas(self, parent):
        self._df_edit_id = None
        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Painel esquerdo: formulário ──────────────────────────────────────
        left = tk.Frame(root, bg=COLORS["bg_card"],
                        highlightthickness=1, highlightbackground=COLORS["border"],
                        width=320)
        left.pack(side="left", fill="y", padx=(0,12))
        left.pack_propagate(False)
        tk.Frame(left, bg=COLORS["orange"], height=4).pack(fill="x")
        tk.Label(left, text="\u25c8  Despesa Fixa",
                 font=("Helvetica",11,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                 ).pack(anchor="w", padx=16, pady=(12,8))

        def lbl(txt):
            tk.Label(left, text=txt, font=("Helvetica",8,"bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]
                     ).pack(anchor="w", padx=16)

        lbl("Descri\u00e7\u00e3o \u2605")
        self._df_desc_var = tk.StringVar()
        tk.Entry(left, textvariable=self._df_desc_var,
                 font=("Helvetica",10), bg=COLORS["bg_main"],
                 fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["blue"]
                 ).pack(fill="x", padx=16, pady=(4,10), ipady=6)

        lbl("Valor (\u00a5) \u2605")
        self._df_valor_entry = self._make_yen_entry(left, width=28)
        self._df_valor_entry.pack(padx=16, pady=(4,10), ipady=6)

        lbl("Categoria")
        self._df_cat_var = tk.StringVar(value="Geral")
        self._df_cat_combo = ttk.Combobox(left, textvariable=self._df_cat_var,
                     font=("Helvetica",9), state="readonly"
                     )
        self._df_cat_combo.pack(fill="x", padx=16, pady=(4,10), ipady=4)
        self._df_cat_combo.bind("<ButtonPress>", lambda e: self._df_refresh_categorias())
        self._df_refresh_categorias()

        lbl("M\\u00eas de Refer\\u00eancia")
        _df_ref_f = tk.Frame(left, bg=COLORS["bg_card"])
        _df_ref_f.pack(fill="x", padx=16, pady=(4,10))
        self._df_mes_var = tk.StringVar(value="03")
        self._df_ano_var = tk.StringVar(value="2026")
        _meses_opts = [f"{m:02d}" for m in range(1,13)]
        _anos_opts  = [str(y) for y in range(2026+2, 2019, -1)]
        ttk.Combobox(_df_ref_f, textvariable=self._df_mes_var,
                     values=_meses_opts, state="readonly",
                     font=("Helvetica",10), width=5
                     ).pack(side="left", ipady=5)
        tk.Label(_df_ref_f, text="/", font=("Helvetica",11,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]
                 ).pack(side="left", padx=4)
        ttk.Combobox(_df_ref_f, textvariable=self._df_ano_var,
                     values=_anos_opts, state="readonly",
                     font=("Helvetica",10), width=7
                     ).pack(side="left", ipady=5)
        self._df_data_var = tk.StringVar(value="03/2026")
        def _sync_df_ref(*_):
            self._df_data_var.set(f"{self._df_mes_var.get()}/{self._df_ano_var.get()}")
        self._df_mes_var.trace_add("write", _sync_df_ref)
        self._df_ano_var.trace_add("write", _sync_df_ref)

        self._df_rec_var = tk.IntVar(value=0)
        tk.Checkbutton(left, text="Recorrente (mensal)",
                       variable=self._df_rec_var,
                       font=("Helvetica",9),
                       bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                       selectcolor=COLORS["bg_main"],
                       activebackground=COLORS["bg_card"]
                       ).pack(anchor="w", padx=16, pady=(0,12))

        self._df_status_lbl = tk.Label(left, text="",
                                       font=("Helvetica",8,"bold"),
                                       bg=COLORS["bg_card"], fg=COLORS["green"])
        self._df_status_lbl.pack(anchor="w", padx=16, pady=(0,8))

        btn_row = tk.Frame(left, bg=COLORS["bg_card"])
        btn_row.pack(fill="x", padx=16, pady=(0,16))
        tk.Button(btn_row, text="  Salvar Despesa  ",
                  font=("Helvetica",10,"bold"),
                  bg=COLORS["orange"], fg="white", relief="flat", cursor="hand2",
                  command=self._df_salvar
                  ).pack(side="left", ipady=7, ipadx=4)
        tk.Button(btn_row, text="Limpar",
                  font=("Helvetica",9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._df_limpar
                  ).pack(side="left", padx=(8,0), ipady=7, ipadx=4)

        # ── Painel direito: hist\u00f3rico ────────────────────────────────────────
        right = tk.Frame(root, bg=COLORS["bg_card"],
                         highlightthickness=1, highlightbackground=COLORS["border"])
        right.pack(side="right", fill="both", expand=True)
        right.pack_propagate(True)
        tk.Frame(right, bg=COLORS["orange"], height=4).pack(fill="x")

        hdr_r = tk.Frame(right, bg=COLORS["bg_card"])
        hdr_r.pack(fill="x", padx=14, pady=(10,6))
        tk.Label(hdr_r, text="Hist\u00f3rico de Despesas",
                 font=("Helvetica",10,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        self._df_total_lbl = tk.Label(hdr_r, text="",
                                      font=("Helvetica",8),
                                      bg=COLORS["bg_card"],
                                      fg=COLORS["text_muted"])
        self._df_total_lbl.pack(side="left", padx=10)
        tk.Button(hdr_r, text="\u21ba", font=("Helvetica",9,"bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._df_refresh
                  ).pack(side="right")

        # Filtro m\u00eas
        ff = tk.Frame(right, bg=COLORS["bg_main"],
                      highlightthickness=1, highlightbackground=COLORS["border"])
        ff.pack(fill="x", padx=14, pady=(0,6), ipady=3)
        tk.Label(ff, text="M\u00eas:", font=("Helvetica",8),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]
                 ).pack(side="left", padx=(8,4))
        self._df_filtro_mes = tk.StringVar(value="Todos")
        import datetime as _dt2
        _anos = [str(y) for y in range(_dt2.date.today().year, 2020, -1)]
        _meses_ref = ["Todos"] + [f"{m:02d}/{y}"
                                   for y in _anos
                                   for m in range(1, 13)]
        ttk.Combobox(ff, textvariable=self._df_filtro_mes,
                     values=_meses_ref[:25], state="readonly",
                     font=("Helvetica",8), width=10
                     ).pack(side="left", ipady=2)
        self._df_filtro_mes.trace_add("write", lambda *_: self._df_refresh())

        # Header da tabela
        hdr_df = tk.Frame(right, bg=COLORS["bg_main"])
        hdr_df.pack(fill="x", padx=14)
        for txt, w in [("#",4),("Descri\u00e7\u00e3o",24),("Categoria",14),
                       ("Ref",10),("Rec.",6),("Valor",11),("A\u00e7\u00f5es",8)]:
            tk.Label(hdr_df, text=txt, font=("Helvetica",7,"bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w", padx=3
                     ).pack(side="left", pady=3)

        # Área de scroll para as linhas
        tbl_outer = tk.Frame(right, bg=COLORS["bg_card"])
        tbl_outer.pack(fill="both", expand=True, padx=14, pady=(0,10))

        vsb_df = tk.Scrollbar(tbl_outer, orient="vertical")
        vsb_df.pack(side="right", fill="y")

        body_cv = tk.Canvas(tbl_outer, bg=COLORS["bg_card"],
                            highlightthickness=0,
                            yscrollcommand=vsb_df.set)
        body_cv.pack(side="left", fill="both", expand=True)
        vsb_df.configure(command=body_cv.yview)

        self._df_rows_frame = tk.Frame(body_cv, bg=COLORS["bg_card"])
        self._df_rows_frame.bind("<Configure>",
            lambda e: body_cv.configure(scrollregion=body_cv.bbox("all")))
        _df_win = body_cv.create_window((0,0), window=self._df_rows_frame, anchor="nw")
        body_cv.bind("<Configure>",
            lambda e: body_cv.itemconfig(_df_win, width=e.width))

        body_cv.bind("<Enter>", lambda e: body_cv.bind_all(
            "<MouseWheel>",
            lambda ev: body_cv.yview_scroll(int(-1*(ev.delta/120)),"units")))
        body_cv.bind("<Leave>", lambda e: body_cv.unbind_all("<MouseWheel>"))

        self.after(300, self._df_refresh)
    def _df_refresh_categorias(self):
        """Recarrega categorias do BD no combo de Despesas Fixas."""
        try:
            cats = [r[0] for r in self.conn.execute(
                "SELECT nome FROM categorias_despesas ORDER BY nome").fetchall()]
            if not cats:
                cats = ["Geral"]
            self._df_cat_combo["values"] = cats
            if self._df_cat_var.get() not in cats:
                self._df_cat_var.set(cats[0])
        except Exception as e:
            logger.error(f"Erro ao carregar categorias: {e}", exc_info=True)

    def _df_salvar(self):
        desc  = self._df_desc_var.get().strip()
        valor = self._yen_raw(self._df_valor_entry)
        cat   = self._df_cat_var.get().strip() or "Geral"
        ref   = self._df_data_var.get().strip()
        rec   = int(self._df_rec_var.get())
        if not desc:
            self._df_status_lbl.configure(text="⚠ Informe a descrição.", fg=COLORS["red"]); return
        if not valor:
            self._df_status_lbl.configure(text="⚠ Informe o valor.", fg=COLORS["red"]); return
        if self._df_edit_id:
            self.conn.execute(
                "UPDATE despesas_fixas SET descricao=?,valor=?,categoria=?,data_ref=?,recorrente=? WHERE id=?",
                (desc, valor, cat, ref, rec, self._df_edit_id))
            self._df_edit_id = None
            self._df_status_lbl.configure(text="✔ Despesa atualizada!", fg=COLORS["green"])
        else:
            self.conn.execute(
                "INSERT INTO despesas_fixas (descricao,valor,categoria,data_ref,recorrente) VALUES (?,?,?,?,?)",
                (desc, valor, cat, ref, rec))
            self._df_status_lbl.configure(text="✔ Despesa registrada!", fg=COLORS["green"])
        self.conn.commit()
        self._df_limpar()
        self._df_refresh()

    def _df_limpar(self):
        self._df_edit_id = None
        self._df_desc_var.set("")
        self._df_valor_entry.delete(0, "end")
        self._df_cat_var.set("Geral")
        self._df_data_var.set("")
        self._df_rec_var.set(0)

    def _df_refresh(self):
        for w in self._df_rows_frame.winfo_children():
            w.destroy()
        fmes = getattr(self, "_df_filtro_mes", None)
        fmes = fmes.get() if fmes else "Todos"
        rows = [dict(r) for r in self.conn.execute(
            "SELECT * FROM despesas_fixas ORDER BY id DESC").fetchall()]
        if fmes != "Todos":
            rows = [r for r in rows if (r.get("data_ref") or "").startswith(fmes)]

        def _int(s):
            try: return int(str(s or "0").replace(",",""))
            except: return 0
        total = sum(_int(r.get("valor")) for r in rows)
        if hasattr(self, "_df_total_lbl"):
            self._df_total_lbl.configure(
                text=f"{len(rows)} registro(s)  |  Total: ¥ {total:,}")

        for i, r in enumerate(rows):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            tk.Frame(self._df_rows_frame, bg=COLORS["border"], height=1).pack(fill="x")
            row_f = tk.Frame(self._df_rows_frame, bg=rb)
            row_f.pack(fill="x")
            for txt, w, fg in [
                (str(r["id"]),                          4,  COLORS["text_muted"]),
                ((r.get("descricao") or "")[:24],       24, COLORS["text_primary"]),
                ((r.get("categoria") or "")[:14],       14, COLORS["text_secondary"]),
                ((r.get("data_ref") or "—")[:9],        10, COLORS["text_muted"]),
                ("✓" if r.get("recorrente") else "—",   6,  COLORS["green"]),
                (self._fmt_yen_display(r.get("valor")), 11, COLORS["orange"]),
            ]:
                tk.Label(row_f, text=str(txt), font=("Helvetica",8),
                         bg=rb, fg=fg, width=w, anchor="w", padx=3
                         ).pack(side="left", pady=4)
            af = tk.Frame(row_f, bg=rb)
            af.pack(side="left", padx=2)
            def _editar_df(rid=r["id"], rdesc=r.get("descricao",""),
                           rval=r.get("valor",""), rcat=r.get("categoria","Geral"),
                           rref=r.get("data_ref",""), rrec=r.get("recorrente",0)):
                self._df_edit_id = rid
                self._df_desc_var.set(rdesc)
                self._df_valor_entry.delete(0,"end")
                try:
                    v = int(str(rval or "0").replace(",",""))
                    self._df_valor_entry.insert(0, f"{v:,}")
                except Exception:
                    self._df_valor_entry.insert(0, rval or "")
                self._df_cat_var.set(rcat or "Geral")
                self._df_data_var.set(rref or "")
                if hasattr(self, "_df_mes_var") and rref and "/" in rref:
                    parts = rref.split("/")
                    if len(parts) == 2:
                        self._df_mes_var.set(parts[0].strip())
                        self._df_ano_var.set(parts[1].strip())
                self._df_rec_var.set(rrec or 0)
            def _excluir_df(rid=r["id"]):
                import tkinter.messagebox as _mb
                if _mb.askyesno("Excluir", "Excluir esta despesa?"):
                    self.conn.execute("DELETE FROM despesas_fixas WHERE id=?", (rid,))
                    self.conn.commit()
                    self._df_refresh()
            tk.Button(af, text="✎", font=("Helvetica",8),
                      bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                      padx=3, command=_editar_df
                      ).pack(side="left", padx=(3,1), pady=3)
            tk.Button(af, text="✕", font=("Helvetica",8),
                      bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                      padx=3, command=_excluir_df
                      ).pack(side="left", pady=3)

    # ── Confirmação saída sem salvar ──────────────────────────────────────────
    def _confirm_unsaved(self, form_name="formulário"):
        """Retorna True se pode prosseguir (sem dados ou usuário confirmou)."""
        import tkinter.messagebox as _mb
        return _mb.askyesno(
            "Dados não salvos",
            f"Há dados preenchidos no {form_name}.\n\nDeseja sair sem salvar?",
            icon="warning")

    # ── Cadastros ─────────────────────────────────────────────────────────────
    def _build_cadastros_subs(self, parent):
        subs = SUBMENUS["Cadastros"]
        for sub in subs:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Cadastros"][sub] = frame
            if sub == "Clientes":
                self._build_clientes(frame)
            elif sub == "Carros":
                self._build_carros(frame)
            elif sub == "Tipos de Custo":
                self._build_tipos_custo_cadastro(frame)
            elif sub == "Tipos de Serviço":
                self._build_tipos_servico_cadastro(frame)
            elif sub == "Categoria Desp. Fixas":
                self._build_categorias_despesas_cadastro(frame)
            else:
                self._build_placeholder(frame, sub)

        # Mostra o primeiro por padrão
        first = subs[0]
        self.sub_pages["Cadastros"][first].place(in_=parent,
                                                  x=0, y=0, relwidth=1, relheight=1)

    def _build_clientes(self, parent):
        self._cliente_edit_id = None
        self._cliente_filtro_var = tk.StringVar(value="")
        self._cliente_ordem_az   = tk.BooleanVar(value=True)   # True = A→Z
        self.clientes_data = [dict(r) for r in
                              self.conn.execute("SELECT * FROM clientes ORDER BY nome COLLATE NOCASE").fetchall()]

        container = tk.Frame(parent, bg=COLORS["bg_content"])
        container.pack(fill="both", expand=True, padx=24, pady=20)

        # ── FORMULÁRIO ────────────────────────────────────────────────────────
        form_card = tk.Frame(container, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        form_card.pack(side="left", fill="y", padx=(0, 16), ipadx=10)

        tk.Frame(form_card, bg=COLORS["accent"], height=4).pack(fill="x")
        tk.Label(form_card, text="◈  Cadastro de Cliente",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=20, pady=(14, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=(0, 16))

        tk.Label(form_card, text="Nome do Cliente",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        self.entry_nome = tk.Entry(form_card, font=("Helvetica", 11),
                                   bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                   insertbackground=COLORS["text_primary"],
                                   relief="flat", bd=0,
                                   highlightthickness=1,
                                   highlightbackground=COLORS["border"],
                                   highlightcolor=COLORS["accent"], width=28)
        self.entry_nome.pack(padx=20, pady=(4, 14), ipady=7)

        tk.Label(form_card, text="Telefone (ex: (080) 0000-0000)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        tel_frame = tk.Frame(form_card, bg=COLORS["bg_card"])
        tel_frame.pack(padx=20, pady=(4, 20), fill="x")
        self.entry_tel = tk.Entry(tel_frame, font=("Helvetica", 11),
                                  bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                  insertbackground=COLORS["text_primary"],
                                  relief="flat", bd=0,
                                  highlightthickness=1,
                                  highlightbackground=COLORS["border"],
                                  highlightcolor=COLORS["accent"], width=28)
        self.entry_tel.pack(ipady=7)
        self.entry_tel.bind("<KeyRelease>", self._fmt_telefone_jp)

        tk.Label(form_card, text="Cidade",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        self.entry_cidade = tk.Entry(form_card, font=("Helvetica", 11),
                                     bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                     insertbackground=COLORS["text_primary"],
                                     relief="flat", bd=0,
                                     highlightthickness=1,
                                     highlightbackground=COLORS["border"],
                                     highlightcolor=COLORS["accent"], width=28)
        self.entry_cidade.pack(padx=20, pady=(4, 14), ipady=7)

        self.lbl_status = tk.Label(form_card, text="", font=("Helvetica", 8),
                                   bg=COLORS["bg_card"], fg=COLORS["red"])
        self.lbl_status.pack(padx=20)

        btn_frame = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_frame.pack(padx=20, pady=(8, 20), fill="x")
        self.btn_salvar = tk.Button(btn_frame, text="  Salvar  ",
                                    font=("Helvetica", 10, "bold"),
                                    bg=COLORS["accent"], fg="white",
                                    relief="flat", cursor="hand2",
                                    activebackground=COLORS["accent2"],
                                    activeforeground="white",
                                    command=self._salvar_cliente)
        self.btn_salvar.pack(side="left", ipady=6, ipadx=6)
        tk.Button(btn_frame, text="  Limpar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  activebackground=COLORS["border_dark"],
                  command=self._limpar_form_cliente
                  ).pack(side="left", padx=(10, 0), ipady=6, ipadx=6)

        # ── TABELA ────────────────────────────────────────────────────────────
        table_card = tk.Frame(container, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True)
        tk.Frame(table_card, bg=COLORS["accent"], height=4).pack(fill="x")

        tbl_header = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_header.pack(fill="x", padx=20, pady=(14, 4))
        tk.Label(tbl_header, text="▦  Clientes Cadastrados",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_header, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_tabela_clientes).pack(side="right", padx=(8,0))
        self.lbl_total = tk.Label(tbl_header, text="0 registros", font=("Helvetica", 8),
                                  bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self.lbl_total.pack(side="right")

        # Barra de filtro
        filt_bar = tk.Frame(table_card, bg=COLORS["bg_main"])
        filt_bar.pack(fill="x", padx=20, pady=(0, 4))
        tk.Label(filt_bar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=6)
        cli_filtro_entry = tk.Entry(filt_bar, textvariable=self._cliente_filtro_var,
                                    font=("Helvetica", 9),
                                    bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                                    insertbackground=COLORS["text_primary"],
                                    relief="flat", bd=0,
                                    highlightthickness=1,
                                    highlightbackground=COLORS["border"],
                                    highlightcolor=COLORS["accent"], width=22)
        cli_filtro_entry.pack(side="left", ipady=5)
        self._cliente_filtro_var.trace_add("write", lambda *_: self._refresh_tabela_clientes())
        tk.Button(filt_bar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=5,
                  command=lambda: self._cliente_filtro_var.set("")
                  ).pack(side="left", padx=(4, 0), ipady=4)

        # Botão ordenação A→Z / Z→A
        self._btn_ordem_cli = tk.Button(
            filt_bar, text="A→Z", font=("Helvetica", 8, "bold"),
            bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2", padx=6,
            command=self._toggle_ordem_clientes)
        self._btn_ordem_cli.pack(side="right", ipady=4, padx=(0, 4))

        tk.Frame(table_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20)
        col_frame = tk.Frame(table_card, bg=COLORS["bg_main"])
        col_frame.pack(fill="x", padx=20)
        for txt, w in [("#", 4), ("Nome", 28), ("Telefone", 16), ("Ações", 10)]:
            tk.Label(col_frame, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=4, pady=6)

        scroll_frame = tk.Frame(table_card, bg=COLORS["bg_card"])
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=(0, 14))
        canvas = tk.Canvas(scroll_frame, bg=COLORS["bg_card"], highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
        self.rows_frame = tk.Frame(canvas, bg=COLORS["bg_card"])
        self.rows_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self._refresh_tabela_clientes()

    def _salvar_cliente(self):
        nome   = self.entry_nome.get().strip()
        tel    = self.entry_tel.get().strip()
        cidade = self.entry_cidade.get().strip()

        if not nome:
            self.lbl_status.configure(text="⚠ Informe o nome do cliente.", fg=COLORS["red"])
            return
        if len(tel) < 14:
            self.lbl_status.configure(text="⚠ Telefone inválido.", fg=COLORS["red"])
            return

        # Garante que coluna cidade existe
        try:
            self.conn.execute("ALTER TABLE clientes ADD COLUMN cidade TEXT DEFAULT ''")
            self.conn.commit()
        except Exception:
            pass

        if self._cliente_edit_id is not None:
            self.conn.execute("UPDATE clientes SET nome=?, telefone=?, cidade=? WHERE id=?",
                              (nome, tel, cidade, self._cliente_edit_id))
            self.conn.commit()
            self.lbl_status.configure(text="✔ Cliente atualizado!", fg=COLORS["green"])
            self._cliente_edit_id = None
            self.btn_salvar.configure(text="  Salvar  ")
        else:
            self.conn.execute("INSERT INTO clientes (nome, telefone, cidade) VALUES (?, ?, ?)", (nome, tel, cidade))
            self.conn.commit()
            self.lbl_status.configure(text="✔ Cliente cadastrado!", fg=COLORS["green"])

        self.clientes_data = [dict(r) for r in
                              self.conn.execute("SELECT * FROM clientes ORDER BY id").fetchall()]
        self._limpar_form_cliente(clear_status=False)
        self._refresh_tabela_clientes()

    def _limpar_form_cliente(self, clear_status=True):
        self.entry_nome.delete(0, tk.END)
        self.entry_tel.delete(0, tk.END)
        self.entry_cidade.delete(0, tk.END)
        self._cliente_edit_id = None
        self.btn_salvar.configure(text="  Salvar  ")
        if clear_status:
            self.lbl_status.configure(text="")

    def _editar_cliente(self, cid):
        for c in self.clientes_data:
            if c["id"] == cid:
                self._cliente_edit_id = cid
                self.entry_nome.delete(0, tk.END)
                self.entry_nome.insert(0, c["nome"])
                self.entry_tel.delete(0, tk.END)
                self.entry_tel.insert(0, c["telefone"])
                self.entry_cidade.delete(0, tk.END)
                self.entry_cidade.insert(0, c.get("cidade") or "")
                self.btn_salvar.configure(text="  Atualizar  ")
                self.lbl_status.configure(text="", fg=COLORS["red"])
                break

    # ── Cadastro de Carros ────────────────────────────────────────────────────
    def _build_carros(self, parent):
        self._carro_edit_id = None
        self._carro_filtro  = tk.StringVar(value="Todos")
        self._status_colors = {
            "Estoque": COLORS["green"],
            "Cliente": COLORS["blue"],
            "Daisha":  COLORS["orange"],
            "Inativo": COLORS["red"],
        }
        STATUS_OPTS = ["Estoque", "Cliente", "Daisha", "Inativo"]
        self.carros_data = [dict(r) for r in
                            self.conn.execute("SELECT * FROM carros ORDER BY id DESC").fetchall()]

        container = tk.Frame(parent, bg=COLORS["bg_content"])
        container.pack(fill="both", expand=True, padx=24, pady=20)

        # ── FORMULÁRIO ────────────────────────────────────────────────────────
        form_card = tk.Frame(container, bg=COLORS["bg_card"],
                             highlightthickness=1,
                             highlightbackground=COLORS["border"])
        form_card.pack(side="left", fill="y", padx=(0, 16), ipadx=10)

        tk.Frame(form_card, bg=COLORS["accent"], height=4).pack(fill="x")
        tk.Label(form_card, text="◈  Cadastro de Carro",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(
                     anchor="w", padx=20, pady=(14, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(
            fill="x", padx=20, pady=(0, 12))

        def lbl(txt):
            tk.Label(form_card, text=txt, font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)

        def entry(var_name, width=28):
            e = tk.Entry(form_card, font=("Helvetica", 11),
                         bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                         insertbackground=COLORS["text_primary"],
                         relief="flat", bd=0,
                         highlightthickness=1,
                         highlightbackground=COLORS["border"],
                         highlightcolor=COLORS["accent"],
                         width=width)
            e.pack(padx=20, pady=(4, 12), ipady=7)
            setattr(self, var_name, e)

        lbl("Carro (Marca / Modelo)"); entry("carro_entry_carro")
        lbl("Ano");                    entry("carro_entry_ano", width=10)
        lbl("Cor");                    entry("carro_entry_cor", width=18)
        lbl("Placa");                  entry("carro_entry_placa", width=14)
        lbl("Chassi");                 entry("carro_entry_chassi")

        tk.Label(form_card, text="Status  ★ obrigatório",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["accent"]).pack(anchor="w", padx=20)

        self.carro_status_var = tk.StringVar(value="")
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TCombobox",
                        fieldbackground=COLORS["bg_main"],
                        background=COLORS["bg_main"],
                        foreground=COLORS["text_primary"],
                        selectbackground=COLORS["accent"],
                        selectforeground="white")
        self._status_combo_carro = ttk.Combobox(form_card, textvariable=self.carro_status_var,
                     values=STATUS_OPTS, state="readonly",
                     font=("Helvetica", 10), width=26)
        self._status_combo_carro.pack(padx=20, pady=(4, 10), ipady=4)
        self.carro_status_var.trace_add("write", lambda *_: self._toggle_cliente_combo())

        # Campo Cliente — aparece apenas quando status = Cliente
        self._cliente_frame = tk.Frame(form_card, bg=COLORS["bg_card"])
        tk.Label(self._cliente_frame, text="Cliente  ★ obrigatório",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["accent"]).pack(anchor="w", padx=20)
        self.carro_cliente_var = tk.StringVar(value="")
        self.carro_cliente_combo = ttk.Combobox(self._cliente_frame,
                                                textvariable=self.carro_cliente_var,
                                                state="readonly",
                                                font=("Helvetica", 10), width=26)
        self.carro_cliente_combo.pack(padx=20, pady=(4, 10), ipady=4)
        self.carro_cliente_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_clientes_carro())

        self.lbl_carro_status = tk.Label(form_card, text="",
                                         font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["red"])
        self.lbl_carro_status.pack(padx=20)

        btn_f = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_f.pack(padx=20, pady=(6, 20), fill="x")
        self.btn_salvar_carro = tk.Button(
            btn_f, text="  Salvar  ", font=("Helvetica", 10, "bold"),
            bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
            activebackground=COLORS["accent2"], activeforeground="white",
            command=self._salvar_carro)
        self.btn_salvar_carro.pack(side="left", ipady=6, ipadx=6)
        tk.Button(btn_f, text="  Limpar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  activebackground=COLORS["border_dark"],
                  command=self._limpar_form_carro
                  ).pack(side="left", padx=(10, 0), ipady=6, ipadx=6)

        # ── TABELA ────────────────────────────────────────────────────────────
        table_card = tk.Frame(container, bg=COLORS["bg_card"],
                              highlightthickness=1,
                              highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True)
        tk.Frame(table_card, bg=COLORS["accent"], height=4).pack(fill="x")

        tbl_h = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_h.pack(fill="x", padx=20, pady=(14, 6))
        tk.Label(tbl_h, text="▦  Carros Cadastrados",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_h, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=6, pady=2,
                  command=self._refresh_tabela_carros
                  ).pack(side="right", padx=(0,8))
        self.lbl_total_carros = tk.Label(tbl_h, text="0 registros",
                                         font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self.lbl_total_carros.pack(side="right")

        # ── Barra de filtros ──────────────────────────────────────────────────
        filter_bar = tk.Frame(table_card, bg=COLORS["bg_main"])
        filter_bar.pack(fill="x", padx=20, pady=(0, 4))
        tk.Label(filter_bar, text="Filtrar:", font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(
                     side="left", padx=(0, 6), pady=6)

        self._filter_btns = {}
        for opt in ["Todos", "Estoque", "Cliente", "Daisha", "Inativo"]:
            c_bg = COLORS["accent"] if opt == "Todos" else self._status_colors.get(opt, COLORS["border"])
            fb = tk.Button(filter_bar, text=opt, font=("Helvetica", 8, "bold"),
                           bg=COLORS["accent"] if opt == "Todos" else COLORS["border"],
                           fg="white" if opt == "Todos" else COLORS["text_secondary"],
                           relief="flat", cursor="hand2", padx=8, pady=3,
                           command=lambda o=opt: self._aplicar_filtro_carros(o))
            fb.pack(side="left", padx=(0, 4))
            self._filter_btns[opt] = fb

        tk.Frame(table_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20)

        # Container com scroll horizontal e vertical
        tbl_outer = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_outer.pack(fill="both", expand=True, padx=20, pady=(0,14))

        vsb_c = tk.Scrollbar(tbl_outer, orient="vertical")
        hsb_c = tk.Scrollbar(tbl_outer, orient="horizontal")
        vsb_c.pack(side="right", fill="y")
        hsb_c.pack(side="bottom", fill="x")

        canvas2 = tk.Canvas(tbl_outer, bg=COLORS["bg_card"], highlightthickness=0,
                            yscrollcommand=vsb_c.set, xscrollcommand=hsb_c.set)
        canvas2.pack(side="left", fill="both", expand=True)
        vsb_c.configure(command=canvas2.yview)
        hsb_c.configure(command=canvas2.xview)

        # Frame interno com header + rows juntos (scroll em bloco)
        inner_tbl = tk.Frame(canvas2, bg=COLORS["bg_card"])
        canvas2.create_window((0,0), window=inner_tbl, anchor="nw")
        inner_tbl.bind("<Configure>",
            lambda e: canvas2.configure(scrollregion=canvas2.bbox("all")))

        col_f = tk.Frame(inner_tbl, bg=COLORS["bg_main"])
        col_f.pack(fill="x")
        for txt, w in [("#",4),("Carro",16),("Ano",5),("Cor",8),
                       ("Placa",15),("Chassi",12),("Status",9),("Cliente",12),
                       ("T.Compra",10),("Comprado de",14),("T.Venda",16),("Vendido para",16),
                       ("Shaken",10),("St.SK",8),("Ações",8)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=3, pady=6)

        self.carros_rows_frame = tk.Frame(inner_tbl, bg=COLORS["bg_card"])
        self.carros_rows_frame.pack(fill="x")

        self._refresh_tabela_carros()

    def _toggle_cliente_combo(self):
        """Mostra/oculta o campo Cliente conforme status selecionado."""
        if self.carro_status_var.get() == "Cliente":
            self._cliente_frame.pack(after=self._status_combo_carro, fill="x")
            self._atualizar_combo_clientes_carro()
        else:
            self._cliente_frame.pack_forget()
            self.carro_cliente_var.set("")

    def _atualizar_combo_clientes_carro(self):
        clientes = [dict(r) for r in
                    self.conn.execute("SELECT * FROM clientes ORDER BY nome").fetchall()]
        opts = [f"{c['id']} — {c['nome']}" for c in clientes]
        self.carro_cliente_combo["values"] = opts if opts else ["Nenhum cliente cadastrado"]

    def _salvar_carro(self):
        carro  = self.carro_entry_carro.get().strip()
        ano    = self.carro_entry_ano.get().strip()
        cor    = self.carro_entry_cor.get().strip()
        placa  = self.carro_entry_placa.get().strip()
        chassi = self.carro_entry_chassi.get().strip()
        status = self.carro_status_var.get().strip()

        if not carro:
            self.lbl_carro_status.configure(text="⚠ Informe o carro.", fg=COLORS["red"]); return
        if not status:
            self.lbl_carro_status.configure(text="⚠ Selecione o Status.", fg=COLORS["red"]); return

        # Cliente obrigatório quando status = Cliente
        cliente_id = None
        if status == "Cliente":
            sel = self.carro_cliente_var.get().strip()
            if not sel or sel == "Nenhum cliente cadastrado":
                self.lbl_carro_status.configure(
                    text="⚠ Vincule um cliente para status Cliente.", fg=COLORS["red"]); return
            try:
                cliente_id = int(sel.split("—")[0].strip())
            except Exception:
                self.lbl_carro_status.configure(text="⚠ Cliente inválido.", fg=COLORS["red"]); return

        if self._carro_edit_id is not None:
            cid = self._carro_edit_id
            # Registrar historico_status se o status mudou
            import datetime as _dt2
            row_ant = self.conn.execute(
                "SELECT status, historico_status FROM carros WHERE id=?", (cid,)).fetchone()
            if row_ant and row_ant[0] != status:
                hist_atual = row_ant[1] or ""
                entradas = [e.strip() for e in hist_atual.split("|") if e.strip()]
                nova_entrada = f"{row_ant[0]} ({_dt2.date.today().strftime('%d/%m/%Y')})"
                if nova_entrada not in entradas:
                    entradas.append(nova_entrada)
                novo_hist = " | ".join(entradas)
                self.conn.execute(
                    "UPDATE carros SET historico_status=? WHERE id=?", (novo_hist, cid))
            self.conn.execute(
                "UPDATE carros SET carro=?, ano=?, cor=?, placa=?, chassi=?, status=?, cliente_id=? WHERE id=?",
                (carro, ano, cor, placa, chassi, status, cliente_id, cid))
            self.conn.commit()

            # ── Propaga alterações para todas as telas vinculadas ──────────────
            # 1. Compras: recalcula texto carro (campo carro armazena "ID — Nome Ano")
            compras_vinc = self.conn.execute(
                "SELECT id FROM compras WHERE carro_id=?", (cid,)).fetchall()
            for cp in compras_vinc:
                novo_nome = f"{cid} — {carro} {ano}"
                self.conn.execute(
                    "UPDATE compras SET carro=?, cor=?, placa=? WHERE id=?",
                    (novo_nome, cor, placa, cp[0]))

            # 2. Vendas e Serviços (tabelas futuras — silencioso se não existirem)
            for tabela in ["vendas", "servicos"]:
                try:
                    self.conn.execute(
                        f"UPDATE {tabela} SET carro=?, cor=?, placa=? WHERE carro_id=?",
                        (carro, cor, placa, cid))
                except Exception:
                    pass

            self.conn.commit()

            # Recarrega compras_data em memória
            self.compras_data = [dict(r) for r in
                self.conn.execute("SELECT * FROM compras ORDER BY id DESC").fetchall()]

            self.lbl_carro_status.configure(text="✔ Carro atualizado!", fg=COLORS["green"])
            self._carro_edit_id = None
            self.btn_salvar_carro.configure(text="  Salvar  ")

            # Atualiza telas abertas
            try: self._refresh_tabela_compras()
            except Exception: pass
            try: self._refresh_estoque_veiculos()
            except Exception: pass

        else:
            self.conn.execute(
                "INSERT INTO carros (carro, ano, cor, placa, chassi, status, cliente_id) VALUES (?,?,?,?,?,?,?)",
                (carro, ano, cor, placa, chassi, status, cliente_id))
            self.conn.commit()
            self.lbl_carro_status.configure(text="✔ Carro cadastrado!", fg=COLORS["green"])

        self.carros_data = [dict(r) for r in
                            self.conn.execute("SELECT * FROM carros ORDER BY id DESC").fetchall()]
        self._limpar_form_carro(clear_status=False)
        self._refresh_tabela_carros()

    def _limpar_form_carro(self, clear_status=True):
        for f in ["carro_entry_carro","carro_entry_ano","carro_entry_cor",
                  "carro_entry_placa","carro_entry_chassi"]:
            getattr(self, f).delete(0, tk.END)
        self.carro_status_var.set("")
        self.carro_cliente_var.set("")
        self._cliente_frame.pack_forget()
        self._carro_edit_id = None
        self.btn_salvar_carro.configure(text="  Salvar  ")
        if clear_status:
            self.lbl_carro_status.configure(text="")

    def _editar_carro(self, cid):
        for c in self.carros_data:
            if c["id"] == cid:
                self._carro_edit_id = cid
                self.carro_entry_carro.delete(0, tk.END);  self.carro_entry_carro.insert(0, c["carro"] or "")
                self.carro_entry_ano.delete(0, tk.END);    self.carro_entry_ano.insert(0, c["ano"] or "")
                self.carro_entry_cor.delete(0, tk.END);    self.carro_entry_cor.insert(0, c["cor"] or "")
                self.carro_entry_placa.delete(0, tk.END);  self.carro_entry_placa.insert(0, c["placa"] or "")
                self.carro_entry_chassi.delete(0, tk.END); self.carro_entry_chassi.insert(0, c["chassi"] or "")
                self.carro_status_var.set(c["status"])
                if c["status"] == "Cliente" and c.get("cliente_id"):
                    self._atualizar_combo_clientes_carro()
                    cli = self.conn.execute(
                        "SELECT * FROM clientes WHERE id=?", (c["cliente_id"],)).fetchone()
                    if cli:
                        self.carro_cliente_var.set(f"{cli['id']} — {cli['nome']}")
                    self._cliente_frame.pack(after=self._status_combo_carro, fill="x")
                else:
                    self.carro_cliente_var.set("")
                    self._cliente_frame.pack_forget()
                self.btn_salvar_carro.configure(text="  Atualizar  ")
                self.lbl_carro_status.configure(text="")
                break
    def _excluir_carro(self, cid):
        """Bloqueia exclusão se houver qualquer registro vinculado ao carro."""
        vinculos = []

        n_compras = self.conn.execute(
            "SELECT COUNT(*) FROM compras WHERE carro_id=?", (cid,)).fetchone()[0]
        if n_compras:
            vinculos.append(f"• {n_compras} compra(s)")

        for tabela, label in [("vendas","venda(s)"),("servicos","serviço(s)")]:
            try:
                n = self.conn.execute(
                    f"SELECT COUNT(*) FROM {tabela} WHERE carro_id=?", (cid,)).fetchone()[0]
                if n: vinculos.append(f"• {n} {label}")
            except Exception: pass

        # Shaken vinculados
        try:
            sk_ids = [r[0] for r in self.conn.execute(
                "SELECT id FROM shaken WHERE carro_id=?", (cid,)).fetchall()]
            if sk_ids:
                vinculos.append(f"• {len(sk_ids)} shaken registrado(s)")
                # Custos SK vinculados
                for sk_id in sk_ids:
                    n_csk = self.conn.execute(
                        "SELECT COUNT(*) FROM custos_sk WHERE shaken_id=?", (sk_id,)).fetchone()[0]
                    if n_csk:
                        vinculos.append(f"  — {n_csk} custo(s) SK do shaken")
                        break
        except Exception: pass

        # Custos de compras vinculadas
        try:
            comp_ids = [r[0] for r in self.conn.execute(
                "SELECT id FROM compras WHERE carro_id=?", (cid,)).fetchall()]
            for cmp_id in comp_ids:
                n_cus = self.conn.execute(
                    "SELECT COUNT(*) FROM custos WHERE compra_id=?", (cmp_id,)).fetchone()[0]
                if n_cus:
                    vinculos.append(f"  — {n_cus} custo(s) da compra #{cmp_id}")
                    break
        except Exception: pass

        if vinculos:
            msgbox.showwarning(
                "Exclusão Bloqueada",
                "Este carro não pode ser excluído pois possui:\n\n"
                + "\n".join(vinculos)
                + "\n\nRemova todos os registros antes de excluir.")
            return

        # Confirmação com digitação de "excluir"
        carro_row = self.conn.execute("SELECT carro FROM carros WHERE id=?", (cid,)).fetchone()
        carro_nome = carro_row[0] if carro_row else f"#{cid}"
        dlg = tk.Toplevel(self)
        dlg.title("Confirmar Exclusão")
        dlg.configure(bg=COLORS["bg_card"])
        dlg.resizable(False, False)
        dlg.lift()
        dlg.attributes("-topmost", True)
        dlg.after(100, lambda: dlg.attributes("-topmost", False))
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 380) // 2
        y = self.winfo_y() + (self.winfo_height() - 200) // 2
        dlg.geometry(f"380x200+{x}+{y}")

        tk.Frame(dlg, bg=COLORS["red"], height=4).pack(fill="x")
        tk.Label(dlg, text=f"Excluir carro: {carro_nome}?",
                 font=("Helvetica",11,"bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(pady=(14,4))
        tk.Label(dlg, text="Digite  excluir  para confirmar:",
                 font=("Helvetica",9),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack()
        conf_entry = tk.Entry(dlg, font=("Helvetica",11),
                              bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                              insertbackground=COLORS["text_primary"],
                              relief="flat", bd=0,
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=18, justify="center")
        conf_entry.pack(pady=(6,4), ipady=6)
        conf_entry.focus_set()
        lbl_err = tk.Label(dlg, text="", font=("Helvetica",8),
                           bg=COLORS["bg_card"], fg=COLORS["red"])
        lbl_err.pack()
        btn_f = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_f.pack(pady=(4,12))

        def confirmar():
            if conf_entry.get().strip().lower() != "excluir":
                lbl_err.configure(text="⚠ Digite exatamente  excluir")
                return
            dlg.destroy()
            self.conn.execute("DELETE FROM carros WHERE id=?", (cid,))
            self.conn.commit()
            self.carros_data = [dict(r) for r in
                self.conn.execute("SELECT * FROM carros ORDER BY id DESC").fetchall()]
            if self._carro_edit_id == cid:
                self._limpar_form_carro()
            self._refresh_tabela_carros()

        tk.Button(btn_f, text="  Excluir  ", font=("Helvetica",10,"bold"),
                  bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                  command=confirmar).pack(side="left", ipady=5, ipadx=6)
        tk.Button(btn_f, text="  Cancelar  ", font=("Helvetica",10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(side="left", padx=(8,0), ipady=5, ipadx=6)
        conf_entry.bind("<Return>", lambda e: confirmar())

    def _aplicar_filtro_carros(self, opt):
        self._carro_filtro.set(opt)
        # Atualiza visual dos botões
        for o, btn in self._filter_btns.items():
            if o == opt:
                active_bg = COLORS["accent"] if o == "Todos" else self._status_colors.get(o, COLORS["accent"])
                btn.configure(bg=active_bg, fg="white")
            else:
                btn.configure(bg=COLORS["border"], fg=COLORS["text_secondary"])
        self._refresh_tabela_carros()

    def _refresh_tabela_carros(self):
        for w in self.carros_rows_frame.winfo_children():
            w.destroy()

        filtro_val = self._carro_filtro.get() if hasattr(self, "_carro_filtro") else "Todos"

        # Busca carros com nome do cliente via JOIN
        rows = [dict(r) for r in self.conn.execute(
            "SELECT ca.*, cl.nome AS cliente_nome FROM carros ca "
            "LEFT JOIN clientes cl ON ca.cliente_id = cl.id ORDER BY ca.id DESC"
        ).fetchall()]

        # Inativo fora da listagem padrão
        if filtro_val == "Todos":
            lista = [c for c in rows if c["status"] != "Inativo"]
        else:
            lista = [c for c in rows if c["status"] == filtro_val]

        self.lbl_total_carros.configure(text=f"{len(lista)} registro(s)")

        if not lista:
            msg = "Nenhum carro inativo." if filtro_val == "Inativo" else "Nenhum carro cadastrado."
            tk.Label(self.carros_rows_frame, text=msg,
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        for i, c in enumerate(lista):
            row_bg = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self.carros_rows_frame, bg=row_bg)
            row.pack(fill="x")

            for val, w in [(str(c["id"]), 4), (c["carro"] or "", 16), (c["ano"] or "", 5),
                           (c["cor"] or "", 8), (c["placa"] or "", 15), (c["chassi"] or "", 12)]:
                tk.Label(row, text=val, font=("Helvetica", 9),
                         bg=row_bg, fg=COLORS["text_primary"],
                         width=w, anchor="w").pack(side="left", padx=3, pady=7)

            sc = self._status_colors.get(c["status"], COLORS["text_muted"])
            tk.Label(row, text=c["status"], font=("Helvetica", 8, "bold"),
                     bg=sc, fg="white", width=9,
                     anchor="center", padx=4).pack(side="left", padx=3, pady=4)

            # Nome do cliente vinculado (só aparece se status=Cliente)
            cliente_txt = c.get("cliente_nome") or "—"
            cli_fg = COLORS["blue"] if c["status"] == "Cliente" else COLORS["text_muted"]
            tk.Label(row, text=cliente_txt, font=("Helvetica", 8),
                     bg=row_bg, fg=cli_fg,
                     width=12, anchor="w").pack(side="left", padx=3)

            # Colunas extra: tipo compra, comprado de, tipo venda, vendido para
            try:
                comp_row = self.conn.execute(
                    "SELECT co.tipo, cl.nome FROM compras co "
                    "LEFT JOIN clientes cl ON co.cliente_id=cl.id "
                    "WHERE co.carro_id=? ORDER BY co.id DESC LIMIT 1", (c["id"],)).fetchone()
                tipo_compra_txt  = comp_row[0] if comp_row else "—"
                comprado_de_txt  = comp_row[1] if (comp_row and comp_row[1]) else "—"
            except Exception:
                tipo_compra_txt = comprado_de_txt = "—"
            try:
                vend_row = self.conn.execute(
                    "SELECT v.tipo_venda, cl.nome FROM vendas v "
                    "LEFT JOIN clientes cl ON v.cliente_id=cl.id "
                    "WHERE v.carro_id=? ORDER BY v.id DESC LIMIT 1", (c["id"],)).fetchone()
                tipo_venda_txt  = vend_row[0] if vend_row else "—"
                vendido_para_txt = vend_row[1] if (vend_row and vend_row[1]) else "—"
            except Exception:
                tipo_venda_txt = vendido_para_txt = "—"

            for txt, w, fg in [
                (tipo_compra_txt[:10],  10, COLORS["orange"]),
                (comprado_de_txt[:14],  14, COLORS["text_secondary"]),
                (tipo_venda_txt[:16],   16, COLORS["green"]),
                (vendido_para_txt[:16], 16, COLORS["accent"]),
            ]:
                tk.Label(row, text=txt, font=("Helvetica", 8),
                         bg=row_bg, fg=fg,
                         width=w, anchor="w").pack(side="left", padx=3)

            # Shaken date + status
            sk_data = c.get("data_shaken") or "—"
            sk_st_txt, sk_st_cor = self._sk_status_info(c.get("data_shaken"), c.get("por_conta_shaken", 0))
            tk.Label(row, text=sk_data[:10], font=("Helvetica", 8),
                     bg=row_bg, fg=COLORS["blue"], width=10, anchor="w"
                     ).pack(side="left", padx=3)
            tk.Label(row, text=sk_st_txt, font=("Helvetica", 7, "bold"),
                     bg=sk_st_cor, fg="white", width=8, anchor="center"
                     ).pack(side="left", padx=3, pady=4)

            acts = tk.Frame(row, bg=row_bg)
            acts.pack(side="left", padx=4)
            tk.Button(acts, text="✏", font=("Helvetica", 9),
                      bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                      padx=6, pady=2,
                      command=lambda cid=c["id"]: self._editar_carro(cid)
                      ).pack(side="left", padx=(0, 4))
            tk.Button(acts, text="✕", font=("Helvetica", 9),
                      bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                      padx=6, pady=2,
                      command=lambda cid=c["id"]: self._excluir_carro(cid)
                      ).pack(side="left", padx=(0, 4))
            is_inativo = c.get("status") == "Inativo"
            tk.Button(acts, text="⛔ Inativar", font=("Helvetica", 7, "bold"),
                      bg=COLORS["text_muted"] if is_inativo else "#7F1D1D",
                      fg="white", relief="flat",
                      cursor="arrow" if is_inativo else "hand2",
                      padx=4, pady=2,
                      state="disabled" if is_inativo else "normal",
                      command=(lambda cid=c["id"]: self._inativar_carro(cid))
                      ).pack(side="left")

    def _inativar_carro(self, cid):
        """Abre popup para registrar observação de inativação e inativa o carro."""
        import tkinter.messagebox as mb
        import datetime as _dti

        carro_row = self.conn.execute("SELECT carro, status FROM carros WHERE id=?", (cid,)).fetchone()
        if not carro_row:
            return
        carro_nome, status_atual = carro_row
        if status_atual == "Inativo":
            mb.showinfo("Já Inativo", f"O carro {carro_nome} já está inativo.")
            return

        # Popup de confirmação com campo de observação
        popup = tk.Toplevel(self)
        popup.title("Inativar Veículo")
        popup.configure(bg=COLORS["bg_card"])
        popup.resizable(False, False)
        popup.grab_set()

        # Centraliza
        popup.update_idletasks()
        x = self.winfo_x() + self.winfo_width()//2 - 220
        y = self.winfo_y() + self.winfo_height()//2 - 160
        popup.geometry(f"440x320+{x}+{y}")

        tk.Frame(popup, bg=COLORS["red"], height=4).pack(fill="x")
        tk.Label(popup, text="⛔  Inativar Veículo",
                 font=("Helvetica", 12, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["red"]).pack(anchor="w", padx=20, pady=(14,4))
        tk.Label(popup, text=f"Veículo: {carro_nome}",
                 font=("Helvetica", 9), bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        tk.Frame(popup, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=8)

        tk.Label(popup, text="Motivo / Observação da inativação:",
                 font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        obs_text = tk.Text(popup, font=("Helvetica", 9), height=6,
                           bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                           insertbackground=COLORS["text_primary"],
                           relief="flat", bd=0,
                           highlightthickness=1, highlightbackground=COLORS["border"])
        obs_text.pack(fill="x", padx=20, pady=(4, 12), ipady=4)

        status_lbl = tk.Label(popup, text="", font=("Helvetica", 8),
                              bg=COLORS["bg_card"], fg=COLORS["red"])
        status_lbl.pack(anchor="w", padx=20)

        btn_f = tk.Frame(popup, bg=COLORS["bg_card"])
        btn_f.pack(padx=20, pady=(0,16), fill="x")

        def _confirmar():
            obs = obs_text.get("1.0", tk.END).strip()
            hoje_str = _dti.date.today().strftime("%d/%m/%Y")
            # Garante coluna obs_inativacao existe
            try:
                self.conn.execute("ALTER TABLE carros ADD COLUMN obs_inativacao TEXT")
                self.conn.commit()
            except Exception:
                pass
            try:
                self.conn.execute("ALTER TABLE carros ADD COLUMN data_inativacao TEXT")
                self.conn.commit()
            except Exception:
                pass
            self.conn.execute(
                "UPDATE carros SET status='Inativo', obs_inativacao=?, data_inativacao=? WHERE id=?",
                (obs or None, hoje_str, cid))
            self.conn.commit()
            # Recarrega dados
            self.carros_data = [dict(r) for r in self.conn.execute(
                "SELECT * FROM carros ORDER BY id DESC").fetchall()]
            self._refresh_tabela_carros()
            popup.destroy()

        tk.Button(btn_f, text="  Confirmar Inativação  ",
                  font=("Helvetica", 10, "bold"),
                  bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                  command=_confirmar).pack(side="left", ipady=6, ipadx=4)
        tk.Button(btn_f, text="Cancelar",
                  font=("Helvetica", 9), bg=COLORS["border"],
                  fg=COLORS["text_secondary"], relief="flat", cursor="hand2",
                  command=popup.destroy).pack(side="left", padx=(10,0), ipady=6, ipadx=4)

    def _toggle_ordem_clientes(self):
        self._cliente_ordem_az.set(not self._cliente_ordem_az.get())
        lbl = "A→Z" if self._cliente_ordem_az.get() else "Z→A"
        self._btn_ordem_cli.configure(text=lbl)
        self._refresh_tabela_clientes()

    def _refresh_tabela_clientes(self):
        for w in self.rows_frame.winfo_children():
            w.destroy()

        termo = (self._cliente_filtro_var.get() if hasattr(self, "_cliente_filtro_var")
                 else "").strip().lower()
        lista = [c for c in self.clientes_data
                 if not termo or
                 termo in c["nome"].lower() or
                 termo in (c["telefone"] or "").lower()]

        az = self._cliente_ordem_az.get() if hasattr(self, "_cliente_ordem_az") else True
        lista = sorted(lista, key=lambda c: c["nome"].lower(), reverse=not az)

        self.lbl_total.configure(text=f"{len(lista)} registro(s)")

        if not lista:
            tk.Label(self.rows_frame,
                     text="Nenhum cliente encontrado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        for i, c in enumerate(lista):
            row_bg = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self.rows_frame, bg=row_bg)
            row.pack(fill="x")

            tk.Label(row, text=str(c["id"]), font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_muted"],
                     width=4, anchor="w").pack(side="left", padx=4, pady=7)

            tk.Label(row, text=c["nome"], font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_primary"],
                     width=28, anchor="w").pack(side="left", padx=4)

            tk.Label(row, text=c["telefone"], font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_secondary"],
                     width=16, anchor="w").pack(side="left", padx=4)

            acts = tk.Frame(row, bg=row_bg)
            acts.pack(side="left", padx=4)

            tk.Button(acts, text="✏", font=("Helvetica", 9),
                      bg=COLORS["blue"], fg="white", relief="flat",
                      cursor="hand2", padx=6, pady=2,
                      command=lambda cid=c["id"]: self._editar_cliente(cid)
                      ).pack(side="left", padx=(0, 4))

            tk.Button(acts, text="✕", font=("Helvetica", 9),
                      bg=COLORS["red"], fg="white", relief="flat",
                      cursor="hand2", padx=6, pady=2,
                      command=lambda cid=c["id"]: self._excluir_cliente(cid)
                      ).pack(side="left")

    def _excluir_cliente(self, cid):
        # Verifica vínculos em vendas e serviços no banco
        n_vendas   = self.conn.execute(
            "SELECT COUNT(*) FROM vendas   WHERE cliente_id=?", (cid,)).fetchone()[0]
        n_servicos = self.conn.execute(
            "SELECT COUNT(*) FROM servicos WHERE cliente_id=?", (cid,)).fetchone()[0]
        # Tabela vendas/servicos podem não existir ainda — tratar erro silencioso
        try:
            pass
        except Exception:
            n_vendas = n_servicos = 0
        if n_vendas or n_servicos:
            msgbox.showwarning(
                "Exclusão Bloqueada",
                f"Este cliente está vinculado a {n_vendas} venda(s) e/ou "
                f"{n_servicos} serviço(s) e não pode ser excluído.")
            return
        if not msgbox.askyesno("Confirmar", "Deseja excluir este cliente?"):
            return
        self.conn.execute("DELETE FROM clientes WHERE id=?", (cid,))
        self.conn.commit()
        self.clientes_data = [dict(r) for r in
                              self.conn.execute("SELECT * FROM clientes ORDER BY nome").fetchall()]
        if self._cliente_edit_id == cid:
            self._limpar_form_cliente()
        self._refresh_tabela_clientes()
    def _build_entrada_subs(self, parent):
        subs = SUBMENUS["Entrada/Compra"]
        for sub in subs:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Entrada/Compra"][sub] = frame
            if sub == "Nova Compra":
                self._build_nova_compra(frame)
            elif sub == "Custos":
                self._build_custos(frame)
            elif sub == "Despesas Fixas":
                self._build_despesas_fixas(frame)
            else:
                self._build_placeholder(frame, sub)
        self.sub_pages["Entrada/Compra"][subs[0]].place(
            in_=parent, x=0, y=0, relwidth=1, relheight=1)

    def _build_custos(self, parent):
        self._custo_edit_id      = None
        self._custo_filtro_carro = tk.StringVar(value="Todos")
        self._custo_filtro_tipo  = tk.StringVar(value="Todos")
        self._custo_filtro_mes   = tk.StringVar(value="Todos")
        self._custo_filtro_ano   = tk.StringVar(value="Todos")

        container = tk.Frame(parent, bg=COLORS["bg_content"])
        container.pack(fill="both", expand=True, padx=24, pady=20)

        # ── FORMULÁRIO ────────────────────────────────────────────────────────
        left_col = tk.Frame(container, bg=COLORS["bg_content"])
        left_col.pack(side="left", fill="y", padx=(0, 16))

        form_card = tk.Frame(left_col, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        form_card.pack(fill="x", ipadx=10)
        tk.Frame(form_card, bg=COLORS["accent"], height=4).pack(fill="x")
        tk.Label(form_card, text="◈  Adicionar Custo", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=20, pady=(12, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=(0, 10))

        def lbl(txt, required=False):
            fg = COLORS["accent"] if required else COLORS["text_secondary"]
            tk.Label(form_card, text=txt + ("  ★" if required else ""),
                     font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"], fg=fg
                     ).pack(anchor="w", padx=20)

        lbl("Carro", required=True)
        self.custo_compra_var = tk.StringVar(value="")
        self.custo_compra_combo = ttk.Combobox(form_card, textvariable=self.custo_compra_var,
                                               state="readonly", font=("Helvetica", 10), width=32)
        self.custo_compra_combo.pack(padx=20, pady=(4, 12), ipady=4)
        self.custo_compra_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_compras())

        lbl("Tipo de Custo", required=True)
        self.custo_tipo_var = tk.StringVar(value="")
        self.custo_tipo_combo = ttk.Combobox(form_card, textvariable=self.custo_tipo_var,
                                             state="readonly", font=("Helvetica", 10), width=32)
        self.custo_tipo_combo.pack(padx=20, pady=(4, 12), ipady=4)
        self._atualizar_combo_tipos_custo()

        lbl("Data do Custo")
        date_frame_c = tk.Frame(form_card, bg=COLORS["bg_card"])
        date_frame_c.pack(anchor="w", padx=20, pady=(4, 12))
        self.custo_dia, self.custo_mes, self.custo_ano = self._make_date_row(date_frame_c)

        lbl("Descrição (opcional)")
        self.custo_desc_entry = tk.Entry(form_card, font=("Helvetica", 10),
                                         bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                         insertbackground=COLORS["text_primary"],
                                         relief="flat", bd=0,
                                         highlightthickness=1,
                                         highlightbackground=COLORS["border"],
                                         highlightcolor=COLORS["accent"], width=32)
        self.custo_desc_entry.pack(padx=20, pady=(4, 12), ipady=6)

        lbl("Valor (¥)")
        self.custo_valor_entry = tk.Entry(form_card, font=("Helvetica", 11),
                                          bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                          insertbackground=COLORS["text_primary"],
                                          relief="flat", bd=0,
                                          highlightthickness=1,
                                          highlightbackground=COLORS["border"],
                                          highlightcolor=COLORS["accent"], width=22)
        self.custo_valor_entry.pack(padx=20, pady=(4, 12), ipady=7)
        self.custo_valor_entry.bind("<KeyRelease>", lambda e: self._fmt_yen(self.custo_valor_entry))

        self.lbl_custo_status = tk.Label(form_card, text="", font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["red"])
        self.lbl_custo_status.pack(padx=20)

        btn_f = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_f.pack(padx=20, pady=(6, 16), fill="x")
        self.btn_salvar_custo = tk.Button(btn_f, text="  Salvar Custo  ",
                                          font=("Helvetica", 10, "bold"),
                                          bg=COLORS["accent"], fg="white",
                                          relief="flat", cursor="hand2",
                                          activebackground=COLORS["accent2"],
                                          command=self._salvar_custo)
        self.btn_salvar_custo.pack(side="left", ipady=6, ipadx=6)
        tk.Button(btn_f, text="  Limpar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_custo
                  ).pack(side="left", padx=(8, 0), ipady=6, ipadx=6)

        # ── TABELA de Custos ──────────────────────────────────────────────────
        table_card = tk.Frame(container, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True)
        tk.Frame(table_card, bg=COLORS["accent"], height=4).pack(fill="x")

        tbl_h = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_h.pack(fill="x", padx=20, pady=(14, 4))
        tk.Label(tbl_h, text="▦  Custos Registrados", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_h, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_tabela_custos
                  ).pack(side="right", padx=(8,0))
        self.lbl_total_custos = tk.Label(tbl_h, text="0 registros", font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self.lbl_total_custos.pack(side="right")

        # Filtros: por carro e por tipo de custo
        filt_bar = tk.Frame(table_card, bg=COLORS["bg_main"])
        filt_bar.pack(fill="x", padx=20, pady=(0, 4))

        tk.Label(filt_bar, text="🔍 Carro:", font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=6)
        self.custo_filtro_carro_combo = ttk.Combobox(filt_bar, textvariable=self._custo_filtro_carro,
                                                      state="readonly", font=("Helvetica", 9), width=22)
        self.custo_filtro_carro_combo.pack(side="left", padx=(0, 10), ipady=3)
        self.custo_filtro_carro_combo.bind("<ButtonPress>", lambda e: self._atualizar_filtros_custo())

        tk.Label(filt_bar, text="Tipo:", font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4))
        self.custo_filtro_tipo_combo = ttk.Combobox(filt_bar, textvariable=self._custo_filtro_tipo,
                                                     state="readonly", font=("Helvetica", 9), width=16)
        self.custo_filtro_tipo_combo.pack(side="left", ipady=3)

        self._custo_filtro_carro.trace_add("write", lambda *_: self._refresh_tabela_custos())
        self._custo_filtro_tipo.trace_add("write",  lambda *_: self._refresh_tabela_custos())

        tk.Button(filt_bar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=6,
                  command=lambda: [self._custo_filtro_carro.set("Todos"),
                                   self._custo_filtro_tipo.set("Todos"),
                                   self._custo_filtro_mes.set("Todos"),
                                   self._custo_filtro_ano.set("Todos")]
                  ).pack(side="left", padx=(8, 0), ipady=3)

        # Filtro Mês / Ano
        import datetime as _custo_dt
        _anos_c  = ["Todos"] + [str(y) for y in range(_custo_dt.date.today().year, 2019, -1)]
        _meses_c = ["Todos","01","02","03","04","05","06","07","08","09","10","11","12"]
        filt_bar2 = tk.Frame(table_card, bg=COLORS["bg_main"])
        filt_bar2.pack(fill="x", padx=20, pady=(0, 4))
        tk.Label(filt_bar2, text="Mês:", font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4))
        ttk.Combobox(filt_bar2, textvariable=self._custo_filtro_mes,
                     values=_meses_c, state="readonly",
                     font=("Helvetica", 9), width=5
                     ).pack(side="left", padx=(0,8), ipady=3)
        tk.Label(filt_bar2, text="Ano:", font=("Helvetica", 8, "bold"),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4))
        ttk.Combobox(filt_bar2, textvariable=self._custo_filtro_ano,
                     values=_anos_c, state="readonly",
                     font=("Helvetica", 9), width=7
                     ).pack(side="left", padx=(0,8), ipady=3)
        self._custo_filtro_mes.trace_add("write", lambda *_: self._refresh_tabela_custos())
        self._custo_filtro_ano.trace_add("write", lambda *_: self._refresh_tabela_custos())

        tk.Frame(table_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20)

        col_f = tk.Frame(table_card, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=20)
        for txt, w in [("#",3),("Carro",16),("Cor",7),("Placa",12),
                       ("Tipo Custo",14),("Descrição",18),("Valor (¥)",11),("Ações",9)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=3, pady=6)

        scroll_f = tk.Frame(table_card, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=20, pady=(0, 14))
        canvas4 = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb4 = tk.Scrollbar(scroll_f, orient="vertical", command=canvas4.yview)
        self.custos_rows_frame = tk.Frame(canvas4, bg=COLORS["bg_card"])
        self.custos_rows_frame.bind(
            "<Configure>", lambda e: canvas4.configure(scrollregion=canvas4.bbox("all")))
        canvas4.create_window((0, 0), window=self.custos_rows_frame, anchor="nw")
        canvas4.configure(yscrollcommand=sb4.set)
        canvas4.pack(side="left", fill="both", expand=True)
        sb4.pack(side="right", fill="y")
        self._atualizar_filtros_custo()
        self._refresh_tabela_custos()

    # ── Helpers Custos ────────────────────────────────────────────────────────
    def _atualizar_combo_compras(self):
        # Inclui TODAS as compras, inclusive de carros que mudaram de Estoque → Cliente
        compras = [dict(r) for r in self.conn.execute(
            "SELECT co.*, COALESCE(ca.status, '?') as c_status "
            "FROM compras co "
            "LEFT JOIN carros ca ON co.carro_id=ca.id "
            "ORDER BY co.id DESC").fetchall()]
        opts = []
        for c in compras:
            carro_txt = c["carro"]
            if " — " in carro_txt: carro_txt = carro_txt.split(" — ", 1)[1]
            carro_txt = carro_txt.split("|")[0].strip()
            status = c.get("c_status","")
            opts.append(f"#{c['id']}  {carro_txt}  [{status}]")
        self.custo_compra_combo["values"] = opts if opts else ["Nenhuma compra cadastrada"]

    def _atualizar_combo_tipos_custo(self):
        tipos = [r[0] for r in self.conn.execute("SELECT nome FROM tipos_custo ORDER BY nome").fetchall()]
        if hasattr(self, "custo_tipo_combo"):
            self.custo_tipo_combo["values"] = tipos

    def _atualizar_filtros_custo(self):
        carros = ["Todos"] + sorted(set(
            r[0] for r in self.conn.execute("SELECT DISTINCT co.carro FROM custos cu "
                                            "LEFT JOIN compras co ON cu.compra_id=co.id "
                                            "WHERE co.carro IS NOT NULL").fetchall()))
        tipos  = ["Todos"] + sorted(set(
            r[0] for r in self.conn.execute("SELECT DISTINCT tipo_custo FROM custos").fetchall()))
        self.custo_filtro_carro_combo["values"] = carros
        self.custo_filtro_tipo_combo["values"]  = tipos

    def _salvar_custo(self):
        compra_sel = self.custo_compra_var.get().strip()
        tipo       = self.custo_tipo_var.get().strip()
        desc       = self.custo_desc_entry.get().strip()
        valor      = self._yen_raw(self.custo_valor_entry)
        data_custo = self._get_date_from_entries(self.custo_dia, self.custo_mes, self.custo_ano)

        if not compra_sel or compra_sel == "Nenhuma compra cadastrada":
            self.lbl_custo_status.configure(text="⚠ Selecione uma compra.", fg=COLORS["red"]); return
        if not tipo:
            self.lbl_custo_status.configure(text="⚠ Selecione o tipo de custo.", fg=COLORS["red"]); return

        compra_id = int(compra_sel.lstrip("#").split()[0])

        if self._custo_edit_id is not None:
            self.conn.execute(
                "UPDATE custos SET compra_id=?, tipo_custo=?, descricao=?, valor=?, data_custo=? WHERE id=?",
                (compra_id, tipo, desc, valor, data_custo, self._custo_edit_id))
            self.conn.commit()
            self.lbl_custo_status.configure(text="✔ Custo atualizado!", fg=COLORS["green"])
            self._custo_edit_id = None
            self.btn_salvar_custo.configure(text="  Salvar Custo  ")
        else:
            self.conn.execute(
                "INSERT INTO custos (compra_id, tipo_custo, descricao, valor, data_custo) VALUES (?,?,?,?,?)",
                (compra_id, tipo, desc, valor, data_custo))
            self.conn.commit()
            self.lbl_custo_status.configure(text="✔ Custo registrado!", fg=COLORS["green"])

        self._limpar_form_custo(clear_status=False)
        self._atualizar_filtros_custo()
        self._refresh_tabela_custos()
        try:
            self._refresh_tabela_compras()
        except Exception:
            pass

    def _limpar_form_custo(self, clear_status=True):
        import datetime
        self.custo_compra_var.set("")
        self.custo_tipo_var.set("")
        self.custo_desc_entry.delete(0, tk.END)
        self.custo_valor_entry.delete(0, tk.END)
        self._custo_edit_id = None
        self.btn_salvar_custo.configure(text="  Salvar Custo  ")
        hoje = datetime.date.today()
        for e, val in [(self.custo_dia, str(hoje.day).zfill(2)),
                       (self.custo_mes, str(hoje.month).zfill(2)),
                       (self.custo_ano, str(hoje.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        if clear_status:
            self.lbl_custo_status.configure(text="")

    def _editar_custo(self, cid):
        row = self.conn.execute("SELECT * FROM custos WHERE id=?", (cid,)).fetchone()
        if not row: return
        c = dict(row)
        self._atualizar_combo_compras()
        compra = self.conn.execute("SELECT * FROM compras WHERE id=?", (c["compra_id"],)).fetchone()
        if compra:
            co = dict(compra)
            sel = f"#{co['id']}  {co['carro']}"
            self.custo_compra_var.set(sel)
        self.custo_tipo_var.set(c["tipo_custo"])
        self.custo_desc_entry.delete(0, tk.END)
        self.custo_desc_entry.insert(0, c.get("descricao") or "")
        self.custo_valor_entry.delete(0, tk.END)
        if c["valor"]:
            try: self.custo_valor_entry.insert(0, f"{int(c['valor']):,}")
            except: self.custo_valor_entry.insert(0, c["valor"])
        # Data do custo
        data = c.get("data_custo") or ""
        partes = data.split("/")
        import datetime; hoje = datetime.date.today()
        dd = partes[0] if len(partes) > 0 and partes[0] else str(hoje.day).zfill(2)
        mm = partes[1] if len(partes) > 1 and partes[1] else str(hoje.month).zfill(2)
        aa = partes[2] if len(partes) > 2 and partes[2] else str(hoje.year)
        for e, val in [(self.custo_dia,dd),(self.custo_mes,mm),(self.custo_ano,aa)]:
            e.delete(0,tk.END); e.insert(0,val); e.configure(fg=COLORS["text_primary"])
        self._custo_edit_id = cid
        self.btn_salvar_custo.configure(text="  Atualizar Custo  ")
        self.lbl_custo_status.configure(text="")

    def _excluir_custo(self, cid):
        if not msgbox.askyesno("Confirmar", "Deseja excluir este custo?"):
            return
        self.conn.execute("DELETE FROM custos WHERE id=?", (cid,))
        self.conn.commit()
        self._atualizar_filtros_custo()
        self._refresh_tabela_custos()
        try:
            self._refresh_tabela_compras()
        except Exception:
            pass

    def _refresh_tabela_custos(self):
        for w in self.custos_rows_frame.winfo_children():
            w.destroy()

        filtro_carro = getattr(self, "_custo_filtro_carro", None)
        filtro_tipo  = getattr(self, "_custo_filtro_tipo",  None)
        fc = filtro_carro.get() if filtro_carro else "Todos"
        ft = filtro_tipo.get()  if filtro_tipo  else "Todos"

        rows = [dict(r) for r in self.conn.execute(
            "SELECT cu.*, co.carro, co.cor, co.placa FROM custos cu "
            "LEFT JOIN compras co ON cu.compra_id=co.id ORDER BY cu.id DESC"
        ).fetchall()]

        fmes_c = getattr(self, "_custo_filtro_mes", None)
        fano_c = getattr(self, "_custo_filtro_ano", None)
        fmes_c = fmes_c.get() if fmes_c else "Todos"
        fano_c = fano_c.get() if fano_c else "Todos"

        if fc != "Todos":
            rows = [r for r in rows if (r.get("carro") or "") == fc]
        if ft != "Todos":
            rows = [r for r in rows if r.get("tipo_custo") == ft]
        if fmes_c != "Todos" or fano_c != "Todos":
            def _match_data_custo(r):
                d = r.get("data_custo") or ""
                pts = d.split("/")
                rm = pts[1] if len(pts) > 1 else ""
                ra = pts[2] if len(pts) > 2 else ""
                if fmes_c != "Todos" and rm != fmes_c: return False
                if fano_c != "Todos" and ra != fano_c: return False
                return True
            rows = [r for r in rows if _match_data_custo(r)]

        self.lbl_total_custos.configure(text=f"{len(rows)} registro(s)")
        if not rows:
            tk.Label(self.custos_rows_frame, text="Nenhum custo registrado.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(pady=30)
            return

        for i, c in enumerate(rows):
            row_bg = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self.custos_rows_frame, bg=row_bg)
            row.pack(fill="x")
            tk.Label(row, text=str(c["id"]), font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_muted"],
                     width=3, anchor="w").pack(side="left", padx=3, pady=7)
            tk.Label(row, text=c.get("carro") or "—", font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_primary"],
                     width=16, anchor="w").pack(side="left", padx=3)
            tk.Label(row, text=c.get("cor")   or "—", font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_secondary"],
                     width=7, anchor="w").pack(side="left", padx=3)
            tk.Label(row, text=c.get("placa") or "—", font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_secondary"],
                     width=8, anchor="w").pack(side="left", padx=3)
            tk.Label(row, text=c["tipo_custo"], font=("Helvetica", 8, "bold"),
                     bg=COLORS["blue"], fg="white", width=14,
                     anchor="center").pack(side="left", padx=3, pady=4)
            tk.Label(row, text=c.get("descricao") or "—", font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_secondary"],
                     width=18, anchor="w").pack(side="left", padx=3)
            tk.Label(row, text=self._fmt_yen_display(c["valor"]),
                     font=("Helvetica", 9), bg=row_bg, fg=COLORS["text_primary"],
                     width=11, anchor="w").pack(side="left", padx=3)
            acts = tk.Frame(row, bg=row_bg)
            acts.pack(side="left", padx=4)
            tk.Button(acts, text="✏", font=("Helvetica", 9),
                      bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                      padx=5, pady=2,
                      command=lambda cid=c["id"]: self._editar_custo(cid)
                      ).pack(side="left", padx=(0, 3))
            tk.Button(acts, text="✕", font=("Helvetica", 9),
                      bg=COLORS["red"], fg="white", relief="flat", cursor="hand2",
                      padx=5, pady=2,
                      command=lambda cid=c["id"]: self._excluir_custo(cid)
                      ).pack(side="left")

    # ── Cadastro de Tipos de Custo (submenu Cadastros) ────────────────────────
    def _build_tipos_custo_cadastro(self, parent):
        container = tk.Frame(parent, bg=COLORS["bg_content"])
        container.pack(fill="both", expand=True, padx=24, pady=20)

        # ── CARD FORMULÁRIO ───────────────────────────────────────────────────
        form_card = tk.Frame(container, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        form_card.pack(side="left", fill="y", padx=(0, 16), ipadx=10)
        tk.Frame(form_card, bg=COLORS["blue"], height=4).pack(fill="x")
        tk.Label(form_card, text="⊕  Tipos de Custo", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=20, pady=(14, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=(0, 12))

        tk.Label(form_card, text="Nome do Tipo", font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=20)
        self.tc_nome_entry = tk.Entry(form_card, font=("Helvetica", 11),
                                      bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                                      insertbackground=COLORS["text_primary"],
                                      relief="flat", bd=0,
                                      highlightthickness=1,
                                      highlightbackground=COLORS["border"],
                                      highlightcolor=COLORS["blue"], width=26)
        self.tc_nome_entry.pack(padx=20, pady=(4, 14), ipady=7)

        self.lbl_tc_status = tk.Label(form_card, text="", font=("Helvetica", 8),
                                      bg=COLORS["bg_card"], fg=COLORS["red"])
        self.lbl_tc_status.pack(padx=20)

        btn_f = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_f.pack(padx=20, pady=(6, 20), fill="x")
        self.btn_salvar_tc = tk.Button(btn_f, text="  Salvar  ",
                                       font=("Helvetica", 10, "bold"),
                                       bg=COLORS["blue"], fg="white",
                                       relief="flat", cursor="hand2",
                                       command=self._salvar_tipo_custo)
        self.btn_salvar_tc.pack(side="left", ipady=6, ipadx=6)
        tk.Button(btn_f, text="  Limpar  ", font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_tc
                  ).pack(side="left", padx=(8, 0), ipady=6, ipadx=6)
        self._tc_edit_nome = None

        # ── TABELA DE TIPOS ───────────────────────────────────────────────────
        table_card = tk.Frame(container, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True)
        tk.Frame(table_card, bg=COLORS["blue"], height=4).pack(fill="x")

        tbl_h = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_h.pack(fill="x", padx=20, pady=(14, 4))
        tk.Label(tbl_h, text="▦  Tipos Cadastrados", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_h, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_tabela_tc).pack(side="right", padx=(8,0))
        self.lbl_total_tc = tk.Label(tbl_h, text="", font=("Helvetica", 8),
                                     bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self.lbl_total_tc.pack(side="right")

        tk.Frame(table_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20)
        col_f = tk.Frame(table_card, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=20)
        for txt, w in [("#",4),("Nome do Tipo",30),("Em Uso",8),("Ações",10)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=4, pady=6)

        scroll_f = tk.Frame(table_card, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=20, pady=(0, 14))
        canvas5 = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb5 = tk.Scrollbar(scroll_f, orient="vertical", command=canvas5.yview)
        self.tc_rows_frame = tk.Frame(canvas5, bg=COLORS["bg_card"])
        self.tc_rows_frame.bind(
            "<Configure>", lambda e: canvas5.configure(scrollregion=canvas5.bbox("all")))
        canvas5.create_window((0, 0), window=self.tc_rows_frame, anchor="nw")
        canvas5.configure(yscrollcommand=sb5.set)
        canvas5.pack(side="left", fill="both", expand=True)
        sb5.pack(side="right", fill="y")
        self._refresh_tabela_tc()

    def _tipo_em_uso(self, nome):
        """Retorna True se o tipo de custo já foi usado em algum custo."""
        count = self.conn.execute(
            "SELECT COUNT(*) FROM custos WHERE tipo_custo=?", (nome,)).fetchone()[0]
        return count > 0

    def _salvar_tipo_custo(self):
        nome = self.tc_nome_entry.get().strip()
        if not nome:
            self.lbl_tc_status.configure(text="⚠ Informe o nome.", fg=COLORS["red"]); return

        if self._tc_edit_nome is not None:
            # Verifica se o nome antigo está em uso
            if self._tipo_em_uso(self._tc_edit_nome):
                self.lbl_tc_status.configure(
                    text="⚠ Tipo em uso — não pode ser editado.", fg=COLORS["red"]); return
            try:
                self.conn.execute("UPDATE tipos_custo SET nome=? WHERE nome=?",
                                  (nome, self._tc_edit_nome))
                self.conn.commit()
                self.lbl_tc_status.configure(text="✔ Tipo atualizado!", fg=COLORS["green"])
            except Exception:
                self.lbl_tc_status.configure(text="⚠ Nome já existe.", fg=COLORS["red"]); return
            self._tc_edit_nome = None
            self.btn_salvar_tc.configure(text="  Salvar  ")
        else:
            try:
                self.conn.execute("INSERT INTO tipos_custo (nome) VALUES (?)", (nome,))
                self.conn.commit()
                self.lbl_tc_status.configure(text="✔ Tipo adicionado!", fg=COLORS["green"])
            except Exception:
                self.lbl_tc_status.configure(text="⚠ Tipo já existe.", fg=COLORS["red"]); return

        self._limpar_form_tc(clear_status=False)
        self._atualizar_combo_tipos_custo()
        self._refresh_tabela_tc()

    def _limpar_form_tc(self, clear_status=True):
        self.tc_nome_entry.delete(0, tk.END)
        self._tc_edit_nome = None
        self.btn_salvar_tc.configure(text="  Salvar  ")
        if clear_status:
            self.lbl_tc_status.configure(text="")

    def _editar_tipo_custo(self, nome):
        if self._tipo_em_uso(nome):
            msgbox.showwarning("Em Uso", f"'{nome}' já foi usado em custos e não pode ser editado.")
            return
        self._tc_edit_nome = nome
        self.tc_nome_entry.delete(0, tk.END)
        self.tc_nome_entry.insert(0, nome)
        self.btn_salvar_tc.configure(text="  Atualizar  ")
        self.lbl_tc_status.configure(text="")

    def _excluir_tipo_custo(self, nome):
        if self._tipo_em_uso(nome):
            msgbox.showwarning("Em Uso", f"'{nome}' já foi usado em custos e não pode ser excluído.")
            return
        if not msgbox.askyesno("Confirmar", f"Excluir tipo '{nome}'?"):
            return
        self.conn.execute("DELETE FROM tipos_custo WHERE nome=?", (nome,))
        self.conn.commit()
        self._atualizar_combo_tipos_custo()
        self._refresh_tabela_tc()

    def _refresh_tabela_tc(self):
        for w in self.tc_rows_frame.winfo_children():
            w.destroy()
        tipos = [r[0] for r in self.conn.execute(
            "SELECT nome FROM tipos_custo ORDER BY nome").fetchall()]
        self.lbl_total_tc.configure(text=f"{len(tipos)} tipo(s)")
        for i, t in enumerate(tipos):
            row_bg = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self.tc_rows_frame, bg=row_bg)
            row.pack(fill="x")
            tk.Label(row, text=str(i+1), font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_muted"],
                     width=4, anchor="w").pack(side="left", padx=4, pady=7)
            tk.Label(row, text=t, font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_primary"],
                     width=30, anchor="w").pack(side="left", padx=4)
            em_uso = self._tipo_em_uso(t)
            uso_lbl = tk.Label(row, text="● sim" if em_uso else "○ não",
                               font=("Helvetica", 8, "bold"),
                               bg=row_bg,
                               fg=COLORS["green"] if em_uso else COLORS["text_muted"],
                               width=8, anchor="w")
            uso_lbl.pack(side="left", padx=4)
            acts = tk.Frame(row, bg=row_bg)
            acts.pack(side="left", padx=4)
            e_bg = COLORS["border"] if em_uso else COLORS["blue"]
            e_fg = COLORS["text_muted"] if em_uso else "white"
            tk.Button(acts, text="✏", font=("Helvetica", 9),
                      bg=e_bg, fg=e_fg, relief="flat",
                      cursor="hand2" if not em_uso else "arrow",
                      padx=5, pady=2,
                      command=lambda n=t: self._editar_tipo_custo(n)
                      ).pack(side="left", padx=(0, 3))
            tk.Button(acts, text="✕", font=("Helvetica", 9),
                      bg=COLORS["border"] if em_uso else COLORS["red"],
                      fg=COLORS["text_muted"] if em_uso else "white",
                      relief="flat",
                      cursor="hand2" if not em_uso else "arrow",
                      padx=5, pady=2,
                      command=lambda n=t: self._excluir_tipo_custo(n)
                      ).pack(side="left")

    # ── Cadastro de Categorias de Despesas Fixas ──────────────────────────────
    def _build_categorias_despesas_cadastro(self, parent):
        container = tk.Frame(parent, bg=COLORS["bg_content"])
        container.pack(fill="both", expand=True, padx=24, pady=20)

        # ── CARD FORMULÁRIO ───────────────────────────────────────────────────
        form_card = tk.Frame(container, bg=COLORS["bg_card"],
                             highlightthickness=1, highlightbackground=COLORS["border"])
        form_card.pack(side="left", fill="y", padx=(0, 16), ipadx=10)
        tk.Frame(form_card, bg=COLORS["orange"], height=4).pack(fill="x")
        tk.Label(form_card, text="💸  Categoria Desp. Fixas",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                 ).pack(anchor="w", padx=20, pady=(14, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1
                 ).pack(fill="x", padx=20, pady=(0, 12))

        tk.Label(form_card, text="Nome da Categoria",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]
                 ).pack(anchor="w", padx=20)
        self._catdf_nome_entry = tk.Entry(
            form_card, font=("Helvetica", 11),
            bg=COLORS["bg_main"], fg=COLORS["text_primary"],
            insertbackground=COLORS["text_primary"],
            relief="flat", bd=0,
            highlightthickness=1,
            highlightbackground=COLORS["border"],
            highlightcolor=COLORS["orange"], width=26)
        self._catdf_nome_entry.pack(padx=20, pady=(4, 14), ipady=7)
        self._catdf_nome_entry.bind("<Return>", lambda e: self._salvar_categoria_despesa())

        self._lbl_catdf_status = tk.Label(form_card, text="",
                                          font=("Helvetica", 8),
                                          bg=COLORS["bg_card"], fg=COLORS["red"])
        self._lbl_catdf_status.pack(padx=20)

        btn_f = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_f.pack(padx=20, pady=(6, 20), fill="x")
        self._btn_salvar_catdf = tk.Button(
            btn_f, text="  Salvar  ",
            font=("Helvetica", 10, "bold"),
            bg=COLORS["orange"], fg="white",
            relief="flat", cursor="hand2",
            command=self._salvar_categoria_despesa)
        self._btn_salvar_catdf.pack(side="left", ipady=6, ipadx=6)
        tk.Button(btn_f, text="  Limpar  ",
                  font=("Helvetica", 10),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_catdf
                  ).pack(side="left", padx=(8, 0), ipady=6, ipadx=6)
        self._catdf_edit_nome = None

        # ── TABELA ────────────────────────────────────────────────────────────
        table_card = tk.Frame(container, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True)
        tk.Frame(table_card, bg=COLORS["orange"], height=4).pack(fill="x")

        tbl_h = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_h.pack(fill="x", padx=20, pady=(14, 4))
        tk.Label(tbl_h, text="▦  Categorias Cadastradas",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_h, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_tabela_catdf).pack(side="right", padx=(8,0))
        self._lbl_total_catdf = tk.Label(tbl_h, text="",
                                         font=("Helvetica", 8),
                                         bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self._lbl_total_catdf.pack(side="right")

        tk.Frame(table_card, bg=COLORS["border"], height=1).pack(fill="x", padx=20)
        col_f = tk.Frame(table_card, bg=COLORS["bg_main"])
        col_f.pack(fill="x", padx=20)
        for txt, w in [("#", 4), ("Nome da Categoria", 30), ("Em Uso", 8), ("Ações", 10)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=4, pady=6)

        scroll_f = tk.Frame(table_card, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True, padx=20, pady=(0, 14))
        canvas_cat = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb_cat = tk.Scrollbar(scroll_f, orient="vertical", command=canvas_cat.yview)
        self._catdf_rows_frame = tk.Frame(canvas_cat, bg=COLORS["bg_card"])
        self._catdf_rows_frame.bind(
            "<Configure>",
            lambda e: canvas_cat.configure(scrollregion=canvas_cat.bbox("all")))
        canvas_cat.create_window((0, 0), window=self._catdf_rows_frame, anchor="nw")
        canvas_cat.configure(yscrollcommand=sb_cat.set)
        canvas_cat.pack(side="left", fill="both", expand=True)
        sb_cat.pack(side="right", fill="y")
        self._refresh_tabela_catdf()

    def _catdf_em_uso(self, nome):
        count = self.conn.execute(
            "SELECT COUNT(*) FROM despesas_fixas WHERE categoria=?", (nome,)).fetchone()[0]
        return count > 0

    def _salvar_categoria_despesa(self):
        nome = self._catdf_nome_entry.get().strip()
        if not nome:
            self._lbl_catdf_status.configure(
                text="⚠ Informe o nome.", fg=COLORS["red"]); return

        if self._catdf_edit_nome is not None:
            if self._catdf_em_uso(self._catdf_edit_nome):
                self._lbl_catdf_status.configure(
                    text="⚠ Categoria em uso — não pode ser editada.", fg=COLORS["red"]); return
            try:
                self.conn.execute("UPDATE categorias_despesas SET nome=? WHERE nome=?",
                                  (nome, self._catdf_edit_nome))
                # Atualiza referências nas despesas existentes
                self.conn.execute("UPDATE despesas_fixas SET categoria=? WHERE categoria=?",
                                  (nome, self._catdf_edit_nome))
                self.conn.commit()
                self._lbl_catdf_status.configure(
                    text="✔ Categoria atualizada!", fg=COLORS["green"])
            except Exception:
                self._lbl_catdf_status.configure(
                    text="⚠ Nome já existe.", fg=COLORS["red"]); return
            self._catdf_edit_nome = None
            self._btn_salvar_catdf.configure(text="  Salvar  ")
        else:
            try:
                self.conn.execute(
                    "INSERT INTO categorias_despesas (nome) VALUES (?)", (nome,))
                self.conn.commit()
                self._lbl_catdf_status.configure(
                    text="✔ Categoria adicionada!", fg=COLORS["green"])
            except Exception:
                self._lbl_catdf_status.configure(
                    text="⚠ Categoria já existe.", fg=COLORS["red"]); return

        self._limpar_form_catdf(clear_status=False)
        self._refresh_tabela_catdf()

    def _limpar_form_catdf(self, clear_status=True):
        self._catdf_nome_entry.delete(0, tk.END)
        self._catdf_edit_nome = None
        self._btn_salvar_catdf.configure(text="  Salvar  ")
        if clear_status:
            self._lbl_catdf_status.configure(text="")

    def _editar_categoria_despesa(self, nome):
        if self._catdf_em_uso(nome):
            msgbox.showwarning("Em Uso",
                f"'{nome}' já está em uso em despesas e não pode ser editada.")
            return
        self._catdf_edit_nome = nome
        self._catdf_nome_entry.delete(0, tk.END)
        self._catdf_nome_entry.insert(0, nome)
        self._btn_salvar_catdf.configure(text="  Atualizar  ")
        self._lbl_catdf_status.configure(text="")

    def _excluir_categoria_despesa(self, nome):
        if self._catdf_em_uso(nome):
            msgbox.showwarning("Em Uso",
                f"'{nome}' já está em uso em despesas e não pode ser excluída.")
            return
        if not msgbox.askyesno("Confirmar", f"Excluir categoria '{nome}'?"):
            return
        self.conn.execute("DELETE FROM categorias_despesas WHERE nome=?", (nome,))
        self.conn.commit()
        self._refresh_tabela_catdf()

    def _refresh_tabela_catdf(self):
        for w in self._catdf_rows_frame.winfo_children():
            w.destroy()
        cats = [r[0] for r in self.conn.execute(
            "SELECT nome FROM categorias_despesas ORDER BY nome").fetchall()]
        self._lbl_total_catdf.configure(text=f"{len(cats)} categoria(s)")
        for i, cat in enumerate(cats):
            row_bg = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._catdf_rows_frame, bg=row_bg)
            row.pack(fill="x")
            tk.Label(row, text=str(i + 1), font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_muted"],
                     width=4, anchor="w").pack(side="left", padx=4, pady=7)
            tk.Label(row, text=cat, font=("Helvetica", 9),
                     bg=row_bg, fg=COLORS["text_primary"],
                     width=30, anchor="w").pack(side="left", padx=4)
            em_uso = self._catdf_em_uso(cat)
            tk.Label(row, text="● sim" if em_uso else "○ não",
                     font=("Helvetica", 8, "bold"),
                     bg=row_bg,
                     fg=COLORS["green"] if em_uso else COLORS["text_muted"],
                     width=8, anchor="w").pack(side="left", padx=4)
            acts = tk.Frame(row, bg=row_bg)
            acts.pack(side="left", padx=4)
            e_bg = COLORS["border"] if em_uso else COLORS["orange"]
            e_fg = COLORS["text_muted"] if em_uso else "white"
            tk.Button(acts, text="✏", font=("Helvetica", 9),
                      bg=e_bg, fg=e_fg, relief="flat",
                      cursor="hand2" if not em_uso else "arrow",
                      padx=5, pady=2,
                      command=lambda n=cat: self._editar_categoria_despesa(n)
                      ).pack(side="left", padx=(0, 3))
            tk.Button(acts, text="✕", font=("Helvetica", 9),
                      bg=COLORS["border"] if em_uso else COLORS["red"],
                      fg=COLORS["text_muted"] if em_uso else "white",
                      relief="flat",
                      cursor="hand2" if not em_uso else "arrow",
                      padx=5, pady=2,
                      command=lambda n=cat: self._excluir_categoria_despesa(n)
                      ).pack(side="left")

    def _build_nova_compra(self, parent):
        self._compra_edit_id    = None
        self._compra_detalhe_id = None
        TIPO_OPTS = ["Leilão", "Troca", "Compra Direta"]
        self.compras_data = [dict(r) for r in
                             self.conn.execute("SELECT * FROM compras ORDER BY id DESC").fetchall()]
        self._compra_filtro_var = tk.StringVar(value="")

        # Layout 3 colunas fixas — ocupa toda a altura disponível
        # Coluna esquerda: formulário (largura fixa, scroll interno)
        # Coluna centro:   histórico de compras (expand)
        # Coluna direita:  custos do carro selecionado (mesma largura do form, sidebar)

        root_frame = tk.Frame(parent, bg=COLORS["bg_content"])
        root_frame.pack(fill="both", expand=True, padx=14, pady=14)

        # ── COLUNA ESQUERDA: formulário com scroll interno ────────────────────
        form_outer = tk.Frame(root_frame, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              width=270)
        form_outer.pack(side="left", fill="y", padx=(0, 8))
        form_outer.pack_propagate(False)

        # Canvas de scroll — ocupa toda a coluna
        form_canvas = tk.Canvas(form_outer, bg=COLORS["bg_card"], highlightthickness=0, width=268)
        form_sb = tk.Scrollbar(form_outer, orient="vertical", command=form_canvas.yview)
        form_canvas.configure(yscrollcommand=form_sb.set)
        form_sb.pack(side="right", fill="y")
        form_canvas.pack(side="left", fill="both", expand=True)

        form_card = tk.Frame(form_canvas, bg=COLORS["bg_card"])
        form_canvas.create_window((0, 0), window=form_card, anchor="nw")
        form_card.bind("<Configure>",
            lambda e: form_canvas.configure(scrollregion=form_canvas.bbox("all")))

        def _scroll_form(ev):
            form_canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        form_outer.bind("<Enter>",  lambda e: form_outer.bind_all("<MouseWheel>", _scroll_form))
        form_outer.bind("<Leave>",  lambda e: form_outer.unbind_all("<MouseWheel>"))

        tk.Frame(form_card, bg=COLORS["accent"], height=4).pack(fill="x")
        self._lbl_form_compra_title = tk.Label(
            form_card, text="⊕  Nova Compra", font=("Helvetica", 11, "bold"),
            bg=COLORS["bg_card"], fg=COLORS["text_primary"])
        self._lbl_form_compra_title.pack(anchor="w", padx=18, pady=(14, 4))
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 12))

        def lbl(txt, required=False):
            fg = COLORS["accent"] if required else COLORS["text_secondary"]
            tk.Label(form_card, text=txt + ("  ★" if required else ""),
                     font=("Helvetica", 9, "bold"), bg=COLORS["bg_card"], fg=fg
                     ).pack(anchor="w", padx=18)

        lbl("Carro", required=True)
        self.compra_carro_var = tk.StringVar(value="")
        self.compra_carro_combo = ttk.Combobox(form_card, textvariable=self.compra_carro_var,
                                               state="readonly", font=("Helvetica", 10), width=28)
        self.compra_carro_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self.compra_carro_combo.bind("<ButtonPress>", lambda e: self._atualizar_combo_carros())

        lbl("Tipo de Compra", required=True)
        self.compra_tipo_var = tk.StringVar(value="")
        self._compra_tipo_combo = ttk.Combobox(form_card, textvariable=self.compra_tipo_var,
                     values=TIPO_OPTS, state="readonly", font=("Helvetica", 10), width=28)
        self._compra_tipo_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self.compra_tipo_var.trace_add("write", lambda *_: self._toggle_taxa_leilao())

        # Cliente — obrigatório para Compra Direta e Troca
        self._compra_cliente_lbl_req = tk.Label(
            form_card, text="Cliente  ★",
            font=("Helvetica", 9, "bold"),
            bg=COLORS["bg_card"], fg=COLORS["accent"])
        self._compra_cliente_lbl_req.pack(anchor="w", padx=18, pady=(4,0))
        self.compra_cliente_var = tk.StringVar(value="")
        self.compra_cliente_combo = ttk.Combobox(
            form_card, textvariable=self.compra_cliente_var,
            state="readonly", font=("Helvetica", 10), width=28)
        self.compra_cliente_combo.pack(padx=18, pady=(4, 12), ipady=4)
        self.compra_cliente_combo.bind("<ButtonPress>",
            lambda e: self._atualizar_combo_clientes_compra())
        # Atualiza visibilidade quando tipo muda
        self.compra_tipo_var.trace_add("write", lambda *_: self._toggle_compra_cliente())

        lbl("Data de Entrada")
        date_frame = tk.Frame(form_card, bg=COLORS["bg_card"])
        date_frame.pack(anchor="w", padx=18, pady=(4, 12))
        self.compra_dia, self.compra_mes, self.compra_ano = self._make_date_row(date_frame)

        lbl("Valor de Compra (¥)")
        self.compra_valor_entry = self._make_yen_entry(form_card, width=20)
        self.compra_valor_entry.pack(padx=18, pady=(4, 12), ipady=7)

        # Bloco de taxas — só aparece quando tipo = Leilão
        self._taxa_leilao_frame = tk.Frame(form_card, bg=COLORS["bg_card"],
                                           highlightthickness=1,
                                           highlightbackground=COLORS["orange"])
        # cabeçalho do bloco
        bloco_hdr = tk.Frame(self._taxa_leilao_frame, bg=COLORS["orange"])
        bloco_hdr.pack(fill="x")
        tk.Label(bloco_hdr, text="  Taxas de Leilão",
                 font=("Helvetica", 8, "bold"),
                 bg=COLORS["orange"], fg="white").pack(side="left", pady=4)

        tk.Label(self._taxa_leilao_frame, text="Taxa de Leilão (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=14, pady=(8, 0))
        self.compra_taxa_leilao_entry = self._make_yen_entry(self._taxa_leilao_frame, width=20)
        self.compra_taxa_leilao_entry.pack(padx=14, pady=(4, 8), ipady=7)

        tk.Label(self._taxa_leilao_frame, text="Taxa de Reciclagem (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=14)
        self.compra_taxa_rec_entry = self._make_yen_entry(self._taxa_leilao_frame, width=20)
        self.compra_taxa_rec_entry.pack(padx=14, pady=(4, 12), ipady=7)
        # começa oculto

        self.lbl_compra_status = tk.Label(form_card, text="", font=("Helvetica", 8),
                                          bg=COLORS["bg_card"], fg=COLORS["red"])
        self.lbl_compra_status.pack(padx=18, pady=(4, 8))

        # Botões dentro do form_card — abaixo do bloco de taxas
        tk.Frame(form_card, bg=COLORS["border"], height=1).pack(fill="x", padx=18, pady=(0, 10))
        btn_f = tk.Frame(form_card, bg=COLORS["bg_card"])
        btn_f.pack(padx=18, pady=(0, 20), fill="x")
        self.btn_salvar_compra = tk.Button(
            btn_f, text="  Registrar Compra  ", font=("Helvetica", 10, "bold"),
            bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
            activebackground=COLORS["accent2"], activeforeground="white",
            command=self._salvar_compra)
        self.btn_salvar_compra.pack(side="left", ipady=7, ipadx=4, fill="x", expand=True)
        tk.Button(btn_f, text="Limpar", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._limpar_form_compra
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=6)
        tk.Button(btn_f, text="↺", font=("Helvetica", 10, "bold"),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_tabela_compras
                  ).pack(side="left", padx=(8, 0), ipady=7, ipadx=6)

        # ── COLUNA DIREITA: sidebar de custos (pack ANTES do centro para reservar espaço) ──
        self._detalhe_panel = tk.Frame(root_frame, bg=COLORS["bg_card"],
                                       highlightthickness=1, highlightbackground=COLORS["border"],
                                       width=270)
        self._detalhe_panel.pack(side="right", fill="y")
        self._detalhe_panel.pack_propagate(False)
        self._detalhe_placeholder()

        # ── COLUNA CENTRO: histórico de compras — expand entre as duas colunas ──
        table_card = tk.Frame(root_frame, bg=COLORS["bg_card"],
                              highlightthickness=1, highlightbackground=COLORS["border"])
        table_card.pack(side="left", fill="both", expand=True, padx=8)
        tk.Frame(table_card, bg=COLORS["accent"], height=4).pack(fill="x")

        tbl_h = tk.Frame(table_card, bg=COLORS["bg_card"])
        tbl_h.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(tbl_h, text="⊕  Histórico de Compras", font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(tbl_h, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=6, pady=2,
                  command=self._refresh_tabela_compras
                  ).pack(side="right", padx=(0,8))
        self.lbl_total_compras = tk.Label(tbl_h, text="0 registros", font=("Helvetica", 8),
                                          bg=COLORS["bg_card"], fg=COLORS["text_muted"])
        self.lbl_total_compras.pack(side="right")

        filt_bar = tk.Frame(table_card, bg=COLORS["bg_main"])
        filt_bar.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(filt_bar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        self.compra_filtro_entry = tk.Entry(
            filt_bar, textvariable=self._compra_filtro_var, font=("Helvetica", 9),
            bg=COLORS["bg_card"], fg=COLORS["text_primary"],
            insertbackground=COLORS["text_primary"], relief="flat", bd=0,
            highlightthickness=1, highlightbackground=COLORS["border"],
            highlightcolor=COLORS["accent"], width=24)
        self.compra_filtro_entry.pack(side="left", ipady=4)
        self._compra_filtro_var.trace_add("write", lambda *_: self._refresh_tabela_compras())
        tk.Button(filt_bar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=5,
                  command=lambda: self._compra_filtro_var.set("")
                  ).pack(side="left", padx=(4, 0), ipady=3)

        # ── Tabela com canvas dual (header + body) para alinhamento pixel-perfect ──
        COMP_COLS = [
            ("#", 35), ("Data", 80), ("Carro", 120), ("Cor", 55), ("Placa", 65),
            ("Cliente", 100), ("Tipo", 75), ("V.Compra", 82), ("Custo", 82),
            ("Total", 86), ("Shaken", 86), ("St.SK", 65), ("Ações", 120),
        ]
        self._comp_cols = COMP_COLS
        COMP_TOTAL_W = sum(c[1] for c in COMP_COLS)

        comp_table_area = tk.Frame(table_card, bg=COLORS["bg_card"])
        comp_table_area.pack(fill="both", expand=True, padx=14, pady=(0, 12))

        vsb_comp = tk.Scrollbar(comp_table_area, orient="vertical")
        hsb_comp = tk.Scrollbar(comp_table_area, orient="horizontal")
        self._comp_hdr_canvas  = tk.Canvas(comp_table_area, bg=COLORS["bg_main"],
                                            height=28, highlightthickness=0)
        self._comp_body_canvas = tk.Canvas(comp_table_area, bg=COLORS["bg_card"],
                                            highlightthickness=0,
                                            yscrollcommand=vsb_comp.set)

        def _comp_sync_x(*a):
            self._comp_hdr_canvas.xview(*a)
            self._comp_body_canvas.xview(*a)
        hsb_comp.config(command=_comp_sync_x)
        vsb_comp.config(command=self._comp_body_canvas.yview)

        self._comp_hdr_canvas.grid(row=0, column=0, sticky="ew")
        self._comp_body_canvas.grid(row=1, column=0, sticky="nsew")
        vsb_comp.grid(row=1, column=1, sticky="ns")
        hsb_comp.grid(row=2, column=0, sticky="ew")
        comp_table_area.grid_rowconfigure(1, weight=1)
        comp_table_area.grid_columnconfigure(0, weight=1)

        comp_hdr_frame = tk.Frame(self._comp_hdr_canvas, bg=COLORS["bg_main"])
        comp_hdr_win = self._comp_hdr_canvas.create_window(
            (0, 0), window=comp_hdr_frame, anchor="nw")
        for col_name, col_w in COMP_COLS:
            f = tk.Frame(comp_hdr_frame, bg=COLORS["bg_main"], width=col_w, height=28)
            f.pack_propagate(False); f.pack(side="left")
            tk.Label(f, text=col_name, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     anchor="w").pack(fill="both", padx=4)
        tk.Frame(comp_hdr_frame, bg=COLORS["border"], height=1,
                 width=COMP_TOTAL_W).pack(fill="x")

        self.compras_rows_frame = tk.Frame(self._comp_body_canvas, bg=COLORS["bg_card"])
        comp_body_win = self._comp_body_canvas.create_window(
            (0, 0), window=self.compras_rows_frame, anchor="nw")
        self.compras_rows_frame.bind("<Configure>",
            lambda e: self._comp_body_canvas.configure(
                scrollregion=self._comp_body_canvas.bbox("all")))

        def _comp_canvas_resize(e, _bw=comp_body_win, _hw=comp_hdr_win):
            w = max(e.width, COMP_TOTAL_W)
            self._comp_body_canvas.itemconfig(_bw, width=w)
            self._comp_hdr_canvas.itemconfig(_hw, width=w)
            self._comp_hdr_canvas.configure(scrollregion=(0, 0, w, 28))
        self._comp_body_canvas.bind("<Configure>", _comp_canvas_resize)

        def _comp_xmove(first, last):
            hsb_comp.set(first, last)
            self._comp_hdr_canvas.xview_moveto(first)
        self._comp_body_canvas.configure(xscrollcommand=_comp_xmove)

        def _scroll_table(ev):
            self._comp_body_canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        table_card.bind("<Enter>", lambda e: table_card.bind_all("<MouseWheel>", _scroll_table))
        table_card.bind("<Leave>", lambda e: table_card.unbind_all("<MouseWheel>"))

    # ── Helpers data / yen ────────────────────────────────────────────────────
    def _make_date_row(self, parent, default_today=True):
        import datetime
        hoje = datetime.date.today()

        def make_de(master, w, ph, maxlen, default_val):
            e = tk.Entry(master, font=("Helvetica", 11), width=w,
                         bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                         insertbackground=COLORS["text_primary"],
                         relief="flat", bd=0,
                         highlightthickness=1, highlightbackground=COLORS["border"],
                         highlightcolor=COLORS["accent"], justify="center")
            val = default_val if default_today else ph
            e.insert(0, val)
            e.configure(fg=COLORS["text_primary"] if default_today else COLORS["text_muted"])

            def fi(ev, entry=e, p=ph):
                if entry.get() == p:
                    entry.delete(0, tk.END)
                    entry.configure(fg=COLORS["text_primary"])

            def fo(ev, entry=e, p=ph):
                if not entry.get():
                    entry.insert(0, p)
                    entry.configure(fg=COLORS["text_muted"])

            def kr(ev, entry=e, ml=maxlen):
                raw = "".join(filter(str.isdigit, entry.get()))[:ml]
                entry.delete(0, tk.END)
                entry.insert(0, raw)
                entry.configure(fg=COLORS["text_primary"])

            e.bind("<FocusIn>", fi)
            e.bind("<FocusOut>", fo)
            e.bind("<KeyRelease>", kr)
            return e

        d = make_de(parent, 4, "DD",   2, str(hoje.day).zfill(2))
        d.pack(side="left", ipady=7)
        tk.Label(parent, text="/", font=("Helvetica", 13, "bold"),
                 bg=parent["bg"], fg=COLORS["text_secondary"]).pack(side="left", padx=2)
        m = make_de(parent, 4, "MM",   2, str(hoje.month).zfill(2))
        m.pack(side="left", ipady=7)
        tk.Label(parent, text="/", font=("Helvetica", 13, "bold"),
                 bg=parent["bg"], fg=COLORS["text_secondary"]).pack(side="left", padx=2)
        a = make_de(parent, 6, "AAAA", 4, str(hoje.year))
        a.pack(side="left", ipady=7)
        return d, m, a

    def _make_yen_entry(self, parent, width=22):
        e = tk.Entry(parent, font=("Helvetica", 11), width=width,
                     bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                     insertbackground=COLORS["text_primary"],
                     relief="flat", bd=0,
                     highlightthickness=1, highlightbackground=COLORS["border"],
                     highlightcolor=COLORS["accent"])
        e.bind("<KeyRelease>", lambda ev: self._fmt_yen(e))
        return e

    def _fmt_telefone_jp(self, event=None):
        """Formata campo telefone no padrão japonês: (0XX) 0000-0000"""
        e = self.entry_tel
        raw = "".join(filter(str.isdigit, e.get()))[:11]
        if len(raw) <= 3:
            fmt = f"({raw}"
        elif len(raw) <= 7:
            fmt = f"({raw[:3]}) {raw[3:]}"
        else:
            fmt = f"({raw[:3]}) {raw[3:7]}-{raw[7:]}"
        pos = e.index(tk.INSERT)
        e.delete(0, tk.END)
        e.insert(0, fmt)

    def _toggle_taxa_leilao(self):
        if self.compra_tipo_var.get() == "Leilão":
            self._taxa_leilao_frame.pack(fill="x", padx=18, pady=(0, 12))
        else:
            self._taxa_leilao_frame.pack_forget()
            self.compra_taxa_leilao_entry.delete(0, tk.END)
            self.compra_taxa_rec_entry.delete(0, tk.END)
        self._toggle_compra_cliente()

    def _toggle_compra_cliente(self):
        """Mostra/oculta label obrigatório do cliente conforme tipo."""
        if not hasattr(self, "_compra_cliente_lbl_req"):
            return
        tipo = self.compra_tipo_var.get()
        required = tipo in ("Compra Direta", "Troca")
        fg = COLORS["accent"] if required else COLORS["text_secondary"]
        txt = "Cliente  ★" if required else "Cliente"
        self._compra_cliente_lbl_req.configure(text=txt, fg=fg)

    def _atualizar_combo_clientes_compra(self):
        """Popula combo de clientes no form de compra."""
        clientes = [dict(r) for r in self.conn.execute(
            "SELECT id, nome FROM clientes ORDER BY nome").fetchall()]
        opts = [f"{c['id']} — {c['nome']}" for c in clientes]
        self.compra_cliente_combo["values"] = opts if opts else ["Nenhum cliente"]

    def _get_date_from_entries(self, e_dia, e_mes, e_ano):
        dia = e_dia.get().strip()
        mes = e_mes.get().strip()
        ano = e_ano.get().strip()
        dia = "" if dia == "DD"   else dia
        mes = "" if mes == "MM"   else mes
        ano = "" if ano == "AAAA" else ano
        return f"{dia.zfill(2)}/{mes.zfill(2)}/{ano}" if dia and mes and ano else ""

    def _atualizar_combo_carros(self):
        # IDs de carros já com compra registrada (exceto a compra sendo editada)
        query = "SELECT DISTINCT carro_id FROM compras WHERE carro_id IS NOT NULL"
        if self._compra_edit_id:
            query += f" AND id != {self._compra_edit_id}"
        ids_com_compra = {row[0] for row in self.conn.execute(query).fetchall()}

        carros_estoque = [
            f"{c['id']} — {c['carro']} {c['ano']} | {c['cor'] or '—'} | {c['placa'] or '—'}"
            for c in getattr(self, "carros_data", [])
            if c["status"] == "Estoque" and c["id"] not in ids_com_compra
        ]

        # Ao editar: garante que o carro da compra atual aparece
        if self._compra_edit_id:
            row = self.conn.execute(
                "SELECT * FROM compras WHERE id=?", (self._compra_edit_id,)).fetchone()
            if row:
                cid = row["carro_id"]
                already = any(f"{cid} —" in v for v in carros_estoque)
                if not already:
                    c = next((x for x in getattr(self, "carros_data", [])
                              if x["id"] == cid), None)
                    if c:
                        carros_estoque.insert(0,
                            f"{c['id']} — {c['carro']} {c['ano']} | {c['cor'] or '—'} | {c['placa'] or '—'}")

        self.compra_carro_combo["values"] = carros_estoque if carros_estoque else ["Nenhum carro disponível"]

    def _fmt_yen(self, entry):
        raw = "".join(filter(str.isdigit,
                             entry.get().replace(",","").replace(".","").replace(" ","")))
        entry.delete(0, tk.END)
        entry.insert(0, f"{int(raw):,}" if raw else "")

    def _yen_raw(self, entry):
        return entry.get().replace(",", "").strip()

    def _salvar_compra(self):
        carro_sel = self.compra_carro_var.get().strip()
        tipo      = self.compra_tipo_var.get().strip()
        valor     = self._yen_raw(self.compra_valor_entry)
        taxa_l    = self._yen_raw(self.compra_taxa_leilao_entry) if tipo == "Leilão" else ""
        taxa_r    = self._yen_raw(self.compra_taxa_rec_entry)

        if not carro_sel or carro_sel == "Nenhum carro em Estoque":
            self.lbl_compra_status.configure(text="⚠ Selecione um carro.", fg=COLORS["red"]); return
        if not tipo:
            self.lbl_compra_status.configure(text="⚠ Selecione o tipo de compra.", fg=COLORS["red"]); return

        # Validação de cliente obrigatório para Compra Direta e Troca
        cliente_id_compra = None
        if hasattr(self, "compra_cliente_var"):
            cli_sel = self.compra_cliente_var.get().strip()
            if cli_sel and "—" in cli_sel:
                try: cliente_id_compra = int(cli_sel.split("—")[0].strip())
                except Exception: pass
            if tipo in ("Compra Direta", "Troca") and not cliente_id_compra:
                self.lbl_compra_status.configure(
                    text="⚠ Selecione o Cliente para este tipo de compra.", fg=COLORS["red"])
                return

        data_entrada = self._get_date_from_entries(self.compra_dia, self.compra_mes, self.compra_ano)
        partes     = carro_sel.split("|")
        cor_sel    = partes[1].strip() if len(partes) > 1 else ""
        placa_sel  = partes[2].strip() if len(partes) > 2 else ""
        carro_nome = carro_sel.split("|")[0].strip()
        # Extrai carro_id do texto "ID — Nome Ano | ..."
        try:
            carro_id = int(carro_sel.split("—")[0].strip())
        except (ValueError, IndexError):
            carro_id = None

        # ── BLOQUEIO DUPLICIDADE: verifica direto no banco ──────────────────
        # Permite nova compra do mesmo `carro_id` caso a compra anterior já tenha
        # sido vendida (exista uma venda vinculada). Bloqueia apenas se existir
        # uma compra atual sem venda associada.
        if carro_id is not None:
            query = (
                "SELECT c.id FROM compras c LEFT JOIN vendas v ON v.compra_id = c.id "
                "WHERE c.carro_id=? AND v.id IS NULL"
            )
            params = [carro_id]
            if self._compra_edit_id is not None:
                query += " AND c.id != ?"
                params.append(self._compra_edit_id)
            duplicado = self.conn.execute(query, params).fetchone()
            if duplicado:
                self.lbl_compra_status.configure(
                    text=f"⚠ Este carro já tem compra registrada (#{duplicado[0]}).",
                    fg=COLORS["red"])
                return

        if self._compra_edit_id is not None:
            self.conn.execute(
                "UPDATE compras SET carro_id=?,carro=?,cor=?,placa=?,tipo=?,valor=?,data_entrada=?,"
                "taxa_leilao=?,taxa_reciclagem=?,cliente_id=?,data_compra=? WHERE id=?",
                (carro_id, carro_nome, cor_sel, placa_sel, tipo, valor, data_entrada,
                 taxa_l, taxa_r, cliente_id_compra, data_entrada, self._compra_edit_id))
            self.conn.commit()
            cid = self._compra_edit_id
            self._compra_edit_id = None
            self.btn_salvar_compra.configure(text="  Registrar Compra  ")
            self._lbl_form_compra_title.configure(text="⊕  Nova Compra")
            self._recriar_custos_taxa(cid, taxa_l, taxa_r, data_entrada)
            self.lbl_compra_status.configure(text="✔ Compra atualizada!", fg=COLORS["green"])
        else:
            cur = self.conn.execute(
                "INSERT INTO compras (carro_id,carro,cor,placa,tipo,valor,data_entrada,"
                "taxa_leilao,taxa_reciclagem,cliente_id,data_compra) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (carro_id, carro_nome, cor_sel, placa_sel, tipo, valor, data_entrada,
                 taxa_l, taxa_r, cliente_id_compra, data_entrada))
            self.conn.commit()
            cid = cur.lastrowid
            self._recriar_custos_taxa(cid, taxa_l, taxa_r, data_entrada)
            self.lbl_compra_status.configure(text="✔ Compra registrada!", fg=COLORS["green"])

        self.compras_data = [dict(r) for r in
                             self.conn.execute("SELECT * FROM compras ORDER BY id DESC").fetchall()]
        self._limpar_form_compra(clear_status=False)
        self._refresh_tabela_compras()

    def _recriar_custos_taxa(self, compra_id, taxa_leilao, taxa_reciclagem, data_entrada):
        self.conn.execute(
            "DELETE FROM custos WHERE compra_id=? AND tipo_custo IN ('Taxa Leilão','Taxa Reciclagem')",
            (compra_id,))
        if taxa_leilao:
            self.conn.execute(
                "INSERT INTO custos (compra_id,tipo_custo,descricao,valor,data_custo) VALUES (?,?,?,?,?)",
                (compra_id, "Taxa Leilão", "Automático", taxa_leilao, data_entrada))
        if taxa_reciclagem:
            self.conn.execute(
                "INSERT INTO custos (compra_id,tipo_custo,descricao,valor,data_custo) VALUES (?,?,?,?,?)",
                (compra_id, "Taxa Reciclagem", "Automático", taxa_reciclagem, data_entrada))
        self.conn.commit()

    def _limpar_form_compra(self, clear_status=True):
        import datetime
        self.compra_carro_var.set("")
        self.compra_tipo_var.set("")
        self.compra_valor_entry.delete(0, tk.END)
        self.compra_taxa_rec_entry.delete(0, tk.END)
        self.compra_taxa_leilao_entry.delete(0, tk.END)
        self._taxa_leilao_frame.pack_forget()
        self._compra_edit_id = None
        self.btn_salvar_compra.configure(text="  Registrar Compra  ")
        self._lbl_form_compra_title.configure(text="⊕  Nova Compra")
        hoje = datetime.date.today()
        for e, val in [(self.compra_dia, str(hoje.day).zfill(2)),
                       (self.compra_mes, str(hoje.month).zfill(2)),
                       (self.compra_ano, str(hoje.year))]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        if clear_status:
            self.lbl_compra_status.configure(text="")

    def _editar_compra(self, cid):
        c = next((x for x in self.compras_data if x["id"] == cid), None)
        if not c: return
        self._compra_edit_id = cid
        self._lbl_form_compra_title.configure(text="✏  Editar Compra")
        self.btn_salvar_compra.configure(text="  Salvar Alterações  ")
        self._atualizar_combo_carros()
        carro_nome = c["carro"]
        vals = list(self.compra_carro_combo["values"])
        match = next((v for v in vals if carro_nome.split(" — ")[-1].split("|")[0].strip() in v), None)
        self.compra_carro_var.set(match or carro_nome)
        self.compra_tipo_var.set(c["tipo"])
        self.compra_valor_entry.delete(0, tk.END)
        if c["valor"]:
            try: self.compra_valor_entry.insert(0, f"{int(c['valor']):,}")
            except: self.compra_valor_entry.insert(0, c["valor"])
        self.compra_taxa_leilao_entry.delete(0, tk.END)
        if c["tipo"] == "Leilão":
            self._taxa_leilao_frame.pack(fill="x")
            if c.get("taxa_leilao"):
                try: self.compra_taxa_leilao_entry.insert(0, f"{int(c['taxa_leilao']):,}")
                except: self.compra_taxa_leilao_entry.insert(0, c["taxa_leilao"])
        else:
            self._taxa_leilao_frame.pack_forget()
        self.compra_taxa_rec_entry.delete(0, tk.END)
        if c.get("taxa_reciclagem"):
            try: self.compra_taxa_rec_entry.insert(0, f"{int(c['taxa_reciclagem']):,}")
            except: self.compra_taxa_rec_entry.insert(0, c["taxa_reciclagem"])
        data = c.get("data_entrada") or ""
        partes = data.split("/")
        import datetime; hoje = datetime.date.today()
        dd = partes[0] if len(partes) > 0 and partes[0] else str(hoje.day).zfill(2)
        mm = partes[1] if len(partes) > 1 and partes[1] else str(hoje.month).zfill(2)
        aa = partes[2] if len(partes) > 2 and partes[2] else str(hoje.year)
        for e, val in [(self.compra_dia, dd), (self.compra_mes, mm), (self.compra_ano, aa)]:
            e.delete(0, tk.END); e.insert(0, val); e.configure(fg=COLORS["text_primary"])
        # Carrega cliente se existir
        if hasattr(self, "compra_cliente_var") and c.get("cliente_id"):
            self._atualizar_combo_clientes_compra()
            vals = list(self.compra_cliente_combo["values"])
            match = next((v for v in vals if v.startswith(f"{c['cliente_id']} —")), None)
            if match:
                self.compra_cliente_var.set(match)
        self.lbl_compra_status.configure(text="")

    def _detalhe_placeholder(self):
        for w in self._detalhe_panel.winfo_children():
            w.destroy()
        tk.Frame(self._detalhe_panel, bg=COLORS["blue"], height=4).pack(fill="x")
        tk.Label(self._detalhe_panel,
                 text="▦  Custos do Carro",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]
                 ).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Frame(self._detalhe_panel, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(0, 0))
        tk.Label(self._detalhe_panel,
                 text="Clique em  ▼ Custos\nno histórico para ver\nos custos de uma compra.",
                 font=("Helvetica", 9), justify="center",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]
                 ).pack(expand=True)

    def _ver_custos_compra(self, cid):
        if self._compra_detalhe_id == cid:
            self._compra_detalhe_id = None
            self._detalhe_placeholder()
            self._refresh_tabela_compras()
            return
        self._compra_detalhe_id = cid
        self._build_detalhe_custos(cid)
        self._refresh_tabela_compras()

    def _build_detalhe_custos(self, compra_id):
        for w in self._detalhe_panel.winfo_children():
            w.destroy()

        compra = self.conn.execute("SELECT * FROM compras WHERE id=?", (compra_id,)).fetchone()
        if not compra:
            self._detalhe_placeholder()
            return
        c = dict(compra)

        # Header fixo
        tk.Frame(self._detalhe_panel, bg=COLORS["blue"], height=4).pack(fill="x")
        hdr = tk.Frame(self._detalhe_panel, bg=COLORS["bg_card"])
        hdr.pack(fill="x", padx=14, pady=(10, 6))
        tk.Label(hdr, text=f"▦  {c['carro']}",
                 font=("Helvetica", 10, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 wraplength=200, anchor="w", justify="left").pack(side="left")
        tk.Button(hdr, text="✕", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=5,
                  command=lambda: self._ver_custos_compra(compra_id)
                  ).pack(side="right", ipady=2)

        tk.Frame(self._detalhe_panel, bg=COLORS["border"], height=1).pack(fill="x", padx=14)

        # Área scrollável para a lista de custos
        det_canvas = tk.Canvas(self._detalhe_panel, bg=COLORS["bg_card"], highlightthickness=0)
        det_sb = tk.Scrollbar(self._detalhe_panel, orient="vertical", command=det_canvas.yview)
        det_canvas.configure(yscrollcommand=det_sb.set)
        det_inner = tk.Frame(det_canvas, bg=COLORS["bg_card"])
        det_canvas.create_window((0, 0), window=det_inner, anchor="nw")
        det_inner.bind("<Configure>",
            lambda e: det_canvas.configure(scrollregion=det_canvas.bbox("all")))

        def _scroll_det(ev):
            det_canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        self._detalhe_panel.bind("<Enter>",
            lambda e: self._detalhe_panel.bind_all("<MouseWheel>", _scroll_det))
        self._detalhe_panel.bind("<Leave>",
            lambda e: self._detalhe_panel.unbind_all("<MouseWheel>"))

        # Busca custos
        rows = [dict(r) for r in self.conn.execute(
            "SELECT * FROM custos WHERE compra_id=? ORDER BY id DESC", (compra_id,)).fetchall()]

        if not rows:
            tk.Label(det_inner, text="Nenhum custo registrado.",
                     font=("Helvetica", 9), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).pack(padx=14, pady=20)
        else:
            for i, cu in enumerate(rows):
                rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
                item = tk.Frame(det_inner, bg=rb, highlightthickness=1,
                                highlightbackground=COLORS["border"])
                item.pack(fill="x", padx=10, pady=(4, 0))

                top = tk.Frame(item, bg=rb)
                top.pack(fill="x", padx=10, pady=(6, 2))
                tk.Label(top, text=cu["tipo_custo"],
                         font=("Helvetica", 8, "bold"),
                         bg=COLORS["blue"], fg="white",
                         padx=6, pady=2).pack(side="left")
                tk.Label(top, text=cu.get("data_custo") or "—",
                         font=("Helvetica", 8),
                         bg=rb, fg=COLORS["text_muted"]).pack(side="right")

                bot = tk.Frame(item, bg=rb)
                bot.pack(fill="x", padx=10, pady=(0, 6))
                desc = cu.get("descricao") or ""
                if desc:
                    tk.Label(bot, text=desc, font=("Helvetica", 8),
                             bg=rb, fg=COLORS["text_secondary"],
                             wraplength=220, anchor="w", justify="left").pack(side="left")
                tk.Label(bot, text=self._fmt_yen_display(cu["valor"]),
                         font=("Helvetica", 9, "bold"),
                         bg=rb, fg=COLORS["text_primary"]).pack(side="right")

        det_canvas.pack(side="left", fill="both", expand=True)
        det_sb.pack(side="right", fill="y")

        # Footer totalizador — fixo no fundo
        custo_total = self._get_custo_total(compra_id)
        try: vcp = int(str(c["valor"]).replace(",","")) if c["valor"] else 0
        except: vcp = 0

        sep = tk.Frame(self._detalhe_panel, bg=COLORS["border"], height=1)
        sep.pack(fill="x", padx=14, side="bottom")

        tot = tk.Frame(self._detalhe_panel, bg=COLORS["bg_main"])
        tot.pack(fill="x", side="bottom", pady=0)

        for lt, val, fg in [
            ("Total:", self._fmt_yen_display(vcp + custo_total), COLORS["green"]),
            ("Custos:", self._fmt_yen_display(custo_total), COLORS["orange"]),
            ("Compra:", self._fmt_yen_display(c["valor"]), COLORS["text_primary"]),
        ]:
            row_t = tk.Frame(tot, bg=COLORS["bg_main"])
            row_t.pack(fill="x", padx=14, pady=2)
            tk.Label(row_t, text=lt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_secondary"],
                     anchor="w").pack(side="left")
            tk.Label(row_t, text=val, font=("Helvetica", 9, "bold"),
                     bg=COLORS["bg_main"], fg=fg,
                     anchor="e").pack(side="right")

    def _excluir_compra(self, compra):
        """Abre dialog de confirmação — usuário deve digitar 'excluir' para confirmar."""
        cid        = compra["id"]
        carro      = compra.get("carro", "—")
        tipo       = compra.get("tipo", "—")
        data       = compra.get("data_entrada") or "—"
        valor      = self._fmt_yen_display(compra.get("valor"))
        custo_tot  = self._get_custo_total(cid)
        try: vcp = int(str(compra["valor"]).replace(",","")) if compra["valor"] else 0
        except: vcp = 0
        total      = self._fmt_yen_display(vcp + custo_tot)

        # ── Janela de confirmação ─────────────────────────────────────────────
        dlg = tk.Toplevel(self)
        dlg.title("Excluir Compra")
        dlg.resizable(False, False)
        dlg.configure(bg=COLORS["bg_card"])
        dlg.grab_set()

        # Centraliza
        dlg.update_idletasks()
        w, h = 420, 320
        x = self.winfo_x() + (self.winfo_width()  - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

        # Header vermelho
        tk.Frame(dlg, bg=COLORS["red"], height=4).pack(fill="x")
        tk.Label(dlg, text="⚠  Excluir Compra",
                 font=("Helvetica", 12, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["red"]).pack(pady=(16, 4))

        # Dados da compra
        info_f = tk.Frame(dlg, bg=COLORS["bg_content"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
        info_f.pack(fill="x", padx=24, pady=(4, 12))
        for label, val in [
            ("Carro:",        carro),
            ("Tipo:",         tipo),
            ("Data entrada:", data),
            ("Valor compra:", valor),
            ("Total (c/custos):", total),
        ]:
            r = tk.Frame(info_f, bg=COLORS["bg_content"])
            r.pack(fill="x", padx=12, pady=2)
            tk.Label(r, text=label, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_secondary"],
                     width=16, anchor="w").pack(side="left")
            tk.Label(r, text=val, font=("Helvetica", 9),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"],
                     anchor="w").pack(side="left")

        # Aviso custos
        n_custos = self.conn.execute(
            "SELECT COUNT(*) FROM custos WHERE compra_id=?", (cid,)).fetchone()[0]
        if n_custos:
            tk.Label(dlg, text=f"⚠ {n_custos} custo(s) vinculado(s) também serão excluídos.",
                     font=("Helvetica", 8), bg=COLORS["bg_card"],
                     fg=COLORS["orange"]).pack(pady=(0, 4))

        # Campo de confirmação
        tk.Label(dlg, text='Digite  "excluir"  para confirmar:',
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack()
        conf_var = tk.StringVar()
        conf_entry = tk.Entry(dlg, textvariable=conf_var,
                              font=("Helvetica", 11), width=18,
                              justify="center",
                              bg=COLORS["bg_main"], fg=COLORS["red"],
                              insertbackground=COLORS["red"],
                              relief="flat", bd=0,
                              highlightthickness=2,
                              highlightbackground=COLORS["border"],
                              highlightcolor=COLORS["red"])
        conf_entry.pack(pady=(6, 12), ipady=6)
        conf_entry.focus_set()

        lbl_err = tk.Label(dlg, text="", font=("Helvetica", 8),
                           bg=COLORS["bg_card"], fg=COLORS["red"])
        lbl_err.pack()

        def confirmar():
            if conf_var.get().strip().lower() != "excluir":
                lbl_err.configure(text='Digite exatamente "excluir" para confirmar.')
                conf_entry.configure(highlightbackground=COLORS["red"])
                return
            # Exclui custos e compra
            self.conn.execute("DELETE FROM custos  WHERE compra_id=?", (cid,))
            self.conn.execute("DELETE FROM compras WHERE id=?", (cid,))
            self.conn.commit()
            dlg.destroy()
            # Fecha painel de custos se estava aberto
            if self._compra_detalhe_id == cid:
                self._compra_detalhe_id = None
                self._detalhe_placeholder()
            if self._compra_edit_id == cid:
                self._compra_edit_id = None
                self._limpar_form_compra()
            self.compras_data = [dict(r) for r in
                self.conn.execute("SELECT * FROM compras ORDER BY id DESC").fetchall()]
            self._refresh_tabela_compras()

        btn_row = tk.Frame(dlg, bg=COLORS["bg_card"])
        btn_row.pack(pady=(0, 16))
        tk.Button(btn_row, text="  Cancelar  ", font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=dlg.destroy).pack(side="left", padx=(0, 10), ipady=5, ipadx=4)
        tk.Button(btn_row, text="  Excluir Compra  ", font=("Helvetica", 9, "bold"),
                  bg=COLORS["red"], fg="white",
                  relief="flat", cursor="hand2",
                  command=confirmar).pack(side="left", ipady=5, ipadx=4)
        dlg.bind("<Return>", lambda e: confirmar())

    def _get_custo_total(self, compra_id):
        """Retorna total de custos da compra (inclui Serviço Shaken de Estoque/Daisha)."""
        row = self.conn.execute(
            "SELECT SUM(CAST(REPLACE(valor,',','') AS REAL)) FROM custos WHERE compra_id=?",
            (compra_id,)).fetchone()
        return int(row[0]) if row[0] else 0

    def _fmt_yen_display(self, valor_str):
        try:
            return f"¥ {int(str(valor_str).replace(',','')):,}"
        except:
            return "—"

    def _refresh_tabela_compras(self):
        for w in self.compras_rows_frame.winfo_children():
            w.destroy()

        COMP_COLS = getattr(self, "_comp_cols", [
            ("#",32),("Data",72),("Carro",115),("Cor",52),("Placa",62),
            ("Cliente",95),("Tipo",70),("V.Compra",78),("Custo",78),
            ("Total",82),("Shaken",82),("St.SK",62),("Ações",88),
        ])
        for ci, (_, pw) in enumerate(COMP_COLS):
            self.compras_rows_frame.grid_columnconfigure(ci, minsize=pw, weight=0)

        filtro = getattr(self, "_compra_filtro_var", None)
        termo  = filtro.get().strip().lower() if filtro else ""
        lista  = self.compras_data

        # Enriquece lista com cliente_nome
        lista_rich = []
        for c in lista:
            r = dict(c)
            cli_row = None
            if r.get("cliente_id"):
                cli_row = self.conn.execute(
                    "SELECT nome FROM clientes WHERE id=?",
                    (r["cliente_id"],)).fetchone()
            r["_cli_nome"] = cli_row[0] if cli_row else ""
            lista_rich.append(r)

        if termo:
            lista_rich = [c for c in lista_rich if
                     termo in (c.get("carro") or "").lower() or
                     termo in (c.get("cor")   or "").lower() or
                     termo in (c.get("placa") or "").lower() or
                     termo in (c.get("tipo")  or "").lower() or
                     termo in (c.get("_cli_nome") or "").lower() or
                     termo in (c.get("data_entrada") or "").lower()]

        self.lbl_total_compras.configure(text=f"{len(lista_rich)} registro(s)")
        if not lista_rich:
            tk.Label(self.compras_rows_frame, text="Nenhuma compra encontrada.",
                     font=("Helvetica",10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"]).grid(
                         row=0, column=0, pady=30, columnspan=len(COMP_COLS))
            return

        TIPO_COLORS = {"Leilão": COLORS["orange"], "Troca": COLORS["blue"],
                       "Compra Direta": COLORS["green"]}

        for i, c in enumerate(lista_rich):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            tk.Frame(self.compras_rows_frame, bg=COLORS["border"], height=1
                     ).grid(row=i*2, column=0, columnspan=len(COMP_COLS), sticky="ew")
            ri = i*2 + 1
            self.compras_rows_frame.grid_rowconfigure(ri, minsize=32)

            custo_total = self._get_custo_total(c["id"])
            try: vcp = int(str(c["valor"]).replace(",","")) if c["valor"] else 0
            except: vcp = 0
            carro_txt = c["carro"]
            if " — " in carro_txt: carro_txt = carro_txt.split(" — ", 1)[1]
            carro_txt = carro_txt.split("|")[0].strip()
            cli_disp = (c["_cli_nome"] or "—")[:14]

            col = 0
            tk.Label(self.compras_rows_frame, text=str(c["id"]),
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_muted"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(self.compras_rows_frame, text=c.get("data_entrada") or "—",
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_secondary"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            c_disp = carro_txt[:17]+"…" if len(carro_txt)>18 else carro_txt
            tk.Label(self.compras_rows_frame, text=c_disp,
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_primary"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(self.compras_rows_frame, text=c.get("cor") or "—",
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_secondary"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(self.compras_rows_frame, text=c.get("placa") or "—",
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_secondary"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(self.compras_rows_frame, text=cli_disp,
                     font=("Helvetica",8), bg=rb, fg=COLORS["accent"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Tipo badge
            tc_f = tk.Frame(self.compras_rows_frame, bg=rb)
            tc_f.grid(row=ri, column=col, sticky="nsew"); col+=1
            tc = TIPO_COLORS.get(c["tipo"], COLORS["text_muted"])
            tk.Label(tc_f, text=c["tipo"], font=("Helvetica",7,"bold"),
                     bg=tc, fg="white", anchor="center"
                     ).pack(fill="both", expand=True, padx=3, pady=5)
            tk.Label(self.compras_rows_frame, text=self._fmt_yen_display(c["valor"]),
                     font=("Helvetica",8), bg=rb, fg=COLORS["text_primary"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(self.compras_rows_frame, text=self._fmt_yen_display(custo_total),
                     font=("Helvetica",8), bg=rb, fg=COLORS["orange"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            tk.Label(self.compras_rows_frame,
                     text=self._fmt_yen_display(vcp+custo_total),
                     font=("Helvetica",8,"bold"), bg=rb, fg=COLORS["green"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            # Shaken date + status for this carro
            try:
                sk_row_c = self.conn.execute(
                    "SELECT data_vencimento, por_conta FROM shaken "
                    "WHERE carro_id=? AND renovado=0 ORDER BY id DESC LIMIT 1",
                    (c.get("carro_id"),)).fetchone()
                sk_dv_c = sk_row_c[0] if sk_row_c else None
                sk_pc_c = sk_row_c[1] if sk_row_c else 0
            except Exception:
                sk_dv_c, sk_pc_c = None, 0
            sk_txt_c, sk_cor_c = self._sk_status_info(sk_dv_c, sk_pc_c)
            sk_disp_c = (sk_dv_c or "—")[:10] if sk_dv_c and sk_dv_c != "Por Conta" else (sk_dv_c or "—")
            tk.Label(self.compras_rows_frame, text=sk_disp_c,
                     font=("Helvetica",8), bg=rb, fg=COLORS["blue"],
                     anchor="w", padx=4
                     ).grid(row=ri, column=col, sticky="nsew"); col+=1
            sk_f_c = tk.Frame(self.compras_rows_frame, bg=rb)
            sk_f_c.grid(row=ri, column=col, sticky="nsew"); col+=1
            bfg_c = "white" if sk_cor_c not in ("#C9A800","#D4AC0D") else "#333"
            tk.Label(sk_f_c, text=sk_txt_c, font=("Helvetica",7,"bold"),
                     bg=sk_cor_c, fg=bfg_c, anchor="center"
                     ).pack(fill="both", expand=True, padx=2, pady=5)
            # Ações
            acts_f = tk.Frame(self.compras_rows_frame, bg=rb)
            acts_f.grid(row=ri, column=col, sticky="nsew")
            tk.Button(acts_f, text="✏", font=("Helvetica",7), bg=COLORS["blue"], fg="white",
                      relief="flat", cursor="hand2", padx=3, pady=1,
                      command=lambda cid=c["id"]: self._editar_compra(cid)
                      ).pack(side="left", padx=(3,1), pady=4)
            tk.Button(acts_f, text="🗑", font=("Helvetica",7), bg=COLORS["red"], fg="white",
                      relief="flat", cursor="hand2", padx=3, pady=1,
                      command=lambda cc=c: self._excluir_compra(cc)
                      ).pack(side="left", padx=(0,1), pady=4)
            is_venda = bool(c.get("a_venda"))
            tk.Button(acts_f,
                      text="✔Venda" if is_venda else "À Venda",
                      font=("Helvetica",7,"bold"),
                      bg=COLORS["green"] if is_venda else COLORS["red"],
                      fg="white", relief="flat", cursor="hand2", padx=2, pady=1,
                      command=lambda cid=c["id"], av=is_venda: self._toggle_a_venda(cid, av)
                      ).pack(side="left", padx=(0,1), pady=4)
            is_open = (self._compra_detalhe_id == c["id"])
            tk.Button(acts_f, text="▲" if is_open else "▼",
                      font=("Helvetica",7),
                      bg=COLORS["border"] if is_open else COLORS["orange"],
                      fg=COLORS["text_secondary"] if is_open else "white",
                      relief="flat", cursor="hand2", padx=3, pady=1,
                      command=lambda cid=c["id"]: self._ver_custos_compra(cid)
                      ).pack(side="left", pady=4)

        # linha final
        tk.Frame(self.compras_rows_frame, bg=COLORS["border"], height=1
                 ).grid(row=len(lista_rich)*2, column=0,
                        columnspan=len(COMP_COLS), sticky="ew")

    def _toggle_a_venda(self, compra_id, currently_venda):
        """Alterna status À Venda da compra e atualiza o estoque de veículos."""
        novo = 0 if currently_venda else 1
        self.conn.execute("UPDATE compras SET a_venda=? WHERE id=?", (novo, compra_id))
        self.conn.commit()
        self.compras_data = [dict(r) for r in
            self.conn.execute("SELECT * FROM compras ORDER BY id DESC").fetchall()]
        self._refresh_tabela_compras()
        # Atualiza tela de estoque se estiver aberta
        try:
            self._refresh_estoque_veiculos()
        except Exception:
            pass

    # ── Estoque ───────────────────────────────────────────────────────────────
    def _build_estoque_subs(self, parent):
        subs = SUBMENUS["Estoque"]
        for sub in subs:
            frame = tk.Frame(parent, bg=COLORS["bg_content"])
            self.sub_pages["Estoque"][sub] = frame
            if sub == "Veículos":
                self._build_estoque_veiculos(frame)
            else:
                self._build_placeholder(frame, sub)
        self.sub_pages["Estoque"][subs[0]].place(
            in_=parent, x=0, y=0, relwidth=1, relheight=1)

    def _build_estoque_veiculos(self, parent):
        self._estoque_filtro_var = tk.StringVar(value="")

        root = tk.Frame(parent, bg=COLORS["bg_content"])
        root.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg=COLORS["bg_content"])
        hdr.pack(fill="x", pady=(0, 10))
        tk.Label(hdr, text="▦  Veículos À Venda",
                 font=("Helvetica", 14, "bold"),
                 bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")
        tk.Button(hdr, text="↺ Atualizar", font=("Helvetica", 8),
                  bg=COLORS["accent"], fg="white", relief="flat", cursor="hand2",
                  padx=8, pady=3, command=self._refresh_estoque_veiculos).pack(side="right", padx=(8,0))
        self.lbl_estoque_total = tk.Label(hdr, text="0 veículos",
                                          font=("Helvetica", 9),
                                          bg=COLORS["bg_content"], fg=COLORS["text_muted"])
        self.lbl_estoque_total.pack(side="right", padx=4)

        # ── Painel de edição de preço (sidebar direita) ───────────────────────
        self._estoque_edit_panel = tk.Frame(root, bg=COLORS["bg_card"],
                                            highlightthickness=1,
                                            highlightbackground=COLORS["border"],
                                            width=280)
        self._estoque_edit_panel.pack(side="right", fill="y", padx=(10, 0))
        self._estoque_edit_panel.pack_propagate(False)
        self._estoque_edit_placeholder()

        # ── Área principal: cards de veículos ─────────────────────────────────
        main = tk.Frame(root, bg=COLORS["bg_content"])
        main.pack(side="left", fill="both", expand=True)

        # Filtro
        filt_bar = tk.Frame(main, bg=COLORS["bg_main"])
        filt_bar.pack(fill="x", pady=(0, 8))
        tk.Label(filt_bar, text="🔍", font=("Helvetica", 9),
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"]).pack(side="left", padx=(0,4), pady=5)
        tk.Entry(filt_bar, textvariable=self._estoque_filtro_var,
                 font=("Helvetica", 9), width=28,
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COLORS["border"],
                 highlightcolor=COLORS["accent"]
                 ).pack(side="left", ipady=5)
        self._estoque_filtro_var.trace_add("write", lambda *_: self._refresh_estoque_veiculos())
        tk.Button(filt_bar, text="📄 PDF Cliente", font=("Helvetica",8,"bold"),
                  bg=COLORS["blue"], fg="white", relief="flat", cursor="hand2",
                  command=self._pdf_estoque_cliente
                  ).pack(side="right", ipady=4, padx=(0,4))
        tk.Button(filt_bar, text="✕", font=("Helvetica", 8),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2", padx=5,
                  command=lambda: self._estoque_filtro_var.set("")
                  ).pack(side="left", padx=(4, 0), ipady=4)

        # Tabela
        tk.Frame(main, bg=COLORS["border"], height=1).pack(fill="x")
        col_f = tk.Frame(main, bg=COLORS["bg_main"])
        col_f.pack(fill="x")
        for txt, w in [("#",3),("Veículo",18),("Ano",5),("Cor",7),("Placa",8),
                       ("Chassi",14),("V.Compra",11),("Custo",11),("Total",11),("P.Venda",12),("P.Avista",12),("Shaken",10),("St.SK",8),("Ação",7)]:
            tk.Label(col_f, text=txt, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_main"], fg=COLORS["text_muted"],
                     width=w, anchor="w").pack(side="left", padx=2, pady=6)

        scroll_f = tk.Frame(main, bg=COLORS["bg_card"])
        scroll_f.pack(fill="both", expand=True)
        canvas = tk.Canvas(scroll_f, bg=COLORS["bg_card"], highlightthickness=0)
        sb = tk.Scrollbar(scroll_f, orient="vertical", command=canvas.yview)
        self._estoque_rows_frame = tk.Frame(canvas, bg=COLORS["bg_card"])
        self._estoque_rows_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._estoque_rows_frame, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)

        def _scroll_est(ev):
            canvas.yview_scroll(int(-1*(ev.delta/120)), "units")
        main.bind("<Enter>", lambda e: main.bind_all("<MouseWheel>", _scroll_est))
        main.bind("<Leave>", lambda e: main.unbind_all("<MouseWheel>"))

        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._refresh_estoque_veiculos()

    def _estoque_edit_placeholder(self):
        for w in self._estoque_edit_panel.winfo_children():
            w.destroy()
        tk.Frame(self._estoque_edit_panel, bg=COLORS["green"], height=4).pack(fill="x")
        tk.Label(self._estoque_edit_panel, text="◆  Preço de Venda",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Frame(self._estoque_edit_panel, bg=COLORS["border"], height=1).pack(fill="x", padx=16)
        tk.Label(self._estoque_edit_panel,
                 text="Clique em ✎ em um\nveículo para definir\no preço sugerido.",
                 font=("Helvetica", 9), justify="center",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack(expand=True)

    def _estoque_edit_preco(self, compra):
        """Abre painel lateral para editar preço de venda do veículo."""
        for w in self._estoque_edit_panel.winfo_children():
            w.destroy()

        cid = compra["id"]
        carro_txt = compra["carro"]
        if " — " in carro_txt:
            carro_txt = carro_txt.split(" — ", 1)[1]
        carro_txt = carro_txt.split("|")[0].strip()

        tk.Frame(self._estoque_edit_panel, bg=COLORS["green"], height=4).pack(fill="x")
        tk.Label(self._estoque_edit_panel, text="◆  Preço de Venda",
                 font=("Helvetica", 11, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14,4))
        tk.Frame(self._estoque_edit_panel, bg=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(0,10))

        tk.Label(self._estoque_edit_panel, text=carro_txt,
                 font=("Helvetica", 10, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["accent"],
                 wraplength=240, justify="left").pack(anchor="w", padx=16, pady=(0,12))

        # Info resumida
        info_f = tk.Frame(self._estoque_edit_panel, bg=COLORS["bg_content"],
                          highlightthickness=1, highlightbackground=COLORS["border"])
        info_f.pack(fill="x", padx=16, pady=(0, 12))
        custo_total = self._get_custo_total(cid)
        try: vcp = int(str(compra["valor"]).replace(",","")) if compra["valor"] else 0
        except: vcp = 0
        for label, val in [
            ("Compra:", self._fmt_yen_display(compra["valor"])),
            ("Custos:", self._fmt_yen_display(custo_total)),
            ("Total:", self._fmt_yen_display(vcp + custo_total)),
        ]:
            r = tk.Frame(info_f, bg=COLORS["bg_content"])
            r.pack(fill="x", padx=10, pady=2)
            tk.Label(r, text=label, font=("Helvetica", 8, "bold"),
                     bg=COLORS["bg_content"], fg=COLORS["text_secondary"],
                     width=8, anchor="w").pack(side="left")
            tk.Label(r, text=val, font=("Helvetica", 9),
                     bg=COLORS["bg_content"], fg=COLORS["text_primary"]).pack(side="left")

        tk.Label(self._estoque_edit_panel, text="Preço Sugerido de Venda (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        preco_entry = self._make_yen_entry(self._estoque_edit_panel, width=20)
        preco_entry.pack(padx=16, pady=(4, 16), ipady=8)
        # Preenche valor existente
        if compra.get("preco_venda"):
            try:
                preco_entry.insert(0, f"{int(str(compra['preco_venda']).replace(',','')):,}")
            except:
                preco_entry.insert(0, compra["preco_venda"])

        tk.Label(self._estoque_edit_panel, text="Preço Sugerido À Vista (¥)",
                 font=("Helvetica", 9, "bold"),
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        avista_entry = self._make_yen_entry(self._estoque_edit_panel, width=20)
        avista_entry.pack(padx=16, pady=(4, 16), ipady=8)
        if compra.get("preco_avista"):
            try:
                avista_entry.insert(0, f"{int(str(compra['preco_avista']).replace(',','')):,}")
            except:
                avista_entry.insert(0, compra["preco_avista"])

        lbl_ok = tk.Label(self._estoque_edit_panel, text="",
                          font=("Helvetica", 8), bg=COLORS["bg_card"], fg=COLORS["green"])
        lbl_ok.pack()

        def salvar_preco():
            val  = self._yen_raw(preco_entry)
            aval = self._yen_raw(avista_entry)
            self.conn.execute("UPDATE compras SET preco_venda=?, preco_avista=? WHERE id=?",
                              (val, aval, cid))
            self.conn.commit()
            self.compras_data = [dict(r) for r in
                self.conn.execute("SELECT * FROM compras ORDER BY id DESC").fetchall()]
            lbl_ok.configure(text="✔ Preços salvos!")
            self._refresh_estoque_veiculos()

        tk.Button(self._estoque_edit_panel, text="  Salvar Preço  ",
                  font=("Helvetica", 10, "bold"),
                  bg=COLORS["green"], fg="white",
                  relief="flat", cursor="hand2",
                  command=salvar_preco
                  ).pack(padx=16, pady=(8, 8), ipady=7, fill="x")
        tk.Button(self._estoque_edit_panel, text="Fechar",
                  font=("Helvetica", 9),
                  bg=COLORS["border"], fg=COLORS["text_secondary"],
                  relief="flat", cursor="hand2",
                  command=self._estoque_edit_placeholder
                  ).pack(padx=16, ipady=5, fill="x")

    def _refresh_estoque_veiculos(self):
        if not hasattr(self, "_estoque_rows_frame"):
            return
        for w in self._estoque_rows_frame.winfo_children():
            w.destroy()

        termo = self._estoque_filtro_var.get().strip().lower() if hasattr(self, "_estoque_filtro_var") else ""
        veiculos = [dict(r) for r in self.conn.execute(
            "SELECT * FROM compras WHERE a_venda=1 ORDER BY id DESC").fetchall()]

        if termo:
            veiculos = [v for v in veiculos if
                        termo in (v.get("carro") or "").lower() or
                        termo in (v.get("cor")   or "").lower() or
                        termo in (v.get("placa") or "").lower()]

        self.lbl_estoque_total.configure(text=f"{len(veiculos)} veículo(s)")

        if not veiculos:
            tk.Label(self._estoque_rows_frame,
                     text="Nenhum veículo à venda.\nMarque compras como 'À Venda' no Histórico de Compras.",
                     font=("Helvetica", 10), bg=COLORS["bg_card"],
                     fg=COLORS["text_muted"], justify="center").pack(pady=40)
            return

        for i, v in enumerate(veiculos):
            rb = COLORS["bg_card"] if i % 2 == 0 else COLORS["bg_content"]
            row = tk.Frame(self._estoque_rows_frame, bg=rb)
            row.pack(fill="x")

            custo_total = self._get_custo_total(v["id"])
            try: vcp = int(str(v["valor"]).replace(",","")) if v["valor"] else 0
            except: vcp = 0

            # Busca dados detalhados do carro cadastrado
            carro_info = {}
            if v.get("carro_id"):
                row_c = self.conn.execute(
                    "SELECT * FROM carros WHERE id=?", (v["carro_id"],)).fetchone()
                if row_c:
                    carro_info = dict(row_c)

            # Nome limpo
            carro_txt = v["carro"]
            if " — " in carro_txt:
                carro_txt = carro_txt.split(" — ", 1)[1]
            carro_txt = carro_txt.split("|")[0].strip()

            ano    = carro_info.get("ano")    or v.get("ano")    or "—"
            cor    = carro_info.get("cor")    or v.get("cor")    or "—"
            placa  = carro_info.get("placa")  or v.get("placa")  or "—"
            chassi = carro_info.get("chassi") or "—"
            preco  = self._fmt_yen_display(v.get("preco_venda")) if v.get("preco_venda") else "—"

            tk.Label(row, text=str(v["id"]), font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_muted"], width=3, anchor="w").pack(side="left",padx=2,pady=7)
            tk.Label(row, text=carro_txt, font=("Helvetica",9,"bold"), bg=rb,
                     fg=COLORS["text_primary"], width=18, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=ano,    font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"], width=5, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=cor,    font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"], width=7, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=placa,  font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"], width=8, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=chassi, font=("Helvetica",9), bg=rb, fg=COLORS["text_secondary"], width=14, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=self._fmt_yen_display(v["valor"]), font=("Helvetica",9), bg=rb,
                     fg=COLORS["text_primary"], width=11, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=self._fmt_yen_display(custo_total), font=("Helvetica",9), bg=rb,
                     fg=COLORS["orange"], width=11, anchor="w").pack(side="left",padx=2)
            tk.Label(row, text=self._fmt_yen_display(vcp+custo_total), font=("Helvetica",9,"bold"), bg=rb,
                     fg=COLORS["text_primary"], width=11, anchor="w").pack(side="left",padx=2)
            # Preço de venda com destaque
            fg_p = COLORS["green"] if v.get("preco_venda") else COLORS["text_muted"]
            tk.Label(row, text=preco, font=("Helvetica",9,"bold"), bg=rb,
                     fg=fg_p, width=12, anchor="w").pack(side="left",padx=2)
            avista = self._fmt_yen_display(v.get("preco_avista")) if v.get("preco_avista") else "—"
            fg_av = COLORS["accent"] if v.get("preco_avista") else COLORS["text_muted"]
            tk.Label(row, text=avista, font=("Helvetica",9,"bold"), bg=rb,
                     fg=fg_av, width=12, anchor="w").pack(side="left",padx=2)
            # Shaken data + status
            try:
                sk_est = self.conn.execute(
                    "SELECT data_vencimento, por_conta FROM shaken "
                    "WHERE carro_id=? AND renovado=0 ORDER BY id DESC LIMIT 1",
                    (v.get("carro_id"),)).fetchone()
                sk_dv_e = sk_est[0] if sk_est else None
                sk_pc_e = sk_est[1] if sk_est else 0
            except Exception:
                sk_dv_e, sk_pc_e = None, 0
            sk_txt_e, sk_cor_e = self._sk_status_info(sk_dv_e, sk_pc_e)
            sk_disp_e = (sk_dv_e or "—")[:10]
            tk.Label(row, text=sk_disp_e, font=("Helvetica",8), bg=rb,
                     fg=COLORS["blue"], width=10, anchor="w").pack(side="left",padx=2)
            bfg_e = "white" if sk_cor_e not in ("#C9A800","#D4AC0D") else "#333"
            tk.Label(row, text=sk_txt_e, font=("Helvetica",7,"bold"),
                     bg=sk_cor_e, fg=bfg_e, width=8, anchor="center"
                     ).pack(side="left", padx=2, pady=4)
            tk.Button(row, text="✎", font=("Helvetica",9), bg=COLORS["green"], fg="white",
                      relief="flat", cursor="hand2", padx=5, pady=1,
                      command=lambda vv=v: self._estoque_edit_preco(vv)
                      ).pack(side="left", padx=4)

    def _build_dashboard_cards(self, parent):
        cards_frame = tk.Frame(parent, bg=COLORS["bg_content"])
        cards_frame.pack(pady=10)

        stats = [
            ("Veículos em Estoque", "—", COLORS["blue"]),
            ("Vendas do Mês",       "—", COLORS["green"]),
            ("Serviços Abertos",    "—", COLORS["orange"]),
            ("Parcelas Vencidas",   "—", COLORS["red"]),
        ]

        for i, (label, value, color) in enumerate(stats):
            card = tk.Frame(cards_frame, bg=COLORS["bg_card"],
                            bd=0, padx=22, pady=16,
                            highlightthickness=1,
                            highlightbackground=COLORS["border"])
            card.grid(row=0, column=i, padx=8)

            tk.Frame(card, bg=color, height=3).pack(fill="x")
            tk.Label(card, text=value, font=("Georgia", 28, "bold"),
                     bg=COLORS["bg_card"], fg=color).pack(pady=(10, 2))
            tk.Label(card, text=label, font=("Helvetica", 8),
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"]).pack()

    # ── Navigation ────────────────────────────────────────────────────────────
    def _show_page(self, page):
        # Dirty check ao trocar de página principal
        dirty = self._is_form_dirty()
        if dirty:
            curr_page = self.current_page.get()
            form_main_pages = {"Venda","Entrada/Compra","Serviço"}
            if curr_page in form_main_pages and curr_page != page:
                if not self._confirm_unsaved(dirty):
                    return
        self.current_page.set(page)
        self._nav_set_active(page)
        self._build_submenu(page)
        self.topbar_title.configure(
            text=f"  {NAV_ICONS.get(page,'•')}  {page}")

        for frame in self.pages.values():
            frame.place_forget()

        self.pages[page].place(in_=self.page_container,
                               x=0, y=0, relwidth=1, relheight=1)

        # Mostra sub-tela ativa se existir
        sub_pages = getattr(self, "sub_pages", {})
        if page in sub_pages and sub_pages[page]:
            active_sub = self.current_sub[page]
            for k, f in sub_pages[page].items():
                f.place_forget()
            target = sub_pages[page].get(active_sub) or list(sub_pages[page].values())[0]
            target.place(in_=self.pages[page], x=0, y=0, relwidth=1, relheight=1)


# ─── Entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        app = KMCars()
        app.protocol("WM_DELETE_WINDOW", lambda: (app.conn.close(), app.destroy()))
        app.mainloop()
    except KeyboardInterrupt:
        pass  # Fechamento via Ctrl+C — encerra silenciosamente
    except Exception as e:
        logger.critical(f"Erro fatal na inicialização: {e}", exc_info=True)
        raise